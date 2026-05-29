from __future__ import annotations

import re
import tempfile
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from importlib import resources
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from hr_toolkit.common.excel import apply_row_snapshot, snapshot_row


TOOL_NAME = "需求6-异动表汇总"
ROSTER_TOOL_NAME = "需求6-花名册更新"
DEFAULT_OUTPUT_FILENAME = "异动汇总表.xlsx"
DEFAULT_SUMMARY_TEMPLATE_RESOURCE = "personnel_change_summary_template.xlsx"
DATE_NUMBER_FORMAT = "yyyy/m/d"
TARGET_SHEETS = ("增员", "减员", "转正", "调动")
HEADER_SERIAL = "序号"
HEADER_NAME = "姓名"
HEADER_ID_CARD = "身份证号码"
LEAVE_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
THIN_BLACK_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

SOURCE_SHEET_ALIASES = {
    "增员": ("增员", "增补表"),
    "减员": ("减员", "离职"),
    "转正": ("转正",),
    "调动": ("调动", "调整"),
}

FIELD_ALIASES = {
    "入职公司": ("入职公司", "公司"),
    "公司": ("公司", "入职公司"),
    "地市": ("地市", "部门（片区）"),
    "部门（片区）": ("部门（片区）", "部门片区"),
    "身份证号码": ("身份证号码", "身份证"),
    "岗位": ("岗位", "职务"),
    "职务": ("职务", "岗位"),
    "所属专业": ("所属专业", "专业"),
    "联系方式": ("联系方式", "工作联系电话"),
    "毕业学校": ("毕业学校", "毕业院校"),
    "毕业院校": ("毕业院校", "毕业学校"),
    "入职时间": ("入职时间", "入职日期"),
    "入职日期": ("入职日期", "入职时间"),
}

FORMULA_TARGET_HEADERS = {"出生日期", "年龄"}
SUMMARY_FOOTER_KEYWORDS = ("对应项目部", "对应异动", "审批人", "审核人", "核对人", "制表人")

DEFAULT_HEADERS_BY_SHEET = {
    "增员": [
        "序号",
        "入职公司",
        "地市",
        "姓名",
        "身份证号码",
        "出生日期",
        "年龄",
        "性别",
        "岗位",
        "人员分类",
        "所属专业",
        "联系方式",
        "学历",
        "毕业学校",
        "专业",
        "毕业时间",
        "婚否",
        "家庭住址",
        "入职日期",
        "用工状态",
        "试用期工资",
        "备注",
    ],
    "减员": ["序号", "公司", "地市", "姓名", "身份证号码", "入职日期", "离职日期", "薪资结算日期", "备注"],
    "转正": ["序号", "公司", "部门（片区）", "姓名", "职务", "入职日期", "转正日期", "试用期工资（元）", "转正后岗级", "转正后工资（元）", "备注"],
    "调动": ["序号", "公司", "地市", "部门（片区）", "姓名", "原部门", "原职位", "原岗级", "原金额（元）", "现部门", "现职位", "异动类型", "调整后金额（元）", "增减额度（元）", "调整日期", "备注"],
}

KEY_HEADERS_BY_SHEET = {
    "增员": ("身份证号码", "姓名", "入职日期"),
    "减员": ("身份证号码", "姓名", "离职日期"),
    "转正": ("姓名", "入职日期", "转正日期"),
    "调动": ("姓名", "原部门", "现部门", "现职位", "异动类型", "调整日期"),
}

PERIOD_FIELD_BY_SHEET = {
    "增员": ("入职日期", "入职时间"),
    "减员": ("离职日期",),
    "转正": ("转正日期",),
    "调动": ("调整日期",),
}


@dataclass
class ChangeRow:
    sheet_name: str
    period: str
    values: dict[str, Any]
    source_file: str
    source_row: int


@dataclass
class PersonnelChangeMergeResult:
    input_dir: Path
    output_dir: Path
    output_file: Path | None = None
    output_files: list[Path] = field(default_factory=list)
    roster_output_file: Path | None = None
    dry_run: bool = False
    period: str | None = None
    append_mode: bool = False
    source_files: list[str] = field(default_factory=list)
    sheet_counts: dict[str, int] = field(default_factory=dict)
    period_counts: dict[str, dict[str, int]] = field(default_factory=dict)
    record_count: int = 0
    inserted_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    roster_added_count: int = 0
    roster_marked_count: int = 0
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "tool_name": TOOL_NAME,
            "input_dir": str(self.input_dir),
            "output_dir": str(self.output_dir),
            "output_file": None if self.output_file is None else str(self.output_file),
            "output_files": [str(path) for path in self.output_files],
            "roster_output_file": None if self.roster_output_file is None else str(self.roster_output_file),
            "dry_run": self.dry_run,
            "period": self.period,
            "append_mode": self.append_mode,
            "source_files": self.source_files,
            "source_file_count": len(self.source_files),
            "sheet_counts": self.sheet_counts,
            "period_counts": self.period_counts,
            "record_count": self.record_count,
            "inserted_count": self.inserted_count,
            "updated_count": self.updated_count,
            "skipped_count": self.skipped_count,
            "roster_added_count": self.roster_added_count,
            "roster_marked_count": self.roster_marked_count,
            "warnings": self.warnings,
        }


@dataclass
class RosterUpdateResult:
    summary_input: Path
    analysis_template_path: Path
    output_dir: Path
    output_file: Path | None = None
    dry_run: bool = False
    period: str | None = None
    source_files: list[str] = field(default_factory=list)
    sheet_counts: dict[str, int] = field(default_factory=dict)
    record_count: int = 0
    roster_added_count: int = 0
    roster_marked_count: int = 0
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "tool_name": ROSTER_TOOL_NAME,
            "summary_input": str(self.summary_input),
            "analysis_template_path": str(self.analysis_template_path),
            "output_dir": str(self.output_dir),
            "output_file": None if self.output_file is None else str(self.output_file),
            "dry_run": self.dry_run,
            "period": self.period,
            "source_files": self.source_files,
            "source_file_count": len(self.source_files),
            "sheet_counts": self.sheet_counts,
            "record_count": self.record_count,
            "roster_added_count": self.roster_added_count,
            "roster_marked_count": self.roster_marked_count,
            "warnings": self.warnings,
        }


@dataclass(frozen=True)
class ChangeSheetLayout:
    sheet_name: str
    header_row: int
    data_start_row: int
    footer_start_row: int
    max_column: int
    headers: dict[str, int]


def merge_personnel_changes(
    input_dir: str | Path | list[str | Path],
    output_dir: str | Path,
    *,
    template_path: str | Path | None = None,
    analysis_template_path: str | Path | None = None,
    dry_run: bool = False,
) -> PersonnelChangeMergeResult:
    input_paths = _normalize_input_paths(input_dir)
    display_input = input_paths[0] if len(input_paths) == 1 else input_paths[0].parent
    output_dir = Path(output_dir).expanduser().resolve()
    warnings: list[str] = []
    for path in input_paths:
        if not path.exists():
            raise FileNotFoundError(f"异动表文件、压缩包或文件夹不存在：{path}")

    with tempfile.TemporaryDirectory(prefix="hr_change_merge_") as temp_root:
        temp_dir = Path(temp_root)
        source_files = _find_change_files(input_paths, temp_dir, warnings)
        if not source_files:
            raise ValueError("未找到 .xlsx 异动表")

        summary_sources = _resolve_summary_sources(template_path)
        analysis_template = _resolve_analysis_template_path(analysis_template_path, input_paths)
        rows_by_period = _empty_period_sheet_map()
        used_files: list[str] = []

        for file_path in source_files:
            file_rows, file_warnings = _read_change_file(file_path)
            warnings.extend(file_warnings)
            if any(file_rows.values()):
                used_files.append(str(file_path))
            for sheet_name, rows in file_rows.items():
                for row in rows:
                    rows_by_period.setdefault(row.period, _empty_sheet_map())[sheet_name].append(row)

        rows_by_period = {period: rows for period, rows in sorted(rows_by_period.items()) if sum(len(items) for items in rows.values())}
        period_counts = {
            period: {sheet_name: len(rows.get(sheet_name, [])) for sheet_name in TARGET_SHEETS}
            for period, rows in rows_by_period.items()
        }
        sheet_counts = {sheet_name: sum(period_counts[period][sheet_name] for period in period_counts) for sheet_name in TARGET_SHEETS}
        record_count = sum(sheet_counts.values())
        result = PersonnelChangeMergeResult(
            input_dir=display_input,
            output_dir=output_dir,
            dry_run=dry_run,
            period=next(iter(rows_by_period), None) if len(rows_by_period) == 1 else None,
            append_mode=bool(summary_sources),
            source_files=used_files,
            sheet_counts=sheet_counts,
            period_counts=period_counts,
            record_count=record_count,
            warnings=warnings,
        )
        if dry_run:
            return result

        output_dir.mkdir(parents=True, exist_ok=True)
        inserted_count = 0
        updated_count = 0
        skipped_count = 0
        output_files: list[Path] = []
        for period, rows_by_sheet in rows_by_period.items():
            summary_path = summary_sources.get(period)
            output_file = output_dir / _output_filename(period)
            summary_result = _write_summary_workbook(summary_path, output_file, rows_by_sheet, period=period)
            inserted_count += summary_result["inserted_count"]
            updated_count += summary_result["updated_count"]
            skipped_count += summary_result["skipped_count"]
            output_files.append(output_file)

        result.output_files = output_files
        result.output_file = output_files[0] if len(output_files) == 1 else None
        result.inserted_count = inserted_count
        result.updated_count = updated_count
        result.skipped_count = skipped_count
        if analysis_template is not None:
            all_rows = _merge_period_rows(rows_by_period)
            roster_output_file = output_dir / _analysis_output_filename(result.period)
            roster_result = _write_updated_roster(analysis_template, roster_output_file, all_rows, warnings)
            result.roster_output_file = roster_output_file
            result.roster_added_count = roster_result["added_count"]
            result.roster_marked_count = roster_result["marked_count"]
        return result


def update_roster_from_change_summaries(
    summary_input: str | Path | list[str | Path],
    analysis_template_path: str | Path,
    output_dir: str | Path,
    *,
    dry_run: bool = False,
) -> RosterUpdateResult:
    input_paths = _normalize_input_paths(summary_input)
    display_input = input_paths[0] if len(input_paths) == 1 else input_paths[0].parent
    analysis_template = Path(analysis_template_path).expanduser().resolve()
    output_dir = Path(output_dir).expanduser().resolve()
    warnings: list[str] = []

    for path in input_paths:
        if not path.exists():
            raise FileNotFoundError(f"异动汇总表文件或文件夹不存在：{path}")
    if not analysis_template.exists() or not analysis_template.is_file():
        raise FileNotFoundError(f"人力资源花名册不存在：{analysis_template}")
    if analysis_template.suffix.lower() != ".xlsx":
        raise ValueError("人力资源花名册目前只支持 .xlsx 文件。")

    summary_files = _find_summary_files(input_paths, warnings)
    if not summary_files:
        raise ValueError("未找到 .xlsx 异动汇总表")

    rows_by_sheet = _empty_sheet_map()
    used_files: list[str] = []
    periods: set[str] = set()
    for file_path in summary_files:
        file_rows, file_warnings = _read_summary_change_file(file_path)
        warnings.extend(file_warnings)
        if any(file_rows.values()):
            used_files.append(str(file_path))
        for sheet_name, rows in file_rows.items():
            rows_by_sheet[sheet_name].extend(rows)
            periods.update(row.period for row in rows if row.period)

    sheet_counts = {sheet_name: len(rows_by_sheet.get(sheet_name, [])) for sheet_name in TARGET_SHEETS}
    record_count = sum(sheet_counts.values())
    period = next(iter(periods)) if len(periods) == 1 else None
    result = RosterUpdateResult(
        summary_input=display_input,
        analysis_template_path=analysis_template,
        output_dir=output_dir,
        dry_run=dry_run,
        period=period,
        source_files=used_files,
        sheet_counts=sheet_counts,
        record_count=record_count,
        warnings=warnings,
    )
    if dry_run:
        return result

    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / _analysis_output_filename(period)
    roster_result = _write_updated_roster(analysis_template, output_file, rows_by_sheet, warnings)
    result.output_file = output_file
    result.roster_added_count = roster_result["added_count"]
    result.roster_marked_count = roster_result["marked_count"]
    return result


def _normalize_input_paths(input_path: str | Path | list[str | Path]) -> list[Path]:
    raw_paths = input_path if isinstance(input_path, list) else [input_path]
    paths = [Path(path).expanduser().resolve() for path in raw_paths]
    if not paths:
        raise ValueError("请选择异动表文件、压缩包或文件夹。")
    return paths


def _empty_sheet_map() -> dict[str, list[ChangeRow]]:
    return {sheet_name: [] for sheet_name in TARGET_SHEETS}


def _empty_period_sheet_map() -> dict[str, dict[str, list[ChangeRow]]]:
    return {}


def _merge_period_rows(rows_by_period: dict[str, dict[str, list[ChangeRow]]]) -> dict[str, list[ChangeRow]]:
    merged = _empty_sheet_map()
    for rows_by_sheet in rows_by_period.values():
        for sheet_name, rows in rows_by_sheet.items():
            merged[sheet_name].extend(rows)
    return merged


def _find_change_files(input_paths: list[Path], temp_dir: Path, warnings: list[str]) -> list[Path]:
    files: list[Path] = []
    seen: set[Path] = set()
    for path in input_paths:
        for file_path in _iter_input_files(path, temp_dir, warnings):
            resolved = file_path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            files.append(file_path)
    return sorted(files)


def _iter_input_files(path: Path, temp_dir: Path, warnings: list[str]) -> list[Path]:
    if path.is_file():
        suffix = path.suffix.lower()
        if suffix == ".xlsx" and not path.name.startswith("~$"):
            return [path]
        if suffix == ".zip":
            return _extract_zip_change_files(path, temp_dir, warnings)
        return []
    if path.is_dir():
        files: list[Path] = []
        for child in sorted(path.rglob("*")):
            if not child.is_file() or child.name.startswith("~$"):
                continue
            if child.suffix.lower() == ".xlsx":
                files.append(child)
            elif child.suffix.lower() == ".zip":
                files.extend(_extract_zip_change_files(child, temp_dir, warnings))
        return files
    return []


def _extract_zip_change_files(zip_path: Path, temp_dir: Path, warnings: list[str]) -> list[Path]:
    extract_dir = temp_dir / f"zip_{len(list(temp_dir.glob('zip_*'))) + 1}"
    extract_dir.mkdir(parents=True, exist_ok=True)
    try:
        with zipfile.ZipFile(zip_path) as archive:
            for member in archive.infolist():
                target = extract_dir / member.filename
                if not target.resolve().is_relative_to(extract_dir.resolve()):
                    warnings.append(f"{zip_path.name} 中存在不安全路径，已跳过：{member.filename}")
                    continue
                archive.extract(member, extract_dir)
    except Exception as exc:
        warnings.append(f"{zip_path.name} 解压失败，已跳过：{exc}")
        return []
    return sorted(path for path in extract_dir.rglob("*.xlsx") if path.is_file() and not path.name.startswith("~$"))


def _find_summary_files(input_paths: list[Path], warnings: list[str]) -> list[Path]:
    files: list[Path] = []
    seen: set[Path] = set()
    for path in input_paths:
        if path.is_file():
            candidates = [path]
        elif path.is_dir():
            candidates = sorted(item for item in path.rglob("*.xlsx") if item.is_file() and not item.name.startswith("~$"))
        else:
            candidates = []
        for candidate in candidates:
            if candidate.suffix.lower() != ".xlsx" or candidate.name.startswith("~$"):
                continue
            resolved = candidate.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            if _looks_like_summary_file(candidate):
                files.append(candidate)
    return sorted(files)


def _looks_like_summary_file(path: Path) -> bool:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return all(sheet_name in workbook.sheetnames for sheet_name in TARGET_SHEETS)
        finally:
            workbook.close()
    except Exception:
        return False


def _resolve_summary_sources(template_path: str | Path | None) -> dict[str, Path]:
    if not template_path:
        return {}
    path = Path(template_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"异动汇总表文件或文件夹不存在：{path}")
    candidates = [path] if path.is_file() else sorted(item for item in path.glob("*.xlsx") if item.is_file() and not item.name.startswith("~$"))
    summary_sources: dict[str, Path] = {}
    for candidate in candidates:
        if candidate.suffix.lower() != ".xlsx":
            continue
        period = _detect_summary_period(candidate)
        if period:
            summary_sources[period] = candidate
    return summary_sources


def _read_change_file(file_path: Path) -> tuple[dict[str, list[ChangeRow]], list[str]]:
    warnings: list[str] = []
    rows_by_sheet: dict[str, list[ChangeRow]] = {sheet_name: [] for sheet_name in TARGET_SHEETS}
    workbook = load_workbook(file_path, data_only=True)
    try:
        file_period = _detect_period([file_path])
        for sheet_name in TARGET_SHEETS:
            ws = _find_source_sheet(workbook, sheet_name)
            if ws is None:
                if _looks_like_change_workbook(workbook):
                    warnings.append(f"{file_path.name} 缺少工作表：{sheet_name}")
                continue
            layout = _detect_sheet_layout(ws)
            rows_by_sheet[sheet_name].extend(
                _read_data_rows(ws, layout, file_path.name, target_sheet=sheet_name, file_period=file_period, warnings=warnings)
            )
    finally:
        workbook.close()
    return rows_by_sheet, warnings


def _read_summary_change_file(file_path: Path) -> tuple[dict[str, list[ChangeRow]], list[str]]:
    warnings: list[str] = []
    rows_by_sheet: dict[str, list[ChangeRow]] = {sheet_name: [] for sheet_name in TARGET_SHEETS}
    workbook = load_workbook(file_path, data_only=True)
    try:
        file_period = _detect_summary_period(file_path)
        for sheet_name in TARGET_SHEETS:
            if sheet_name not in workbook.sheetnames:
                warnings.append(f"{file_path.name} 缺少工作表：{sheet_name}")
                continue
            ws = workbook[sheet_name]
            layout = _detect_sheet_layout(ws)
            for row_index in range(layout.data_start_row, layout.footer_start_row):
                if not _is_existing_summary_data_row(ws, layout, row_index):
                    continue
                values = {header: ws.cell(row_index, col_index).value for header, col_index in layout.headers.items()}
                period = _detect_change_row_period(values, sheet_name, file_period) or file_period or ""
                rows_by_sheet[sheet_name].append(
                    ChangeRow(
                        sheet_name=sheet_name,
                        period=period,
                        values=values,
                        source_file=file_path.name,
                        source_row=row_index,
                    )
                )
    finally:
        workbook.close()
    return rows_by_sheet, warnings


def _detect_sheet_layout(ws: Worksheet) -> ChangeSheetLayout:
    header_row = _find_header_row(ws)
    return ChangeSheetLayout(
        sheet_name=ws.title,
        header_row=header_row,
        data_start_row=header_row + 1,
        footer_start_row=_find_summary_footer_start(ws, header_row + 1),
        max_column=_last_header_column(ws, header_row),
        headers=_read_headers(ws, header_row),
    )


def _find_header_row(ws: Worksheet) -> int:
    for row_index in range(1, min(ws.max_row, 20) + 1):
        for col_index in range(1, ws.max_column + 1):
            value = ws.cell(row_index, col_index).value
            if str(value or "").strip() == HEADER_SERIAL:
                return row_index
    raise ValueError(f"{ws.title} 未在前 20 行找到“{HEADER_SERIAL}”表头")


def _last_header_column(ws: Worksheet, header_row: int) -> int:
    last_column = 1
    for col_index in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_index).value
        if value not in (None, ""):
            last_column = col_index
    return last_column


def _read_headers(ws: Worksheet, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_index in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_index).value
        if value not in (None, ""):
            headers[_normalize_header(value)] = col_index
    return headers


def _read_data_rows(
    ws: Worksheet,
    layout: ChangeSheetLayout,
    source_file: str,
    *,
    target_sheet: str,
    file_period: str | None,
    warnings: list[str],
) -> list[ChangeRow]:
    rows: list[ChangeRow] = []
    for row_index in range(layout.data_start_row, layout.footer_start_row):
        values = {
            header: ws.cell(row_index, col_index).value
            for header, col_index in layout.headers.items()
        }
        if not _is_filled_change_row(values):
            continue
        period = _detect_change_row_period(values, target_sheet, file_period)
        if not period:
            warnings.append(f"{source_file} {ws.title} 第 {row_index} 行缺少可识别日期，已跳过。")
            continue
        rows.append(
            ChangeRow(
                sheet_name=target_sheet,
                period=period,
                values=values,
                source_file=source_file,
                source_row=row_index,
            )
        )
    return rows


def _find_summary_footer_start(ws: Worksheet, start_row: int) -> int:
    for row_index in range(start_row, ws.max_row + 1):
        row_text = " ".join(_cell_text(ws.cell(row_index, col_index).value) for col_index in range(1, min(ws.max_column, 12) + 1))
        if any(keyword in row_text for keyword in SUMMARY_FOOTER_KEYWORDS):
            return row_index
    return ws.max_row + 1


def _is_filled_change_row(values: dict[str, Any]) -> bool:
    # 模板中常预填“序号”，但其他列为空；这种行不是有效异动记录。
    return any(_has_value(value) for header, value in values.items() if header != HEADER_SERIAL)


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def _write_summary_workbook(
    summary_path: Path | None,
    output_file: Path,
    rows_by_sheet: dict[str, list[ChangeRow]],
    *,
    period: str | None,
) -> dict[str, int]:
    workbook = load_workbook(summary_path) if summary_path is not None else _create_default_summary_workbook(period)
    inserted_count = 0
    updated_count = 0
    skipped_count = 0
    try:
        missing_sheets = [sheet_name for sheet_name in TARGET_SHEETS if sheet_name not in workbook.sheetnames]
        if missing_sheets:
            raise ValueError(f"异动汇总表缺少工作表：{'、'.join(missing_sheets)}")
        if summary_path is None:
            _update_period_titles(workbook, period)
        for sheet_name in TARGET_SHEETS:
            ws = workbook[sheet_name]
            sheet_result = _append_sheet_rows(ws, rows_by_sheet.get(sheet_name, []))
            inserted_count += sheet_result["inserted_count"]
            updated_count += sheet_result["updated_count"]
            skipped_count += sheet_result["skipped_count"]
        workbook.save(output_file)
    finally:
        workbook.close()
    return {"inserted_count": inserted_count, "updated_count": updated_count, "skipped_count": skipped_count}


def _append_sheet_rows(ws: Worksheet, rows: list[ChangeRow]) -> dict[str, int]:
    inserted_count = 0
    updated_count = 0
    skipped_count = 0
    for row in rows:
        layout = _detect_sheet_layout(ws)
        existing = _existing_change_index(ws, layout, row.sheet_name)
        row_key = _change_key_from_source(row, layout)
        if row_key and row_key in existing:
            if _merge_existing_change_row(ws, layout, existing[row_key], row):
                updated_count += 1
            else:
                skipped_count += 1
            continue
        target_row = _next_summary_write_row(ws, layout)
        layout = _detect_sheet_layout(ws)
        _write_change_row(ws, layout, target_row, row)
        inserted_count += 1
    _renumber_summary_sheet(ws, _detect_sheet_layout(ws))
    return {"inserted_count": inserted_count, "updated_count": updated_count, "skipped_count": skipped_count}


def _existing_change_index(ws: Worksheet, layout: ChangeSheetLayout, sheet_name: str) -> dict[tuple[str, ...], int]:
    existing: dict[tuple[str, ...], int] = {}
    for row_index in range(layout.data_start_row, layout.footer_start_row):
        if not _is_existing_summary_data_row(ws, layout, row_index):
            continue
        key = _change_key_from_target(ws, layout, row_index, sheet_name)
        if key:
            existing[key] = row_index
    return existing


def _next_summary_write_row(ws: Worksheet, layout: ChangeSheetLayout) -> int:
    blank_row = _first_blank_summary_row(ws, layout)
    if blank_row is not None:
        return blank_row
    if layout.footer_start_row <= ws.max_row:
        target_row = layout.footer_start_row
        template_row = max(layout.data_start_row, target_row - 1)
        template_snapshot = snapshot_row(ws, template_row, layout.max_column)
        ws.insert_rows(target_row, 1)
        apply_row_snapshot(ws, target_row, template_snapshot, translate_formulas=True)
        _clear_row_values(ws, target_row, layout.max_column)
        return target_row
    target_row = ws.max_row + 1
    template_row = max(layout.data_start_row, target_row - 1)
    template_snapshot = snapshot_row(ws, template_row, layout.max_column)
    apply_row_snapshot(ws, target_row, template_snapshot, translate_formulas=True)
    _clear_row_values(ws, target_row, layout.max_column)
    return target_row


def _first_blank_summary_row(ws: Worksheet, layout: ChangeSheetLayout) -> int | None:
    for row_index in range(layout.data_start_row, layout.footer_start_row):
        if not _is_existing_summary_data_row(ws, layout, row_index):
            return row_index
    return None


def _is_existing_summary_data_row(ws: Worksheet, layout: ChangeSheetLayout, row_index: int) -> bool:
    for header, col_index in layout.headers.items():
        if header == HEADER_SERIAL or header in FORMULA_TARGET_HEADERS:
            continue
        if _has_value(ws.cell(row_index, col_index).value):
            return True
    return False


def _write_change_row(ws: Worksheet, layout: ChangeSheetLayout, row_index: int, row: ChangeRow) -> None:
    template_snapshot = snapshot_row(ws, row_index, layout.max_column)
    apply_row_snapshot(ws, row_index, template_snapshot, translate_formulas=True)
    _clear_row_values(ws, row_index, layout.max_column)
    for header, col_index in layout.headers.items():
        if header == HEADER_SERIAL or header in FORMULA_TARGET_HEADERS:
            continue
        value = _mapped_change_value(row.values, header, row.sheet_name)
        if _has_value(value):
            cell = ws.cell(row_index, col_index)
            cell.value = value
            _apply_date_number_format(cell, header)
    _write_summary_formulas(ws, layout, row_index)
    _center_row_cells(ws, row_index, layout.max_column)


def _merge_existing_change_row(ws: Worksheet, layout: ChangeSheetLayout, row_index: int, row: ChangeRow) -> bool:
    changed = False
    for header, col_index in layout.headers.items():
        if header == HEADER_SERIAL or header in FORMULA_TARGET_HEADERS:
            continue
        value = _mapped_change_value(row.values, header, row.sheet_name)
        cell = ws.cell(row_index, col_index)
        if _has_value(value) and not _has_value(cell.value):
            cell.value = value
            _apply_date_number_format(cell, header)
            changed = True
    before = tuple(ws.cell(row_index, col_index).value for header, col_index in layout.headers.items() if header in FORMULA_TARGET_HEADERS)
    _write_summary_formulas(ws, layout, row_index)
    after = tuple(ws.cell(row_index, col_index).value for header, col_index in layout.headers.items() if header in FORMULA_TARGET_HEADERS)
    _center_row_cells(ws, row_index, layout.max_column)
    return changed or before != after


def _clear_row_values(ws: Worksheet, row_index: int, max_column: int) -> None:
    for col_index in range(1, max_column + 1):
        ws.cell(row_index, col_index).value = None


def _write_summary_formulas(ws: Worksheet, layout: ChangeSheetLayout, row_index: int) -> None:
    id_col = layout.headers.get(HEADER_ID_CARD)
    birth_col = layout.headers.get("出生日期")
    if id_col is None:
        return
    id_ref = f"{get_column_letter(id_col)}{row_index}"
    if birth_col is not None:
        cell = ws.cell(row_index, birth_col)
        cell.value = f'=MIDB({id_ref},7,4)&"-"&MIDB({id_ref},11,2)&"-"&MIDB({id_ref},13,2)'
        _apply_date_number_format(cell, "出生日期")
    age_col = layout.headers.get("年龄")
    if age_col is not None and birth_col is not None:
        birth_ref = f"{get_column_letter(birth_col)}{row_index}"
        ws.cell(row_index, age_col).value = f'=DATEDIF({birth_ref},TODAY(),"Y")'


def _center_row_cells(ws: Worksheet, row_index: int, max_column: int) -> None:
    for col_index in range(1, max_column + 1):
        cell = ws.cell(row_index, col_index)
        cell.border = THIN_BLACK_BORDER
        current = cell.alignment
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            textRotation=current.textRotation,
            wrapText=current.wrapText,
            shrinkToFit=current.shrinkToFit,
            indent=current.indent,
            relativeIndent=current.relativeIndent,
            justifyLastLine=current.justifyLastLine,
            readingOrder=current.readingOrder,
        )


def _renumber_summary_sheet(ws: Worksheet, layout: ChangeSheetLayout) -> None:
    serial_col = layout.headers.get(HEADER_SERIAL)
    if serial_col is None:
        return
    serial = 1
    for row_index in range(layout.data_start_row, layout.footer_start_row):
        if _is_existing_summary_data_row(ws, layout, row_index):
            ws.cell(row_index, serial_col).value = serial
            _center_row_cells(ws, row_index, layout.max_column)
            serial += 1
        else:
            ws.cell(row_index, serial_col).value = None


def _change_key_from_source(row: ChangeRow, layout: ChangeSheetLayout) -> tuple[str, ...] | None:
    return _build_change_key(row.sheet_name, lambda header: _mapped_change_value(row.values, header, row.sheet_name), layout.headers)


def _change_key_from_target(ws: Worksheet, layout: ChangeSheetLayout, row_index: int, sheet_name: str) -> tuple[str, ...] | None:
    return _build_change_key(sheet_name, lambda header: _target_cell_value(ws, layout, row_index, header), layout.headers)


def _build_change_key(sheet_name: str, getter, target_headers: dict[str, int]) -> tuple[str, ...] | None:
    id_card = _normalize_id_card(getter(HEADER_ID_CARD))
    if id_card:
        date_header = {
            "增员": "入职日期",
            "减员": "离职日期",
        }.get(sheet_name)
        date_part = _normalize_key_value(getter(date_header), date_header) if date_header else ""
        return (sheet_name, "id", id_card, date_part)
    key_headers = KEY_HEADERS_BY_SHEET.get(sheet_name, (HEADER_NAME,))
    parts = tuple(_normalize_key_value(getter(header), header) for header in key_headers if header in target_headers or _has_value(getter(header)))
    if any(parts):
        return (sheet_name, "fields", *parts)
    return None


def _target_cell_value(ws: Worksheet, layout: ChangeSheetLayout, row_index: int, header: str) -> Any:
    col_index = layout.headers.get(header)
    if col_index is None:
        return None
    return ws.cell(row_index, col_index).value


def _normalize_key_value(value: Any, header: str | None = None) -> str:
    if header and any(keyword in header for keyword in ("日期", "时间")) and isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            pass
    if hasattr(value, "date") and callable(value.date):
        try:
            return value.date().isoformat()
        except Exception:
            pass
    return _cell_text(value).upper()


def _create_default_summary_workbook(period: str | None) -> Workbook:
    workbook = _load_packaged_summary_template()
    if workbook is not None:
        return workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in TARGET_SHEETS:
        headers = DEFAULT_HEADERS_BY_SHEET[sheet_name]
        ws = workbook.create_sheet(sheet_name)
        title = _default_sheet_title(sheet_name, period)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(1, 1).value = title
        ws.cell(1, 1).font = Font(name="宋体", bold=True, size=14)
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        header_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(2, col_index)
            cell.value = header
            cell.font = Font(name="宋体", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            cell.border = border
            ws.cell(3, col_index).border = border
            ws.cell(3, col_index).alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
        for col_index, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_index)].width = _default_column_width(header)
    return workbook


def _load_packaged_summary_template() -> Workbook | None:
    try:
        template = resources.files("hr_toolkit.templates").joinpath(DEFAULT_SUMMARY_TEMPLATE_RESOURCE)
        with template.open("rb") as handle:
            return load_workbook(handle)
    except Exception:
        return None


def _default_sheet_title(sheet_name: str, period: str | None) -> str:
    prefix = period or ""
    title_map = {
        "转正": "员工转正表",
        "调动": "人员调整表",
        "奖罚扣补": "奖、罚、扣、补表",
    }
    return f"{prefix}{title_map.get(sheet_name, sheet_name + '表')}"


def _default_column_width(header: str) -> int:
    if header in {"身份证号码", "家庭住址"}:
        return 20
    if header in {"入职公司", "部门（片区）", "联系方式", "毕业学校"}:
        return 14
    if header in {"序号", "年龄", "性别"}:
        return 8
    return 12


def _find_source_sheet(workbook, target_sheet: str) -> Worksheet | None:
    for alias in SOURCE_SHEET_ALIASES[target_sheet]:
        if alias in workbook.sheetnames:
            return workbook[alias]
    normalized_aliases = {_normalize_header(alias) for alias in SOURCE_SHEET_ALIASES[target_sheet]}
    for ws in workbook.worksheets:
        if _normalize_header(ws.title) in normalized_aliases:
            return ws
    return None


def _looks_like_change_workbook(workbook) -> bool:
    for target_sheet in TARGET_SHEETS:
        if _find_source_sheet(workbook, target_sheet) is not None:
            return True
    return False


def _mapped_change_value(source_values: dict[str, Any], target_header: str, target_sheet: str) -> Any:
    if target_header in FORMULA_TARGET_HEADERS:
        return None
    if target_sheet == "调动" and target_header == "地市":
        return _first_value(source_values, ("地市",))
    candidates = FIELD_ALIASES.get(target_header, (target_header,))
    return _first_value(source_values, candidates)


def _first_value(values: dict[str, Any], headers: tuple[str, ...]) -> Any:
    normalized_values = {_normalize_header(header): value for header, value in values.items()}
    for header in headers:
        value = normalized_values.get(_normalize_header(header))
        if _has_value(value):
            return value
    return None


def _detect_change_row_period(values: dict[str, Any], sheet_name: str, file_period: str | None) -> str | None:
    for header in PERIOD_FIELD_BY_SHEET[sheet_name]:
        period = _period_from_value(_first_value(values, (header,)), file_period)
        if period:
            return period
    return None


def _period_from_value(value: Any, fallback_period: str | None = None) -> str | None:
    if not _has_value(value):
        return None
    if isinstance(value, datetime):
        return f"{value.year}年{value.month}月"
    if isinstance(value, date):
        return f"{value.year}年{value.month}月"
    if isinstance(value, (int, float)):
        try:
            value_date = from_excel(value)
            return f"{value_date.year}年{value_date.month}月"
        except Exception:
            return None
    text = _cell_text(value)
    match = re.search(r"(20\d{2})[年/\-. ]+([01]?\d)月?", text)
    if match:
        year, month = match.groups()
        return f"{year}年{int(month)}月"
    match = re.search(r"([01]?\d)月", text)
    if match and fallback_period:
        year_match = re.search(r"(20\d{2})年", fallback_period)
        if year_match:
            return f"{year_match.group(1)}年{int(match.group(1))}月"
    return None


def _resolve_analysis_template_path(template_path: str | Path | None, input_paths: list[Path]) -> Path | None:
    if template_path:
        path = Path(template_path).expanduser().resolve()
        if not path.exists() or not path.is_file():
            raise FileNotFoundError(f"人力资源分析表不存在：{path}")
        return path
    for input_path in input_paths:
        if input_path.is_file():
            candidates = [input_path]
        elif input_path.is_dir():
            candidates = sorted(input_path.glob("*.xlsx"))
        else:
            candidates = []
        for path in candidates:
            if path.name.startswith("~$"):
                continue
            try:
                workbook = load_workbook(path, read_only=True, data_only=True)
                try:
                    if "花名册" in workbook.sheetnames:
                        return path
                finally:
                    workbook.close()
            except Exception:
                continue
    return None


def _detect_period(paths: list[Path | None]) -> str | None:
    pattern = re.compile(r"(20\d{2})年?([01]?\d)月")
    for path in paths:
        if path is None:
            continue
        match = pattern.search(path.name)
        if match:
            year, month = match.groups()
            return f"{year}年{int(month)}月"
    return None


def _detect_summary_period(path: Path) -> str | None:
    period = _detect_period([path])
    if period:
        return period
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in TARGET_SHEETS:
                if sheet_name not in workbook.sheetnames:
                    continue
                ws = workbook[sheet_name]
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 3), values_only=True):
                    for value in row:
                        if isinstance(value, str):
                            match = re.search(r"(20\d{2})年?([01]?\d)月", value)
                            if match:
                                year, month = match.groups()
                                return f"{year}年{int(month)}月"
        finally:
            workbook.close()
    except Exception:
        return None
    return None


def _output_filename(period: str | None) -> str:
    return f"{period}异动汇总表.xlsx" if period else DEFAULT_OUTPUT_FILENAME


def _analysis_output_filename(period: str | None) -> str:
    return f"{period}人力资源分析表_更新后.xlsx" if period else "人力资源分析表_更新后.xlsx"


def _update_period_titles(workbook, period: str | None) -> None:
    if not period:
        return
    for sheet_name in TARGET_SHEETS:
        ws = workbook[sheet_name]
        value = ws.cell(1, 1).value
        if isinstance(value, str):
            if re.search(r"20\d{2}年[01]?\d月", value):
                ws.cell(1, 1).value = re.sub(r"20\d{2}年[01]?\d月", period, value, count=1)
            elif value:
                ws.cell(1, 1).value = f"{period}{value}"


def _write_updated_roster(
    analysis_template: Path,
    output_file: Path,
    rows_by_sheet: dict[str, list[ChangeRow]],
    warnings: list[str],
) -> dict[str, int]:
    workbook = load_workbook(analysis_template)
    added_count = 0
    marked_count = 0
    try:
        if "花名册" not in workbook.sheetnames:
            raise ValueError("人力资源分析表缺少“花名册”工作表。")
        ws = workbook["花名册"]
        layout = _detect_roster_layout(ws)
        existing = _roster_existing_records(ws, layout)
        for row in rows_by_sheet.get("增员", []):
            id_card = _normalize_id_card(_mapped_change_value(row.values, HEADER_ID_CARD, "增员"))
            if not id_card:
                warnings.append(f"{row.source_file} 第 {row.source_row} 行增员缺少身份证号码，未写入花名册。")
                continue
            if id_card in existing:
                warnings.append(f"花名册已存在增员身份证 {id_card}，未重复写入。")
                continue
            inserted_row = _insert_roster_addition(ws, layout, row)
            existing[id_card] = inserted_row
            added_count += 1
        layout = _detect_roster_layout(ws)
        existing = _roster_existing_records(ws, layout)
        for row in rows_by_sheet.get("减员", []):
            id_card = _normalize_id_card(_mapped_change_value(row.values, HEADER_ID_CARD, "减员"))
            if not id_card:
                continue
            row_index = existing.get(id_card)
            if row_index is None:
                warnings.append(f"减员人员未在花名册找到：{id_card}")
                continue
            _mark_roster_leave(ws, layout, row_index)
            marked_count += 1
        _renumber_roster(ws, _detect_roster_layout(ws))
        workbook.save(output_file)
    finally:
        workbook.close()
    return {"added_count": added_count, "marked_count": marked_count}


@dataclass(frozen=True)
class RosterLayout:
    header_row: int
    data_start_row: int
    footer_start_row: int
    max_column: int
    headers: dict[str, int]


def _detect_roster_layout(ws: Worksheet) -> RosterLayout:
    header_row = _find_header_row(ws)
    headers = _read_headers(ws, header_row)
    footer_start = _find_roster_footer_start(ws, header_row + 1)
    return RosterLayout(
        header_row=header_row,
        data_start_row=header_row + 1,
        footer_start_row=footer_start,
        max_column=_last_header_column(ws, header_row),
        headers=headers,
    )


def _find_roster_footer_start(ws: Worksheet, start_row: int) -> int:
    for row_index in range(start_row, ws.max_row + 1):
        row_text = " ".join(_cell_text(ws.cell(row_index, col_index).value) for col_index in range(1, 6))
        if any(keyword in row_text for keyword in ("对应异动", "根据异动汇总表")):
            return row_index
    return ws.max_row + 1


def _roster_existing_records(ws: Worksheet, layout: RosterLayout) -> dict[str, int]:
    id_col = layout.headers.get(HEADER_ID_CARD)
    if id_col is None:
        return {}
    records: dict[str, int] = {}
    for row_index in range(layout.data_start_row, layout.footer_start_row):
        id_card = _normalize_id_card(ws.cell(row_index, id_col).value)
        if id_card:
            records[id_card] = row_index
    return records


def _insert_roster_addition(ws: Worksheet, layout: RosterLayout, row: ChangeRow) -> int:
    project = _cell_text(_mapped_change_value(row.values, "地市", "增员"))
    insert_at = _roster_insert_row(ws, layout, project)
    template_row = max(layout.data_start_row, insert_at - 1)
    if template_row >= layout.footer_start_row:
        template_row = layout.data_start_row
    template_snapshot = snapshot_row(ws, template_row, layout.max_column)
    ws.insert_rows(insert_at, 1)
    apply_row_snapshot(ws, insert_at, template_snapshot, translate_formulas=True)
    values = _roster_values_for_addition(ws, layout, row, project, template_row)
    for header, value in values.items():
        col_index = layout.headers.get(header)
        if col_index is not None:
            ws.cell(insert_at, col_index).value = value
            _apply_date_number_format(ws.cell(insert_at, col_index), header)
    _write_roster_formulas(ws, layout, insert_at)
    _format_roster_data_row(ws, layout, insert_at)
    return insert_at


def _roster_insert_row(ws: Worksheet, layout: RosterLayout, project: str) -> int:
    project_col = layout.headers.get("部门/项目")
    if project and project_col is not None:
        last_match = None
        for row_index in range(layout.data_start_row, layout.footer_start_row):
            if _cell_text(ws.cell(row_index, project_col).value) == project:
                last_match = row_index
        if last_match is not None:
            return last_match + 1
    return layout.footer_start_row


def _roster_values_for_addition(ws: Worksheet, layout: RosterLayout, row: ChangeRow, project: str, template_row: int) -> dict[str, Any]:
    department = None
    department_col = layout.headers.get("部门")
    project_col = layout.headers.get("部门/项目")
    if department_col is not None and project_col is not None and project:
        for row_index in range(layout.data_start_row, layout.footer_start_row):
            if _cell_text(ws.cell(row_index, project_col).value) == project:
                department = ws.cell(row_index, department_col).value
                break
    if department is None and department_col is not None:
        department = ws.cell(template_row, department_col).value
    return {
        "部门": department,
        "部门/项目": project,
        "姓名": _mapped_change_value(row.values, "姓名", "增员"),
        "身份证号码": _mapped_change_value(row.values, "身份证号码", "增员"),
        "性别": _mapped_change_value(row.values, "性别", "增员"),
        "岗位": _mapped_change_value(row.values, "岗位", "增员"),
        "人员分类": _mapped_change_value(row.values, "人员分类", "增员"),
        "所属专业": _mapped_change_value(row.values, "所属专业", "增员"),
        "联系方式": _mapped_change_value(row.values, "联系方式", "增员"),
        "入职公司": _mapped_change_value(row.values, "入职公司", "增员"),
        "学历": _mapped_change_value(row.values, "学历", "增员"),
        "毕业院校": _mapped_change_value(row.values, "毕业院校", "增员"),
        "专业": _mapped_change_value(row.values, "专业", "增员"),
        "入职时间": _mapped_change_value(row.values, "入职日期", "增员"),
    }


def _write_roster_formulas(ws: Worksheet, layout: RosterLayout, row_index: int) -> None:
    formulas = {
        "出生日期": f'=MIDB(E{row_index},7,4)&"-"&MIDB(E{row_index},11,2)&"-"&MIDB(E{row_index},13,2)',
        "年龄": f'=DATEDIF(F{row_index},TODAY(),"Y")',
    }
    for header, formula in formulas.items():
        col_index = layout.headers.get(header)
        if col_index is not None:
            cell = ws.cell(row_index, col_index)
            cell.value = formula
            _apply_date_number_format(cell, header)


def _mark_roster_leave(ws: Worksheet, layout: RosterLayout, row_index: int) -> None:
    for col_index in range(1, layout.max_column + 1):
        ws.cell(row_index, col_index).fill = LEAVE_FILL
    _format_roster_data_row(ws, layout, row_index)


def _format_roster_data_row(ws: Worksheet, layout: RosterLayout, row_index: int) -> None:
    for col_index in range(1, layout.max_column + 1):
        cell = ws.cell(row_index, col_index)
        header = _header_by_column(layout.headers, col_index)
        if header:
            _apply_date_number_format(cell, header)
        cell.border = THIN_BLACK_BORDER
        current = cell.alignment
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            textRotation=current.textRotation,
            wrapText=current.wrapText,
            shrinkToFit=current.shrinkToFit,
            indent=current.indent,
            relativeIndent=current.relativeIndent,
            justifyLastLine=current.justifyLastLine,
            readingOrder=current.readingOrder,
        )


def _renumber_roster(ws: Worksheet, layout: RosterLayout) -> None:
    serial_col = layout.headers.get(HEADER_SERIAL)
    if serial_col is None:
        return
    serial = 1
    for row_index in range(layout.data_start_row, layout.footer_start_row):
        name_col = layout.headers.get(HEADER_NAME)
        id_col = layout.headers.get(HEADER_ID_CARD)
        if name_col and id_col and not _has_value(ws.cell(row_index, name_col).value) and not _has_value(ws.cell(row_index, id_col).value):
            continue
        ws.cell(row_index, serial_col).value = serial
        _format_roster_data_row(ws, layout, row_index)
        serial += 1


def _apply_date_number_format(cell, header: str) -> None:
    if any(keyword in header for keyword in ("日期", "时间")):
        cell.number_format = DATE_NUMBER_FORMAT


def _header_by_column(headers: dict[str, int], col_index: int) -> str | None:
    for header, header_col_index in headers.items():
        if header_col_index == col_index:
            return header
    return None


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip())


def _normalize_id_card(value: Any) -> str:
    return _cell_text(value).upper()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
