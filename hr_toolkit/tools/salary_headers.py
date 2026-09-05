"""工资合并的表头定位与用户确认规则；不参与金额计算。"""

from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from collections import Counter
from typing import Any

from openpyxl.utils import get_column_letter

FIELD_LABELS = {"name": "姓名", "id_card": "身份证号码", "amount": "应发工资"}
ALIASES = {
    "name": ("姓名",),
    "id_card": ("身份证号码", "身份证号"),
    "amount": ("应发小计", "本月应发工资"),
}
MAX_PROFILES = 200
MAX_HEADER_ROW = 200
MAX_COLUMNS = 512


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(value or ""))).casefold()


def _digest(value: Any) -> str:
    return hashlib.sha256(json.dumps(value, ensure_ascii=False, sort_keys=True).encode()).hexdigest()


def fields_for(role: str) -> tuple[str, ...]:
    return ("name", "id_card") if role == "summary" else ("name", "id_card", "amount")


def _rows(ws, last: int) -> list[list[Any]]:
    return [list(row) for row in ws.iter_rows(
        min_row=1, max_row=last, max_col=min(ws.max_column or MAX_COLUMNS, MAX_COLUMNS), values_only=True,
    )]


def _auto_header(rows: list[list[Any]], role: str) -> int:
    # 先沿用身份证列定位方式；未知字段全部改名时，展示最可能的文字表头供人确认。
    identity_aliases = {normalize_header(x) for x in ALIASES["id_card"]}
    for index, row in enumerate(rows[:20], 1):
        if any(normalize_header(value) in identity_aliases for value in row):
            return index
    def score(row):
        return sum(isinstance(value, str) and bool(value.strip()) and not value.startswith("=") for value in row)
    return max(range(min(20, len(rows))), key=lambda index: score(rows[index]), default=0) + 1


def _describe(ws, rows: list[list[Any]], role: str, first: int, last: int) -> dict[str, Any]:
    if not 1 <= first <= last <= MAX_HEADER_ROW or last - first > 5:
        raise ValueError("请选择 1—200 行内的表头，连续表头最多 6 行")
    if last > len(rows):
        raise ValueError("所选表头超出工作表范围")
    columns = []
    width = max((len(row) for row in rows[first - 1:last]), default=0)
    for col in range(width):
        labels = []
        for row in rows[first - 1:last]:
            value = row[col] if col < len(row) else None
            text = str(value or "").strip()
            if text and (not labels or text != labels[-1]):
                labels.append(text)
        if not labels:
            continue
        label = " / ".join(labels)
        samples = []
        for row in rows[last:last + 5]:
            value = row[col] if col < len(row) else None
            if value is not None and str(value).strip():
                samples.append(str(value)[:80])
            if len(samples) >= 3:
                break
        columns.append({"column": col + 1, "label": label, "key": normalize_header(label),
                        "leaves": [normalize_header(x) for x in labels],
                        "display": f"{get_column_letter(col + 1)}列：{label}", "samples": "；".join(samples)})
    keys = [column["key"] for column in columns]
    # 顺序调整不改变模板身份；重名列仍需通过完整顺序指纹核对具体位置。
    order = _digest([(column["column"], column["key"]) for column in columns])
    duplicated = any(count > 1 for count in Counter(keys).values())
    key = _digest([role, normalize_header(ws.title), last - first, sorted(keys), order if duplicated else ""])
    selections = {}
    for field in fields_for(role):
        selections[field] = 0
        for alias in ALIASES[field]:
            candidates = [c for c in columns if normalize_header(alias) in c["leaves"]]
            if candidates:
                # 同一名称出现多次时，必须由用户选择实际业务列。
                if len(candidates) == 1:
                    selections[field] = candidates[0]["column"]
                break
    return {"key": key, "order": order, "role": role, "sheet": ws.title,
            "header_row": first, "header_bottom": last, "columns": columns,
            "selections": selections, "builtin_selections": dict(selections),
            "ready": all(selections.values()), "saved": False, "problem": "", "files": []}


def _apply_profile(group: dict[str, Any], profile: dict[str, Any]) -> bool:
    selections = {}
    for field in fields_for(group["role"]):
        locator = profile.get("fields", {}).get(field, {})
        candidates = [c for c in group["columns"] if c["key"] == locator.get("header")]
        if len(candidates) > 1 and profile.get("order") == group["order"]:
            candidates = [c for c in candidates if c["column"] == locator.get("column")]
        if len(candidates) != 1:
            group["problem"] = "保存的对应列无法唯一定位，请重新确认"
            group["ready"] = False
            return False
        selections[field] = candidates[0]["column"]
    if len(set(selections.values())) != len(selections):
        group["problem"] = "不同字段不能对应同一列，请重新选择"
        group["ready"] = False
        return False
    group.update(selections=selections, ready=True, saved=True)
    return True


def inspect_workbook(
    workbook, *, role: str = "detail", profiles: dict[str, Any] | None = None,
    hint: dict[str, Any] | None = None,
) -> dict[str, Any]:
    profiles = profiles or {}
    keyword = "汇总" if role == "summary" else "明细"
    names = workbook.sheetnames
    preferred = next((name for name in names if keyword in name), names[0])
    cache: dict[str, list[list[Any]]] = {}

    def describe(sheet: str, first: int = 0, bottom: int = 0):
        if sheet not in names:
            raise ValueError("所选工作表已不存在，请重新选择")
        last_needed = max(25, bottom + 5, first + 5)
        if last_needed > MAX_HEADER_ROW + 5:
            raise ValueError("表头行不能超过 200")
        rows = cache.get(sheet)
        if rows is None or len(rows) < last_needed:
            rows = _rows(workbook[sheet], last_needed)
            cache[sheet] = rows
        first = first or _auto_header(rows, role)
        group = _describe(workbook[sheet], rows, role, first, bottom or first)
        group["sheet_names"] = names
        return group

    if hint:
        group = describe(str(hint.get("sheet") or preferred), int(hint.get("header_row") or 0),
                         int(hint.get("header_bottom") or 0))
        if group["key"] in profiles:
            _apply_profile(group, profiles[group["key"]])
    else:
        candidates = []
        seen = set()
        # 使用已确认工作表/表头位置；规则仍须匹配本次真实表头指纹。
        for profile in list(profiles.values())[:MAX_PROFILES]:
            if not isinstance(profile, dict) or profile.get("role") != role or profile.get("sheet") not in names:
                continue
            spec = (profile["sheet"], int(profile.get("header_row") or 0), int(profile.get("header_bottom") or 0))
            if spec in seen:
                continue
            seen.add(spec)
            candidate = describe(*spec)
            if candidate["key"] in profiles:
                _apply_profile(candidate, profiles[candidate["key"]])
                candidates.append(candidate)
        group = describe(preferred)
        if group["key"] in profiles:
            _apply_profile(group, profiles[group["key"]])
            if not any(c["key"] == group["key"] for c in candidates):
                candidates.append(group)
        if len(candidates) == 1:
            group = candidates[0]
        elif len(candidates) > 1:
            group.update(ready=False, problem="多张工作表匹配已保存设置，请选择本次工资明细")
        elif keyword not in preferred and role == "detail":
            group.update(ready=False, problem="未找到明细工作表，请确认工作表与对应列")
        # 内置字段已明确识别时，旧模板的变化不应再次要求用户确认。
        elif not candidates and not group["ready"]:
            current = Counter(c["key"] for c in group["columns"])
            for profile_key, profile in profiles.items():
                if not isinstance(profile, dict) or profile.get("role") != role or profile.get("sheet") != group["sheet"]:
                    continue
                previous = Counter(profile.get("headers") or [])
                selected = {item.get("header") for item in profile.get("fields", {}).values() if isinstance(item, dict)}
                if previous and selected and len(selected & set(current)) >= max(1, len(selected) - 1):
                    group.update(ready=False, problem="模板结构与已保存设置有变化，请重新确认对应列")
                    group.setdefault("related_profile_keys", []).append(profile_key)
                    break
    if not group["ready"] and not group["problem"]:
        missing = [FIELD_LABELS[field] for field, col in group["selections"].items() if not col]
        group["problem"] = "请选择对应列：" + "、".join(missing)
    group["hinted"] = bool(hint)
    return group


def profile_from_selection(group: dict[str, Any], selections: dict[str, Any]) -> dict[str, Any]:
    fields = {}
    for field in fields_for(group["role"]):
        try:
            col = int(selections.get(field) or 0)
        except (TypeError, ValueError):
            col = 0
        column = next((c for c in group["columns"] if c["column"] == col), None)
        if column is None:
            raise ValueError(f"请选择{FIELD_LABELS[field]}对应的原表列")
        fields[field] = {"header": column["key"], "column": col}
    if len({v["column"] for v in fields.values()}) != len(fields):
        raise ValueError("姓名、身份证号码和应发工资必须对应不同的列")
    return {"role": group["role"], "sheet": group["sheet"], "header_row": group["header_row"],
            "header_bottom": group["header_bottom"], "order": group["order"], "fields": fields,
            "headers": [c["key"] for c in group["columns"]]}
