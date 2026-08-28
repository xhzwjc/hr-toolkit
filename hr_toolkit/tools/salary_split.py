from __future__ import annotations

import json
import re
import tempfile
from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..common.excel import (
    RowSnapshot,
    apply_row_snapshot,
    snapshot_row,
    unmerge_ranges_from_row,
)
from ..common.excel_compat import SUPPORTED_EXCEL_SUFFIXES, ensure_xlsx_workbook
from ..common.filenames import safe_filename


TOOL_NAME = "需求4-工资表按入职公司拆分"
DETAIL_SHEET_KEYWORD = "明细"
SUMMARY_SHEET_KEYWORD = "汇总"
HEADER_COMPANY_SYNONYMS = ("入职公司", "归属公司", "所属公司", "签约公司", "公司")
HEADER_PROJECT_SYNONYMS = ("项目", "项目名称", "所属项目", "区域", "部门")
HEADER_NAME_SYNONYMS = ("姓名", "员工姓名", "人员姓名")
HEADER_ID_CARD_SYNONYMS = ("身份证号码", "身份证号", "证件号码", "身份证")
HEADER_SEQ_SYNONYMS = ("序号", "行号", "No", "NO", "NO.", "no")
TRAILING_METADATA_EXACT = {
    "区域",
    "所属区域",
    "项目",
    "项目名称",
    "所属项目",
    "岗位",
    "岗位名称",
    "职务",
    "部门",
    "入职公司",
    "归属公司",
    "所属公司",
    "签约公司",
    "公司",
    "卡号",
    "银行卡号",
    "银行账号",
    "账号",
    "开户行",
    "开户银行",
    "开户支行",
    "支行",
    "手机号",
    "手机号码",
    "电话",
    "联系电话",
    "备注",
}


@dataclass
class EmployeeRow:
    source_row: int
    company: str
    section: str
    category: str | None
    snapshot: RowSnapshot


@dataclass(frozen=True)
class LeafSection:
    label: str
    source_start_row: int
    source_end_row: int
    subtotal_row: int | None
    subtotal_snapshot: RowSnapshot | None
    category_label: str | None = None
    merged_ranges: tuple[tuple[int, int], ...] = ()


@dataclass(frozen=True)
class CategoryGroup:
    label: str
    total_row: int
    total_snapshot: RowSnapshot
    leaf_labels: tuple[str, ...]
    merged_ranges: tuple[tuple[int, int], ...] = ()


@dataclass(frozen=True)
class GrandTotalInfo:
    label: str
    total_row: int
    total_snapshot: RowSnapshot
    merged_ranges: tuple[tuple[int, int], ...] = ()


@dataclass(frozen=True)
class DetailHierarchy:
    leaves: list[LeafSection]
    categories: list[CategoryGroup]
    grand_total: GrandTotalInfo | None


@dataclass(frozen=True)
class RenderedLeaf:
    label: str
    source_subtotal_row: int | None
    new_subtotal_row: int | None
    data_start_row: int
    data_end_row: int
    employee_count: int


@dataclass(frozen=True)
class RenderedGroup:
    label: str
    source_total_row: int
    new_total_row: int
    rendered_leaves: list[RenderedLeaf]


@dataclass(frozen=True)
class RebuiltDetailResult:
    rendered_leaves: list[RenderedLeaf]
    rendered_groups: list[RenderedGroup]
    rendered_grand_total_row: int | None
    row_map: dict[int, int]


@dataclass
class CompanyOutput:
    company: str
    employee_count: int
    sections: list[str]
    file_path: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "company": self.company,
            "employee_count": self.employee_count,
            "sections": self.sections,
            "projects": self.sections,
            "file_path": self.file_path,
        }


@dataclass
class SalarySplitResult:
    input_path: Path
    output_dir: Path
    dry_run: bool
    outputs: list[CompanyOutput] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "tool_name": TOOL_NAME,
            "input_path": str(self.input_path),
            "output_dir": str(self.output_dir),
            "dry_run": self.dry_run,
            "company_count": len(self.outputs),
            "employee_count": sum(item.employee_count for item in self.outputs),
            "outputs": [item.to_dict() for item in self.outputs],
        }


@dataclass(frozen=True)
class SalarySheetLayout:
    detail_sheet_name: str
    summary_sheet_name: str
    header_row: int
    data_start_row: int
    max_column: int
    seq_col: int
    name_col: int
    id_card_col: int
    project_col: int | None
    company_col: int
    amount_end_col: int


def split_salary_by_company(
    input_path: str | Path,
    output_dir: str | Path,
    *,
    dry_run: bool = False,
    write_manifest: bool = False,
    cancelled: Callable[[], bool] | None = None,
) -> SalarySplitResult:
    """Split one salary workbook into one workbook per hiring company."""
    _check_cancelled(cancelled)
    input_path = Path(input_path).expanduser().resolve()
    output_dir = Path(output_dir).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")
    if input_path.suffix.lower() not in SUPPORTED_EXCEL_SUFFIXES:
        raise ValueError("当前工资拆分工具仅支持 .xlsx 或 .xls 文件")

    with tempfile.TemporaryDirectory(prefix="hr_salary_split_") as temp_root:
        working_input_path = ensure_xlsx_workbook(input_path, Path(temp_root))
        workbook = load_workbook(working_input_path, data_only=False)
        try:
            layout = _detect_layout(workbook)
            detail_ws = workbook[layout.detail_sheet_name]
            hierarchy = _detect_detail_hierarchy(detail_ws, layout)
            employees = _collect_employees(
                detail_ws,
                layout,
                hierarchy,
                cancelled=cancelled,
            )
            groups = _group_by_company(employees)
        finally:
            workbook.close()

        result = SalarySplitResult(input_path=input_path, output_dir=output_dir, dry_run=dry_run)
        for company, rows in groups.items():
            _check_cancelled(cancelled)
            result.outputs.append(
                CompanyOutput(
                    company=company,
                    employee_count=len(rows),
                    sections=list(_group_sections(rows).keys()),
                )
            )

        if dry_run:
            return result

        output_dir.mkdir(parents=True, exist_ok=True)
        for company_output in result.outputs:
            _check_cancelled(cancelled)
            rows = groups[company_output.company]
            output_path = output_dir / f"{safe_filename(company_output.company)}-工资表.xlsx"
            _write_company_workbook(
                working_input_path,
                layout,
                hierarchy,
                rows,
                output_path,
                cancelled=cancelled,
            )
            company_output.file_path = str(output_path)

        if write_manifest:
            manifest_path = output_dir / "_salary_split_manifest.json"
            manifest_path.write_text(
                json.dumps(result.to_dict(), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
    return result


def _check_cancelled(cancelled: Callable[[], bool] | None) -> None:
    if cancelled is not None and cancelled():
        raise RuntimeError("本次处理已停止。")


def _find_detail_max_col(ws: Worksheet, header_row: int) -> int:
    max_col = 1
    worksheet_max_column = ws.max_column
    for r in range(max(1, header_row - 2), min(ws.max_row, header_row + 3) + 1):
        for c in range(1, worksheet_max_column + 1):
            if ws.cell(r, c).value is not None:
                max_col = max(max_col, c)
    return max_col


def _detect_layout(workbook) -> SalarySheetLayout:
    detail_sheet_name = _find_sheet_name(workbook.sheetnames, DETAIL_SHEET_KEYWORD)
    summary_sheet_name = _find_sheet_name(workbook.sheetnames, SUMMARY_SHEET_KEYWORD)
    detail_ws = workbook[detail_sheet_name]

    header_row = _find_header_row(detail_ws, HEADER_COMPANY_SYNONYMS)
    headers = _read_headers(detail_ws, header_row)

    company_col = _match_header_col(headers, HEADER_COMPANY_SYNONYMS)
    if company_col is None:
        raise ValueError("明细表缺少“入职公司”字段")

    name_col = _match_header_col(headers, HEADER_NAME_SYNONYMS)
    if name_col is None:
        raise ValueError("明细表缺少“姓名”字段")

    id_card_col = _match_header_col(headers, HEADER_ID_CARD_SYNONYMS)
    if id_card_col is None:
        raise ValueError("明细表缺少“身份证号码”字段")

    seq_col = _match_header_col(headers, HEADER_SEQ_SYNONYMS) or 1
    project_col = _match_header_col(headers, HEADER_PROJECT_SYNONYMS)
    detail_max_col = _find_detail_max_col(detail_ws, header_row)

    return SalarySheetLayout(
        detail_sheet_name=detail_sheet_name,
        summary_sheet_name=summary_sheet_name,
        header_row=header_row,
        data_start_row=_find_data_start_row(detail_ws, header_row),
        max_column=detail_max_col,
        seq_col=seq_col,
        name_col=name_col,
        id_card_col=id_card_col,
        project_col=project_col,
        company_col=company_col,
        amount_end_col=_find_amount_end_col(headers, detail_max_col),
    )


def _find_sheet_name(sheetnames: list[str], keyword: str) -> str:
    for sheetname in sheetnames:
        if keyword in sheetname:
            return sheetname
    raise ValueError(f"未找到包含“{keyword}”的工作表")


def _find_header_row(ws: Worksheet, required_headers: tuple[str, ...]) -> int:
    max_column = ws.max_column
    for row_index in range(1, min(ws.max_row, 20) + 1):
        values = [
            str(ws.cell(row_index, col).value or "").strip()
            for col in range(1, max_column + 1)
        ]
        if any(header in values for header in required_headers):
            return row_index
    raise ValueError(f"未在明细表前 20 行找到字段：{required_headers[0]}")


def _read_headers(ws: Worksheet, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_index in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_index).value
        if value is None:
            continue
        text = str(value).strip()
        if text and text not in headers:
            headers[text] = col_index
    return headers


def _match_header_col(headers: dict[str, int], synonyms: tuple[str, ...]) -> int | None:
    for syn in synonyms:
        if syn in headers:
            return headers[syn]
    for syn in synonyms:
        for header, col in headers.items():
            if syn in header:
                return col
    return None


def _find_amount_end_col(headers: dict[str, int], max_column: int) -> int:
    trailing_cols = [
        col_index
        for header, col_index in headers.items()
        if col_index >= 5 and header.strip() in TRAILING_METADATA_EXACT
    ]
    if not trailing_cols:
        return max_column
    return min(trailing_cols) - 1


def _find_data_start_row(ws: Worksheet, header_row: int) -> int:
    bottom = header_row
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= header_row <= merged_range.max_row:
            bottom = max(bottom, merged_range.max_row)
    return bottom + 1


def _is_grand_total_label(label: str) -> bool:
    cleaned = label.strip()
    return "总计" in cleaned or "全公司" in cleaned


def _is_subtotal_label(label: str) -> bool:
    cleaned = label.strip()
    return "小计" in cleaned


def _is_category_total_label(label: str) -> bool:
    cleaned = label.strip()
    return "合计" in cleaned or "大计" in cleaned or "专业合计" in cleaned


def _get_row_merged_ranges(ws: Worksheet, row_index: int) -> tuple[tuple[int, int], ...]:
    ranges = []
    for m in ws.merged_cells.ranges:
        if m.min_row <= row_index <= m.max_row:
            ranges.append((m.min_col, m.max_col))
    return tuple(sorted(ranges))


def _detect_detail_hierarchy(ws: Worksheet, layout: SalarySheetLayout) -> DetailHierarchy:
    """Classify rows in detail sheet into hierarchical leaf sections and totals."""
    grand_total_row: int | None = None
    grand_total_info: GrandTotalInfo | None = None

    for row_index in range(ws.max_row, layout.data_start_row - 1, -1):
        label = _cell_text(ws, row_index, 1)
        if label and (_is_grand_total_label(label) or "合计" in label or "汇总" in label):
            grand_total_row = row_index
            grand_total_info = GrandTotalInfo(
                label=label,
                total_row=row_index,
                total_snapshot=snapshot_row(ws, row_index, layout.max_column),
                merged_ranges=_get_row_merged_ranges(ws, row_index),
            )
            break

    end_row = (grand_total_row - 1) if grand_total_row is not None else ws.max_row

    leaves: list[LeafSection] = []
    categories: list[CategoryGroup] = []
    current_category_leaves: list[LeafSection] = []
    section_start = layout.data_start_row

    for row_index in range(layout.data_start_row, end_row + 1):
        label = _cell_text(ws, row_index, 1)
        if not label:
            continue

        if _is_subtotal_label(label):
            leaf = LeafSection(
                label=label,
                source_start_row=section_start,
                source_end_row=row_index - 1,
                subtotal_row=row_index,
                subtotal_snapshot=snapshot_row(ws, row_index, layout.max_column),
                merged_ranges=_get_row_merged_ranges(ws, row_index),
            )
            leaves.append(leaf)
            current_category_leaves.append(leaf)
            section_start = row_index + 1
        elif _is_category_total_label(label):
            if current_category_leaves:
                cat = CategoryGroup(
                    label=label,
                    total_row=row_index,
                    total_snapshot=snapshot_row(ws, row_index, layout.max_column),
                    leaf_labels=tuple(item.label for item in current_category_leaves),
                    merged_ranges=_get_row_merged_ranges(ws, row_index),
                )
                categories.append(cat)
                current_category_leaves = []
                section_start = row_index + 1
            else:
                leaf = LeafSection(
                    label=label,
                    source_start_row=section_start,
                    source_end_row=row_index - 1,
                    subtotal_row=row_index,
                    subtotal_snapshot=snapshot_row(ws, row_index, layout.max_column),
                    merged_ranges=_get_row_merged_ranges(ws, row_index),
                )
                leaves.append(leaf)
                section_start = row_index + 1

    if section_start <= end_row:
        has_any_data = False
        for r in range(section_start, end_row + 1):
            if _cell_text(ws, r, layout.company_col) or _cell_text(ws, r, layout.name_col):
                has_any_data = True
                break
        if has_any_data:
            leaf = LeafSection(
                label="全体员工" if not leaves else f"分段{len(leaves) + 1}",
                source_start_row=section_start,
                source_end_row=end_row,
                subtotal_row=None,
                subtotal_snapshot=None,
            )
            leaves.append(leaf)

    if not leaves and grand_total_info is not None:
        leaf = LeafSection(
            label=grand_total_info.label,
            source_start_row=layout.data_start_row,
            source_end_row=end_row,
            subtotal_row=None,
            subtotal_snapshot=None,
        )
        leaves.append(leaf)

    if not leaves:
        raise ValueError("未识别到明细表中的数据或小计分段")

    return DetailHierarchy(leaves=leaves, categories=categories, grand_total=grand_total_info)



def _collect_employees(
    ws: Worksheet,
    layout: SalarySheetLayout,
    hierarchy: DetailHierarchy,
    *,
    cancelled: Callable[[], bool] | None = None,
) -> list[EmployeeRow]:
    employees: list[EmployeeRow] = []
    leaf_to_cat: dict[str, str] = {}
    for cat in hierarchy.categories:
        for leaf_label in cat.leaf_labels:
            leaf_to_cat[leaf_label] = cat.label

    for row_index in range(layout.data_start_row, ws.max_row + 1):
        if row_index % 500 == 0:
            _check_cancelled(cancelled)
        company = _cell_text(ws, row_index, layout.company_col)
        name = _cell_text(ws, row_index, layout.name_col)
        id_card = _cell_text(ws, row_index, layout.id_card_col)
        if not company:
            continue
        if not name and not id_card:
            continue
        leaf = _find_leaf_for_row(row_index, hierarchy.leaves)
        employees.append(
            EmployeeRow(
                source_row=row_index,
                company=company,
                section=leaf.label,
                category=leaf_to_cat.get(leaf.label),
                snapshot=snapshot_row(ws, row_index, layout.max_column),
            )
        )
    if not employees:
        raise ValueError("未识别到可拆分的员工数据，请检查明细表的“入职公司”列")
    return employees


def _find_leaf_for_row(row_index: int, leaves: list[LeafSection]) -> LeafSection:
    for leaf in leaves:
        if leaf.source_start_row <= row_index <= leaf.source_end_row:
            return leaf
    for leaf in leaves:
        if leaf.subtotal_row is not None and row_index <= leaf.subtotal_row:
            return leaf
    return leaves[-1]


def _group_by_company(employees: list[EmployeeRow]) -> OrderedDict[str, list[EmployeeRow]]:
    groups: OrderedDict[str, list[EmployeeRow]] = OrderedDict()
    for employee in employees:
        groups.setdefault(employee.company, []).append(employee)
    return groups


def _group_sections(rows: list[EmployeeRow]) -> OrderedDict[str, list[EmployeeRow]]:
    sections: OrderedDict[str, list[EmployeeRow]] = OrderedDict()
    for row in rows:
        sections.setdefault(row.section, []).append(row)
    return sections


def _apply_row_merged_ranges(
    ws: Worksheet,
    target_row: int,
    source_merged_ranges: tuple[tuple[int, int], ...],
    default_end_col: int,
) -> None:
    if source_merged_ranges:
        for min_col, max_col in source_merged_ranges:
            ws.merge_cells(
                start_row=target_row,
                start_column=min_col,
                end_row=target_row,
                end_column=max_col,
            )
    else:
        ws.merge_cells(
            start_row=target_row,
            start_column=1,
            end_row=target_row,
            end_column=default_end_col,
        )


def _write_company_workbook(
    input_path: Path,
    layout: SalarySheetLayout,
    hierarchy: DetailHierarchy,
    rows: list[EmployeeRow],
    output_path: Path,
    *,
    cancelled: Callable[[], bool] | None = None,
) -> None:
    _check_cancelled(cancelled)
    workbook = load_workbook(input_path, data_only=False)
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

        detail_ws = workbook[layout.detail_sheet_name]
        summary_ws = workbook[layout.summary_sheet_name]

        rebuilt = _rebuild_detail_sheet(
            detail_ws,
            layout,
            hierarchy,
            rows,
            cancelled=cancelled,
        )
        _check_cancelled(cancelled)
        _rebuild_summary_sheet(summary_ws, layout, hierarchy, rebuilt)
        _check_cancelled(cancelled)
        workbook.save(output_path)
    finally:
        workbook.close()


def _rebuild_detail_sheet(
    ws: Worksheet,
    layout: SalarySheetLayout,
    hierarchy: DetailHierarchy,
    rows: list[EmployeeRow],
    *,
    cancelled: Callable[[], bool] | None = None,
) -> RebuiltDetailResult:
    unmerge_ranges_from_row(ws, layout.data_start_row)
    ws.delete_rows(layout.data_start_row, ws.max_row - layout.data_start_row + 1)

    current_row = layout.data_start_row
    employee_seq = 1
    rendered_leaves: list[RenderedLeaf] = []
    rendered_groups: list[RenderedGroup] = []
    row_map: dict[int, int] = {}

    rows_by_leaf = _group_sections(rows)
    leaf_by_label = {leaf.label: leaf for leaf in hierarchy.leaves}
    default_merge_end = max(1, layout.id_card_col - 1)

    if hierarchy.categories:
        leaves_in_categories = set()
        for cat in hierarchy.categories:
            cat_rendered_leaves: list[RenderedLeaf] = []
            for leaf_label in cat.leaf_labels:
                leaves_in_categories.add(leaf_label)
                leaf = leaf_by_label.get(leaf_label)
                if leaf is None:
                    continue
                section_rows = rows_by_leaf.get(leaf.label, [])
                if not section_rows:
                    continue
                data_start_row = current_row
                for emp in section_rows:
                    if employee_seq % 500 == 0:
                        _check_cancelled(cancelled)
                    apply_row_snapshot(ws, current_row, emp.snapshot, translate_formulas=True)
                    ws.cell(current_row, layout.seq_col).value = employee_seq
                    current_row += 1
                    employee_seq += 1
                data_end_row = current_row - 1
                if leaf.subtotal_snapshot is not None:
                    subtotal_row = current_row
                    apply_row_snapshot(ws, subtotal_row, leaf.subtotal_snapshot, translate_formulas=False)
                    ws.cell(subtotal_row, 1).value = leaf.label
                    _write_detail_section_total_formulas(ws, layout, subtotal_row, data_start_row, data_end_row)
                    _apply_row_merged_ranges(ws, subtotal_row, leaf.merged_ranges, default_merge_end)
                    if leaf.subtotal_row is not None:
                        row_map[leaf.subtotal_row] = subtotal_row
                    current_row += 1
                else:
                    subtotal_row = None
                r_leaf = RenderedLeaf(
                    label=leaf.label,
                    source_subtotal_row=leaf.subtotal_row,
                    new_subtotal_row=subtotal_row,
                    data_start_row=data_start_row,
                    data_end_row=data_end_row,
                    employee_count=len(section_rows),
                )
                rendered_leaves.append(r_leaf)
                cat_rendered_leaves.append(r_leaf)

            if cat_rendered_leaves:
                group_total_row = current_row
                apply_row_snapshot(ws, group_total_row, cat.total_snapshot, translate_formulas=False)
                ws.cell(group_total_row, 1).value = cat.label
                sub_rows = [l.new_subtotal_row for l in cat_rendered_leaves if l.new_subtotal_row is not None]
                _write_detail_group_total_formulas(ws, layout, group_total_row, sub_rows)
                _apply_row_merged_ranges(ws, group_total_row, cat.merged_ranges, default_merge_end)
                row_map[cat.total_row] = group_total_row
                rendered_groups.append(
                    RenderedGroup(
                        label=cat.label,
                        source_total_row=cat.total_row,
                        new_total_row=group_total_row,
                        rendered_leaves=cat_rendered_leaves,
                    )
                )
                current_row += 1

        for leaf in hierarchy.leaves:
            if leaf.label in leaves_in_categories:
                continue
            section_rows = rows_by_leaf.get(leaf.label, [])
            if not section_rows:
                continue
            data_start_row = current_row
            for emp in section_rows:
                if employee_seq % 500 == 0:
                    _check_cancelled(cancelled)
                apply_row_snapshot(ws, current_row, emp.snapshot, translate_formulas=True)
                ws.cell(current_row, layout.seq_col).value = employee_seq
                current_row += 1
                employee_seq += 1
            data_end_row = current_row - 1
            if leaf.subtotal_snapshot is not None:
                subtotal_row = current_row
                apply_row_snapshot(ws, subtotal_row, leaf.subtotal_snapshot, translate_formulas=False)
                ws.cell(subtotal_row, 1).value = leaf.label
                _write_detail_section_total_formulas(ws, layout, subtotal_row, data_start_row, data_end_row)
                _apply_row_merged_ranges(ws, subtotal_row, leaf.merged_ranges, default_merge_end)
                if leaf.subtotal_row is not None:
                    row_map[leaf.subtotal_row] = subtotal_row
                current_row += 1
            else:
                subtotal_row = None
            r_leaf = RenderedLeaf(
                label=leaf.label,
                source_subtotal_row=leaf.subtotal_row,
                new_subtotal_row=subtotal_row,
                data_start_row=data_start_row,
                data_end_row=data_end_row,
                employee_count=len(section_rows),
            )
            rendered_leaves.append(r_leaf)
    else:
        for leaf in hierarchy.leaves:
            section_rows = rows_by_leaf.get(leaf.label, [])
            if not section_rows:
                continue
            data_start_row = current_row
            for emp in section_rows:
                if employee_seq % 500 == 0:
                    _check_cancelled(cancelled)
                apply_row_snapshot(ws, current_row, emp.snapshot, translate_formulas=True)
                ws.cell(current_row, layout.seq_col).value = employee_seq
                current_row += 1
                employee_seq += 1
            data_end_row = current_row - 1
            if leaf.subtotal_snapshot is not None:
                subtotal_row = current_row
                apply_row_snapshot(ws, subtotal_row, leaf.subtotal_snapshot, translate_formulas=False)
                ws.cell(subtotal_row, 1).value = leaf.label
                _write_detail_section_total_formulas(ws, layout, subtotal_row, data_start_row, data_end_row)
                _apply_row_merged_ranges(ws, subtotal_row, leaf.merged_ranges, default_merge_end)
                if leaf.subtotal_row is not None:
                    row_map[leaf.subtotal_row] = subtotal_row
                current_row += 1
            else:
                subtotal_row = None
            r_leaf = RenderedLeaf(
                label=leaf.label,
                source_subtotal_row=leaf.subtotal_row,
                new_subtotal_row=subtotal_row,
                data_start_row=data_start_row,
                data_end_row=data_end_row,
                employee_count=len(section_rows),
            )
            rendered_leaves.append(r_leaf)

    rendered_grand_total_row: int | None = None
    if hierarchy.grand_total is not None:
        grand_total_row = current_row
        apply_row_snapshot(ws, grand_total_row, hierarchy.grand_total.total_snapshot, translate_formulas=False)
        ws.cell(grand_total_row, 1).value = hierarchy.grand_total.label

        if rendered_groups:
            rows_to_sum = [g.new_total_row for g in rendered_groups]
            _write_detail_group_total_formulas(ws, layout, grand_total_row, rows_to_sum)
        elif any(l.new_subtotal_row is not None for l in rendered_leaves):
            rows_to_sum = [l.new_subtotal_row for l in rendered_leaves if l.new_subtotal_row is not None]
            _write_detail_group_total_formulas(ws, layout, grand_total_row, rows_to_sum)
        elif rendered_leaves:
            first_data = min(l.data_start_row for l in rendered_leaves)
            last_data = max(l.data_end_row for l in rendered_leaves)
            _write_detail_section_total_formulas(ws, layout, grand_total_row, first_data, last_data)

        _apply_row_merged_ranges(ws, grand_total_row, hierarchy.grand_total.merged_ranges, default_merge_end)
        row_map[hierarchy.grand_total.total_row] = grand_total_row
        rendered_grand_total_row = grand_total_row
        current_row += 1

    return RebuiltDetailResult(
        rendered_leaves=rendered_leaves,
        rendered_groups=rendered_groups,
        rendered_grand_total_row=rendered_grand_total_row,
        row_map=row_map,
    )


def _write_detail_section_total_formulas(
    ws: Worksheet,
    layout: SalarySheetLayout,
    subtotal_row: int,
    first_data_row: int,
    last_data_row: int,
) -> None:
    for col_index in range(4, layout.max_column + 1):
        col_letter = get_column_letter(col_index)
        cell = ws.cell(subtotal_row, col_index)
        if isinstance(cell, MergedCell):
            continue
        if col_index == layout.id_card_col:
            cell.value = None
        elif col_index <= layout.amount_end_col:
            if last_data_row >= first_data_row:
                cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
            else:
                cell.value = None
        else:
            cell.value = None


def _write_detail_group_total_formulas(
    ws: Worksheet,
    layout: SalarySheetLayout,
    target_row: int,
    sub_rows: list[int],
) -> None:
    for col_index in range(4, layout.max_column + 1):
        col_letter = get_column_letter(col_index)
        cell = ws.cell(target_row, col_index)
        if isinstance(cell, MergedCell):
            continue
        if col_index == layout.id_card_col:
            cell.value = None
        elif col_index <= layout.amount_end_col:
            if len(sub_rows) > 1:
                cell.value = "=" + "+".join(f"{col_letter}{r}" for r in sub_rows)
            elif len(sub_rows) == 1:
                cell.value = f"={col_letter}{sub_rows[0]}"
            else:
                cell.value = None
        else:
            cell.value = None


def _find_summary_max_col(ws: Worksheet, start_row: int, total_template_row: int) -> int:
    max_col = 1
    worksheet_max_column = ws.max_column
    for r in range(1, start_row):
        for c in range(1, worksheet_max_column + 1):
            if ws.cell(r, c).value is not None:
                max_col = max(max_col, c)
    for r in range(start_row, min(ws.max_row, total_template_row + 1) + 1):
        for c in range(1, worksheet_max_column + 1):
            if ws.cell(r, c).value is not None:
                max_col = max(max_col, c)
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= total_template_row + 2:
            max_col = max(max_col, merged_range.max_col)
    return max_col


def _rebuild_summary_sheet(
    ws: Worksheet,
    layout: SalarySheetLayout,
    hierarchy: DetailHierarchy,
    rebuilt: RebuiltDetailResult,
) -> None:
    start_row = 6
    for r in range(1, min(ws.max_row, 15) + 1):
        c1 = str(ws.cell(r, 1).value or "").strip()
        if "项目名称" in c1:
            start_row = r + 1
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row <= r <= merged_range.max_row:
                    start_row = max(start_row, merged_range.max_row + 1)
            break

    total_template_row = _find_summary_total_row(ws, start_row)
    sign_template_row = total_template_row + 1
    summary_max_col = _find_summary_max_col(ws, start_row, total_template_row)

    template_rows: list[tuple[int, str, RowSnapshot, list[Any]]] = []
    for r in range(start_row, total_template_row):
        lbl = _cell_text(ws, r, 1)
        snap = snapshot_row(ws, r, summary_max_col)
        raw_vals = [ws.cell(r, c).value for c in range(1, summary_max_col + 1)]
        template_rows.append((r, lbl, snap, raw_vals))

    total_snapshot = snapshot_row(ws, total_template_row, summary_max_col)
    total_label = _cell_text(ws, total_template_row, 1) or "合计"

    sign_snapshot: RowSnapshot | None = None
    signature_text: Any = None
    if sign_template_row <= ws.max_row:
        sign_snapshot = snapshot_row(ws, sign_template_row, summary_max_col)
        signature_text = ws.cell(sign_template_row, 1).value

    unmerge_ranges_from_row(ws, start_row)
    ws.delete_rows(start_row, ws.max_row - start_row + 1)

    rendered_leaf_map = {l.label: l for l in rebuilt.rendered_leaves}
    rendered_group_map = {g.label: g for g in rebuilt.rendered_groups}

    source_items: list[tuple[str, Any]] = []
    leaf_by_label = {l.label: l for l in hierarchy.leaves}
    if hierarchy.categories:
        leaves_in_cat = set()
        for cat in hierarchy.categories:
            for l_lbl in cat.leaf_labels:
                leaves_in_cat.add(l_lbl)
                leaf = leaf_by_label.get(l_lbl)
                if leaf is not None:
                    source_items.append(("leaf", leaf))
            source_items.append(("group", cat))
        for leaf in hierarchy.leaves:
            if leaf.label not in leaves_in_cat:
                source_items.append(("leaf", leaf))
    else:
        for leaf in hierarchy.leaves:
            source_items.append(("leaf", leaf))

    template_item_map: dict[int, tuple[str, Any]] = {}
    for idx, (src_r, lbl, _snap, raw_vals) in enumerate(template_rows):
        matched_item = None
        for val in raw_vals:
            if isinstance(val, str) and ("!" in val or "明细" in val):
                for item_kind, item in source_items:
                    target_r = item.subtotal_row if item_kind == "leaf" else item.total_row
                    if target_r is not None and (f"!P{target_r}" in val or f"!U{target_r}" in val or f"!{target_r}" in val or f"{target_r}" in val):
                        matched_item = (item_kind, item)
                        break
            if matched_item is not None:
                break

        if matched_item is None and lbl:
            for item_kind, item in source_items:
                if item.label and (item.label in lbl or lbl in item.label):
                    matched_item = (item_kind, item)
                    break

        if matched_item is None and len(template_rows) == len(source_items):
            matched_item = source_items[idx]

        if matched_item is not None:
            template_item_map[src_r] = matched_item

    current_row = start_row
    rendered_sub_rows_in_summary: list[int] = []
    rendered_group_rows_in_summary: list[int] = []

    for src_r, lbl, snap, _raw_vals in template_rows:
        mapped = template_item_map.get(src_r)
        target_leaf: RenderedLeaf | None = None
        target_group: RenderedGroup | None = None
        is_rendered = False

        if mapped is not None:
            item_kind, item = mapped
            if item_kind == "leaf" and item.label in rendered_leaf_map:
                is_rendered = True
                target_leaf = rendered_leaf_map[item.label]
            elif item_kind == "group" and item.label in rendered_group_map:
                is_rendered = True
                target_group = rendered_group_map[item.label]
        else:
            if lbl in rendered_leaf_map:
                is_rendered = True
                target_leaf = rendered_leaf_map[lbl]
            elif lbl in rendered_group_map:
                is_rendered = True
                target_group = rendered_group_map[lbl]
            else:
                for l_lbl, r_leaf in rendered_leaf_map.items():
                    if l_lbl and (l_lbl in lbl or lbl in l_lbl):
                        is_rendered = True
                        target_leaf = r_leaf
                        break
                if not is_rendered:
                    for g_lbl, r_group in rendered_group_map.items():
                        if g_lbl and (g_lbl in lbl or lbl in g_lbl):
                            is_rendered = True
                            target_group = r_group
                            break

        if not is_rendered:
            continue

        apply_row_snapshot(ws, current_row, snap, translate_formulas=False)
        ws.cell(current_row, 1).value = lbl

        is_group_row = (target_group is not None) or _is_category_total_label(lbl)
        if not is_group_row:
            rendered_sub_rows_in_summary.append(current_row)
            _translate_summary_leaf_formulas(
                ws,
                current_row,
                src_r,
                layout.detail_sheet_name,
                target_leaf,
                rebuilt.row_map,
                summary_max_col,
            )
        else:
            rendered_group_rows_in_summary.append(current_row)
            _translate_summary_group_formulas(
                ws,
                current_row,
                src_r,
                rendered_sub_rows_in_summary,
                rendered_group_rows_in_summary,
                summary_max_col,
            )
        current_row += 1

    total_row = current_row
    apply_row_snapshot(ws, total_row, total_snapshot, translate_formulas=False)
    ws.cell(total_row, 1).value = total_label
    _write_summary_total_formulas(
        ws,
        start_row,
        total_row,
        rendered_group_rows_in_summary or rendered_sub_rows_in_summary,
        summary_max_col,
    )

    if sign_snapshot is not None:
        sign_row = total_row + 1
        apply_row_snapshot(ws, sign_row, sign_snapshot, translate_formulas=False)
        ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=summary_max_col)
        ws.cell(sign_row, 1).value = signature_text


def _find_summary_total_row(ws: Worksheet, start_row: int) -> int:
    for row_index in range(ws.max_row, start_row - 1, -1):
        label = _cell_text(ws, row_index, 1)
        if label and (_is_grand_total_label(label) or "合计" in label or "汇总" in label or "共计" in label):
            return row_index
    for row_index in range(start_row, ws.max_row + 1):
        label = _cell_text(ws, row_index, 1)
        if "合计" in label or "总计" in label:
            return row_index
    return ws.max_row


def _translate_sheet_references(formula: str, target_sheet: str, row_map: dict[int, int]) -> str:
    target_clean = target_sheet.replace("'", "").strip()
    pattern = re.compile(
        r"(?:'([^']+)'|([A-Za-z0-9_\u4e00-\u9fa5\(\)\uff08\uff09-]+))!\$?([A-Za-z]{1,3})\$?(\d+)"
    )

    def repl(m: re.Match) -> str:
        s1, s2, col, row_str = m.groups()
        sheet = (s1 or s2 or "").replace("'", "").strip()
        if sheet.lower() == target_clean.lower():
            old_r = int(row_str)
            new_r = row_map.get(old_r, old_r)
            return f"'{target_sheet}'!{col}{new_r}"
        return m.group(0)

    return pattern.sub(repl, formula)


def _translate_summary_leaf_formulas(
    ws: Worksheet,
    current_row: int,
    src_template_row: int,
    detail_sheet_name: str,
    target_leaf: RenderedLeaf | None,
    row_map: dict[int, int],
    max_column: int,
) -> None:
    detail_ref = _formula_sheet_name(detail_sheet_name)
    subtotal_row = target_leaf.new_subtotal_row if target_leaf is not None else None

    has_any_formula = any(
        isinstance(ws.cell(current_row, c).value, str) and ws.cell(current_row, c).value.startswith("=")
        for c in range(2, max_column + 1)
    )

    if not has_any_formula and subtotal_row is not None and target_leaf is not None:
        defaults = {
            2: (
                f"=COUNT({detail_ref}!$A${target_leaf.data_start_row}:$A${target_leaf.data_end_row})"
                if target_leaf.data_end_row >= target_leaf.data_start_row
                else str(target_leaf.employee_count)
            ),
            3: f"={detail_ref}!P{subtotal_row}",
            4: f"={detail_ref}!U{subtotal_row}",
            5: f"={detail_ref}!M{subtotal_row}+{detail_ref}!N{subtotal_row}+{detail_ref}!O{subtotal_row}+{detail_ref}!S{subtotal_row}",
            6: f"={detail_ref}!R{subtotal_row}",
            7: f"={detail_ref}!X{subtotal_row}",
            8: f"={detail_ref}!Y{subtotal_row}",
            9: f"={detail_ref}!AA{subtotal_row}",
            10: f"={detail_ref}!AB{subtotal_row}",
            11: f"={detail_ref}!AD{subtotal_row}",
            12: f"={detail_ref}!AE{subtotal_row}",
            13: f"={detail_ref}!AG{subtotal_row}",
            14: f"={detail_ref}!AH{subtotal_row}",
            15: f"={detail_ref}!AJ{subtotal_row}",
            16: f"={detail_ref}!AM{subtotal_row}",
            17: f"={detail_ref}!AN{subtotal_row}",
            18: f"={detail_ref}!AO{subtotal_row}",
            19: f"={detail_ref}!AK{subtotal_row}",
            20: f"={detail_ref}!AP{subtotal_row}",
            21: f"=SUM(D{current_row}:T{current_row})",
        }
        for col_index, formula in defaults.items():
            if col_index <= max_column:
                ws.cell(current_row, col_index).value = formula
        return

    for col_index in range(2, max_column + 1):
        val = ws.cell(current_row, col_index).value
        if col_index == 2:
            if target_leaf is not None and target_leaf.data_end_row >= target_leaf.data_start_row:
                ws.cell(current_row, col_index).value = (
                    f"=COUNT({detail_ref}!$A${target_leaf.data_start_row}:$A${target_leaf.data_end_row})"
                )
            elif target_leaf is not None:
                ws.cell(current_row, col_index).value = target_leaf.employee_count
            continue

        if isinstance(val, str) and val.startswith("="):
            if detail_sheet_name in val:
                ws.cell(current_row, col_index).value = _translate_sheet_references(
                    val, detail_sheet_name, row_map
                )
            elif re.search(rf"\b[A-Za-z]{{1,3}}{src_template_row}\b", val):
                adjusted = re.sub(
                    rf"\b([A-Za-z]{{1,3}}){src_template_row}\b",
                    rf"\g<1>{current_row}",
                    val,
                )
                ws.cell(current_row, col_index).value = adjusted
        elif target_leaf is not None and target_leaf.new_subtotal_row is not None:
            pass


def _translate_summary_group_formulas(
    ws: Worksheet,
    current_row: int,
    src_template_row: int,
    rendered_sub_rows: list[int],
    rendered_group_rows: list[int],
    max_column: int,
) -> None:
    for col_index in range(2, max_column + 1):
        val = ws.cell(current_row, col_index).value
        col_letter = get_column_letter(col_index)
        if rendered_sub_rows:
            first_r = min(rendered_sub_rows)
            last_r = max(rendered_sub_rows)
            if len(rendered_sub_rows) > 1:
                ws.cell(current_row, col_index).value = f"=SUM({col_letter}{first_r}:{col_letter}{last_r})"
            else:
                ws.cell(current_row, col_index).value = f"={col_letter}{first_r}"
        elif isinstance(val, str) and val.startswith("="):
            adjusted = re.sub(
                rf"\b([A-Za-z]{{1,3}}){src_template_row}\b",
                rf"\g<1>{current_row}",
                val,
            )
            ws.cell(current_row, col_index).value = adjusted


def _write_summary_total_formulas(
    ws: Worksheet,
    first_row: int,
    total_row: int,
    rows_to_sum: list[int],
    max_column: int,
) -> None:
    for col_index in range(2, max_column + 1):
        col_letter = get_column_letter(col_index)
        if rows_to_sum:
            if len(rows_to_sum) == (max(rows_to_sum) - min(rows_to_sum) + 1):
                ws.cell(total_row, col_index).value = (
                    f"=SUM({col_letter}{min(rows_to_sum)}:{col_letter}{max(rows_to_sum)})"
                )
            else:
                ws.cell(total_row, col_index).value = "=" + "+".join(
                    f"{col_letter}{r}" for r in rows_to_sum
                )
        elif total_row > first_row:
            ws.cell(total_row, col_index).value = f"=SUM({col_letter}{first_row}:{col_letter}{total_row - 1})"
        else:
            ws.cell(total_row, col_index).value = None



def _cell_text(ws: Worksheet, row_index: int, col_index: int) -> str:
    value = ws.cell(row_index, col_index).value
    if value is None:
        return ""
    return str(value).strip()


def _formula_sheet_name(sheet_name: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"
