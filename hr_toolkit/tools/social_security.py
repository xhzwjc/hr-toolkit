from __future__ import annotations

import re


# 预编译正则
_PERIOD_FROM_FILENAME = re.compile(r"(20\d{2})[-年._ ]?([01]?\d)")
_PERIOD_FROM_TEXT = re.compile(r"(20\d{2})[-年/. ]?([01]?\d)")
_FILENAME_HAS_PERIOD = re.compile(r"^\d{4}[-年._ ]?[01]?\d")
_FILENAME_STRIP_PERIOD = re.compile(r"20\d{2}.*")
_BRACKET_CONTENT = re.compile(r"[（(]([^（）()]+)[）)]")
_SHEET_TITLE_INVALID = re.compile(r"[:\\/?*\[\]]")
_FILENAME_INVALID = re.compile(r"[\\/:*?\"<>|]")
_HEADER_WHITESPACE = re.compile(r"\s+")
import shutil
import tempfile
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from hr_toolkit.common.resources import open_template_resource
from hr_toolkit.common.excel import (
    SheetGrid,
    apply_row_snapshot,
    cached_style_id,
    cell_text as _cell_text,
    insert_rows,
    set_style_ids,
    snapshot_row,
)
from hr_toolkit.common.excel_compat import ensure_xlsx_workbook, is_supported_excel_file
from hr_toolkit.common.inputs import (
    archive_stem,
    extract_archive_excel_files,
    is_supported_archive_file,
    normalize_input_paths,
)


TOOL_NAME = "需求1-社保明细汇总"
DETAIL_TEMPLATE_RESOURCE = "social_security_detail_template.xlsx"
SUMMARY_TEMPLATE_RESOURCE = "social_security_summary_template.xlsx"
DETAIL_OUTPUT_FILENAME = "社保明细表.xlsx"
SUMMARY_OUTPUT_FILENAME = "社保汇总表.xlsx"
SPLIT_DETAIL_DIR_NAME = "按参保单位参保地拆分明细"

TAX_RATE = 0.0672
STANDARD_HOURS_PER_DAY = 7

ROSTER_NAME = "*姓名.简体中文"
ROSTER_ID = "*身份证"
ROSTER_STATUS = "*参保状态"
ROSTER_START_DATE = "*参保日期"
ROSTER_PLAN = "*参保方案.名称"
ROSTER_UNIT = "*参保单位.名称"
ROSTER_DEPARTMENT = "*责任部门.名称"
ROSTER_PROJECT = "项目.项目名称"
ROSTER_COST_CENTER = "成本中心.名称"
ROSTER_MANAGEMENT_FEE = "管理费"

ACCOUNT_DISPLAY_ALIASES = {
    "北京春苗抚州": "北京抚州",
}

COMPANY_ALIASES = {
    "唐人数智科技股份有限公司": "唐人数智",
    "春苗人力资源（北京）有限公司": "北京春苗",
}

SOCIAL_CATEGORIES = ("养老", "医疗", "失业", "工伤", "补充工伤", "生育", "大病医疗")

PAYMENT_NORMAL = "normal"
PAYMENT_ARREARS = "arrears"
PAYMENT_DIFFERENCE = "difference"
PAYMENT_UNKNOWN = "unknown"

FEE_PERIOD_START_HEADERS = (
    "费款所属日期起",
    "费款所属期起",
    "费款所属期",
    "费用所属期",
    "所属月份",
    "缴费月份",
)
FEE_PERIOD_END_HEADERS = ("费款所属日期止", "费款所属期止")
PAYMENT_NATURE_HEADERS = (
    "缴费性质",
    "业务类型",
    "业务类别",
    "申报类型",
    "征缴类型",
    "款项性质",
    "补缴类型",
    "调整类型",
    "摘要",
    "备注",
    "征收子目",
)

DETAIL_COLUMNS = {
    "序号": 1,
    "公司": 2,
    "参保地": 3,
    "项目": 4,
    "姓名": 5,
    "身份证": 6,
    "参保时间": 7,
    "账单期": 8,
    "养老基数": 9,
    "养老个人比例": 10,
    "养老个人金额": 11,
    "养老单位比例": 12,
    "养老单位金额": 13,
    "医疗基数": 14,
    "医疗个人比例": 15,
    "医疗个人金额": 16,
    "医疗单位比例": 17,
    "医疗单位金额": 18,
    "失业基数": 19,
    "失业个人比例": 20,
    "失业个人金额": 21,
    "失业单位比例": 22,
    "失业单位金额": 23,
    "工伤基数": 24,
    "工伤单位比例": 25,
    "工伤单位金额": 26,
    "补充工伤基数": 27,
    "补充工伤单位比例": 28,
    "补充工伤单位金额": 29,
    "生育基数": 30,
    "生育单位比例": 31,
    "生育单位金额": 32,
    "大病医疗基数": 33,
    "大病医疗个人比例": 34,
    "大病医疗个人金额": 35,
    "大病医疗单位比例": 36,
    "大病医疗单位金额": 37,
    "个人社保补缴合计": 62,
    "单位社保补缴合计": 63,
    "个人公积金补缴合计": 64,
    "单位公积金补缴合计": 65,
    "单位补缴滞纳金": 66,
    "个人社保合计": 67,
    "单位社保合计": 68,
    "社保缴纳合计": 69,
    "个人公积金合计": 70,
    "单位公积金合计": 71,
    "公积金缴纳合计": 72,
    "管理费": 73,
    "税金": 74,
    "费用总计": 75,
    "备注": 76,
}

DIFFERENCE_COLUMNS = {
    "养老": {"基数": 44, "个人比例": 45, "个人金额": 46, "单位比例": 47, "单位金额": 48},
    "失业": {"基数": 49, "个人比例": 50, "个人金额": 51, "单位比例": 52, "单位金额": 53},
    "工伤": {"基数": 54, "单位比例": 55, "单位金额": 56},
    "医疗": {"基数": 57, "个人比例": 58, "个人金额": 59, "单位比例": 60, "单位金额": 61},
}


@dataclass(frozen=True)
class RosterPerson:
    name: str
    id_card: str
    status: str
    start_period: str
    account: str
    account_display: str
    company: str
    insured_place: str
    department: str
    project: str
    project_display: str
    cost_center: str
    management_fee: float | None


@dataclass(frozen=True)
class SourceContext:
    label: str
    file_period: str
    container_period: str
    billing_period_hint: str
    period_split_file: bool
    account_hint: str
    company_hint: str
    nature_hint: str | None


@dataclass(frozen=True)
class SocialPaymentLine:
    source_file: str
    source_row: int
    name: str
    id_card: str
    fee_period: str
    fee_period_end: str
    billing_period_hint: str
    period_split_file: bool
    account_hint: str
    company_hint: str
    category: str
    side: str
    wage: float | None
    base: float | None
    rate: float | None
    amount: float
    nature_hint: str | None = None


@dataclass
class DetailRecord:
    id_card: str
    name: str
    period: str
    billing_period: str
    period_split_input: bool
    account: str
    account_display: str
    company: str
    insured_place: str
    project: str
    project_display: str
    cost_center: str
    start_period: str
    management_fee: float | None
    amounts: dict[str, dict[str, float]] = field(default_factory=dict)
    bases: dict[str, float] = field(default_factory=dict)
    base_period_counts: dict[str, int] = field(default_factory=dict)
    rates: dict[str, dict[str, float]] = field(default_factory=dict)
    arrears_amounts: dict[str, dict[str, float]] = field(default_factory=dict)
    difference_amounts: dict[str, dict[str, float]] = field(default_factory=dict)
    difference_bases: dict[str, set[float]] = field(default_factory=dict)
    difference_rates: dict[str, dict[str, set[float]]] = field(default_factory=dict)
    main_periods: set[str] = field(default_factory=set)
    difference_periods: dict[str, set[str]] = field(default_factory=dict)
    unknown_periods: dict[str, set[str]] = field(default_factory=dict)
    source_files: set[str] = field(default_factory=set)
    warnings: list[str] = field(default_factory=list)

    @property
    def personal_total(self) -> float:
        normal = sum(side_amounts.get("个人", 0.0) for side_amounts in self.amounts.values())
        arrears = sum(side_amounts.get("个人", 0.0) for side_amounts in self.arrears_amounts.values())
        difference = sum(side_amounts.get("个人", 0.0) for side_amounts in self.difference_amounts.values())
        return round(normal + arrears + difference, 2)

    @property
    def unit_total(self) -> float:
        normal = sum(side_amounts.get("单位", 0.0) for side_amounts in self.amounts.values())
        arrears = sum(side_amounts.get("单位", 0.0) for side_amounts in self.arrears_amounts.values())
        difference = sum(side_amounts.get("单位", 0.0) for side_amounts in self.difference_amounts.values())
        return round(normal + arrears + difference, 2)

    @property
    def social_total(self) -> float:
        return round(self.personal_total + self.unit_total, 2)

    @property
    def tax(self) -> float:
        return round((self.social_total + (self.management_fee or 0.0)) * TAX_RATE, 2)

    @property
    def total_fee(self) -> float:
        return round(self.social_total + (self.management_fee or 0.0) + self.tax, 2)


@dataclass
class SocialSecurityResult:
    input_path: Path
    roster_path: Path
    output_dir: Path
    dry_run: bool = False
    source_files: list[str] = field(default_factory=list)
    source_record_count: int = 0
    detail_record_count: int = 0
    employee_count: int = 0
    account_counts: dict[str, int] = field(default_factory=dict)
    period_counts: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    detail_output_file: Path | None = None
    detail_output_files: list[Path] = field(default_factory=list)
    summary_output_file: Path | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "tool_name": TOOL_NAME,
            "input_path": str(self.input_path),
            "roster_path": str(self.roster_path),
            "output_dir": str(self.output_dir),
            "dry_run": self.dry_run,
            "source_files": self.source_files,
            "source_file_count": len(self.source_files),
            "source_record_count": self.source_record_count,
            "detail_record_count": self.detail_record_count,
            "employee_count": self.employee_count,
            "account_counts": self.account_counts,
            "period_counts": self.period_counts,
            "warnings": self.warnings,
            "detail_output_file": None if self.detail_output_file is None else str(self.detail_output_file),
            "detail_output_files": [str(path) for path in self.detail_output_files],
            "summary_output_file": None if self.summary_output_file is None else str(self.summary_output_file),
        }


def generate_social_security_reports(
    input_path: str | Path | list[str | Path],
    roster_path: str | Path,
    output_dir: str | Path,
    *,
    dry_run: bool = False,
    cancelled: Callable[[], bool] | None = None,
) -> SocialSecurityResult:
    _check_cancelled(cancelled)
    input_paths = _normalize_input_paths(input_path)
    display_input = input_paths[0] if len(input_paths) == 1 else input_paths[0].parent
    roster = Path(roster_path).expanduser().resolve()
    output = Path(output_dir).expanduser().resolve()
    warnings: list[str] = []

    for path in input_paths:
        _check_cancelled(cancelled)
        if not path.exists():
            raise FileNotFoundError(f"社保缴费清单文件、压缩包或文件夹不存在：{path}")
    if not roster.exists() or not roster.is_file():
        raise FileNotFoundError(f"参保人员花名册不存在：{roster}")
    if not is_supported_excel_file(roster):
        raise ValueError("参保人员花名册仅支持 .xlsx 或 .xls 文件。")

    with tempfile.TemporaryDirectory(prefix="hr_social_security_") as temp_root:
        temp_dir = Path(temp_root)
        roster_people = _read_roster(roster, temp_dir)
        source_files = _find_social_security_files(input_paths, temp_dir, warnings)
        _check_cancelled(cancelled)
        if not source_files:
            raise ValueError("未找到 .xlsx 或 .xls 社保缴费清单。")

        lines: list[SocialPaymentLine] = []
        used_files: list[str] = []
        for source_file in source_files:
            _check_cancelled(cancelled)
            try:
                file_lines = _read_payment_file(source_file)
            except ValueError as exc:
                if len(input_paths) == 1 and input_paths[0].is_file():
                    raise
                warnings.append(f"{source_file.name} 不是可识别的社保缴费清单，已跳过：{exc}")
                continue
            if file_lines:
                used_files.append(str(source_file))
            lines.extend(file_lines)

        detail_records = _build_detail_records(lines, roster_people, warnings)
        result = SocialSecurityResult(
            input_path=display_input,
            roster_path=roster,
            output_dir=output,
            dry_run=dry_run,
            source_files=used_files,
            source_record_count=len(lines),
            detail_record_count=len(detail_records),
            employee_count=len({record.id_card for record in detail_records}),
            account_counts=_count_records(detail_records, "account_display"),
            period_counts=_count_records(detail_records, "period"),
            warnings=warnings,
        )
        if dry_run:
            return result

        _check_cancelled(cancelled)
        output.mkdir(parents=True, exist_ok=True)
        detail_output = output / DETAIL_OUTPUT_FILENAME
        summary_output = output / SUMMARY_OUTPUT_FILENAME
        _write_detail_workbook(
            detail_records,
            detail_output,
            temp_dir,
            cancelled=cancelled,
        )
        _check_cancelled(cancelled)
        split_detail_outputs = _write_split_detail_workbooks(
            detail_records,
            output,
            temp_dir,
            cancelled=cancelled,
        )
        _check_cancelled(cancelled)
        _write_summary_workbook(
            detail_records,
            summary_output,
            temp_dir,
            warnings,
            cancelled=cancelled,
        )
        result.detail_output_file = detail_output
        result.detail_output_files = split_detail_outputs
        result.summary_output_file = summary_output
        return result


def _check_cancelled(cancelled: Callable[[], bool] | None) -> None:
    if cancelled is not None and cancelled():
        raise RuntimeError("本次处理已停止。")


def _normalize_input_paths(input_path: str | Path | list[str | Path]) -> list[Path]:
    return normalize_input_paths(input_path, "请选择社保缴费清单文件、压缩包或文件夹。")


def _find_social_security_files(input_paths: list[Path], temp_dir: Path, warnings: list[str]) -> list[Path]:
    files: list[Path] = []
    seen: set[Path] = set()
    for input_path in input_paths:
        for file_path in _iter_input_files(input_path, temp_dir, warnings):
            resolved = file_path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            files.append(file_path)
    return sorted(files)


def _iter_input_files(input_path: Path, temp_dir: Path, warnings: list[str]) -> list[Path]:
    if input_path.is_file():
        if is_supported_excel_file(input_path):
            return [input_path]
        if is_supported_archive_file(input_path):
            return _extract_archive_files(input_path, temp_dir, warnings)
        return []
    if not input_path.is_dir():
        raise FileNotFoundError(f"路径不存在：{input_path}")
    files: list[Path] = []
    for path in sorted(input_path.rglob("*")):
        if not path.is_file() or path.name.startswith(("~$", ".~")):
            continue
        if is_supported_excel_file(path) and not _is_non_source_excel(path):
            files.append(path)
        elif is_supported_archive_file(path):
            files.extend(_extract_archive_files(path, temp_dir, warnings))
    return files


def _extract_archive_files(archive_path: Path, temp_dir: Path, warnings: list[str]) -> list[Path]:
    files = extract_archive_excel_files(
        archive_path,
        temp_dir,
        warnings,
        subdir=_safe_file_stem(archive_stem(archive_path)),
    )
    return [path for path in files if not _is_non_source_excel(path)]


def _read_roster(roster_path: Path, temp_dir: Path) -> dict[str, RosterPerson]:
    if roster_path.suffix.lower() == ".xls" and _is_binary_xls(roster_path):
        return _read_xls_roster(roster_path)
    working_path = ensure_xlsx_workbook(roster_path, temp_dir)
    workbook = load_workbook(working_path, data_only=True, read_only=True)
    try:
        # read_only 工作表随机访问是 O(行数²)，先单遍读入内存再处理
        ws = SheetGrid(workbook[workbook.sheetnames[0]])
        header_row = _find_header_row(ws, (ROSTER_NAME, ROSTER_ID))
        headers = _read_headers(ws, header_row)
        people: dict[str, RosterPerson] = {}
        for row_index in range(header_row + 1, (ws.max_row or 0) + 1):
            values = {header: ws.cell(row_index, col_index).value for header, col_index in headers.items()}
            id_card = _normalize_id_card(values.get(ROSTER_ID))
            name = _cell_text(values.get(ROSTER_NAME))
            if not id_card or not name:
                continue
            account = _cell_text(values.get(ROSTER_PLAN))
            unit = _cell_text(values.get(ROSTER_UNIT))
            department = _cell_text(values.get(ROSTER_DEPARTMENT))
            project = _cell_text(values.get(ROSTER_PROJECT))
            cost_center = _cell_text(values.get(ROSTER_COST_CENTER))
            people[id_card] = RosterPerson(
                name=name,
                id_card=id_card,
                status=_cell_text(values.get(ROSTER_STATUS)),
                start_period=_period_from_value(values.get(ROSTER_START_DATE)) or "",
                account=account,
                account_display=_account_display(account),
                company=_company_display(unit, account),
                insured_place=_insured_place(account),
                department=department,
                project=project,
                project_display=_project_display(project, cost_center, department),
                cost_center=cost_center,
                management_fee=_to_number(values.get(ROSTER_MANAGEMENT_FEE)),
            )
        if not people:
            raise ValueError("参保人员花名册中未识别到人员数据。")
        return people
    finally:
        workbook.close()


def _read_xls_roster(roster_path: Path) -> dict[str, RosterPerson]:
    try:
        import xlrd  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RuntimeError("读取 .xls 参保人员花名册需要 xlrd，请先安装依赖。") from exc
    book = xlrd.open_workbook(roster_path)
    for sheet in book.sheets():
        header_row = _find_xls_roster_header_row(sheet)
        if header_row is None:
            continue
        headers = {_normalize_header(value): index for index, value in enumerate(sheet.row_values(header_row)) if _cell_text(value)}
        people: dict[str, RosterPerson] = {}
        for row_index in range(header_row + 1, sheet.nrows):
            values = {header: sheet.cell_value(row_index, col_index) for header, col_index in headers.items()}
            id_card = _normalize_id_card(values.get(ROSTER_ID))
            name = _cell_text(values.get(ROSTER_NAME))
            if not id_card or not name:
                continue
            account = _cell_text(values.get(ROSTER_PLAN))
            unit = _cell_text(values.get(ROSTER_UNIT))
            department = _cell_text(values.get(ROSTER_DEPARTMENT))
            project = _cell_text(values.get(ROSTER_PROJECT))
            cost_center = _cell_text(values.get(ROSTER_COST_CENTER))
            people[id_card] = RosterPerson(
                name=name,
                id_card=id_card,
                status=_cell_text(values.get(ROSTER_STATUS)),
                start_period=_period_from_value(values.get(ROSTER_START_DATE)) or "",
                account=account,
                account_display=_account_display(account),
                company=_company_display(unit, account),
                insured_place=_insured_place(account),
                department=department,
                project=project,
                project_display=_project_display(project, cost_center, department),
                cost_center=cost_center,
                management_fee=_to_number(values.get(ROSTER_MANAGEMENT_FEE)),
            )
        if people:
            return people
    raise ValueError("参保人员花名册中未识别到人员数据。")


def _find_xls_roster_header_row(sheet) -> int | None:
    required = {_normalize_header(ROSTER_NAME), _normalize_header(ROSTER_ID)}
    for row_index in range(min(sheet.nrows, 30)):
        values = {_normalize_header(value) for value in sheet.row_values(row_index)}
        if required.issubset(values):
            return row_index
    return None


def _read_payment_file(file_path: Path) -> list[SocialPaymentLine]:
    context = _source_context(file_path)
    if file_path.suffix.lower() == ".xls":
        return _read_xls_payment_file(file_path, context)
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        # read_only 工作表随机访问是 O(行数²)，先单遍读入内存再处理
        ws = SheetGrid(workbook[workbook.sheetnames[0]])
        header_row = _find_payment_header_row(ws)
        headers = _read_headers(ws, header_row)
        if "参保费种" in headers and ("征收品目" in headers or "险种" in headers):
            return _read_long_sheet(ws, headers, header_row, context)
        if _is_single_kind_sheet(headers):
            return _read_single_kind_sheet(ws, headers, header_row, context)
        if _has_wide_amount_columns(headers):
            return _read_wide_sheet(ws, headers, header_row, context)
        return _read_single_kind_sheet(ws, headers, header_row, context)
    finally:
        workbook.close()


def _read_xls_payment_file(file_path: Path, context: SourceContext) -> list[SocialPaymentLine]:
    try:
        import xlrd  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RuntimeError("读取 .xls 社保清单需要 xlrd，请先安装依赖。") from exc
    book = xlrd.open_workbook(file_path)
    for sheet in book.sheets():
        header_row = _find_xls_header_row(sheet)
        if header_row is None:
            continue
        headers = {_normalize_header(value): index for index, value in enumerate(sheet.row_values(header_row)) if _cell_text(value)}
        if "参保费种" in headers and ("征收品目" in headers or "险种" in headers):
            return _read_xls_long_sheet(sheet, headers, header_row, context)
        if _is_single_kind_sheet(headers):
            return _read_xls_single_kind_sheet(sheet, headers, header_row, context)
        if _has_wide_amount_columns(headers):
            return _read_xls_wide_sheet(sheet, headers, header_row, context)
        return _read_xls_single_kind_sheet(sheet, headers, header_row, context)
    raise ValueError("未找到包含姓名和身份证的表头。")


def _find_xls_header_row(sheet) -> int | None:
    for row_index in range(min(sheet.nrows, 20)):
        values = {_normalize_header(value) for value in sheet.row_values(row_index)}
        if "姓名" in values and ("身份证件号码" in values or "证件号码" in values):
            return row_index
    return None


def _read_xls_long_sheet(sheet, headers: dict[str, int], header_row: int, context: SourceContext) -> list[SocialPaymentLine]:
    lines: list[SocialPaymentLine] = []
    for row_index in range(header_row + 1, sheet.nrows):
        row = {header: sheet.cell_value(row_index, col_index) for header, col_index in headers.items()}
        name = _cell_text(row.get("姓名"))
        id_card = _normalize_id_card(row.get("身份证件号码") or row.get("证件号码"))
        if not _valid_person_row(name, id_card):
            continue
        category, side = _classify_insurance_item(_cell_text(row.get("参保费种")), _cell_text(row.get("征收品目")))
        if category is None:
            continue
        fee_period, fee_period_end = _payment_periods_from_row(row, context)
        lines.append(
            SocialPaymentLine(
                source_file=context.label,
                source_row=row_index + 1,
                name=name,
                id_card=id_card,
                fee_period=fee_period,
                fee_period_end=fee_period_end,
                billing_period_hint=context.billing_period_hint,
                period_split_file=context.period_split_file,
                account_hint=context.account_hint,
                company_hint=context.company_hint,
                category=category,
                side=side,
                wage=_to_number(row.get("本人工资")),
                base=_to_number(row.get("缴费基数")),
                rate=_rate_to_decimal(row.get("费率")),
                amount=_to_number(row.get("本期应缴费额")) or 0.0,
                nature_hint=_nature_hint_from_row(row, context),
            )
        )
    return lines


def _read_xls_single_kind_sheet(sheet, headers: dict[str, int], header_row: int, context: SourceContext) -> list[SocialPaymentLine]:
    category, side = _classify_insurance_item(context.label, context.label)
    if category is None:
        raise ValueError("文件名中未识别险种。")
    lines: list[SocialPaymentLine] = []
    for row_index in range(header_row + 1, sheet.nrows):
        row = {header: sheet.cell_value(row_index, col_index) for header, col_index in headers.items()}
        name = _cell_text(row.get("姓名"))
        id_card = _normalize_id_card(row.get("身份证件号码") or row.get("证件号码"))
        if not _valid_person_row(name, id_card):
            continue
        fee_period, fee_period_end = _payment_periods_from_row(row, context)
        lines.append(
            SocialPaymentLine(
                source_file=context.label,
                source_row=row_index + 1,
                name=name,
                id_card=id_card,
                fee_period=fee_period,
                fee_period_end=fee_period_end,
                billing_period_hint=context.billing_period_hint,
                period_split_file=context.period_split_file,
                account_hint=context.account_hint,
                company_hint=context.company_hint,
                category=category,
                side=side,
                wage=_to_number(row.get("缴费工资") or row.get("本人工资")),
                base=_to_number(row.get("缴费基数")),
                rate=_rate_to_decimal(row.get("费率")),
                amount=_to_number(row.get("应缴费额(元)") or row.get("本期应缴费额")) or 0.0,
                nature_hint=_nature_hint_from_row(row, context),
            )
        )
    return lines


def _read_xls_wide_sheet(sheet, headers: dict[str, int], header_row: int, context: SourceContext) -> list[SocialPaymentLine]:
    amount_columns: list[tuple[str, int, str, str]] = []
    for header, col_index in headers.items():
        if "应缴费额" not in header:
            continue
        category, side = _classify_insurance_item(header, header)
        if category is not None:
            amount_columns.append((header, col_index, category, side))
    if not amount_columns:
        raise ValueError("宽表中未识别险种金额列。")
    lines: list[SocialPaymentLine] = []
    for row_index in range(header_row + 1, sheet.nrows):
        name = _cell_text(_xls_header_value(sheet, row_index, headers, "姓名"))
        id_card = _normalize_id_card(_xls_header_value(sheet, row_index, headers, "证件号码") or _xls_header_value(sheet, row_index, headers, "身份证件号码"))
        if not _valid_person_row(name, id_card):
            continue
        row = {header: sheet.cell_value(row_index, col_index) for header, col_index in headers.items()}
        fee_period, fee_period_end = _payment_periods_from_row(row, context)
        for _header, col_index, category, side in amount_columns:
            amount = _to_number(sheet.cell_value(row_index, col_index))
            if not amount:
                continue
            lines.append(
                SocialPaymentLine(
                    source_file=context.label,
                    source_row=row_index + 1,
                    name=name,
                    id_card=id_card,
                    fee_period=fee_period,
                    fee_period_end=fee_period_end,
                    billing_period_hint=context.billing_period_hint,
                    period_split_file=context.period_split_file,
                    account_hint=context.account_hint,
                    company_hint=context.company_hint,
                    category=category,
                    side=side,
                    wage=None,
                    base=None,
                    rate=None,
                    amount=amount,
                    nature_hint=_nature_hint_from_row(row, context, _header),
                )
            )
    return lines


def _read_long_sheet(ws: Worksheet, headers: dict[str, int], header_row: int, context: SourceContext) -> list[SocialPaymentLine]:
    lines: list[SocialPaymentLine] = []
    for row_index in range(header_row + 1, (ws.max_row or 0) + 1):
        row = {header: ws.cell(row_index, col_index).value for header, col_index in headers.items()}
        name = _cell_text(row.get("姓名"))
        id_card = _normalize_id_card(row.get("身份证件号码") or row.get("证件号码"))
        if not _valid_person_row(name, id_card):
            continue
        category, side = _classify_insurance_item(_cell_text(row.get("参保费种")), _cell_text(row.get("征收品目") or row.get("险种")))
        if category is None:
            continue
        fee_period, fee_period_end = _payment_periods_from_row(row, context)
        lines.append(
            SocialPaymentLine(
                source_file=context.label,
                source_row=row_index,
                name=name,
                id_card=id_card,
                fee_period=fee_period,
                fee_period_end=fee_period_end,
                billing_period_hint=context.billing_period_hint,
                period_split_file=context.period_split_file,
                account_hint=context.account_hint,
                company_hint=context.company_hint,
                category=category,
                side=side,
                wage=_to_number(row.get("本人工资") or row.get("缴费工资")),
                base=_to_number(row.get("缴费基数")),
                rate=_rate_to_decimal(row.get("费率")),
                amount=_to_number(row.get("本期应缴费额") or row.get("应缴费额(元)")) or 0.0,
                nature_hint=_nature_hint_from_row(row, context),
            )
        )
    return lines


def _read_single_kind_sheet(ws: Worksheet, headers: dict[str, int], header_row: int, context: SourceContext) -> list[SocialPaymentLine]:
    category, side = _classify_insurance_item(context.label, context.label)
    if category is None:
        raise ValueError("文件名中未识别险种。")
    lines: list[SocialPaymentLine] = []
    for row_index in range(header_row + 1, (ws.max_row or 0) + 1):
        row = {header: ws.cell(row_index, col_index).value for header, col_index in headers.items()}
        name = _cell_text(row.get("姓名"))
        id_card = _normalize_id_card(row.get("身份证件号码") or row.get("证件号码"))
        if not _valid_person_row(name, id_card):
            continue
        fee_period, fee_period_end = _payment_periods_from_row(row, context)
        lines.append(
            SocialPaymentLine(
                source_file=context.label,
                source_row=row_index,
                name=name,
                id_card=id_card,
                fee_period=fee_period,
                fee_period_end=fee_period_end,
                billing_period_hint=context.billing_period_hint,
                period_split_file=context.period_split_file,
                account_hint=context.account_hint,
                company_hint=context.company_hint,
                category=category,
                side=side,
                wage=_to_number(row.get("缴费工资") or row.get("本人工资")),
                base=_to_number(row.get("缴费基数")),
                rate=_rate_to_decimal(row.get("费率")),
                amount=_to_number(row.get("应缴费额(元)") or row.get("本期应缴费额")) or 0.0,
                nature_hint=_nature_hint_from_row(row, context),
            )
        )
    return lines


def _read_wide_sheet(ws: Worksheet, headers: dict[str, int], header_row: int, context: SourceContext) -> list[SocialPaymentLine]:
    amount_columns: list[tuple[str, int, str, str]] = []
    for header, col_index in headers.items():
        if "应缴费额" not in header:
            continue
        category, side = _classify_insurance_item(header, header)
        if category is not None:
            amount_columns.append((header, col_index, category, side))
    if not amount_columns:
        raise ValueError("宽表中未识别险种金额列。")

    lines: list[SocialPaymentLine] = []
    for row_index in range(header_row + 1, (ws.max_row or 0) + 1):
        name = _cell_text(ws.cell(row_index, headers.get("姓名", 0)).value if "姓名" in headers else "")
        id_card = _normalize_id_card(ws.cell(row_index, headers.get("证件号码", headers.get("身份证件号码", 0))).value if ("证件号码" in headers or "身份证件号码" in headers) else "")
        if not _valid_person_row(name, id_card):
            continue
        row = {header: ws.cell(row_index, col_index).value for header, col_index in headers.items()}
        fee_period, fee_period_end = _payment_periods_from_row(row, context)
        for _header, col_index, category, side in amount_columns:
            amount = _to_number(ws.cell(row_index, col_index).value)
            if not amount:
                continue
            lines.append(
                SocialPaymentLine(
                    source_file=context.label,
                    source_row=row_index,
                    name=name,
                    id_card=id_card,
                    fee_period=fee_period,
                    fee_period_end=fee_period_end,
                    billing_period_hint=context.billing_period_hint,
                    period_split_file=context.period_split_file,
                    account_hint=context.account_hint,
                    company_hint=context.company_hint,
                    category=category,
                    side=side,
                    wage=None,
                    base=None,
                    rate=None,
                    amount=amount,
                    nature_hint=_nature_hint_from_row(row, context, _header),
                )
            )
    return lines


def _build_detail_records(
    lines: list[SocialPaymentLine],
    roster_people: dict[str, RosterPerson],
    warnings: list[str],
) -> list[DetailRecord]:
    resolved: list[tuple[SocialPaymentLine, RosterPerson, str, str, str]] = []
    account_by_line: dict[SocialPaymentLine, str] = {}
    mismatch_warning_keys: set[tuple[str, str, str, str, str]] = set()
    for line in lines:
        person = roster_people.get(line.id_card)
        if person is None:
            warnings.append(f"{line.source_file} 第 {line.source_row} 行人员 {line.name} 未在参保人员花名册中找到，已跳过。")
            continue
        if _normalize_name(line.name) != _normalize_name(person.name):
            warnings.append(f"{line.source_file} 第 {line.source_row} 行身份证 {line.id_card} 姓名与花名册不同：{line.name} / {person.name}。")
        source_account = _account_display(line.account_hint)
        source_company = _company_display(line.company_hint, line.account_hint)
        source_place = _insured_place(line.account_hint)
        _append_source_mismatch_warnings(line, person, source_account, source_company, source_place, warnings, mismatch_warning_keys)
        account_display = source_account or person.account_display
        account_by_line[line] = account_display
        resolved.append((line, person, account_display, source_company, source_place))

    billing_periods = _billing_periods_by_account(resolved)
    natures = _classify_payment_natures([item[0] for item in resolved], account_by_line, billing_periods)
    resolved_groups: OrderedDict[
        tuple[str, str, str],
        list[tuple[SocialPaymentLine, RosterPerson, str, str, str]],
    ] = OrderedDict()
    normalized_natures: dict[SocialPaymentLine, str] = {}
    unsupported_difference_warning_keys: set[tuple[str, str, str]] = set()
    for item in resolved:
        line, person, account_display, _source_company, _source_place = item
        billing_period = billing_periods.get(account_display, "") or line.billing_period_hint or line.fee_period or "未识别账单期"
        key = (line.id_card, account_display, billing_period)
        nature = natures.get(line, PAYMENT_NORMAL)
        if nature == PAYMENT_DIFFERENCE and line.category not in DIFFERENCE_COLUMNS:
            warning_key = (line.id_card, line.category, billing_period)
            if warning_key not in unsupported_difference_warning_keys:
                unsupported_difference_warning_keys.add(warning_key)
                warnings.append(
                    f"{person.name} {line.category}来源明确为补差，但模板没有对应补差明细列；"
                    "金额已保留在普通缴费区，请人工确认。"
                )
            nature = PAYMENT_NORMAL
        normalized_natures[line] = nature
        resolved_groups.setdefault(key, []).append(item)
        if billing_period == "未识别账单期":
            warnings.append(f"{line.source_file} 第 {line.source_row} 行未识别账单期，已写入“未识别账单期”。")

    records: list[DetailRecord] = []
    for (_id_card, _group_account_display, billing_period), group_items in resolved_groups.items():
        front_items = [item for item in group_items if normalized_natures[item[0]] != PAYMENT_DIFFERENCE]
        difference_items = [item for item in group_items if normalized_natures[item[0]] == PAYMENT_DIFFERENCE]
        group_by_period = any(normalized_natures[item[0]] == PAYMENT_ARREARS for item in front_items)
        front_groups = _group_front_payment_items(front_items) if group_by_period else [front_items]
        if not front_groups:
            front_groups = [[]]
        primary_index = _primary_front_group_index(front_groups, billing_period)
        period_split_input = any(item[0].period_split_file for item in group_items)

        group_records: list[DetailRecord] = []
        for group_index, front_group in enumerate(front_groups):
            context_item = front_group[0] if front_group else group_items[0]
            line, person, account_display, source_company, source_place = context_item
            record = DetailRecord(
                id_card=line.id_card,
                name=person.name,
                period=billing_period,
                billing_period=billing_period,
                period_split_input=period_split_input,
                account=line.account_hint or person.account,
                account_display=account_display,
                company=source_company or person.company,
                insured_place=source_place or person.insured_place,
                project=person.project,
                project_display=person.project_display,
                cost_center=person.cost_center,
                start_period=person.start_period,
                management_fee=person.management_fee if group_index == primary_index else None,
            )
            _add_front_payment_lines(
                record,
                front_group,
                normalized_natures,
                accumulate_bases=group_by_period and _front_group_source_period_count(front_group) > 1,
            )
            if group_by_period and record.main_periods:
                record.period = _format_period_span(record.main_periods)
            if group_index == primary_index:
                _add_difference_payment_lines(record, difference_items)
            _append_payment_nature_warnings(record, warnings)
            group_records.append(record)
        records.extend(group_records)
    return records


def _group_front_payment_items(
    items: list[tuple[SocialPaymentLine, RosterPerson, str, str, str]],
) -> list[list[tuple[SocialPaymentLine, RosterPerson, str, str, str]]]:
    buckets: OrderedDict[
        tuple[str, ...],
        list[tuple[SocialPaymentLine, RosterPerson, str, str, str]],
    ] = OrderedDict()
    for item in items:
        periods = tuple(sorted(_line_periods(item[0])))
        buckets.setdefault(periods, []).append(item)

    ordered_buckets = sorted(
        buckets.items(),
        key=lambda item: (item[0][0] if item[0] else "999999", item[0]),
    )
    groups: list[list[tuple[SocialPaymentLine, RosterPerson, str, str, str]]] = []
    group_periods: list[set[str]] = []
    group_profiles: list[tuple[tuple[str, str, float, float], ...] | None] = []
    for periods, bucket_items in ordered_buckets:
        profile = _front_payment_profile(bucket_items) if len(periods) == 1 else None
        can_merge = (
            bool(groups)
            and profile is not None
            and group_profiles[-1] == profile
            and len(periods) == 1
            and _periods_are_adjacent(group_periods[-1], periods[0])
        )
        if can_merge:
            groups[-1].extend(bucket_items)
            group_periods[-1].update(periods)
            continue
        groups.append(list(bucket_items))
        group_periods.append(set(periods))
        group_profiles.append(profile)
    return groups


def _front_payment_profile(
    items: list[tuple[SocialPaymentLine, RosterPerson, str, str, str]],
) -> tuple[tuple[str, str, float, float], ...] | None:
    profile: list[tuple[str, str, float, float]] = []
    bases_by_category: dict[str, list[float]] = {}
    for line, _person, _account_display, _source_company, _source_place in items:
        if line.category not in SOCIAL_CATEGORIES or line.base is None or line.rate is None:
            return None
        bases_by_category.setdefault(line.category, []).append(line.base)
        profile.append((line.category, line.side, round(line.base, 4), round(line.rate, 8)))
    for values in bases_by_category.values():
        if any(not _amount_close(values[0], value) for value in values[1:]):
            return None
    return tuple(sorted(profile))


def _periods_are_adjacent(existing_periods: set[str], next_period: str) -> bool:
    if not existing_periods or not _is_period(next_period):
        return False
    last_period = max(existing_periods)
    return _period_sequence(last_period, next_period) == [last_period, next_period]


def _primary_front_group_index(
    groups: list[list[tuple[SocialPaymentLine, RosterPerson, str, str, str]]],
    billing_period: str,
) -> int:
    for index, group in enumerate(groups):
        periods = set().union(*(_line_periods(item[0]) for item in group)) if group else set()
        if billing_period in periods:
            return index
    return max(len(groups) - 1, 0)


def _front_group_source_period_count(
    items: list[tuple[SocialPaymentLine, RosterPerson, str, str, str]],
) -> int:
    periods = {
        (item[0].fee_period, item[0].fee_period_end or item[0].fee_period)
        for item in items
        if item[0].fee_period
    }
    return len(periods)


def _add_front_payment_lines(
    record: DetailRecord,
    items: list[tuple[SocialPaymentLine, RosterPerson, str, str, str]],
    natures: dict[SocialPaymentLine, str],
    *,
    accumulate_bases: bool,
) -> None:
    bases_by_category_period: dict[str, dict[tuple[str, str], list[float]]] = {}
    for line, _person, _account_display, _source_company, _source_place in items:
        if line.category not in SOCIAL_CATEGORIES:
            record.warnings.append(f"未识别险种：{line.category}")
            continue
        record.source_files.add(line.source_file)
        line_periods = _line_periods(line)
        side_amounts = record.amounts.setdefault(line.category, {"个人": 0.0, "单位": 0.0})
        side_amounts[line.side] = round(side_amounts.get(line.side, 0.0) + line.amount, 2)
        record.main_periods.update(line_periods)
        if line.base is not None:
            if accumulate_bases:
                period_key = (line.fee_period, line.fee_period_end or line.fee_period)
                bases_by_category_period.setdefault(line.category, {}).setdefault(period_key, []).append(line.base)
            else:
                record.bases[line.category] = line.base
        if line.rate is not None:
            record.rates.setdefault(line.category, {})[line.side] = line.rate
        if natures[line] == PAYMENT_UNKNOWN:
            record.unknown_periods.setdefault(line.category, set()).update(line_periods)

    if not accumulate_bases:
        return
    for category, period_values in bases_by_category_period.items():
        total_base = 0.0
        valid = True
        for values in period_values.values():
            if any(not _amount_close(values[0], value) for value in values[1:]):
                valid = False
                break
            total_base += values[0]
        if valid:
            record.bases[category] = round(total_base, 4)
            record.base_period_counts[category] = len(period_values)


def _add_difference_payment_lines(
    record: DetailRecord,
    items: list[tuple[SocialPaymentLine, RosterPerson, str, str, str]],
) -> None:
    for line, _person, _account_display, _source_company, _source_place in items:
        if line.category not in DIFFERENCE_COLUMNS:
            continue
        record.source_files.add(line.source_file)
        side_amounts = record.difference_amounts.setdefault(line.category, {"个人": 0.0, "单位": 0.0})
        side_amounts[line.side] = round(side_amounts.get(line.side, 0.0) + line.amount, 2)
        record.difference_periods.setdefault(line.category, set()).update(_line_periods(line))
        if line.base is not None:
            record.difference_bases.setdefault(line.category, set()).add(line.base)
        if line.rate is not None:
            record.difference_rates.setdefault(line.category, {}).setdefault(line.side, set()).add(line.rate)


def _billing_periods_by_account(
    resolved: list[tuple[SocialPaymentLine, RosterPerson, str, str, str]],
) -> dict[str, str]:
    fee_periods: OrderedDict[str, set[str]] = OrderedDict()
    fallback_periods: OrderedDict[str, set[str]] = OrderedDict()
    for line, _person, account_display, _source_company, _source_place in resolved:
        if line.fee_period:
            fee_periods.setdefault(account_display, set()).add(line.fee_period_end or line.fee_period)
        if line.billing_period_hint:
            fallback_periods.setdefault(account_display, set()).add(line.billing_period_hint)
    accounts = OrderedDict.fromkeys([item[2] for item in resolved])
    result: dict[str, str] = {}
    for account in accounts:
        candidates = fallback_periods.get(account) or fee_periods.get(account) or set()
        result[account] = max(candidates) if candidates else ""
    return result


def _classify_payment_natures(
    lines: list[SocialPaymentLine],
    account_by_line: dict[SocialPaymentLine, str],
    billing_periods: dict[str, str],
) -> dict[SocialPaymentLine, str]:
    result: dict[SocialPaymentLine, str] = {}
    current_by_person: dict[tuple[str, str, str, str], list[SocialPaymentLine]] = {}
    current_by_account: dict[tuple[str, str, str], list[SocialPaymentLine]] = {}
    history_groups: dict[tuple[str, str, str, str], list[SocialPaymentLine]] = {}

    for line in lines:
        account = account_by_line[line]
        billing_period = billing_periods.get(account, "")
        if line.nature_hint:
            result[line] = line.nature_hint
        if line.fee_period and billing_period and line.fee_period < billing_period:
            history_groups.setdefault((line.id_card, account, line.category, line.side), []).append(line)
        elif line.fee_period and line.fee_period == billing_period:
            current_by_person.setdefault((line.id_card, account, line.category, line.side), []).append(line)
            current_by_account.setdefault((account, line.category, line.side), []).append(line)

    difference_candidates = _cohort_difference_candidates(history_groups, current_by_person, current_by_account)

    for line in lines:
        if line in result:
            continue
        account = account_by_line[line]
        billing_period = billing_periods.get(account, "")
        if not line.fee_period or not billing_period or line.fee_period >= billing_period:
            result[line] = PAYMENT_NORMAL
            continue

        person_key = (line.id_card, account, line.category, line.side)
        person_current = current_by_person.get(person_key, [])
        account_current = current_by_account.get((account, line.category, line.side), [])
        if any(_same_contribution_profile(line, current, require_basis_evidence=True) for current in person_current):
            result[line] = PAYMENT_ARREARS
            continue
        if any(_same_contribution_profile(line, current, require_basis_evidence=True) for current in account_current):
            result[line] = PAYMENT_ARREARS
            continue

        if line in difference_candidates:
            result[line] = PAYMENT_DIFFERENCE
            continue
        result[line] = PAYMENT_UNKNOWN
    return result


def _cohort_difference_candidates(
    history_groups: dict[tuple[str, str, str, str], list[SocialPaymentLine]],
    current_by_person: dict[tuple[str, str, str, str], list[SocialPaymentLine]],
    current_by_account: dict[tuple[str, str, str], list[SocialPaymentLine]],
) -> set[SocialPaymentLine]:
    """无明确性质时，仅在同一调整模式呈群体特征时推断为补差。"""
    candidates: dict[tuple[Any, ...], list[SocialPaymentLine]] = {}
    for person_key, group_lines in history_groups.items():
        _id_card, account, category, side = person_key
        history = [line for line in group_lines if line.nature_hint is None]
        current = current_by_person.get(person_key, [])
        periods = tuple(sorted({line.fee_period for line in history if line.fee_period}))
        if len(periods) < 2 or not current or not _same_history_profile(history):
            continue
        if not all(any(_is_lower_adjustment_profile(line, item) for item in current) for line in history):
            continue
        signature = (
            account,
            category,
            side,
            periods,
            _contribution_profile_key(history[0]),
            tuple(sorted({_contribution_profile_key(line) for line in current})),
        )
        candidates.setdefault(signature, []).extend(history)

    accepted: set[SocialPaymentLine] = set()
    for signature, lines in candidates.items():
        account, category, side = signature[:3]
        affected_people = {line.id_card for line in lines}
        current_people = {line.id_card for line in current_by_account.get((account, category, side), [])}
        # 单个人的历史费率/基数变化也可能是历史政策差异，不能仅凭金额自动当成补差。
        if len(affected_people) < 2:
            continue
        # 补差应呈现同账户同险种的群体调整特征；零散记录保留为待确认，避免误伤补缴。
        if current_people and len(affected_people) * 2 < len(current_people):
            continue
        accepted.update(lines)
    return accepted


def _contribution_profile_key(line: SocialPaymentLine) -> tuple[str, str, str]:
    return (
        "" if line.base is None else f"{line.base:.4f}",
        "" if line.rate is None else f"{line.rate:.8f}",
        f"{line.amount:.2f}",
    )


def _same_contribution_profile(
    left: SocialPaymentLine,
    right: SocialPaymentLine,
    *,
    require_basis_evidence: bool = False,
) -> bool:
    if not _amount_close(left.amount, right.amount):
        return False
    comparable_base = left.base is not None and right.base is not None
    comparable_rate = left.rate is not None and right.rate is not None
    if require_basis_evidence and not (comparable_base or comparable_rate):
        return False
    if comparable_base and not _amount_close(left.base, right.base):
        return False
    if comparable_rate and not _rate_close(left.rate, right.rate):
        return False
    return True


def _same_history_profile(lines: list[SocialPaymentLine]) -> bool:
    if not lines:
        return False
    first = lines[0]
    return all(_same_contribution_profile(first, line) for line in lines[1:])


def _is_lower_adjustment_profile(history: SocialPaymentLine, current: SocialPaymentLine) -> bool:
    if history.amount >= current.amount - 0.005:
        return False
    lower_rate = history.rate is not None and current.rate is not None and history.rate < current.rate - 0.0000001
    lower_base = history.base is not None and current.base is not None and history.base < current.base - 0.005
    return lower_rate or lower_base


def _amount_close(left: float, right: float) -> bool:
    return abs(left - right) <= 0.01


def _rate_close(left: float, right: float) -> bool:
    return abs(left - right) <= 0.0000001


def _line_periods(line: SocialPaymentLine) -> set[str]:
    if not line.fee_period:
        return set()
    return set(_period_sequence(line.fee_period, line.fee_period_end or line.fee_period))


def _append_payment_nature_warnings(record: DetailRecord, warnings: list[str]) -> None:
    difference_categories = [category for category in SOCIAL_CATEGORIES if category in record.difference_periods]
    if difference_categories:
        for category in difference_categories:
            multiple_base = len(record.difference_bases.get(category, set())) > 1
            multiple_rate = any(len(values) > 1 for values in record.difference_rates.get(category, {}).values())
            if multiple_base or multiple_rate:
                warnings.append(f"{record.name} {category}补差含多种基数或比例，基数/比例留空，金额按源表汇总。")
    unknown_categories = [category for category in SOCIAL_CATEGORIES if category in record.unknown_periods]
    if unknown_categories:
        periods = set().union(*(record.unknown_periods[category] for category in unknown_categories))
        message = f"待确认历史缴费：{_format_period_span(periods)}（{'、'.join(unknown_categories)}）"
        warnings.append(f"{record.name} {message}，金额已保留在普通缴费区，未自动归入补差。")


def _append_source_mismatch_warnings(
    line: SocialPaymentLine,
    person: RosterPerson,
    source_account: str,
    source_company: str,
    source_place: str,
    warnings: list[str],
    seen: set[tuple[str, str, str, str, str]],
) -> None:
    checks = [
        ("参保账户", source_account, person.account_display),
        ("公司", source_company, person.company),
        ("参保地", source_place, person.insured_place),
    ]
    for label, source_value, roster_value in checks:
        if not source_value or not roster_value or source_value == roster_value:
            continue
        key = (line.source_file, line.id_card, label, source_value, roster_value)
        if key in seen:
            continue
        seen.add(key)
        warnings.append(
            f"{line.source_file} 人员 {person.name} 的{label}与花名册不一致：账单/文件夹识别为“{source_value}”，花名册为“{roster_value}”。已按账单/文件夹识别结果写入。"
        )


def _write_detail_workbook(
    records: list[DetailRecord],
    output_file: Path,
    temp_dir: Path,
    *,
    cancelled: Callable[[], bool] | None = None,
) -> None:
    template_path = _copy_template(DETAIL_TEMPLATE_RESOURCE, temp_dir)
    workbook = load_workbook(template_path)
    try:
        ws = workbook[workbook.sheetnames[0]]
        ws.title = "社保明细表"
        title = _detail_title(records)
        if title:
            ws["A1"].value = title
        _write_difference_headers(ws, records)
        worksheet_max_column = ws.max_column
        template_snapshot = snapshot_row(ws, 4, worksheet_max_column)
        data_start = 4
        if records:
            insert_rows(ws, data_start + 1, max(len(records) - 1, 0))
        for offset, record in enumerate(records):
            if offset % 500 == 0:
                _check_cancelled(cancelled)
            row_index = data_start + offset
            apply_row_snapshot(ws, row_index, template_snapshot, translate_formulas=True)
            _clear_row_values(ws, row_index, worksheet_max_column)
            _write_detail_row(ws, row_index, offset + 1, record)
        if not records:
            _clear_row_values(ws, data_start, worksheet_max_column)
        _check_cancelled(cancelled)
        workbook.save(output_file)
    finally:
        workbook.close()


def _write_split_detail_workbooks(
    records: list[DetailRecord],
    output_dir: Path,
    temp_dir: Path,
    *,
    cancelled: Callable[[], bool] | None = None,
) -> list[Path]:
    grouped: OrderedDict[str, list[DetailRecord]] = OrderedDict()
    for index, record in enumerate(records):
        if index % 500 == 0:
            _check_cancelled(cancelled)
        grouped.setdefault(_detail_group_name(record), []).append(record)
    if not grouped:
        return []

    split_dir = output_dir / SPLIT_DETAIL_DIR_NAME
    split_dir.mkdir(parents=True, exist_ok=True)
    outputs: list[Path] = []
    for group_name, group_records in grouped.items():
        _check_cancelled(cancelled)
        output_file = split_dir / f"{_safe_file_stem(group_name)}-社保明细表.xlsx"
        _write_detail_workbook(
            group_records,
            output_file,
            temp_dir,
            cancelled=cancelled,
        )
        outputs.append(output_file)
    return outputs


def _write_detail_row(ws: Worksheet, row_index: int, sequence: int, record: DetailRecord) -> None:
    values: dict[int, Any] = {
        DETAIL_COLUMNS["序号"]: sequence,
        DETAIL_COLUMNS["公司"]: record.company,
        DETAIL_COLUMNS["参保地"]: record.insured_place,
        DETAIL_COLUMNS["项目"]: record.project_display,
        DETAIL_COLUMNS["姓名"]: record.name,
        DETAIL_COLUMNS["身份证"]: record.id_card,
        DETAIL_COLUMNS["参保时间"]: record.start_period,
        DETAIL_COLUMNS["账单期"]: record.period,
        DETAIL_COLUMNS["管理费"]: record.management_fee,
        DETAIL_COLUMNS["备注"]: "；".join(record.warnings),
    }
    for col_index, value in values.items():
        ws.cell(row_index, col_index).value = value

    _write_category_cells(ws, row_index, record, "养老", "养老")
    _write_category_cells(ws, row_index, record, "医疗", "医疗")
    _write_category_cells(ws, row_index, record, "失业", "失业")
    _write_category_cells(ws, row_index, record, "工伤", "工伤")
    _write_category_cells(ws, row_index, record, "补充工伤", "补充工伤")
    _write_category_cells(ws, row_index, record, "生育", "生育")
    _write_category_cells(ws, row_index, record, "大病医疗", "大病医疗")
    _write_difference_cells(ws, row_index, record)
    _write_difference_rollups(ws, row_index, record)
    _write_detail_totals(ws, row_index)
    _format_difference_rate_cells(ws, row_index, record)


def _write_category_cells(ws: Worksheet, row_index: int, record: DetailRecord, category: str, prefix: str) -> None:
    base_col = DETAIL_COLUMNS.get(f"{prefix}基数")
    if base_col is not None:
        ws.cell(row_index, base_col).value = record.bases.get(category)
    personal_rate_col = DETAIL_COLUMNS.get(f"{prefix}个人比例")
    if personal_rate_col is not None:
        ws.cell(row_index, personal_rate_col).value = record.rates.get(category, {}).get("个人")
    personal_amount_col = DETAIL_COLUMNS.get(f"{prefix}个人金额")
    if personal_amount_col is not None:
        _write_template_amount(
            ws,
            row_index,
            base_col,
            personal_rate_col,
            personal_amount_col,
            record.bases.get(category),
            record.rates.get(category, {}).get("个人"),
            record.amounts.get(category, {}).get("个人"),
        )
    unit_rate_col = DETAIL_COLUMNS.get(f"{prefix}单位比例")
    if unit_rate_col is not None:
        ws.cell(row_index, unit_rate_col).value = record.rates.get(category, {}).get("单位")
    unit_amount_col = DETAIL_COLUMNS.get(f"{prefix}单位金额")
    if unit_amount_col is not None:
        _write_template_amount(
            ws,
            row_index,
            base_col,
            unit_rate_col,
            unit_amount_col,
            record.bases.get(category),
            record.rates.get(category, {}).get("单位"),
            record.amounts.get(category, {}).get("单位"),
        )


def _write_template_amount(
    ws: Worksheet,
    row_index: int,
    base_col: int | None,
    rate_col: int | None,
    amount_col: int,
    base: float | None,
    rate: float | None,
    amount: float | None,
) -> None:
    """Use the client's formula when it reproduces the source amount exactly."""
    if not amount:
        ws.cell(row_index, amount_col).value = None
        return
    if base_col is not None and rate_col is not None and base is not None and rate is not None:
        calculated = round(base * rate, 2)
        if abs(amount - calculated) < 0.005:
            ws.cell(row_index, amount_col).value = (
                f"=ROUND({_cell_ref(base_col, row_index)}*{_cell_ref(rate_col, row_index)},2)"
            )
            return
    ws.cell(row_index, amount_col).value = round(amount, 2)


def _write_difference_headers(ws: Worksheet, records: list[DetailRecord]) -> None:
    periods: set[str] = set()
    for record in records:
        for category_periods in record.difference_periods.values():
            periods.update(category_periods)
    if not periods:
        return
    period_text = _format_chinese_period_span(periods)
    for category, column in (("养老", 44), ("失业", 49), ("工伤", 54), ("医疗", 57)):
        ws.cell(2, column).value = f"{category}{period_text}补差"


def _write_difference_cells(ws: Worksheet, row_index: int, record: DetailRecord) -> None:
    if not record.difference_amounts:
        return
    for category, columns in DIFFERENCE_COLUMNS.items():
        base = _single_number(record.difference_bases.get(category, set()))
        if base is not None:
            ws.cell(row_index, columns["基数"]).value = base
        personal_rate = _single_number(record.difference_rates.get(category, {}).get("个人", set()))
        if personal_rate is not None and "个人比例" in columns:
            ws.cell(row_index, columns["个人比例"]).value = personal_rate
        personal_amount = record.difference_amounts.get(category, {}).get("个人")
        if personal_amount and "个人金额" in columns:
            ws.cell(row_index, columns["个人金额"]).value = personal_amount
        unit_rate = _single_number(record.difference_rates.get(category, {}).get("单位", set()))
        if unit_rate is not None and "单位比例" in columns:
            ws.cell(row_index, columns["单位比例"]).value = unit_rate
        unit_amount = record.difference_amounts.get(category, {}).get("单位")
        if unit_amount and "单位金额" in columns:
            ws.cell(row_index, columns["单位金额"]).value = unit_amount


def _write_difference_rollups(ws: Worksheet, row_index: int, record: DetailRecord) -> None:
    personal_difference = round(
        record.difference_amounts.get("养老", {}).get("个人", 0.0)
        + record.difference_amounts.get("失业", {}).get("个人", 0.0),
        2,
    )
    unit_difference = round(
        record.difference_amounts.get("养老", {}).get("单位", 0.0)
        + record.difference_amounts.get("失业", {}).get("单位", 0.0)
        + record.difference_amounts.get("工伤", {}).get("单位", 0.0)
        + record.difference_amounts.get("医疗", {}).get("单位", 0.0),
        2,
    )
    _write_template_rollup(
        ws,
        row_index,
        DETAIL_COLUMNS["个人社保补缴合计"],
        [DIFFERENCE_COLUMNS["养老"]["个人金额"], DIFFERENCE_COLUMNS["失业"]["个人金额"]],
        personal_difference,
    )
    _write_template_rollup(
        ws,
        row_index,
        DETAIL_COLUMNS["单位社保补缴合计"],
        [
            DIFFERENCE_COLUMNS["养老"]["单位金额"],
            DIFFERENCE_COLUMNS["失业"]["单位金额"],
            DIFFERENCE_COLUMNS["工伤"]["单位金额"],
            DIFFERENCE_COLUMNS["医疗"]["单位金额"],
        ],
        unit_difference,
    )


def _write_template_rollup(
    ws: Worksheet,
    row_index: int,
    target_col: int,
    source_cols: list[int],
    difference_amount: float,
) -> None:
    if not difference_amount:
        ws.cell(row_index, target_col).value = None
        return
    formula = "+".join(_cell_ref(col, row_index) for col in source_cols)
    ws.cell(row_index, target_col).value = f"={formula}"


def _format_difference_rate_cells(ws: Worksheet, row_index: int, record: DetailRecord) -> None:
    """Only newly populated supplement rates use the client's percentage format."""
    for category, columns in DIFFERENCE_COLUMNS.items():
        rates = record.difference_rates.get(category, {})
        for side, key in (("个人", "个人比例"), ("单位", "单位比例")):
            if key not in columns or _single_number(rates.get(side, set())) is None:
                continue
            ws.cell(row_index, columns[key]).number_format = "0.00%"


def _single_number(values: set[float]) -> float | None:
    if len(values) != 1:
        return None
    return next(iter(values))


def _write_detail_totals(ws: Worksheet, row_index: int) -> None:
    personal_total = DETAIL_COLUMNS["个人社保合计"]
    unit_total = DETAIL_COLUMNS["单位社保合计"]
    social_total = DETAIL_COLUMNS["社保缴纳合计"]
    management_fee = DETAIL_COLUMNS["管理费"]
    tax = DETAIL_COLUMNS["税金"]
    total_fee = DETAIL_COLUMNS["费用总计"]
    ws.cell(row_index, personal_total).value = (
        f"={_cell_ref(DETAIL_COLUMNS['养老个人金额'], row_index)}"
        f"+{_cell_ref(DETAIL_COLUMNS['医疗个人金额'], row_index)}"
        f"+{_cell_ref(DETAIL_COLUMNS['失业个人金额'], row_index)}"
        f"+{_cell_ref(DETAIL_COLUMNS['大病医疗个人金额'], row_index)}"
        f"+{_cell_ref(DETAIL_COLUMNS['个人社保补缴合计'], row_index)}"
    )
    ws.cell(row_index, unit_total).value = (
        f"={_cell_ref(DETAIL_COLUMNS['养老单位金额'], row_index)}"
        f"+{_cell_ref(DETAIL_COLUMNS['医疗单位金额'], row_index)}"
        f"+{_cell_ref(DETAIL_COLUMNS['失业单位金额'], row_index)}"
        f"+{_cell_ref(DETAIL_COLUMNS['工伤单位金额'], row_index)}"
        f"+{_cell_ref(DETAIL_COLUMNS['大病医疗单位金额'], row_index)}"
        f"+{_cell_ref(DETAIL_COLUMNS['单位社保补缴合计'], row_index)}"
        f"+{_cell_ref(DETAIL_COLUMNS['补充工伤单位金额'], row_index)}"
        f"+{_cell_ref(DETAIL_COLUMNS['单位补缴滞纳金'], row_index)}"
    )
    ws.cell(row_index, social_total).value = f"={_cell_ref(personal_total, row_index)}+{_cell_ref(unit_total, row_index)}"
    ws.cell(row_index, tax).value = (
        f"=ROUND(({_cell_ref(social_total, row_index)}+{_cell_ref(management_fee, row_index)})*6.72%,2)"
    )
    ws.cell(row_index, total_fee).value = (
        f"={_cell_ref(social_total, row_index)}+{_cell_ref(management_fee, row_index)}+{_cell_ref(tax, row_index)}"
    )


def _write_summary_workbook(
    records: list[DetailRecord],
    output_file: Path,
    temp_dir: Path,
    warnings: list[str],
    *,
    cancelled: Callable[[], bool] | None = None,
) -> None:
    template_path = _copy_template(SUMMARY_TEMPLATE_RESOURCE, temp_dir)
    workbook = load_workbook(template_path)
    try:
        for sheet_name in list(workbook.sheetnames):
            workbook.remove(workbook[sheet_name])
        _create_summary_sheet(
            workbook,
            "社保汇总表",
            "社保汇总表",
            records,
            cancelled=cancelled,
        )
        records_by_company: dict[str, list[DetailRecord]] = {}
        for index, record in enumerate(records):
            if index % 500 == 0:
                _check_cancelled(cancelled)
            if record.company:
                records_by_company.setdefault(record.company, []).append(record)
        for company, company_records in sorted(records_by_company.items()):
            _check_cancelled(cancelled)
            _create_summary_sheet(
                workbook,
                company,
                f"{company}社保汇总表",
                company_records,
                cancelled=cancelled,
            )
        _check_cancelled(cancelled)
        _create_analysis_sheet(workbook, records)
        _create_warning_sheet(workbook, warnings)
        _check_cancelled(cancelled)
        workbook.save(output_file)
    finally:
        workbook.close()


def _create_summary_sheet(
    workbook,
    sheet_name: str,
    title: str,
    records: list[DetailRecord],
    *,
    cancelled: Callable[[], bool] | None = None,
) -> None:
    ws = workbook.create_sheet(_safe_sheet_title(sheet_name, workbook.sheetnames))
    accounts = sorted({record.account_display for record in records if record.account_display})
    if not accounts:
        accounts = ["未识别账户"]
    max_col = 5 + len(accounts) * 3
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(1, 1).value = title
    base_headers = ["项目/部门", "成本中心", "社保缴纳人数", "公积金缴纳人数", "费用合计"]
    for col_index, header in enumerate(base_headers, start=1):
        ws.cell(2, col_index).value = header
    for index, account in enumerate(accounts):
        start_col = 6 + index * 3
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 2)
        ws.cell(2, start_col).value = account
        ws.cell(3, start_col).value = "人数"
        ws.cell(3, start_col + 1).value = "个人"
        ws.cell(3, start_col + 2).value = "单位"
    grouped = _group_summary_records(records)
    for row_offset, ((project, cost_center), group_records) in enumerate(grouped.items(), start=4):
        if row_offset % 500 == 0:
            _check_cancelled(cancelled)
        ws.cell(row_offset, 1).value = project
        ws.cell(row_offset, 2).value = cost_center
        ws.cell(row_offset, 3).value = len({record.id_card for record in group_records})
        ws.cell(row_offset, 4).value = 0
        ws.cell(row_offset, 5).value = round(sum(record.total_fee for record in group_records), 2)
        group_records_by_account: dict[str, list[DetailRecord]] = {}
        for record in group_records:
            group_records_by_account.setdefault(record.account_display, []).append(record)
        for index, account in enumerate(accounts):
            account_records = group_records_by_account.get(account, [])
            start_col = 6 + index * 3
            ws.cell(row_offset, start_col).value = len({record.id_card for record in account_records})
            ws.cell(row_offset, start_col + 1).value = round(sum(record.personal_total for record in account_records), 2)
            ws.cell(row_offset, start_col + 2).value = round(sum(record.unit_total for record in account_records), 2)
    total_row = 4 + len(grouped)
    ws.cell(total_row, 1).value = "总计"
    if grouped:
        for col_index in range(3, max_col + 1):
            ws.cell(total_row, col_index).value = f"=SUM({get_column_letter(col_index)}4:{get_column_letter(col_index)}{total_row - 1})"
    _format_summary_sheet(ws, max(total_row, 4), max_col)


def _create_analysis_sheet(workbook, records: list[DetailRecord]) -> None:
    ws = workbook.create_sheet("数据分析")
    row_index = 1
    row_index = _write_analysis_section(ws, row_index, "按缴纳单位统计", records, "account_display", "缴纳单位")
    row_index += 2
    row_index = _write_place_difference_analysis(ws, row_index, records)
    row_index += 2
    row_index = _write_category_analysis(ws, row_index, records)
    row_index += 2
    row_index = _write_project_analysis(ws, row_index, records)
    _format_summary_sheet(ws, max(row_index, 1), 8)


def _write_analysis_section(ws: Worksheet, start_row: int, title: str, records: list[DetailRecord], attr: str, dimension_header: str = "维度") -> int:
    ws.cell(start_row, 1).value = title
    headers = [dimension_header, "人数", "个人金额", "单位金额", "社保合计", "税金", "费用总计"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(start_row + 1, col_index).value = header
    grouped: dict[str, list[DetailRecord]] = OrderedDict()
    for record in records:
        key = getattr(record, attr) or "未填写"
        grouped.setdefault(key, []).append(record)
    row_index = start_row + 2
    for key, group_records in grouped.items():
        ws.cell(row_index, 1).value = key
        ws.cell(row_index, 2).value = len({record.id_card for record in group_records})
        ws.cell(row_index, 3).value = round(sum(record.personal_total for record in group_records), 2)
        ws.cell(row_index, 4).value = round(sum(record.unit_total for record in group_records), 2)
        ws.cell(row_index, 5).value = round(sum(record.social_total for record in group_records), 2)
        ws.cell(row_index, 6).value = round(sum(record.tax for record in group_records), 2)
        ws.cell(row_index, 7).value = round(sum(record.total_fee for record in group_records), 2)
        row_index += 1
    return row_index


def _write_place_difference_analysis(ws: Worksheet, start_row: int, records: list[DetailRecord]) -> int:
    ws.cell(start_row, 1).value = "按参保地差异分析"
    headers = ["参保地", "人数", "个人金额", "单位金额", "总金额", "人均总金额", "与最低人均差异"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(start_row + 1, col_index).value = header
    grouped: OrderedDict[str, list[DetailRecord]] = OrderedDict()
    for record in records:
        grouped.setdefault(record.insured_place or "未填写", []).append(record)
    per_capita_values: dict[str, float] = {}
    for place, group_records in grouped.items():
        people_count = len({record.id_card for record in group_records})
        total = round(sum(record.social_total for record in group_records), 2)
        per_capita_values[place] = round(total / people_count, 2) if people_count else 0.0
    min_per_capita = min(per_capita_values.values()) if per_capita_values else 0.0
    row_index = start_row + 2
    for place, group_records in grouped.items():
        people_count = len({record.id_card for record in group_records})
        personal = round(sum(record.personal_total for record in group_records), 2)
        unit = round(sum(record.unit_total for record in group_records), 2)
        total = round(sum(record.social_total for record in group_records), 2)
        per_capita = per_capita_values[place]
        ws.cell(row_index, 1).value = place
        ws.cell(row_index, 2).value = people_count
        ws.cell(row_index, 3).value = personal
        ws.cell(row_index, 4).value = unit
        ws.cell(row_index, 5).value = total
        ws.cell(row_index, 6).value = per_capita
        ws.cell(row_index, 7).value = round(per_capita - min_per_capita, 2)
        row_index += 1
    return row_index


def _write_category_analysis(ws: Worksheet, start_row: int, records: list[DetailRecord]) -> int:
    ws.cell(start_row, 1).value = "按险种统计"
    headers = ["险种", "人数", "平均基数", "个人金额", "单位金额", "金额合计"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(start_row + 1, col_index).value = header
    row_index = start_row + 2
    for category in SOCIAL_CATEGORIES:
        category_records = [
            record
            for record in records
            if category in record.amounts or category in record.arrears_amounts or category in record.difference_amounts
        ]
        if not category_records:
            continue
        bases: list[float] = []
        for record in category_records:
            if category in record.bases:
                period_count = max(record.base_period_counts.get(category, 1), 1)
                bases.append(record.bases[category] / period_count)
            else:
                difference_base = _single_number(record.difference_bases.get(category, set()))
                if difference_base is not None:
                    bases.append(difference_base)
        personal = sum(
            record.amounts.get(category, {}).get("个人", 0.0)
            + record.arrears_amounts.get(category, {}).get("个人", 0.0)
            + record.difference_amounts.get(category, {}).get("个人", 0.0)
            for record in category_records
        )
        unit = sum(
            record.amounts.get(category, {}).get("单位", 0.0)
            + record.arrears_amounts.get(category, {}).get("单位", 0.0)
            + record.difference_amounts.get(category, {}).get("单位", 0.0)
            for record in category_records
        )
        ws.cell(row_index, 1).value = category
        ws.cell(row_index, 2).value = len({record.id_card for record in category_records})
        ws.cell(row_index, 3).value = round(sum(bases) / len(bases), 2) if bases else ""
        ws.cell(row_index, 4).value = round(personal, 2)
        ws.cell(row_index, 5).value = round(unit, 2)
        ws.cell(row_index, 6).value = round(personal + unit, 2)
        row_index += 1
    return row_index


def _write_project_analysis(ws: Worksheet, start_row: int, records: list[DetailRecord]) -> int:
    ws.cell(start_row, 1).value = "按项目统计社保费用"
    headers = ["项目/部门", "参保地数量", "参保地", "人数", "个人金额", "单位金额", "社保合计", "费用总计"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(start_row + 1, col_index).value = header
    grouped: OrderedDict[str, list[DetailRecord]] = OrderedDict()
    for record in records:
        grouped.setdefault(record.project_display or "未填写", []).append(record)
    row_index = start_row + 2
    for project, group_records in grouped.items():
        places = sorted({record.insured_place for record in group_records if record.insured_place})
        ws.cell(row_index, 1).value = project
        ws.cell(row_index, 2).value = len(places)
        ws.cell(row_index, 3).value = "、".join(places)
        ws.cell(row_index, 4).value = len({record.id_card for record in group_records})
        ws.cell(row_index, 5).value = round(sum(record.personal_total for record in group_records), 2)
        ws.cell(row_index, 6).value = round(sum(record.unit_total for record in group_records), 2)
        ws.cell(row_index, 7).value = round(sum(record.social_total for record in group_records), 2)
        ws.cell(row_index, 8).value = round(sum(record.total_fee for record in group_records), 2)
        row_index += 1
    return row_index


def _create_warning_sheet(workbook, warnings: list[str]) -> None:
    ws = workbook.create_sheet("异常提醒")
    ws.cell(1, 1).value = "序号"
    ws.cell(1, 2).value = "提醒内容"
    for index, warning in enumerate(warnings, start=1):
        ws.cell(index + 1, 1).value = index
        ws.cell(index + 1, 2).value = warning
    _format_summary_sheet(ws, max(len(warnings) + 1, 2), 2)
    ws.column_dimensions["B"].width = 90


def _group_summary_records(records: list[DetailRecord]) -> OrderedDict[tuple[str, str], list[DetailRecord]]:
    grouped: OrderedDict[tuple[str, str], list[DetailRecord]] = OrderedDict()
    for record in records:
        key = (record.project_display or "未填写", record.cost_center)
        grouped.setdefault(key, []).append(record)
    return grouped


def _copy_template(resource_name: str, temp_dir: Path) -> Path:
    target = temp_dir / resource_name
    with open_template_resource(resource_name) as source, target.open("wb") as output:
        shutil.copyfileobj(source, output)
    return target


def _find_header_row(ws: Worksheet, required_headers: tuple[str, ...]) -> int:
    normalized_required = {_normalize_header(header) for header in required_headers}
    max_column = ws.max_column or 0
    for row_index in range(1, min(ws.max_row or 0, 30) + 1):
        values = {
            _normalize_header(ws.cell(row_index, col_index).value)
            for col_index in range(1, max_column + 1)
        }
        if normalized_required.issubset(values):
            return row_index
    raise ValueError(f"{ws.title} 未找到表头：{'、'.join(required_headers)}")


def _find_payment_header_row(ws: Worksheet) -> int:
    max_column = ws.max_column or 0
    for row_index in range(1, min(ws.max_row or 0, 20) + 1):
        values = {
            _normalize_header(ws.cell(row_index, col_index).value)
            for col_index in range(1, max_column + 1)
        }
        if "姓名" in values and ("证件号码" in values or "身份证件号码" in values):
            return row_index
    raise ValueError("未找到包含姓名和证件号码的表头。")


def _read_headers(ws: Worksheet, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_index in range(1, (ws.max_column or 0) + 1):
        header = _normalize_header(ws.cell(header_row, col_index).value)
        if header:
            headers[header] = col_index
    return headers


def _is_non_source_excel(path: Path) -> bool:
    return any(keyword in path.name for keyword in ("模板", "花名册"))


def _is_single_kind_sheet(headers: dict[str, int]) -> bool:
    return {"姓名", "缴费基数", "费率"}.issubset(headers) and any(header in headers for header in ("应缴费额(元)", "本期应缴费额"))


def _has_wide_amount_columns(headers: dict[str, int]) -> bool:
    amount_headers = [header for header in headers if "应缴费额" in header]
    return len(amount_headers) > 1 or any("保险" in header for header in amount_headers)


def _header_value(ws: Worksheet, row_index: int, headers: dict[str, int], header: str) -> Any:
    col_index = headers.get(header)
    if col_index is None:
        return None
    return ws.cell(row_index, col_index).value


def _xls_header_value(sheet, row_index: int, headers: dict[str, int], header: str) -> Any:
    col_index = headers.get(header)
    if col_index is None:
        return None
    return sheet.cell_value(row_index, col_index)


def _classify_insurance_item(type_text: str, item_text: str) -> tuple[str | None, str]:
    text = f"{type_text} {item_text}"
    side = "个人" if "个人" in text else "单位"
    if "补充工伤" in text:
        return "补充工伤", "单位"
    if "大额" in text or "大病" in text or "互助" in text:
        return "大病医疗", side
    if "养老" in text:
        return "养老", side
    if "医疗" in text:
        return "医疗", side
    if "失业" in text:
        return "失业", side
    if "工伤" in text:
        return "工伤", "单位"
    if "生育" in text:
        return "生育", "单位"
    return None, side


def _source_context(file_path: Path) -> SourceContext:
    candidates = _source_name_candidates(file_path)
    hint_text = " ".join(candidates)
    file_period = _period_from_filename(file_path.name) or ""
    period_split_file = bool(_FILENAME_HAS_PERIOD.search(file_path.name))
    container_period = ""
    for candidate in candidates[:-1]:
        container_period = _period_from_filename(candidate) or ""
        if container_period:
            break
    account_hint = ""
    for candidate in candidates:
        account_hint = _known_account_hint(candidate)
        if account_hint:
            break
    return SourceContext(
        label=file_path.name,
        file_period=file_period,
        container_period=container_period,
        billing_period_hint=(file_period if file_period and not period_split_file else "") or (container_period if not file_period else ""),
        period_split_file=period_split_file,
        account_hint=account_hint,
        company_hint=_company_hint_from_filename(hint_text),
        nature_hint=_payment_nature_from_text(file_path.name),
    )


def _source_name_candidates(file_path: Path) -> list[str]:
    candidates: list[str] = []
    for parent in (file_path.parent, file_path.parent.parent, file_path.parent.parent.parent):
        name = parent.name
        if not name or name.startswith(("zip_", "hr_social_security_", "xls_converted")):
            continue
        if re.fullmatch(r"[a-f0-9]{12}", name):
            continue
        candidates.append(name)
    candidates.append(file_path.name)
    return candidates


def _period_from_filename(file_name: str) -> str | None:
    match = _PERIOD_FROM_FILENAME.search(file_name)
    if not match:
        return None
    year = int(match.group(1))
    month = int(match.group(2))
    if 1 <= month <= 12:
        return f"{year}{month:02d}"
    return None


def _period_from_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return f"{value.year}{value.month:02d}"
    if isinstance(value, (int, float)) and 20000 <= float(value) <= 80000:
        converted = datetime(1899, 12, 30) + timedelta(days=int(value))
        return f"{converted.year}{converted.month:02d}"
    text = _cell_text(value)
    if not text:
        return None
    match = _PERIOD_FROM_TEXT.search(text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return f"{year}{month:02d}"
    return None


def _payment_periods_from_row(row: dict[str, Any], context: SourceContext) -> tuple[str, str]:
    start = ""
    for header in FEE_PERIOD_START_HEADERS:
        start = _period_from_value(row.get(header)) or ""
        if start:
            break
    end = ""
    for header in FEE_PERIOD_END_HEADERS:
        end = _period_from_value(row.get(header)) or ""
        if end:
            break
    start = start or context.file_period or context.container_period
    end = end or start
    if start and end and end < start:
        start, end = end, start
    return start, end


def _nature_hint_from_row(row: dict[str, Any], context: SourceContext, *extra_values: Any) -> str | None:
    values = [row.get(header) for header in PAYMENT_NATURE_HEADERS]
    values.extend(extra_values)
    return _payment_nature_from_text(*values) or context.nature_hint


def _payment_nature_from_text(*values: Any) -> str | None:
    text = " ".join(_cell_text(value) for value in values if value is not None)
    if not text:
        return None
    difference = any(keyword in text for keyword in ("补差", "差额", "基数调整", "基数补差", "调整补收"))
    explicit_arrears = any(keyword in text for keyword in ("补缴", "追缴"))
    if difference and not explicit_arrears:
        return PAYMENT_DIFFERENCE
    if explicit_arrears and not difference:
        return PAYMENT_ARREARS
    if "补收" in text and not difference:
        return PAYMENT_ARREARS
    if any(keyword in text for keyword in ("正常缴费", "正常征收", "当期缴费")):
        return PAYMENT_NORMAL
    return None


def _period_sequence(start: str, end: str) -> list[str]:
    if not _is_period(start) or not _is_period(end):
        return [start] if start else []
    start_index = int(start[:4]) * 12 + int(start[4:]) - 1
    end_index = int(end[:4]) * 12 + int(end[4:]) - 1
    if end_index < start_index:
        start_index, end_index = end_index, start_index
    periods: list[str] = []
    for index in range(start_index, end_index + 1):
        year, month_index = divmod(index, 12)
        periods.append(f"{year}{month_index + 1:02d}")
    return periods


def _format_period_span(periods: set[str]) -> str:
    normalized = sorted(period for period in periods if _is_period(period))
    if not normalized:
        return ""
    if len(normalized) == 1:
        return normalized[0]
    contiguous = normalized == _period_sequence(normalized[0], normalized[-1])
    if contiguous:
        return f"{normalized[0]}-{normalized[-1]}"
    return "、".join(normalized)


def _format_chinese_period_span(periods: set[str]) -> str:
    normalized = sorted(period for period in periods if _is_period(period))
    if not normalized:
        return ""
    start = normalized[0]
    end = normalized[-1]
    if start == end:
        return f"{start[:4]}年{int(start[4:]):d}月"
    if start[:4] == end[:4] and normalized == _period_sequence(start, end):
        return f"{start[:4]}年{int(start[4:]):d}月-{int(end[4:]):d}月"
    return f"{start[:4]}年{int(start[4:]):d}月-{end[:4]}年{int(end[4:]):d}月"


def _is_period(value: str) -> bool:
    return bool(re.fullmatch(r"20\d{4}", value)) and 1 <= int(value[4:]) <= 12


def _account_hint_from_filename(file_name: str) -> str:
    known = _known_account_hint(file_name)
    if known:
        return known
    if _FILENAME_HAS_PERIOD.search(file_name):
        return ""
    if any(keyword in file_name for keyword in SOCIAL_CATEGORIES):
        return ""
    cleaned = _FILENAME_STRIP_PERIOD.sub("", file_name).strip(" -_")
    return cleaned


def _known_account_hint(file_name: str) -> str:
    for keyword in ("北京春苗抚州", "唐人四川", "唐人长春", "北京春苗", "唐人数智"):
        if keyword in file_name:
            return keyword
    return ""


def _company_hint_from_filename(file_name: str) -> str:
    if "春苗" in file_name:
        return "北京春苗"
    if "唐人" in file_name:
        return "唐人数智"
    return ""


def _account_display(account: str) -> str:
    if not account:
        return ""
    if account in ACCOUNT_DISPLAY_ALIASES:
        return ACCOUNT_DISPLAY_ALIASES[account]
    return account


def _company_display(unit: str, account: str) -> str:
    for full_name, short_name in COMPANY_ALIASES.items():
        if full_name in unit:
            return short_name
    if "春苗" in unit or "春苗" in account:
        return "北京春苗"
    if "唐人" in unit or "唐人" in account:
        return "唐人数智"
    return unit or account


def _insured_place(account: str) -> str:
    for place in ("四川", "长春", "抚州", "成都", "南昌", "北京"):
        if place in account:
            return place
    return account


def _project_display(project: str, cost_center: str, department: str) -> str:
    if project:
        bracket_match = _BRACKET_CONTENT.search(project)
        if bracket_match:
            return bracket_match.group(1).strip()
        return project
    return cost_center or department or "未填写"


def _valid_person_row(name: str, id_card: str) -> bool:
    if not name or name in {"合计", "总计"}:
        return False
    if not id_card or id_card in {"——", "--", "-"}:
        return False
    return True


def _count_records(records: list[DetailRecord], attr: str) -> dict[str, int]:
    counts: dict[str, set[str]] = OrderedDict()
    for record in records:
        key = getattr(record, attr) or "未填写"
        counts.setdefault(key, set()).add(record.id_card)
    return {key: len(values) for key, values in counts.items()}


def _detail_title(records: list[DetailRecord]) -> str:
    companies = sorted({record.company for record in records if record.company})
    periods = sorted({record.billing_period for record in records if _is_period(record.billing_period)})
    company_part = "、".join(companies) if companies else "社保"
    if len(periods) == 1:
        period_text = f"{periods[0][:4]}年{int(periods[0][4:]):d}月"
    elif periods:
        period_text = f"{periods[0][:4]}年{int(periods[0][4:]):d}月-{periods[-1][:4]}年{int(periods[-1][4:]):d}月"
    else:
        period_text = ""
    return f"{company_part}{period_text}社保明细表"


def _clear_row_values(ws: Worksheet, row_index: int, max_column: int) -> None:
    for col_index in range(1, max_column + 1):
        ws.cell(row_index, col_index).value = None


_DATA_ROW_SIDE = Side(style="thin", color="000000")
_DATA_ROW_BORDER = Border(left=_DATA_ROW_SIDE, right=_DATA_ROW_SIDE, top=_DATA_ROW_SIDE, bottom=_DATA_ROW_SIDE)
_DATA_ROW_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ROW_FONT = Font(name="宋体", size=10)


def _format_data_row(ws: Worksheet, row_index: int, max_column: int) -> None:
    # 整行是固定样式，逐格赋值会让 openpyxl 每次都递归哈希样式对象去查表；
    # 样式表只增不改，下标查一次即可重复使用。
    border_id = cached_style_id(ws, "border", "detail_thin", lambda: _DATA_ROW_BORDER)
    alignment_id = cached_style_id(ws, "alignment", "detail_center", lambda: _DATA_ROW_ALIGNMENT)
    font_id = cached_style_id(ws, "font", "detail_song10", lambda: _DATA_ROW_FONT)
    for col_index in range(1, max_column + 1):
        cell = ws.cell(row_index, col_index)
        set_style_ids(cell, border_id=border_id, alignment_id=alignment_id, font_id=font_id)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0.00"


_SUMMARY_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
_SUMMARY_TITLE_FILL = PatternFill("solid", fgColor="FFFFFF")
_SUMMARY_TITLE_FONT = Font(name="宋体", size=14, bold=True)
_SUMMARY_HEADER_FONT = Font(name="宋体", size=10, bold=True)


def _format_summary_sheet(ws: Worksheet, max_row: int, max_col: int) -> None:
    border_id = cached_style_id(ws, "border", "detail_thin", lambda: _DATA_ROW_BORDER)
    alignment_id = cached_style_id(ws, "alignment", "detail_center", lambda: _DATA_ROW_ALIGNMENT)
    title_font_id = cached_style_id(ws, "font", "summary_title", lambda: _SUMMARY_TITLE_FONT)
    header_font_id = cached_style_id(ws, "font", "summary_header", lambda: _SUMMARY_HEADER_FONT)
    normal_font_id = cached_style_id(ws, "font", "detail_song10", lambda: _DATA_ROW_FONT)
    title_fill_id = cached_style_id(ws, "fill", "summary_title_fill", lambda: _SUMMARY_TITLE_FILL)
    header_fill_id = cached_style_id(ws, "fill", "summary_header_fill", lambda: _SUMMARY_HEADER_FILL)
    for row_index in range(1, max_row + 1):
        is_title = row_index == 1
        is_header = row_index in {2, 3}
        font_id = title_font_id if is_title else header_font_id if is_header else normal_font_id
        fill_id = title_fill_id if is_title else header_fill_id if is_header else None
        for col_index in range(1, max_col + 1):
            cell = ws.cell(row_index, col_index)
            set_style_ids(cell, border_id=border_id, alignment_id=alignment_id, font_id=font_id, fill_id=fill_id)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    for col_index in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col_index)].width = 16 if col_index <= 5 else 12
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A4"


def _sum_formula(cols: list[int], row_index: int) -> str:
    refs = [_cell_ref(col, row_index) for col in cols]
    return f"=SUM({','.join(refs)})"


def _cell_ref(col_index: int, row_index: int) -> str:
    return f"{get_column_letter(col_index)}{row_index}"


def _safe_sheet_title(title: str, existing_titles: list[str]) -> str:
    cleaned = _SHEET_TITLE_INVALID.sub("_", title).strip()[:31] or "未命名"
    if cleaned not in existing_titles:
        return cleaned
    base = cleaned[:28]
    counter = 1
    while f"{base}_{counter}" in existing_titles:
        counter += 1
    return f"{base}_{counter}"


def _detail_group_name(record: DetailRecord) -> str:
    account = record.account_display or record.company or "未识别参保单位"
    place = record.insured_place or "未识别参保地"
    if place and place not in account:
        return f"{account}-{place}"
    return account


def _safe_file_stem(name: str) -> str:
    cleaned = _FILENAME_INVALID.sub("_", name).strip()
    return cleaned or "未命名"


def _normalize_header(value: Any) -> str:
    return _HEADER_WHITESPACE.sub("", _cell_text(value))


def _normalize_id_card(value: Any) -> str:
    return _cell_text(value).upper()


def _normalize_name(value: Any) -> str:
    return _HEADER_WHITESPACE.sub("", _cell_text(value))



def _is_binary_xls(path: Path) -> bool:
    with path.open("rb") as file:
        return file.read(8).startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1")


def _to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _cell_text(value)
    if not text:
        return None
    text = text.replace(",", "").replace("，", "").replace("元", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def _rate_to_decimal(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _cell_text(value)
    if not text:
        return None
    try:
        if text.endswith("%"):
            return float(text.rstrip("%")) / 100
        return float(text)
    except ValueError:
        return None


def _zero_blank(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value, 2) if value else None
