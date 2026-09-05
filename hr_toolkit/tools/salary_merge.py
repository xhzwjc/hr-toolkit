from __future__ import annotations

import ast
import hashlib
import re


# 预编译正则
_PERIOD_COMPACT = re.compile(r"(20\d{2})\D{0,3}([01]?\d)")
_PERIOD_PLAIN = re.compile(r"(20\d{2})([01]\d)")
import tempfile
from collections import Counter, OrderedDict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from hr_toolkit.common.excel import SheetGrid, cached_style_id, set_style_ids
from hr_toolkit.common.excel_compat import is_supported_excel_file, ensure_xlsx_workbook
from hr_toolkit.common.inputs import (
    extract_archive_excel_files,
    is_supported_archive_file,
    normalize_input_paths,
)
from .salary_headers import inspect_workbook


TOOL_NAME = "需求5-多月工资合并个人薪资汇总"
DETAIL_SHEET_KEYWORD = "明细"
SUMMARY_SHEET_KEYWORD = "汇总"
SUMMARY_TITLE = "唐人数智科技股份有限公司广东分公司（河源项目部）个人应发工资汇总表"
HEADER_NAME = "姓名"
HEADER_ID_CARD = "身份证号码"
HEADER_ID_CARD_ALIASES = (HEADER_ID_CARD, "身份证号")
HEADER_AMOUNT = "应发小计"
HEADER_AMOUNT_ALIASES = (HEADER_AMOUNT, "本月应发工资")
AMOUNT_NUMBER_FORMAT = '#,##0.00;[Red]-#,##0.00;0'
_MONTHLY_AMOUNT_FORMULA = re.compile(
    r"=ROUND\(SUM\(([A-Z]{1,3}\d+):([A-Z]{1,3}\d+)\)"
    r"-SUM\(([A-Z]{1,3}\d+):([A-Z]{1,3}\d+)\),2\)"
)
_SALARY_CELL_REF = re.compile(r"([A-Z]{1,3})(\d+)")


@dataclass
class SalaryRecord:
    month: str
    name: str
    id_card: str
    amount: float
    source_file: str
    source_row: int


@dataclass
class SalaryMergeResult:
    input_dir: Path
    output_dir: Path
    existing_summary_path: Path | None = None
    output_file: Path | None = None
    dry_run: bool = False
    source_files: list[str] = field(default_factory=list)
    months: list[str] = field(default_factory=list)
    employee_count: int = 0
    record_count: int = 0
    applied_record_count: int = 0
    skipped_record_count: int = 0
    warnings: list[str] = field(default_factory=list)
    excluded_files: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "tool_name": TOOL_NAME,
            "input_dir": str(self.input_dir),
            "output_dir": str(self.output_dir),
            "existing_summary_path": None
            if self.existing_summary_path is None
            else str(self.existing_summary_path),
            "output_file": None if self.output_file is None else str(self.output_file),
            "dry_run": self.dry_run,
            "source_files": self.source_files,
            "source_file_count": len(self.source_files),
            "months": self.months,
            "employee_count": self.employee_count,
            "record_count": self.record_count,
            "applied_record_count": self.applied_record_count,
            "skipped_record_count": self.skipped_record_count,
            "warnings": self.warnings,
            "excluded_files": self.excluded_files,
        }


@dataclass
class MergedEmployee:
    name: str
    id_card: str
    amounts: OrderedDict[str, float]


@dataclass(frozen=True)
class SalarySourceLayout:
    detail_sheet_name: str
    header_row: int
    data_start_row: int
    name_col: int
    id_card_col: int
    amount_col: int
    legacy_amount_fallback: bool = True


def merge_monthly_salary(
    input_dir: str | Path | list[str | Path],
    output_dir: str | Path,
    *,
    existing_summary_path: str | Path | None = None,
    year: int | None = None,
    dry_run: bool = False,
    cancelled: Callable[[], bool] | None = None,
    header_profiles: dict[str, Any] | None = None,
    strict_headers: bool = False,
    skipped_source_keys: list[str] | None = None,
    expected_sources: list[str] | None = None,
    confirmed_layouts: dict[str, dict[str, Any]] | None = None,
) -> SalaryMergeResult:
    _check_cancelled(cancelled)
    input_paths = _normalize_input_paths(input_dir)
    display_input = input_paths[0] if len(input_paths) == 1 else input_paths[0].parent
    output_dir = Path(output_dir).expanduser().resolve()
    summary_path = None
    if existing_summary_path:
        summary_path = Path(existing_summary_path).expanduser().resolve()
        if not summary_path.exists() or not summary_path.is_file():
            raise FileNotFoundError(f"已有汇总表不存在：{summary_path}")

    for input_path in input_paths:
        _check_cancelled(cancelled)
        if not input_path.exists():
            raise FileNotFoundError(f"工资表文件、压缩包或文件夹不存在：{input_path}")

    warnings: list[str] = []
    with tempfile.TemporaryDirectory(prefix="hr_salary_merge_") as temp_root:
        temp_dir = Path(temp_root)
        working_summary_path = None if summary_path is None else ensure_xlsx_workbook(summary_path, temp_dir)
        source_keys: dict[Path, str] = {}
        identify_sources = strict_headers or expected_sources is not None or bool(skipped_source_keys) or bool(confirmed_layouts)
        salary_files = _find_salary_files(
            input_paths, temp_dir, summary_path, warnings,
            source_keys=source_keys if identify_sources else None, cancelled=cancelled,
        )
        if strict_headers and warnings:
            raise ValueError("部分输入未能完整读取，请处理后重新合并：\n" + "\n".join(warnings[:10]))
        actual_sources = list(source_keys.values())
        summary_key = ""
        if identify_sources and summary_path is not None:
            summary_key = "summary:" + _salary_source_key(summary_path, cancelled)
            actual_sources.append(summary_key)
        if expected_sources is not None and Counter(actual_sources) != Counter(expected_sources):
            raise ValueError("所选资料在列头确认后发生变化，请重新点击开始合并并确认列头")
        _check_cancelled(cancelled)
        if not salary_files:
            raise ValueError("未在所选路径中找到 .xlsx 或 .xls 工资表")

        records: list[SalaryRecord] = []
        source_files: list[str] = []
        excluded_files: list[str] = []
        skipped_keys = set(skipped_source_keys or [])
        for file_path in salary_files:
            _check_cancelled(cancelled)
            if source_keys.get(file_path) in skipped_keys:
                excluded_files.append(file_path.name)
                warnings.append(f"用户选择不合并：{file_path.name}")
                continue
            try:
                month = _detect_month(file_path)
                file_records, file_warnings = _read_salary_file(
                    file_path,
                    month,
                    cancelled=cancelled,
                    header_profiles=header_profiles,
                    layout_hint=(confirmed_layouts or {}).get(source_keys.get(file_path, "")),
                )
            except ValueError as exc:
                if strict_headers:
                    raise ValueError(f"{file_path.name} 未完成合并，本次未生成汇总结果：{exc}") from exc
                warnings.append(f"{file_path.name} 不是有效月度工资表，已跳过：{exc}")
                continue
            records.extend(file_records)
            warnings.extend(file_warnings)
            source_files.append(str(file_path))

        if not records:
            message = "未识别到可合并的月度工资记录，请确认路径中包含需求4格式的月度工资表"
            if warnings:
                message += "\n" + "\n".join(warnings[:3])
                if len(warnings) > 3:
                    message += f"\n另有 {len(warnings) - 3} 条提示"
            raise ValueError(message)

        existing_employees: list[MergedEmployee] | None = None
        existing_months: list[str] | None = None
        if working_summary_path is not None:
            _check_cancelled(cancelled)
            existing_months, existing_employees, summary_warnings = _read_existing_summary(
                working_summary_path, header_profiles=header_profiles,
                layout_hint=(confirmed_layouts or {}).get(summary_key),
            )
            warnings.extend(summary_warnings)

        months = _build_output_months(records, year, existing_months)
        employees, merge_warnings, applied_count, skipped_count = _merge_records(
            records,
            months,
            existing_employees=existing_employees,
        )
        warnings.extend(merge_warnings)

        result = SalaryMergeResult(
            input_dir=display_input,
            output_dir=output_dir,
            existing_summary_path=summary_path,
            dry_run=dry_run,
            source_files=source_files,
            months=months,
            employee_count=len(employees),
            record_count=len(records),
            applied_record_count=applied_count,
            skipped_record_count=skipped_count,
            warnings=warnings,
            excluded_files=excluded_files,
        )

        if dry_run:
            return result

        _check_cancelled(cancelled)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_file = output_dir / "个人薪资汇总表.xlsx"
        _write_output_workbook(
            output_file,
            employees,
            months,
            cancelled=cancelled,
        )
        if excluded_files:
            workbook = load_workbook(output_file)
            try:
                sheet = workbook.create_sheet("未合并文件")
                sheet.append(["文件名", "原因"])
                sheet.column_dimensions["A"].width = 60
                sheet.column_dimensions["B"].width = 45
                for filename in excluded_files:
                    sheet.append([filename, "用户在列头对应窗口中明确选择不合并"])
                workbook.save(output_file)
            finally:
                workbook.close()
        result.output_file = output_file
        return result


def _normalize_input_paths(input_path: str | Path | list[str | Path]) -> list[Path]:
    return normalize_input_paths(input_path, "请选择工资表文件、压缩包或文件夹。")


def _find_salary_files(
    input_paths: list[Path], temp_dir: Path, existing_summary_path: Path | None = None,
    warnings: list[str] | None = None, *, source_keys: dict[Path, str] | None = None,
    cancelled: Callable[[], bool] | None = None,
) -> list[Path]:
    warnings = [] if warnings is None else warnings
    excluded = None if existing_summary_path is None else existing_summary_path.resolve()
    files: list[Path] = []
    seen: set[Path] = set()
    for input_path in input_paths:
        for path in _iter_salary_files(input_path, temp_dir, warnings):
            _check_cancelled(cancelled)
            if path.name == "个人薪资汇总表.xlsx":
                continue
            if excluded is not None and path.resolve() == excluded:
                continue
            working_path = ensure_xlsx_workbook(path, temp_dir)
            resolved = working_path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            files.append(working_path)
            if source_keys is not None:
                source_keys[working_path] = _salary_source_key(path, cancelled)
    return sorted(files)


def _salary_source_key(path: Path, cancelled=None) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while True:
            _check_cancelled(cancelled)
            chunk = stream.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    # 相同内容可能按不同文件名代表不同月份，跳过决定不能误伤另一月份。
    return digest.hexdigest() + ":" + (_month_from_text(path.stem) or "")


def inspect_salary_templates(
    input_dir: str | Path | list[str | Path], *, existing_summary_path: str | Path | None = None,
    header_profiles: dict[str, Any] | None = None, layout_hints: dict[str, Any] | None = None,
    cancelled: Callable[[], bool] | None = None, progress_callback=None,
) -> dict[str, Any]:
    """后台顺序预检，返回表头与少量样例；不保留工作簿或生成工资结果。"""
    profiles, hints = header_profiles or {}, layout_hints or {}
    groups: dict[str, dict[str, Any]] = {}
    issues, expected_sources = [], []
    summary = Path(existing_summary_path).expanduser().resolve() if existing_summary_path else None
    with tempfile.TemporaryDirectory(prefix="hr_salary_headers_") as temp_root:
        temp = Path(temp_root)
        source_keys: dict[Path, str] = {}
        warnings: list[str] = []
        files = _find_salary_files(_normalize_input_paths(input_dir), temp, summary, warnings,
                                   source_keys=source_keys, cancelled=cancelled)
        if not files:
            raise ValueError("未在所选路径中找到 .xlsx 或 .xls 工资表")
        if warnings:
            raise ValueError("部分输入未能完整读取，请检查：\n" + "\n".join(warnings[:10]))
        sources = [(path, source_keys[path], "detail") for path in files]
        if summary is not None:
            sources.append((ensure_xlsx_workbook(summary, temp), "summary:" + _salary_source_key(summary, cancelled), "summary"))
        for index, (path, source_key, role) in enumerate(sources):
            _check_cancelled(cancelled)
            expected_sources.append(source_key)
            if progress_callback:
                progress_callback(index, len(sources), f"正在检查列头 {index + 1}/{len(sources)}：{path.name}")
            try:
                if role == "detail":
                    _detect_month(path)
                workbook = load_workbook(path, read_only=True, data_only=False)
                try:
                    group = inspect_workbook(workbook, role=role, profiles=profiles, hint=hints.get(source_key))
                    if role == "summary" and not _read_month_columns(workbook[group["sheet"]], group["header_bottom"]):
                        raise ValueError("已有汇总表未找到月份列，请确认表头包含 202601 这类月份")
                finally:
                    workbook.close()
                key = group["key"]
                # 同名列调整顺序后不得沿用另一文件中的物理列号。
                if key in groups and group["order"] != groups[key]["order"]:
                    counts = Counter(c["key"] for c in group["columns"])
                    if any(count > 1 for count in counts.values()):
                        key += ":" + group["order"]
                if key not in groups:
                    group["group_id"] = key
                    groups[key] = group
                groups[key]["files"].append({
                    "key": source_key, "name": path.name,
                    "layout": {field: group[field] for field in ("sheet", "header_row", "header_bottom")},
                })
            except Exception as exc:
                _check_cancelled(cancelled)
                issues.append({"key": source_key, "name": path.name, "message": str(exc), "skippable": role == "detail"})
            if progress_callback:
                progress_callback(index + 1, len(sources), f"已检查列头 {index + 1}/{len(sources)}")
    return {"groups": list(groups.values()), "issues": issues, "expected_sources": expected_sources}


def _iter_salary_files(input_path: Path, temp_dir: Path, warnings: list[str]) -> list[Path]:
    if input_path.is_file():
        if is_supported_excel_file(input_path):
            return [input_path]
        if is_supported_archive_file(input_path):
            return extract_archive_excel_files(input_path, temp_dir, warnings)
        return []
    if not input_path.is_dir():
        return []
    files: list[Path] = []
    for child in sorted(input_path.rglob("*")):
        if not child.is_file() or child.name.startswith("~$"):
            continue
        if is_supported_excel_file(child):
            files.append(child)
        elif is_supported_archive_file(child):
            files.extend(extract_archive_excel_files(child, temp_dir, warnings))
    return files


def _check_cancelled(cancelled: Callable[[], bool] | None) -> None:
    if cancelled is not None and cancelled():
        raise RuntimeError("本次处理已停止。")


def _read_salary_file(
    file_path: Path,
    month: str,
    *,
    cancelled: Callable[[], bool] | None = None,
    header_profiles: dict[str, Any] | None = None,
    layout_hint: dict[str, Any] | None = None,
) -> tuple[list[SalaryRecord], list[str]]:
    warnings: list[str] = []
    formula_wb = load_workbook(file_path, data_only=False)
    try:
        value_wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            layout = _detect_source_layout(formula_wb, header_profiles=header_profiles, layout_hint=layout_hint)
            formula_ws = formula_wb[layout.detail_sheet_name]
            value_ws = SheetGrid(value_wb[layout.detail_sheet_name])
        finally:
            value_wb.close()

        records: list[SalaryRecord] = []
        for row_index in range(layout.data_start_row, formula_ws.max_row + 1):
            if row_index % 500 == 0:
                _check_cancelled(cancelled)
            id_card = _cell_text(value_ws, row_index, layout.id_card_col)
            name = _cell_text(value_ws, row_index, layout.name_col)
            if not id_card and not name:
                continue
            if not id_card:
                warnings.append(f"{file_path.name} 第 {row_index} 行缺少身份证号码，已跳过")
                continue
            amount = _amount_value(value_ws, formula_ws, row_index, layout)
            if amount is None:
                if not layout.legacy_amount_fallback:
                    raise ValueError(f"{layout.detail_sheet_name}!{get_column_letter(layout.amount_col)}{row_index} 未读到有效金额，请核对所选应发工资列及公式计算结果")
                warnings.append(f"{file_path.name} 第 {row_index} 行未识别到应发工资，按 0 处理")
                amount = 0
            records.append(
                SalaryRecord(
                    month=month,
                    name=name,
                    id_card=id_card,
                    amount=amount,
                    source_file=file_path.name,
                    source_row=row_index,
                )
            )
        return records, warnings
    finally:
        formula_wb.close()


def _detect_source_layout(
    workbook, *, header_profiles: dict[str, Any] | None = None, layout_hint: dict[str, Any] | None = None,
) -> SalarySourceLayout:
    if header_profiles is not None:
        group = inspect_workbook(workbook, profiles=header_profiles, hint=layout_hint)
        if not group["ready"]:
            raise ValueError(group["problem"])
        selected = group["selections"]
        return SalarySourceLayout(
            group["sheet"], group["header_row"],
            _find_data_start_row(workbook[group["sheet"]], group["header_bottom"]),
            selected["name"], selected["id_card"], selected["amount"], not group["saved"],
        )
    detail_sheet_name = _find_sheet_name(workbook.sheetnames, DETAIL_SHEET_KEYWORD)
    ws = workbook[detail_sheet_name]
    header_row = _find_header_row_any(ws, HEADER_ID_CARD_ALIASES, sheet_label="明细表")
    headers = _read_headers(ws, header_row)
    try:
        return SalarySourceLayout(
            detail_sheet_name=detail_sheet_name,
            header_row=header_row,
            data_start_row=_find_data_start_row(ws, header_row),
            name_col=headers[HEADER_NAME],
            id_card_col=_first_header_col(headers, HEADER_ID_CARD_ALIASES, sheet_label="明细表"),
            amount_col=_first_header_col(headers, HEADER_AMOUNT_ALIASES, sheet_label="明细表"),
        )
    except KeyError as exc:
        raise ValueError(f"{workbook.path if hasattr(workbook, 'path') else ''}明细表缺少必要字段：{exc.args[0]}") from exc


def _find_sheet_name(sheetnames: list[str], keyword: str) -> str:
    for sheetname in sheetnames:
        if keyword in sheetname:
            return sheetname
    raise ValueError(f"未找到包含“{keyword}”的工作表")


def _find_header_row(ws: Worksheet, required_header: str) -> int:
    max_column = ws.max_column
    for row_index in range(1, min(ws.max_row, 20) + 1):
        values = [
            str(ws.cell(row_index, col).value or "").strip()
            for col in range(1, max_column + 1)
        ]
        if required_header in values:
            return row_index
    raise ValueError(f"未在明细表前 20 行找到字段：{required_header}")


def _read_headers(ws: Worksheet, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_index in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_index).value
        if value is None:
            continue
        text = str(value).strip()
        if text and text not in headers:
            headers[text] = col_index
    return headers


def _find_data_start_row(ws: Worksheet, header_row: int) -> int:
    bottom = header_row
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= header_row <= merged_range.max_row:
            bottom = max(bottom, merged_range.max_row)
    return bottom + 1


def _detect_month(file_path: Path) -> str:
    month = _month_from_text(file_path.stem)
    if month:
        return month

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        for ws in workbook.worksheets:
            max_row = ws.max_row or 5
            max_col = ws.max_column or 5
            for row in ws.iter_rows(min_row=1, max_row=min(max_row, 5), max_col=min(max_col, 5)):
                for cell in row:
                    month = _month_from_value(cell.value)
                    if month:
                        return month
    finally:
        workbook.close()
    raise ValueError(f"无法识别工资表月份：{file_path.name}。请在文件名中包含 202604 或 2026年4月")


def _month_from_text(text: str) -> str | None:
    compact = _PERIOD_COMPACT.search(text)
    if compact:
        year = int(compact.group(1))
        month = int(compact.group(2))
        if 1 <= month <= 12:
            return f"{year}{month:02d}"

    plain = _PERIOD_PLAIN.search(text)
    if plain:
        year = int(plain.group(1))
        month = int(plain.group(2))
        if 1 <= month <= 12:
            return f"{year}{month:02d}"
    return None


def _month_from_value(value: Any) -> str | None:
    if isinstance(value, datetime):
        return f"{value.year}{value.month:02d}"
    if isinstance(value, date):
        return f"{value.year}{value.month:02d}"
    if isinstance(value, str):
        return _month_from_text(value)
    if isinstance(value, (int, float)) and 40000 <= value <= 60000:
        dt = from_excel(value)
        return f"{dt.year}{dt.month:02d}"
    if isinstance(value, (int, float)):
        integer = int(value)
        year = integer // 100
        month = integer % 100
        if 2000 <= year <= 2099 and 1 <= month <= 12:
            return f"{year}{month:02d}"
    return None


def _amount_value(
    value_ws: Worksheet,
    formula_ws: Worksheet,
    row_index: int,
    layout: SalarySourceLayout,
) -> float | None:
    cached = value_ws.cell(row_index, layout.amount_col).value
    parsed = _number(cached)
    if parsed is not None:
        return parsed

    formula = formula_ws.cell(row_index, layout.amount_col).value
    if isinstance(formula, str) and formula.startswith("="):
        amount_header = _cell_text(formula_ws, layout.header_row, layout.amount_col)
        if amount_header != HEADER_AMOUNT or not layout.legacy_amount_fallback:
            return _monthly_amount_from_formula(value_ws, formula_ws, row_index, layout, formula)
        return _fallback_amount_from_row(value_ws, row_index)
    return None


def _monthly_amount_from_formula(
    value_ws: Worksheet,
    formula_ws: Worksheet,
    row_index: int,
    layout: SalarySourceLayout,
    formula: str,
) -> float:
    """Read the new template's actual ranges when splitting removed Excel caches.

    Deliberately limited to ROUND(SUM(row range)-SUM(row range),2), with
    basic arithmetic in its input cells. Never guess columns or execute Excel
    expressions; other uncached formulas require recalculation in Excel/WPS.
    The legacy template keeps its existing fallback unchanged.
    """
    values: dict[int, Decimal] = {}
    visiting: set[int] = set()

    def column(ref: str) -> int:
        match = _SALARY_CELL_REF.fullmatch(ref)
        if match is None or int(match[2]) != row_index:
            raise ValueError("not a same-row reference")
        col = column_index_from_string(match[1])
        if not 1 <= col < layout.amount_col:
            raise ValueError("reference outside amount inputs")
        return col

    def arithmetic(node: ast.AST) -> Decimal:
        if isinstance(node, ast.Constant) and type(node.value) in (int, float):
            return Decimal(str(node.value))
        if isinstance(node, ast.Name):
            return cell_value(column(node.id))
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            number = arithmetic(node.operand)
            return number if isinstance(node.op, ast.UAdd) else -number
        if isinstance(node, ast.BinOp):
            left, right = arithmetic(node.left), arithmetic(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if isinstance(node.op, ast.Div):
                return left / right
        raise ValueError("unsupported arithmetic")

    def cell_value(col: int) -> Decimal:
        if col in values:
            return values[col]
        if col in visiting or len(visiting) >= 32:
            raise ValueError("cyclic or excessive formula dependencies")
        visiting.add(col)
        cached = value_ws.cell(row_index, col).value
        parsed = _number(cached)
        source = formula_ws.cell(row_index, col).value
        if parsed is not None:
            result = Decimal(str(parsed))
        elif isinstance(source, str) and source.startswith("="):
            if len(source) > 512:
                raise ValueError("excessive formula length")
            tree = ast.parse(source[1:].replace("$", "").strip().upper(), mode="eval")
            if sum(1 for _ in ast.walk(tree)) > 64:
                raise ValueError("excessive formula complexity")
            result = arithmetic(tree.body)
        elif source is None or source == "":
            result = Decimal(0)
        else:
            raise ValueError("non-numeric formula input")
        if not result.is_finite():
            raise ValueError("non-finite formula input")
        values[col] = result
        visiting.remove(col)
        return result

    def range_sum(start: str, end: str) -> Decimal:
        first, last = column(start), column(end)
        if first > last:
            raise ValueError("reversed amount range")
        return sum((cell_value(col) for col in range(first, last + 1)), Decimal(0))

    try:
        if len(formula) > 512:
            raise ValueError("excessive amount formula length")
        match = _MONTHLY_AMOUNT_FORMULA.fullmatch(re.sub(r"\s+", "", formula).replace("$", "").upper())
        if match is None:
            raise ValueError("unsupported amount formula")
        amount = range_sum(match[1], match[2]) - range_sum(match[3], match[4])
        return float(amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except (ValueError, SyntaxError, ArithmeticError, RecursionError) as exc:
        coordinate = f"{get_column_letter(layout.amount_col)}{row_index}"
        raise ValueError(
            f"{layout.detail_sheet_name}!{coordinate} 应发工资公式没有可用缓存且无法安全计算，"
            "请用 Excel/WPS 重新计算并保存后再合并"
        ) from exc


def _fallback_amount_from_row(ws: Worksheet, row_index: int) -> float:
    gross = sum(_number(ws.cell(row_index, col).value) or 0 for col in range(5, 13))
    deductions = sum(_number(ws.cell(row_index, col).value) or 0 for col in range(13, 16))
    return round(gross - deductions, 2)


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if cleaned in {"", "-"}:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _cell_text(ws: Worksheet, row_index: int, col_index: int) -> str:
    value = ws.cell(row_index, col_index).value
    if value is None:
        return ""
    return str(value).strip()


def _build_output_months(
    records: list[SalaryRecord],
    year: int | None,
    existing_months: list[str] | None = None,
) -> list[str]:
    if existing_months:
        if year is not None:
            months = [f"{year}{month:02d}" for month in range(1, 13)]
            for month in [*existing_months, *sorted({record.month for record in records})]:
                if month not in months:
                    months.append(month)
            return months

        months = list(existing_months)
        for month in sorted({record.month for record in records}):
            if month not in months:
                months.append(month)
        return months

    if year is None:
        years = sorted({int(record.month[:4]) for record in records})
        if len(years) == 1:
            year = years[0]
    if year is not None:
        return [f"{year}{month:02d}" for month in range(1, 13)]
    return sorted({record.month for record in records})


def _read_existing_summary(
    summary_path: Path,
    *, header_profiles: dict[str, Any] | None = None,
    layout_hint: dict[str, Any] | None = None,
) -> tuple[list[str], list[MergedEmployee], list[str]]:
    warnings: list[str] = []
    workbook = load_workbook(summary_path, data_only=True)
    try:
        if header_profiles is None:
            sheet_name = _find_sheet_name_or_active(workbook.sheetnames, SUMMARY_SHEET_KEYWORD)
            ws = workbook[sheet_name]
            header_row = _find_header_row_any(ws, HEADER_ID_CARD_ALIASES, sheet_label="汇总表")
            headers = _read_headers(ws, header_row)
            name_col = _first_header_col(headers, (HEADER_NAME,), sheet_label="汇总表")
            id_card_col = _first_header_col(headers, HEADER_ID_CARD_ALIASES, sheet_label="汇总表")
        else:
            group = inspect_workbook(workbook, role="summary", profiles=header_profiles, hint=layout_hint)
            if not group["ready"]:
                raise ValueError(group["problem"])
            ws = workbook[group["sheet"]]
            header_row = group["header_bottom"]
            name_col, id_card_col = group["selections"]["name"], group["selections"]["id_card"]
        month_columns = _read_month_columns(ws, header_row)
        if not month_columns:
            raise ValueError("已有汇总表未找到月份列，请确认表头包含 202601 这类月份")

        data_start_row = _find_data_start_row(ws, header_row)
        employees: OrderedDict[str, MergedEmployee] = OrderedDict()
        for row_index in range(data_start_row, ws.max_row + 1):
            id_card = _cell_text(ws, row_index, id_card_col)
            name = _cell_text(ws, row_index, name_col)
            if not id_card and not name:
                continue
            if not id_card:
                warnings.append(f"{summary_path.name} 第 {row_index} 行缺少身份证号码，已跳过")
                continue
            amounts = OrderedDict(
                (month, _number(ws.cell(row_index, col_index).value) or 0.0)
                for month, col_index in month_columns.items()
            )
            if id_card in employees:
                warnings.append(f"已有汇总表中身份证 {id_card} 出现重复行，仅保留第一行")
                continue
            employees[id_card] = MergedEmployee(name=name, id_card=id_card, amounts=amounts)
        return list(month_columns.keys()), list(employees.values()), warnings
    finally:
        workbook.close()


def _find_sheet_name_or_active(sheetnames: list[str], keyword: str) -> str:
    try:
        return _find_sheet_name(sheetnames, keyword)
    except ValueError:
        return sheetnames[0]


def _find_header_row_any(ws: Worksheet, required_headers: tuple[str, ...], *, sheet_label: str) -> int:
    max_column = ws.max_column
    for row_index in range(1, min(ws.max_row, 20) + 1):
        values = [
            str(ws.cell(row_index, col).value or "").strip()
            for col in range(1, max_column + 1)
        ]
        if any(required_header in values for required_header in required_headers):
            return row_index
    joined = " / ".join(required_headers)
    raise ValueError(f"未在{sheet_label}前 20 行找到字段：{joined}")


def _first_header_col(headers: dict[str, int], names: tuple[str, ...], *, sheet_label: str) -> int:
    for name in names:
        if name in headers:
            return headers[name]
    joined = " / ".join(names)
    raise ValueError(f"{sheet_label}缺少必要字段：{joined}")


def _read_month_columns(ws: Worksheet, header_row: int) -> OrderedDict[str, int]:
    month_columns: OrderedDict[str, int] = OrderedDict()
    for col_index in range(1, ws.max_column + 1):
        month = _month_from_value(ws.cell(header_row, col_index).value)
        if month and month not in month_columns:
            month_columns[month] = col_index
    return month_columns


def _merge_records(
    records: list[SalaryRecord],
    months: list[str],
    existing_employees: list[MergedEmployee] | None = None,
) -> tuple[list[MergedEmployee], list[str], int, int]:
    warnings: list[str] = []
    employees: OrderedDict[str, MergedEmployee] = OrderedDict()
    protected_months: set[tuple[str, str]] = set()
    if existing_employees:
        for existing_employee in existing_employees:
            amounts = OrderedDict((month, 0.0) for month in months)
            for month, amount in existing_employee.amounts.items():
                if month in amounts:
                    amounts[month] = amount
                    if _has_existing_amount(amount):
                        protected_months.add((existing_employee.id_card, month))
            employees[existing_employee.id_card] = MergedEmployee(
                name=existing_employee.name,
                id_card=existing_employee.id_card,
                amounts=amounts,
            )

    seen_month_rows: set[tuple[str, str]] = set()
    applied_count = 0
    skipped_count = 0
    for record in records:
        if record.month not in months:
            warnings.append(f"{record.source_file} 第 {record.source_row} 行月份 {record.month} 不在输出月份中，已跳过")
            continue
        if record.id_card not in employees:
            employees[record.id_card] = MergedEmployee(
                name=record.name,
                id_card=record.id_card,
                amounts=OrderedDict((month, 0.0) for month in months),
            )
        employee = employees[record.id_card]
        if employee.name and record.name and employee.name != record.name:
            warnings.append(f"身份证 {record.id_card} 出现多个姓名：{employee.name} / {record.name}")
        key = (record.id_card, record.month)
        if key in protected_months:
            skipped_count += 1
            warnings.append(
                f"身份证 {record.id_card} {record.month} 已存在金额，未覆盖"
                f"（来源：{record.source_file} 第 {record.source_row} 行）"
            )
            continue
        if key in seen_month_rows:
            warnings.append(f"身份证 {record.id_card} 在 {record.month} 出现重复记录，金额已累加")
        seen_month_rows.add(key)
        employee.amounts[record.month] = round(employee.amounts[record.month] + record.amount, 2)
        applied_count += 1
    return list(employees.values()), warnings, applied_count, skipped_count


def _has_existing_amount(value: Any) -> bool:
    parsed = _number(value)
    return parsed is not None and parsed != 0


def _write_output_workbook(
    output_file: Path,
    employees: list[MergedEmployee],
    months: list[str],
    *,
    cancelled: Callable[[], bool] | None = None,
) -> None:
    workbook = Workbook()
    try:
        ws = workbook.active
        ws.title = "汇总"

        max_col = 3 + len(months)
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=max_col)
        ws["A1"] = SUMMARY_TITLE
        ws["A1"].font = Font(name="宋体", size=20, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        headers: list[Any] = ["序号", "姓名", "身份证号码", *(int(month) for month in months)]
        subheaders = ["", "", "", *(["应发工资"] * len(months))]
        for col_index, value in enumerate(headers, start=1):
            ws.cell(3, col_index).value = value
        for col_index, value in enumerate(subheaders, start=1):
            ws.cell(4, col_index).value = value
        ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
        ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
        ws.merge_cells(start_row=3, start_column=3, end_row=4, end_column=3)
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 22
        ws.freeze_panes = "D5"

        for row_index, employee in enumerate(employees, start=5):
            if row_index % 500 == 0:
                _check_cancelled(cancelled)
            values: list[Any] = [row_index - 4, employee.name, employee.id_card]
            values.extend(employee.amounts[month] for month in months)
            ws.append(values)

        _format_output_sheet(
            ws,
            max_col,
            len(employees),
            cancelled=cancelled,
        )
        _check_cancelled(cancelled)
        workbook.save(output_file)
    finally:
        workbook.close()


_OUTPUT_SIDE = Side(style="thin", color="000000")
_OUTPUT_BORDER = Border(left=_OUTPUT_SIDE, right=_OUTPUT_SIDE, top=_OUTPUT_SIDE, bottom=_OUTPUT_SIDE)
_OUTPUT_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_OUTPUT_SONG_BOLD = Font(name="宋体", size=10, bold=True)
_OUTPUT_SONG_PLAIN = Font(name="宋体", size=10, bold=False)
_OUTPUT_TIMES_PLAIN = Font(name="Times New Roman", size=10, bold=False)
_OUTPUT_TIMES_BODY = Font(name="Times New Roman", size=10)


def _format_output_sheet(
    ws: Worksheet,
    max_col: int,
    employee_count: int,
    *,
    cancelled: Callable[[], bool] | None = None,
) -> None:
    # 员工行数可能上千，逐格新建 Font/Alignment 再交给 openpyxl 哈希查表开销很大；
    # 这里先把用到的几种样式解析成下标，再直接写入单元格样式下标。
    border_id = cached_style_id(ws, "border", "output_thin", lambda: _OUTPUT_BORDER)
    alignment_id = cached_style_id(ws, "alignment", "output_center", lambda: _OUTPUT_ALIGNMENT)
    song_bold_id = cached_style_id(ws, "font", "output_song_bold", lambda: _OUTPUT_SONG_BOLD)
    song_plain_id = cached_style_id(ws, "font", "output_song_plain", lambda: _OUTPUT_SONG_PLAIN)
    times_plain_id = cached_style_id(ws, "font", "output_times_plain", lambda: _OUTPUT_TIMES_PLAIN)
    times_body_id = cached_style_id(ws, "font", "output_times_body", lambda: _OUTPUT_TIMES_BODY)
    for row in ws.iter_rows(min_row=3, max_row=4, max_col=max_col):
        for cell in row:
            if cell.row == 3 and cell.column >= 4:
                font_id = times_plain_id
            elif cell.row == 4 or cell.column <= 3:
                font_id = song_bold_id
            else:
                font_id = song_plain_id
            set_style_ids(cell, border_id=border_id, alignment_id=alignment_id, font_id=font_id)

    for row in ws.iter_rows(min_row=5, max_row=4 + employee_count, max_col=max_col):
        if row and row[0].row % 500 == 0:
            _check_cancelled(cancelled)
        for cell in row:
            set_style_ids(cell, border_id=border_id, alignment_id=alignment_id, font_id=times_body_id)
            if cell.column >= 4:
                cell.number_format = AMOUNT_NUMBER_FORMAT

    widths = {1: 5.5, 2: 8, 3: 20}
    for col_index in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col_index)].width = widths.get(col_index, 13)
