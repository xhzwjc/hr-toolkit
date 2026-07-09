from __future__ import annotations

import re


# 预编译正则
_POLICY_NO_PATTERN = re.compile(r"(?:保单号码|保险单号)\s*[:：]?\s*([A-Z0-9]{10,})")
_FILENAME_POLICY_PATTERN = re.compile(r"([A-Z]{3,}[A-Z0-9]{8,})")
_HEADER_WHITESPACE = re.compile(r"\s+")
_ID_CARD_STRIP_PATTERN = re.compile(r"[^0-9.\-]")
import shutil
import tempfile
from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from hr_toolkit.common.resources import open_template_resource
from hr_toolkit.common.excel import SheetGrid, apply_row_snapshot, snapshot_row
from hr_toolkit.common.excel_compat import ensure_xlsx_workbook, is_supported_excel_file
from hr_toolkit.common.inputs import extract_zip_excel_files, normalize_input_paths


TOOL_NAME = "需求3-保险台账"
TEMPLATE_RESOURCE = "insurance_ledger_template.xlsx"
OUTPUT_FILENAME = "保险台账.xlsx"
ROSTER_WARNING_OUTPUT_FILENAME = "人力资源分析表_保险预警.xlsx"
PEAC_DEFAULT_AMOUNT = 60.0
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
WARNING_FONT = Font(name="宋体", size=10, color="C00000", bold=True)
THIN_BLACK_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


@dataclass(frozen=True)
class RosterPerson:
    name: str
    id_card: str
    department: str
    status: str
    active: bool
    source_file: str
    sheet_name: str
    source_row: int


@dataclass(frozen=True)
class PolicyEntry:
    policy_no: str
    name: str
    id_card: str
    amount: float
    source_file: str
    source_row: int


@dataclass
class LedgerPerson:
    name: str
    id_card: str
    department: str = ""
    policies: OrderedDict[str, float] = field(default_factory=OrderedDict)
    warning: str = ""

    @property
    def total_amount(self) -> float:
        return round(sum(self.policies.values()), 2)


@dataclass(frozen=True)
class InsuranceWarning:
    warning_type: str
    name: str
    id_card: str
    department: str
    message: str


@dataclass
class InsuranceLedgerResult:
    input_path: Path
    roster_path: Path
    output_dir: Path
    dry_run: bool = False
    source_files: list[str] = field(default_factory=list)
    policy_count: int = 0
    insured_person_count: int = 0
    roster_person_count: int = 0
    add_warning_count: int = 0
    reduce_warning_count: int = 0
    warnings: list[str] = field(default_factory=list)
    output_file: Path | None = None
    roster_warning_file: Path | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "tool_name": TOOL_NAME,
            "input_path": str(self.input_path),
            "roster_path": str(self.roster_path),
            "output_dir": str(self.output_dir),
            "dry_run": self.dry_run,
            "source_files": self.source_files,
            "source_file_count": len(self.source_files),
            "policy_count": self.policy_count,
            "insured_person_count": self.insured_person_count,
            "roster_person_count": self.roster_person_count,
            "add_warning_count": self.add_warning_count,
            "reduce_warning_count": self.reduce_warning_count,
            "warnings": self.warnings,
            "output_file": None if self.output_file is None else str(self.output_file),
            "roster_warning_file": None if self.roster_warning_file is None else str(self.roster_warning_file),
        }


def generate_insurance_ledger(
    input_path: str | Path | list[str | Path],
    roster_path: str | Path,
    output_dir: str | Path,
    *,
    dry_run: bool = False,
) -> InsuranceLedgerResult:
    input_paths = _normalize_input_paths(input_path)
    display_input = input_paths[0] if len(input_paths) == 1 else input_paths[0].parent
    roster = Path(roster_path).expanduser().resolve()
    output = Path(output_dir).expanduser().resolve()
    warnings: list[str] = []

    for path in input_paths:
        if not path.exists():
            raise FileNotFoundError(f"保单人员清单、压缩包或文件夹不存在：{path}")
    if not roster.exists() or not roster.is_file():
        raise FileNotFoundError(f"人力资源分析表不存在：{roster}")
    if not is_supported_excel_file(roster):
        raise ValueError("人力资源分析表仅支持 .xlsx 或 .xls 文件。")

    with tempfile.TemporaryDirectory(prefix="hr_insurance_ledger_") as temp_root:
        temp_dir = Path(temp_root)
        working_roster = ensure_xlsx_workbook(roster, temp_dir)
        roster_people = _read_roster(working_roster, roster.name, warnings)
        policy_files = _find_policy_files(input_paths, temp_dir, warnings)
        if not policy_files:
            raise ValueError("未找到 .xlsx 或 .xls 保单人员清单。")

        entries: list[PolicyEntry] = []
        used_files: list[str] = []
        for file_path in policy_files:
            file_entries = _read_policy_file(file_path, warnings)
            if file_entries:
                used_files.append(str(file_path))
                entries.extend(file_entries)
            else:
                warnings.append(f"{file_path.name} 未识别到保单人员，已跳过。")

        if not entries:
            raise ValueError("未识别到保单人员，请确认保单清单格式。")

        ledger_people, policy_order, personnel_warnings = _build_ledger(entries, roster_people, warnings)
        add_count = sum(1 for item in personnel_warnings if item.warning_type == "需加保")
        reduce_count = sum(1 for item in personnel_warnings if item.warning_type == "需减保")

        result = InsuranceLedgerResult(
            input_path=display_input,
            roster_path=roster,
            output_dir=output,
            dry_run=dry_run,
            source_files=used_files,
            policy_count=len(policy_order),
            insured_person_count=len(ledger_people),
            roster_person_count=sum(1 for item in roster_people.values() if item.active),
            add_warning_count=add_count,
            reduce_warning_count=reduce_count,
            warnings=warnings,
        )
        if dry_run:
            return result

        output.mkdir(parents=True, exist_ok=True)
        output_file = output / OUTPUT_FILENAME
        _write_output_workbook(output_file, ledger_people, policy_order, personnel_warnings, temp_dir)
        result.output_file = output_file
        if add_count:
            roster_warning_file = output / ROSTER_WARNING_OUTPUT_FILENAME
            _write_roster_warning_workbook(working_roster, roster_warning_file, roster_people, personnel_warnings)
            result.roster_warning_file = roster_warning_file
        return result


def _normalize_input_paths(input_path: str | Path | list[str | Path]) -> list[Path]:
    return normalize_input_paths(input_path, "请选择保单人员清单、压缩包或文件夹。")


def _find_policy_files(input_paths: list[Path], temp_dir: Path, warnings: list[str]) -> list[Path]:
    files: list[Path] = []
    seen: set[Path] = set()
    for input_path in input_paths:
        for path in _iter_input_files(input_path, temp_dir, warnings):
            if _is_non_source_file(path):
                continue
            working_path = ensure_xlsx_workbook(path, temp_dir)
            resolved = working_path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            files.append(working_path)
    return sorted(files)


def _iter_input_files(input_path: Path, temp_dir: Path, warnings: list[str]) -> list[Path]:
    if input_path.is_file():
        if is_supported_excel_file(input_path):
            return [input_path]
        if input_path.suffix.lower() == ".zip":
            return extract_zip_excel_files(input_path, temp_dir, warnings)
        return []
    if not input_path.is_dir():
        raise FileNotFoundError(f"路径不存在：{input_path}")
    files: list[Path] = []
    for path in sorted(input_path.rglob("*")):
        if not path.is_file() or path.name.startswith(("~$", ".~")):
            continue
        if is_supported_excel_file(path):
            files.append(path)
        elif path.suffix.lower() == ".zip":
            files.extend(extract_zip_excel_files(path, temp_dir, warnings))
    return files


def _is_non_source_file(path: Path) -> bool:
    return any(keyword in path.name for keyword in ("模板", "台账", "汇总"))


def _read_roster(workbook_path: Path, source_name: str, warnings: list[str]) -> OrderedDict[str, RosterPerson]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    people: OrderedDict[str, RosterPerson] = OrderedDict()
    try:
        worksheets = [workbook["花名册"]] if "花名册" in workbook.sheetnames else workbook.worksheets
        for ws in worksheets:
            header_row = _find_roster_header_row(ws)
            if header_row is None:
                continue
            footer_start = _find_roster_footer_start(ws, header_row + 1)
            headers = _read_headers_first(ws, header_row)
            for row_index in range(header_row + 1, footer_start):
                name = _cell_text(_header_value_any(ws, row_index, headers, ("*姓名.简体中文", "姓名", "员工姓名", "人员姓名")))
                id_card = _normalize_id_card(_header_value_any(ws, row_index, headers, ("*身份证", "身份证", "身份证号码", "证件号")))
                if not name or not id_card or name in {"姓名", "员工姓名", "人员姓名"} or id_card in {"身份证", "身份证号码", "证件号"}:
                    continue
                department = _cell_text(
                    _header_value_any(
                        ws,
                        row_index,
                        headers,
                        (
                            "部门/项目",
                            "项目/部门",
                            "部门（片区）",
                            "部门片区",
                            "成本中心.名称",
                            "*责任部门.名称",
                            "责任部门.名称",
                            "部门名称",
                            "部门",
                            "项目.项目名称",
                        ),
                    )
                )
                status = _cell_text(
                    _header_value_any(
                        ws,
                        row_index,
                        headers,
                        ("状态", "员工状态", "在职状态", "*参保状态", "参保状态"),
                    )
                )
                person = RosterPerson(
                    name=name,
                    id_card=id_card,
                    department=department or "未填写",
                    status=status,
                    active=_is_active_status(status) and not _is_leave_marked_row(ws, row_index, headers),
                    source_file=source_name,
                    sheet_name=ws.title,
                    source_row=row_index,
                )
                if id_card in people and people[id_card].name != name:
                    warnings.append(f"花名册身份证重复但姓名不同：{id_card}，已保留首次记录 {people[id_card].name}。")
                    continue
                people[id_card] = person
    finally:
        workbook.close()
    if not people:
        raise ValueError("人力资源分析表的“花名册”未识别到姓名和身份证号码，请确认表头。")
    return people


def _find_roster_footer_start(ws: Worksheet, start_row: int) -> int:
    for row_index in range(start_row, (ws.max_row or 0) + 1):
        row_text = " ".join(_cell_text(ws.cell(row_index, col_index).value) for col_index in range(1, 8))
        if any(keyword in row_text for keyword in ("对应异动", "根据异动汇总表", "对应项目部")):
            return row_index
    return (ws.max_row or start_row) + 1


def _find_roster_header_row(ws: Worksheet) -> int | None:
    max_col = min(ws.max_column or 0, 120)
    id_headers = {"*身份证", "身份证", "身份证号码", "证件号"}
    name_headers = {"*姓名.简体中文", "姓名", "员工姓名", "人员姓名"}
    for row_index in range(1, min(ws.max_row or 0, 30) + 1):
        values = {_normalize_header(ws.cell(row_index, col_index).value) for col_index in range(1, max_col + 1)}
        if values & id_headers and values & name_headers:
            return row_index
    return None


def _is_leave_marked_row(ws: Worksheet, row_index: int, headers: dict[str, int]) -> bool:
    max_col = max(headers.values()) if headers else min(ws.max_column or 0, 40)
    max_col = min(max_col, 80)
    marked_cells = 0
    filled_cells = 0
    for col_index in range(1, max_col + 1):
        cell = ws.cell(row_index, col_index)
        if cell.value is not None:
            filled_cells += 1
        color = _cell_fill_rgb(cell)
        if color in {"FFF2CC", "00FFF2CC", "FFC7CE", "00FFC7CE", "FF0000", "00FF0000"}:
            marked_cells += 1
    return marked_cells >= 2 or (filled_cells > 0 and marked_cells == filled_cells)


def _cell_fill_rgb(cell) -> str:
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return ""
    color = fill.fgColor
    if color is None or color.type != "rgb":
        return ""
    return (color.rgb or "").upper()


def _read_policy_file(file_path: Path, warnings: list[str]) -> list[PolicyEntry]:
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    entries: list[PolicyEntry] = []
    seen: set[tuple[str, str]] = set()
    try:
        for worksheet in workbook.worksheets:
            # read_only 工作表随机访问是 O(行数²)，先单遍读入内存再处理
            ws = SheetGrid(worksheet)
            header_row = _find_policy_header_row(ws)
            if header_row is None:
                continue
            policy_no = _find_policy_no(ws, file_path)
            headers = _read_headers_first(ws, header_row)
            name_col = _first_header_col(headers, ("雇员姓名", "姓名", "被保险人姓名"))
            id_col = _first_header_col(headers, ("身份证号码", "证件号", "证件号码", "身份证号"))
            amount_col = _first_header_col(headers, ("每人伤残死亡限额", "伤残死亡限额", "死亡伤残限额"))
            if name_col is None or id_col is None:
                continue
            for row_index in range(header_row + 1, (ws.max_row or 0) + 1):
                name = _cell_text(ws.cell(row_index, name_col).value)
                id_card = _normalize_id_card(ws.cell(row_index, id_col).value)
                if not name or not id_card:
                    continue
                amount = _policy_amount(ws, row_index, amount_col, policy_no, warnings, file_path.name)
                if amount <= 0:
                    warnings.append(f"{file_path.name} 第 {row_index} 行未识别保额，已跳过。")
                    continue
                key = (id_card, policy_no)
                if key in seen:
                    continue
                seen.add(key)
                entries.append(
                    PolicyEntry(
                        policy_no=policy_no,
                        name=name,
                        id_card=id_card,
                        amount=amount,
                        source_file=file_path.name,
                        source_row=row_index,
                    )
                )
    finally:
        workbook.close()
    return entries


def _find_policy_header_row(ws: SheetGrid) -> int | None:
    max_col = min(ws.max_column or 0, 80)
    id_headers = {"身份证号码", "证件号", "证件号码", "身份证号"}
    name_headers = {"雇员姓名", "姓名", "被保险人姓名"}
    for row_index in range(1, min(ws.max_row or 0, 30) + 1):
        values = {_normalize_header(ws.cell(row_index, col_index).value) for col_index in range(1, max_col + 1)}
        if values & id_headers and values & name_headers:
            return row_index
    return None


def _find_policy_no(ws: SheetGrid, file_path: Path) -> str:
    for row_index in range(1, min(ws.max_row or 0, 15) + 1):
        for col_index in range(1, min(ws.max_column or 0, 30) + 1):
            text = _cell_text(ws.cell(row_index, col_index).value)
            match = _POLICY_NO_PATTERN.search(text.replace("\xa0", " "))
            if match:
                return match.group(1)
    match = _FILENAME_POLICY_PATTERN.search(file_path.stem.upper())
    return match.group(1) if match else file_path.stem


def _policy_amount(
    ws: SheetGrid,
    row_index: int,
    amount_col: int | None,
    policy_no: str,
    warnings: list[str],
    file_name: str,
) -> float:
    if amount_col is not None:
        amount = _to_number(ws.cell(row_index, amount_col).value)
        if amount:
            return _format_amount_wan(amount / 10000)
    if policy_no.upper().startswith("PEAC"):
        return PEAC_DEFAULT_AMOUNT
    warnings.append(f"{file_name} 的保单 {policy_no} 未找到“每人伤残死亡限额”，该行保额无法计算。")
    return 0.0


def _build_ledger(
    entries: list[PolicyEntry],
    roster_people: OrderedDict[str, RosterPerson],
    warnings: list[str],
) -> tuple[list[LedgerPerson], list[str], list[InsuranceWarning]]:
    ledger_by_id: OrderedDict[str, LedgerPerson] = OrderedDict()
    policy_order: list[str] = []
    policy_seen: set[str] = set()
    for entry in entries:
        if entry.policy_no not in policy_seen:
            policy_seen.add(entry.policy_no)
            policy_order.append(entry.policy_no)
        roster_person = roster_people.get(entry.id_card)
        person = ledger_by_id.setdefault(
            entry.id_card,
            LedgerPerson(
                name=roster_person.name if roster_person else entry.name,
                id_card=entry.id_card,
                department=roster_person.department if roster_person else "",
            ),
        )
        if roster_person and person.name != roster_person.name:
            person.name = roster_person.name
        elif not roster_person and person.name != entry.name:
            warnings.append(f"保单身份证 {entry.id_card} 出现不同姓名：{person.name}、{entry.name}，已保留首次姓名。")
        person.policies[entry.policy_no] = entry.amount

    personnel_warnings: list[InsuranceWarning] = []
    insured_ids = set(ledger_by_id)
    for person in roster_people.values():
        if person.active and person.id_card not in insured_ids:
            personnel_warnings.append(
                InsuranceWarning(
                    warning_type="需加保",
                    name=person.name,
                    id_card=person.id_card,
                    department=person.department,
                    message="花名册有该在职人员，但保单清单中未找到。",
                )
            )
    for id_card, person in ledger_by_id.items():
        roster_person = roster_people.get(id_card)
        if roster_person is None:
            person.warning = "需减保"
            personnel_warnings.append(
                InsuranceWarning(
                    warning_type="需减保",
                    name=person.name,
                    id_card=id_card,
                    department=person.department,
                    message="保单清单有该人员，但人力资源分析表花名册中未找到。",
                )
            )
        elif not roster_person.active:
            person.warning = "需减保"
            personnel_warnings.append(
                InsuranceWarning(
                    warning_type="需减保",
                    name=person.name,
                    id_card=id_card,
                    department=person.department,
                    message=f"保单清单有该人员，但花名册状态为“{roster_person.status or '非在职'}”。",
                )
            )

    return list(ledger_by_id.values()), _sort_policy_order(policy_order), personnel_warnings


def _sort_policy_order(policy_order: list[str]) -> list[str]:
    return sorted(policy_order, key=lambda item: (item.upper().startswith("PEAC"), item))


def _write_output_workbook(
    output_file: Path,
    ledger_people: list[LedgerPerson],
    policy_order: list[str],
    personnel_warnings: list[InsuranceWarning],
    temp_dir: Path,
) -> None:
    template_path = _copy_template(temp_dir)
    workbook = load_workbook(template_path)
    try:
        ws = workbook["保险台账"] if "保险台账" in workbook.sheetnames else workbook.active
        _write_ledger_sheet(ws, ledger_people, policy_order)
        _write_warning_sheet(workbook, personnel_warnings)
        workbook.save(output_file)
    finally:
        workbook.close()


def _copy_template(temp_dir: Path) -> Path:
    target = temp_dir / TEMPLATE_RESOURCE
    with open_template_resource(TEMPLATE_RESOURCE) as source, target.open("wb") as output:
        shutil.copyfileobj(source, output)
    return target


def _write_ledger_sheet(ws: Worksheet, ledger_people: list[LedgerPerson], policy_order: list[str]) -> None:
    headers = _build_headers(policy_order)
    max_col = len(headers)
    row_snapshot = snapshot_row(ws, 3, min(ws.max_column, 9))
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    # 清除模板多余列，避免残留模板数据
    if ws.max_column > max_col:
        ws.delete_cols(max_col + 1, ws.max_column - max_col)
    _write_ledger_headers(ws, policy_order)
    for offset, person in enumerate(ledger_people):
        row_index = 3 + offset
        apply_row_snapshot(ws, row_index, row_snapshot, translate_formulas=False)
        for col_index in range(1, max_col + 1):
            ws.cell(row_index, col_index).value = None
        ws.cell(row_index, 1).value = offset + 1
        ws.cell(row_index, 2).value = person.name
        ws.cell(row_index, 3).value = person.id_card
        ws.cell(row_index, 4).value = person.department
        for policy_index, policy_no in enumerate(policy_order, start=1):
            policy_col = 3 + policy_index * 2
            amount_col = policy_col + 1
            if policy_no in person.policies:
                ws.cell(row_index, policy_col).value = policy_no
                ws.cell(row_index, amount_col).value = person.policies[policy_no]
        ws.cell(row_index, max_col - 1).value = person.total_amount
        ws.cell(row_index, max_col).value = person.warning
    if not ledger_people:
        for col_index in range(1, max_col + 1):
            ws.cell(3, col_index).value = None
    _format_table(ws, 2, max(3, 2 + len(ledger_people)), max_col)
    for row_index in range(3, max(3, 2 + len(ledger_people)) + 1):
        ws.cell(row_index, 3).number_format = "@"
    ws.freeze_panes = "A3"


def _build_headers(policy_order: list[str]) -> list[str]:
    headers = ["序号", "姓名", "身份证号码", "项目/部门"]
    for index, _policy_no in enumerate(policy_order, start=1):
        headers.extend([f"保单号{index}", "保额"])
    headers.extend(["保额合计", "预警"])
    return headers


def _write_ledger_headers(ws: Worksheet, policy_order: list[str]) -> None:
    headers = _build_headers(policy_order)
    for col_index, header in enumerate(headers, start=1):
        ws.cell(2, col_index).value = header
    if ws.max_column > len(headers):
        ws.delete_cols(len(headers) + 1, ws.max_column - len(headers))


def _write_warning_sheet(workbook, personnel_warnings: list[InsuranceWarning]) -> None:
    if "人员增减预警" in workbook.sheetnames:
        del workbook["人员增减预警"]
    ws = workbook.create_sheet("人员增减预警")
    headers = ["序号", "预警类型", "姓名", "身份证号码", "项目/部门", "说明"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(1, col_index).value = header
    for index, item in enumerate(personnel_warnings, start=1):
        values = [index, item.warning_type, item.name, item.id_card, item.department, item.message]
        for col_index, value in enumerate(values, start=1):
            ws.cell(index + 1, col_index).value = value
        ws.cell(index + 1, 4).number_format = "@"
    _format_table(ws, 1, max(2, len(personnel_warnings) + 1), len(headers))
    ws.column_dimensions["F"].width = 42
    ws.freeze_panes = "A2"


def _write_roster_warning_workbook(
    source_workbook: Path,
    output_file: Path,
    roster_people: OrderedDict[str, RosterPerson],
    personnel_warnings: list[InsuranceWarning],
) -> None:
    add_warning_ids = {item.id_card for item in personnel_warnings if item.warning_type == "需加保"}
    if not add_warning_ids:
        return
    workbook = load_workbook(source_workbook)
    try:
        # 按 sheet 缓存 header_row 和 headers，避免重复扫描
        header_cache: dict[str, tuple[int, dict[str, int]]] = {}
        for person in roster_people.values():
            if person.id_card not in add_warning_ids:
                continue
            if person.sheet_name not in workbook.sheetnames:
                continue
            ws = workbook[person.sheet_name]
            if person.sheet_name not in header_cache:
                header_row = _find_roster_header_row(ws)
                if header_row is None:
                    header_cache[person.sheet_name] = (0, None, 0)
                    continue
                headers = _read_headers_first(ws, header_row)
                header_cache[person.sheet_name] = (header_row, headers, 0)
            header_row, headers, warning_col = header_cache[person.sheet_name]
            if header_row == 0:
                continue
            if warning_col == 0:
                warning_col = _last_header_column(ws, header_row) + 1
                ws.cell(header_row, warning_col).value = "保险预警"
                ws.cell(header_row, warning_col).fill = PatternFill("solid", fgColor="FCE4D6")
                ws.cell(header_row, warning_col).font = Font(name="宋体", size=10, bold=True)
                ws.cell(header_row, warning_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.column_dimensions[get_column_letter(warning_col)].width = 16
                # 缓存刚创建的新列号，headers 字典保持只读不被破坏
                header_cache[person.sheet_name] = (header_row, headers, warning_col)
            warning_col = header_cache[person.sheet_name][2]
            cell = ws.cell(person.source_row, warning_col)
            cell.value = "需加保"
            cell.fill = WARNING_FILL
            cell.font = WARNING_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BLACK_BORDER
        workbook.save(output_file)
    finally:
        workbook.close()


def _last_header_column(ws: Worksheet, header_row: int) -> int:
    last_col = 1
    for col_index in range(1, min(ws.max_column or 0, 220) + 1):
        if _cell_text(ws.cell(header_row, col_index).value):
            last_col = col_index
    return last_col


def _format_table(ws: Worksheet, min_row: int, max_row: int, max_col: int) -> None:
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    header_fill = PatternFill("solid", fgColor="FCE4D6")
    header_font = Font(name="宋体", size=10, bold=True)
    normal_font = Font(name="宋体", size=10)
    for row_index in range(min_row, max_row + 1):
        for col_index in range(1, max_col + 1):
            cell = ws.cell(row_index, col_index)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = header_font if row_index == min_row else normal_font
            if row_index == min_row:
                cell.fill = header_fill
            if isinstance(cell.value, float):
                cell.number_format = "0.##"
    for col_index in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col_index)].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18


def _read_headers_first(ws: Worksheet | SheetGrid, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    max_col = min(ws.max_column or 0, 160)
    for col_index in range(1, max_col + 1):
        header = _normalize_header(ws.cell(header_row, col_index).value)
        if header and header not in headers:
            headers[header] = col_index
    return headers


def _header_value_any(ws: Worksheet | SheetGrid, row_index: int, headers: dict[str, int], candidates: tuple[str, ...]) -> Any:
    col_index = _first_header_col(headers, candidates)
    if col_index is None:
        return None
    return ws.cell(row_index, col_index).value


def _first_header_col(headers: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    for header in candidates:
        col_index = headers.get(_normalize_header(header))
        if col_index is not None:
            return col_index
    return None


def _normalize_header(value: Any) -> str:
    return _HEADER_WHITESPACE.sub("", _cell_text(value).replace("\xa0", ""))


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().replace("\xa0", " ")


def _normalize_id_card(value: Any) -> str:
    text = _cell_text(value)
    text = _HEADER_WHITESPACE.sub("", text).replace("'", "").upper()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def _to_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = _cell_text(value)
    text = text.replace(",", "").replace("，", "")
    text = _ID_CARD_STRIP_PATTERN.sub("", text)
    try:
        return float(text)
    except ValueError:
        return 0.0


def _format_amount_wan(value: float) -> float:
    value = round(value, 2)
    return int(value) if float(value).is_integer() else value


def _is_active_status(status: str) -> bool:
    text = status.strip()
    if not text:
        return True
    inactive_keywords = ("离职", "停保", "不在职", "已离", "停用", "无效")
    return not any(keyword in text for keyword in inactive_keywords)
