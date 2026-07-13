from __future__ import annotations

import re


# 预编译正则
_SHEET_TITLE_INVALID = re.compile(r"[:\\/?*\[\]]")
_COMPANY_FILE_INVALID = re.compile(r'[<>:"/\\|?*\r\n]+')
_HEADER_WHITESPACE = re.compile(r"\s+")

import shutil
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from hr_toolkit.common.resources import open_template_resource
from hr_toolkit.common.excel_compat import is_supported_excel_file, ensure_xlsx_workbook
from hr_toolkit.common.excel import apply_row_snapshot, cell_text as _cell_text, insert_rows, snapshot_row
from hr_toolkit.common.inputs import extract_zip_excel_files, normalize_input_paths


TOOL_NAME = "需求7-档案入库"
EXPORT_TOOL_NAME = "需求7-档案表生成"
OUTPUT_FILENAME = "档案表汇总表.xlsx"
DEFAULT_ARCHIVE_SUMMARY_TEMPLATE_RESOURCE = "archive_summary_template.xlsx"
DEFAULT_ARCHIVE_COMPANY_TEMPLATE_RESOURCE = "archive_company_template.xlsx"
HEADER_COMPANY = "公司"
HEADER_NAME = "姓名"
HEADER_ID_CARD = "身份证"
HEADER_OTHER = "其他"
PLACEHOLDER_SHEET_TITLES = {"模板", "公司……", "公司..."}

REGION_CODES = {
    "总部": "00",
    "南昌": "01",
    "南昌分公司": "01",
    "抚州": "02",
    "鹰潭": "03",
    "达州": "04",
    "乐山": "05",
    "成都": "06",
    "广州": "07",
    "河源": "08",
    "云浮": "09",
    "阳江": "10",
    "茂名": "11",
    "普洱": "12",
    "德宏": "13",
    "上海": "14",
    "新疆": "15",
    "青岛": "16",
    "青海": "17",
    "研发部": "18",
    "南京": "19",
    "福建": "20",
    "河南": "21",
    "湖南": "22",
    "北京": "23",
    "江西工程": "24",
    "惠州": "25",
    "陕西": "26",
    "贵州": "27",
    "攀枝花": "28",
    "山东": "29",
    "西藏": "30",
    "中山": "31",
    "临沧": "32",
    "上饶": "33",
    "公路事业部": "34",
    "九江": "35",
    "湖州": "36",
    "舟山": "37",
    "绍兴": "38",
}

DIRECT_FIELD_MAP = {
    "姓名": "姓名",
    "身份证": "身份证",
    "入职时间": "入职时间",
    "入职登记表": "员工入职表",
    "身份证复印件": "身份证复印件",
    "银行卡复印件": "银行卡复印件",
    "体检报告单": "体检报告单",
    "学历证书": "学历证书",
    "学位证书": "学位证书",
    "相关资格证书": "相关资格证书",
    "劳动合同": "劳动合同",
    "电子照片": "照片",
    "离职证明\n（前司）": "离职证明",
    "离职证明（前司）": "离职证明",
    "入职须知": "入职员工须知",
    "员工手册签收单": "员工手册签收单",
    "安全责任书": "安全生产责任书",
    "保密协议": "保密协议",
    "竞业协议": "竞业协议",
    "员工三级安全教育": "三级安全教育登记（登记卡+试卷）",
    "务工人员健康调查表": "员工健康情况调查表",
    "人员进场记录": "员工进场记录",
    "员工异动审批表": "员工异动审批表",
    "入职考试试卷": "入职考试试卷",
    "员工转正审批表": "员工转正审批表",
    "转正考试试卷": "转正考试试卷",
    "增购社保申请单": "增购社保申请单",
    "离职申请单、交接清单": "离职申请单",
}

FORMULA_HEADERS = {"出生日期", "年齡", "年龄", "入职公式", "出生年月公式", "档案号"}
BLANK_HEADERS = {"序号", "档案柜号"}
SOURCE_SKIP_HEADERS = {HEADER_COMPANY, "出生日期", "年齡", "年龄", "入职公式", "出生年月公式"}


@dataclass(frozen=True)
class ArchiveTransferRecord:
    company: str
    name: str
    id_card: str
    values: dict[str, Any]
    source_file: str
    source_title: str
    source_row: int


@dataclass
class ArchiveImportResult:
    input_path: Path
    target_path: Path | None
    output_dir: Path
    output_file: Path | None = None
    dry_run: bool = False
    using_default_template: bool = False
    source_files: list[str] = field(default_factory=list)
    source_record_count: int = 0
    inserted_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    company_counts: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "tool_name": TOOL_NAME,
            "input_path": str(self.input_path),
            "target_path": None if self.target_path is None else str(self.target_path),
            "output_dir": str(self.output_dir),
            "output_file": None if self.output_file is None else str(self.output_file),
            "dry_run": self.dry_run,
            "using_default_template": self.using_default_template,
            "source_files": self.source_files,
            "source_file_count": len(self.source_files),
            "source_record_count": self.source_record_count,
            "inserted_count": self.inserted_count,
            "updated_count": self.updated_count,
            "skipped_count": self.skipped_count,
            "company_counts": self.company_counts,
            "warnings": self.warnings,
        }


@dataclass
class ArchiveExportResult:
    summary_path: Path
    output_dir: Path
    existing_archive_path: Path | None = None
    output_files: list[Path] = field(default_factory=list)
    dry_run: bool = False
    summary_files: list[str] = field(default_factory=list)
    existing_archive_files: list[str] = field(default_factory=list)
    company_counts: dict[str, int] = field(default_factory=dict)
    created_count: int = 0
    inserted_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "tool_name": EXPORT_TOOL_NAME,
            "summary_path": str(self.summary_path),
            "existing_archive_path": None if self.existing_archive_path is None else str(self.existing_archive_path),
            "output_dir": str(self.output_dir),
            "output_files": [str(path) for path in self.output_files],
            "output_file_count": len(self.output_files),
            "dry_run": self.dry_run,
            "summary_files": self.summary_files,
            "summary_file_count": len(self.summary_files),
            "existing_archive_files": self.existing_archive_files,
            "existing_archive_file_count": len(self.existing_archive_files),
            "company_counts": self.company_counts,
            "created_count": self.created_count,
            "inserted_count": self.inserted_count,
            "updated_count": self.updated_count,
            "skipped_count": self.skipped_count,
            "warnings": self.warnings,
        }


@dataclass(frozen=True)
class ArchiveSheetLayout:
    header_row: int
    data_start_row: int
    footer_start_row: int
    max_column: int
    headers: dict[str, int]


def import_archive_transfers(
    input_path: str | Path | list[str | Path],
    target_path: str | Path | None,
    output_dir: str | Path,
    *,
    dry_run: bool = False,
) -> ArchiveImportResult:
    input_paths = _normalize_input_paths(input_path)
    display_input = input_paths[0] if len(input_paths) == 1 else input_paths[0].parent
    target = None if target_path is None else Path(target_path).expanduser().resolve()
    output_dir = Path(output_dir).expanduser().resolve()
    warnings: list[str] = []
    for path in input_paths:
        if not path.exists():
            raise FileNotFoundError(f"档案移交表文件、压缩包或文件夹不存在：{path}")
    if target is not None:
        if not target.exists() or not target.is_file():
            raise FileNotFoundError(f"档案汇总表不存在：{target}")
        if not is_supported_excel_file(target):
            raise ValueError("档案汇总表目前只支持 .xlsx 或 .xls 文件。")

    with tempfile.TemporaryDirectory(prefix="hr_archive_import_") as temp_root:
        temp_dir = Path(temp_root)
        source_files = _find_source_files(input_paths, temp_dir, warnings)
        if not source_files:
            raise ValueError("未找到 .xlsx 或 .xls 档案移交表。")

        records: list[ArchiveTransferRecord] = []
        source_names: list[str] = []
        for source_file in source_files:
            try:
                file_records, file_warnings = _read_transfer_file(source_file)
            except ValueError:
                if len(input_paths) == 1 and input_paths[0].is_file() and is_supported_excel_file(input_paths[0]):
                    raise
                warnings.append(f"{source_file.name} 不是档案移交表，已跳过。")
                continue
            warnings.extend(file_warnings)
            if file_records:
                source_names.append(str(source_file))
            records.extend(file_records)

        result = ArchiveImportResult(
            input_path=display_input,
            target_path=target,
            output_dir=output_dir,
            dry_run=dry_run,
            using_default_template=target is None,
            source_files=source_names,
            source_record_count=len(records),
            company_counts=_count_by_company(records),
            warnings=warnings,
        )
        if dry_run:
            result.inserted_count = len(records)
            return result

        output_dir.mkdir(parents=True, exist_ok=True)
        output_file = output_dir / OUTPUT_FILENAME
        target_for_write = ensure_xlsx_workbook(target, temp_dir) if target is not None else _copy_default_archive_template(temp_dir)
        apply_result = _write_archive_summary(
            target_for_write,
            output_file,
            records,
            warnings,
            remove_template_sheet=target is None,
        )
        result.output_file = output_file
        result.inserted_count = apply_result["inserted_count"]
        result.updated_count = apply_result["updated_count"]
        result.skipped_count = apply_result["skipped_count"]
        return result


def export_company_archive_tables(
    summary_path: str | Path | list[str | Path],
    output_dir: str | Path,
    *,
    existing_archive_path: str | Path | list[str | Path] | None = None,
    dry_run: bool = False,
) -> ArchiveExportResult:
    summary_paths = _normalize_input_paths(summary_path)
    display_summary = summary_paths[0] if len(summary_paths) == 1 else summary_paths[0].parent
    existing_paths = [] if existing_archive_path is None else _normalize_input_paths(existing_archive_path)
    display_existing = None if not existing_paths else (existing_paths[0] if len(existing_paths) == 1 else existing_paths[0].parent)
    output_dir = Path(output_dir).expanduser().resolve()
    for path in summary_paths:
        if not path.exists():
            raise FileNotFoundError(f"档案汇总表文件、压缩包或文件夹不存在：{path}")
    for path in existing_paths:
        if not path.exists():
            raise FileNotFoundError(f"已有公司档案表文件、压缩包或文件夹不存在：{path}")

    warnings: list[str] = []
    with tempfile.TemporaryDirectory(prefix="hr_archive_export_") as temp_root:
        temp_dir = Path(temp_root)
        summary_files = _find_excel_input_files(summary_paths, temp_dir, warnings)
        if not summary_files:
            raise ValueError("未找到 .xlsx 或 .xls 档案汇总表。")
        records = _read_archive_summary_records(summary_files, warnings)
        company_counts = _count_by_company(records)
        existing_files = _find_excel_input_files(existing_paths, temp_dir, warnings) if existing_paths else []
        existing_by_company = _find_existing_company_archives(existing_files, company_counts.keys(), warnings)
        result = ArchiveExportResult(
            summary_path=display_summary,
            existing_archive_path=display_existing,
            output_dir=output_dir,
            dry_run=dry_run,
            summary_files=[str(path) for path in summary_files],
            existing_archive_files=[str(path) for path in existing_files],
            company_counts=company_counts,
            warnings=warnings,
        )
        if dry_run:
            return result
        if not company_counts:
            raise ValueError("档案汇总表中没有可生成档案表的公司数据。")

        output_dir.mkdir(parents=True, exist_ok=True)
        output_files: list[Path] = []
        for company, company_records in _group_by_company(records).items():
            output_file = output_dir / f"{_safe_filename(company)}-档案表.xlsx"
            write_result = _write_company_archive_file(
                company,
                company_records,
                existing_by_company.get(company),
                output_file,
                temp_dir,
                warnings,
            )
            output_files.append(output_file)
            result.created_count += 1 if write_result["created"] else 0
            result.inserted_count += write_result["inserted_count"]
            result.updated_count += write_result["updated_count"]
            result.skipped_count += write_result["skipped_count"]
        result.output_files = output_files
        return result


def _normalize_input_paths(input_path: str | Path | list[str | Path]) -> list[Path]:
    return normalize_input_paths(input_path, "请选择档案移交表文件、压缩包或文件夹。")


def _find_source_files(input_paths: list[Path], temp_dir: Path, warnings: list[str]) -> list[Path]:
    files: list[Path] = []
    seen: set[Path] = set()
    for input_path in input_paths:
        for file_path in _iter_source_files(input_path, temp_dir, warnings):
            working_path = ensure_xlsx_workbook(file_path, temp_dir)
            resolved = working_path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            files.append(working_path)
    return sorted(files)


def _iter_source_files(input_path: Path, temp_dir: Path, warnings: list[str]) -> list[Path]:
    if input_path.is_file():
        suffix = input_path.suffix.lower()
        if is_supported_excel_file(input_path):
            return [input_path]
        if suffix == ".zip":
            return extract_zip_excel_files(input_path, temp_dir, warnings)
        return []
    if not input_path.is_dir():
        raise FileNotFoundError(f"档案移交表路径不存在：{input_path}")
    files: list[Path] = []
    for path in sorted(input_path.rglob("*")):
        if not path.is_file() or path.name.startswith(("~$", ".~")):
            continue
        if is_supported_excel_file(path) and path.name != OUTPUT_FILENAME:
            files.append(path)
        elif path.suffix.lower() == ".zip":
            files.extend(extract_zip_excel_files(path, temp_dir, warnings))
    return files


def _find_excel_input_files(input_paths: list[Path], temp_dir: Path, warnings: list[str]) -> list[Path]:
    files: list[Path] = []
    seen: set[Path] = set()
    for input_path in input_paths:
        for file_path in _iter_excel_input_files(input_path, temp_dir, warnings):
            working_path = ensure_xlsx_workbook(file_path, temp_dir)
            resolved = working_path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            files.append(working_path)
    return sorted(files)


def _iter_excel_input_files(input_path: Path, temp_dir: Path, warnings: list[str]) -> list[Path]:
    if input_path.is_file():
        suffix = input_path.suffix.lower()
        if is_supported_excel_file(input_path):
            return [input_path]
        if suffix == ".zip":
            return extract_zip_excel_files(input_path, temp_dir, warnings)
        return []
    if not input_path.is_dir():
        raise FileNotFoundError(f"路径不存在：{input_path}")
    files: list[Path] = []
    for path in sorted(input_path.rglob("*")):
        if not path.is_file() or path.name.startswith("~$") or path.name.startswith(".~"):
            continue
        if is_supported_excel_file(path):
            files.append(path)
        elif path.suffix.lower() == ".zip":
            files.extend(extract_zip_excel_files(path, temp_dir, warnings))
    return files


def _read_transfer_file(file_path: Path) -> tuple[list[ArchiveTransferRecord], list[str]]:
    workbook = load_workbook(file_path, data_only=False)
    warnings: list[str] = []
    try:
        ws = _find_transfer_sheet(workbook)
        header_row = _find_header_row(ws, (HEADER_COMPANY, HEADER_NAME, HEADER_ID_CARD))
        headers = _read_headers(ws, header_row)
        required = [HEADER_COMPANY, HEADER_NAME, HEADER_ID_CARD]
        missing = [header for header in required if header not in headers]
        if missing:
            raise ValueError(f"{file_path.name} 缺少字段：{'、'.join(missing)}")
        records: list[ArchiveTransferRecord] = []
        source_title = _cell_text(ws.cell(1, 1).value)
        for row_index in range(header_row + 1, ws.max_row + 1):
            values = {header: ws.cell(row_index, col_index).value for header, col_index in headers.items()}
            company = _cell_text(values.get(HEADER_COMPANY))
            name = _cell_text(values.get(HEADER_NAME))
            id_card = _normalize_id_card(values.get(HEADER_ID_CARD))
            if not company and not name and not id_card:
                if _is_instruction_row(ws, row_index):
                    continue
                continue
            if not company or not name or not id_card:
                if _is_instruction_row(ws, row_index):
                    continue
                warnings.append(f"{file_path.name} 第 {row_index} 行缺少公司、姓名或身份证，已跳过。")
                continue
            records.append(
                ArchiveTransferRecord(
                    company=company,
                    name=name,
                    id_card=id_card,
                    values=values,
                    source_file=file_path.name,
                    source_title=source_title,
                    source_row=row_index,
                )
            )
        return records, warnings
    finally:
        workbook.close()


def _find_transfer_sheet(workbook) -> Worksheet:
    for ws in workbook.worksheets:
        try:
            _find_header_row(ws, (HEADER_COMPANY, HEADER_NAME, HEADER_ID_CARD))
            return ws
        except ValueError:
            continue
    raise ValueError("未找到包含“公司、姓名、身份证”的档案移交表。")


def _write_archive_summary(
    target_path: Path,
    output_file: Path,
    records: list[ArchiveTransferRecord],
    warnings: list[str],
    *,
    remove_template_sheet: bool = False,
) -> dict[str, int]:
    workbook = load_workbook(target_path)
    inserted_count = 0
    updated_count = 0
    skipped_count = 0
    try:
        template_sheet = workbook[workbook.sheetnames[0]]
        for company, company_records in _group_by_company(records).items():
            ws = _get_or_create_company_sheet(
                workbook,
                company,
                template_sheet,
                warnings,
                warn_created_sheet=not remove_template_sheet,
            )
            layout = _detect_archive_layout(ws)
            existing = _existing_records(ws, layout)
            new_records: list[ArchiveTransferRecord] = []
            pending_ids: set[str] = set()
            for record in company_records:
                if _detect_region_code(record) is None:
                    warnings.append(f"{record.source_file} 第 {record.source_row} 行未识别项目地区，编号已留空。")
                existing_row = existing.get(record.id_card)
                if existing_row is None:
                    if record.id_card in pending_ids:
                        warnings.append(f"{company} 身份证 {record.id_card} 在本次导入中重复，已跳过后续记录。")
                        skipped_count += 1
                        continue
                    pending_ids.add(record.id_card)
                    new_records.append(record)
                    continue
                if _cell_text(ws.cell(existing_row, layout.headers[HEADER_NAME]).value) != record.name:
                    warnings.append(f"{company} 身份证 {record.id_card} 已存在，但姓名不同：{record.name}")
                if _merge_existing_archive_row(ws, layout, existing_row, record):
                    updated_count += 1
                else:
                    skipped_count += 1
            if new_records:
                _append_archive_rows(ws, layout, new_records)
                inserted_count += len(new_records)
        if remove_template_sheet:
            _remove_unused_template_sheet(workbook)
        workbook.save(output_file)
    finally:
        workbook.close()
    return {"inserted_count": inserted_count, "updated_count": updated_count, "skipped_count": skipped_count}


def _get_or_create_company_sheet(
    workbook,
    company: str,
    template_sheet: Worksheet,
    warnings: list[str],
    *,
    warn_created_sheet: bool,
) -> Worksheet:
    if company in workbook.sheetnames:
        return workbook[company]
    copied = workbook.copy_worksheet(template_sheet)
    copied.title = _safe_sheet_title(company, workbook.sheetnames)
    _clear_archive_data_rows(copied)
    if warn_created_sheet:
        warnings.append(f"档案汇总表缺少工作表：{company}，已按模板自动创建。")
    return copied


def _safe_sheet_title(title: str, existing_titles: list[str]) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", title).strip()[:31] or "未命名公司"
    if cleaned not in existing_titles:
        return cleaned
    base = cleaned[:28]
    counter = 1
    while f"{base}_{counter}" in existing_titles:
        counter += 1
    return f"{base}_{counter}"


def _detect_archive_layout(ws: Worksheet) -> ArchiveSheetLayout:
    header_row = _find_header_row(ws, (HEADER_NAME, HEADER_ID_CARD))
    headers = _read_headers(ws, header_row)
    footer_start_row = _find_footer_start_row(ws, header_row + 1)
    return ArchiveSheetLayout(
        header_row=header_row,
        data_start_row=header_row + 1,
        footer_start_row=footer_start_row,
        max_column=_last_header_column(ws, header_row),
        headers=headers,
    )


def _find_header_row(ws: Worksheet, required_headers: tuple[str, ...]) -> int:
    normalized_required = {_normalize_header(header) for header in required_headers}
    for row_index in range(1, min(ws.max_row, 20) + 1):
        headers = {_normalize_header(ws.cell(row_index, col_index).value) for col_index in range(1, ws.max_column + 1)}
        if normalized_required.issubset(headers):
            return row_index
    raise ValueError(f"{ws.title} 未找到表头：{'、'.join(required_headers)}")


def _read_headers(ws: Worksheet, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_index in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_index).value
        if value in (None, ""):
            continue
        headers[_normalize_header(value)] = col_index
    return headers


def _last_header_column(ws: Worksheet, header_row: int) -> int:
    last_col = 1
    for col_index in range(1, ws.max_column + 1):
        if ws.cell(header_row, col_index).value not in (None, ""):
            last_col = col_index
    return last_col


def _find_footer_start_row(ws: Worksheet, start_row: int) -> int:
    for row_index in range(start_row, ws.max_row + 1):
        row_text = " ".join(_cell_text(ws.cell(row_index, col_index).value) for col_index in range(1, min(ws.max_column, 6) + 1))
        if any(keyword in row_text for keyword in ("对应行", "汇总表中的公司", "交接档案室", "备注：")):
            return row_index
    return ws.max_row + 1


def _clear_archive_data_rows(ws: Worksheet) -> None:
    layout = _detect_archive_layout(ws)
    for row_index in range(layout.data_start_row, layout.footer_start_row):
        for col_index in range(1, layout.max_column + 1):
            ws.cell(row_index, col_index).value = None


def _existing_records(ws: Worksheet, layout: ArchiveSheetLayout) -> dict[str, int]:
    id_col = layout.headers[HEADER_ID_CARD]
    records: dict[str, int] = {}
    for row_index in range(layout.data_start_row, layout.footer_start_row):
        id_card = _normalize_id_card(ws.cell(row_index, id_col).value)
        if id_card:
            records[id_card] = row_index
    return records


def _append_archive_rows(ws: Worksheet, layout: ArchiveSheetLayout, records: list[ArchiveTransferRecord]) -> None:
    template_row = _template_row(ws, layout)
    template_snapshot = snapshot_row(ws, template_row, layout.max_column)
    target_rows = _blank_data_rows(ws, layout)
    remaining_count = len(records) - len(target_rows)
    if remaining_count > 0:
        insert_at = layout.footer_start_row
        insert_rows(ws, insert_at, remaining_count)
        target_rows.extend(range(insert_at, insert_at + remaining_count))
    for row_index, record in zip(target_rows, records):
        apply_row_snapshot(ws, row_index, template_snapshot, translate_formulas=True)
        _clear_archive_record_values(ws, layout, row_index)
        _write_archive_record(ws, layout, row_index, record)
        _format_archive_data_row(ws, layout, row_index)


def _clear_archive_record_values(ws: Worksheet, layout: ArchiveSheetLayout, row_index: int) -> None:
    for col_index in range(1, layout.max_column + 1):
        ws.cell(row_index, col_index).value = None


def _blank_data_rows(ws: Worksheet, layout: ArchiveSheetLayout) -> list[int]:
    rows: list[int] = []
    name_col = layout.headers.get(HEADER_NAME)
    id_col = layout.headers.get(HEADER_ID_CARD)
    if name_col is None or id_col is None:
        return rows
    for row_index in range(layout.data_start_row, layout.footer_start_row):
        if not _has_value(ws.cell(row_index, name_col).value) and not _has_value(ws.cell(row_index, id_col).value):
            rows.append(row_index)
    return rows


def _template_row(ws: Worksheet, layout: ArchiveSheetLayout) -> int:
    if layout.data_start_row < layout.footer_start_row:
        return layout.data_start_row
    return layout.header_row


def _merge_existing_archive_row(ws: Worksheet, layout: ArchiveSheetLayout, row_index: int, record: ArchiveTransferRecord) -> bool:
    changed = False
    target_values = _target_values_for_record(record, layout.headers)
    for header, value in target_values.items():
        col_index = layout.headers.get(header)
        if col_index is None or header in FORMULA_HEADERS or header in BLANK_HEADERS:
            continue
        cell = ws.cell(row_index, col_index)
        if header == HEADER_OTHER:
            changed = _merge_other_cell(cell, value) or changed
        elif _has_value(value) and not _has_value(cell.value):
            cell.value = value
            changed = True
    return changed


def _merge_other_cell(cell, value: Any) -> bool:
    value_text = _cell_text(value)
    if not value_text:
        return False
    current = _cell_text(cell.value)
    if not current:
        cell.value = value_text
        return True
    if value_text in current:
        return False
    cell.value = f"{current}；{value_text}"
    return True


def _write_archive_record(ws: Worksheet, layout: ArchiveSheetLayout, row_index: int, record: ArchiveTransferRecord) -> None:
    target_values = _target_values_for_record(record, layout.headers)
    for header, col_index in layout.headers.items():
        if header in BLANK_HEADERS:
            ws.cell(row_index, col_index).value = None
        elif header in target_values:
            ws.cell(row_index, col_index).value = target_values[header]
    _write_archive_formulas(ws, layout, row_index)


def _target_values_for_record(record: ArchiveTransferRecord, target_headers: dict[str, int]) -> dict[str, Any]:
    values: dict[str, Any] = {}
    code = _detect_region_code(record)
    if "编号" in target_headers:
        values["编号"] = _cell_text(record.values.get("编号")) or code
    other_parts: list[str] = []
    for source_header, value in record.values.items():
        normalized_source = _normalize_header(source_header)
        if normalized_source in SOURCE_SKIP_HEADERS or normalized_source in FORMULA_HEADERS:
            continue
        if normalized_source in target_headers:
            values[normalized_source] = value
            continue
        target_header = DIRECT_FIELD_MAP.get(normalized_source)
        if target_header and target_header in target_headers:
            values[target_header] = value
            continue
        if normalized_source in DIRECT_FIELD_MAP and DIRECT_FIELD_MAP[normalized_source] not in target_headers:
            other_parts.append(_format_other_part(normalized_source, value))
        elif normalized_source not in DIRECT_FIELD_MAP and _has_value(value):
            other_parts.append(_format_other_part(normalized_source, value))
    other_text = "；".join(part for part in other_parts if part)
    if other_text and HEADER_OTHER in target_headers:
        values[HEADER_OTHER] = other_text
    return values


def _write_archive_formulas(ws: Worksheet, layout: ArchiveSheetLayout, row_index: int) -> None:
    id_col = _column_ref(layout, HEADER_ID_CARD, row_index)
    birth_col = _column_ref(layout, "出生日期", row_index)
    entry_date_col = _column_ref(layout, "入职时间", row_index)
    entry_formula_col = _column_ref(layout, "入职公式", row_index)
    code_col = _column_ref(layout, "编号", row_index)
    birth_month_col = _column_ref(layout, "出生年月公式", row_index)
    serial_col = _column_ref(layout, "序号", row_index)
    # 人事提供的档案号公式为“编号-入职公式-出生年月公式-序号”，不同模板列序可能不同。
    formulas = {
        "出生日期": f'=MIDB({id_col},7,4)&"-"&MIDB({id_col},11,2)&"-"&MIDB({id_col},13,2)',
        "年齡": f"=(TODAY()-{birth_col})/365",
        "年龄": f"=(TODAY()-{birth_col})/365",
        "入职公式": f'=TEXT({entry_date_col},"yyyymmdd")',
        "档案号": f'={code_col}&"-"&TEXT({entry_formula_col},"00000000")&"-"&TEXT({birth_month_col},"00")&"-"&{serial_col}',
        "出生年月公式": f'=CONCATENATE(MID({id_col},7,4),"",MID({id_col},11,2))',
    }
    for header, formula in formulas.items():
        col_index = layout.headers.get(header)
        if col_index is not None:
            ws.cell(row_index, col_index).value = formula


def _format_archive_data_row(
    ws: Worksheet,
    layout: ArchiveSheetLayout,
    row_index: int,
    *,
    clear_fill: bool = True,
) -> None:
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    font = Font(name="宋体", size=10, bold=False, color="000000")
    empty_fill = PatternFill(fill_type=None)
    for col_index in range(1, layout.max_column + 1):
        cell = ws.cell(row_index, col_index)
        cell.border = border
        cell.alignment = alignment
        cell.font = font
        if clear_fill:
            cell.fill = empty_fill


def _normalize_archive_output_sheet(ws: Worksheet, layout: ArchiveSheetLayout) -> None:
    name_col = layout.headers.get(HEADER_NAME)
    id_col = layout.headers.get(HEADER_ID_CARD)
    if name_col is None or id_col is None:
        return
    for row_index in range(layout.data_start_row, layout.footer_start_row):
        if _has_value(ws.cell(row_index, name_col).value) or _has_value(ws.cell(row_index, id_col).value):
            _format_archive_data_row(ws, layout, row_index, clear_fill=False)
    _apply_archive_column_widths(ws, layout)


def _apply_archive_column_widths(ws: Worksheet, layout: ArchiveSheetLayout) -> None:
    widths = {
        "编号": 8,
        HEADER_NAME: 12,
        HEADER_ID_CARD: 24,
        "出生日期": 13,
        "年齡": 8,
        "年龄": 8,
        "入职时间": 13,
        "入职公式": 12,
        "序号": 8,
        "档案号": 24,
        "出生年月公式": 12,
        "档案柜号": 10,
    }
    for header, width in widths.items():
        col_index = layout.headers.get(header)
        if col_index is None:
            continue
        letter = get_column_letter(col_index)
        current = ws.column_dimensions[letter].width or 0
        ws.column_dimensions[letter].width = max(current, width)


def _column_ref(layout: ArchiveSheetLayout, header: str, row_index: int) -> str:
    col_index = layout.headers.get(header)
    if col_index is None:
        return f"A{row_index}"
    return f"{get_column_letter(col_index)}{row_index}"


def _detect_region_code(record: ArchiveTransferRecord) -> str | None:
    haystack = f"{record.source_file} {record.source_title} {record.company}"
    for name, code in sorted(REGION_CODES.items(), key=lambda item: len(item[0]), reverse=True):
        if name in haystack:
            return code
    return None


def _copy_default_archive_template(temp_dir: Path) -> Path:
    target = temp_dir / DEFAULT_ARCHIVE_SUMMARY_TEMPLATE_RESOURCE
    with open_template_resource(DEFAULT_ARCHIVE_SUMMARY_TEMPLATE_RESOURCE) as source, target.open("wb") as output:
        shutil.copyfileobj(source, output)
    return target


def _copy_default_company_archive_template(temp_dir: Path) -> Path:
    target = temp_dir / DEFAULT_ARCHIVE_COMPANY_TEMPLATE_RESOURCE
    with open_template_resource(DEFAULT_ARCHIVE_COMPANY_TEMPLATE_RESOURCE) as source, target.open("wb") as output:
        shutil.copyfileobj(source, output)
    return target


def _read_archive_summary_records(summary_files: list[Path], warnings: list[str]) -> list[ArchiveTransferRecord]:
    records: list[ArchiveTransferRecord] = []
    for summary_file in summary_files:
        workbook = load_workbook(summary_file, data_only=False)
        try:
            for ws in workbook.worksheets:
                if _is_placeholder_sheet_title(ws.title):
                    continue
                try:
                    layout = _detect_archive_layout(ws)
                except ValueError:
                    warnings.append(f"{summary_file.name} 的 {ws.title} 未识别到档案表表头，已跳过。")
                    continue
                for row_index in range(layout.data_start_row, layout.footer_start_row):
                    name = _cell_text(ws.cell(row_index, layout.headers[HEADER_NAME]).value)
                    id_card = _normalize_id_card(ws.cell(row_index, layout.headers[HEADER_ID_CARD]).value)
                    if not name and not id_card:
                        continue
                    if not name or not id_card:
                        warnings.append(f"{summary_file.name} 的 {ws.title} 第 {row_index} 行缺少姓名或身份证，已跳过。")
                        continue
                    values = {header: ws.cell(row_index, col_index).value for header, col_index in layout.headers.items()}
                    values[HEADER_COMPANY] = ws.title
                    records.append(
                        ArchiveTransferRecord(
                            company=ws.title,
                            name=name,
                            id_card=id_card,
                            values=values,
                            source_file=summary_file.name,
                            source_title=ws.title,
                            source_row=row_index,
                        )
                    )
        finally:
            workbook.close()
    return records


def _find_existing_company_archives(
    archive_files: list[Path],
    companies: Any,
    warnings: list[str],
) -> dict[str, Path]:
    company_names = list(companies)
    matched: dict[str, Path] = {}
    for archive_file in archive_files:
        for company in _match_company_archive_file(archive_file, company_names, warnings):
            if company in matched:
                warnings.append(f"{company} 匹配到多个已有档案表，已使用第一个：{Path(matched[company]).name}")
                continue
            matched[company] = archive_file
    return matched


def _match_company_archive_file(archive_file: Path, companies: list[str], warnings: list[str]) -> list[str]:
    workbook = load_workbook(archive_file, data_only=False)
    try:
        archive_sheet_titles: list[str] = []
        for ws in workbook.worksheets:
            try:
                _detect_archive_layout(ws)
            except ValueError:
                continue
            archive_sheet_titles.append(ws.title)
        if not archive_sheet_titles:
            warnings.append(f"{archive_file.name} 未识别到公司档案表表头，已跳过。")
            return []
        # 精确匹配
        matched = [company for company in companies if company in archive_sheet_titles]
        if matched:
            return matched
        # 模糊匹配：规范化名称相同
        for company in companies:
            company_normalized = _normalize_company_name(company)
            for title in archive_sheet_titles:
                if company_normalized == _normalize_company_name(title):
                    matched.append(company)
                    break
        if matched:
            return matched
        filename_matches = [company for company in companies if company and company in archive_file.stem]
        if filename_matches:
            return [max(filename_matches, key=len)]
        return []
    finally:
        workbook.close()


def _write_company_archive_file(
    company: str,
    records: list[ArchiveTransferRecord],
    existing_file: Path | None,
    output_file: Path,
    temp_dir: Path,
    warnings: list[str],
) -> dict[str, int | bool]:
    source_file = existing_file or _copy_default_company_archive_template(temp_dir)
    workbook = load_workbook(source_file)
    created = existing_file is None
    try:
        ws = _select_company_archive_sheet(workbook, company)
        for other_ws in list(workbook.worksheets):
            if other_ws is not ws:
                workbook.remove(other_ws)
        ws.title = _safe_sheet_title(company, [])
        if ws["A1"].value:
            ws["A1"].value = f"{company}人员档案编号表"
        layout = _detect_archive_layout(ws)
        write_counts = _append_or_merge_archive_records(ws, layout, records, warnings)
        _normalize_archive_output_sheet(ws, layout)
        workbook.save(output_file)
        return {
            "created": created,
            "inserted_count": write_counts["inserted_count"],
            "updated_count": write_counts["updated_count"],
            "skipped_count": write_counts["skipped_count"],
        }
    finally:
        workbook.close()


def _select_company_archive_sheet(workbook, company: str) -> Worksheet:
    if company in workbook.sheetnames:
        return workbook[company]
    archive_sheets: list[Worksheet] = []
    for ws in workbook.worksheets:
        try:
            _detect_archive_layout(ws)
        except ValueError:
            continue
        archive_sheets.append(ws)
    if not archive_sheets:
        raise ValueError(f"{company} 的已有档案表未识别到表头。")
    return archive_sheets[0]


def _append_or_merge_archive_records(
    ws: Worksheet,
    layout: ArchiveSheetLayout,
    records: list[ArchiveTransferRecord],
    warnings: list[str],
) -> dict[str, int]:
    existing = _existing_records(ws, layout)
    new_records: dict[str, ArchiveTransferRecord] = {}
    updated_count = 0
    skipped_count = 0
    for record in records:
        existing_row = existing.get(record.id_card)
        if existing_row is not None:
            if _cell_text(ws.cell(existing_row, layout.headers[HEADER_NAME]).value) != record.name:
                warnings.append(f"{record.company} 身份证 {record.id_card} 已存在，但姓名不同：{record.name}")
            if _merge_existing_archive_row(ws, layout, existing_row, record):
                updated_count += 1
            else:
                skipped_count += 1
            continue
        if record.id_card in new_records:
            new_records[record.id_card] = _merge_archive_records(new_records[record.id_card], record)
            warnings.append(f"{record.company} 身份证 {record.id_card} 在汇总表中重复，已合并为一条。")
            skipped_count += 1
            continue
        new_records[record.id_card] = record
    if new_records:
        _append_archive_rows(ws, layout, list(new_records.values()))
    return {
        "inserted_count": len(new_records),
        "updated_count": updated_count,
        "skipped_count": skipped_count,
    }


def _merge_archive_records(base: ArchiveTransferRecord, extra: ArchiveTransferRecord) -> ArchiveTransferRecord:
    values = dict(base.values)
    for header, value in extra.values.items():
        if _has_value(value) and not _has_value(values.get(header)):
            values[header] = value
    return ArchiveTransferRecord(
        company=base.company,
        name=base.name,
        id_card=base.id_card,
        values=values,
        source_file=base.source_file,
        source_title=base.source_title,
        source_row=base.source_row,
    )


def _scan_exportable_archive_sheets(summary_path: Path, warnings: list[str]) -> dict[str, int]:
    workbook = load_workbook(summary_path, data_only=False)
    company_counts: dict[str, int] = {}
    try:
        for ws in workbook.worksheets:
            if _is_placeholder_sheet_title(ws.title):
                continue
            try:
                layout = _detect_archive_layout(ws)
            except ValueError:
                warnings.append(f"{ws.title} 未识别到档案表表头，已跳过。")
                continue
            record_count = _count_archive_records(ws, layout)
            if record_count:
                company_counts[ws.title] = record_count
    finally:
        workbook.close()
    return company_counts


def _count_archive_records(ws: Worksheet, layout: ArchiveSheetLayout) -> int:
    name_col = layout.headers.get(HEADER_NAME)
    id_col = layout.headers.get(HEADER_ID_CARD)
    if name_col is None or id_col is None:
        return 0
    count = 0
    for row_index in range(layout.data_start_row, layout.footer_start_row):
        if _has_value(ws.cell(row_index, name_col).value) and _has_value(ws.cell(row_index, id_col).value):
            count += 1
    return count


def _remove_unused_template_sheet(workbook) -> None:
    if len(workbook.worksheets) <= 1:
        return
    for ws in list(workbook.worksheets):
        if not _is_placeholder_sheet_title(ws.title):
            continue
        try:
            layout = _detect_archive_layout(ws)
        except ValueError:
            continue
        if _count_archive_records(ws, layout) == 0 and len(workbook.worksheets) > 1:
            workbook.remove(ws)


def _is_placeholder_sheet_title(title: str) -> bool:
    return title.strip() in PLACEHOLDER_SHEET_TITLES


def _safe_filename(name: str) -> str:
    cleaned = _COMPANY_FILE_INVALID.sub("_", name).strip()
    return cleaned or "未命名公司"


def _format_other_part(header: str, value: Any) -> str:
    if not _has_value(value):
        return ""
    return f"{header}：{_cell_text(value)}"


def _count_by_company(records: list[ArchiveTransferRecord]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for record in records:
        counts[record.company] = counts.get(record.company, 0) + 1
    return counts


def _group_by_company(records: list[ArchiveTransferRecord]) -> dict[str, list[ArchiveTransferRecord]]:
    grouped: dict[str, list[ArchiveTransferRecord]] = {}
    # 建立规范化名称到标准名称的映射，用于模糊匹配
    canonical_names: dict[str, str] = {}
    for record in records:
        normalized = _normalize_company_name(record.company)
        if normalized in canonical_names:
            # 使用已有的标准名称
            canonical = canonical_names[normalized]
        else:
            # 新的规范化名称，使用当前公司名作为标准名称
            canonical = record.company
            canonical_names[normalized] = canonical
        grouped.setdefault(canonical, []).append(record)
    return grouped


# 规范化公司名称时,移除空白、常见分隔符
_COMPANY_NAME_STRIP_CHARS = str.maketrans("", "", " 　·.。、，,（）()-—_")

def _normalize_company_name(name: str) -> str:
    """规范化公司名称：移除空白和常见分隔符，然后字符排序，用于模糊匹配。

    注：只剔除空白和分隔符，不做语义处理；字符排序会让"北京春苗"和"春苗北京"
    这类字符组成相同的名称产生相同的 key，由调用方决定是否进一步去歧义。
    """
    cleaned = name.translate(_COMPANY_NAME_STRIP_CHARS)
    # 字符排序，使"北京春苗"和"春苗北京"得到相同的规范化结果
    return "".join(sorted(cleaned))


def _normalize_header(value: Any) -> str:
    return _HEADER_WHITESPACE.sub("", str(value or "").strip())


def _is_instruction_row(ws: Worksheet, row_index: int) -> bool:
    row_text = " ".join(_cell_text(ws.cell(row_index, col_index).value) for col_index in range(1, min(ws.max_column, 8) + 1))
    return any(keyword in row_text for keyword in ("备注：", "对应", "不用手动填写", "归档档案", "同一行", "如有交接清单", "其他档案"))


def _normalize_id_card(value: Any) -> str:
    return _cell_text(value).upper()



def _has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True
