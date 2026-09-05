from __future__ import annotations

import hashlib
import json
import math
import os
import re
import shutil
import threading
import time
import zipfile
import xml.etree.ElementTree as ET
from array import array
from collections import deque
from contextlib import contextmanager
from dataclasses import dataclass, field, replace
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover - Win7 构建改用 PDFium
    PdfReader = None  # type: ignore[assignment,misc]

try:
    import pypdfium2 as pdfium
except ImportError:  # pragma: no cover - 现代构建使用纯 Python pypdf
    pdfium = None  # type: ignore[assignment]

from ..common.excel import SheetGrid
from ..common.excel_compat import ensure_xlsx_workbook, is_supported_excel_file
from ..common.filenames import safe_filename

try:
    from zoneinfo import ZoneInfo
    _BEIJING_TZ = ZoneInfo("Asia/Shanghai")
except Exception:  # pragma: no cover - 兼容性回退
    _BEIJING_TZ = timezone(timedelta(hours=8))


def _beijing_now_str() -> str:
    """统一使用北京时间生成标准格式时间字符串 YYYY-MM-DD HH:MM:SS。"""
    return datetime.now(tz=_BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S")


# OCR 引擎全局单例与线程安全锁
_OCR_ENGINE = None
_OCR_ATTEMPTED = False
_OCR_LOCK = threading.Lock()
_PDF_LOCK = threading.RLock()


def _ocr_runtime_options() -> dict[str, int]:
    """Reserve one CPU for the desktop event loop on supported low-core machines."""

    cpu_count = max(1, int(os.cpu_count() or 1))
    if cpu_count > 4:
        # Keep the OCR package's tuned defaults on modern hardware.
        return {}
    return {
        "intra_op_num_threads": max(1, cpu_count - 1),
        "inter_op_num_threads": 1,
    }


@contextmanager
def _rapidocr_low_memory_session_options() -> Iterator[None]:
    """Disable ONNX's retained input-shape allocation plan during engine init.

    RapidOCR 1.4.4 creates its three ONNX sessions through one private factory.
    ONNX Runtime's default memory-pattern planner can retain hundreds of MB
    after the second high-resolution image even though RapidOCR already turns
    off the CPU arena.  Disabling only that allocator optimization leaves the
    model, preprocessing, tensors, and recognition rules unchanged.  Keep the
    hook scoped to construction and fall back silently if a future compatible
    RapidOCR version moves the factory or an older runtime lacks the option.
    """

    try:
        from rapidocr_onnxruntime.utils.infer_engine import OrtInferSession

        original_descriptor = OrtInferSession.__dict__.get("_init_sess_opts")
        original_factory = getattr(OrtInferSession, "_init_sess_opts")
        if original_descriptor is None or not callable(original_factory):
            yield
            return
    except Exception:
        yield
        return

    def _bounded_session_options(config):
        session_options = original_factory(config)
        try:
            session_options.enable_mem_pattern = False
        except Exception:
            pass
        return session_options

    try:
        OrtInferSession._init_sess_opts = staticmethod(_bounded_session_options)
    except Exception:
        yield
        return
    try:
        yield
    finally:
        try:
            OrtInferSession._init_sess_opts = original_descriptor
        except Exception:
            pass


def _get_ocr_engine():
    global _OCR_ENGINE, _OCR_ATTEMPTED
    with _OCR_LOCK:
        if not _OCR_ATTEMPTED:
            _OCR_ATTEMPTED = True
            try:
                from rapidocr_onnxruntime import RapidOCR
                options = _ocr_runtime_options()
                with _rapidocr_low_memory_session_options():
                    try:
                        _OCR_ENGINE = RapidOCR(**options)
                    except TypeError:
                        # Forward-compatible fallback if a future RapidOCR removes
                        # the thread-control keyword arguments.
                        _OCR_ENGINE = RapidOCR()
            except Exception:
                _OCR_ENGINE = None
        return _OCR_ENGINE


# OCR 智能索引缓存：写入资料库根目录的隐藏 JSON 文件
_OCR_CACHE_FILE_NAME = ".hr_material_index_cache.json"
_OCR_CACHE_VERSION = 6
_OCR_CACHE_TEXT_SNIPPET_MAX = 4096
_OCR_CACHE_FILE_MAX_BYTES = 64 * 1024 * 1024
_OCR_CACHE_FILE_TRIM_BYTES = 48 * 1024 * 1024
_OCR_CACHE_FILE_LOAD_MAX_BYTES = 256 * 1024 * 1024
_OCR_CACHE_ENTRY_MAX_AGE_DAYS = 90
_OCR_CACHE_HASH_WINDOW = 1 * 1024 * 1024
_OCR_CACHE_HASH_TRIGGER_SIZE = 10 * 1024 * 1024
_OCR_MEMORY_CACHE_MIN_BYTES = 1 * 1024 * 1024
_OCR_MEMORY_CACHE_LOCK = threading.RLock()
_OCR_MEMORY_CACHE_PATH: str | None = None
_OCR_MEMORY_CACHE_SIGNATURE: tuple[int, int, int, int, int] | None = None
_OCR_MEMORY_CACHE_DATA: dict[str, Any] | None = None
_OFFICE_ARCHIVE_MAX_MEMBERS = 20_000
_OFFICE_XML_MEMBER_MAX_BYTES = 16 * 1024 * 1024
_OFFICE_XML_TOTAL_MAX_BYTES = 64 * 1024 * 1024
_OFFICE_XML_MAX_COMPRESSION_RATIO = 500
_PDF_MAX_FILE_BYTES = 64 * 1024 * 1024
_PDF_MAX_PAGES = 100
_PDF_MAX_TEXT_CHARS = 2_000_000
_PDF_MAX_XOBJECTS_PER_PAGE = 64
_PDF_MAX_IMAGES_PER_PAGE = 8
_PDF_MAX_IMAGE_PIXELS = 12_000_000
_PDF_MAX_DECODED_IMAGE_BYTES = 64 * 1024 * 1024
_PDF_MAX_TOTAL_IMAGE_PIXELS = 120_000_000
_PDF_MAX_TOTAL_DECODED_IMAGE_BYTES = 512 * 1024 * 1024
_PDF_MAX_ENCODED_IMAGE_BYTES = 32 * 1024 * 1024
_PDF_PAGE_RENDER_SCALE = 2.0
_PDF_BACKEND = "pypdf" if PdfReader is not None else "pdfium"

if PdfReader is not None:
    try:
        import pypdf.filters as _pypdf_filters

        _pypdf_filters.ZLIB_MAX_OUTPUT_LENGTH = min(
            int(getattr(_pypdf_filters, "ZLIB_MAX_OUTPUT_LENGTH", 75_000_000)),
            _PDF_MAX_DECODED_IMAGE_BYTES,
        )
    except Exception:
        pass


class PDFRecognitionError(ValueError):
    """PDF 无法被安全、完整地读取或识别。"""


class PDFResourceLimitError(PDFRecognitionError):
    """PDF 超过为老旧电脑设置的资源安全门禁。"""


class MaterialCollectionCancelled(RuntimeError):
    """用户请求停止本次资料收集。"""


# ---------------------------------------------------------------------------
# OCR 智能索引缓存层：纯函数（无副作用，便于单测）
# ---------------------------------------------------------------------------


def _get_engine_signature() -> str:
    """获取当前 OCR 引擎的版本签名；用于缓存条目与引擎版本一致性校验。"""
    try:
        import rapidocr_onnxruntime

        version = getattr(rapidocr_onnxruntime, "__version__", "unknown")
        return f"rapidocr_onnxruntime@{version}"
    except Exception:
        return "rapidocr_onnxruntime@unknown"


def _get_pdf_backend_signature() -> str:
    """区分现代与 Win7 PDF 后端，切换安装包时只失效 PDF 缓存。"""
    if _PDF_BACKEND == "pdfium" and pdfium is not None:
        version = getattr(getattr(pdfium, "PYPDFIUM_INFO", None), "version", "unknown")
        return f"pdfium@{version}"
    if PdfReader is not None:
        try:
            import pypdf

            return f"pypdf@{getattr(pypdf, '__version__', 'unknown')}"
        except Exception:
            pass
    return f"{_PDF_BACKEND}@unknown"


def _compute_file_fingerprint(file_path: Path) -> tuple[int, float, str] | None:
    """旧文件夹模式的缓存指纹；大文件保持原有的前 1MB hash 行为。"""
    try:
        stat = file_path.stat()
    except (FileNotFoundError, OSError):
        return None

    size = stat.st_size
    mtime = stat.st_mtime
    sha = hashlib.sha256()
    try:
        with open(file_path, "rb") as fp:
            if size <= _OCR_CACHE_HASH_TRIGGER_SIZE:
                while True:
                    chunk = fp.read(64 * 1024)
                    if not chunk:
                        break
                    sha.update(chunk)
            else:
                sha.update(fp.read(_OCR_CACHE_HASH_WINDOW))
    except OSError:
        return None
    return (size, mtime, sha.hexdigest())


def _windows_file_change_time(file_path: Path) -> int | None:
    """Return the Windows file change time, not ``stat().st_ctime``.

    Python 3.12 still exposes file creation time through ``st_ctime`` on
    Windows.  Creation time does not change when an existing file is
    overwritten, so it cannot safely guard the metadata-only cache fast path.
    ``FILE_BASIC_INFO.ChangeTime`` is the NT file change token we need here.
    """
    if os.name != "nt":
        return None
    try:
        import ctypes
        from ctypes import wintypes

        class _FileBasicInfo(ctypes.Structure):
            _fields_ = [
                ("CreationTime", ctypes.c_longlong),
                ("LastAccessTime", ctypes.c_longlong),
                ("LastWriteTime", ctypes.c_longlong),
                ("ChangeTime", ctypes.c_longlong),
                ("FileAttributes", wintypes.DWORD),
            ]

        kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
        create_file = kernel32.CreateFileW
        create_file.argtypes = [
            wintypes.LPCWSTR,
            wintypes.DWORD,
            wintypes.DWORD,
            wintypes.LPVOID,
            wintypes.DWORD,
            wintypes.DWORD,
            wintypes.HANDLE,
        ]
        create_file.restype = wintypes.HANDLE
        get_file_information = kernel32.GetFileInformationByHandleEx
        get_file_information.argtypes = [
            wintypes.HANDLE,
            ctypes.c_int,
            wintypes.LPVOID,
            wintypes.DWORD,
        ]
        get_file_information.restype = wintypes.BOOL
        close_handle = kernel32.CloseHandle
        close_handle.argtypes = [wintypes.HANDLE]
        close_handle.restype = wintypes.BOOL

        file_read_attributes = 0x0080
        file_share_all = 0x0001 | 0x0002 | 0x0004
        open_existing = 3
        file_attribute_normal = 0x0080
        invalid_handle = wintypes.HANDLE(-1).value
        handle = create_file(
            os.path.abspath(os.fspath(file_path)),
            file_read_attributes,
            file_share_all,
            None,
            open_existing,
            file_attribute_normal,
            None,
        )
        if handle == invalid_handle:
            return None
        try:
            info = _FileBasicInfo()
            file_basic_info = 0
            if not get_file_information(
                handle,
                file_basic_info,
                ctypes.byref(info),
                ctypes.sizeof(info),
            ):
                return None
            change_time = int(info.ChangeTime)
            return change_time if change_time > 0 else None
        finally:
            close_handle(handle)
    except Exception:  # pragma: no cover - platform API failure must degrade safely
        # FAT/network shares or restricted handles may not expose ChangeTime.
        # Callers treat None as unsafe for metadata-only reuse and hash instead.
        return None


def _file_change_token(file_path: Path, stat_result: os.stat_result) -> int | None:
    if os.name == "nt":
        return _windows_file_change_time(file_path)
    return stat_result.st_ctime_ns


def _stream_full_sha256(file_path: Path) -> str | None:
    sha = hashlib.sha256()
    try:
        with open(file_path, "rb") as fp:
            while True:
                chunk = fp.read(1024 * 1024)
                if not chunk:
                    break
                sha.update(chunk)
    except OSError:
        return None
    return sha.hexdigest()


def _compute_full_file_fingerprint(file_path: Path) -> tuple[int, float, str] | None:
    """TASK-8 无序索引专用完整指纹，保证大文件后半段变化也能失效。

    返回 None 表示文件无法访问。首次索引流式计算完整 SHA-256；后续同一路径
    会先用 size/mtime/可靠变更标记命中路径索引，避免重复读取大文件。
    """
    initial_metadata = _flat_path_metadata(file_path)
    if initial_metadata is None:
        return None
    size, mtime_ns, change_token = initial_metadata
    sha = _stream_full_sha256(file_path)
    if sha is None:
        return None
    final_metadata = _flat_path_metadata(file_path)
    if final_metadata != initial_metadata:
        return None

    if change_token is None:
        # Without a reliable OS change token, verify the bytes twice.  This
        # fallback is slower but prevents same-size, restored-mtime changes
        # from producing a stable-looking fingerprint on unsupported volumes.
        verified_sha = _stream_full_sha256(file_path)
        if verified_sha != sha or _flat_path_metadata(file_path) != final_metadata:
            return None

    return (size, mtime_ns / 1_000_000_000, sha)


def _compute_cache_key(
    file_path: Path,
    employee_key: str = "",
) -> str | None:
    """根据文件二进制内容哈希 + 文件大小生成唯一指纹（完全脱离文件名与路径）。

    不管文件被重命名为任何名称、移动到任何子目录，只要内容未变，指纹恒定不变。
    """
    fingerprint = _compute_file_fingerprint(file_path)
    if fingerprint is None:
        return None
    size, _mtime, sha = fingerprint
    return f"{sha[:24]}_{size}"


def _mask_id_card(id_card: str) -> str:
    """身份证号脱敏：仅保留前 4 与后 4 位；空串直接返回。"""
    if not id_card:
        return ""
    if len(id_card) <= 8:
        return id_card[:2] + "*" * (len(id_card) - 4) + id_card[-2:]
    return id_card[:4] + "*" * (len(id_card) - 8) + id_card[-4:]


def _hash_id_card(id_card: str) -> str:
    """身份证号 sha256；用于 mismatch 校验但不在缓存文件留明文。"""
    if not id_card:
        return ""
    return hashlib.sha256(id_card.encode("utf-8")).hexdigest()[:16]


def _hash_phone(phone: str) -> str:
    if not phone:
        return ""
    return hashlib.sha256(phone.encode("utf-8")).hexdigest()[:16]


def _new_ocr_cache() -> dict[str, Any]:
    now = _beijing_now_str()
    return {
        "version": _OCR_CACHE_VERSION,
        "engine_signature": _get_engine_signature(),
        "pdf_backend_signature": _get_pdf_backend_signature(),
        "created_at": now,
        "updated_at": now,
        "entries": {},
        "paths": {},
    }


def _ocr_cache_path_key(cache_path: Path) -> str:
    return os.path.abspath(os.fspath(cache_path))


def _ocr_cache_file_signature(stat_result: os.stat_result) -> tuple[int, int, int, int, int]:
    return (
        int(getattr(stat_result, "st_dev", 0) or 0),
        int(getattr(stat_result, "st_ino", 0) or 0),
        int(stat_result.st_size),
        int(getattr(stat_result, "st_mtime_ns", int(stat_result.st_mtime * 1_000_000_000))),
        int(getattr(stat_result, "st_ctime_ns", int(stat_result.st_ctime * 1_000_000_000))),
    )


def _clear_ocr_memory_cache(cache_path: Path | None = None) -> None:
    global _OCR_MEMORY_CACHE_PATH, _OCR_MEMORY_CACHE_SIGNATURE, _OCR_MEMORY_CACHE_DATA
    path_key = _ocr_cache_path_key(cache_path) if cache_path is not None else None
    with _OCR_MEMORY_CACHE_LOCK:
        if path_key is not None and _OCR_MEMORY_CACHE_PATH != path_key:
            return
        _OCR_MEMORY_CACHE_PATH = None
        _OCR_MEMORY_CACHE_SIGNATURE = None
        _OCR_MEMORY_CACHE_DATA = None


def _remember_ocr_memory_cache(cache_path: Path, data: dict[str, Any]) -> None:
    global _OCR_MEMORY_CACHE_PATH, _OCR_MEMORY_CACHE_SIGNATURE, _OCR_MEMORY_CACHE_DATA
    try:
        stat_result = cache_path.stat()
    except OSError:
        _clear_ocr_memory_cache(cache_path)
        return
    if stat_result.st_size < _OCR_MEMORY_CACHE_MIN_BYTES:
        _clear_ocr_memory_cache(cache_path)
        return
    with _OCR_MEMORY_CACHE_LOCK:
        _OCR_MEMORY_CACHE_PATH = _ocr_cache_path_key(cache_path)
        _OCR_MEMORY_CACHE_SIGNATURE = _ocr_cache_file_signature(stat_result)
        _OCR_MEMORY_CACHE_DATA = data


def _load_ocr_cache(cache_path: Path) -> dict[str, Any]:
    """读取缓存 JSON；损坏时返回空结构并由上层决定是否重建。"""
    path_key = _ocr_cache_path_key(cache_path)
    with _OCR_MEMORY_CACHE_LOCK:
        if _OCR_MEMORY_CACHE_PATH is not None and _OCR_MEMORY_CACHE_PATH != path_key:
            _clear_ocr_memory_cache()
    if cache_path.is_symlink() or not cache_path.is_file():
        _clear_ocr_memory_cache(cache_path)
        return _new_ocr_cache()

    try:
        stat_result = cache_path.stat()
        if stat_result.st_size > _OCR_CACHE_FILE_LOAD_MAX_BYTES:
            _clear_ocr_memory_cache(cache_path)
            return _new_ocr_cache()
        signature = _ocr_cache_file_signature(stat_result)
        with _OCR_MEMORY_CACHE_LOCK:
            if _OCR_MEMORY_CACHE_PATH is not None and _OCR_MEMORY_CACHE_PATH != path_key:
                _clear_ocr_memory_cache()
            if stat_result.st_size >= _OCR_MEMORY_CACHE_MIN_BYTES:
                if (
                    _OCR_MEMORY_CACHE_PATH == path_key
                    and _OCR_MEMORY_CACHE_SIGNATURE == signature
                    and _OCR_MEMORY_CACHE_DATA is not None
                ):
                    return _OCR_MEMORY_CACHE_DATA
                # Only one parsed large cache is retained.  Drop the previous
                # one before decoding another so their peaks do not overlap.
                _clear_ocr_memory_cache()
        # Avoid retaining a second full-size text copy after decoding a large
        # cache.  The parsed dictionary remains exactly the same.
        with cache_path.open("r", encoding="utf-8") as cache_file:
            data = json.load(cache_file)
    except (OSError, UnicodeError, json.JSONDecodeError):
        _clear_ocr_memory_cache(cache_path)
        return _new_ocr_cache()

    if not isinstance(data, dict):
        _clear_ocr_memory_cache(cache_path)
        return _new_ocr_cache()

    # 缺失版本号的旧缓存必须按最早版本迁移，不能误当成当前版本继续复用。
    data.setdefault("version", 0)
    if not isinstance(data.get("entries"), dict):
        data["entries"] = {}
    if not isinstance(data.get("paths"), dict):
        data["paths"] = {}
    _remember_ocr_memory_cache(cache_path, data)
    return data


def _drop_pdf_cache_entries(data: dict[str, Any]) -> int:
    """只删除 PDF 内容索引及其路径引用，保留图片和其他文档缓存。"""
    entries: dict[str, Any] = data.get("entries") or {}
    paths: dict[str, Any] = data.get("paths") or {}
    removed_keys: set[str] = set()
    for key, entry in entries.items():
        if not isinstance(entry, dict):
            continue
        sample_name = str(entry.get("sample_filename") or "")
        source_relpath = str(entry.get("source_relpath") or "")
        if sample_name.lower().endswith(".pdf") or source_relpath.lower().endswith(".pdf"):
            removed_keys.add(str(key))

    # 早期平铺缓存可能只在 paths 中保留扩展名；先反查其内容键再统一清理。
    for rel_path, path_entry in paths.items():
        if not str(rel_path).lower().endswith(".pdf"):
            continue
        if isinstance(path_entry, dict) and path_entry.get("cache_key"):
            removed_keys.add(str(path_entry["cache_key"]))

    for cache_key in removed_keys:
        entries.pop(cache_key, None)
    for rel_path, path_entry in list(paths.items()):
        cache_key = path_entry.get("cache_key") if isinstance(path_entry, dict) else None
        if str(rel_path).lower().endswith(".pdf") or str(cache_key) in removed_keys:
            paths.pop(rel_path, None)
    return len(removed_keys)


def _invalidate_legacy_pdf_cache_entries(data: dict[str, Any]) -> int:
    """PDF 解析升级时只失效 PDF 条目，保留现有图片和文档 OCR 缓存。"""
    try:
        cache_version = int(data.get("version") or 0)
    except (TypeError, ValueError):
        cache_version = 0
    if cache_version >= _OCR_CACHE_VERSION:
        return 0

    removed_count = _drop_pdf_cache_entries(data)
    data["version"] = _OCR_CACHE_VERSION
    return removed_count


def _invalidate_changed_pdf_backend_entries(data: dict[str, Any]) -> int:
    current_signature = _get_pdf_backend_signature()
    previous_signature = str(data.get("pdf_backend_signature") or "")
    if previous_signature == current_signature:
        return 0
    removed_count = _drop_pdf_cache_entries(data)
    data["pdf_backend_signature"] = current_signature
    return removed_count


def _save_ocr_cache(cache_path: Path, data: dict[str, Any]) -> bool:
    """原子写缓存：tmp + os.replace；返回是否成功。

    任何 OSError / PermissionError 都被捕获并返回 False，由调用方降级。
    """
    data["version"] = _OCR_CACHE_VERSION
    data["engine_signature"] = _get_engine_signature()
    data["pdf_backend_signature"] = _get_pdf_backend_signature()
    data["updated_at"] = _beijing_now_str()
    data.setdefault("created_at", data["updated_at"])
    data.setdefault("entries", {})
    data.setdefault("paths", {})

    if cache_path.is_symlink():
        _clear_ocr_memory_cache(cache_path)
        return False
    tmp_path = cache_path.with_name(
        f".{cache_path.name}.{os.getpid()}.{time.time_ns()}.tmp"
    )
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        if cache_path.parent.is_symlink():
            _clear_ocr_memory_cache(cache_path)
            return False
        with tmp_path.open("x", encoding="utf-8") as temp_file:
            # This is an internal machine cache rather than a user document.
            # Compact separators reduce repeated checkpoint I/O without
            # changing keys, values, migration, or atomic replacement.
            json.dump(data, temp_file, ensure_ascii=False, separators=(",", ":"))
            temp_file.write("\n")
        try:
            tmp_path.chmod(0o600)
        except OSError:
            pass
    except (FileExistsError, OSError):
        _clear_ocr_memory_cache(cache_path)
        return False

    try:
        os.replace(tmp_path, cache_path)
    except OSError:
        # 兜底：清理残留 tmp
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass
        _clear_ocr_memory_cache(cache_path)
        return False
    _remember_ocr_memory_cache(cache_path, data)
    return True


def _trim_cache_by_age_and_size(data: dict[str, Any]) -> None:
    """对缓存做 LRU + 体积治理：清理过期与超量条目。

    注：当前 cache key 基于 (content_hash + size)，不依赖文件路径，
因此文件改名/移动不会导致缓存失效；文件被删除后条目会自然过期（90 天未命中后清理）。
    """
    entries: dict[str, Any] = data.get("entries") or {}
    paths: dict[str, Any] = data.get("paths") or {}
    if not entries:
        paths.clear()
        return

    # 1. 清理超过 90 天未验证的条目
    now = datetime.now(tz=_BEIJING_TZ)
    cutoff = now - timedelta(days=_OCR_CACHE_ENTRY_MAX_AGE_DAYS)

    def _parse_ts(ts: str) -> datetime:
        try:
            return datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").replace(tzinfo=_BEIJING_TZ)
        except (TypeError, ValueError):
            try:
                d = datetime.fromisoformat(ts)
                return d if d.tzinfo is not None else d.replace(tzinfo=_BEIJING_TZ)
            except (TypeError, ValueError):
                return now

    stale_keys: list[str] = []
    for key, entry in entries.items():
        ts_str = entry.get("verified_at") if isinstance(entry, dict) else None
        if not ts_str:
            stale_keys.append(key)
            continue
        ts = _parse_ts(ts_str)
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=_BEIJING_TZ)
        if ts < cutoff:
            stale_keys.append(key)

    for key in stale_keys:
        entries.pop(key, None)

    def _drop_orphan_paths() -> None:
        for rel_path, path_entry in list(paths.items()):
            if not isinstance(path_entry, dict) or path_entry.get("cache_key") not in entries:
                paths.pop(rel_path, None)

    _drop_orphan_paths()

    # 2. 估算大小并按 verified_at 升序删除
    serialized_len = sum(
        len(json.dumps(v, ensure_ascii=False).encode("utf-8")) + len(k.encode("utf-8"))
        for k, v in entries.items()
    )
    if serialized_len <= _OCR_CACHE_FILE_MAX_BYTES:
        return

    sorted_keys = sorted(
        entries.keys(),
        key=lambda k: entries[k].get("verified_at", "") if isinstance(entries[k], dict) else "",
    )
    for key in sorted_keys:
        if serialized_len <= _OCR_CACHE_FILE_TRIM_BYTES:
            break
        try:
            serialized_len -= (
                len(json.dumps(entries[key], ensure_ascii=False).encode("utf-8"))
                + len(key.encode("utf-8"))
            )
        except (TypeError, ValueError):
            serialized_len -= 1024
        entries.pop(key, None)
    _drop_orphan_paths()


TOOL_NAME = "需求9-员工资料自动打包与信息提取"

MODE_BY_EMPLOYEE = "by_employee"
MODE_BY_MATERIAL = "by_material"
MODE_FLAT = "flat"
MODES = {MODE_BY_EMPLOYEE, MODE_BY_MATERIAL, MODE_FLAT}

MODE_LABELS = {
    "按员工归类（每人一个文件夹）": MODE_BY_EMPLOYEE,
    "按材料归类（每类材料一个文件夹）": MODE_BY_MATERIAL,
    "平铺输出（所有文件在同一文件夹）": MODE_FLAT,
}
MODE_LABELS_REVERSE = {v: k for k, v in MODE_LABELS.items()}

# 资料库的组织形式与上面的“输出归类模式”是两个独立维度。默认值始终走旧逻辑，
# 避免 TASK-8 影响当前已经投入使用的三种输出结构。
LIBRARY_MODE_PERSON_FOLDER = "person_folder"
LIBRARY_MODE_FLAT_OCR = "flat_ocr"
LIBRARY_MODES = {LIBRARY_MODE_PERSON_FOLDER, LIBRARY_MODE_FLAT_OCR}
LIBRARY_MODE_LABELS = {
    "按人员文件夹查找（原模式）": LIBRARY_MODE_PERSON_FOLDER,
    "无序平铺资料库（OCR 索引）": LIBRARY_MODE_FLAT_OCR,
}
LIBRARY_MODE_LABELS_REVERSE = {v: k for k, v in LIBRARY_MODE_LABELS.items()}

# 预置常见材料类型及其多维度别名同义词库（互斥排他分类，绝不串门）
MATERIAL_SYNONYMS: dict[str, list[str]] = {
    "身份证": [
        "身份证", "身分证", "sfz", "idcard", "id_card", "id", "identity",
        "正面", "反面", "人像面", "国徽面", "人像", "国徽", "A面", "B面", "正反面", "正反",
        "zhengmian", "fanmian", "zm", "fm",
        "身份证正面", "身份证反面", "身份证正反面", "身份证复印件", "身份证照片", "身份证件", "证件",
    ],
    "劳动合同": [
        "劳动合同", "劳动协议", "劳务合同", "劳务协议", "用工合同", "用工协议", "聘用合同", "聘用协议",
        "续签合同", "续签协议", "合同", "协议", "contract", "hetong", "ht", "劳动关系", "劳动手册", "协议书", "聘书", "用工",
    ],
    "学历证明": [
        "学历证", "毕业证", "学位证", "学历证书", "毕业证书", "学位证书", "学信网", "备案表", "教育部",
        "学历证明", "学历认证", "学历", "毕业", "学位", "文凭", "xueli", "biye", "xuexin",
        "学籍", "大专", "本科", "硕士", "博士", "中专", "高中",
    ],
    "资格证书": [
        "资格证", "职业资格证", "职称证", "技能证", "驾驶证", "驾照", "上岗证", "从业资格", "资格", "职称", "技能", "证书", "zige", "jineng",
        "certificate", "license",
    ],
    "安全员证": [
        "安全员证", "安全员", "安全员证书", "安全考核合格证", "建安C证", "建安A证", "建安B证",
        "安管人员", "安全生产考核", "C证", "A证", "B证", "安全考核", "安全员合格证", "anquanyuan",
    ],
    "特种证书": [
        "特种证书", "特种作业证", "特种作业操作证", "特种作业证书", "特种作业", "特种设备",
        "特种操作证", "特种工", "高处作业", "电工作业", "焊接作业", "电工证", "焊工证", "登高证",
        "操作证", "tezhong",
    ],
    "证件照片": [
        "一寸照", "二寸照", "证件照", "寸照", "登记照", "蓝底", "白底", "红底", "个人照片", "照片",
        "相片", "头像", "个人照", "photo", "pic", "avatar", "head",
    ],
    "银行卡": [
        "银行卡", "工资卡", "卡号", "开户行", "存折", "银行账号", "bank", "card", "yinhang",
    ],
}

# 只允许业务含义明确、不会扩张匹配范围的正式名称映射。
# MATERIAL_SYNONYMS 中含“证书/合同/协议/照片/id”等检索关键词，不能直接
# 当作请求类型等价表，否则用户新增同名自定义材料时会发生跨类型误收集。
_SAFE_REQUEST_MATERIAL_ALIASES: dict[str, frozenset[str]] = {
    "特种证书": frozenset({
        "中华人民共和国特种作业操作证",
        "特种作业证",
        "特种作业证书",
        "特种作业操作证",
        "特种作业人员操作证",
        "特种操作证",
        "特种设备作业人员证",
    }),
}

_NOISE_WORDS: set[str] = {
    "序号", "姓名", "员工", "人员", "名字", "部门", "项目", "项目部", "所属部门", "归属部门",
    "身份证", "身份证号码", "身份证号", "证件号码", "工号", "员工编号", "职务", "岗位", "状态",
    "备注", "合计", "总计", "花名册", "统计表", "名单", "汇总", "公司", "集团", "制表",
    "日期", "未命名", "需要材料", "材料", "需求", "所需材料", "资料类型", "全部", "正面", "反面",
}

_ALL_MATERIAL_KEYWORDS: set[str] = set()
for _syns in MATERIAL_SYNONYMS.values():
    for _s in _syns:
        _ALL_MATERIAL_KEYWORDS.add(_s)
for _k in MATERIAL_SYNONYMS:
    _ALL_MATERIAL_KEYWORDS.add(_k)

SUPPORTED_FILE_EXTENSIONS = {
    ".pdf", ".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp",
    ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".txt",
}

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp"}
DOC_EXTENSIONS = {".doc", ".docx", ".pdf", ".txt"}

# 无序资料库中，仅对已确认存在跨图片需求的劳动合同启用保守分组。
# 采用有界连续窗口，避免在老电脑上进行 O(n^2) 全量文件比较。
_DOCUMENT_GROUPABLE_MATERIALS = frozenset({"劳动合同"})
_DOCUMENT_GROUP_MAX_PAGES = 32
_DOCUMENT_GROUP_MAX_TIME_GAP_SECONDS = 5 * 60
_DOCUMENT_GROUP_TEXT_MAX_CHARS = 20_000
_DOCUMENT_GROUP_MAX_CANDIDATES = 64

_IGNORED_FILENAMES = {
    ".ds_store", "thumbs.db", "desktop.ini", ".localized", "ehthumbs.db",
}

_PHONE_RE = re.compile(r"^1[3-9]\d{9}$")
_ID_CARD_RE = re.compile(r"^\d{17}[\dXx]$|^\d{15}$")


def _is_junk_or_temp_file(path: Path | str) -> bool:
    """检查是否为系统垃圾文件或 Office 临时锁文件。"""
    name = Path(path).name.lower()
    if name.startswith(".") or name.startswith("~$") or name in _IGNORED_FILENAMES:
        return True
    return False


def _is_path_nested(child: Path, parent: Path) -> bool:
    """检查 child 路径是否处于 parent 路径内部或为同一路径。"""
    try:
        child_res = child.resolve()
        parent_res = parent.resolve()
        if child_res == parent_res:
            return True
        child_res.relative_to(parent_res)
        return True
    except (ValueError, RuntimeError):
        return False


def _get_file_signature(path: Path) -> tuple[int, str]:
    """计算文件大小及首尾采样哈希，作为同员工去重特征。"""
    try:
        size = path.stat().st_size
        with open(path, "rb") as f:
            digest = hashlib.sha256()
            digest.update(f.read(65536))
            if size > 65536:
                f.seek(max(size - 65536, 0))
                digest.update(f.read(65536))
        return size, digest.hexdigest()
    except Exception:
        return -1, str(path)


@dataclass(frozen=True)
class TargetEmployee:
    name: str
    id_card: str = ""
    employee_no: str = ""
    department: str = ""
    phone: str = ""
    per_person_materials: tuple[str, ...] = ()
    extra: dict[str, Any] = field(default_factory=dict)

    @property
    def identity_key(self) -> str:
        """员工唯一复合主键，防止同名员工覆盖。"""
        suffix = f"_{self.employee_no}" if self.employee_no else ""
        return f"{self.name}_{self.id_card}_{self.phone}{suffix}"

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "id_card": self.id_card,
            "employee_no": self.employee_no,
            "department": self.department,
            "phone": self.phone,
            "per_person_materials": list(self.per_person_materials),
        }


@dataclass(frozen=True)
class MaterialFileMatch:
    employee_name: str
    material_type: str
    source_path: Path
    relative_source_path: str
    matched_by: str  # "filename", "ocr", "doc_content", "id_card", "phone", "read_failed"
    target_filename: str = ""
    target_path: Path | None = None
    extracted_person_name: str = ""
    extracted_id_card: str = ""
    mismatch_warning: str = ""
    cache_hit: bool = False  # 新增：OCR 缓存命中标记
    employee_identity_key: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "employee_name": self.employee_name,
            "material_type": self.material_type,
            "source_path": str(self.source_path),
            "relative_source_path": self.relative_source_path,
            "matched_by": self.matched_by,
            "target_filename": self.target_filename,
            "target_path": str(self.target_path) if self.target_path else "",
            "extracted_person_name": self.extracted_person_name,
            "extracted_id_card": self.extracted_id_card,
            "mismatch_warning": self.mismatch_warning,
            "cache_hit": self.cache_hit,
            "employee_identity_key": self.employee_identity_key,
        }


@dataclass
class MaterialCollectResult:
    library_dir: Path
    output_dir: Path
    zip_path: Path | None = None
    report_path: Path | None = None
    mode: str = MODE_BY_EMPLOYEE
    library_mode: str = LIBRARY_MODE_PERSON_FOLDER
    target_employees: list[TargetEmployee] = field(default_factory=list)
    requested_materials: list[str] = field(default_factory=list)
    matches: list[MaterialFileMatch] = field(default_factory=list)
    missing_records: dict[str, list[str]] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    folder_match_counts: dict[str, int] = field(default_factory=dict)
    employee_result_keys: list[str] = field(default_factory=list, repr=False)

    # === 新增：OCR 缓存层指标（默认值保持向后兼容） ===
    ocr_cache_enabled: bool = True
    ocr_cache_hits: int = 0
    ocr_cache_misses: int = 0
    ocr_cache_invalidated: int = 0
    ocr_cache_path: str | None = None
    ocr_cache_skipped_reason: str | None = None

    # === 占位字段：未来隐私开关（本次不实现行为，仅留接口） ===
    # TODO: 当 HR 提报隐私报送需求时启用，本次保持 None / False 以保证报送数据完整
    zip_password: str | None = None
    mask_sensitive: bool = False
    review_path: Path | None = None

    @property
    def total_employees(self) -> int:
        return len(self.target_employees)

    @property
    def matched_file_count(self) -> int:
        return len(self.matches)

    @property
    def complete_employee_count(self) -> int:
        keys = (
            self.employee_result_keys
            if len(self.employee_result_keys) == len(self.target_employees)
            else [employee.name for employee in self.target_employees]
        )
        pending = {
            match.employee_identity_key or match.employee_name
            for match in self.matches if "待确认" in match.mismatch_warning
        }
        return sum(
            1 for key, employee in zip(keys, self.target_employees)
            if not self.missing_records.get(key)
            and employee.identity_key not in pending and employee.name not in pending
        )

    def to_dict(self, *, include_matches: bool = True) -> dict[str, Any]:
        payload = {
            "tool_name": TOOL_NAME,
            "library_dir": str(self.library_dir),
            "output_dir": str(self.output_dir),
            "zip_path": str(self.zip_path) if self.zip_path else None,
            "report_path": str(self.report_path) if self.report_path else None,
            "review_path": str(self.review_path) if self.review_path else None,
            "mode": self.mode,
            "library_mode": self.library_mode,
            "total_employees": self.total_employees,
            "matched_file_count": self.matched_file_count,
            "complete_employee_count": self.complete_employee_count,
            "requested_materials": self.requested_materials,
            "missing_records": self.missing_records,
            "warnings": self.warnings,
            # 缓存层指标
            "ocr_cache_enabled": self.ocr_cache_enabled,
            "ocr_cache_hits": self.ocr_cache_hits,
            "ocr_cache_misses": self.ocr_cache_misses,
            "ocr_cache_invalidated": self.ocr_cache_invalidated,
            "ocr_cache_path": self.ocr_cache_path,
            "ocr_cache_skipped_reason": self.ocr_cache_skipped_reason,
            # 占位字段
            "zip_password": self.zip_password,
            "mask_sensitive": self.mask_sensitive,
        }
        if include_matches:
            payload["matches"] = [match.to_dict() for match in self.matches]
        return payload


# ---------------------------------------------------------------------------
# 材料需求识别：将自由文本映射到标准材料类型
# ---------------------------------------------------------------------------

def _resolve_material_text(text: str) -> list[str]:
    """将自由文本（如"身份证，合同"）解析为标准材料类型列表。"""
    if not text:
        return []
    result: list[str] = []
    parts = re.split(r"[,，、;；\s]+", text.strip())
    for part in parts:
        part = part.strip()
        if not part:
            continue
        matched = False
        for mat_type, synonyms in MATERIAL_SYNONYMS.items():
            if part == mat_type or any(syn == part or syn in part for syn in synonyms):
                if mat_type not in result:
                    result.append(mat_type)
                matched = True
                break
        if not matched and part not in result:
            result.append(part)
    return result


def _is_valid_person_name(name: str) -> bool:
    """校验是否为一个合法的员工姓名（过滤数字、表头、噪音词）。"""
    name = str(name or "").strip()
    if not name:
        return False
    if name in _NOISE_WORDS:
        return False
    if re.match(r"^\d+$", name):
        return False
    if any(noise in name for noise in ("花名册", "统计表", "汇总表", "总人数", "部门：", "公司：", "制表人")):
        return False
    if len(name) > 30:
        return False
    return True


# ---------------------------------------------------------------------------
# 目标员工名单解析（支持直接输入单人/多人文本、或 Excel 文件）
# ---------------------------------------------------------------------------

def _parse_single_text_item(item_str: str) -> TargetEmployee | None:
    """解析单个字符串条目（如 '张三', '张三 440111199001011234', '440111199001011234', '张三 身份证'）。"""
    item_str = item_str.strip()
    if not item_str or item_str.startswith("#"):
        return None

    parts = [p for p in re.split(r"[\s\t]+", item_str) if p]
    if not parts:
        return None

    if _ID_CARD_RE.match(parts[0]):
        id_card = parts[0]
        name = parts[1] if len(parts) > 1 and _is_valid_person_name(parts[1]) else id_card
        mat_text = " ".join(parts[2:]) if len(parts) > 2 else ""
        per_mats = tuple(_resolve_material_text(mat_text))
        return TargetEmployee(name=name, id_card=id_card, per_person_materials=per_mats)

    if _PHONE_RE.match(parts[0]):
        phone = parts[0]
        name = parts[1] if len(parts) > 1 and _is_valid_person_name(parts[1]) else phone
        return TargetEmployee(name=name, phone=phone)

    name = parts[0]
    if not _is_valid_person_name(name):
        return None

    id_card = ""
    phone = ""
    emp_no = ""
    materials_parts: list[str] = []

    for p in parts[1:]:
        if _ID_CARD_RE.match(p) and not id_card:
            id_card = p
        elif _PHONE_RE.match(p) and not phone:
            phone = p
        elif any(kw in p for kw in _ALL_MATERIAL_KEYWORDS):
            materials_parts.append(p)
        elif re.match(r"^[A-Za-z0-9_-]+$", p) and len(p) <= 10 and not emp_no:
            emp_no = p
        else:
            materials_parts.append(p)

    per_mats = tuple(_resolve_material_text(" ".join(materials_parts)))
    return TargetEmployee(
        name=name,
        id_card=id_card,
        phone=phone,
        employee_no=emp_no,
        per_person_materials=per_mats,
    )


def parse_employee_roster(
    source: str | Path | list[dict[str, Any]] | list[str],
) -> list[TargetEmployee]:
    """Parse employee list from an Excel workbook, text content, or structured list."""
    employees: list[TargetEmployee] = []
    seen_keys: set[str] = set()

    def _add_emp(emp: TargetEmployee | None) -> None:
        if not emp or not emp.name:
            return
        key = emp.identity_key
        if key not in seen_keys:
            seen_keys.add(key)
            employees.append(emp)

    if isinstance(source, list):
        for item in source:
            if isinstance(item, dict):
                name = str(item.get("name") or item.get("姓名") or "").strip()
                id_card = str(item.get("id_card") or item.get("身份证号码") or item.get("身份证") or "").strip()
                emp_no = str(item.get("employee_no") or item.get("工号") or "").strip()
                dept = str(item.get("department") or item.get("部门") or item.get("项目") or "").strip()
                phone = str(item.get("phone") or item.get("手机号") or item.get("电话") or "").strip()
                mat_text = str(item.get("materials") or item.get("材料") or item.get("需要材料") or "").strip()
                per_mats = tuple(_resolve_material_text(mat_text))
                if _is_valid_person_name(name) or _ID_CARD_RE.match(id_card):
                    _add_emp(TargetEmployee(
                        name=name or id_card, id_card=id_card, employee_no=emp_no,
                        department=dept, phone=phone, per_person_materials=per_mats,
                    ))
            elif isinstance(item, str):
                _add_emp(_parse_single_text_item(item))
        return employees

    source_path = Path(source) if isinstance(source, (str, Path)) else None
    if source_path and source_path.is_file() and is_supported_excel_file(source_path):
        import tempfile
        with tempfile.TemporaryDirectory() as temp_dir_str:
            working_path = ensure_xlsx_workbook(source_path, Path(temp_dir_str))
            wb = load_workbook(working_path, data_only=True, read_only=True)
            try:
                # Read-only worksheets are cheap to open but random ``cell`` access
                # reparses XML from the beginning.  A single compact value pass keeps
                # the existing parser contract while bounding memory on large rosters.
                ws = SheetGrid(wb.active)
            finally:
                wb.close()

            name_col: int | None = None
            id_card_col: int | None = None
            emp_no_col: int | None = None
            dept_col: int | None = None
            phone_col: int | None = None
            material_col: int | None = None
            header_row_idx = 1

            name_synonyms = ("姓名", "员工姓名", "人员姓名", "名字")
            id_card_synonyms = ("身份证号码", "身份证号", "身份证", "证件号码", "证件号")
            emp_no_synonyms = ("工号", "员工编号", "员工号", "人员编号")
            dept_synonyms = ("部门", "项目", "项目部", "所属部门", "归属部门", "所属项目")
            phone_synonyms = ("手机号", "手机号码", "电话", "联系电话", "联系方式")
            material_col_synonyms = ("需要材料", "材料", "需求", "所需材料", "资料类型")

            max_r = ws.max_row or 1
            max_c = ws.max_column or 1

            for r in range(1, min(max_r, 15) + 1):
                for c in range(1, max_c + 1):
                    val = str(ws.cell(r, c).value or "").strip()
                    if not val:
                        continue
                    if name_col is None and val in name_synonyms:
                        name_col = c
                    if id_card_col is None and val in id_card_synonyms:
                        id_card_col = c
                    if emp_no_col is None and val in emp_no_synonyms:
                        emp_no_col = c
                    if dept_col is None and val in dept_synonyms:
                        dept_col = c
                    if phone_col is None and val in phone_synonyms:
                        phone_col = c
                    if material_col is None and any(syn in val for syn in material_col_synonyms):
                        material_col = c
                if name_col is not None:
                    header_row_idx = r
                    break

            if name_col is None:
                first_val = str(ws.cell(1, 1).value or "").strip()
                if _is_valid_person_name(first_val) and not any(noise in first_val for noise in ("花名册", "表", "单", "人员", "员工")):
                    name_col = 1
                    header_row_idx = 0
                else:
                    name_col = 1
                    header_row_idx = 1

            if material_col is None and max_c >= name_col + 1:
                candidate_col = name_col + 1
                if candidate_col not in (id_card_col, emp_no_col, dept_col, phone_col):
                    hits = 0
                    checked = 0
                    for r in range(header_row_idx + 1, min(max_r, header_row_idx + 20) + 1):
                        cell_val = str(ws.cell(r, candidate_col).value or "").strip()
                        if not cell_val:
                            continue
                        checked += 1
                        if any(kw in cell_val for kw in _ALL_MATERIAL_KEYWORDS):
                            hits += 1
                    if checked > 0 and hits / checked >= 0.3:
                        material_col = candidate_col

            for r in range(header_row_idx + 1, max_r + 1):
                name_val = str(ws.cell(r, name_col).value or "").strip()
                if not name_val or not _is_valid_person_name(name_val):
                    continue
                id_card_val = str(ws.cell(r, id_card_col).value or "").strip() if id_card_col else ""
                emp_no_val = str(ws.cell(r, emp_no_col).value or "").strip() if emp_no_col else ""
                dept_val = str(ws.cell(r, dept_col).value or "").strip() if dept_col else ""
                phone_val = str(ws.cell(r, phone_col).value or "").strip() if phone_col else ""

                mat_text = ""
                if material_col:
                    mat_text = str(ws.cell(r, material_col).value or "").strip()
                per_mats = tuple(_resolve_material_text(mat_text))

                _add_emp(TargetEmployee(
                    name=name_val,
                    id_card=id_card_val,
                    phone=phone_val,
                    employee_no=emp_no_val,
                    department=dept_val,
                    per_person_materials=per_mats,
                ))
        return employees

    raw_text = str(source)
    raw_items = re.split(r"[\n\r;；,，]+", raw_text)
    for item in raw_items:
        _add_emp(_parse_single_text_item(item))

    return employees


def _is_direct_employee_input(
    source: str | Path | list[dict[str, Any]] | list[str],
) -> bool:
    """Return whether the roster came from the GUI's direct employee input."""
    if not isinstance(source, str):
        return False
    try:
        source_path = Path(source).expanduser()
        return not (source_path.is_file() and is_supported_excel_file(source_path))
    except (OSError, ValueError):
        # Free-form employee text can contain characters or lengths that are not
        # valid filesystem paths.  It is still a direct roster input.
        return True


# ---------------------------------------------------------------------------
# 文件夹匹配核心算法
# ---------------------------------------------------------------------------


class _AhoCandidateMatcher:
    """Compact multi-pattern matcher used only to find possible employees.

    The established business predicates remain the final authority.  This
    matcher merely avoids testing every employee against every folder or OCR
    record, which otherwise becomes quadratic for large batches.
    """

    __slots__ = ("_transitions", "_failures", "_outputs", "_patterns", "_payloads")

    def __init__(self, pattern_payloads: dict[str, set[int]]) -> None:
        self._transitions: list[dict[str, int]] = [{}]
        self._failures: list[int] = [0]
        self._outputs: list[list[int]] = [[]]
        self._patterns: list[str] = []
        self._payloads: list[tuple[int, ...]] = []

        for pattern, payloads in pattern_payloads.items():
            if not pattern or not payloads:
                continue
            pattern_index = len(self._patterns)
            self._patterns.append(pattern)
            self._payloads.append(tuple(sorted(payloads)))
            state = 0
            for character in pattern:
                next_state = self._transitions[state].get(character)
                if next_state is None:
                    next_state = len(self._transitions)
                    self._transitions[state][character] = next_state
                    self._transitions.append({})
                    self._failures.append(0)
                    self._outputs.append([])
                state = next_state
            self._outputs[state].append(pattern_index)

        pending: deque[int] = deque()
        for child in self._transitions[0].values():
            pending.append(child)
        while pending:
            state = pending.popleft()
            for character, child in self._transitions[state].items():
                pending.append(child)
                fallback = self._failures[state]
                while fallback and character not in self._transitions[fallback]:
                    fallback = self._failures[fallback]
                self._failures[child] = self._transitions[fallback].get(character, 0)
                inherited = self._outputs[self._failures[child]]
                if inherited:
                    self._outputs[child].extend(inherited)

    @staticmethod
    def _is_cjk(character: str) -> bool:
        return "\u3400" <= character <= "\u9fff"

    def match_payloads(self, text: str, *, cjk_boundaries: bool = False) -> set[int]:
        if not text or len(self._transitions) == 1:
            return set()
        matched: set[int] = set()
        state = 0
        for position, character in enumerate(text):
            while state and character not in self._transitions[state]:
                state = self._failures[state]
            state = self._transitions[state].get(character, 0)
            for pattern_index in self._outputs[state]:
                if cjk_boundaries:
                    pattern = self._patterns[pattern_index]
                    start = position - len(pattern) + 1
                    if start > 0 and self._is_cjk(text[start - 1]):
                        continue
                    next_position = position + 1
                    if next_position < len(text) and self._is_cjk(text[next_position]):
                        continue
                matched.update(self._payloads[pattern_index])
        return matched


class _FolderEmployeeCandidateIndex:
    """Map each employee to candidate folder positions in original scan order."""

    __slots__ = ("_employees", "_folder_entries", "_candidate_positions")

    def __init__(
        self,
        folder_index: dict[str, list[Path]],
        employees: list[TargetEmployee],
        *,
        cancelled: Callable[[], bool] | None = None,
    ) -> None:
        pattern_payloads: dict[str, set[int]] = {}
        for employee_index, employee in enumerate(employees):
            patterns: set[str] = set()
            employee_id = employee.id_card.strip()
            if employee_id and len(employee_id) >= 15:
                patterns.add(employee_id)
            employee_phone = employee.phone.strip()
            if employee_phone and len(employee_phone) == 11:
                patterns.add(employee_phone)
            employee_name = employee.name.strip()
            if employee_name and (
                _is_valid_person_name(employee_name)
                or _ID_CARD_RE.match(employee_name)
            ):
                patterns.add(employee_name)
            for pattern in patterns:
                pattern_payloads.setdefault(pattern, set()).add(employee_index)

        matcher = _AhoCandidateMatcher(pattern_payloads)
        self._employees = employees
        self._folder_entries = list(folder_index.items())
        # ``array('I')`` keeps a 10k x 10k sparse match set compact on low-memory PCs.
        self._candidate_positions = [array("I") for _ in employees]
        for folder_position, (folder_name, _paths) in enumerate(self._folder_entries):
            if folder_position % 512 == 0:
                _raise_if_cancelled(cancelled)
            for employee_index in matcher.match_payloads(folder_name.strip()):
                self._candidate_positions[employee_index].append(folder_position)

    def matches_for(self, employee_index: int) -> list[tuple[Path, str]]:
        employee = self._employees[employee_index]
        matched_folders: list[tuple[Path, str]] = []
        for folder_position in self._candidate_positions[employee_index]:
            folder_name, paths = self._folder_entries[folder_position]
            reason = _match_folder_to_employee(folder_name, employee)
            if reason:
                matched_folders.extend((path, reason) for path in paths)
        return matched_folders

def _match_folder_to_employee(folder_name: str, emp: TargetEmployee) -> str | None:
    """判断一个文件夹名是否属于某员工，返回匹配依据或 None。"""
    f_name = folder_name.strip()
    if not f_name:
        return None

    emp_id = emp.id_card.strip()
    if emp_id and len(emp_id) >= 15 and emp_id in f_name:
        return "id_card"

    emp_phone = emp.phone.strip()
    if emp_phone and len(emp_phone) == 11 and emp_phone in f_name:
        return "phone"

    emp_name = emp.name.strip()
    if not emp_name or not _is_valid_person_name(emp_name):
        return None

    if _ID_CARD_RE.match(emp_name) and emp_name in f_name:
        return "id_card"

    if f_name == emp_name:
        return "exact_name"

    pattern = r"(?:^|[\d_\s\-\(\)（）\[\]【】#])" + re.escape(emp_name) + r"(?:[\d_\s\-\(\)（）\[\]【】#]|$)"
    if re.search(pattern, f_name):
        return "name"

    if re.fullmatch(re.escape(emp_name) + r"(?:的)?(?:资料|材料|档案)(?:夹)?", f_name):
        return "name"

    return None


# ---------------------------------------------------------------------------
# 文档正文提取 & 本地离线 OCR 识图引擎
# ---------------------------------------------------------------------------

def _read_office_xml_member(
    archive: zipfile.ZipFile,
    member_name: str,
    *,
    remaining_bytes: int = _OFFICE_XML_TOTAL_MAX_BYTES,
) -> bytes:
    if len(archive.infolist()) > _OFFICE_ARCHIVE_MAX_MEMBERS:
        raise ValueError("Office 文档包含过多压缩条目")
    member = archive.getinfo(member_name)
    limit = min(_OFFICE_XML_MEMBER_MAX_BYTES, remaining_bytes)
    if member.file_size > limit:
        raise ValueError("Office XML 条目体积异常")
    if member.file_size and (
        member.compress_size <= 0
        or member.file_size / member.compress_size > _OFFICE_XML_MAX_COMPRESSION_RATIO
    ):
        raise ValueError("Office XML 条目压缩比异常")
    with archive.open(member) as source:
        payload = source.read(limit + 1)
    if len(payload) > limit or len(payload) != member.file_size:
        raise ValueError("Office XML 条目实际体积异常")
    if re.search(br"<!\s*(?:DOCTYPE|ENTITY)", payload, flags=re.IGNORECASE):
        raise ValueError("Office XML 包含不支持的实体声明")
    return payload


def _raise_if_cancelled(cancelled: Callable[[], bool] | None) -> None:
    if cancelled is not None and cancelled():
        raise MaterialCollectionCancelled("本次处理已停止。")


def _validate_pdf_source_size(file_path: Path) -> int:
    try:
        file_size = file_path.stat().st_size
    except OSError as exc:
        raise PDFRecognitionError(f"PDF 文件无法读取：{file_path.name}") from exc
    if file_size <= 0:
        raise PDFRecognitionError(f"PDF 文件为空或已损坏：{file_path.name}")
    if file_size > _PDF_MAX_FILE_BYTES:
        raise PDFResourceLimitError(
            f"PDF 文件体积超过安全上限：{file_path.name}，"
            f"允许不超过 {_PDF_MAX_FILE_BYTES // (1024 * 1024)} MB"
        )
    return file_size


@contextmanager
def _open_pdf_reader(file_path: Path) -> Iterator[tuple[Any, int]]:
    """以流式文件句柄打开 PDF，并在解析前执行体积、加密和页数门禁。"""
    _validate_pdf_source_size(file_path)
    if PdfReader is None:
        raise PDFRecognitionError("当前安装包缺少 PDF 文字解析组件，请重新安装完整版本")

    try:
        stream = file_path.open("rb")
    except OSError as exc:
        raise PDFRecognitionError(f"PDF 文件无法读取：{file_path.name}") from exc
    try:
        try:
            # 禁止宽松恢复恶意/畸形交叉引用；真实损坏文件应明确提示用户修复。
            reader = PdfReader(stream, strict=True)
        except Exception as exc:
            raise PDFRecognitionError(
                f"PDF 文件损坏或格式异常：{file_path.name}"
            ) from exc
        try:
            if reader.is_encrypted:
                raise PDFRecognitionError(
                    f"PDF 文件已加密，无法自动识别：{file_path.name}；"
                    "请先在可信的 PDF 工具中解除密码后重试"
                )
            page_count = len(reader.pages)
        except PDFRecognitionError:
            raise
        except Exception as exc:
            raise PDFRecognitionError(
                f"PDF 文件损坏，无法读取页面结构：{file_path.name}"
            ) from exc
        if page_count <= 0:
            raise PDFRecognitionError(f"PDF 文件没有可读取页面：{file_path.name}")
        if page_count > _PDF_MAX_PAGES:
            raise PDFResourceLimitError(
                f"PDF 页数超过安全上限：{file_path.name}，"
                f"共 {page_count} 页，允许不超过 {_PDF_MAX_PAGES} 页"
            )
        yield reader, page_count
    finally:
        stream.close()


@contextmanager
def _open_pdfium_document(file_path: Path) -> Iterator[tuple[Any, int]]:
    """在 Win7 兼容包中使用受维护的 PDFium 安全加载并及时释放句柄。"""
    _validate_pdf_source_size(file_path)
    if pdfium is None:
        raise PDFRecognitionError("当前安装包缺少 PDF 页面解析组件，请重新安装完整版本")

    _PDF_LOCK.acquire()
    document = None
    try:
        try:
            document = pdfium.PdfDocument(file_path)
        except Exception as exc:
            password_code = getattr(getattr(pdfium, "raw", None), "FPDF_ERR_PASSWORD", 4)
            if getattr(exc, "err_code", None) == password_code:
                raise PDFRecognitionError(
                    f"PDF 文件已加密，无法自动识别：{file_path.name}；"
                    "请先在可信的 PDF 工具中解除密码后重试"
                ) from exc
            raise PDFRecognitionError(
                f"PDF 文件损坏或格式异常：{file_path.name}"
            ) from exc
        try:
            page_count = len(document)
        except Exception as exc:
            raise PDFRecognitionError(
                f"PDF 文件损坏，无法读取页面结构：{file_path.name}"
            ) from exc
        if page_count <= 0:
            raise PDFRecognitionError(f"PDF 文件没有可读取页面：{file_path.name}")
        if page_count > _PDF_MAX_PAGES:
            raise PDFResourceLimitError(
                f"PDF 页数超过安全上限：{file_path.name}，"
                f"共 {page_count} 页，允许不超过 {_PDF_MAX_PAGES} 页"
            )
        yield document, page_count
    finally:
        if document is not None:
            try:
                document.close()
            except Exception:
                pass
        _PDF_LOCK.release()


@contextmanager
def _open_pdf_document(file_path: Path) -> Iterator[tuple[Any, int]]:
    if _PDF_BACKEND == "pdfium":
        with _open_pdfium_document(file_path) as opened:
            yield opened
        return
    with _open_pdf_reader(file_path) as opened:
        yield opened


def _extract_pdf_text(
    file_path: Path,
    *,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
) -> str:
    """逐页提取完整 PDF 文字层，避免固定字节截断和二进制乱码。"""
    if _PDF_BACKEND == "pdfium":
        return _extract_pdfium_text(
            file_path,
            progress_callback=progress_callback,
            cancelled=cancelled,
        )

    chunks: list[str] = []
    character_count = 0
    with _open_pdf_reader(file_path) as (reader, page_count):
        for page_number, page in enumerate(reader.pages, start=1):
            _raise_if_cancelled(cancelled)
            if progress_callback is not None:
                progress_callback(
                    page_number,
                    page_count,
                    f"正在读取 PDF 文字层：{file_path.name}（{page_number}/{page_count}）",
                )
            try:
                page_text = page.extract_text() or ""
            except Exception as exc:
                raise PDFRecognitionError(
                    f"PDF 第 {page_number} 页文字层损坏：{file_path.name}"
                ) from exc
            if page_text:
                character_count += len(page_text)
                if character_count > _PDF_MAX_TEXT_CHARS:
                    raise PDFResourceLimitError(
                        f"PDF 文字量超过安全上限：{file_path.name}，"
                        f"允许不超过 {_PDF_MAX_TEXT_CHARS} 个字符"
                    )
                chunks.append(page_text)
        _raise_if_cancelled(cancelled)
    return "\n".join(chunks)


def _extract_pdfium_text(
    file_path: Path,
    *,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
) -> str:
    """使用 PDFium 逐页提取 Win7 兼容包中的完整文字层。"""
    chunks: list[str] = []
    character_count = 0
    with _open_pdfium_document(file_path) as (document, page_count):
        for page_index in range(page_count):
            page_number = page_index + 1
            _raise_if_cancelled(cancelled)
            if progress_callback is not None:
                progress_callback(
                    page_number,
                    page_count,
                    f"正在读取 PDF 文字层：{file_path.name}（{page_number}/{page_count}）",
                )
            page = None
            text_page = None
            try:
                page = document[page_index]
                text_page = page.get_textpage()
                # pypdfium2 4.27 is the final Win7-compatible binary. Its
                # upstream range helper has a known allocation regression;
                # bounded extraction is the documented full-page safe path.
                page_text = text_page.get_text_bounded() or ""
            except Exception as exc:
                raise PDFRecognitionError(
                    f"PDF 第 {page_number} 页文字层损坏：{file_path.name}"
                ) from exc
            finally:
                if text_page is not None:
                    try:
                        text_page.close()
                    except Exception:
                        pass
                if page is not None:
                    try:
                        page.close()
                    except Exception:
                        pass
            if page_text:
                character_count += len(page_text)
                if character_count > _PDF_MAX_TEXT_CHARS:
                    raise PDFResourceLimitError(
                        f"PDF 文字量超过安全上限：{file_path.name}，"
                        f"允许不超过 {_PDF_MAX_TEXT_CHARS} 个字符"
                    )
                chunks.append(page_text)
        _raise_if_cancelled(cancelled)
    return "\n".join(chunks)


def _iter_pdf_xobject_images(
    owner: Any,
    *,
    ancestors: tuple[str, ...] = (),
    seen: set[tuple[int, int] | int] | None = None,
    counter: list[int] | None = None,
    depth: int = 0,
) -> Iterator[tuple[tuple[str, ...], Any]]:
    """有界遍历页面与 Form XObject 中的图片，跳过循环引用。"""
    if depth > 8:
        raise PDFResourceLimitError("PDF 页面对象嵌套层级超过安全上限")
    if seen is None:
        seen = set()
    if counter is None:
        counter = [0]
    try:
        resources = owner.get("/Resources")
        if resources is None:
            return
        resources = resources.get_object()
        xobjects = resources.get("/XObject")
        if xobjects is None:
            return
        xobjects = xobjects.get_object()
    except Exception as exc:
        raise PDFRecognitionError("PDF 页面资源结构损坏") from exc

    for name in xobjects:
        counter[0] += 1
        if counter[0] > _PDF_MAX_XOBJECTS_PER_PAGE:
            raise PDFResourceLimitError("PDF 单页图像对象数量超过安全上限")
        try:
            candidate = xobjects[name].get_object()
        except Exception as exc:
            raise PDFRecognitionError("PDF 页面图像对象损坏") from exc
        reference = getattr(candidate, "indirect_reference", None)
        if reference is not None:
            identity: tuple[int, int] | int = (
                int(reference.idnum),
                int(reference.generation),
            )
        else:
            identity = id(candidate)
        if identity in seen:
            continue
        seen.add(identity)
        subtype = str(candidate.get("/Subtype") or "")
        image_path = (*ancestors, str(name))
        if subtype == "/Image":
            yield image_path, candidate
        elif subtype == "/Form":
            yield from _iter_pdf_xobject_images(
                candidate,
                ancestors=image_path,
                seen=seen,
                counter=counter,
                depth=depth + 1,
            )


def _pdf_image_dimensions(image_object: Any) -> tuple[int, int]:
    try:
        width = int(image_object.get("/Width") or 0)
        height = int(image_object.get("/Height") or 0)
    except (TypeError, ValueError) as exc:
        raise PDFRecognitionError("PDF 页面图像尺寸无效") from exc
    if width <= 0 or height <= 0:
        raise PDFRecognitionError("PDF 页面图像缺少有效尺寸")
    return width, height


def _release_pdf_image_decode_cache(
    image_object: Any,
    *,
    seen: set[int] | None = None,
) -> None:
    """释放 pypdf 为当前图片保留的解码缓存，避免多页扫描件线性占用内存。"""
    if seen is None:
        seen = set()
    identity = id(image_object)
    if identity in seen:
        return
    seen.add(identity)

    if hasattr(image_object, "decoded_self"):
        try:
            image_object.decoded_self = None
        except Exception:
            pass
    try:
        soft_mask = image_object.get("/SMask")
        if soft_mask is not None:
            _release_pdf_image_decode_cache(soft_mask.get_object(), seen=seen)
    except Exception:
        pass


def _iter_pdfium_ocr_images(
    file_path: Path,
    *,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
) -> Iterator[bytes]:
    """按页渲染扫描 PDF，覆盖内嵌图、表单及复杂页面组合。"""
    total_pixels = 0
    total_decoded_bytes = 0
    with _open_pdfium_document(file_path) as (document, page_count):
        for page_index in range(page_count):
            page_number = page_index + 1
            _raise_if_cancelled(cancelled)
            if progress_callback is not None:
                progress_callback(
                    page_number,
                    page_count,
                    f"正在识别扫描 PDF：{file_path.name}（{page_number}/{page_count}）",
                )

            page = None
            bitmap = None
            image = None
            try:
                page = document[page_index]
                page_width, page_height = page.get_size()
                if (
                    not math.isfinite(page_width)
                    or not math.isfinite(page_height)
                    or page_width <= 0
                    or page_height <= 0
                ):
                    raise PDFRecognitionError(
                        f"PDF 第 {page_number} 页尺寸无效：{file_path.name}"
                    )
                width = max(1, math.ceil(page_width * _PDF_PAGE_RENDER_SCALE))
                height = max(1, math.ceil(page_height * _PDF_PAGE_RENDER_SCALE))
                pixels = width * height
                decoded_bytes = pixels * 4
                if pixels > _PDF_MAX_IMAGE_PIXELS:
                    raise PDFResourceLimitError(
                        f"PDF 第 {page_number} 页渲染像素超过安全上限："
                        f"{pixels} > {_PDF_MAX_IMAGE_PIXELS}"
                    )
                if decoded_bytes > _PDF_MAX_DECODED_IMAGE_BYTES:
                    raise PDFResourceLimitError(
                        f"PDF 第 {page_number} 页估算解码内存超过安全上限："
                        f"{decoded_bytes} > {_PDF_MAX_DECODED_IMAGE_BYTES} bytes"
                    )
                total_pixels += pixels
                total_decoded_bytes += decoded_bytes
                if total_pixels > _PDF_MAX_TOTAL_IMAGE_PIXELS:
                    raise PDFResourceLimitError(
                        f"PDF 累计页面像素超过安全上限：{file_path.name}"
                    )
                if total_decoded_bytes > _PDF_MAX_TOTAL_DECODED_IMAGE_BYTES:
                    raise PDFResourceLimitError(
                        f"PDF 累计估算解码内存超过安全上限：{file_path.name}"
                    )

                bitmap = page.render(scale=_PDF_PAGE_RENDER_SCALE)
                image = bitmap.to_pil()
                with BytesIO() as output:
                    image.save(output, format="PNG", compress_level=1)
                    payload = output.getvalue()
            except (MaterialCollectionCancelled, PDFRecognitionError):
                raise
            except Exception as exc:
                raise PDFRecognitionError(
                    f"PDF 第 {page_number} 页损坏或无法安全渲染：{file_path.name}"
                ) from exc
            finally:
                if image is not None:
                    try:
                        image.close()
                    except Exception:
                        pass
                if bitmap is not None:
                    try:
                        bitmap.close()
                    except Exception:
                        pass
                if page is not None:
                    try:
                        page.close()
                    except Exception:
                        pass

            _raise_if_cancelled(cancelled)
            if not payload:
                raise PDFRecognitionError(
                    f"PDF 第 {page_number} 页渲染结果为空：{file_path.name}"
                )
            if len(payload) > _PDF_MAX_ENCODED_IMAGE_BYTES:
                raise PDFResourceLimitError(
                    f"PDF 第 {page_number} 页渲染图像体积超过安全上限："
                    f"{len(payload)} > {_PDF_MAX_ENCODED_IMAGE_BYTES} bytes"
                )
            yield payload
        _raise_if_cancelled(cancelled)


def _iter_pdf_ocr_images(
    file_path: Path,
    *,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
    render_pages: bool = False,
) -> Iterator[bytes]:
    """逐页、逐图像产出 OCR 输入；每次只保留当前图像的解码结果。"""
    if _PDF_BACKEND == "pdfium":
        yield from _iter_pdfium_ocr_images(
            file_path,
            progress_callback=progress_callback,
            cancelled=cancelled,
        )
        return
    if render_pages and pdfium is not None:
        # 显式整页渲染通道：渲染结果包含页面全部视觉信息（标题、照片等），
        # 用于内嵌图通道无法定论时的二次识别。
        yield from _iter_pdfium_ocr_images(
            file_path,
            progress_callback=progress_callback,
            cancelled=cancelled,
        )
        return

    total_pixels = 0
    total_decoded_bytes = 0
    image_count = 0
    with _open_pdf_reader(file_path) as (reader, page_count):
        for page_number, page in enumerate(reader.pages, start=1):
            _raise_if_cancelled(cancelled)
            if progress_callback is not None:
                progress_callback(
                    page_number,
                    page_count,
                    f"正在识别扫描 PDF：{file_path.name}（{page_number}/{page_count}）",
                )
            candidates = list(_iter_pdf_xobject_images(page))
            if len(candidates) > _PDF_MAX_IMAGES_PER_PAGE:
                raise PDFResourceLimitError(
                    f"PDF 第 {page_number} 页图像数量超过安全上限："
                    f"{len(candidates)} > {_PDF_MAX_IMAGES_PER_PAGE}"
                )
            for image_path, image_object in candidates:
                _raise_if_cancelled(cancelled)
                width, height = _pdf_image_dimensions(image_object)
                pixels = width * height
                decoded_bytes = pixels * 4
                if pixels > _PDF_MAX_IMAGE_PIXELS:
                    raise PDFResourceLimitError(
                        f"PDF 第 {page_number} 页图像像素超过安全上限："
                        f"{pixels} > {_PDF_MAX_IMAGE_PIXELS}"
                    )
                if decoded_bytes > _PDF_MAX_DECODED_IMAGE_BYTES:
                    raise PDFResourceLimitError(
                        f"PDF 第 {page_number} 页图像估算解码内存超过安全上限："
                        f"{decoded_bytes} > {_PDF_MAX_DECODED_IMAGE_BYTES} bytes"
                    )
                total_pixels += pixels
                total_decoded_bytes += decoded_bytes
                if total_pixels > _PDF_MAX_TOTAL_IMAGE_PIXELS:
                    raise PDFResourceLimitError(
                        f"PDF 累计图像像素超过安全上限：{file_path.name}"
                    )
                if total_decoded_bytes > _PDF_MAX_TOTAL_DECODED_IMAGE_BYTES:
                    raise PDFResourceLimitError(
                        f"PDF 累计估算解码内存超过安全上限：{file_path.name}"
                    )
                image_key: str | tuple[str, ...]
                image_key = image_path[0] if len(image_path) == 1 else image_path
                image_file = None
                try:
                    image_file = page.images[image_key]
                    payload = bytes(image_file.data)
                except ImportError as exc:
                    raise PDFRecognitionError(
                        "PDF 图片解码组件不可用，请重新安装完整版本后重试"
                    ) from exc
                except Exception as exc:
                    raise PDFRecognitionError(
                        f"PDF 第 {page_number} 页图像损坏或格式不受支持："
                        f"{file_path.name}"
                    ) from exc
                finally:
                    if image_file is not None:
                        pil_image = getattr(image_file, "image", None)
                        if pil_image is not None:
                            try:
                                pil_image.close()
                            except Exception:
                                pass
                    _release_pdf_image_decode_cache(image_object)
                if not payload:
                    raise PDFRecognitionError(
                        f"PDF 第 {page_number} 页图像为空：{file_path.name}"
                    )
                if len(payload) > _PDF_MAX_ENCODED_IMAGE_BYTES:
                    raise PDFResourceLimitError(
                        f"PDF 第 {page_number} 页解码后图像体积超过安全上限："
                        f"{len(payload)} > {_PDF_MAX_ENCODED_IMAGE_BYTES} bytes"
                    )
                image_count += 1
                yield payload
        _raise_if_cancelled(cancelled)
    if image_count == 0 and pdfium is not None:
        # 无可提取内嵌图时整页渲染兜底，覆盖标题为图形的电子证书等场景。
        yield from _iter_pdfium_ocr_images(
            file_path,
            progress_callback=progress_callback,
            cancelled=cancelled,
        )
        return
    if image_count == 0:
        raise PDFRecognitionError(
            f"扫描型 PDF 未找到可安全提取的页面图片：{file_path.name}；"
            "如该文件由特殊渲染器生成，请先另存为标准 PDF 或逐页图片后重试"
        )


def _extract_document_text(
    file_path: Path,
    *,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
) -> str:
    """提取文档内部文本（支持 .docx, .txt, .pdf, .doc 纯文本搜索）。"""
    ext = file_path.suffix.lower()
    if ext == ".docx":
        try:
            with zipfile.ZipFile(file_path) as zf:
                if "word/document.xml" in zf.namelist():
                    xml_content = _read_office_xml_member(zf, "word/document.xml")
                    tree = ET.fromstring(xml_content)
                    return "".join(tree.itertext())
        except Exception:
            pass
    elif ext in (".txt", ".csv"):
        try:
            return file_path.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            pass
    elif ext == ".pdf":
        return _extract_pdf_text(
            file_path,
            progress_callback=progress_callback,
            cancelled=cancelled,
        )
    elif ext == ".doc":
        try:
            with open(file_path, "rb") as f:
                content = f.read(150000)
            return content.decode("utf-8", errors="ignore")
        except Exception:
            pass
    return ""


def _build_doc_format_hint(file_path: Path) -> str | None:
    """针对旧版 .doc 文件给出友好提示，避免静默失败。

    .doc 为二进制 OLE 容器，直接 utf-8 解码几乎全部乱码；本工具不引入外部依赖，
    因此对该格式仅作识别提示，让用户主动另存为 .docx 后重跑。
    """
    if file_path.suffix.lower() != ".doc":
        return None
    return (
        f"⚠️ 旧版 Word 文件 {file_path.name} 为 .doc 格式，"
        "工具无法保证识别准确性；建议另存为 .docx 后重试。"
    )


def _build_employee_key(emp: TargetEmployee) -> str:
    """构造 (姓名|身份证) 维度的员工键，用于缓存与同名员工隔离。"""
    name = (emp.name or "").strip()
    id_card = (emp.id_card or "").strip()
    return f"{name}|{id_card}"


def _lookup_ocr_cache(
    cache: dict[str, Any],
    file_path: Path,
    employee_key: str = "",
    rel_path: str = "",
) -> tuple[str, str, str, str, str, str, bool, dict[str, Any]] | None:
    """按内容指纹查询缓存，并返回分类、脱敏 OCR 文字与分析状态。

    只要文件二进制内容没变，无论文件名如何修改、移动到何处，都能 100% 瞬间命中。
    """
    entries: dict[str, Any] = cache.get("entries") or {}
    if not entries:
        return None

    target_key = _compute_cache_key(file_path, employee_key)
    if not target_key:
        return None

    entry = entries.get(target_key)
    if not isinstance(entry, dict):
        return None

    return (
        entry.get("material_type") or "",
        entry.get("match_method") or "ocr_cached",
        entry.get("subtype") or "",
        entry.get("extracted_name") or "",
        entry.get("extracted_id_hash") or "",
        entry.get("ocr_text") or "",
        entry.get("analysis_state") == "complete",
        entry,
    )


def _store_ocr_cache(
    cache: dict[str, Any],
    file_path: Path,
    material_type: str,
    match_method: str,
    subtype: str,
    extracted_name: str,
    extracted_id: str,
    employee_key: str = "",
    rel_path: str = "",
    *,
    extracted_text: str = "",
    analysis_complete: bool | None = None,
    visual_ocr_query_signature: str = "",
) -> None:
    """OCR 分析完成后按内容指纹缓存标准分类及脱敏文字。"""
    fingerprint = _compute_file_fingerprint(file_path)
    if fingerprint is None:
        return
    size, mtime, sha = fingerprint

    cache_key = f"{sha[:24]}_{size}"
    entry = {
        "content_hash": sha[:24],
        "source_size": size,
        "source_mtime": mtime,
        "material_type": material_type,
        "match_method": match_method,
        "subtype": subtype,
        "extracted_name": extracted_name,
        "extracted_id_hash": _hash_id_card(extracted_id),
        "verified_at": _beijing_now_str(),
        "sample_filename": file_path.name,
    }
    if analysis_complete is not None:
        entry["analysis_state"] = "complete" if analysis_complete else "incomplete"
        entry["ocr_text"] = _sanitize_cached_text(extracted_text)
        entry["index_scope"] = LIBRARY_MODE_PERSON_FOLDER
        entry["document_page_marker"] = _footer_page_marker(extracted_text)
    entries: dict[str, Any] = cache.setdefault("entries", {})
    previous_entry = entries.get(cache_key)
    previous_queries = (
        previous_entry.get("visual_ocr_queries")
        if isinstance(previous_entry, dict)
        else []
    )
    if isinstance(previous_queries, list):
        entry["visual_ocr_queries"] = [
            str(item) for item in previous_queries[-31:] if str(item or "")
        ]
    if visual_ocr_query_signature:
        queries = entry.setdefault("visual_ocr_queries", [])
        if visual_ocr_query_signature not in queries:
            queries.append(visual_ocr_query_signature)
    entries[cache_key] = entry


_DOC_CONTENT_PATTERNS: dict[str, list[str]] = {
    "劳动合同": ["劳动合同", "用工合同", "劳务合同", "聘用合同", "用工协议", "劳动期限", "工作内容", "劳动报酬", "劳动争议", "解除劳动合同", "劳动法", "甲乙双方根据", "试用期"],
    "学历证明": ["毕业证书", "学位证书", "教育部学历证书", "学信网", "普通高等学校", "学士学位", "硕士学位", "博士学位"],
    "安全员证": ["安全生产考核合格证书", "建筑施工企业项目负责人安全生产考核合格证书", "建筑施工企业专职安全生产管理人员安全生产考核合格证书", "安全员考核合格证", "安全员C证", "安全员A证", "安全员B证"],
    "特种证书": ["特种作业操作证", "特种作业人员操作证", "特种设备作业人员证", "特种作业", "特种设备作业"],
    "资格证书": ["职业资格证书", "专业技术职务资格证书", "职称证书", "技能等级证书", "中华人民共和国机动车驾驶证"],
}


def _sanitize_cached_text(text: str) -> str:
    """缓存 OCR 摘要前脱敏证件号、手机号，并限制体积。"""
    compact = re.sub(r"\s+", " ", str(text or "")).strip()
    compact = re.sub(
        r"(?<!\d)(\d{17}[\dXx]|\d{15})(?!\d)",
        lambda match: _mask_id_card(match.group(1)),
        compact,
    )
    compact = re.sub(
        r"(?<!\d)(1[3-9]\d{9})(?!\d)",
        lambda match: match.group(1)[:3] + "****" + match.group(1)[-4:],
        compact,
    )
    return compact[:_OCR_CACHE_TEXT_SNIPPET_MAX]


def _normalize_person_name(name: str) -> str:
    return re.sub(r"[\s·•・]", "", str(name or "")).strip()


def _extract_person_names(text: str) -> list[str]:
    """从证件、合同等正文提取明确带字段标签的人名，不依赖文件名。"""
    normalized = re.sub(r"[\r\n\t]+", " ", str(text or ""))
    labels = (
        "劳动者姓名", "员工姓名", "人员姓名", "持证人姓名", "申请人姓名",
        "姓名", "持证人", "劳动者", "乙方", "受聘人", "开户人",
    )
    names: list[str] = []
    suffix_noise = (
        "性别", "民族", "出生", "住址", "身份证", "公民身份", "证件号码",
        "电话", "手机号", "合同", "岗位", "部门", "文化程度", "有效期",
    )
    for label in labels:
        pattern = rf"{re.escape(label)}\s*(?:名称|[（(](?:签名|签字)[）)])?\s*[:：]?\s*([\u3400-\u9fff·•・]{{2,10}})"
        for match in re.finditer(pattern, normalized):
            candidate = match.group(1).strip()
            for suffix in suffix_noise:
                if suffix in candidate:
                    candidate = candidate.split(suffix, 1)[0]
            candidate = candidate.strip()
            if candidate in {"签字", "签名", "签字姓名", "年月日", "甲方", "乙方"}:
                continue
            if not (2 <= len(_normalize_person_name(candidate)) <= 8):
                continue
            if not _is_valid_person_name(candidate):
                continue
            if candidate not in names:
                names.append(candidate)
    return names


def _extract_id_card(text: str) -> str:
    match = re.search(r"(?<!\d)(\d{17}[\dXx]|\d{15})(?!\d)", str(text or ""))
    return match.group(1) if match else ""


def _extract_phone(text: str) -> str:
    match = re.search(r"(?<!\d)(1[3-9]\d{9})(?!\d)", str(text or ""))
    return match.group(1) if match else ""


def _compact_for_window(text: str) -> str:
    """去除空白与标点，仅保留中英文与数字，用于连续短语匹配。"""
    return re.sub(r"[^0-9A-Za-z一-鿿]+", "", str(text or ""))


def _canonical_material_for_request_label(material: str) -> str | None:
    """将明确的正式请求名称映射到内部标准类型；宽泛关键词不参与。"""
    label = str(material or "").strip()
    if label in MATERIAL_SYNONYMS:
        return label
    compact_label = _compact_for_window(label)
    for canonical, aliases in _SAFE_REQUEST_MATERIAL_ALIASES.items():
        if any(_compact_for_window(alias) == compact_label for alias in aliases):
            return canonical
    return None


def _requested_label_for_detected_material(
    detected_material: str,
    requested_materials: list[str] | tuple[str, ...],
) -> str | None:
    """把内部标准分类安全地还原为用户本次请求的显示名称。"""
    detected = str(detected_material or "").strip()
    if not detected:
        return None
    for requested in requested_materials:
        label = str(requested or "").strip()
        if label == detected:
            return label
    detected_canonical = _canonical_material_for_request_label(detected) or detected
    for requested in requested_materials:
        label = str(requested or "").strip()
        if (
            label
            and _canonical_material_for_request_label(label) == detected_canonical
        ):
            return label
    return None


def _has_reordered_name_window(compact_text: str, compact_name: str) -> bool:
    """仅允许把材料名拆成两段后前后互换；不接受任意字符排列。

    例如“我的证书天谴”包含“证书天谴”，视为“天谴证书”的乱序命中；
    “我天，天谴的证书”去标点后仍被“的”打断，不命中。
    """
    length = len(compact_name)
    if (
        length < 4
        or length > 12
        or len(compact_text) < length
        or re.fullmatch(r"[一-鿿]+", compact_name) is None
    ):
        return False
    searchable_text = compact_text[:20_000]
    for split_at in range(1, length):
        if compact_name[split_at:] + compact_name[:split_at] in searchable_text:
            return True
    return False


def _has_special_certificate_evidence(text: str) -> bool:
    """识别特种证书的强证据，发证机关名称本身不能单独定类。"""
    full_text = str(text or "")
    if any(
        keyword in full_text
        for keyword in (
            "特种作业操作证",
            "特种作业人员操作证",
            "特种设备作业人员证",
            "特种作业人员",
            "特种设备作业",
            "特种作业",
        )
    ):
        return True
    return "应急管理局" in full_text and any(
        field in full_text
        for field in ("作业类别", "准操项目", "操作项目", "作业项目", "复审日期")
    )


def _has_identity_card_front_evidence(text: str) -> bool:
    full_text = str(text or "")
    return "公民身份号码" in full_text or (
        "姓名" in full_text
        and any(keyword in full_text for keyword in ("住址", "民族", "出生"))
    )


def _is_weak_cached_material_result(
    material: str,
    method: str,
    cached_text: str,
) -> bool:
    """识别当前及旧缓存中的弱分类，不把真实身份证强证据降级。"""
    if not material or material == "其他材料":
        return True
    if "id_number_fallback" in method:
        return True
    return (
        material == "身份证"
        and method.endswith("id_front")
        and not _has_identity_card_front_evidence(cached_text)
    )


def _visual_ocr_query_signature(requested_types: list[str] | None) -> str:
    normalized = sorted({
        str(material or "").strip()
        for material in (requested_types or [])
        if str(material or "").strip()
    })
    if not normalized:
        return ""
    return hashlib.sha256("\0".join(normalized).encode("utf-8")).hexdigest()[:16]


def _visual_ocr_query_was_attempted(
    entry: dict[str, Any],
    requested_types: list[str] | None,
) -> bool:
    signature = _visual_ocr_query_signature(requested_types)
    queries = entry.get("visual_ocr_queries")
    return bool(signature and isinstance(queries, list) and signature in queries)


def _record_visual_ocr_query(
    entry: dict[str, Any],
    requested_types: list[str] | None,
) -> None:
    signature = _visual_ocr_query_signature(requested_types)
    if not signature:
        return
    raw_queries = entry.get("visual_ocr_queries")
    queries = [
        str(item) for item in raw_queries[-31:] if str(item or "")
    ] if isinstance(raw_queries, list) else []
    if signature not in queries:
        queries.append(signature)
    entry["visual_ocr_queries"] = queries


def _classify_requested_material_text(
    text: str,
    requested_types: list[str] | None,
    *,
    method_prefix: str,
) -> tuple[str | None, str, str]:
    """按当前请求的完整材料名称匹配正文，不扩展或猜测同义词。"""
    full_text = str(text or "")
    compact_text = _compact_for_window(full_text)
    for requested in sorted(requested_types or [], key=len, reverse=True):
        material = str(requested or "").strip()
        compact_material = _compact_for_window(material)
        if compact_material and compact_material in compact_text:
            return material, f"{method_prefix}_custom", ""
    window_text = _compact_for_window(full_text)
    for requested in sorted(requested_types or [], key=len, reverse=True):
        material = str(requested or "").strip()
        window_name = _compact_for_window(material)
        if _has_reordered_name_window(window_text, window_name):
            return material, f"{method_prefix}_custom_reordered", ""
    return None, "", ""


def _classify_text_content(
    text: str,
    *,
    requested_types: list[str] | None = None,
    method_prefix: str = "ocr",
    allow_weak_id_fallback: bool = True,
) -> tuple[str | None, str, str]:
    """按正文判定材料类型；返回 (类型, 匹配方式, 正反面子类型)。

    先应用已有标准材料强规则，再匹配用户显式输入的完整自定义名称；
    仅含 18 位号码的身份证推断最后执行，避免证书编号抢占真实分类。
    """
    full_text = str(text or "")
    if not full_text.strip():
        return None, "", ""

    if _has_special_certificate_evidence(full_text):
        return "特种证书", f"{method_prefix}_special_cert", ""
    if any(keyword in full_text for keyword in ("安全生产考核", "安全考核合格", "建安C证", "建安A证", "建安B证")):
        return "安全员证", f"{method_prefix}_safety_cert", ""
    if (
        any(keyword in full_text for keyword in ("劳动合同", "用工合同", "劳务合同", "聘用合同"))
        or ("甲方" in full_text and "乙方" in full_text and any(keyword in full_text for keyword in ("劳动", "报酬", "工作内容")))
    ):
        return "劳动合同", f"{method_prefix}_contract", ""
    if any(keyword in full_text for keyword in ("毕业证书", "学位证书", "学信网", "学历证书", "普通高等学校")):
        return "学历证明", f"{method_prefix}_degree", ""
    if any(keyword in full_text for keyword in ("机动车驾驶证", "驾驶证", "职业资格证书", "职业资格", "职称证书", "技能等级证书")):
        return "资格证书", f"{method_prefix}_certificate", ""
    if re.search(r"\d{16,19}", full_text) and any(keyword in full_text for keyword in ("银行", "银联", "Bank")):
        return "银行卡", f"{method_prefix}_bank", ""
    if "居民身份证" in full_text or (
        "签发机关" in full_text and "有效期限" in full_text
        and "特种" not in full_text and "安全" not in full_text
    ):
        return "身份证", f"{method_prefix}_id_back", "反面"
    if _has_identity_card_front_evidence(full_text):
        return "身份证", f"{method_prefix}_id_front", "正面"

    custom_types = [
        material_type
        for material_type in (requested_types or [])
        if str(material_type or "").strip() and material_type not in MATERIAL_SYNONYMS
    ]
    custom_material, custom_method, custom_subtype = _classify_requested_material_text(
        full_text, custom_types, method_prefix=method_prefix,
    )
    if custom_material:
        return custom_material, custom_method, custom_subtype

    # 标准请求名称仍保留既有完整正文匹配；乱序仅用于自定义材料。
    for requested in sorted(requested_types or [], key=len, reverse=True):
        if requested and requested in full_text:
            return requested, f"{method_prefix}_custom", ""
    if allow_weak_id_fallback and _extract_id_card(full_text):
        return "身份证", f"{method_prefix}_id_number_fallback", "正面"
    return None, "", ""


def _iter_ocr_targets(
    file_path: Path,
    *,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
    render_pages: bool = False,
) -> Iterator[str | bytes]:
    if file_path.suffix.lower() == ".pdf":
        yield from _iter_pdf_ocr_images(
            file_path,
            progress_callback=progress_callback,
            cancelled=cancelled,
            render_pages=render_pages,
        )
        return
    _raise_if_cancelled(cancelled)
    yield str(file_path)


def _extract_ocr_result_texts(result: Any) -> list[str]:
    """只保留 OCR 文字，立即丢弃坐标框与置信度等大对象。"""
    texts: list[str] = []
    for item in result or []:
        if not isinstance(item, (list, tuple)):
            continue
        if len(item) >= 3 and isinstance(item[1], str):
            texts.append(item[1])
        elif len(item) == 2 and isinstance(item[0], str) and isinstance(item[1], str):
            # 一些本地适配器以 [字段名, 字段值] 返回。
            texts.append(f"{item[0]}：{item[1]}")
        elif len(item) >= 2 and isinstance(item[1], str):
            texts.append(item[1])
    return texts


def _collect_ocr_texts(
    file_path: Path,
    *,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
    render_pages: bool = False,
) -> tuple[list[str], bool]:
    """逐页 OCR，只累计文字；页面坐标结果在下一页前即可释放。"""
    engine = _get_ocr_engine()
    if engine is None:
        return [], False

    texts: list[str] = []
    try:
        for target_input in _iter_ocr_targets(
            file_path,
            progress_callback=progress_callback,
            cancelled=cancelled,
            render_pages=render_pages,
        ):
            _raise_if_cancelled(cancelled)
            with _OCR_LOCK:
                result, _ = engine(target_input)
            if result:
                texts.extend(_extract_ocr_result_texts(result))
    except (MaterialCollectionCancelled, PDFRecognitionError):
        raise
    except Exception:
        return [], False
    return texts, True


def _read_ocr_text(
    file_path: Path,
    *,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
    render_pages: bool = False,
) -> tuple[str, list[str], bool]:
    """调用本地 OCR 并兼容 RapidOCR 与测试/旧适配器的两种结果形态。"""
    texts, analysis_complete = _collect_ocr_texts(
        file_path,
        progress_callback=progress_callback,
        cancelled=cancelled,
        render_pages=render_pages,
    )
    if not texts:
        return "", [], analysis_complete
    return "\n".join(texts), texts, analysis_complete


def _analyze_ocr_file(
    file_path: Path,
    requested_types: list[str] | None = None,
    *,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
    render_pages: bool = False,
    allow_weak_id_fallback: bool = True,
) -> tuple[str | None, str, str, str, str, str, list[str], bool]:
    full_text, _texts, analysis_complete = _read_ocr_text(
        file_path,
        progress_callback=progress_callback,
        cancelled=cancelled,
        render_pages=render_pages,
    )
    names = _extract_person_names(full_text)
    extracted_name = names[0] if names else ""
    extracted_id = _extract_id_card(full_text)
    material, method, subtype = _classify_text_content(
        full_text,
        requested_types=requested_types,
        method_prefix="ocr",
        allow_weak_id_fallback=allow_weak_id_fallback,
    )
    return material, method, subtype, extracted_name, extracted_id, full_text, names, analysis_complete


def _analyze_folder_ocr_file(
    file_path: Path,
    requested_types: list[str] | None = None,
    *,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
    render_pages: bool = False,
    allow_weak_id_fallback: bool = True,
) -> tuple[str | None, str, str, str, str, str, bool]:
    """保持既有标准材料规则，并返回可供自定义材料重分类的 OCR 文字。"""
    texts, analysis_complete = _collect_ocr_texts(
        file_path,
        progress_callback=progress_callback,
        cancelled=cancelled,
        render_pages=render_pages,
    )
    if not texts:
        return None, "", "", "", "", "", analysis_complete
    full_text = " ".join(texts)

    extracted_name = ""
    extracted_id = ""
    id_match = re.search(r"\b\d{17}[\dxX]\b", full_text)
    if id_match:
        extracted_id = id_match.group(0)

    for idx, text in enumerate(texts):
        match = re.search(r"姓名(?:[/A-Za-z\s:：]*)([\u4e00-\u9fa5]{2,10})", text)
        if match:
            extracted_name = match.group(1).strip()
            break
        if re.search(r"^姓名(?:[/A-Za-z\s:：]*)$", text.strip()):
            if idx + 1 < len(texts) and re.match(r"^[\u4e00-\u9fa5]{2,10}$", texts[idx + 1].strip()):
                extracted_name = texts[idx + 1].strip()
                break

    material: str | None = None
    method = ""
    subtype = ""
    if _has_special_certificate_evidence(full_text):
        material, method = "特种证书", "ocr_special_cert"
    elif (
        "安全生产考核" in full_text or "安全考核合格" in full_text
        or "建安C证" in full_text or "建安A证" in full_text
        or "建安B证" in full_text or "安全员" in full_text
    ):
        material, method = "安全员证", "ocr_safety_cert"
    elif (
        "劳动合同" in full_text or "用工合同" in full_text
        or ("甲方" in full_text and "乙方" in full_text and (
            "劳动" in full_text or "报酬" in full_text or "工作内容" in full_text
        ))
    ):
        material, method = "劳动合同", "ocr_contract"
    elif (
        "毕业证书" in full_text or "学位证书" in full_text or "学信网" in full_text
        or "学历证书" in full_text or "普通高等学校" in full_text
    ):
        material, method = "学历证明", "ocr_degree"
    elif (
        "机动车驾驶证" in full_text or "驾驶证" in full_text
        or "职业资格证书" in full_text or "职业资格" in full_text
    ):
        material, method = "资格证书", "ocr_certificate"
    elif re.search(r"\d{16,19}", full_text) and (
        "银行" in full_text or "银联" in full_text or "Bank" in full_text
    ):
        material, method = "银行卡", "ocr_bank"
    elif "居民身份证" in full_text or (
        "签发机关" in full_text and "有效期限" in full_text
        and "特种" not in full_text and "安全" not in full_text
    ):
        material, method, subtype = "身份证", "ocr_id_back", "反面"
    elif (
        "公民身份号码" in full_text
        or ("姓名" in full_text and ("住址" in full_text or "民族" in full_text or "出生" in full_text))
    ):
        material, method, subtype = "身份证", "ocr_id_front", "正面"
    else:
        custom_types = [
            material_type
            for material_type in (requested_types or [])
            if material_type not in MATERIAL_SYNONYMS
        ]
        material, method, subtype = _classify_requested_material_text(
            full_text,
            custom_types,
            method_prefix="ocr",
        )
        if material is None and extracted_id and allow_weak_id_fallback:
            material, method, subtype = (
                "身份证",
                "ocr_id_number_fallback",
                "正面",
            )

    return (
        material,
        method,
        subtype,
        extracted_name,
        extracted_id,
        full_text,
        analysis_complete,
    )


def _classify_by_ocr(file_path: Path) -> tuple[str | None, str, str, str, str]:
    """通过本地离线 OCR 识别图片文字并执行既有标准材料分类。"""
    material, method, subtype, name, extracted_id, _text, _complete = (
        _analyze_folder_ocr_file(file_path)
    )
    return material, method, subtype, name, extracted_id


def _classify_material_type(
    file_path: Path,
    filename: str,
    requested_types: list[str],
    *,
    employee_key: str = "",
    rel_path: str = "",
    cache: dict[str, Any] | None = None,
    use_cache: bool = True,
    cache_stats: dict[str, int] | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
    analysis_records: dict[Path, dict[str, Any]] | None = None,
) -> tuple[str | None, str, str, str, str, bool]:
    """Classify file into a material type using filenames, document contents, and local OCR.

    Returns: (matched_material_type or None, match_method, subtype_label,
              extracted_name, extracted_id, cache_hit)
    cache_hit=True 表示本次分类结果来自 OCR 缓存命中。
    """
    _raise_if_cancelled(cancelled)
    is_pdf = file_path.suffix.lower() == ".pdf"
    if is_pdf:
        _validate_pdf_source_size(file_path)
    stem = Path(filename).stem.lower()

    # 1. 优先匹配当前请求列表中明确包含在文件名里的材料（标准或自定义，按关键词长度降序最长优先匹配）
    sorted_req_types = sorted(
        [r for r in requested_types if r and r.strip()],
        key=lambda x: len(x),
        reverse=True,
    )
    for req_type in sorted_req_types:
        syns = MATERIAL_SYNONYMS.get(req_type, [req_type])
        # 按同义词长度降序，最长精确匹配优先（例如"保密协议"优先于"协议"）
        for syn in sorted(syns, key=lambda s: len(s), reverse=True):
            if syn.lower() in stem:
                sub = "正面" if "正面" in stem or "人像" in stem else ("反面" if "反面" in stem or "国徽" in stem else "")
                if is_pdf:
                    # 文件名已足够判型时仍需拒绝损坏、加密或超限 PDF。
                    with _open_pdf_document(file_path):
                        pass
                return req_type, "filename_keyword", sub, "", "", False

    # 2b. 全量同义词库匹配（按优先级互斥判断）
    for mat_type, synonyms in MATERIAL_SYNONYMS.items():
        for syn in sorted(synonyms, key=lambda s: len(s), reverse=True):
            if syn.lower() in stem:
                sub = "正面" if "正面" in stem or "人像" in stem else ("反面" if "反面" in stem or "国徽" in stem else "")
                if is_pdf:
                    with _open_pdf_document(file_path):
                        pass
                requested_label = _requested_label_for_detected_material(
                    mat_type, requested_types,
                )
                return requested_label or mat_type, "filename_keyword", sub, "", "", False

    # 3. 文档内部文本内容深度检索（针对文件名如 01.docx, file.pdf 等非标准命名）
    doc_text = _extract_document_text(
        file_path,
        progress_callback=progress_callback,
        cancelled=cancelled,
    )
    if doc_text:
        if _has_special_certificate_evidence(doc_text):
            requested_label = _requested_label_for_detected_material(
                "特种证书", requested_types,
            )
            return requested_label or "特种证书", "doc_content", "", "", "", False
        for mat_type, content_keywords in _DOC_CONTENT_PATTERNS.items():
            for kw in content_keywords:
                if kw in doc_text:
                    requested_label = _requested_label_for_detected_material(
                        mat_type, requested_types,
                    )
                    return requested_label or mat_type, "doc_content", "", "", "", False
        custom_types = [
            material_type
            for material_type in requested_types
            if material_type not in MATERIAL_SYNONYMS
        ]
        custom_mat, custom_method, custom_subtype = _classify_requested_material_text(
            doc_text,
            custom_types,
            method_prefix="doc_content",
        )
        if custom_mat:
            return custom_mat, custom_method, custom_subtype, "", "", False

    # 4. 本地离线 OCR 视觉图文识别（针对纯哈希/随机命名的图片或扫描版 PDF）
    ext = file_path.suffix.lower()
    if ext in IMAGE_EXTENSIONS or ext == ".pdf":
        # 4a. 先查缓存（基于文件二进制内容 SHA256 哈希指纹，改名或移动均能 100% 秒级命中）
        if use_cache and cache is not None:
            hit = _lookup_ocr_cache(cache, file_path, employee_key=employee_key, rel_path=rel_path)
            if hit is not None:
                (
                    mat,
                    method,
                    sub,
                    name,
                    _id_hash,
                    cached_text,
                    analysis_complete,
                    cached_entry,
                ) = hit
                if analysis_records is not None:
                    analysis_records[file_path] = cached_entry
                weak_cached_result = _is_weak_cached_material_result(
                    mat,
                    method,
                    cached_text,
                )
                matched_request = _requested_label_for_detected_material(
                    mat, requested_types,
                )
                has_alternative_request = any(
                    _requested_label_for_detected_material(mat, [requested]) is None
                    for requested in requested_types
                )
                visual_query_attempted = _visual_ocr_query_was_attempted(
                    cached_entry,
                    requested_types,
                )
                if matched_request is not None and (
                    not weak_cached_result
                    or not has_alternative_request
                    or visual_query_attempted
                ):
                    if cache_stats is not None:
                        cache_stats["hits"] = cache_stats.get("hits", 0) + 1
                    return matched_request, method, sub, name, "", True
                if analysis_complete:
                    cached_mat, cached_method, cached_sub = _classify_text_content(
                        cached_text,
                        requested_types=requested_types,
                        method_prefix="cached_text",
                        allow_weak_id_fallback=False,
                    )
                    cached_request = _requested_label_for_detected_material(
                        cached_mat or "", requested_types,
                    )
                    if cached_request is not None:
                        if cache_stats is not None:
                            cache_stats["hits"] = cache_stats.get("hits", 0) + 1
                        return cached_request, cached_method, cached_sub, name, "", True
                    should_retry_visual_ocr = (
                        is_pdf
                        and pdfium is not None
                        and weak_cached_result
                        and not visual_query_attempted
                        and (matched_request is None or has_alternative_request)
                    )
                    if not should_retry_visual_ocr:
                        if cache_stats is not None:
                            cache_stats["hits"] = cache_stats.get("hits", 0) + 1
                        return matched_request or mat or None, method, sub, name, "", True
                if cache_stats is not None:
                    cache_stats["invalidated"] = cache_stats.get("invalidated", 0) + 1
            if cache_stats is not None:
                cache_stats["misses"] = cache_stats.get("misses", 0) + 1

        # 4b. 缓存未命中或不可用 → 真实 OCR
        used_page_render = (
            is_pdf
            and pdfium is not None
            and (bool(doc_text) or _PDF_BACKEND == "pdfium")
        )
        completed_visual_ocr = False
        (
            ocr_mat,
            ocr_method,
            ocr_sub,
            ocr_name,
            ocr_id,
            ocr_text,
            analysis_complete,
        ) = _analyze_folder_ocr_file(
            file_path,
            requested_types=requested_types,
            progress_callback=progress_callback,
            cancelled=cancelled,
            render_pages=used_page_render,
            allow_weak_id_fallback=False,
        )
        completed_visual_ocr = used_page_render and analysis_complete
        combined_text = " ".join(part for part in (doc_text, ocr_text) if part)
        combined_mat, combined_method, combined_sub = _classify_text_content(
            combined_text,
            requested_types=requested_types,
            method_prefix="ocr",
            allow_weak_id_fallback=False,
        )
        if combined_mat:
            ocr_mat, ocr_method, ocr_sub = combined_mat, combined_method, combined_sub
        if not ocr_name:
            combined_names = _extract_person_names(combined_text)
            ocr_name = combined_names[0] if combined_names else ""
        ocr_id = ocr_id or _extract_id_card(combined_text)
        ocr_text = combined_text
        if (
            is_pdf
            and pdfium is not None
            and not ocr_mat
            and not used_page_render
        ):
            (
                render_mat,
                render_method,
                render_sub,
                render_name,
                render_id,
                render_text,
                render_complete,
            ) = _analyze_folder_ocr_file(
                file_path,
                requested_types=requested_types,
                progress_callback=progress_callback,
                cancelled=cancelled,
                render_pages=True,
                allow_weak_id_fallback=False,
            )
            analysis_complete = render_complete
            completed_visual_ocr = render_complete
            if render_complete:
                ocr_text = " ".join(part for part in (combined_text, render_text) if part)
                ocr_name = ocr_name or render_name
                ocr_id = ocr_id or render_id or _extract_id_card(ocr_text)
                ocr_mat, ocr_method, ocr_sub = _classify_text_content(
                    ocr_text,
                    requested_types=requested_types,
                    method_prefix="ocr",
                    allow_weak_id_fallback=False,
                )
                if not ocr_mat and render_mat:
                    ocr_mat, ocr_method, ocr_sub = render_mat, render_method, render_sub
        if not ocr_mat and ocr_text:
            ocr_mat, ocr_method, ocr_sub = _classify_text_content(
                ocr_text,
                requested_types=requested_types,
                method_prefix="ocr",
                allow_weak_id_fallback=True,
            )
        if analysis_records is not None and analysis_complete:
            analysis_records[file_path] = {
                "material_type": ocr_mat or "其他材料", "match_method": ocr_method,
                "extracted_names": _extract_person_names(ocr_text),
                "extracted_id_hash": _hash_id_card(ocr_id),
                "ocr_text": _sanitize_cached_text(ocr_text),
                "document_page_marker": _footer_page_marker(ocr_text),
            }
        if use_cache and cache is not None and analysis_complete:
            stores_standard_result = ocr_mat in MATERIAL_SYNONYMS
            _store_ocr_cache(
                cache,
                file_path,
                ocr_mat if stores_standard_result else "其他材料",
                ocr_method if stores_standard_result else "unrecognized",
                ocr_sub if stores_standard_result else "",
                ocr_name,
                ocr_id,
                employee_key=employee_key,
                rel_path=rel_path,
                extracted_text=ocr_text,
                analysis_complete=True,
                visual_ocr_query_signature=(
                    _visual_ocr_query_signature(requested_types)
                    if completed_visual_ocr
                    else ""
                ),
            )
        if ocr_mat:
            requested_label = _requested_label_for_detected_material(
                ocr_mat, requested_types,
            )
            return requested_label or ocr_mat, ocr_method, ocr_sub, ocr_name, ocr_id, False

    return None, "", "", "", "", False


def _scan_folder_index(
    lib_path: Path,
    max_depth: int = 1,
    skip_dir: Path | None = None,
) -> dict[str, list[Path]]:
    """扫描资料库，建立"文件夹名 → 文件夹路径列表"的索引，主动跳过输出目录。"""
    folder_index: dict[str, list[Path]] = {}

    def _scan(parent: Path, depth: int) -> None:
        if depth > max_depth:
            return
        if skip_dir and _is_path_nested(parent, skip_dir):
            return
        try:
            entries = list(os.scandir(parent))
        except PermissionError:
            return
        for entry in entries:
            if entry.is_dir(follow_symlinks=False):
                name = entry.name
                if name.startswith(".") or name in _IGNORED_FILENAMES:
                    continue
                path = Path(entry.path)
                if skip_dir and _is_path_nested(path, skip_dir):
                    continue
                folder_index.setdefault(name, []).append(path)
                if depth < max_depth:
                    _scan(path, depth + 1)

    _scan(lib_path, 1)
    return folder_index


@dataclass(frozen=True)
class _FlatIndexedFile:
    source_path: Path
    relative_path: str
    cache_key: str
    material_type: str
    match_method: str
    subtype: str
    extracted_names: tuple[str, ...]
    extracted_id_hash: str
    extracted_phone_hash: str = ""
    text_snippet: str = ""
    extracted_id_card: str = ""
    cache_hit: bool = False
    document_group_id: str = ""
    document_page_number: int = 0
    document_page_count: int = 0
    document_warning: str = ""
    filename_names: tuple[str, ...] = ()
    filename_id_hash: str = ""
    filename_employee_no: str = ""
    ocr_page_marker: tuple[int, int | None] | None = None
    confirmed_employee_key: str = ""


def _scan_flat_library_files(
    lib_path: Path,
    skip_dir: Path | None = None,
    *,
    cancelled: Callable[[], bool] | None = None,
) -> list[Path]:
    """递归扫描无序资料库中的可识别文件，跳过隐藏项、链接目录和输出目录。"""
    _raise_if_cancelled(cancelled)
    keyed_paths: list[tuple[str, Path]] = []
    pending: list[tuple[Path, str]] = [(lib_path, "")]
    scanned_entries = 0

    while pending:
        parent, relative_parent = pending.pop()
        try:
            entries = list(os.scandir(parent))
        except OSError:
            continue

        child_directories: list[tuple[Path, str]] = []
        for entry in entries:
            scanned_entries += 1
            if scanned_entries % 512 == 0:
                _raise_if_cancelled(cancelled)
            try:
                is_directory = entry.is_dir(follow_symlinks=False)
                is_file = entry.is_file(follow_symlinks=False)
            except OSError:
                continue

            source_path = Path(entry.path)
            relative_path = (
                f"{relative_parent}/{entry.name}"
                if relative_parent
                else entry.name
            )
            if is_directory:
                if (
                    entry.name.startswith(".")
                    or entry.name.lower() in _IGNORED_FILENAMES
                    or (skip_dir and _is_path_nested(source_path, skip_dir))
                ):
                    continue
                child_directories.append((source_path, relative_path))
                continue
            if not is_file or _is_junk_or_temp_file(entry.name):
                continue
            if source_path.suffix.lower() not in SUPPORTED_FILE_EXTENSIONS:
                continue
            keyed_paths.append((relative_path.casefold(), source_path))

        # os.walk visits child directories in scandir order.  The stack is
        # reversed so equal case-folded paths retain the same stable ordering.
        pending.extend(reversed(child_directories))

    keyed_paths.sort(key=lambda item: item[0])
    _raise_if_cancelled(cancelled)
    return [path for _sort_key, path in keyed_paths]


def _extract_flat_document_text(
    file_path: Path,
    *,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
) -> str:
    """提取无序库文档正文；该扩展能力仅用于 TASK-8，不改变原文件夹模式。"""
    ext = file_path.suffix.lower()
    if ext == ".docx":
        try:
            with zipfile.ZipFile(file_path) as archive:
                tree = ET.fromstring(_read_office_xml_member(archive, "word/document.xml"))
                return " ".join(text for text in tree.itertext() if text)
        except Exception:
            return ""
    if ext in DOC_EXTENSIONS:
        return _extract_document_text(
            file_path,
            progress_callback=progress_callback,
            cancelled=cancelled,
        )
    if ext == ".xlsx":
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            chunks: list[str] = []
            character_count = 0
            try:
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        for value in row:
                            if value is not None:
                                text = str(value)
                                chunks.append(text)
                                character_count += len(text)
                                if character_count >= 20000:
                                    return " ".join(chunks)
            finally:
                workbook.close()
            return " ".join(chunks)
        except Exception:
            return ""
    if ext == ".xls":
        try:
            import xlrd

            workbook = xlrd.open_workbook(str(file_path), on_demand=True)
            chunks: list[str] = []
            character_count = 0
            try:
                for sheet in workbook.sheets():
                    for row_index in range(sheet.nrows):
                        for value in sheet.row_values(row_index):
                            if value not in (None, ""):
                                text = str(value)
                                chunks.append(text)
                                character_count += len(text)
                                if character_count >= 20000:
                                    return " ".join(chunks)
            finally:
                workbook.release_resources()
            return " ".join(chunks)
        except Exception:
            return ""
    if ext == ".pptx":
        try:
            with zipfile.ZipFile(file_path) as archive:
                if len(archive.infolist()) > _OFFICE_ARCHIVE_MAX_MEMBERS:
                    return ""
                xml_names = [
                    name for name in archive.namelist()
                    if name.endswith(".xml") and (name.startswith("ppt/slides/") or name == "word/document.xml")
                ]
                total_xml_bytes = sum(archive.getinfo(name).file_size for name in xml_names)
                if total_xml_bytes > _OFFICE_XML_TOTAL_MAX_BYTES:
                    return ""
                chunks: list[str] = []
                remaining_bytes = _OFFICE_XML_TOTAL_MAX_BYTES
                for name in xml_names:
                    payload = _read_office_xml_member(
                        archive,
                        name,
                        remaining_bytes=remaining_bytes,
                    )
                    remaining_bytes -= len(payload)
                    tree = ET.fromstring(payload)
                    chunks.extend(text for text in tree.itertext() if text)
                return " ".join(chunks)
        except Exception:
            return ""
    return ""


def _classify_material_from_filename(filename: str, requested_types: list[str]) -> tuple[str | None, str]:
    """正文无法判型时，仅用文件名补充材料类型；绝不以文件名判断人员。"""
    stem = Path(filename).stem.casefold()
    types = list(dict.fromkeys([*requested_types, *MATERIAL_SYNONYMS.keys()]))
    weak_tokens = {"id", "a面", "b面", "正", "反", "照片", "证件"}
    for material in sorted(types, key=len, reverse=True):
        synonyms = MATERIAL_SYNONYMS.get(material, [material])
        for synonym in sorted(synonyms, key=len, reverse=True):
            token = synonym.strip().casefold()
            if not token or token in weak_tokens or len(token) < 2:
                continue
            if token in stem:
                return material, "filename_material_only"
    return None, ""


def _flat_cache_entry_usable(entry: Any) -> bool:
    if not isinstance(entry, dict):
        return False
    if entry.get("analysis_state") == "complete":
        return True
    # 兼容旧缓存：只有明确识别出人员的成功条目才可用于无序库，不能把旧空结果当负缓存。
    return bool(entry.get("material_type") and entry.get("extracted_name"))


def _is_portrait_material(material: str) -> bool:
    return material in {"证件照片", "证件照", "一寸照", "二寸照", "登记照", "个人照片"}


def _portrait_filename_identity(filename: str) -> tuple[tuple[str, ...], str, str]:
    """只解析照片名称中的完整身份词；结果随路径重算，不能写入内容 OCR 缓存。"""
    stem = Path(filename).stem
    for keyword in sorted(MATERIAL_SYNONYMS["证件照片"], key=len, reverse=True):
        stem = re.sub(re.escape(keyword), " ", stem, flags=re.IGNORECASE)
    identity = _extract_id_card(stem)
    if identity:
        stem = stem.replace(identity, " ")
    names = tuple(dict.fromkeys(
        _normalize_person_name(token) for token in re.findall(r"[\u3400-\u9fff·]{2,10}", stem)
        if _is_valid_person_name(token)
        and token not in {"入职", "正面", "反面", "原图", "副本", "扫描件", "电子版"}
    ))
    numbers = re.findall(r"(?:工号|员工编号)\s*[:：]?\s*(\d{3,16})(?!\d)", stem)
    if not numbers and not names:
        numbers = re.findall(r"(?<!\w)\d{3,16}(?!\w)", re.sub(r"[_（）()\-]", " ", stem))
    employee_no = numbers[0] if len(numbers) == 1 else ""
    return names, _hash_id_card(identity), employee_no


def _named_portrait(file_path: Path, requested_types: list[str]) -> tuple[str | None, str]:
    if file_path.suffix.lower() not in IMAGE_EXTENSIONS:
        return None, ""
    material, _method = _classify_material_from_filename(file_path.name, requested_types)
    if material and _is_portrait_material(material) and any(_portrait_filename_identity(file_path.name)):
        return material, "filename_portrait"
    return None, ""


def _flat_path_metadata(file_path: Path) -> tuple[int, int, int | None] | None:
    try:
        stat = file_path.stat()
    except OSError:
        return None
    return stat.st_size, stat.st_mtime_ns, _file_change_token(file_path, stat)


def _lookup_flat_index_cache(
    cache: dict[str, Any],
    file_path: Path,
    rel_path: str,
    cache_stats: dict[str, int],
) -> tuple[dict[str, Any] | None, str | None, tuple[int, float, str] | None]:
    """路径元数据先行、内容 hash 兜底；返回缓存条目、key 与指纹。"""
    entries: dict[str, Any] = cache.setdefault("entries", {})
    paths: dict[str, Any] = cache.setdefault("paths", {})
    metadata = _flat_path_metadata(file_path)
    if metadata is None:
        return None, None, None
    size, mtime_ns, change_token = metadata

    previous = paths.get(rel_path)
    if isinstance(previous, dict) and (
        change_token is not None
        and previous.get("source_size") == size
        and previous.get("source_mtime_ns") == mtime_ns
        and previous.get("source_change_token") == change_token
    ):
        previous_key = previous.get("cache_key")
        entry = entries.get(previous_key)
        if _flat_cache_entry_usable(entry):
            entry["verified_at"] = _beijing_now_str()
            cache_stats["hits"] = cache_stats.get("hits", 0) + 1
            return entry, str(previous_key), None

    fingerprint = _compute_full_file_fingerprint(file_path)
    if fingerprint is None:
        return None, None, None
    fp_size, mtime, sha = fingerprint
    cache_key = f"{sha}_{fp_size}"
    entry = entries.get(cache_key)
    if not _flat_cache_entry_usable(entry):
        legacy_key = f"{sha[:24]}_{fp_size}"
        legacy_entry = entries.get(legacy_key)
        if _flat_cache_entry_usable(legacy_entry):
            entry = dict(legacy_entry)
            entry["content_hash"] = sha
            entries[cache_key] = entry
            entries.pop(legacy_key, None)

    old_key = previous.get("cache_key") if isinstance(previous, dict) else None
    if old_key and old_key != cache_key:
        cache_stats["invalidated"] = cache_stats.get("invalidated", 0) + 1

    paths[rel_path] = {
        "cache_key": cache_key,
        "source_size": size,
        "source_mtime_ns": mtime_ns,
        "source_change_token": change_token,
    }
    if _flat_cache_entry_usable(entry):
        entry["verified_at"] = _beijing_now_str()
        cache_stats["hits"] = cache_stats.get("hits", 0) + 1
        return entry, cache_key, fingerprint

    cache_stats["misses"] = cache_stats.get("misses", 0) + 1
    return None, cache_key, fingerprint


def _store_flat_index_entry(
    cache: dict[str, Any],
    file_path: Path,
    rel_path: str,
    cache_key: str,
    fingerprint: tuple[int, float, str],
    *,
    material_type: str,
    match_method: str,
    subtype: str,
    extracted_names: list[str],
    extracted_id: str,
    extracted_text: str,
) -> dict[str, Any]:
    size, mtime, sha = fingerprint
    entry = {
        "content_hash": sha,
        "source_size": size,
        "source_mtime": mtime,
        "material_type": material_type or "其他材料",
        "match_method": match_method or "unrecognized",
        "subtype": subtype,
        "extracted_name": extracted_names[0] if extracted_names else "",
        "extracted_names": list(dict.fromkeys(extracted_names)),
        "extracted_id_hash": _hash_id_card(extracted_id),
        "extracted_phone_hash": _hash_phone(_extract_phone(extracted_text)),
        "ocr_text": _sanitize_cached_text(extracted_text),
        "verified_at": _beijing_now_str(),
        "sample_filename": file_path.name,
        "source_relpath": rel_path,
        "analysis_state": "complete",
        "index_scope": LIBRARY_MODE_FLAT_OCR,
        "document_page_marker": _footer_page_marker(extracted_text),
    }
    cache.setdefault("entries", {})[cache_key] = entry
    return entry


def _analyze_flat_source(
    file_path: Path,
    requested_types: list[str],
    *,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
) -> tuple[str, str, str, list[str], str, str, bool]:
    """分析单个无序库文件，返回类型、方式、子类型、姓名列表、证件号和正文。"""
    portrait, portrait_method = _named_portrait(file_path, requested_types)
    if portrait:
        return portrait, portrait_method, "", [], "", "", True
    text = _extract_flat_document_text(
        file_path,
        progress_callback=progress_callback,
        cancelled=cancelled,
    )
    method_prefix = "doc_content"
    analysis_complete = True
    names = _extract_person_names(text)
    extracted_id = _extract_id_card(text)
    # 文字层首轮分类禁用“18 位证号→身份证”弱兜底：电子证书的证号常嵌入
    # 身份证号，会抢占真实分类；弱兜底放到视觉识别之后统一执行。
    material, method, subtype = _classify_text_content(
        text,
        requested_types=requested_types,
        method_prefix=method_prefix,
        allow_weak_id_fallback=False,
    )

    ext = file_path.suffix.lower()
    needs_visual_ocr = (
        (ext in IMAGE_EXTENSIONS and not text)
        or (ext == ".pdf" and (not material or not names))
    )
    used_page_render = False
    if needs_visual_ocr:
        used_page_render = (
            ext == ".pdf"
            and pdfium is not None
            and (bool(text) or _PDF_BACKEND == "pdfium")
        )
        (
            ocr_material,
            ocr_method,
            ocr_subtype,
            _name,
            ocr_id,
            ocr_text,
            ocr_names,
            analysis_complete,
        ) = _analyze_ocr_file(
            file_path,
            requested_types=requested_types,
            progress_callback=progress_callback,
            cancelled=cancelled,
            render_pages=used_page_render,
            allow_weak_id_fallback=False,
        )
        if analysis_complete:
            names = list(dict.fromkeys([*names, *ocr_names]))
            extracted_id = extracted_id or ocr_id
            text = " ".join(part for part in (text, ocr_text) if part)
            combined_material, combined_method, combined_subtype = _classify_text_content(
                text,
                requested_types=requested_types,
                method_prefix=method_prefix,
                allow_weak_id_fallback=False,
            )
            if combined_material:
                material, method, subtype = (
                    combined_material,
                    combined_method,
                    combined_subtype,
                )
            elif not material and ocr_material:
                material, method, subtype = ocr_material, ocr_method, ocr_subtype

    if (
        ext == ".pdf"
        and pdfium is not None
        and not used_page_render
        and not material
    ):
        # 内嵌图通道（照片/国徽等局部图）无法定论时，整页渲染二次识别，
        # 让图形标题进入正文参与分类与缓存。
        (
            render_material,
            render_method,
            render_subtype,
            _render_name,
            render_id,
            render_text,
            render_names,
            render_complete,
        ) = _analyze_ocr_file(
            file_path,
            requested_types=requested_types,
            progress_callback=progress_callback,
            cancelled=cancelled,
            render_pages=True,
            allow_weak_id_fallback=False,
        )
        if render_complete:
            names = list(dict.fromkeys([*names, *render_names]))
            extracted_id = extracted_id or render_id
            text = " ".join(part for part in (text, render_text) if part)
            combined_material, combined_method, combined_subtype = _classify_text_content(
                text,
                requested_types=requested_types,
                method_prefix=method_prefix,
                allow_weak_id_fallback=False,
            )
            if combined_material:
                material, method, subtype = (
                    combined_material,
                    combined_method,
                    combined_subtype,
                )
            elif render_material:
                material, method, subtype = render_material, render_method, render_subtype
            analysis_complete = render_complete

    if not material and text:
        weak_material, weak_method, weak_subtype = _classify_text_content(
            text,
            requested_types=requested_types,
            method_prefix=method_prefix,
            allow_weak_id_fallback=True,
        )
        if weak_material:
            material, method, subtype = weak_material, weak_method, weak_subtype

    if not material:
        material, filename_method = _classify_material_from_filename(file_path.name, requested_types)
        if material:
            method = filename_method if not method else f"{method}+{filename_method}"
    return material or "其他材料", method or "unrecognized", subtype, names, extracted_id, text, analysis_complete


def _entry_to_flat_indexed_file(
    file_path: Path,
    rel_path: str,
    cache_key: str,
    entry: dict[str, Any],
    *,
    cache_hit: bool,
    extracted_id_card: str = "",
) -> _FlatIndexedFile:
    raw_names = entry.get("extracted_names")
    if not isinstance(raw_names, list):
        raw_names = [entry.get("extracted_name")] if entry.get("extracted_name") else []
    names = tuple(
        str(name).strip() for name in raw_names
        if str(name or "").strip()
    )
    material = str(entry.get("material_type") or "其他材料")
    photo_names, photo_id, photo_no = (
        _portrait_filename_identity(file_path.name)
        if _is_portrait_material(material) else ((), "", "")
    )
    raw_marker = entry.get("document_page_marker")
    page_marker = None
    if isinstance(raw_marker, (list, tuple)) and len(raw_marker) == 2:
        if isinstance(raw_marker[0], int) and 0 < raw_marker[0] <= _DOCUMENT_GROUP_MAX_PAGES:
            total = raw_marker[1]
            if total is None or isinstance(total, int) and raw_marker[0] <= total <= _DOCUMENT_GROUP_MAX_PAGES:
                page_marker = (raw_marker[0], total)
    return _FlatIndexedFile(
        source_path=file_path,
        relative_path=rel_path,
        cache_key=cache_key,
        material_type=material,
        match_method=str(entry.get("match_method") or "unrecognized"),
        subtype=str(entry.get("subtype") or ""),
        extracted_names=names,
        extracted_id_hash=str(entry.get("extracted_id_hash") or ""),
        extracted_phone_hash=str(entry.get("extracted_phone_hash") or ""),
        text_snippet=str(entry.get("ocr_text") or ""),
        extracted_id_card=extracted_id_card,
        cache_hit=cache_hit,
        filename_names=photo_names,
        filename_id_hash=photo_id,
        filename_employee_no=photo_no,
        ocr_page_marker=page_marker,
    )


def _natural_file_key(path: Path) -> tuple[Any, ...]:
    """按数字自然排序文件名，保证 scan_2 位于 scan_10 之前。"""
    return tuple(
        int(part) if part.isdigit() else part.casefold()
        for part in re.split(r"(\d+)", path.name)
    )


def _filename_page_sequence(path: Path) -> tuple[str, int] | None:
    """提取连续扫描文件名末尾的短序号；身份证号等长数字不会被当页码。"""
    stem = path.stem.casefold()
    if re.fullmatch(r"[0-9a-f]{20,64}", stem):
        return None
    match = re.search(r"(?<!\d)(\d{1,6})$", stem)
    if match is None:
        return None
    prefix = stem[:match.start()].rstrip(" _-.()（）[]【】")
    return prefix, int(match.group(1))


def _footer_page_marker(text: str) -> tuple[int, int | None] | None:
    """单独的尾行数字只作页码线索，不以正文数字或日期充当页码。"""
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    if len(lines) < 2:
        return None
    match = re.fullmatch(r"[-—–]?\s*(\d{1,2})\s*[-—–]?", lines[-1])
    if match and 0 < int(match.group(1)) <= _DOCUMENT_GROUP_MAX_PAGES:
        return int(match.group(1)), None
    return None


def _explicit_page_marker(item: _FlatIndexedFile) -> tuple[int, int | None] | None:
    """从文件名和 OCR 摘要提取“第 N 页”或“N/总页数”标记。"""
    if item.ocr_page_marker is not None:
        return item.ocr_page_marker
    sample = f"{item.source_path.stem} {item.text_snippet}"
    patterns = (
        r"第\s*(\d{1,3})\s*页(?:\s*[/共]\s*(\d{1,3})\s*页?)?",
        r"(?<![\d/.-])(\d{1,3})\s*/\s*(\d{1,3})(?![\d/.-])",
        r"\bpage\s*(\d{1,3})(?:\s*(?:of|/)\s*(\d{1,3}))?\b",
    )
    for pattern in patterns:
        match = re.search(pattern, sample, flags=re.IGNORECASE)
        if match is None:
            continue
        page_number = int(match.group(1))
        total_pages = int(match.group(2)) if match.group(2) else None
        if page_number <= 0 or total_pages is not None and (
            total_pages <= 0 or page_number > total_pages
        ):
            continue
        return page_number, total_pages
    return _footer_page_marker(item.text_snippet)


def _effective_group_material(item: _FlatIndexedFile) -> str:
    """弱身份证兜底不作为跨页分组边界，避免合同证件号抢占页面。"""
    if _is_weak_cached_material_result(
        item.material_type,
        item.match_method,
        item.text_snippet,
    ):
        return ""
    return item.material_type if item.material_type in MATERIAL_SYNONYMS else ""


def _looks_like_document_start(item: _FlatIndexedFile) -> bool:
    material = _effective_group_material(item)
    if material not in _DOCUMENT_GROUPABLE_MATERIALS:
        return False
    marker = _explicit_page_marker(item)
    if marker is not None:
        return marker[0] == 1
    text = item.text_snippet
    compact = re.sub(r"\s+", "", text)
    return compact.startswith(("劳动合同", "用工合同", "劳务合同", "聘用合同"))


def _has_document_end_evidence(item: _FlatIndexedFile) -> bool:
    marker = _explicit_page_marker(item)
    if marker is not None and marker[1] is not None and marker[0] == marker[1]:
        return True
    text = item.text_snippet
    return any(
        keyword in text
        for keyword in (
            "乙方签字", "乙方（签字", "乙方(签字", "劳动者签字",
            "本人签字", "签字盖章", "签署日期", "合同签订日期",
            "乙方（签名", "乙方(签名", "乙方签名",
        )
    )


def _has_contract_continuation_evidence(item: _FlatIndexedFile) -> bool:
    """扫描时间/尺寸属于弱信号，正文还必须具有合同连续页特征。"""
    text = item.text_snippet
    return any(
        keyword in text
        for keyword in (
            "劳动", "合同", "甲方", "乙方", "用人单位", "工作地点",
            "工作内容", "劳动报酬", "工资", "社会保险", "合同期限",
            "违约责任", "劳动争议", "签字", "签署",
        )
    )


def _normalized_group_names(item: _FlatIndexedFile) -> set[str]:
    return {
        normalized
        for normalized in (
            _normalize_person_name(name) for name in item.extracted_names
        )
        if normalized
    }


def _read_image_dimensions(file_path: Path) -> tuple[int, int] | None:
    """只读取图片头部尺寸，不解码完整像素；失败时降级为其他连续性信号。"""
    try:
        from PIL import Image

        with Image.open(file_path) as image:
            width, height = image.size
        if width <= 0 or height <= 0:
            return None
        return int(width), int(height)
    except Exception:
        return None


def _image_group_continuity(
    previous: _FlatIndexedFile,
    current: _FlatIndexedFile,
    *,
    dimensions: dict[Path, tuple[int, int] | None],
    mtimes: dict[Path, float | None],
) -> tuple[str, ...]:
    """返回相邻页面连续性证据；至少命中一个组合信号才进入候选组。"""
    reasons: list[str] = []
    previous_marker = _explicit_page_marker(previous)
    current_marker = _explicit_page_marker(current)
    if (
        previous_marker is not None
        and current_marker is not None
        and current_marker[0] == previous_marker[0] + 1
        and (
            previous_marker[1] is None
            or current_marker[1] is None
            or previous_marker[1] == current_marker[1]
        )
    ):
        reasons.append("page_number")

    previous_sequence = _filename_page_sequence(previous.source_path)
    current_sequence = _filename_page_sequence(current.source_path)
    if (
        previous_sequence is not None
        and current_sequence is not None
        and previous_sequence[0] == current_sequence[0]
        and current_sequence[1] == previous_sequence[1] + 1
    ):
        reasons.append("filename_sequence")

    # 强序号已经与“合同首页 + 明确结束页”共同构成保守门禁，无需再打开
    # 图片读取尺寸；这使常见扫描批次在老电脑上仅做文件名比较。
    if reasons:
        return tuple(reasons)

    def _mtime(item: _FlatIndexedFile) -> float | None:
        if item.source_path not in mtimes:
            try:
                mtimes[item.source_path] = item.source_path.stat().st_mtime
            except OSError:
                mtimes[item.source_path] = None
        return mtimes[item.source_path]

    previous_mtime = _mtime(previous)
    current_mtime = _mtime(current)
    if (
        previous_mtime is not None
        and current_mtime is not None
        and abs(current_mtime - previous_mtime)
        <= _DOCUMENT_GROUP_MAX_TIME_GAP_SECONDS
    ):
        reasons.append("scan_time")

    def _dimensions(item: _FlatIndexedFile) -> tuple[int, int] | None:
        if item.source_path not in dimensions:
            dimensions[item.source_path] = _read_image_dimensions(item.source_path)
        return dimensions[item.source_path]

    previous_dimensions = _dimensions(previous)
    current_dimensions = _dimensions(current)
    if (
        previous_dimensions is not None
        and current_dimensions is not None
        and previous_dimensions == current_dimensions
    ):
        reasons.append("dimensions")

    # 没有页码/连续文件名时，必须同时满足扫描时间和图片尺寸。
    if "scan_time" in reasons and "dimensions" in reasons:
        return tuple(reasons)
    return ()


def _enrich_ordered_document_groups(
    indexed: list[_FlatIndexedFile],
    warnings: list[str],
    requested_types: list[str],
    *,
    cancelled: Callable[[], bool] | None = None,
) -> list[_FlatIndexedFile]:
    """以 O(n log n) 有界窗口合并独立合同图片的组级识别证据。"""
    canonical_requests = {
        _canonical_material_for_request_label(material) or material
        for material in requested_types
        if material
    }
    if not canonical_requests.intersection(_DOCUMENT_GROUPABLE_MATERIALS):
        return indexed
    by_directory: dict[Path, list[_FlatIndexedFile]] = {}
    for item in indexed:
        if item.source_path.suffix.lower() not in IMAGE_EXTENSIONS:
            continue
        by_directory.setdefault(item.source_path.parent, []).append(item)

    replacement_groups: dict[str, list[_FlatIndexedFile]] = {}
    path_to_group: dict[Path, str] = {}

    for directory_items in by_directory.values():
        if not any(_looks_like_document_start(item) for item in directory_items):
            continue
        ordered = sorted(
            directory_items,
            key=lambda item: _natural_file_key(item.source_path),
        )
        position = 0
        while position < len(ordered):
            _raise_if_cancelled(cancelled)
            start = ordered[position]
            if start.source_path in path_to_group or not _looks_like_document_start(start):
                position += 1
                continue

            group = [start]
            group_names = _normalized_group_names(start)
            group_id_hash = start.extracted_id_hash
            group_phone_hash = start.extracted_phone_hash
            closed = _has_document_end_evidence(start)
            conflict_warning = ""
            cursor = position + 1

            while (
                not closed
                and cursor < len(ordered)
                and len(group) < _DOCUMENT_GROUP_MAX_PAGES
            ):
                candidate = ordered[cursor]
                if candidate.source_path in path_to_group:
                    break
                candidate_marker = _explicit_page_marker(candidate)
                if (
                    candidate_marker is not None and candidate_marker[0] == 1
                    or (
                        _looks_like_document_start(candidate)
                        and candidate_marker is None
                    )
                ):
                    # 清晰的新首页就是上一份合同的安全结束边界，不吞入下一份。
                    closed = len(group) > 1
                    break

                candidate_material = _effective_group_material(candidate)
                if candidate_material and candidate_material != start.material_type:
                    break
                candidate_names = _normalized_group_names(candidate)
                if group_names and candidate_names and group_names.isdisjoint(candidate_names):
                    conflict_warning = (
                        "疑似多页劳动合同未自动合并（姓名冲突）："
                        f"{group[0].relative_path} 与 {candidate.relative_path}"
                    )
                    break
                if (
                    group_id_hash
                    and candidate.extracted_id_hash
                    and group_id_hash != candidate.extracted_id_hash
                ):
                    conflict_warning = (
                        "疑似多页劳动合同未自动合并（证件号码冲突）："
                        f"{group[0].relative_path} 与 {candidate.relative_path}"
                    )
                    break
                if (
                    group_phone_hash
                    and candidate.extracted_phone_hash
                    and group_phone_hash != candidate.extracted_phone_hash
                ):
                    conflict_warning = (
                        "疑似多页劳动合同未自动合并（手机号码冲突）："
                        f"{group[0].relative_path} 与 {candidate.relative_path}"
                    )
                    break
                previous_sequence = _filename_page_sequence(group[-1].source_path)
                candidate_sequence = _filename_page_sequence(candidate.source_path)
                if (
                    previous_sequence is None or candidate_sequence is None
                    or previous_sequence[0] != candidate_sequence[0]
                    or candidate_sequence[1] != previous_sequence[1] + 1
                ):
                    break
                previous_marker = _explicit_page_marker(group[-1])
                if (
                    previous_marker and candidate_marker
                    and candidate_marker[0] != previous_marker[0] + 1
                ):
                    break

                group.append(candidate)
                group_names.update(candidate_names)
                group_id_hash = group_id_hash or candidate.extracted_id_hash
                group_phone_hash = group_phone_hash or candidate.extracted_phone_hash
                closed = _has_document_end_evidence(candidate)
                cursor += 1

            if conflict_warning:
                warnings.append(conflict_warning)
            if len(group) > 1 and closed and not conflict_warning:
                group_key = hashlib.sha256(
                    "\0".join(item.relative_path for item in group).encode("utf-8")
                ).hexdigest()[:16]
                combined_text = " ".join(
                    item.text_snippet for item in group if item.text_snippet
                )[:_DOCUMENT_GROUP_TEXT_MAX_CHARS]
                combined_names = tuple(dict.fromkeys(
                    name
                    for item in group
                    for name in item.extracted_names
                    if str(name or "").strip()
                ))
                if not combined_names and not group_id_hash and not group_phone_hash:
                    warnings.append(
                        "已确认多页劳动合同边界，但未识别到可核对的姓名、证件号或手机号，"
                        f"未自动归属员工：{group[0].relative_path}"
                    )
                group_warning = ""
                last_marker = _explicit_page_marker(group[-1])
                if not _has_document_end_evidence(group[-1]) or (
                    last_marker and last_marker[1] and len(group) != last_marker[1]
                ):
                    group_warning = "合同完整性待确认：结束页或总页数不完整"
                if not combined_names and not group_id_hash and not group_phone_hash:
                    group_warning = "合同人员归属待确认：未可靠识别姓名或证件号码"
                enriched: list[_FlatIndexedFile] = []
                for page_number, item in enumerate(group, start=1):
                    method = item.match_method or "unrecognized"
                    if "document_group" not in method:
                        method = f"{method}+document_group"
                    enriched.append(replace(
                        item,
                        material_type=start.material_type,
                        match_method=method,
                        extracted_names=combined_names,
                        extracted_id_hash=group_id_hash,
                        extracted_phone_hash=group_phone_hash,
                        text_snippet=combined_text,
                        document_group_id=group_key,
                        document_page_number=page_number,
                        document_page_count=len(group),
                        document_warning=group_warning,
                    ))
                    path_to_group[item.source_path] = group_key
                replacement_groups[group_key] = enriched
                position += len(group)
                continue
            if len(group) > 1 and not conflict_warning:
                warnings.append(
                    "疑似多页劳动合同未自动合并（无法确认结束页）："
                    f"{group[0].relative_path}；请保留连续页码或签字页标记"
                )
            position += 1

    if not replacement_groups:
        return indexed
    result: list[_FlatIndexedFile] = []
    emitted_groups: set[str] = set()
    for item in indexed:
        group_key = path_to_group.get(item.source_path)
        if group_key is None:
            result.append(item)
            continue
        if group_key not in emitted_groups:
            result.extend(replacement_groups[group_key])
            emitted_groups.add(group_key)
    return result


def _contract_chapters(item: _FlatIndexedFile) -> tuple[int, ...]:
    digits = {char: number for number, char in enumerate("零一二三四五六七八九")}
    def number(text: str) -> int:
        if text.isdigit():
            return int(text)
        if "十" in text:
            left, right = text.split("十", 1)
            return digits.get(left, 1) * 10 + digits.get(right, 0)
        return digits.get(text, 0)
    return tuple(sorted({
        value for token in re.findall(r"第\s*([一二三四五六七八九十零\d]{1,3})\s*[条章]", item.text_snippet)
        if 0 < (value := number(token)) < 100
    }))


def _document_identity_conflicts(left: _FlatIndexedFile, right: _FlatIndexedFile) -> bool:
    if left.extracted_id_hash and right.extracted_id_hash:
        return left.extracted_id_hash != right.extracted_id_hash
    if left.extracted_phone_hash and right.extracted_phone_hash:
        if left.extracted_phone_hash != right.extracted_phone_hash:
            return True
    left_names, right_names = _normalized_group_names(left), _normalized_group_names(right)
    return bool(left_names and right_names and left_names.isdisjoint(right_names))


def _enrich_flat_index_with_document_groups(
    indexed: list[_FlatIndexedFile],
    warnings: list[str],
    requested_types: list[str],
    *,
    cancelled: Callable[[], bool] | None = None,
) -> list[_FlatIndexedFile]:
    """连续文件名走线性分组；无序图片只在最多 64 页的候选块中关联。

    时间/尺寸不决定人员归属。多个合同竞争同一页时整组待确认，避免贪心串人。
    仅使用已有 OCR 摘要，不再次打开图片或调用模型。
    """
    if not any((_canonical_material_for_request_label(t) or t) in _DOCUMENT_GROUPABLE_MATERIALS for t in requested_types):
        return indexed
    blocks: dict[tuple[Path, str], list[_FlatIndexedFile]] = {}
    for position, item in enumerate(indexed):
        if position % 256 == 0:
            _raise_if_cancelled(cancelled)
        if item.source_path.suffix.lower() not in IMAGE_EXTENSIONS:
            continue
        material = _effective_group_material(item)
        if material and material not in _DOCUMENT_GROUPABLE_MATERIALS:
            continue
        sequence = _filename_page_sequence(item.source_path)
        if sequence is None and not (
            _has_contract_continuation_evidence(item) or _explicit_page_marker(item)
        ):
            continue
        prefix = "sequence:" + sequence[0] if sequence else ""
        blocks.setdefault((item.source_path.parent, prefix), []).append(item)

    replacements: dict[Path, _FlatIndexedFile] = {}
    for (_directory, prefix), items in blocks.items():
        _raise_if_cancelled(cancelled)
        if not any(_looks_like_document_start(item) for item in items):
            continue
        if prefix:
            ordered = _enrich_ordered_document_groups(items, warnings, requested_types, cancelled=cancelled)
            for item in ordered:
                if item.document_group_id:
                    replacements[item.source_path] = item
            items = [item for item in items if item.source_path not in replacements]
            if not items:
                continue
        starts = [item for item in items if _looks_like_document_start(item)]
        if not starts:
            continue
        if len(items) > _DOCUMENT_GROUP_MAX_CANDIDATES:
            for item in items:
                replacements[item.source_path] = replace(item, document_warning="合同分组待确认：同目录相似候选过多，请按人员或合同批次分开")
            warnings.append(f"合同分组待确认：{_directory} 有 {len(items)} 张候选，未进行全量两两比较")
            continue
        chapters = {item.source_path: _contract_chapters(item) for item in items}
        markers = {item.source_path: _explicit_page_marker(item) for item in items}
        proposals: list[list[_FlatIndexedFile]] = []
        for start in starts:
            _raise_if_cancelled(cancelled)
            candidates = [start] + [
                item for item in items if item is not start
                and not _looks_like_document_start(item)
                and not _document_identity_conflicts(start, item)
            ]
            if len(candidates) < 2 or len(candidates) > _DOCUMENT_GROUP_MAX_PAGES:
                replacements[start.source_path] = replace(start, document_warning="合同完整性待确认：未确认续页及结束页")
                continue
            if all(chapters[item.source_path] for item in candidates):
                candidates.sort(key=lambda item: chapters[item.source_path][0])
                continuous = all(
                    chapters[left.source_path][-1] < chapters[right.source_path][0]
                    for left, right in zip(candidates, candidates[1:])
                )
            elif all(markers[item.source_path] for item in candidates):
                candidates.sort(key=lambda item: markers[item.source_path][0])
                continuous = all(
                    markers[right.source_path][0] == markers[left.source_path][0] + 1
                    for left, right in zip(candidates, candidates[1:])
                )
            else:
                continuous = False
            if (
                not continuous or candidates[0] is not start
                or not _has_document_end_evidence(candidates[-1])
                or any(_has_document_end_evidence(item) for item in candidates[1:-1])
                or any(_document_identity_conflicts(left, right) for left in candidates for right in candidates)
            ):
                for item in candidates:
                    replacements[item.source_path] = replace(item, document_warning="合同分组待确认：页序、结束位置或人员信息不唯一")
                continue
            proposals.append(candidates)
        uses: dict[Path, int] = {}
        for group in proposals:
            for item in group:
                uses[item.source_path] = uses.get(item.source_path, 0) + 1
        for group in proposals:
            if any(uses[item.source_path] > 1 for item in group):
                for item in group:
                    replacements[item.source_path] = replace(item, document_warning="合同分组待确认：同一页面存在多个可能归属")
                continue
            names = tuple(dict.fromkeys(name for item in group for name in item.extracted_names))
            identity = next((item.extracted_id_hash for item in group if item.extracted_id_hash), "")
            phone = next((item.extracted_phone_hash for item in group if item.extracted_phone_hash), "")
            key = hashlib.sha256("\0".join(item.cache_key for item in group).encode()).hexdigest()[:24]
            page_numbers = [markers[item.source_path] for item in group]
            complete = (
                all(page_numbers) and [marker[0] for marker in page_numbers] == list(range(1, len(group) + 1))
                and page_numbers[-1][1] == len(group)
            )
            warning = "" if complete else "合同完整性待确认：已关联页面，但未识别完整的总页数"
            if not names and not identity and not phone:
                warning = "合同人员归属待确认：未可靠识别姓名或证件号码"
            for page, item in enumerate(group, 1):
                replacements[item.source_path] = replace(
                    item, material_type=group[0].material_type,
                    match_method=item.match_method + "+document_group",
                    extracted_names=names, extracted_id_hash=identity, extracted_phone_hash=phone,
                    document_group_id=key, document_page_number=page, document_page_count=len(group),
                    document_warning=warning,
                )
    replaced = [replacements.get(item.source_path, item) for item in indexed]
    groups = _flat_document_groups(replaced)
    result: list[_FlatIndexedFile] = []
    emitted: set[str] = set()
    for item in replaced:
        if not item.document_group_id:
            result.append(item)
        elif item.document_group_id not in emitted:
            emitted.add(item.document_group_id)
            result.extend(sorted(groups[item.document_group_id], key=lambda page: page.document_page_number))
    pending = sum(bool(item.document_warning) for item in result)
    if pending:
        warnings.append(f"有 {pending} 张合同页面需要确认归属或完整性，详见待确认资料清单")
    return result


def _build_flat_ocr_index(
    lib_path: Path,
    out_path: Path,
    requested_types: list[str],
    cache: dict[str, Any],
    cache_path: Path,
    cache_stats: dict[str, int],
    warnings: list[str],
    progress_callback: Callable[[int, int, str], None] | None,
    cancelled: Callable[[], bool] | None = None,
) -> tuple[list[_FlatIndexedFile], bool]:
    """建立/复用全局 OCR 索引；小批次密集、超大批次分段持久化。"""
    resolved_cache_path = cache_path.resolve()
    source_files = [
        path
        for path in _scan_flat_library_files(
            lib_path,
            skip_dir=out_path,
            cancelled=cancelled,
        )
        if path != resolved_cache_path
    ]
    indexed: list[_FlatIndexedFile] = []
    active_paths: set[str] = set()
    checkpoint_ok = True
    cache_changed_since_checkpoint = False

    for index, source_path in enumerate(source_files, start=1):
        _raise_if_cancelled(cancelled)
        rel_path = source_path.relative_to(lib_path).as_posix()
        active_paths.add(rel_path)
        if progress_callback:
            progress_callback(index, len(source_files), f"正在建立无序资料 OCR 索引：{index}/{len(source_files)}")

        entry, cache_key, fingerprint = _lookup_flat_index_cache(
            cache, source_path, rel_path, cache_stats,
        )
        if cache_key is None:
            warnings.append(f"无法读取资料文件，已跳过：{rel_path}")
            continue

        cache_hit = entry is not None
        portrait, _portrait_method = _named_portrait(source_path, requested_types)
        if entry is not None and portrait and not (
            entry.get("ocr_text") or entry.get("extracted_names") or entry.get("extracted_name")
            or entry.get("extracted_id_hash") or entry.get("extracted_phone_hash")
        ) and entry.get("material_type") in ("其他材料", "证件照片", portrait):
            entry = dict(entry, material_type=portrait, match_method="filename_portrait")
        if entry is not None and entry.get("match_method") == "filename_portrait":
            # 文件名证据不随内容缓存继承；改成随机名或其他材料后需重新分析。
            current_portrait, _ = _named_portrait(source_path, requested_types)
            if current_portrait != entry.get("material_type"):
                entry = None
                cache_hit = False
                cache_stats["hits"] = max(0, cache_stats.get("hits", 0) - 1)
                cache_stats["misses"] = cache_stats.get("misses", 0) + 1
        extracted_id = ""
        if entry is None:
            try:
                material, method, subtype, names, extracted_id, text, analysis_complete = _analyze_flat_source(
                    source_path,
                    requested_types,
                    progress_callback=progress_callback,
                    cancelled=cancelled,
                )
            except MaterialCollectionCancelled:
                raise
            except PDFRecognitionError as exc:
                warnings.append(f"PDF 识别失败，已跳过 {rel_path}：{exc}")
                continue
            if fingerprint is None:
                fingerprint = _compute_full_file_fingerprint(source_path)
            if fingerprint is None:
                warnings.append(f"文件在索引过程中发生变化，已跳过：{rel_path}")
                continue
            if analysis_complete:
                entry = _store_flat_index_entry(
                    cache,
                    source_path,
                    rel_path,
                    cache_key,
                    fingerprint,
                    material_type=material,
                    match_method=method,
                    subtype=subtype,
                    extracted_names=names,
                    extracted_id=extracted_id,
                    extracted_text=text,
                )
                cache_changed_since_checkpoint = True
            else:
                # OCR 引擎不可用/文件暂时无法识别时不能写成永久负结果，下次查询必须重试。
                entry = {
                    "material_type": material,
                    "match_method": method,
                    "subtype": subtype,
                    "extracted_names": names,
                    "extracted_id_hash": _hash_id_card(extracted_id),
                    "analysis_state": "incomplete",
                }
                if not any("暂时无法完成 OCR" in warning for warning in warnings):
                    warnings.append(
                        "部分图片或扫描版 PDF 暂时无法完成 OCR，未写入负缓存；下次查询会自动重试。"
                    )
        elif entry.get("ocr_text"):
            cached_material = str(entry.get("material_type") or "其他材料")
            cached_method = str(entry.get("match_method") or "")
            cached_text = str(entry.get("ocr_text") or "")
            weak_cached_result = _is_weak_cached_material_result(
                cached_material,
                cached_method,
                cached_text,
            )
            matched_request = _requested_label_for_detected_material(
                cached_material, requested_types,
            )
            has_alternative_request = any(
                _requested_label_for_detected_material(
                    cached_material, [requested],
                ) is None
                for requested in requested_types
            )
            should_reclassify = (
                matched_request is None
                or (weak_cached_result and has_alternative_request)
            )
            if not should_reclassify:
                indexed.append(_entry_to_flat_indexed_file(
                    source_path,
                    rel_path,
                    cache_key,
                    entry,
                    cache_hit=cache_hit,
                    extracted_id_card=extracted_id,
                ))
                continue

            # 先用缓存摘要重分类，不重新 OCR；弱结论仍未命中时才尝试整页视觉识别。
            material, method, subtype = _classify_text_content(
                cached_text,
                requested_types=requested_types,
                method_prefix="cached_text",
                allow_weak_id_fallback=False,
            )
            if _requested_label_for_detected_material(
                material or "", requested_types,
            ) is not None:
                entry["material_type"] = material
                entry["match_method"] = method
                entry["subtype"] = subtype
                cache_changed_since_checkpoint = True
            elif (
                source_path.suffix.lower() == ".pdf"
                and pdfium is not None
                and weak_cached_result
                and not _visual_ocr_query_was_attempted(entry, requested_types)
            ):
                # 缓存摘要满足不了当前请求时，整页渲染并记录本次查询签名，
                # 同一查询下次不再重复整本 OCR。
                (
                    fresh_material,
                    fresh_method,
                    fresh_subtype,
                    fresh_names,
                    fresh_id,
                    fresh_text,
                    fresh_complete,
                ) = _analyze_flat_source(
                    source_path,
                    requested_types,
                    progress_callback=progress_callback,
                    cancelled=cancelled,
                )
                cache_hit = False
                cache_stats["hits"] = max(0, cache_stats.get("hits", 0) - 1)
                cache_stats["misses"] = cache_stats.get("misses", 0) + 1
                if fresh_complete:
                    entry["ocr_text"] = _sanitize_cached_text(fresh_text)
                    entry["extracted_id_hash"] = _hash_id_card(fresh_id)
                    entry["extracted_phone_hash"] = _hash_phone(
                        _extract_phone(fresh_text)
                    )
                    if fresh_names:
                        entry["extracted_names"] = list(fresh_names)
                    _record_visual_ocr_query(entry, requested_types)
                    if _requested_label_for_detected_material(
                        fresh_material, requested_types,
                    ) is not None:
                        entry["material_type"] = fresh_material
                        entry["match_method"] = fresh_method
                        entry["subtype"] = fresh_subtype
                    cache_changed_since_checkpoint = True

        indexed.append(_entry_to_flat_indexed_file(
            source_path,
            rel_path,
            cache_key,
            entry,
            cache_hit=cache_hit,
            extracted_id_card=extracted_id,
        ))

        should_checkpoint = (
            (index <= 1000 and index % 100 == 0)
            or (index > 1000 and index % 1000 == 0)
        )
        if should_checkpoint and cache_changed_since_checkpoint:
            _trim_cache_by_age_and_size(cache)
            if not _save_ocr_cache(cache_path, cache):
                checkpoint_ok = False
            cache_changed_since_checkpoint = False

    # 移除已经删除/改名的旧路径；内容条目仅在没有任何现存路径引用时清理。
    paths: dict[str, Any] = cache.setdefault("paths", {})
    for rel_path in list(paths):
        if rel_path not in active_paths:
            paths.pop(rel_path, None)
            cache_changed_since_checkpoint = True
    referenced_keys = {
        item.get("cache_key") for item in paths.values()
        if isinstance(item, dict) and item.get("cache_key")
    }
    entries: dict[str, Any] = cache.setdefault("entries", {})
    for key, entry in list(entries.items()):
        if (
            isinstance(entry, dict)
            and entry.get("index_scope") == LIBRARY_MODE_FLAT_OCR
            and key not in referenced_keys
        ):
            entries.pop(key, None)
            cache_changed_since_checkpoint = True

    if cache_changed_since_checkpoint:
        _trim_cache_by_age_and_size(cache)
        if not _save_ocr_cache(cache_path, cache):
            checkpoint_ok = False
    return _enrich_flat_index_with_document_groups(
        indexed, warnings, requested_types, cancelled=cancelled,
    ), checkpoint_ok


def _flat_file_matches_employee(
    indexed_file: _FlatIndexedFile,
    employee: TargetEmployee,
    duplicate_target_names: set[str],
) -> bool:
    if indexed_file.confirmed_employee_key:
        return indexed_file.confirmed_employee_key == _review_employee_key(employee)
    target_name = _normalize_person_name(employee.name)
    target_id_hash = _hash_id_card(employee.id_card)
    target_phone_hash = _hash_phone(employee.phone)
    if not target_id_hash and _ID_CARD_RE.fullmatch(employee.name.strip()):
        target_id_hash = _hash_id_card(employee.name.strip())

    if target_id_hash and indexed_file.extracted_id_hash:
        return target_id_hash == indexed_file.extracted_id_hash
    if target_phone_hash and indexed_file.extracted_phone_hash:
        return target_phone_hash == indexed_file.extracted_phone_hash
    if _is_portrait_material(indexed_file.material_type) and not (
        indexed_file.extracted_names or indexed_file.extracted_id_hash or indexed_file.extracted_phone_hash
    ):
        if indexed_file.filename_names and indexed_file.filename_names != (target_name,):
            return False
        if indexed_file.filename_id_hash:
            return bool(target_id_hash and target_id_hash == indexed_file.filename_id_hash)
        if indexed_file.filename_employee_no:
            return bool(employee.employee_no and employee.employee_no == indexed_file.filename_employee_no)
        return (
            target_name not in duplicate_target_names
            and indexed_file.filename_names == (target_name,)
        )
    if not target_name or target_name in duplicate_target_names:
        return False
    names = {_normalize_person_name(name) for name in indexed_file.extracted_names}
    if target_name in names:
        return True

    # 部分证书只印姓名本身而没有“姓名：”标签；以非中文边界做全文精确匹配，
    # 避免把“张三”错误命中“张三丰”；照片文件名由上方独立分支处理。
    raw_target = str(employee.name or "").strip()
    if raw_target and indexed_file.text_snippet:
        return bool(re.search(
            rf"(?<![\u3400-\u9fff]){re.escape(raw_target)}(?![\u3400-\u9fff])",
            indexed_file.text_snippet,
        ))
    return False


class _FlatEmployeeCandidateIndex:
    """人员倒排索引；文件名、正文和人工确认均由最终谓词核对。"""

    __slots__ = ("_indexed_files", "_candidate_positions")

    def __init__(
        self,
        indexed_files: list[_FlatIndexedFile],
        employees: list[TargetEmployee],
        duplicate_target_names: set[str],
        *,
        cancelled: Callable[[], bool] | None = None,
    ) -> None:
        id_targets: dict[str, set[int]] = {}
        phone_targets: dict[str, set[int]] = {}
        number_targets: dict[str, set[int]] = {}
        confirmed_targets: dict[str, set[int]] = {}
        name_targets: dict[str, set[int]] = {}
        text_targets: dict[str, set[int]] = {}

        for employee_index, employee in enumerate(employees):
            confirmed_targets.setdefault(_review_employee_key(employee), set()).add(employee_index)
            if employee.employee_no:
                number_targets.setdefault(employee.employee_no, set()).add(employee_index)
            target_id_hash = _hash_id_card(employee.id_card)
            if not target_id_hash and _ID_CARD_RE.fullmatch(employee.name.strip()):
                target_id_hash = _hash_id_card(employee.name.strip())
            if target_id_hash:
                id_targets.setdefault(target_id_hash, set()).add(employee_index)

            target_phone_hash = _hash_phone(employee.phone)
            if target_phone_hash:
                phone_targets.setdefault(target_phone_hash, set()).add(employee_index)

            target_name = _normalize_person_name(employee.name)
            if target_name and target_name not in duplicate_target_names:
                name_targets.setdefault(target_name, set()).add(employee_index)
                raw_target = str(employee.name or "").strip()
                if raw_target:
                    text_targets.setdefault(raw_target, set()).add(employee_index)

        text_matcher = _AhoCandidateMatcher(text_targets)
        self._indexed_files = indexed_files
        self._candidate_positions = [array("I") for _ in employees]
        for file_position, indexed_file in enumerate(indexed_files):
            if file_position % 256 == 0:
                _raise_if_cancelled(cancelled)
            candidate_employees: set[int] = set()
            if indexed_file.confirmed_employee_key:
                candidate_employees.update(confirmed_targets.get(indexed_file.confirmed_employee_key, ()))
            if _is_portrait_material(indexed_file.material_type):
                if indexed_file.filename_id_hash:
                    candidate_employees.update(id_targets.get(indexed_file.filename_id_hash, ()))
                if indexed_file.filename_employee_no:
                    positions = number_targets.get(indexed_file.filename_employee_no, ())
                    if len(positions) == 1:
                        candidate_employees.update(positions)
                for name in indexed_file.filename_names:
                    candidate_employees.update(name_targets.get(name, ()))
            if indexed_file.extracted_id_hash:
                candidate_employees.update(
                    id_targets.get(indexed_file.extracted_id_hash, ())
                )
            if indexed_file.extracted_phone_hash:
                candidate_employees.update(
                    phone_targets.get(indexed_file.extracted_phone_hash, ())
                )
            for extracted_name in indexed_file.extracted_names:
                normalized_name = _normalize_person_name(extracted_name)
                if normalized_name:
                    candidate_employees.update(name_targets.get(normalized_name, ()))
            if indexed_file.text_snippet:
                candidate_employees.update(text_matcher.match_payloads(
                    indexed_file.text_snippet,
                    cjk_boundaries=True,
                ))
            for employee_index in candidate_employees:
                self._candidate_positions[employee_index].append(file_position)

    def files_for(
        self,
        employee_index: int,
        employee: TargetEmployee,
        duplicate_target_names: set[str],
    ) -> list[_FlatIndexedFile]:
        return [
            indexed_file
            for file_position in self._candidate_positions[employee_index]
            if _flat_file_matches_employee(
                indexed_file := self._indexed_files[file_position],
                employee,
                duplicate_target_names,
            )
        ]


def _flat_document_groups(
    indexed_files: list[_FlatIndexedFile],
) -> dict[str, list[_FlatIndexedFile]]:
    groups: dict[str, list[_FlatIndexedFile]] = {}
    for indexed_file in indexed_files:
        if indexed_file.document_group_id:
            groups.setdefault(indexed_file.document_group_id, []).append(indexed_file)
    return groups


def _review_employee_key(employee: TargetEmployee) -> str:
    return hashlib.sha256((employee.identity_key + "_" + employee.employee_no).encode()).hexdigest()


def _review_source_hash(item: _FlatIndexedFile) -> str | None:
    match = re.fullmatch(r"([0-9a-f]{64})_\d+", item.cache_key)
    return match.group(1) if match else _stream_full_sha256(item.source_path)


def _load_document_reviews(
    cache: dict[str, Any] | None, employees: list[TargetEmployee], warnings: list[str],
    *, cancelled: Callable[[], bool] | None = None,
) -> None:
    """每次任务只读取一次用户填写的清单；未明确填写“是”不构成确认。"""
    if not cache or not isinstance(cache.get("review_workbook"), str):
        return
    path = Path(cache["review_workbook"])
    try:
        if not path.is_file() or path.stat().st_size > 8 * 1024 * 1024:
            return
        by_name: dict[str, list[TargetEmployee]] = {}
        for employee in employees:
            by_name.setdefault(employee.name, []).append(employee)
        choices: dict[str, set[str]] = {}
        with path.open("rb") as stream:
            workbook = load_workbook(stream, read_only=True, data_only=False)
            try:
                if "待确认归属" not in workbook.sheetnames:
                    return
                for index, row in enumerate(workbook["待确认归属"].iter_rows(min_row=2, max_col=7, values_only=True)):
                    if index % 128 == 0:
                        _raise_if_cancelled(cancelled)
                    if index >= 20000:
                        warnings.append("待确认清单超过 20000 行，本次仅处理前 20000 行")
                        break
                    group, name, qualifier, confirmed = (str(value or "").strip() for value in row[:4])
                    if confirmed != "是" or not name or name.startswith("="):
                        continue
                    candidates = by_name.get(name, [])
                    if qualifier:
                        candidates = [e for e in candidates if qualifier in (e.employee_no, e.id_card[-4:] if e.id_card else "")]
                    if len(candidates) != 1:
                        warnings.append(f"待确认清单中的人员无法唯一对应当前名单，未应用：{group}")
                        continue
                    choices.setdefault(group, set()).add(_review_employee_key(candidates[0]))
            finally:
                workbook.close()
        approved = cache.setdefault("approved_document_groups", {})
        if not isinstance(approved, dict):
            return
        sources = cache.get("review_sources") or {}
        for group, people in choices.items():
            if len(people) != 1 or group not in sources:
                warnings.append(f"待确认清单存在相互冲突或失效的选择，未应用：{group}")
                continue
            approved[group] = {"employee": next(iter(people)), "sources": sources[group]}
    except MaterialCollectionCancelled:
        raise
    except Exception as exc:
        warnings.append(f"待确认清单无法读取，保留待确认状态：{exc}")


def _apply_document_reviews(
    indexed: list[_FlatIndexedFile], cache: dict[str, Any] | None,
    employees: list[TargetEmployee], warnings: list[str],
    *, cancelled: Callable[[], bool] | None = None,
) -> list[_FlatIndexedFile]:
    approved = (cache or {}).get("approved_document_groups") or {}
    if not isinstance(approved, dict) or not approved:
        return indexed
    people = {_review_employee_key(employee): employee for employee in employees}
    groups: dict[str, list[_FlatIndexedFile]] = {}
    for item in indexed:
        key = item.document_group_id or hashlib.sha256(str(item.source_path).encode()).hexdigest()[:24]
        groups.setdefault(key, []).append(item)
    replacements: dict[Path, _FlatIndexedFile] = {}
    for key, group in groups.items():
        _raise_if_cancelled(cancelled)
        record = approved.get(key)
        if not isinstance(record, dict) or record.get("employee") not in people:
            continue
        sources = record.get("sources") or []
        expected = {source.get("path"): source.get("sha256") for source in sources if isinstance(source, dict)}
        if set(expected) != {str(item.source_path) for item in group}:
            continue
        if any(_review_source_hash(item) != expected[str(item.source_path)] for item in group):
            warnings.append("确认过的资料内容已变化，确认记录失效，请重新核对")
            approved.pop(key, None)
            continue
        employee = people[record["employee"]]
        for item in group:
            replacements[item.source_path] = replace(
                item, extracted_names=(employee.name,), extracted_id_hash=_hash_id_card(employee.id_card),
                extracted_phone_hash=_hash_phone(employee.phone), document_warning="",
                match_method=item.match_method + "+manual_review",
                material_type="劳动合同" if item.document_warning.startswith("合同") else item.material_type,
                confirmed_employee_key=_review_employee_key(employee),
            )
    return [replacements.get(item.source_path, item) for item in indexed]


def _write_document_review(
    output_dir: Path, indexed: list[_FlatIndexedFile], cache: dict[str, Any] | None,
    warnings: list[str], *, cancelled: Callable[[], bool] | None = None,
) -> Path | None:
    """清单可编辑确认；预览按组加载原图，避免 Win7 一次解码几千张图片。"""
    if not indexed:
        return None
    path = output_dir / "资料待确认.xlsx"
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("待确认归属")
    for column, width in (("A", 29), ("B", 16), ("C", 20), ("D", 29), ("E", 8), ("F", 75), ("G", 65)):
        sheet.column_dimensions[column].width = width
    sheet.column_dimensions["C"].number_format = "@"
    sheet.append(["确认编号", "确认员工", "工号或证件尾号", "已核对归属及完整性（填是）", "页序", "源文件", "待确认原因"])
    groups: dict[str, list[_FlatIndexedFile]] = {}
    for item in indexed:
        key = item.document_group_id or hashlib.sha256(str(item.source_path).encode()).hexdigest()[:24]
        groups.setdefault(key, []).append(item)
    source_records: dict[str, list[dict[str, str]]] = {}
    previews = []
    for key, group in groups.items():
        _raise_if_cancelled(cancelled)
        source_records[key] = []
        urls = []
        for item in group:
            _raise_if_cancelled(cancelled)
            digest = _review_source_hash(item)
            if digest is None:
                continue
            source_records[key].append({"path": str(item.source_path), "sha256": digest})
            sheet.append([key, "", "", "", item.document_page_number or "", str(item.source_path), item.document_warning])
            urls.append({"url": item.source_path.as_uri(), "name": item.source_path.name})
        previews.append({"key": key, "files": urls, "reason": group[0].document_warning})
    workbook.save(path)
    workbook.close()
    if cache is not None:
        cache["review_workbook"] = str(path)
        cache["review_sources"] = source_records
    payload = json.dumps(previews, ensure_ascii=True).replace("<", "\\u003c")
    preview = output_dir / "资料待确认预览.html"
    preview.write_text(
        '<!doctype html><meta charset="utf-8"><title>资料待确认</title>'
        '<style>body{font:16px sans-serif;margin:24px}button{display:block;margin:8px 0;padding:8px}img{max-width:320px;max-height:500px;margin:8px;vertical-align:top}</style>'
        '<h2>资料待确认</h2><p>点击分组查看原图。核对后在资料待确认.xlsx 中填写确认员工、必要的工号或证件尾号，并在确认列填“是”。同组填写一行即可。保存后重新运行资料打包，系统核对原文件未变后复用确认。</p>'
        '<p>只填写单页确认编号时仅确认该页，不会将其他待确认页面一起归属。</p>'
        '<div id="groups"></div><div id="pages"></div><script>var data=' + payload + ';'
        'function show(i){var box=document.getElementById("pages");box.innerHTML="";'
        'var p=document.createElement("p");p.textContent=data[i].key+"："+data[i].reason;box.appendChild(p);'
        'for(var j=0;j<data[i].files.length;j++){var a=document.createElement("a");a.href=data[i].files[j].url;a.target="_blank";'
        'var im=document.createElement("img");im.src=a.href;im.alt=data[i].files[j].name;a.appendChild(im);box.appendChild(a);}}'
        'for(var i=0;i<data.length;i++){var b=document.createElement("button");b.textContent=data[i].key+"（"+data[i].files.length+"页）";'
        '(function(n){b.onclick=function(){show(n);};})(i);document.getElementById("groups").appendChild(b);}</script>',
        encoding="utf-8",
    )
    reuse_hint = "填写并保存确认后，重新运行可复用。" if cache is not None else "当前关闭缓存，清单仅供核对；启用缓存后生成的清单才可复用确认。"
    warnings.append(f"待确认清单：{path}；原图预览：{preview}。{reuse_hint}")
    return path


def _flat_employee_result_key(
    employee: TargetEmployee,
    duplicate_target_names: set[str],
) -> str:
    """同名人员的输出/缺件键使用非敏感限定符，防止文件互相覆盖。"""
    if _normalize_person_name(employee.name) not in duplicate_target_names:
        return employee.name
    if employee.employee_no:
        qualifier = f"工号{employee.employee_no}"
    elif employee.id_card:
        qualifier = f"证件尾号{employee.id_card[-4:]}"
    elif employee.phone:
        qualifier = f"手机尾号{employee.phone[-4:]}"
    else:
        qualifier = hashlib.sha256(employee.identity_key.encode("utf-8")).hexdigest()[:6]
    return f"{employee.name}（{qualifier}）"


def _unique_destination(path: Path) -> Path:
    if not path.exists():
        return path
    sequence = 2
    while True:
        candidate = path.with_name(f"{path.stem}_{sequence}{path.suffix}")
        if not candidate.exists():
            return candidate
        sequence += 1


def _collect_from_flat_index(
    employee: TargetEmployee,
    indexed_files: list[_FlatIndexedFile],
    out_path: Path,
    mode: str,
    requested_materials: list[str] | None,
    matches: list[MaterialFileMatch],
    warnings: list[str],
    duplicate_target_names: set[str],
    *,
    candidate_files: list[_FlatIndexedFile] | None = None,
    document_groups: dict[str, list[_FlatIndexedFile]] | None = None,
    cancelled: Callable[[], bool] | None = None,
) -> tuple[list[str], int]:
    person_files = candidate_files if candidate_files is not None else [
        item for item in indexed_files
        if _flat_file_matches_employee(item, employee, duplicate_target_names)
    ]
    verified_group_fingerprints: dict[Path, tuple[int, float, str]] = {}
    relevant_group_ids = {
        item.document_group_id for item in person_files if item.document_group_id
    }
    invalid_group_ids: set[str] = set()
    if relevant_group_ids:
        all_group_members = document_groups or _flat_document_groups(indexed_files)
        for group_id, members in all_group_members.items():
            if group_id not in relevant_group_ids:
                continue
            for member in members:
                _raise_if_cancelled(cancelled)
                fingerprint = _compute_full_file_fingerprint(member.source_path)
                if fingerprint is not None:
                    verified_group_fingerprints[member.source_path] = fingerprint
                if (
                    fingerprint is None
                    or f"{fingerprint[2]}_{fingerprint[0]}" != member.cache_key
                ):
                    invalid_group_ids.add(group_id)
                    warnings.append(
                        "多页文档中的文件在索引后发生变化，整组已跳过并请重新运行："
                        f"{member.relative_path}"
                    )
                    break
        if invalid_group_ids:
            person_files = [
                item for item in person_files
                if item.document_group_id not in invalid_group_ids
            ]
    person_file_count = len(person_files)
    if requested_materials is not None:
        person_files_with_material = [
            (item, _requested_label_for_detected_material(
                item.material_type, requested_materials,
            ))
            for item in person_files
        ]
        person_files_with_material = [
            (item, material)
            for item, material in person_files_with_material
            if material is not None
        ]
    else:
        person_files_with_material = [
            (item, item.material_type or "其他材料")
            for item in person_files
        ]

    material_counts: dict[str, int] = {}
    for _item, material in person_files_with_material:
        material_counts[material] = material_counts.get(material, 0) + 1
    material_sequences: dict[str, int] = {}
    copied_materials: set[str] = set()
    copied_count = 0
    clean_employee = safe_filename(
        _flat_employee_result_key(employee, duplicate_target_names)
    )

    for item, material in person_files_with_material:
        _raise_if_cancelled(cancelled)
        clean_material = safe_filename(material)
        material_sequences[material] = material_sequences.get(material, 0) + 1
        sequence = material_sequences[material]
        extension = item.source_path.suffix

        if requested_materials is None and mode == MODE_BY_EMPLOYEE:
            destination_dir = out_path / clean_employee
            target_name = item.source_path.name
        else:
            if item.subtype:
                suffix = f"_{item.subtype}"
                if material_counts[material] > 1 and sequence > 1:
                    suffix += f"_{sequence}"
            else:
                suffix = f"_{sequence}" if material_counts[material] > 1 else ""
            target_name = f"{clean_employee}_{clean_material}{suffix}{extension}"
            if mode == MODE_BY_EMPLOYEE:
                destination_dir = out_path / clean_employee
            elif mode == MODE_BY_MATERIAL:
                destination_dir = out_path / clean_material
            else:
                destination_dir = out_path

        destination_dir.mkdir(parents=True, exist_ok=True)
        destination = _unique_destination(destination_dir / target_name)
        current_fingerprint = verified_group_fingerprints.get(item.source_path)
        if current_fingerprint is None:
            current_fingerprint = _compute_full_file_fingerprint(item.source_path)
        if current_fingerprint is None:
            warnings.append(f"资料文件在检索后无法读取，已跳过：{item.relative_path}")
            continue
        current_size, _current_mtime, current_sha = current_fingerprint
        if f"{current_sha}_{current_size}" != item.cache_key:
            warnings.append(f"资料文件在索引后发生变化，已跳过并请重新运行：{item.relative_path}")
            continue
        try:
            shutil.copy2(item.source_path, destination)
        except Exception as exc:
            try:
                destination.unlink(missing_ok=True)
            except OSError:
                pass
            warnings.append(f"复制失败：{item.source_path} → {destination}: {exc}")
            continue

        copied_fingerprint = _compute_full_file_fingerprint(destination)
        if (
            copied_fingerprint is None
            or f"{copied_fingerprint[2]}_{copied_fingerprint[0]}" != item.cache_key
        ):
            try:
                destination.unlink(missing_ok=True)
            except OSError:
                pass
            warnings.append(f"资料文件复制后校验不一致，已移除结果并请重新运行：{item.relative_path}")
            continue

        matches.append(MaterialFileMatch(
            employee_name=employee.name,
            material_type=material,
            source_path=item.source_path,
            relative_source_path=item.relative_path,
            matched_by=item.match_method,
            target_filename=destination.name,
            target_path=destination,
            extracted_person_name=employee.name,
            extracted_id_card=item.extracted_id_card,
            mismatch_warning=item.document_warning,
            cache_hit=item.cache_hit,
            employee_identity_key=employee.identity_key,
        ))
        copied_materials.add(material)
        copied_count += 1

    if requested_materials is None:
        return ([] if copied_count else ["（OCR 未识别到该人员资料）"], person_file_count)
    return (
        [material for material in requested_materials if material not in copied_materials],
        person_file_count,
    )


# ---------------------------------------------------------------------------
# 核心：收集员工资料
# ---------------------------------------------------------------------------

def collect_employee_materials(
    library_dir: str | Path,
    output_dir: str | Path,
    *,
    roster_source: str | Path | list[dict[str, Any]] | list[str],
    material_types: list[str] | None = None,
    mode: str = MODE_BY_EMPLOYEE,
    library_mode: str = LIBRARY_MODE_PERSON_FOLDER,
    create_zip: bool = False,
    generate_report: bool = True,
    collect_all: bool = False,
    scan_depth: int = 1,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
    use_ocr_cache: bool = True,
    ocr_cache_path: Path | str | None = None,
) -> MaterialCollectResult:
    """Search, match, extract, and package employee materials from the repository."""
    _raise_if_cancelled(cancelled)
    lib_path = Path(library_dir).expanduser().resolve()
    out_path = Path(output_dir).expanduser().resolve()

    if not lib_path.exists() or not lib_path.is_dir():
        raise FileNotFoundError(f"资料库目录不存在：{lib_path}")

    # P0 防递归死循环：输出目录严禁处于资料库内部
    if _is_path_nested(out_path, lib_path):
        raise ValueError(
            f"保存目录不能在资料库目录内部（会导致循环嵌套复制）：\n"
            f"资料库：{lib_path}\n"
            f"保存目录：{out_path}\n"
            f"请选择一个位于资料库外部的独立文件夹作为保存目录。"
        )

    if mode not in MODES:
        raise ValueError(f"不支持的归类模式：{mode}，可选值：{MODES}")
    if library_mode not in LIBRARY_MODES:
        raise ValueError(f"不支持的资料库形式：{library_mode}，可选值：{LIBRARY_MODES}")
    if library_mode == LIBRARY_MODE_FLAT_OCR and not use_ocr_cache:
        raise ValueError("无序平铺资料库必须启用 OCR 索引缓存，避免每次查询重复识别全部文件")

    employees = parse_employee_roster(roster_source)
    if not employees:
        raise ValueError("未能解析出有效的员工信息，请输入员工姓名/身份证，或上传员工名单表格")

    if material_types is None or len(material_types) == 0:
        global_materials = list(MATERIAL_SYNONYMS.keys())
    else:
        global_materials = list(material_types)

    skip_folder_all_ocr = (
        collect_all
        and library_mode == LIBRARY_MODE_PERSON_FOLDER
        and _is_direct_employee_input(roster_source)
    )
    effective_use_ocr_cache = use_ocr_cache and not skip_folder_all_ocr

    out_path.mkdir(parents=True, exist_ok=True)

    # === OCR 智能索引缓存层：启动期加载 / 引擎升级全量失效 / 只读目录降级 ===
    ocr_cache: dict[str, Any] | None = None
    cache_stats: dict[str, int] = {"hits": 0, "misses": 0, "invalidated": 0}
    cache_path: Path | None = None
    cache_skipped_reason: str | None = None
    cache_write_ok = True
    current_signature = ""
    legacy_pdf_cache_invalidated = 0

    if effective_use_ocr_cache:
        current_signature = _get_engine_signature()
        cache_path = (
            Path(ocr_cache_path).expanduser().resolve()
            if ocr_cache_path is not None
            else (lib_path / _OCR_CACHE_FILE_NAME)
        )
        ocr_cache = _load_ocr_cache(cache_path)
        legacy_pdf_cache_invalidated = _invalidate_legacy_pdf_cache_entries(ocr_cache)
        legacy_pdf_cache_invalidated += _invalidate_changed_pdf_backend_entries(ocr_cache)
        cache_stats["invalidated"] += legacy_pdf_cache_invalidated
        # 引擎升级 → 视为全量失效（彻底重写 entries）
        prev_sig = ocr_cache.get("engine_signature")
        if prev_sig and prev_sig != current_signature:
            ocr_cache["entries"] = {}
            ocr_cache["paths"] = {}
    elif skip_folder_all_ocr:
        cache_skipped_reason = "全部材料按人员文件夹直接复制，无需 OCR 或 OCR 缓存"
    else:
        cache_skipped_reason = "用户已关闭 OCR 识别缓存"

    total_steps = len(employees)
    matches: list[MaterialFileMatch] = []
    missing_records: dict[str, list[str]] = {}
    warnings: list[str] = []
    folder_match_counts: dict[str, int] = {}
    employee_result_keys: list[str] = []
    review_candidates: list[_FlatIndexedFile] = []
    _load_document_reviews(ocr_cache, employees, warnings, cancelled=cancelled)

    if legacy_pdf_cache_invalidated:
        warnings.append(
            f"PDF 识别能力已升级，已安全失效 {legacy_pdf_cache_invalidated} 条旧 PDF 缓存；"
            "图片和其他文档缓存保持不变。"
        )

    # 引擎升级警告：本轮第一次识别时输出一次即可
    if (
        effective_use_ocr_cache
        and ocr_cache is not None
        and ocr_cache.get("engine_signature")
        and ocr_cache.get("engine_signature") != current_signature
    ):
        warnings.append(
            f"OCR 引擎版本变更（{ocr_cache.get('engine_signature')} → {current_signature}），"
            "缓存已全量失效，本次将重新 OCR 识别所有图片。"
        )

    folder_index: dict[str, list[Path]] = {}
    flat_index: list[_FlatIndexedFile] = []
    folder_candidate_index: _FolderEmployeeCandidateIndex | None = None
    flat_candidate_index: _FlatEmployeeCandidateIndex | None = None
    flat_document_groups: dict[str, list[_FlatIndexedFile]] = {}
    duplicate_target_names: set[str] = set()
    if library_mode == LIBRARY_MODE_PERSON_FOLDER:
        _raise_if_cancelled(cancelled)
        if progress_callback:
            progress_callback(0, len(employees), "正在扫描资料库文件夹索引...")
        folder_index = _scan_folder_index(lib_path, max_depth=scan_depth, skip_dir=out_path)
        folder_candidate_index = _FolderEmployeeCandidateIndex(
            folder_index,
            employees,
            cancelled=cancelled,
        )
    else:
        assert ocr_cache is not None and cache_path is not None
        normalized_name_counts: dict[str, int] = {}
        for employee in employees:
            normalized_name = _normalize_person_name(employee.name)
            normalized_name_counts[normalized_name] = normalized_name_counts.get(normalized_name, 0) + 1
        duplicate_target_names = {
            name for name, count in normalized_name_counts.items() if name and count > 1
        }
        if duplicate_target_names:
            warnings.append(
                "名单中存在同名人员；无身份证号可核对的同名资料不会自动归属，请补充身份证号后重试。"
            )
        # 只传真实请求：标准分类不依赖 requested，传全量标准名并集反而会让
        # “缓存类型 ∉ 请求”的重分类/渲染重分析条件永远不成立。
        flat_index, checkpoint_ok = _build_flat_ocr_index(
            lib_path,
            out_path,
            list(dict.fromkeys(global_materials)),
            ocr_cache,
            cache_path,
            cache_stats,
            warnings,
            progress_callback,
            cancelled,
        )
        if not checkpoint_ok:
            cache_write_ok = False
            cache_skipped_reason = "资料库目录只读或无写入权限"
            warnings.append(
                f"OCR 索引缓存写入失败：{cache_path}；本次仍使用内存索引完成检索，下次会重新建立。"
            )
        flat_index = _apply_document_reviews(flat_index, ocr_cache, employees, warnings, cancelled=cancelled)
        for item in flat_index:
            if item.document_warning:
                review_candidates.append(item)
            elif _is_portrait_material(item.material_type) and not (
                item.extracted_names or item.extracted_id_hash or item.extracted_phone_hash
                or item.filename_names or item.filename_id_hash or item.filename_employee_no
            ):
                review_candidates.append(replace(item, document_warning="照片人员归属待确认：名称只有材料类型，缺少人员依据"))
        flat_candidate_index = _FlatEmployeeCandidateIndex(
            flat_index,
            employees,
            duplicate_target_names,
            cancelled=cancelled,
        )
        flat_document_groups = _flat_document_groups(flat_index)

    for idx, emp in enumerate(employees):
        _raise_if_cancelled(cancelled)
        emp_key = (
            _flat_employee_result_key(emp, duplicate_target_names)
            if library_mode == LIBRARY_MODE_FLAT_OCR
            else emp.name
        )
        employee_result_keys.append(emp_key)
        employee_key = _build_employee_key(emp)
        if progress_callback:
            progress_callback(
                idx + 1, total_steps,
                f"[{idx + 1}/{total_steps}] 正在检索与匹配：{emp.name}"
                + (f"（缓存命中 {cache_stats['hits']}）" if cache_stats["hits"] else ""),
            )

        if collect_all:
            emp_materials: list[str] | None = None
        elif emp.per_person_materials:
            emp_materials = list(emp.per_person_materials)
        else:
            emp_materials = global_materials

        if library_mode == LIBRARY_MODE_FLAT_OCR:
            assert flat_candidate_index is not None
            emp_missing, matched_count = _collect_from_flat_index(
                emp,
                flat_index,
                out_path,
                mode,
                emp_materials,
                matches,
                warnings,
                duplicate_target_names,
                candidate_files=flat_candidate_index.files_for(
                    idx,
                    emp,
                    duplicate_target_names,
                ),
                document_groups=flat_document_groups,
                cancelled=cancelled,
            )
            folder_match_counts[emp_key] = 1 if matched_count else 0
            if emp_missing:
                missing_records[emp_key] = emp_missing
            continue

        assert folder_candidate_index is not None
        matched_folders = folder_candidate_index.matches_for(idx)

        folder_match_counts[emp_key] = len(matched_folders)

        # 同名文件夹防错配检测：如果一个名字匹配到多个不同路径的文件夹
        duplicate_folder_warning = ""
        if len(matched_folders) > 1:
            duplicate_folder_warning = f"⚠️ 资料库中存在 {len(matched_folders)} 个同名文件夹，已全部提取归档，请注意核实！"
            warnings.append(f"员工【{emp.name}】：{duplicate_folder_warning}")

        if not matched_folders:
            if emp_materials is not None:
                missing_records[emp_key] = list(emp_materials) if emp_materials else list(global_materials)
            else:
                missing_records[emp_key] = ["（整个文件夹）"]
            continue

        if emp_materials is None:
            _collect_all_from_folders(
                emp, matched_folders, out_path, mode, matches, warnings, duplicate_folder_warning,
                employee_key=employee_key,
                ocr_cache=ocr_cache,
                use_ocr_cache=effective_use_ocr_cache,
                cache_stats=cache_stats,
                verify_image_identity=not skip_folder_all_ocr,
            )
        else:
            emp_missing = _collect_specific_materials(
                emp, matched_folders, out_path, mode, emp_materials, matches, warnings,
                duplicate_folder_warning,
                employee_key=employee_key,
                ocr_cache=ocr_cache,
                use_ocr_cache=effective_use_ocr_cache,
                cache_stats=cache_stats,
                progress_callback=progress_callback,
                cancelled=cancelled,
                review_candidates=review_candidates,
            )
            if emp_missing:
                missing_records[emp_key] = emp_missing

    review_path = _write_document_review(out_path, review_candidates, ocr_cache, warnings, cancelled=cancelled)

    # === OCR 缓存：跑完一轮后汇总写一次（而非每张图片即写）"""
    if (
        effective_use_ocr_cache
        and ocr_cache is not None
        and cache_path is not None
        and (
            ocr_cache.get("entries")
            or library_mode == LIBRARY_MODE_FLAT_OCR
            or legacy_pdf_cache_invalidated
        )
    ):
        _trim_cache_by_age_and_size(ocr_cache)
        if not _save_ocr_cache(cache_path, ocr_cache):
            if cache_write_ok:
                warnings.append(
                    f"OCR 缓存写入失败：{cache_path}（资料库目录可能为只读），本次未持久化识别结果。"
                )
            cache_write_ok = False
            cache_skipped_reason = "资料库目录只读或无写入权限"

    # 缓存指标摘要写进 warnings
    if effective_use_ocr_cache and cache_write_ok and ocr_cache is not None:
        total = cache_stats["hits"] + cache_stats["misses"]
        if total > 0:
            if library_mode == LIBRARY_MODE_FLAT_OCR:
                warnings.append(
                    f"无序资料 OCR 索引：复用 {cache_stats['hits']} 个，新增识别 {cache_stats['misses']} 个"
                    + (f"，缓存文件：{cache_path}" if cache_path else "")
                )
            else:
                warnings.append(
                    f"OCR 智能索引缓存：命中 {cache_stats['hits']} 次，实时识别 {cache_stats['misses']} 次"
                    + (f"，缓存文件：{cache_path}" if cache_path else "")
                )

    zip_path: Path | None = None
    if create_zip:
        _raise_if_cancelled(cancelled)
        zip_path = out_path.parent / f"{out_path.name}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(out_path):
                dirs[:] = [d for d in dirs if not d.startswith(".")]
                for f in files:
                    _raise_if_cancelled(cancelled)
                    if _is_junk_or_temp_file(f):
                        continue
                    full_p = Path(root) / f
                    if f in {"资料待确认.xlsx", "资料待确认预览.html"}:
                        continue
                    arcname = full_p.relative_to(out_path)
                    zf.write(full_p, arcname=str(arcname))

    report_path: Path | None = None
    if collect_all:
        report_materials = ["全部资料"]
    else:
        report_materials = global_materials
    if generate_report:
        report_path = out_path / "《员工资料提取汇总与缺失清单》.xlsx"
        _write_excel_report(
            report_path, employees, report_materials, matches, collect_all, warnings,
            cache_stats=cache_stats, cache_path=cache_path,
            content_verification_skipped=skip_folder_all_ocr,
        )

    return MaterialCollectResult(
        library_dir=lib_path,
        output_dir=out_path,
        zip_path=zip_path,
        report_path=report_path,
        review_path=review_path,
        mode=mode,
        library_mode=library_mode,
        target_employees=employees,
        requested_materials=report_materials,
        matches=matches,
        missing_records=missing_records,
        warnings=warnings,
        folder_match_counts=folder_match_counts,
        employee_result_keys=employee_result_keys,
        ocr_cache_enabled=effective_use_ocr_cache,
        ocr_cache_hits=cache_stats["hits"],
        ocr_cache_misses=cache_stats["misses"],
        ocr_cache_invalidated=cache_stats["invalidated"],
        ocr_cache_path=str(cache_path) if (use_ocr_cache and cache_write_ok and cache_path) else None,
        ocr_cache_skipped_reason=cache_skipped_reason,
    )


# ---------------------------------------------------------------------------
# 收集策略实现
# ---------------------------------------------------------------------------

def _check_mismatch_warning(emp: TargetEmployee, extracted_name: str, extracted_id: str, duplicate_warning: str = "") -> str:
    """核对识别到的证件人名/号码与目标员工是否一致。"""
    warns: list[str] = []
    if duplicate_warning:
        warns.append(duplicate_warning)
    if extracted_name and emp.name and extracted_name != emp.name and emp.name not in extracted_name:
        warns.append(f"⚠️ 证件姓名【{extracted_name}】与目标【{emp.name}】不一致")
    if extracted_id and emp.id_card and extracted_id != emp.id_card:
        warns.append(f"⚠️ 证件号码【{extracted_id}】与目标【{emp.id_card}】不一致")
    return "；".join(warns)


def _collect_all_from_folders(
    emp: TargetEmployee,
    matched_folders: list[tuple[Path, str]],
    out_path: Path,
    mode: str,
    matches: list[MaterialFileMatch],
    warnings: list[str],
    duplicate_warning: str = "",
    *,
    employee_key: str = "",
    ocr_cache: dict[str, Any] | None = None,
    use_ocr_cache: bool = True,
    cache_stats: dict[str, int] | None = None,
    verify_image_identity: bool = True,
) -> None:
    """全部材料模式：将匹配到的文件夹整体拷贝到输出目录。"""
    clean_emp = safe_filename(emp.name)
    seen_hashes: set[tuple[int, str]] = set()

    for folder_idx, (folder_path, match_reason) in enumerate(matched_folders):
        suffix = f"_同名{folder_idx + 1}" if len(matched_folders) > 1 else ""
        dest_name = f"{clean_emp}{suffix}"

        dest_dir = out_path / dest_name
        dest_dir.mkdir(parents=True, exist_ok=True)

        try:
            for root, dirs, files in os.walk(folder_path):
                dirs[:] = [d for d in dirs if not d.startswith(".")]
                rel_root = Path(root).relative_to(folder_path)
                target_root = dest_dir / rel_root
                target_root.mkdir(parents=True, exist_ok=True)

                for f in files:
                    if _is_junk_or_temp_file(f):
                        continue
                    src = Path(root) / f

                    # 同一员工内部跨目录重复文件 Hash 去重
                    sig = _get_file_signature(src)
                    if sig in seen_hashes:
                        continue
                    seen_hashes.add(sig)

                    dst = target_root / f
                    try:
                        shutil.copy2(src, dst)
                    except Exception as e:
                        err_msg = f"复制失败：{src} → {dst}: {e}"
                        warnings.append(err_msg)
                        matches.append(MaterialFileMatch(
                            employee_name=emp.name,
                            material_type="全部",
                            source_path=src,
                            relative_source_path=src.name,
                            matched_by="读取失败",
                            target_filename=f,
                            mismatch_warning=f"⚠️ 文件复制或读取失败: {e}",
                        ))
                        continue

                    try:
                        rel_p = str(src.relative_to(folder_path.parent))
                    except ValueError:
                        rel_p = src.name

                    # 尝试轻量分析图片是否有信息不匹配（优先查缓存）
                    ocr_name, ocr_id = "", ""
                    cache_hit = False
                    if verify_image_identity and src.suffix.lower() in IMAGE_EXTENSIONS:
                        cached = None
                        if use_ocr_cache and ocr_cache is not None:
                            cached = _lookup_ocr_cache(
                                ocr_cache, src,
                                employee_key=employee_key, rel_path=rel_p,
                            )
                        if cached is not None:
                            _, _, _, ocr_name, _, _, _, _ = cached
                            cache_hit = True
                            if cache_stats is not None:
                                cache_stats["hits"] += 1
                        else:
                            ocr_mat, ocr_method, ocr_sub, ocr_name, ocr_id = _classify_by_ocr(src)
                            if cache_stats is not None:
                                cache_stats["misses"] += 1
                            if use_ocr_cache and ocr_cache is not None and ocr_mat:
                                _store_ocr_cache(
                                    ocr_cache, src,
                                    ocr_mat, ocr_method, ocr_sub, ocr_name, ocr_id,
                                    employee_key=employee_key, rel_path=rel_p,
                                )
                    mismatch = _check_mismatch_warning(emp, ocr_name, ocr_id, duplicate_warning)

                    matches.append(MaterialFileMatch(
                        employee_name=emp.name,
                        material_type="全部",
                        source_path=src,
                        relative_source_path=rel_p,
                        matched_by=match_reason,
                        target_filename=f,
                        target_path=dst,
                        extracted_person_name=ocr_name,
                        extracted_id_card=ocr_id,
                        mismatch_warning=mismatch,
                        cache_hit=cache_hit,
                    ))
        except Exception as e:
            warnings.append(f"无法访问文件夹 {folder_path}: {e}")


def _score_file_candidate(
    filename: str,
    requested_materials: list[str],
) -> int:
    """计算文件对当前请求材料的相关性优先级评分（分数越高越优先做 OCR/内容识别）。

    100分: 文件名明确包含请求的材料名称或同义词（如"身份证"、"特种作业"、"安全员"、"劳动合同"）
     80分: 文件名包含编号特征线索（如 T+身份证号、A/B/C+编号、纯身份证号等高疑似文件名）
     50分: 随机/乱码命名的图片、扫描版 PDF 或普通文件
     10分: 文件名明确属于其他【未请求】的材料类型（降级到最后兜底）
    """
    stem = Path(filename).stem.lower()
    canonical_requests = {
        _canonical_material_for_request_label(material) or material
        for material in requested_materials
        if material
    }

    # 1. 文件名直接命中当前请求材料或其同义词 -> 100分
    for req_type in requested_materials:
        if not req_type:
            continue
        syns = MATERIAL_SYNONYMS.get(req_type, [req_type])
        for syn in syns:
            if syn.lower() in stem:
                return 100

    # 2. 文件名包含线索编号特征 -> 80分
    if "特种证书" in canonical_requests or "资格证书" in canonical_requests:
        if re.search(r"(?:^|_)t\d{17}[\dxX]|(?:^|_)t\d{15}", stem):
            return 80
    if "安全员证" in canonical_requests:
        if re.search(r"(?:^|_)[abc]\d{17}[\dxX]|(?:^|_)[abc]\d{15}", stem):
            return 80
    if "身份证" in canonical_requests:
        if re.search(r"(?:^|[^a-z0-9])\d{17}[\dxX](?:[^a-z0-9]|$)|(?:^|[^a-z0-9])\d{15}(?:[^a-z0-9]|$)", stem):
            return 80

    # 3. 检查是否明确包含其他【未请求】材料的同义词 -> 10分
    for other_mat, other_syns in MATERIAL_SYNONYMS.items():
        if other_mat not in canonical_requests:
            for s in other_syns:
                if s.lower() in stem:
                    return 10

    # 4. 其他普通文件（随机命名图片/PDF等） -> 50分
    return 50


def _is_all_requested_materials_satisfied(
    found: dict[str, list[Any]],
    requested_materials: list[str],
) -> bool:
    """检查当前员工所需材料是否已全部找齐（用于触发短路早停，跳过后续无谓 OCR）。"""
    for mat in requested_materials:
        items = found.get(mat)
        if not items:
            return False
        # 如果是身份证且区分正反面，若只有单侧（只有正面无反面，或只有反面无正面），不算完全找齐，继续找另一侧
        if mat == "身份证":
            subtypes = {it[3] for it in items if len(it) > 3}
            if "" not in subtypes:
                if ("正面" in subtypes and "反面" not in subtypes) or ("反面" in subtypes and "正面" not in subtypes):
                    return False
    return True


def _collect_specific_materials(
    emp: TargetEmployee,
    matched_folders: list[tuple[Path, str]],
    out_path: Path,
    mode: str,
    requested_materials: list[str],
    matches: list[MaterialFileMatch],
    warnings: list[str],
    duplicate_warning: str = "",
    *,
    employee_key: str = "",
    ocr_cache: dict[str, Any] | None = None,
    use_ocr_cache: bool = True,
    cache_stats: dict[str, int] | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancelled: Callable[[], bool] | None = None,
    review_candidates: list[_FlatIndexedFile] | None = None,
) -> list[str]:
    """指定材料模式：在匹配到的文件夹中精准搜集对应材料类型的文件。

    使用 启发式优先级排序 + 文件名特征 + 文档内容检索 + 离线视觉 OCR 进行识别：
    1. 优先将高疑似度的文件（如明确命名或含证件编号特征的文件）排在前面进行 OCR / 正文识别；
    2. 劳动合同检查全部候选并关联续页；其他材料在找齐且无高置信度候选后保留早停；
    3. 页面证据在本次扫描内复用，既有持久缓存命中时不重复 OCR。
    """
    clean_emp = safe_filename(emp.name)
    found: dict[str, list[tuple[Path, str, str, str, str, str, bool]]] = {m: [] for m in requested_materials}
    seen_hashes: set[tuple[int, str]] = set()
    contract_labels = [m for m in requested_materials if (_canonical_material_for_request_label(m) or m) == "劳动合同"]
    analysis_records: dict[Path, dict[str, Any]] = {}
    contract_candidates: list[_FlatIndexedFile] = []
    candidate_context: dict[Path, tuple[str, str, str, str, bool]] = {}
    document_warnings: dict[Path, str] = {}

    # 1. 扫描匹配到的所有文件夹，收集所有候选文件
    raw_candidates: list[tuple[Path, str, str]] = []
    for folder_path, folder_reason in matched_folders:
        _raise_if_cancelled(cancelled)
        try:
            for root, dirs, files in os.walk(folder_path):
                dirs[:] = [d for d in dirs if not d.startswith(".")]
                for f in files:
                    _raise_if_cancelled(cancelled)
                    if _is_junk_or_temp_file(f):
                        continue
                    f_path = Path(root) / f
                    ext = f_path.suffix.lower()
                    if ext not in SUPPORTED_FILE_EXTENSIONS:
                        continue
                    try:
                        rel_p = str(f_path.relative_to(folder_path.parent))
                    except ValueError:
                        rel_p = f_path.name
                    raw_candidates.append((f_path, rel_p, folder_reason))
        except MaterialCollectionCancelled:
            raise
        except Exception as e:
            warnings.append(f"无法访问文件夹 {folder_path}: {e}")

    # 2. 按线索优先级评分降序排序（高疑似度文件排在最前面优先做 OCR/内容识别）
    scored_candidates: list[tuple[int, Path, str, str]] = [
        (_score_file_candidate(f_path.name, requested_materials), f_path, rel_p, folder_reason)
        for f_path, rel_p, folder_reason in raw_candidates
    ]
    scored_candidates.sort(key=lambda item: item[0], reverse=True)

    # 3. 按优先级顺序逐个进行精准识别（支持短路早停）
    for idx_cand, (_cand_score, f_path, rel_p, folder_reason) in enumerate(scored_candidates):
        _raise_if_cancelled(cancelled)
        sig = _get_file_signature(f_path)
        if sig in seen_hashes:
            continue

        doc_hint = _build_doc_format_hint(f_path)
        try:
            classified_mat_type, match_method, subtype, ocr_name, ocr_id, cache_hit = _classify_material_type(
                f_path, f_path.name, requested_materials,
                employee_key=employee_key,
                rel_path=rel_p,
                cache=ocr_cache,
                use_cache=use_ocr_cache,
                cache_stats=cache_stats,
                progress_callback=progress_callback,
                cancelled=cancelled,
                analysis_records=analysis_records if contract_labels else None,
            )
        except MaterialCollectionCancelled:
            raise
        except Exception as exc:
            warnings.append(f"文件读取异常 {f_path.name}: {exc}")
            if doc_hint:
                warnings.append(doc_hint)
            matches.append(MaterialFileMatch(
                employee_name=emp.name,
                material_type="未知",
                source_path=f_path,
                relative_source_path=rel_p,
                matched_by="读取失败",
                target_filename=f_path.name,
                mismatch_warning=f"⚠️ 文件读取损坏或异常: {exc}",
            ))
            continue

        requested_label = _requested_label_for_detected_material(
            classified_mat_type or "", requested_materials,
        )
        if classified_mat_type and _is_portrait_material(classified_mat_type):
            photo_names, photo_id, photo_no = _portrait_filename_identity(f_path.name)
            conflict = (
                bool(photo_names and photo_names != (_normalize_person_name(emp.name),))
                or bool(photo_id and emp.id_card and photo_id != _hash_id_card(emp.id_card))
                or bool(photo_no and emp.employee_no and photo_no != emp.employee_no)
            )
            if conflict:
                photo = _FlatIndexedFile(
                    f_path, rel_p, str(f_path), classified_mat_type, match_method, subtype, (), "",
                    document_warning="照片人员归属待确认：文件名与人员目录冲突",
                )
                reviewed = _apply_document_reviews([photo], ocr_cache, [emp], warnings, cancelled=cancelled)[0]
                if reviewed.document_warning:
                    if review_candidates is not None:
                        review_candidates.append(reviewed)
                    warnings.append(f"照片人员归属冲突，未提取：{rel_p}")
                    continue
                match_method = reviewed.match_method
        if contract_labels and f_path.suffix.lower() in IMAGE_EXTENSIONS and classified_mat_type in (None, "其他材料", "劳动合同"):
            entry = analysis_records.get(f_path) or {
                "material_type": classified_mat_type or "其他材料", "match_method": match_method,
                "extracted_name": ocr_name, "extracted_id_hash": _hash_id_card(ocr_id),
            }
            item = _entry_to_flat_indexed_file(f_path, rel_p, str(f_path), entry, cache_hit=cache_hit)
            folder_identity = _FlatIndexedFile(f_path, rel_p, "", "劳动合同", "folder", "", (emp.name,), _hash_id_card(emp.id_card))
            if _document_identity_conflicts(folder_identity, item):
                warnings.append(f"合同人员归属冲突，未提取：{rel_p}")
                continue
            if not item.extracted_names:
                item = replace(item, extracted_names=(emp.name,))
            contract_candidates.append(item)
            candidate_context[f_path] = (rel_p, match_method, ocr_name, ocr_id, cache_hit)
        if requested_label is not None:
            if not any(existing[0] == f_path for existing in found[requested_label]):
                seen_hashes.add(sig)
                found[requested_label].append(
                    (f_path, rel_p, match_method or folder_reason, subtype, ocr_name, ocr_id, cache_hit)
                )

        # 4. 短路早停：如果所有请求的材料都已经找齐，且后续没有高置信度同名候选文件（如多页合同/多个证书），立即停止扫描后续文件！
        if not contract_labels and _is_all_requested_materials_satisfied(found, requested_materials):
            next_score = scored_candidates[idx_cand + 1][0] if idx_cand + 1 < len(scored_candidates) else 0
            if next_score < 80:
                break

    if contract_candidates:
        grouped = _enrich_flat_index_with_document_groups(contract_candidates, warnings, contract_labels, cancelled=cancelled)
        grouped = _apply_document_reviews(grouped, ocr_cache, [emp], warnings, cancelled=cancelled)
        for item in grouped:
            if item.document_warning:
                document_warnings[item.source_path] = item.document_warning
                if review_candidates is not None:
                    review_candidates.append(item)
            if not item.document_group_id and "+manual_review" not in item.match_method:
                continue
            label = contract_labels[0]
            rel_p, method, name, identity, hit = candidate_context[item.source_path]
            found[label] = [row for row in found[label] if row[0] != item.source_path]
            found[label].append((item.source_path, rel_p, item.match_method, "", name, identity, hit))

    # 复制匹配到的真实文件到输出目录
    missing_list: list[str] = []
    for mat_type in requested_materials:
        m_list = found[mat_type]
        if not m_list:
            missing_list.append(mat_type)
            continue

        for seq, (src_path, rel_p, match_reason, subtype, ocr_name, ocr_id, cache_hit) in enumerate(m_list, start=1):
            _raise_if_cancelled(cancelled)
            ext = src_path.suffix
            clean_mat = safe_filename(mat_type)

            # 如果 OCR 或文件名识别出具体子类型（如"正面"、"反面"）
            if subtype:
                target_name = f"{clean_emp}_{clean_mat}_{subtype}{ext}"
            else:
                suffix = f"_{seq}" if len(m_list) > 1 else ""
                target_name = f"{clean_emp}_{clean_mat}{suffix}{ext}"

            if mode == MODE_BY_EMPLOYEE:
                dest_dir = out_path / clean_emp
            elif mode == MODE_BY_MATERIAL:
                dest_dir = out_path / clean_mat
            else:  # FLAT
                dest_dir = out_path

            dest_dir.mkdir(parents=True, exist_ok=True)
            dest_file = _unique_destination(dest_dir / target_name)
            target_name = dest_file.name

            try:
                shutil.copy2(src_path, dest_file)
            except Exception as e:
                warnings.append(f"复制失败：{src_path} → {dest_file}: {e}")
                matches.append(MaterialFileMatch(
                    employee_name=emp.name,
                    material_type=mat_type,
                    source_path=src_path,
                    relative_source_path=rel_p,
                    matched_by="写入失败",
                    target_filename=target_name,
                    mismatch_warning=f"⚠️ 文件复制失败: {e}",
                ))
                continue

            mismatch = _check_mismatch_warning(emp, ocr_name, ocr_id, duplicate_warning)
            mismatch = "；".join(filter(None, (mismatch, document_warnings.get(src_path, ""))))

            matches.append(MaterialFileMatch(
                employee_name=emp.name,
                material_type=mat_type,
                source_path=src_path,
                relative_source_path=rel_p,
                matched_by=match_reason,
                target_filename=target_name,
                target_path=dest_file,
                extracted_person_name=ocr_name,
                extracted_id_card=ocr_id,
                mismatch_warning=mismatch,
                cache_hit=cache_hit,
            ))

    return missing_list


# ---------------------------------------------------------------------------
# Excel 报告生成
# ---------------------------------------------------------------------------

def _write_excel_report(
    report_path: Path,
    employees: list[TargetEmployee],
    requested_materials: list[str],
    all_matches: list[MaterialFileMatch],
    collect_all: bool,
    warnings: list[str] | None = None,
    *,
    cache_stats: dict[str, int] | None = None,
    cache_path: Path | None = None,
    content_verification_skipped: bool = False,
) -> None:
    """Generate structured summary and missing Excel report with optimized columns and wrap text."""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "资料提取汇总与缺失清单"

    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    normal_font = Font(name="微软雅黑", size=10)
    ok_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ok_font = Font(name="微软雅黑", size=10, color="375623")
    missing_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    missing_font = Font(name="微软雅黑", size=10, bold=True, color="C65911")
    warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    warning_font = Font(name="微软雅黑", size=10, bold=True, color="BD8100")
    stat_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    cache_hit_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    cache_hit_font = Font(name="微软雅黑", size=10, color="1F4E79")

    ws_summary["A1"] = "员工资料提取汇总与缺失清单"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = Alignment(vertical="center")

    emp_file_counts: dict[str, int] = {}
    emp_cache_hits: dict[str, int] = {}
    emp_mismatch_warnings: dict[str, list[str]] = {}
    emp_material_counts: dict[str, dict[str, int]] = {}
    emp_pending_materials: dict[str, set[str]] = {}
    uses_identity_keys = any(match.employee_identity_key for match in all_matches)

    def _report_match_key(match: MaterialFileMatch) -> str:
        if uses_identity_keys and match.employee_identity_key:
            return match.employee_identity_key
        return match.employee_name

    def _report_employee_key(employee: TargetEmployee) -> str:
        return employee.identity_key if uses_identity_keys else employee.name

    for m in all_matches:
        match_key = _report_match_key(m)
        emp_file_counts[match_key] = emp_file_counts.get(match_key, 0) + 1
        if m.cache_hit:
            emp_cache_hits[match_key] = emp_cache_hits.get(match_key, 0) + 1
        if m.mismatch_warning:
            emp_mismatch_warnings.setdefault(match_key, []).append(m.mismatch_warning)
            if "待确认" in m.mismatch_warning:
                emp_pending_materials.setdefault(match_key, set()).add(m.material_type)
        if m.material_type in requested_materials:
            material_counts = emp_material_counts.setdefault(match_key, {})
            material_counts[m.material_type] = material_counts.get(m.material_type, 0) + 1

    total_emp = len(employees)
    total_files = len(all_matches)
    found_emp = sum(
        1 for emp in employees if emp_file_counts.get(_report_employee_key(emp), 0) > 0
    )
    not_found_emp = total_emp - found_emp

    # OCR 缓存指标
    hits = cache_stats.get("hits", 0) if cache_stats else 0
    misses = cache_stats.get("misses", 0) if cache_stats else 0
    cache_total = hits + misses
    ws_summary["A3"] = "统计概要"
    ws_summary["A3"].font = Font(name="微软雅黑", size=11, bold=True)

    stats: list[tuple[str, Any, str, Any]] = [
        ("目标员工总数", total_emp, "已提取文件总数", total_files),
        ("已找到员工数", found_emp, "未找到员工数", not_found_emp),
    ]
    # 仅当启用了缓存且有数据时附加缓存统计行
    if cache_stats is not None:
        stats.append(("OCR 缓存命中", hits, "OCR 实时识别", misses))

    for row_idx, (k1, v1, k2, v2) in enumerate(stats, start=4):
        ws_summary[f"A{row_idx}"] = k1
        ws_summary[f"B{row_idx}"] = v1
        ws_summary[f"C{row_idx}"] = k2
        ws_summary[f"D{row_idx}"] = v2
        for col_let in ["A", "B", "C", "D"]:
            c = ws_summary[f"{col_let}{row_idx}"]
            c.border = thin_border
            c.font = normal_font
            if col_let in ["A", "C"]:
                c.fill = stat_fill
                c.font = Font(name="微软雅黑", size=10, bold=True)
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

    # Detail Table
    start_row = 7
    if collect_all:
        headers = ["序号", "员工姓名", "身份证号码", "提取状态", "提取文件数", "OCR 缓存命中", "信息核对预警 / 备注"]
    else:
        headers = ["序号", "员工姓名", "身份证号码", "提取进度", "OCR 缓存命中"] + requested_materials + ["信息核对预警 / 备注"]

    for c_idx, h_text in enumerate(headers, start=1):
        cell = ws_summary.cell(start_row, c_idx, h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    current_r = start_row + 1
    for idx, emp in enumerate(employees, start=1):
        ws_summary.cell(current_r, 1, idx).alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.cell(current_r, 2, emp.name).alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.cell(current_r, 3, emp.id_card).alignment = Alignment(horizontal="center", vertical="center")

        employee_report_key = _report_employee_key(emp)
        file_count = emp_file_counts.get(employee_report_key, 0)
        warn_list = emp_mismatch_warnings.get(employee_report_key, [])
        emp_hits = emp_cache_hits.get(employee_report_key, 0)

        if collect_all:
            status_cell = ws_summary.cell(current_r, 4)
            count_cell = ws_summary.cell(current_r, 5)
            cache_cell = ws_summary.cell(current_r, 6)
            warn_cell = ws_summary.cell(current_r, 7)
            if file_count > 0:
                status_cell.value = "已找到"
                status_cell.fill = ok_fill
                status_cell.font = ok_font
                count_cell.value = file_count
            else:
                status_cell.value = "未找到"
                status_cell.fill = missing_fill
                status_cell.font = missing_font
                count_cell.value = 0
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
            count_cell.alignment = Alignment(horizontal="center", vertical="center")

            cache_cell.value = f"{emp_hits}/{file_count}" if file_count > 0 else "-"
            cache_cell.fill = cache_hit_fill if emp_hits > 0 else stat_fill
            cache_cell.font = cache_hit_font if emp_hits > 0 else normal_font
            cache_cell.alignment = Alignment(horizontal="center", vertical="center")

            if warn_list:
                warn_cell.value = "；".join(sorted(set(warn_list)))
                warn_cell.fill = warning_fill
                warn_cell.font = warning_font
            elif content_verification_skipped and file_count > 0:
                warn_cell.value = "已按文件夹直接复制（未进行 OCR 内容核对）"
                warn_cell.fill = stat_fill
                warn_cell.font = normal_font
            else:
                warn_cell.value = "正常 (信息一致)" if file_count > 0 else "-"
                warn_cell.fill = ok_fill if file_count > 0 else stat_fill
                warn_cell.font = ok_font if file_count > 0 else normal_font
            warn_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            emp_matches_by_type = emp_material_counts.get(employee_report_key, {})

            found_count = sum(1 for m in requested_materials if emp_matches_by_type.get(m, 0) > 0)
            total_req = len(requested_materials)
            status_cell = ws_summary.cell(current_r, 4, f"{found_count}/{total_req}")
            if emp_pending_materials.get(employee_report_key):
                status_cell.value = f"{found_count}/{total_req}（待确认）"
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
            if emp_pending_materials.get(employee_report_key):
                status_cell.fill = warning_fill
                status_cell.font = warning_font
            elif found_count == total_req:
                status_cell.fill = ok_fill
                status_cell.font = ok_font
            else:
                status_cell.fill = missing_fill
                status_cell.font = missing_font

            # OCR 缓存命中列
            cache_cell = ws_summary.cell(current_r, 5)
            cache_cell.value = f"{emp_hits}/{file_count}" if file_count > 0 else "-"
            cache_cell.fill = cache_hit_fill if emp_hits > 0 else stat_fill
            cache_cell.font = cache_hit_font if emp_hits > 0 else normal_font
            cache_cell.alignment = Alignment(horizontal="center", vertical="center")

            for col_offset, mat_type in enumerate(requested_materials, start=6):
                count = emp_matches_by_type.get(mat_type, 0)
                cell = ws_summary.cell(current_r, col_offset)
                cell.border = thin_border
                if count > 0:
                    cell.value = f"已提取({count}份)"
                    cell.fill = ok_fill
                    cell.font = ok_font
                    if mat_type in emp_pending_materials.get(employee_report_key, set()):
                        cell.value += "，待确认"
                        cell.fill = warning_fill
                        cell.font = warning_font
                else:
                    cell.value = "缺失"
                    cell.fill = missing_fill
                    cell.font = missing_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            warn_col_idx = 6 + len(requested_materials)
            warn_cell = ws_summary.cell(current_r, warn_col_idx)
            if warn_list:
                warn_cell.value = "；".join(sorted(set(warn_list)))
                warn_cell.fill = warning_fill
                warn_cell.font = warning_font
            else:
                warn_cell.value = "正常 (信息一致)" if found_count > 0 else "-"
                warn_cell.fill = ok_fill if found_count > 0 else stat_fill
                warn_cell.font = ok_font if found_count > 0 else normal_font
            warn_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for c in range(1, len(headers) + 1):
            ws_summary.cell(current_r, c).border = thin_border
            if c <= 3:
                ws_summary.cell(current_r, c).font = normal_font

        current_r += 1

    # 优化列宽：预警列固定 38 并换行，普通列自适应
    for col in ws_summary.columns:
        col_letter = get_column_letter(col[0].column)
        col_name = str(ws_summary.cell(start_row, col[0].column).value or "")
        if "预警" in col_name or "备注" in col_name:
            ws_summary.column_dimensions[col_letter].width = 38
        elif "缓存命中" in col_name:
            ws_summary.column_dimensions[col_letter].width = 14
        else:
            max_len = max(len(str(cell.value or "")) for cell in col if cell.row >= start_row)
            ws_summary.column_dimensions[col_letter].width = max(min(max_len * 2 + 2, 28), 12)

    # Sheet 2: Matched File List
    ws_files = wb.create_sheet(title="提取文件明细清单")
    file_headers = ["序号", "员工姓名", "材料类型", "目标文件名", "证件识别姓名", "证件识别号码", "信息匹配校验", "匹配依据", "缓存命中", "原始文件路径"]
    for c_idx, h_text in enumerate(file_headers, start=1):
        cell = ws_files.cell(1, c_idx, h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for idx, match_item in enumerate(all_matches, start=1):
        r = idx + 1
        ws_files.cell(r, 1, idx).alignment = Alignment(horizontal="center", vertical="center")
        ws_files.cell(r, 2, match_item.employee_name).alignment = Alignment(horizontal="center", vertical="center")
        ws_files.cell(r, 3, match_item.material_type).alignment = Alignment(horizontal="center", vertical="center")
        ws_files.cell(r, 4, match_item.target_filename).alignment = Alignment(horizontal="left", vertical="center")
        ws_files.cell(r, 5, match_item.extracted_person_name or "-").alignment = Alignment(horizontal="center", vertical="center")
        ws_files.cell(r, 6, match_item.extracted_id_card or "-").alignment = Alignment(horizontal="center", vertical="center")

        chk_cell = ws_files.cell(r, 7)
        if match_item.mismatch_warning:
            chk_cell.value = match_item.mismatch_warning
            chk_cell.fill = warning_fill
            chk_cell.font = warning_font
        elif match_item.extracted_person_name or match_item.extracted_id_card:
            chk_cell.value = "✓ 一致"
            chk_cell.fill = ok_fill
            chk_cell.font = ok_font
        else:
            chk_cell.value = "-"
            chk_cell.font = normal_font
        chk_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws_files.cell(r, 8, match_item.matched_by).alignment = Alignment(horizontal="center", vertical="center")

        # 缓存命中列
        hit_cell = ws_files.cell(r, 9)
        if match_item.cache_hit:
            hit_cell.value = "✓ 命中"
            hit_cell.fill = cache_hit_fill
            hit_cell.font = cache_hit_font
        elif match_item.matched_by and "ocr" in match_item.matched_by.lower():
            hit_cell.value = "✗ 实时"
            hit_cell.font = normal_font
        else:
            hit_cell.value = "-"
            hit_cell.font = normal_font
        hit_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_files.cell(r, 10, match_item.relative_source_path).alignment = Alignment(horizontal="left", vertical="center")

        for c in range(1, len(file_headers) + 1):
            cell = ws_files.cell(r, c)
            cell.border = thin_border
            if c not in (7,):
                cell.font = normal_font

    for col in ws_files.columns:
        col_letter = get_column_letter(col[0].column)
        col_name = str(ws_files.cell(1, col[0].column).value or "")
        if "校验" in col_name or "路径" in col_name:
            ws_files.column_dimensions[col_letter].width = 38
        elif "缓存命中" in col_name:
            ws_files.column_dimensions[col_letter].width = 12
        else:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_files.column_dimensions[col_letter].width = max(min(max_len * 2 + 2, 30), 12)

    # Sheet 3: OCR 缓存指标（如果有缓存数据）
    if cache_stats is not None:
        ws_cache = wb.create_sheet(title="OCR 缓存指标")
        ws_cache["A1"] = "OCR 智能索引缓存指标"
        ws_cache["A1"].font = title_font
        ws_cache["A2"] = "缓存文件"
        ws_cache["B2"] = str(cache_path) if cache_path else "-"
        ws_cache["A3"] = "命中次数"
        ws_cache["B3"] = hits
        ws_cache["A4"] = "实时识别次数"
        ws_cache["B4"] = misses
        ws_cache["A5"] = "命中率"
        ws_cache["B5"] = f"{hits / cache_total * 100:.1f}%" if cache_total > 0 else "-"
        ws_cache["A6"] = "失效次数"
        ws_cache["B6"] = cache_stats.get("invalidated", 0)

        for r in range(1, 7):
            ws_cache.cell(r, 1).font = Font(name="微软雅黑", size=10, bold=True)
            ws_cache.cell(r, 1).fill = stat_fill
            ws_cache.cell(r, 1).alignment = Alignment(horizontal="right", vertical="center")
            ws_cache.cell(r, 1).border = thin_border
            ws_cache.cell(r, 2).font = normal_font
            ws_cache.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws_cache.cell(r, 2).border = thin_border

        ws_cache.column_dimensions["A"].width = 18
        ws_cache.column_dimensions["B"].width = 60

    report_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(report_path)
    finally:
        wb.close()
