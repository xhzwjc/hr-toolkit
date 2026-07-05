from __future__ import annotations

import re
import shutil
import tempfile
import zipfile
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from hr_toolkit.common.resources import open_template_resource
from hr_toolkit.common.excel import apply_row_snapshot, snapshot_row
from hr_toolkit.common.excel_compat import ensure_xlsx_workbook, is_supported_excel_file


TOOL_NAME = "需求2-考勤周月报统计"
TEMPLATE_RESOURCE = "data_statistics_template.xlsx"
OUTPUT_FILENAME = "考勤周月报汇总表.xlsx"
STANDARD_HOURS_PER_DAY = 7
DEFAULT_COMPANY = "总部"


@dataclass
class AttendanceSourceRow:
    source_file: str
    source_row: int
    name: str
    company: str
    department: str
    day: date
    personal_leave_days: float = 0.0
    sick_leave_days: float = 0.0
    paid_leave_days: float = 0.0
    rest_days: float = 0.0
    overtime_days: float = 0.0
    absence_days: float = 0.0
    late_count: int = 0
    early_count: int = 0
    missing_punch_count: int = 0
    expected_hours: float = 0.0
    actual_hours: float = 0.0
    plan_time: str = ""
    punch_record: str = ""
    missing_record: str = ""


@dataclass
class AttendancePersonSummary:
    company: str
    department: str
    name: str
    personal_leave_days: float = 0.0
    sick_leave_days: float = 0.0
    paid_leave_days: float = 0.0
    rest_days: float = 0.0
    month_overtime_days: dict[int, float] = field(default_factory=dict)
    absence_days: float = 0.0
    late_early_count: int = 0
    missing_punch_count: int = 0
    remarks: list[str] = field(default_factory=list)


@dataclass
class AttendanceException:
    name: str
    department: str
    day: date
    exception_type: str
    value: str
    remark: str
    source_file: str
    source_row: int


@dataclass
class ReportRecord:
    report_type: str
    report_no: str
    report_time: datetime
    name: str
    company: str
    department: str
    source_file: str
    source_row: int


@dataclass
class ExpectedReporter:
    name: str
    company: str = DEFAULT_COMPANY
    department: str = "未填写"
    source_file: str = ""
    source_row: int | None = None


@dataclass
class ReportPersonSummary:
    company: str
    department: str
    name: str
    missing_weekly_count: int = 0
    late_weekly_count: int = 0
    missing_monthly_count: int = 0
    late_monthly_count: int = 0
    remarks: list[str] = field(default_factory=list)


@dataclass
class ReportException:
    name: str
    department: str
    report_type: str
    period: str
    exception_type: str
    due_time: datetime
    report_time: datetime | None
    source_file: str = ""
    source_row: int | None = None


@dataclass
class DataStatisticsResult:
    input_dir: Path
    output_dir: Path
    output_file: Path | None = None
    dry_run: bool = False
    source_files: list[str] = field(default_factory=list)
    attendance_source_count: int = 0
    attendance_person_count: int = 0
    attendance_exception_count: int = 0
    weekly_record_count: int = 0
    monthly_record_count: int = 0
    report_person_count: int = 0
    report_exception_count: int = 0
    expected_reporter_count: int = 0
    report_staff_path: Path | None = None
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "tool_name": TOOL_NAME,
            "input_dir": str(self.input_dir),
            "output_dir": str(self.output_dir),
            "output_file": None if self.output_file is None else str(self.output_file),
            "dry_run": self.dry_run,
            "source_files": self.source_files,
            "source_file_count": len(self.source_files),
            "attendance_source_count": self.attendance_source_count,
            "attendance_person_count": self.attendance_person_count,
            "attendance_exception_count": self.attendance_exception_count,
            "weekly_record_count": self.weekly_record_count,
            "monthly_record_count": self.monthly_record_count,
            "report_person_count": self.report_person_count,
            "report_exception_count": self.report_exception_count,
            "expected_reporter_count": self.expected_reporter_count,
            "report_staff_path": None if self.report_staff_path is None else str(self.report_staff_path),
            "warnings": self.warnings,
        }


def generate_data_statistics_reports(
    input_path: str | Path | list[str | Path],
    output_dir: str | Path,
    *,
    report_staff_path: str | Path | None = None,
    dry_run: bool = False,
) -> DataStatisticsResult:
    input_paths = _normalize_input_paths(input_path)
    display_input = input_paths[0] if len(input_paths) == 1 else input_paths[0].parent
    output = Path(output_dir).expanduser().resolve()
    staff_path = Path(report_staff_path).expanduser().resolve() if report_staff_path else None
    warnings: list[str] = []

    for path in input_paths:
        if not path.exists():
            raise FileNotFoundError(f"考勤、周报、月报文件、压缩包或文件夹不存在：{path}")
    if staff_path is not None:
        if not staff_path.exists() or not staff_path.is_file():
            raise FileNotFoundError(f"应汇报人员名单不存在：{staff_path}")
        if staff_path.suffix.lower() not in {".xlsx", ".xls"}:
            raise ValueError("应汇报人员名单只支持 .xlsx 或 .xls 文件。")

    with tempfile.TemporaryDirectory(prefix="hr_data_statistics_") as temp_root:
        temp_dir = Path(temp_root)
        files = _find_source_files(input_paths, temp_dir, warnings)
        if not files:
            raise ValueError("未找到 .xlsx 或 .xls 数据文件。")
        expected_reporters = _read_expected_reporters(staff_path, temp_dir, warnings) if staff_path else []

        attendance_rows: list[AttendanceSourceRow] = []
        weekly_records: list[ReportRecord] = []
        monthly_records: list[ReportRecord] = []
        used_files: list[str] = []
        for file_path in files:
            file_attendance, file_weekly, file_monthly = _read_statistics_file(file_path, warnings)
            if file_attendance or file_weekly or file_monthly:
                used_files.append(str(file_path))
            attendance_rows.extend(file_attendance)
            weekly_records.extend(file_weekly)
            monthly_records.extend(file_monthly)

        if not attendance_rows and not weekly_records and not monthly_records:
            raise ValueError("未识别到考勤结果、周报记录或月报记录，请确认文件格式。")

        attendance_summaries, attendance_exceptions = _summarize_attendance(attendance_rows)
        report_summaries, report_exceptions, report_warnings = _summarize_reports(
            weekly_records,
            monthly_records,
            expected_reporters,
        )
        warnings.extend(report_warnings)

        result = DataStatisticsResult(
            input_dir=display_input,
            output_dir=output,
            dry_run=dry_run,
            source_files=used_files,
            attendance_source_count=len(attendance_rows),
            attendance_person_count=len(attendance_summaries),
            attendance_exception_count=len(attendance_exceptions),
            weekly_record_count=len(weekly_records),
            monthly_record_count=len(monthly_records),
            report_person_count=len(report_summaries),
            report_exception_count=len(report_exceptions),
            expected_reporter_count=len(expected_reporters),
            report_staff_path=staff_path,
            warnings=warnings,
        )
        if dry_run:
            return result

        output.mkdir(parents=True, exist_ok=True)
        output_file = output / OUTPUT_FILENAME
        _write_output_workbook(
            output_file,
            attendance_summaries,
            attendance_exceptions,
            weekly_records,
            monthly_records,
            report_summaries,
            report_exceptions,
            temp_dir,
        )
        result.output_file = output_file
        return result


def _normalize_input_paths(input_path: str | Path | list[str | Path]) -> list[Path]:
    raw_paths = input_path if isinstance(input_path, list) else [input_path]
    paths = [Path(path).expanduser().resolve() for path in raw_paths]
    if not paths:
        raise ValueError("请选择考勤、周报、月报文件、压缩包或文件夹。")
    return paths


def _find_source_files(input_paths: list[Path], temp_dir: Path, warnings: list[str]) -> list[Path]:
    files: list[Path] = []
    seen: set[Path] = set()
    for input_path in input_paths:
        for path in _iter_input_files(input_path, temp_dir, warnings):
            if _is_non_source_file(path):
                continue
            working_path = ensure_xlsx_workbook(path, temp_dir)
            resolved = working_path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            files.append(working_path)
    return sorted(files)


def _iter_input_files(input_path: Path, temp_dir: Path, warnings: list[str]) -> list[Path]:
    if input_path.is_file():
        if is_supported_excel_file(input_path):
            return [input_path]
        if input_path.suffix.lower() == ".zip":
            return _extract_zip_files(input_path, temp_dir, warnings)
        return []
    if not input_path.is_dir():
        raise FileNotFoundError(f"路径不存在：{input_path}")
    files: list[Path] = []
    for path in sorted(input_path.rglob("*")):
        if not path.is_file() or path.name.startswith(("~$", ".~")):
            continue
        if is_supported_excel_file(path):
            files.append(path)
        elif path.suffix.lower() == ".zip":
            files.extend(_extract_zip_files(path, temp_dir, warnings))
    return files


def _extract_zip_files(zip_path: Path, temp_dir: Path, warnings: list[str]) -> list[Path]:
    extract_dir = temp_dir / f"zip_{len(list(temp_dir.glob('zip_*'))) + 1}"
    extract_dir.mkdir(parents=True, exist_ok=True)
    try:
        with zipfile.ZipFile(zip_path) as archive:
            for member in archive.infolist():
                target = extract_dir / member.filename
                if not target.resolve().is_relative_to(extract_dir.resolve()):
                    warnings.append(f"{zip_path.name} 中存在不安全路径，已跳过：{member.filename}")
                    continue
                archive.extract(member, extract_dir)
    except Exception as exc:
        warnings.append(f"{zip_path.name} 解压失败，已跳过：{exc}")
        return []
    return sorted(path for path in extract_dir.rglob("*") if path.is_file() and is_supported_excel_file(path))


def _is_non_source_file(path: Path) -> bool:
    return any(keyword in path.name for keyword in ("模板", "汇总表"))


def _read_statistics_file(file_path: Path, warnings: list[str]) -> tuple[list[AttendanceSourceRow], list[ReportRecord], list[ReportRecord]]:
    attendance_rows: list[AttendanceSourceRow] = []
    weekly_records: list[ReportRecord] = []
    monthly_records: list[ReportRecord] = []
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        for ws in workbook.worksheets:
            header_row = _find_header_row(ws)
            if header_row is None:
                continue
            headers = _read_headers(ws, header_row)
            if _is_attendance_sheet(headers):
                attendance_rows.extend(_read_attendance_sheet(ws, headers, header_row, file_path.name))
            elif _is_summary_attendance_sheet(headers):
                attendance_rows.extend(_read_summary_attendance_sheet(ws, headers, header_row, file_path.name))
            elif _is_report_sheet(headers):
                report_type = _report_type_from_name(file_path.name, ws.title)
                if report_type is None:
                    warnings.append(f"{file_path.name} 未能判断是周报还是月报，已跳过。")
                    continue
                records = _read_report_sheet(ws, headers, header_row, file_path.name, report_type)
                if report_type == "weekly":
                    weekly_records.extend(records)
                else:
                    monthly_records.extend(records)
    finally:
        workbook.close()
    return attendance_rows, weekly_records, monthly_records


def _find_header_row(ws: Worksheet) -> int | None:
    max_col = min(ws.max_column or 0, 80)
    for row_index in range(1, min(ws.max_row or 0, 20) + 1):
        values = {_normalize_header(ws.cell(row_index, col_index).value) for col_index in range(1, max_col + 1)}
        if "姓名" in values and ("日期" in values or "汇报时间" in values):
            return row_index
        if "汇报编号" in values and "汇报人" in values:
            return row_index
        # 汇总格式考勤表：有姓名和应出勤天数/应出勤小时数，但没有日期列
        if "姓名" in values and ("应出勤天数" in values or "应出勤小时数" in values) and "日期" not in values:
            return row_index
        # 汇总格式考勤表（几维等）：有姓名和事假/病假等字段，但没有日期列
        if "姓名" in values and ("事假" in values or "事假（天）" in values or "病假" in values or "病假（天）" in values) and "日期" not in values:
            return row_index
    return None


def _read_headers(ws: Worksheet, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    max_col = min(ws.max_column or 0, 120)
    for col_index in range(1, max_col + 1):
        header = _normalize_header(ws.cell(header_row, col_index).value)
        if header:
            headers[header] = col_index
    return headers


def _is_attendance_sheet(headers: dict[str, int]) -> bool:
    return {"姓名", "日期", "漏打卡次数", "应出勤小时数"}.issubset(headers)


def _is_summary_attendance_sheet(headers: dict[str, int]) -> bool:
    """判断是否为汇总格式的考勤表（没有日期列，直接按人汇总）"""
    if "姓名" not in headers:
        return False
    if "日期" in headers:
        return False
    # 必须有应出勤天数/应出勤小时数，或者有事假/病假等字段
    has_attendance = "应出勤天数" in headers or "应出勤小时数" in headers
    has_leave = any(h in headers for h in ("事假", "事假（天）", "事假\n(天)", "事假\n(小时)", "病假", "病假（天）", "病假\n(天)"))
    return has_attendance or has_leave


def _is_report_sheet(headers: dict[str, int]) -> bool:
    return {"汇报编号", "汇报时间", "汇报人"}.issubset(headers)


def _read_attendance_sheet(ws: Worksheet, headers: dict[str, int], header_row: int, file_name: str) -> list[AttendanceSourceRow]:
    rows: list[AttendanceSourceRow] = []
    for row_index in range(header_row + 1, (ws.max_row or 0) + 1):
        name = _cell_text(_header_value(ws, row_index, headers, "姓名"))
        day = _date_from_value(_header_value(ws, row_index, headers, "日期"))
        if not name or day is None:
            continue
        department_text = _cell_text(_header_value(ws, row_index, headers, "部门名称"))
        company, department = _company_department_from_text(department_text)
        rows.append(
            AttendanceSourceRow(
                source_file=file_name,
                source_row=row_index,
                name=name,
                company=company,
                department=department,
                day=day,
                personal_leave_days=_to_days(_number(_header_value(ws, row_index, headers, "事假"))),
                sick_leave_days=_to_days(_number(_header_value(ws, row_index, headers, "病假天数"))),
                paid_leave_days=_to_days(_number(_header_value(ws, row_index, headers, "年假天数"))),
                rest_days=_to_days(_number(_header_value(ws, row_index, headers, "调休"))),
                overtime_days=_to_days(_number(_header_value(ws, row_index, headers, "加班计调休时长"))),
                absence_days=_number(_header_value(ws, row_index, headers, "旷工天数")),
                late_count=int(_number(_header_value(ws, row_index, headers, "迟到次数"))),
                early_count=int(_number(_header_value(ws, row_index, headers, "早退次数"))),
                missing_punch_count=int(_number(_header_value(ws, row_index, headers, "漏打卡次数"))),
                expected_hours=_number(_header_value(ws, row_index, headers, "应出勤小时数")),
                actual_hours=_number(_header_value(ws, row_index, headers, "实出勤小时数")),
                plan_time=_cell_text(_header_value(ws, row_index, headers, "计划上下班时间")),
                punch_record=_cell_text(_header_value(ws, row_index, headers, "当日刷卡记录")),
                missing_record=_cell_text(_header_value(ws, row_index, headers, "缺卡记录")),
            )
        )
    return rows


def _read_summary_attendance_sheet(ws: Worksheet, headers: dict[str, int], header_row: int, file_name: str) -> list[AttendanceSourceRow]:
    """读取汇总格式的考勤表（没有日期列，直接按人汇总）"""
    rows: list[AttendanceSourceRow] = []
    # 从文件名或标题行推断月份
    month = _infer_month_from_filename(file_name)
    default_date = date(date.today().year, month, 1) if month else date.today()
    for row_index in range(header_row + 1, (ws.max_row or 0) + 1):
        name = _cell_text(_header_value(ws, row_index, headers, "姓名"))
        if not name:
            continue
        # 跳过表头行（姓名列值为"姓名"等）
        if name in ("姓名", "员工姓名", "人员姓名"):
            continue
        # 获取公司和部门
        company_text = _cell_text(_header_value(ws, row_index, headers, "公司"))
        department_text = _cell_text(_header_value(ws, row_index, headers, "部门（片区）"))
        if not department_text:
            department_text = _cell_text(_header_value(ws, row_index, headers, "部门名称"))
        if not department_text:
            department_text = _cell_text(_header_value(ws, row_index, headers, "部门"))
        company, department = _company_department_from_text(department_text)
        if company_text:
            company = company_text
        # 读取各项数据，支持不同的字段名
        personal_leave = _number(_header_value_any(ws, row_index, headers, ("事假", "事假（天）", "事假\n(天)", "事假\n(小时)")))
        sick_leave = _number(_header_value_any(ws, row_index, headers, ("病假天数", "病假", "病假（天）", "病假\n(天)")))
        paid_leave = _number(_header_value_any(ws, row_index, headers, ("年假天数", "年假", "年假\n（天）", "年假\n(天)", "带薪休假", "带薪休假（天）")))
        rest_days = _number(_header_value_any(ws, row_index, headers, ("调休", "调休（小时）", "总调休", "总调休\n(小时)", "总调休（小时）")))
        overtime_days = _number(_header_value_any(ws, row_index, headers, ("加班计调休时长", "当月加班时长", "当月加班（小时）")))
        absence_days = _number(_header_value_any(ws, row_index, headers, ("旷工天数", "旷工", "旷工（天）")))
        late_count = int(_number(_header_value_any(ws, row_index, headers, ("迟到次数", "迟到", "迟到（次）"))))
        early_count = int(_number(_header_value_any(ws, row_index, headers, ("早退次数", "早退", "早退（次）"))))
        missing_punch = int(_number(_header_value_any(ws, row_index, headers, ("漏打卡次数", "漏打卡", "漏打卡（次）"))))
        # 应出勤和实出勤
        expected_hours = _number(_header_value_any(ws, row_index, headers, ("应出勤小时数", "应出勤天数")))
        actual_hours = _number(_header_value_any(ws, row_index, headers, ("实出勤小时数", "实际出勤天数")))
        rows.append(
            AttendanceSourceRow(
                source_file=file_name,
                source_row=row_index,
                name=name,
                company=company,
                department=department,
                day=default_date,
                personal_leave_days=_to_days(personal_leave),
                sick_leave_days=_to_days(sick_leave),
                paid_leave_days=_to_days(paid_leave),
                rest_days=_to_days(rest_days),
                overtime_days=_to_days(overtime_days),
                absence_days=absence_days,
                late_count=late_count,
                early_count=early_count,
                missing_punch_count=missing_punch,
                expected_hours=expected_hours,
                actual_hours=actual_hours,
            )
        )
    return rows


def _infer_month_from_filename(file_name: str) -> int | None:
    """从文件名推断月份"""
    match = re.search(r"(\d{1,2})\s*月", file_name)
    if match:
        month = int(match.group(1))
        if 1 <= month <= 12:
            return month
    return None


def _header_value_any(ws: Worksheet, row_index: int, headers: dict[str, int], candidates: tuple[str, ...]) -> Any:
    """尝试多个表头名，返回第一个匹配的值"""
    for candidate in candidates:
        col_index = headers.get(_normalize_header(candidate))
        if col_index is not None:
            return ws.cell(row_index, col_index).value
    return None


def _read_report_sheet(ws: Worksheet, headers: dict[str, int], header_row: int, file_name: str, report_type: str) -> list[ReportRecord]:
    records: list[ReportRecord] = []
    for row_index in range(header_row + 1, (ws.max_row or 0) + 1):
        name = _cell_text(_header_value(ws, row_index, headers, "汇报人"))
        report_time = _datetime_from_value(_header_value(ws, row_index, headers, "汇报时间"))
        if not name or report_time is None:
            continue
        company, department = _company_department_from_text(_cell_text(_header_value(ws, row_index, headers, "汇报人部门")))
        records.append(
            ReportRecord(
                report_type=report_type,
                report_no=_cell_text(_header_value(ws, row_index, headers, "汇报编号")),
                report_time=report_time,
                name=name,
                company=company,
                department=department,
                source_file=file_name,
                source_row=row_index,
            )
        )
    return records


def _read_expected_reporters(staff_path: Path, temp_dir: Path, warnings: list[str]) -> list[ExpectedReporter]:
    workbook_path = ensure_xlsx_workbook(staff_path, temp_dir)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    reporters: OrderedDict[str, ExpectedReporter] = OrderedDict()
    try:
        for ws in workbook.worksheets:
            header_row = _find_expected_reporter_header_row(ws)
            if header_row is None:
                continue
            headers = _read_headers(ws, header_row)
            for row_index in range(header_row + 1, (ws.max_row or 0) + 1):
                name = _cell_text(_header_value_any(ws, row_index, headers, ("姓名", "汇报人", "员工姓名", "人员姓名")))
                if not name:
                    continue
                company = _cell_text(_header_value_any(ws, row_index, headers, ("公司", "所属公司", "单位")))
                department_text = _cell_text(
                    _header_value_any(
                        ws,
                        row_index,
                        headers,
                        ("部门（片区）", "部门片区", "部门", "部门名称", "汇报人部门", "所属部门"),
                    )
                )
                parsed_company, department = _company_department_from_text(department_text)
                reporter = ExpectedReporter(
                    name=name,
                    company=company or parsed_company,
                    department=department,
                    source_file=staff_path.name,
                    source_row=row_index,
                )
                if name not in reporters:
                    reporters[name] = reporter
        if staff_path and not reporters:
            warnings.append(f"{staff_path.name} 未识别到应汇报人员名单，请确认表头包含“姓名”或“汇报人”。")
    finally:
        workbook.close()
    return list(reporters.values())


def _find_expected_reporter_header_row(ws: Worksheet) -> int | None:
    max_col = min(ws.max_column or 0, 80)
    name_headers = {"姓名", "汇报人", "员工姓名", "人员姓名"}
    for row_index in range(1, min(ws.max_row or 0, 20) + 1):
        values = {_normalize_header(ws.cell(row_index, col_index).value) for col_index in range(1, max_col + 1)}
        if values & name_headers:
            return row_index
    return None


def _summarize_attendance(rows: list[AttendanceSourceRow]) -> tuple[list[AttendancePersonSummary], list[AttendanceException]]:
    summaries: OrderedDict[tuple[str, str, str], AttendancePersonSummary] = OrderedDict()
    exceptions: list[AttendanceException] = []
    for row in rows:
        key = (row.company, row.department, row.name)
        summary = summaries.setdefault(key, AttendancePersonSummary(company=row.company, department=row.department, name=row.name))
        summary.personal_leave_days += row.personal_leave_days
        summary.sick_leave_days += row.sick_leave_days
        summary.paid_leave_days += row.paid_leave_days
        summary.rest_days += row.rest_days
        summary.month_overtime_days[row.day.month] = summary.month_overtime_days.get(row.day.month, 0.0) + row.overtime_days
        summary.absence_days += row.absence_days
        summary.late_early_count += row.late_count + row.early_count
        summary.missing_punch_count += row.missing_punch_count

        remarks = _attendance_row_remarks(row)
        if remarks:
            summary.remarks.append(f"{row.day.month}.{row.day.day}" + "、".join(remarks))
        for exception_type, value, remark in _attendance_row_exceptions(row):
            exceptions.append(
                AttendanceException(
                    name=row.name,
                    department=row.department,
                    day=row.day,
                    exception_type=exception_type,
                    value=value,
                    remark=remark,
                    source_file=row.source_file,
                    source_row=row.source_row,
                )
            )
    return list(summaries.values()), exceptions


def _attendance_row_remarks(row: AttendanceSourceRow) -> list[str]:
    remarks: list[str] = []
    if row.overtime_days:
        remarks.append(f"晚上加班{_format_number(row.overtime_days)}天")
    if row.rest_days:
        prefix = "上午" if row.expected_hours and row.expected_hours <= 3.5 and row.actual_hours == 0 else ""
        remarks.append(f"{prefix}调休{_format_number(row.rest_days)}天")
    if row.personal_leave_days:
        remarks.append(f"事假{_format_number(row.personal_leave_days)}天")
    if row.sick_leave_days:
        remarks.append(f"病假{_format_number(row.sick_leave_days)}天")
    if row.paid_leave_days:
        remarks.append(f"带薪休假{_format_number(row.paid_leave_days)}天")
    if row.absence_days:
        remarks.append(f"旷工{_format_number(row.absence_days)}天")
    if row.late_count:
        remarks.append(f"迟到{row.late_count}次")
    if row.early_count:
        remarks.append(f"早退{row.early_count}次")
    if row.missing_punch_count:
        remarks.append(_missing_punch_remark(row))
    return remarks


def _attendance_row_exceptions(row: AttendanceSourceRow) -> list[tuple[str, str, str]]:
    exceptions: list[tuple[str, str, str]] = []
    if row.missing_punch_count:
        exceptions.append(("漏打卡", str(row.missing_punch_count), _missing_punch_remark(row)))
    if row.late_count:
        exceptions.append(("迟到", str(row.late_count), f"迟到{row.late_count}次"))
    if row.early_count:
        exceptions.append(("早退", str(row.early_count), f"早退{row.early_count}次"))
    if row.absence_days:
        exceptions.append(("旷工", _format_number(row.absence_days), f"旷工{_format_number(row.absence_days)}天"))
    if row.personal_leave_days:
        exceptions.append(("事假", _format_number(row.personal_leave_days), f"事假{_format_number(row.personal_leave_days)}天"))
    if row.sick_leave_days:
        exceptions.append(("病假", _format_number(row.sick_leave_days), f"病假{_format_number(row.sick_leave_days)}天"))
    if row.rest_days:
        exceptions.append(("调休", _format_number(row.rest_days), f"调休{_format_number(row.rest_days)}天"))
    if row.overtime_days:
        exceptions.append(("加班", _format_number(row.overtime_days), f"加班{_format_number(row.overtime_days)}天"))
    return exceptions


def _missing_punch_remark(row: AttendanceSourceRow) -> str:
    if row.missing_record:
        return row.missing_record
    if row.missing_punch_count == 1:
        if row.overtime_days:
            return "下班未打卡"
        plan_start = _first_time_in_text(row.plan_time)
        first_punch = _first_time_in_text(row.punch_record)
        if plan_start is not None and first_punch is not None and first_punch >= plan_start:
            return "上班未打卡"
        return "下班未打卡"
    return f"漏打卡{row.missing_punch_count}次"


def _summarize_reports(
    weekly_records: list[ReportRecord],
    monthly_records: list[ReportRecord],
    expected_reporters: list[ExpectedReporter],
) -> tuple[list[ReportPersonSummary], list[ReportException], list[str]]:
    warnings: list[str] = []
    people = _build_report_people(weekly_records, monthly_records, expected_reporters)
    if not people:
        return [], [], warnings
    summaries = {
        name: ReportPersonSummary(company=company, department=department, name=name)
        for name, (company, department) in people.items()
    }
    exceptions: list[ReportException] = []

    weekly_due_dates, skipped_due_dates = _weekly_due_dates(weekly_records)
    weekly_by_due = _group_weekly_records(weekly_records, weekly_due_dates)
    expected_weekly_people = _expected_weekly_people(
        weekly_records,
        monthly_records,
        expected_reporters,
        skipped_due_dates,
        weekly_due_dates,
    )
    for due_index, due_date in enumerate(weekly_due_dates, start=1):
        due_time = datetime.combine(due_date, time(17, 0))
        period_text = f"第{_chinese_number(due_index)}周"
        for name in sorted(expected_weekly_people):
            person_records = weekly_by_due.get(due_date, {}).get(name, [])
            if not person_records:
                if due_date in skipped_due_dates:
                    continue
                _add_report_exception(
                    summaries[name],
                    exceptions,
                    report_type="周报",
                    period=period_text,
                    exception_type="未写周报",
                    due_time=due_time,
                    report_time=None,
                )
                continue
            first_report = min(person_records, key=lambda item: item.report_time)
            if _is_report_late(first_report.report_time, due_time):
                _add_report_exception(
                    summaries[name],
                    exceptions,
                    report_type="周报",
                    period=period_text,
                    exception_type="周报超时",
                    due_time=due_time,
                    report_time=first_report.report_time,
                    source_file=first_report.source_file,
                    source_row=first_report.source_row,
                )

    monthly_due = _monthly_due_time(weekly_records, monthly_records)
    monthly_by_name: dict[str, list[ReportRecord]] = {}
    for record in monthly_records:
        monthly_by_name.setdefault(record.name, []).append(record)
    monthly_period_text = _monthly_period_text(monthly_due)
    for name, summary in summaries.items():
        records = monthly_by_name.get(name, [])
        if not records:
            _add_report_exception(
                summary,
                exceptions,
                report_type="月报",
                period=monthly_period_text,
                exception_type="未写月报",
                due_time=monthly_due,
                report_time=None,
            )
            continue
        first_report = min(records, key=lambda item: item.report_time)
        if _is_report_late(first_report.report_time, monthly_due):
            _add_report_exception(
                summary,
                exceptions,
                report_type="月报",
                period=monthly_period_text,
                exception_type="月报超时",
                due_time=monthly_due,
                report_time=first_report.report_time,
                source_file=first_report.source_file,
                source_row=first_report.source_row,
            )

    if weekly_records and not weekly_due_dates:
        warnings.append("未能从周报文件名或汇报时间推断周报截止周期，周报未写统计可能不完整。")
    summaries_with_issues = sorted(
        [summary for summary in summaries.values() if _report_issue_count(summary) > 0],
        key=lambda item: (
            item.missing_monthly_count + item.late_monthly_count == 0,
            item.department,
            item.name,
        ),
    )
    return summaries_with_issues, exceptions, warnings


def _build_report_people(
    weekly_records: list[ReportRecord],
    monthly_records: list[ReportRecord],
    expected_reporters: list[ExpectedReporter],
) -> OrderedDict[str, tuple[str, str]]:
    people: OrderedDict[str, tuple[str, str]] = OrderedDict()
    for reporter in expected_reporters:
        people[reporter.name] = (reporter.company, reporter.department)
    for record in weekly_records + monthly_records:
        if record.name not in people or people[record.name][1] == "未填写":
            people[record.name] = (record.company, record.department)
    return people


def _weekly_due_dates(records: list[ReportRecord]) -> tuple[list[date], set[date]]:
    if not records:
        return [], set()
    start, end = _report_range_from_records(records)
    first_monday = start + timedelta(days=(7 - start.weekday()) % 7)
    due_dates: list[date] = []
    current = first_monday
    while current <= end:
        due_dates.append(current)
        current += timedelta(days=7)
    skipped = {due_dates[0]} if due_dates and start.weekday() != 0 else set()
    return due_dates, skipped


def _report_range_from_records(records: list[ReportRecord]) -> tuple[date, date]:
    for record in records:
        parsed = _range_from_filename(record.source_file, record.report_time.year)
        if parsed:
            return parsed
    min_day = min(record.report_time.date() for record in records)
    max_day = max(record.report_time.date() for record in records)
    month_start = min_day.replace(day=1)
    return month_start, max_day


def _range_from_filename(file_name: str, year: int) -> tuple[date, date] | None:
    match = re.search(r"([01]?\d)[.月/-]([0-3]?\d)\s*[-—~至]+\s*([01]?\d)[.月/-]([0-3]?\d)", file_name)
    if not match:
        return None
    start_month, start_day, end_month, end_day = (int(part) for part in match.groups())
    start = date(year, start_month, start_day)
    end_year = year + 1 if end_month < start_month else year
    end = date(end_year, end_month, end_day)
    return start, end


def _group_weekly_records(records: list[ReportRecord], due_dates: list[date]) -> dict[date, dict[str, list[ReportRecord]]]:
    grouped: dict[date, dict[str, list[ReportRecord]]] = {}
    for record in records:
        due_date = _assign_weekly_due_date(record.report_time.date(), due_dates)
        if due_date is None:
            continue
        grouped.setdefault(due_date, {}).setdefault(record.name, []).append(record)
    return grouped


def _assign_weekly_due_date(report_day: date, due_dates: list[date]) -> date | None:
    for due_date in due_dates:
        if report_day <= due_date:
            return due_date
    return due_dates[-1] if due_dates else None


def _expected_weekly_people(
    weekly_records: list[ReportRecord],
    monthly_records: list[ReportRecord],
    expected_reporters: list[ExpectedReporter],
    skipped_due_dates: set[date],
    due_dates: list[date],
) -> set[str]:
    if expected_reporters:
        return {reporter.name for reporter in expected_reporters}
    expected = {record.name for record in monthly_records}
    for record in weekly_records:
        due_date = _assign_weekly_due_date(record.report_time.date(), due_dates)
        if due_date not in skipped_due_dates:
            expected.add(record.name)
    return expected or {record.name for record in weekly_records}


def _is_report_late(report_time: datetime, due_time: datetime) -> bool:
    return report_time >= due_time + timedelta(minutes=1)


def _monthly_due_time(weekly_records: list[ReportRecord], monthly_records: list[ReportRecord]) -> datetime:
    records = monthly_records or weekly_records
    if not records:
        today = date.today()
        return datetime(today.year, today.month, 2, 17, 0)
    start, _end = _report_range_from_records(records)
    year = start.year
    month = start.month
    next_month = 1 if month == 12 else month + 1
    next_year = year + 1 if month == 12 else year
    return datetime(next_year, next_month, 2, 17, 0)


def _monthly_period_text(monthly_due: datetime) -> str:
    month = monthly_due.month - 1
    year = monthly_due.year
    if month == 0:
        month = 12
        year -= 1
    return f"{year}年{month}月"


def _add_report_exception(
    summary: ReportPersonSummary,
    exceptions: list[ReportException],
    *,
    report_type: str,
    period: str,
    exception_type: str,
    due_time: datetime,
    report_time: datetime | None,
    source_file: str = "",
    source_row: int | None = None,
) -> None:
    if exception_type == "未写周报":
        summary.missing_weekly_count += 1
    elif exception_type == "周报超时":
        summary.late_weekly_count += 1
    elif exception_type == "未写月报":
        summary.missing_monthly_count += 1
    elif exception_type == "月报超时":
        summary.late_monthly_count += 1
    summary.remarks.append(f"{period}{exception_type}" if report_type == "周报" else exception_type)
    exceptions.append(
        ReportException(
            name=summary.name,
            department=summary.department,
            report_type=report_type,
            period=period,
            exception_type=exception_type,
            due_time=due_time,
            report_time=report_time,
            source_file=source_file,
            source_row=source_row,
        )
    )


def _report_issue_count(summary: ReportPersonSummary) -> int:
    return summary.missing_weekly_count + summary.late_weekly_count + summary.missing_monthly_count + summary.late_monthly_count


def _write_output_workbook(
    output_file: Path,
    attendance_summaries: list[AttendancePersonSummary],
    attendance_exceptions: list[AttendanceException],
    weekly_records: list[ReportRecord],
    monthly_records: list[ReportRecord],
    report_summaries: list[ReportPersonSummary],
    report_exceptions: list[ReportException],
    temp_dir: Path,
) -> None:
    template_path = _copy_template(temp_dir)
    workbook = load_workbook(template_path)
    try:
        attendance_ws = workbook["考勤表模板"]
        attendance_ws.title = "考勤统计"
        report_ws = workbook["周月报模板"]
        report_ws.title = "周月报统计"
        _write_attendance_sheet(attendance_ws, attendance_summaries)
        _write_report_sheet(report_ws, report_summaries, weekly_records, monthly_records)
        _write_attendance_detail_sheet(workbook, attendance_exceptions)
        _write_report_detail_sheet(workbook, report_exceptions)
        workbook.save(output_file)
    finally:
        workbook.close()


def _copy_template(temp_dir: Path) -> Path:
    target = temp_dir / TEMPLATE_RESOURCE
    with open_template_resource(TEMPLATE_RESOURCE) as source, target.open("wb") as output:
        shutil.copyfileobj(source, output)
    return target


def _write_attendance_sheet(ws: Worksheet, summaries: list[AttendancePersonSummary]) -> None:
    max_month = max([4] + [month for summary in summaries for month in summary.month_overtime_days])
    if max_month > 4:
        ws.insert_cols(13, max_month - 4)
    stat_start_col = 9 + max_month
    headers = ["序号", "公司", "部门（片区）", "姓名", "事假（天）", "病假（天）", "带薪休假（天）", "调休（天）"]
    headers.extend(f"{month}月份加班天数" for month in range(1, max_month + 1))
    headers.extend(["旷工", "迟到/早退（次）", "漏打卡", "累计剩余加班天数", "备注"])
    for col_index, header in enumerate(headers, start=1):
        ws.cell(2, col_index).value = header

    max_col = len(headers)
    template_snapshot = snapshot_row(ws, 3, min(ws.max_column, max_col))
    if len(summaries) > 1:
        ws.insert_rows(4, len(summaries) - 1)
    for offset, summary in enumerate(summaries):
        row_index = 3 + offset
        apply_row_snapshot(ws, row_index, template_snapshot, translate_formulas=True)
        for col_index in range(1, max_col + 1):
            ws.cell(row_index, col_index).value = None
        ws.cell(row_index, 1).value = offset + 1
        ws.cell(row_index, 2).value = summary.company
        ws.cell(row_index, 3).value = summary.department
        ws.cell(row_index, 4).value = summary.name
        ws.cell(row_index, 5).value = _zero_blank(summary.personal_leave_days)
        ws.cell(row_index, 6).value = _zero_blank(summary.sick_leave_days)
        ws.cell(row_index, 7).value = _zero_blank(summary.paid_leave_days)
        ws.cell(row_index, 8).value = _zero_blank(summary.rest_days)
        for month in range(1, max_month + 1):
            ws.cell(row_index, 8 + month).value = _zero_blank(summary.month_overtime_days.get(month, 0.0))
        ws.cell(row_index, stat_start_col).value = _zero_blank(summary.absence_days)
        ws.cell(row_index, stat_start_col + 1).value = summary.late_early_count or None
        ws.cell(row_index, stat_start_col + 2).value = summary.missing_punch_count or None
        start_ref = f"{get_column_letter(9)}{row_index}"
        end_ref = f"{get_column_letter(8 + max_month)}{row_index}"
        rest_ref = f"{get_column_letter(8)}{row_index}"
        ws.cell(row_index, stat_start_col + 3).value = f"=SUM({start_ref}:{end_ref})-{rest_ref}"
        ws.cell(row_index, stat_start_col + 4).value = "；".join(summary.remarks) + ("；" if summary.remarks else "")
        _format_row(ws, row_index, max_col)
    if not summaries:
        for col_index in range(1, max_col + 1):
            ws.cell(3, col_index).value = None
    _format_table(ws, 2, max(3, 2 + len(summaries)), max_col)
    ws.freeze_panes = "A3"


def _write_report_sheet(
    ws: Worksheet,
    summaries: list[ReportPersonSummary],
    weekly_records: list[ReportRecord],
    monthly_records: list[ReportRecord],
) -> None:
    period_title = _report_title(weekly_records, monthly_records)
    ws["A1"].value = period_title
    template_snapshot = snapshot_row(ws, 3, 10)
    if len(summaries) > 2:
        ws.insert_rows(5, len(summaries) - 2)
    for offset, summary in enumerate(summaries):
        row_index = 3 + offset
        apply_row_snapshot(ws, row_index, template_snapshot, translate_formulas=True)
        for col_index in range(1, 11):
            ws.cell(row_index, col_index).value = None
        ws.cell(row_index, 1).value = offset + 1
        ws.cell(row_index, 2).value = summary.company
        ws.cell(row_index, 3).value = summary.department
        ws.cell(row_index, 4).value = summary.name
        ws.cell(row_index, 5).value = summary.missing_weekly_count or None
        ws.cell(row_index, 6).value = summary.late_weekly_count or None
        ws.cell(row_index, 7).value = summary.missing_monthly_count or None
        ws.cell(row_index, 8).value = summary.late_monthly_count or None
        ws.cell(row_index, 9).value = None
        ws.cell(row_index, 10).value = "；".join(summary.remarks)
        _format_row(ws, row_index, 10)
    for row_index in range(3 + len(summaries), 5):
        for col_index in range(1, 11):
            ws.cell(row_index, col_index).value = None
    total_row = 5 + max(0, len(summaries) - 2)
    ws.cell(total_row, 2).value = _report_total_text(weekly_records)
    ws.cell(total_row, 9).value = None
    _format_table(ws, 2, total_row, 10, set_column_widths=False)
    _write_report_footer(ws, total_row, weekly_records, monthly_records)
    if ws.max_column > 10:
        ws.delete_cols(11, ws.max_column - 10)
    ws.freeze_panes = "A3"


def _write_attendance_detail_sheet(workbook, exceptions: list[AttendanceException]) -> None:
    ws = workbook.create_sheet("考勤异常明细")
    headers = ["序号", "部门（片区）", "姓名", "日期", "异常类型", "数值", "说明", "来源文件", "来源行"]
    _write_headers(ws, headers)
    for index, item in enumerate(exceptions, start=1):
        values = [index, item.department, item.name, item.day, item.exception_type, item.value, item.remark, item.source_file, item.source_row]
        for col_index, value in enumerate(values, start=1):
            ws.cell(index + 1, col_index).value = value
        ws.cell(index + 1, 4).number_format = "yyyy/m/d"
    _format_table(ws, 1, max(2, len(exceptions) + 1), len(headers))
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 36


def _write_report_detail_sheet(workbook, exceptions: list[ReportException]) -> None:
    ws = workbook.create_sheet("周月报异常明细")
    headers = ["序号", "部门（片区）", "姓名", "类型", "周期", "异常类型", "截止时间", "汇报时间", "扣款金额", "来源文件", "来源行"]
    _write_headers(ws, headers)
    sorted_exceptions = sorted(exceptions, key=lambda item: (item.report_type != "月报", item.department, item.name, item.period))
    for index, item in enumerate(sorted_exceptions, start=1):
        values = [
            index,
            item.department,
            item.name,
            item.report_type,
            item.period,
            item.exception_type,
            item.due_time,
            item.report_time,
            None,
            item.source_file,
            item.source_row,
        ]
        for col_index, value in enumerate(values, start=1):
            ws.cell(index + 1, col_index).value = value
        ws.cell(index + 1, 7).number_format = "yyyy/m/d h:mm"
        ws.cell(index + 1, 8).number_format = "yyyy/m/d h:mm"
    _format_table(ws, 1, max(2, len(exceptions) + 1), len(headers))
    ws.column_dimensions["J"].width = 36


def _write_headers(ws: Worksheet, headers: list[str]) -> None:
    for col_index, header in enumerate(headers, start=1):
        ws.cell(1, col_index).value = header


def _format_table(ws: Worksheet, min_row: int, max_row: int, max_col: int, *, set_column_widths: bool = True) -> None:
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    header_fill = PatternFill("solid", fgColor="FCE4D6")
    header_font = Font(name="宋体", size=10, bold=True)
    normal_font = Font(name="宋体", size=10)
    for row_index in range(min_row, max_row + 1):
        for col_index in range(1, max_col + 1):
            cell = ws.cell(row_index, col_index)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = header_font if row_index == min_row else normal_font
            if row_index == min_row:
                cell.fill = header_fill
            if isinstance(cell.value, float):
                cell.number_format = "0.##"
    if set_column_widths:
        for col_index in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col_index)].width = 14


def _format_row(ws: Worksheet, row_index: int, max_col: int) -> None:
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    for col_index in range(1, max_col + 1):
        cell = ws.cell(row_index, col_index)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if isinstance(cell.value, float):
            cell.number_format = "0.##"


def _write_report_footer(
    ws: Worksheet,
    total_row: int,
    weekly_records: list[ReportRecord],
    monthly_records: list[ReportRecord],
) -> None:
    approval_row = total_row + 2
    weekly_rule_row = approval_row + 2
    monthly_rule_row = weekly_rule_row + 1

    for row_index in range(total_row + 1, monthly_rule_row + 1):
        for col_index in range(1, 11):
            cell = ws.cell(row_index, col_index)
            cell.value = None
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font(name="宋体", size=11)
            cell.alignment = Alignment()

    _merge_report_range(ws, total_row, 2, 8)
    _merge_report_range(ws, approval_row, 1, 10)
    _merge_report_range(ws, weekly_rule_row, 4, 8)
    _merge_report_range(ws, monthly_rule_row, 4, 8)

    ws.row_dimensions[total_row].height = 29
    ws.row_dimensions[total_row + 1].height = 19
    ws.row_dimensions[weekly_rule_row].height = None
    ws.row_dimensions[monthly_rule_row].height = None

    ws.cell(approval_row, 1).value = "   审批：                                审核：                                            制表："
    ws.cell(weekly_rule_row, 3).value = "汇报规则："
    ws.cell(weekly_rule_row, 4).value = _weekly_rule_text(weekly_records)
    ws.cell(monthly_rule_row, 3).value = "月报规则："
    ws.cell(monthly_rule_row, 4).value = _monthly_rule_text(weekly_records, monthly_records)

    ws.cell(approval_row, 1).font = Font(name="宋体", size=12)
    ws.cell(approval_row, 1).alignment = Alignment(horizontal="left", vertical="center")

    for row_index in (weekly_rule_row, monthly_rule_row):
        ws.cell(row_index, 3).font = Font(name="宋体", size=12)
        ws.cell(row_index, 3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row_index, 4).font = Font(name="宋体", size=12)
        ws.cell(row_index, 4).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    _restore_report_column_widths(ws)


def _merge_report_range(ws: Worksheet, row_index: int, start_col: int, end_col: int) -> None:
    start = f"{get_column_letter(start_col)}{row_index}"
    end = f"{get_column_letter(end_col)}{row_index}"
    target = f"{start}:{end}"
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.min_row <= row_index <= merged_range.max_row
            and not (merged_range.max_col < start_col or merged_range.min_col > end_col)
        ):
            ws.unmerge_cells(str(merged_range))
    ws.merge_cells(target)


def _restore_report_column_widths(ws: Worksheet) -> None:
    widths = {
        "A": 4.82407407407407,
        "B": 6.02777777777778,
        "C": 16.0,
        "D": 14.7777777777778,
        "E": 13.8796296296296,
        "F": 13.2777777777778,
        "G": 13.6388888888889,
        "H": 15.5555555555556,
        "I": 10.5,
        "J": 53.6666666666667,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _report_title(weekly_records: list[ReportRecord], monthly_records: list[ReportRecord]) -> str:
    records = weekly_records or monthly_records
    if not records:
        return "周月报汇总"
    start, _end = _report_range_from_records(records)
    return f"{start.year}年{start.month}月份周月报汇总"


def _report_total_text(weekly_records: list[ReportRecord]) -> str:
    last_due = _last_effective_weekly_due(weekly_records)
    if last_due is None:
        return "总计"
    return f"总计（周报截止时间{_format_due_date(last_due)} 17:00）"


def _weekly_rule_text(weekly_records: list[ReportRecord]) -> str:
    due_dates = _effective_weekly_due_dates(weekly_records)
    base_rule = "周报在每周六中午下班后即可汇报，截止时间为次周周一17:00"
    if not due_dates:
        return base_rule + "；17:00:59前正常，17:01起异常。"
    due_text = "、".join(_format_due_date(day) for day in due_dates)
    return f"{base_rule}；本期截止日期：{due_text}；17:00:59前正常，17:01起异常。"


def _monthly_rule_text(weekly_records: list[ReportRecord], monthly_records: list[ReportRecord]) -> str:
    base_rule = "月报在每月最后一个工作日结束后即可汇报，截止时间为次月2日17:00"
    if not weekly_records and not monthly_records:
        return base_rule + "；17:00:59前正常，17:01起异常。"
    monthly_due = _monthly_due_time(weekly_records, monthly_records)
    return f"{base_rule}；本期截止日期：{_format_due_date(monthly_due.date())} 17:00；17:00:59前正常，17:01起异常。"


def _effective_weekly_due_dates(weekly_records: list[ReportRecord]) -> list[date]:
    due_dates, skipped = _weekly_due_dates(weekly_records)
    return [due_day for due_day in due_dates if due_day not in skipped]


def _last_effective_weekly_due(weekly_records: list[ReportRecord]) -> date | None:
    due_dates = _effective_weekly_due_dates(weekly_records)
    if not due_dates:
        return None
    return due_dates[-1]


def _format_due_date(day: date) -> str:
    return f"{day.year}.{day.month}.{day.day}"


def _report_type_from_name(file_name: str, sheet_name: str) -> str | None:
    text = f"{file_name} {sheet_name}"
    if "周报" in text:
        return "weekly"
    if "月报" in text:
        return "monthly"
    return None


def _company_department_from_text(text: str) -> tuple[str, str]:
    if not text:
        return DEFAULT_COMPANY, "未填写"
    parts = [part.strip() for part in re.split(r"[/／]", text) if part and part.strip()]
    department = parts[-1] if parts else text
    return DEFAULT_COMPANY, department


def _header_value(ws: Worksheet, row_index: int, headers: dict[str, int], header: str) -> Any:
    col_index = headers.get(_normalize_header(header))
    if col_index is None:
        return None
    return ws.cell(row_index, col_index).value


def _header_value_any(ws: Worksheet, row_index: int, headers: dict[str, int], candidates: tuple[str, ...]) -> Any:
    for header in candidates:
        value = _header_value(ws, row_index, headers, header)
        if value is not None:
            return value
    return None


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", _cell_text(value))


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = _cell_text(value).replace(",", "").replace("，", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _to_days(value: float) -> float:
    if not value:
        return 0.0
    if abs(value) <= 1:
        return round(value, 2)
    return round(value / STANDARD_HOURS_PER_DAY, 2)


def _date_from_value(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None
    text = _cell_text(value)
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m.%d"):
        try:
            parsed = datetime.strptime(text, pattern)
            if pattern == "%m.%d":
                parsed = parsed.replace(year=date.today().year)
            return parsed.date()
        except ValueError:
            continue
    return None


def _datetime_from_value(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time())
    if isinstance(value, (int, float)):
        try:
            return from_excel(value)
        except Exception:
            return None
    text = _cell_text(value)
    for pattern in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def _first_time_in_text(text: str) -> time | None:
    match = re.search(r"([0-2]?\d):([0-5]\d)", text)
    if not match:
        return None
    hour, minute = (int(part) for part in match.groups())
    if hour > 23:
        return None
    return time(hour, minute)


def _zero_blank(value: float) -> float | None:
    return round(value, 2) if value else None


def _format_number(value: float) -> str:
    value = round(value, 2)
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def _chinese_number(value: int) -> str:
    numbers = {
        1: "一",
        2: "二",
        3: "三",
        4: "四",
        5: "五",
        6: "六",
        7: "七",
        8: "八",
        9: "九",
        10: "十",
    }
    return numbers.get(value, str(value))
