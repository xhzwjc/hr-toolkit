from __future__ import annotations

import filecmp
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from hr_toolkit.common.excel_compat import ensure_xlsx_workbook, is_supported_excel_file


TOOL_NAME = "需求8-人员资料文件夹改名"
MODE_APPEND = "append"
MODE_REMOVE = "remove"
MODE_REPLACE = "replace"
MODE_EXCEL_BATCH = "excel"
MODES = {MODE_APPEND, MODE_REMOVE, MODE_REPLACE}
WINDOWS_INVALID_CHARS = set('<>:"/\\|?*')
WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    "CLOCK$",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}

# 文件类型分组
FILE_TYPE_FOLDER = "folder"
FILE_TYPE_PDF = "pdf"
FILE_TYPE_IMAGE = "image"
FILE_TYPE_DOCUMENT = "document"
FILE_TYPE_ALL = "all"
FILE_TYPE_EXTENSIONS: dict[str, list[str]] = {
    FILE_TYPE_FOLDER: [],
    FILE_TYPE_PDF: [".pdf"],
    FILE_TYPE_IMAGE: [".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"],
    FILE_TYPE_DOCUMENT: [".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".txt"],
    FILE_TYPE_ALL: [],
}


@dataclass(frozen=True)
class FolderRenameOperation:
    source: Path
    target: Path

    def to_dict(self) -> dict[str, str]:
        return {
            "source": str(self.source),
            "target": str(self.target),
            "source_name": self.source.name,
            "target_name": self.target.name,
        }


@dataclass
class FolderRenameResult:
    root_dir: Path
    mode: str
    dry_run: bool = False
    operations: list[FolderRenameOperation] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def operation_count(self) -> int:
        return len(self.operations)

    def to_dict(self) -> dict[str, Any]:
        return {
            "tool_name": TOOL_NAME,
            "root_dir": str(self.root_dir),
            "mode": self.mode,
            "dry_run": self.dry_run,
            "operation_count": self.operation_count,
            "operations": [operation.to_dict() for operation in self.operations],
            "warnings": self.warnings,
        }


def rename_person_folders(
    root_dir: str | Path,
    *,
    mode: str,
    text: str = "",
    target_name: str = "",
    replacement_name: str = "",
    file_type: str = FILE_TYPE_FOLDER,
    dry_run: bool = False,
) -> FolderRenameResult:
    root_dir = Path(root_dir).expanduser().resolve()
    if mode not in MODES:
        raise ValueError(f"不支持的改名模式：{mode}")
    if not root_dir.exists() or not root_dir.is_dir():
        raise FileNotFoundError(f"文件夹不存在：{root_dir}")

    operations, warnings = _plan_operations(
        root_dir=root_dir,
        mode=mode,
        text=text,
        target_name=target_name,
        replacement_name=replacement_name,
        file_type=file_type,
    )
    result = FolderRenameResult(
        root_dir=root_dir,
        mode=mode,
        dry_run=dry_run,
        operations=operations,
        warnings=warnings,
    )
    if dry_run:
        return result

    completed: list[FolderRenameOperation] = []
    runtime_warnings = list(warnings)
    for operation in operations:
        try:
            operation.source.rename(operation.target)
        except OSError as exc:
            runtime_warnings.append(f"{operation.source.name} 改名失败：{exc}")
            continue
        completed.append(operation)
    result.operations = completed
    result.warnings = runtime_warnings
    return result


def rename_files_by_excel(
    root_dir: str | Path,
    excel_path: str | Path,
    *,
    name_column: str = "姓名",
    header_row: int = 1,
    file_type: str = FILE_TYPE_ALL,
    file_extensions: list[str] | None = None,
    expected_operations: list[tuple[str, str]] | None = None,
    expected_warnings: list[str] | None = None,
    dry_run: bool = False,
) -> FolderRenameResult:
    """按照 Excel 姓名行顺序批量改名目录中的第一层项目。

    Args:
        root_dir: 包含待改名项目的目录。
        excel_path: 包含姓名列的 Excel 文件。
        name_column: 姓名列的表头名称，默认“姓名”。
        header_row: 表头行号，默认 1。
        file_type: 待改名项目类型，支持文件夹、PDF、图片、文档和全部。
        file_extensions: 兼容旧调用的扩展名过滤；传入后仅筛选这些扩展名的文件。
        expected_operations: 已确认预览中的“原名称、目标名称”列表；不一致时拒绝执行。
        expected_warnings: 已确认预览中的提醒列表；不一致时拒绝执行。
        dry_run: 是否只生成预览而不执行。
    """
    root_dir = Path(root_dir).expanduser().resolve()
    excel_path = Path(excel_path).expanduser().resolve()
    if not root_dir.exists() or not root_dir.is_dir():
        raise FileNotFoundError(f"文件夹不存在：{root_dir}")
    if not excel_path.exists() or not excel_path.is_file():
        raise FileNotFoundError(f"Excel文件不存在：{excel_path}")
    if not is_supported_excel_file(excel_path):
        raise ValueError("仅支持 .xlsx 或 .xls 文件。")
    if header_row < 1:
        raise ValueError("表头行号必须大于等于 1。")
    if file_type not in FILE_TYPE_EXTENSIONS:
        raise ValueError(f"不支持的文件类型：{file_type}")

    names = _read_names_from_excel(excel_path, name_column, header_row)
    if not names:
        raise ValueError(f"Excel文件中未找到“{name_column}”列或该列无数据。")

    items = _list_items_for_excel_rename(
        root_dir,
        file_type=file_type,
        file_extensions=file_extensions,
        excel_path=excel_path,
    )
    operations, warnings = _plan_excel_batch_operations(items, names)
    _validate_confirmed_excel_preview(
        operations,
        warnings,
        expected_operations=expected_operations,
        expected_warnings=expected_warnings,
    )

    result = FolderRenameResult(
        root_dir=root_dir,
        mode=MODE_EXCEL_BATCH,
        dry_run=dry_run,
        operations=operations,
        warnings=warnings,
    )
    if dry_run:
        return result

    completed: list[FolderRenameOperation] = []
    runtime_warnings = list(warnings)
    for operation in operations:
        if not operation.source.exists():
            runtime_warnings.append(f"{operation.source.name} 已不存在，已跳过")
            continue
        if operation.target.exists():
            runtime_warnings.append(f"{operation.target.name} 执行前已存在，{operation.source.name} 已跳过，未覆盖原项目")
            continue
        try:
            operation.source.rename(operation.target)
        except OSError as exc:
            runtime_warnings.append(f"{operation.source.name} 改名失败：{exc}")
            continue
        completed.append(operation)
    result.operations = completed
    result.warnings = runtime_warnings
    return result


def _read_names_from_excel(excel_path: Path, name_column: str, header_row: int) -> list[str]:
    """从 Excel 文件中读取姓名列，保留非空数据行的原始顺序。"""
    import tempfile

    with tempfile.TemporaryDirectory() as temp_dir:
        working_path = ensure_xlsx_workbook(excel_path, Path(temp_dir))
        wb = load_workbook(working_path, data_only=True, read_only=True)
        try:
            ws = wb.active
            # 部分第三方工作簿会错误声明 A1:A1；让 read_only 模式按 XML
            # 实际内容迭代。表头只缓存有限单元格，姓名列则流式读取，避免
            # 大花名册被完整复制进内存。
            reset_dimensions = getattr(ws, "reset_dimensions", None)
            if callable(reset_dimensions):
                reset_dimensions()
            name_col = None
            matched_header_row = header_row
            normalized_name_column = _normalize_text(name_column)
            max_header_row = max(20, header_row)
            header_values = {
                row_index: values
                for row_index, values in enumerate(
                    ws.iter_rows(
                        min_row=1,
                        max_row=max_header_row,
                        min_col=1,
                        max_col=50,
                        values_only=True,
                    ),
                    start=1,
                )
            }
            candidate_rows = [header_row, *(row for row in range(1, 21) if row != header_row)]
            for candidate_row in candidate_rows:
                for col, value in enumerate(header_values.get(candidate_row, ()), start=1):
                    header = _normalize_text(value)
                    if header == normalized_name_column:
                        name_col = col
                        matched_header_row = candidate_row
                        break
                if name_col is not None:
                    break
            if name_col is None:
                for candidate_row in candidate_rows:
                    for col, value in enumerate(header_values.get(candidate_row, ()), start=1):
                        header = _normalize_text(value)
                        if normalized_name_column in header or header in ("姓名", "名字", "名称"):
                            name_col = col
                            matched_header_row = candidate_row
                            break
                    if name_col is not None:
                        break
            if name_col is None:
                return []
            names: list[str] = []
            for (value,) in ws.iter_rows(
                min_row=matched_header_row + 1,
                min_col=name_col,
                max_col=name_col,
                values_only=True,
            ):
                if value is not None:
                    name = str(value).strip()
                    if name:
                        names.append(name)
            return names
        finally:
            wb.close()


def _list_items_for_excel_rename(
    root_dir: Path,
    *,
    file_type: str,
    file_extensions: list[str] | None,
    excel_path: Path,
) -> list[Path]:
    """按人类可读的文件名顺序列出目录第一层的待改名项目。"""
    legacy_extensions = None
    if file_extensions is not None:
        legacy_extensions = {
            extension.lower() if extension.startswith(".") else f".{extension.lower()}"
            for extension in file_extensions
        }
    allowed_extensions = set(FILE_TYPE_EXTENSIONS[file_type])

    def matches(path: Path) -> bool:
        if path.name.startswith((".", "~$")):
            return False
        if legacy_extensions is not None:
            return path.is_file() and path.suffix.lower() in legacy_extensions
        if file_type == FILE_TYPE_FOLDER:
            return path.is_dir()
        if file_type == FILE_TYPE_ALL:
            return path.is_dir() or path.is_file()
        return path.is_file() and path.suffix.lower() in allowed_extensions

    items = []
    for path in root_dir.iterdir():
        if not matches(path):
            continue
        if path.is_file() and _is_excel_source_item(path, excel_path):
            continue
        items.append(path)
    return sorted(items, key=_natural_name_sort_key)


def _is_excel_source_item(path: Path, excel_path: Path) -> bool:
    """避免把放在待处理目录中的名单工作簿本身纳入改名。"""
    try:
        if path.samefile(excel_path):
            return True
    except OSError:
        pass

    # 项目执行时，目录副本和名单留存副本位于不同位置；相同文件名和内容
    # 可以确认它仍是名单副本，避免预览与正式执行的候选数量发生变化。
    if path.name != excel_path.name:
        return False
    try:
        if path.stat().st_size != excel_path.stat().st_size:
            return False
        return filecmp.cmp(path, excel_path, shallow=False)
    except OSError:
        return False


def _natural_name_sort_key(path: Path) -> tuple[tuple[int, int | str], ...]:
    parts = re.split(r"(\d+)", path.name.casefold())
    return tuple((1, int(part)) if part.isdigit() else (0, part) for part in parts)


def _plan_excel_batch_operations(
    items: list[Path],
    names: list[str],
) -> tuple[list[FolderRenameOperation], list[str]]:
    """按位置一一配对，冲突只跳过当前配对，不改变后续对应关系。"""
    warnings: list[str] = []

    if len(items) > len(names):
        unmatched_items = [path.name for path in items[len(names) :]]
        warnings.append(
            f"名单比筛选后的项目少 {len(unmatched_items)} 个；以下项目保持原名："
            f"{_warning_item_list(unmatched_items)}"
        )
    elif len(names) > len(items):
        unmatched_names = names[len(items) :]
        warnings.append(
            f"名单比筛选后的项目多 {len(unmatched_names)} 人；以下姓名没有对应项目："
            f"{_warning_item_list(unmatched_names)}"
        )

    planned: list[FolderRenameOperation] = []
    for source, new_name in zip(items, names):
        suffix = source.suffix if source.is_file() else ""
        target_name = f"{new_name}{suffix}"
        try:
            _validate_excel_target_name(target_name)
            planned.append(FolderRenameOperation(source=source, target=source.with_name(target_name)))
        except ValueError as exc:
            warnings.append(f"{source.name} 对应姓名“{new_name}”不能用于改名，已跳过：{exc}")
            continue

    return _filter_invalid_operations(planned, warnings, case_insensitive_targets=True), warnings


def _warning_item_list(values: list[str], *, limit: int = 8) -> str:
    shown = values[:limit]
    text = "、".join(shown)
    if len(values) > limit:
        text += f" 等共 {len(values)} 个"
    return text


def _validate_confirmed_excel_preview(
    operations: list[FolderRenameOperation],
    warnings: list[str],
    *,
    expected_operations: list[tuple[str, str]] | None,
    expected_warnings: list[str] | None,
) -> None:
    if expected_operations is None and expected_warnings is None:
        return
    actual_operations = [(operation.source.name, operation.target.name) for operation in operations]
    operations_changed = (
        expected_operations is not None
        and actual_operations != [tuple(pair) for pair in expected_operations]
    )
    warnings_changed = expected_warnings is not None and warnings != expected_warnings
    if operations_changed or warnings_changed:
        raise RuntimeError("待改名目录或人员名单在预览确认后发生了变化，本次未执行。请重新预览并确认。")


def _normalize_text(value: Any) -> str:
    """规范化文本：移除空白字符"""
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).replace("\xa0", "")


def _plan_operations(
    *,
    root_dir: Path,
    mode: str,
    text: str,
    target_name: str,
    replacement_name: str,
    file_type: str = FILE_TYPE_FOLDER,
) -> tuple[list[FolderRenameOperation], list[str]]:
    text = text.strip()
    target_name = target_name.strip()
    replacement_name = replacement_name.strip()
    warnings: list[str] = []

    if target_name:
        _validate_folder_name(target_name)

    if mode == MODE_APPEND:
        if not text:
            raise ValueError("请填写要追加的文字，建议以 - 或 _ 开头，例如：-劳动合同")
        suffix = _normalize_append_text(text)
        candidates = _iter_target_items(root_dir, target_name, file_type)
        planned = []
        for path in candidates:
            # 获取不含扩展名的名称（文件夹没有扩展名）
            stem = path.stem if path.is_file() else path.name
            ext = path.suffix if path.is_file() else ""
            if _already_has_suffix(stem, suffix, warnings):
                continue
            new_name = f"{stem}{suffix}{ext}"
            planned.append(_build_operation(path, new_name))
    elif mode == MODE_REMOVE:
        if not text:
            raise ValueError("请填写要删除的结尾文字，例如：劳动合同、-劳动合同 或 _身份证")
        candidates = _iter_target_items(root_dir, target_name, file_type)
        planned = []
        suffixes = _remove_suffix_candidates(text)
        for path in candidates:
            # 获取不含扩展名的名称
            stem = path.stem if path.is_file() else path.name
            ext = path.suffix if path.is_file() else ""
            suffix = _matching_remove_suffix(stem, suffixes)
            if suffix is None:
                continue
            new_stem = stem[: -len(suffix)]
            if not new_stem:
                warnings.append(f"{path.name} 删除后名称为空，已跳过")
                continue
            new_name = f"{new_stem}{ext}"
            planned.append(_build_operation(path, new_name))
    else:
        if not target_name:
            raise ValueError("请填写原名称，例如：张三")
        if not replacement_name:
            raise ValueError("请填写替换后的名称，例如：章五")
        source = root_dir / target_name
        if not source.exists():
            raise FileNotFoundError(f"未找到要替换的项目：{target_name}")
        # 对于文件，保留原扩展名
        if source.is_file():
            ext = source.suffix
            if not replacement_name.endswith(ext):
                replacement_name = f"{replacement_name}{ext}"
        planned = [_build_operation(source, replacement_name)]

    operations = _filter_invalid_operations(planned, warnings)
    return operations, warnings


def _iter_target_items(root_dir: Path, target_name: str, file_type: str) -> list[Path]:
    """根据文件类型返回待改名的项目列表"""
    extensions = FILE_TYPE_EXTENSIONS.get(file_type, [])

    def _matches_type(path: Path) -> bool:
        if file_type == FILE_TYPE_FOLDER:
            return path.is_dir()
        if file_type == FILE_TYPE_ALL:
            return path.is_dir() or path.is_file()
        # 按扩展名匹配文件
        return path.is_file() and path.suffix.lower() in extensions

    if target_name:
        target = root_dir / target_name
        if target.exists() and _matches_type(target):
            return [target]
        return [
            path
            for path in sorted(root_dir.iterdir())
            if _matches_type(path) and target_name in path.name
        ]
    return sorted(path for path in root_dir.iterdir() if _matches_type(path) and not path.name.startswith((".", "~$")))


def _normalize_append_text(text: str) -> str:
    return text


def _remove_suffix_candidates(text: str) -> list[str]:
    if text.startswith(("-", "_")):
        base = text[1:]
        candidates = [text, "-" + base, "_" + base, base]
    else:
        candidates = [text, "-" + text, "_" + text]
    return sorted(set(candidates), key=len, reverse=True)


def _matching_remove_suffix(folder_name: str, suffixes: list[str]) -> str | None:
    for suffix in suffixes:
        if suffix and folder_name.endswith(suffix):
            return suffix
    return None


def _already_has_suffix(folder_name: str, suffix: str, warnings: list[str]) -> bool:
    if folder_name.endswith(suffix):
        warnings.append(f"{folder_name} 已包含后缀，已跳过")
        return True
    return False


def _build_operation(source: Path, target_name: str) -> FolderRenameOperation:
    _validate_folder_name(target_name)
    return FolderRenameOperation(source=source, target=source.with_name(target_name))


def _validate_folder_name(name: str) -> None:
    if not name.strip():
        raise ValueError("文件夹名称不能为空")
    if name in {".", ".."}:
        raise ValueError(f"文件夹名称不合法：{name}")
    if any(char in WINDOWS_INVALID_CHARS for char in name):
        raise ValueError(f"文件夹名称包含 Windows 不支持的字符：{name}")
    if any(ord(char) < 32 for char in name):
        raise ValueError(f"文件夹名称包含 Windows 不支持的控制字符：{name}")
    if name.endswith((" ", ".")):
        raise ValueError(f"文件夹名称不能以空格或句点结尾：{name}")
    if name.split(".", 1)[0].upper() in WINDOWS_RESERVED_NAMES:
        raise ValueError(f"文件夹名称是 Windows 保留名称：{name}")


def _validate_excel_target_name(name: str) -> None:
    _validate_folder_name(name)


def _filter_invalid_operations(
    operations: list[FolderRenameOperation],
    warnings: list[str],
    *,
    case_insensitive_targets: bool = False,
) -> list[FolderRenameOperation]:
    valid: list[FolderRenameOperation] = []
    seen_targets: set[object] = set()
    for operation in operations:
        if operation.source == operation.target:
            warnings.append(f"{operation.source.name} 改名前后相同，已跳过")
            continue
        target_key: object = operation.target
        if case_insensitive_targets:
            target_key = (str(operation.target.parent.resolve()).casefold(), operation.target.name.casefold())
        if target_key in seen_targets:
            warnings.append(f"{operation.target.name} 目标名称重复，{operation.source.name} 已跳过")
            continue
        if operation.target.exists():
            warnings.append(f"{operation.target.name} 已存在，{operation.source.name} 已跳过")
            continue
        seen_targets.add(target_key)
        valid.append(operation)
    return valid
