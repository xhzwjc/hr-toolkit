from __future__ import annotations

import re
import weakref
from copy import copy
from dataclasses import dataclass
from typing import Any, NamedTuple

from openpyxl.formula.translate import Translator
from openpyxl.styles.cell_style import StyleArray
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


_DEFAULT_STYLE_ARRAY = StyleArray()


def _workbook_ref(ws: Worksheet) -> Any:
    """Weakly reference the owning workbook so snapshots never keep it alive."""
    workbook = getattr(ws, "parent", None)
    if workbook is None:
        return None
    try:
        return weakref.ref(workbook)
    except TypeError:
        return None


def cached_style_id(ws: Worksheet, kind: str, tag: Any, factory) -> int:
    """Resolve a style object to its workbook style-table index, once per tag.

    ``cell.border = X`` 看着便宜，实际上每次都要对 X 做一次 ``hash()``——
    openpyxl 的 Border/Alignment 是 Serialisable，哈希要递归遍历全部字段。
    整表逐格赋同一个边框时，这个哈希会重复几十万次。样式表是只增不改的，
    下标一旦拿到就长期有效，所以按调用方给的 ``tag`` 缓存下标，之后直接写
    ``StyleArray`` 里的整数即可。

    ``tag`` 必须由调用方保证与样式值一一对应（例如固定常量用字符串标签，
    派生样式用来源下标），不要用 ``id()`` 之类可能被回收复用的键。
    """
    workbook = getattr(ws, "parent", None)
    if workbook is None:
        raise ValueError("工作表没有关联的工作簿，无法解析样式下标。")
    cache = getattr(workbook, "_hr_style_id_cache", None)
    if cache is None:
        cache = {}
        try:
            workbook._hr_style_id_cache = cache
        except AttributeError:
            cache = None
    key = (kind, tag)
    if cache is not None and key in cache:
        return cache[key]
    collection = {
        "border": workbook._borders,
        "alignment": workbook._alignments,
        "font": workbook._fonts,
        "fill": workbook._fills,
    }[kind]
    style_id = collection.add(factory())
    if cache is not None:
        cache[key] = style_id
    return style_id


def style_source_id(cell, kind: str) -> int:
    """Return the cell's current index into the given workbook style table."""
    style_array = cell._style
    if style_array is None:
        return 0
    return getattr(style_array, {"border": "borderId", "alignment": "alignmentId", "font": "fontId", "fill": "fillId"}[kind])


def set_style_ids(
    cell,
    *,
    border_id: int | None = None,
    alignment_id: int | None = None,
    font_id: int | None = None,
    fill_id: int | None = None,
) -> None:
    """Write style-table indices straight into the cell's StyleArray.

    先复制再写：openpyxl 读取工作簿时可能让多个单元格共用同一个 StyleArray，
    原地修改会顺带改掉别的单元格。
    """
    style_array = cell._style
    updated = copy(style_array) if style_array is not None else StyleArray()
    if border_id is not None:
        updated.borderId = border_id
    if alignment_id is not None:
        updated.alignmentId = alignment_id
    if font_id is not None:
        updated.fontId = font_id
    if fill_id is not None:
        updated.fillId = fill_id
    cell._style = updated


def _style_translation_cache(ws: Worksheet, snapshot: "RowSnapshot") -> dict | None:
    """Map a specific source workbook's style indices to ones registered here.

    样式下标只在所属工作簿内有意义，不同来源工作簿的同一组下标含义可能完全
    不同，所以缓存必须按来源工作簿分桶；用弱引用键，来源关闭后自动释放。
    """
    workbook = getattr(ws, "parent", None)
    reference = snapshot.source_workbook
    source = reference() if reference is not None else None
    if workbook is None or source is None:
        return None
    caches = getattr(workbook, "_hr_style_translation", None)
    if caches is None:
        caches = weakref.WeakKeyDictionary()
        try:
            workbook._hr_style_translation = caches
        except AttributeError:
            return None
    cache = caches.get(source)
    if cache is None:
        cache = {}
        caches[source] = cache
    return cache


def _is_same_workbook(ws: Worksheet, snapshot: "RowSnapshot") -> bool:
    reference = snapshot.source_workbook
    if reference is None:
        return False
    source = reference()
    return source is not None and source is getattr(ws, "parent", None)


@dataclass(frozen=True)
class CellSnapshot:
    value: Any
    font: Any
    fill: Any
    border: Any
    alignment: Any
    number_format: str
    protection: Any
    style: Any = None


@dataclass(frozen=True)
class RowSnapshot:
    source_row: int
    height: float | None
    cells: list[CellSnapshot]
    source_workbook: Any = None


def _style_pool(ws: Worksheet) -> tuple[Any, ...] | None:
    """Return the workbook's shared style collections, or None if unavailable."""
    workbook = getattr(ws, "parent", None)
    if workbook is None:
        return None
    try:
        return (
            workbook._fonts,
            workbook._fills,
            workbook._borders,
            workbook._alignments,
            workbook._protections,
        )
    except AttributeError:
        return None


def snapshot_row(ws: Worksheet, row_index: int, max_column: int) -> RowSnapshot:
    """Capture a row's values and formatting.

    openpyxl 单元格只保存指向工作簿共享样式表的整数下标（``cell._style``），
    ``cell.font`` 每次都会新建一个 StyleProxy，再 ``copy()`` 就是一次深拷贝。
    这里改为直接按下标取出共享样式对象：样式对象在 openpyxl 里是只读的
    （StyleProxy 禁止赋值），多个单元格共用同一个实例完全安全，因此不必拷贝。
    同时记录整份 ``StyleArray``，同一工作簿内可以整行直接套用下标，省掉逐格
    的哈希与登记。
    """
    pool = _style_pool(ws)
    cells: list[CellSnapshot] = []
    for col_index in range(1, max_column + 1):
        cell = ws.cell(row_index, col_index)
        style_array = cell._style
        if pool is None:
            font = copy(cell.font)
            fill = copy(cell.fill)
            border = copy(cell.border)
            alignment = copy(cell.alignment)
            protection = copy(cell.protection)
        else:
            fonts, fills, borders, alignments, protections = pool
            indices = style_array if style_array is not None else _DEFAULT_STYLE_ARRAY
            font = fonts[indices.fontId]
            fill = fills[indices.fillId]
            border = borders[indices.borderId]
            alignment = alignments[indices.alignmentId]
            protection = protections[indices.protectionId]
        cells.append(
            CellSnapshot(
                value=cell.value,
                font=font,
                fill=fill,
                border=border,
                alignment=alignment,
                number_format=cell.number_format,
                protection=protection,
                style=copy(style_array) if style_array is not None else None,
            )
        )
    return RowSnapshot(
        source_row=row_index,
        height=ws.row_dimensions[row_index].height,
        cells=cells,
        source_workbook=_workbook_ref(ws),
    )


def apply_row_snapshot(
    ws: Worksheet,
    target_row: int,
    snapshot: RowSnapshot,
    *,
    translate_formulas: bool = True,
) -> None:
    ws.row_dimensions[target_row].height = snapshot.height
    # 样式下标只在所属工作簿内有效；跨工作簿（例如工资表拆分写入新文件）要先把
    # 样式登记进目标工作簿。同一份快照通常要套用到成百上千行，样式组合却只有
    # 寥寥几种，所以按来源下标缓存"已登记好的 StyleArray"，每种只付一次代价。
    same_workbook = _is_same_workbook(ws, snapshot)
    translation = None if same_workbook else _style_translation_cache(ws, snapshot)
    for col_index, snap in enumerate(snapshot.cells, start=1):
        cell = ws.cell(target_row, col_index)
        if same_workbook:
            cell._style = copy(snap.style) if snap.style is not None else StyleArray()
        else:
            key = tuple(snap.style) if snap.style is not None else None
            translated = translation.get(key) if translation is not None else None
            if translated is None:
                cell.font = snap.font
                cell.fill = snap.fill
                cell.border = snap.border
                cell.alignment = snap.alignment
                cell.number_format = snap.number_format
                cell.protection = snap.protection
                if translation is not None:
                    translation[key] = copy(cell._style) if cell._style is not None else StyleArray()
            else:
                cell._style = copy(translated)

        value = snap.value
        if translate_formulas and isinstance(value, str) and value.startswith("="):
            origin = f"{get_column_letter(col_index)}{snapshot.source_row}"
            destination = f"{get_column_letter(col_index)}{target_row}"
            try:
                value = Translator(value, origin=origin).translate_formula(destination)
            except Exception:
                value = _translate_same_row_formula(value, snapshot.source_row, target_row)
        cell.value = value


class _GridCell(NamedTuple):
    value: Any


class SheetGrid:
    """把工作表一次性读入内存的轻量值网格。

    openpyxl 的 read_only 模式只适合顺序读取：ws.cell(row, col) 每次都会
    从头重新解析工作表 XML，随机访问会退化成 O(行数²)，大文件要跑几分钟。
    先用 iter_rows 单遍读完，之后在内存里随机访问，行列号仍为 1 起始。

    同时提供 ws.cell(row, col).value 形式的兼容接口，便于原有按
    Worksheet 编写的读取函数直接换用。
    """

    __slots__ = ("title", "max_row", "max_column", "_rows")

    def __init__(self, ws: Any) -> None:
        self.title: str = ws.title
        self._rows: list[tuple[Any, ...]] = [tuple(row) for row in ws.iter_rows(values_only=True)]
        self.max_row: int = len(self._rows)
        self.max_column: int = max((len(row) for row in self._rows), default=0)

    def value(self, row_index: int, col_index: int) -> Any:
        if not 1 <= row_index <= self.max_row or col_index < 1:
            return None
        row = self._rows[row_index - 1]
        return row[col_index - 1] if col_index <= len(row) else None

    def cell(self, row_index: int, col_index: int) -> _GridCell:
        return _GridCell(self.value(row_index, col_index))


def insert_rows(ws: Worksheet, idx: int, amount: int = 1) -> None:
    """在第 idx 行前插入 amount 个空行，内存占用与已用单元格数成正比。

    openpyxl 自带的 ``ws.insert_rows`` 会先 ``list(iter_rows(min_row=idx))``，
    把插入位置以下、所有列的空单元格全部物化成对象再逐个搬动。几千行 × 几十列
    时会瞬间生成上百万个单元格对象而 MemoryError（见 openpyxl
    ``worksheet._move_cells``）。这里只搬动实际存在的稀疏单元格，行为与
    openpyxl 保持一致——同样不改动已有公式引用、不移动合并单元格与行高。
    """
    # ws.max_row 会把所有单元格的行号收集成 set 再取最大值，逐行插入时这一步
    # 本身就是 O(单元格数)，几千次插入即退化成平方级；改由 _shift_cells 在原本
    # 就要做的那一趟遍历里顺带算出来。
    new_max_row = _shift_cells(ws, min_index=idx, amount=amount, is_row=True)
    if new_max_row is not None:
        ws._current_row = new_max_row


def insert_cols(ws: Worksheet, idx: int, amount: int = 1) -> None:
    """在第 idx 列前插入 amount 个空列，是 :func:`insert_rows` 的列向版本。"""
    _shift_cells(ws, min_index=idx, amount=amount, is_row=False)


def _shift_cells(ws: Worksheet, *, min_index: int, amount: int, is_row: bool) -> int | None:
    """Shift existing cells and return the sheet's resulting max row index."""
    if amount <= 0:
        return None
    cells = ws._cells
    # 只挑出受影响的已存在单元格；从远端向 min_index 方向搬，避免覆盖目标位置
    axis = 0 if is_row else 1
    affected = []
    max_row_before = 0
    for key in cells:
        if key[axis] >= min_index:
            affected.append(key)
        if key[0] > max_row_before:
            max_row_before = key[0]
    affected.sort(key=lambda key: key[axis], reverse=True)
    for row, col in affected:
        cell = cells.pop((row, col))
        if is_row:
            row += amount
        else:
            col += amount
        cell.row = row
        cell.column = col
        cells[(row, col)] = cell
    if not cells:
        return 1
    # 行向插入时，最大行号本身若在插入点之后，会跟着整体下移
    if is_row and max_row_before >= min_index:
        return max_row_before + amount
    return max(max_row_before, 1)


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clone_style(source_cell, target_cell) -> None:
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def unmerge_ranges_from_row(ws: Worksheet, min_row: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= min_row:
            ws.unmerge_cells(str(merged_range))


def _translate_same_row_formula(formula: str, source_row: int, target_row: int) -> str:
    cell_ref_pattern = re.compile(r"(?<![A-Za-z0-9_])(\$?[A-Z]{1,3})(\$?)(\d+)(?![A-Za-z0-9_])")

    def replace_row(match: re.Match[str]) -> str:
        column_ref, row_anchor, row_number = match.groups()
        if row_anchor or int(row_number) != source_row:
            return match.group(0)
        return f"{column_ref}{target_row}"

    return cell_ref_pattern.sub(replace_row, formula)
