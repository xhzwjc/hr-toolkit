from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import hashlib
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Callable
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile


SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xls"}
_SPREADSHEET_XML_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_SPREADSHEET_XML_TAG = f"{{{_SPREADSHEET_XML_NAMESPACE}}}"


@dataclass(frozen=True)
class XlsxSaveCompatibilitySnapshot:
    """保存 openpyxl 无法完整往返的工作簿级样式表。"""

    cell_style_xfs: bytes
    cell_styles: bytes


def is_supported_excel_file(path: Path) -> bool:
    return path.suffix.lower() in SUPPORTED_EXCEL_SUFFIXES and not path.name.startswith(("~$", ".~"))


def capture_xlsx_save_compatibility(path: Path) -> XlsxSaveCompatibilitySnapshot:
    """在 openpyxl 保存前快照其可能重排或丢弃的样式父表。"""

    try:
        with ZipFile(path) as archive:
            styles_xml = archive.read("xl/styles.xml")
        root = ElementTree.fromstring(styles_xml)
        cell_style_xfs = root.find(f"{_SPREADSHEET_XML_TAG}cellStyleXfs")
        cell_styles = root.find(f"{_SPREADSHEET_XML_TAG}cellStyles")
    except (BadZipFile, KeyError, ElementTree.ParseError, OSError) as exc:
        raise RuntimeError(f"无法读取 Excel 样式结构：{path}") from exc
    if cell_style_xfs is None or cell_styles is None:
        raise RuntimeError(f"Excel 文件缺少必要样式结构：{path}")
    return XlsxSaveCompatibilitySnapshot(
        cell_style_xfs=ElementTree.tostring(cell_style_xfs, encoding="utf-8"),
        cell_styles=ElementTree.tostring(cell_styles, encoding="utf-8"),
    )


def finalize_xlsx_after_openpyxl_save(
    output_path: Path,
    snapshot: XlsxSaveCompatibilitySnapshot | None,
) -> None:
    """修复 openpyxl 保存旧版 Excel 模板时产生的 OOXML 兼容性退化。

    openpyxl 会把部分非连续 ``cellStyleXfs`` 压缩掉，却保留单元格原来的
    ``xfId``；同时会把绘图几何中的 ``a:avLst`` 写成错误的默认命名空间。
    这两类问题在新版软件中通常会被静默容错，但旧版 Excel 可能提示修复文件。
    """

    output_path = Path(output_path)
    temp_path: Path | None = None
    try:
        with ZipFile(output_path) as source_archive:
            styles_xml = source_archive.read("xl/styles.xml")
            repaired_styles = _restore_openpyxl_style_tables(styles_xml, snapshot)
            drawing_names = [
                name
                for name in source_archive.namelist()
                if name.startswith("xl/drawings/") and name.endswith(".xml")
            ]
            drawing_repairs = {
                name
                for name in drawing_names
                if b"<avLst" in source_archive.read(name)
            }
            if repaired_styles == styles_xml and not drawing_repairs:
                return

            output_path.parent.mkdir(parents=True, exist_ok=True)
            with tempfile.NamedTemporaryFile(
                prefix=f".{output_path.name}.",
                suffix=".repairing",
                dir=output_path.parent,
                delete=False,
            ) as temp_file:
                temp_path = Path(temp_file.name)
            with ZipFile(temp_path, "w") as target_archive:
                target_archive.comment = source_archive.comment
                for info in source_archive.infolist():
                    data = source_archive.read(info.filename)
                    if info.filename == "xl/styles.xml":
                        data = repaired_styles
                    elif info.filename in drawing_repairs:
                        if b"xmlns:a=" not in data:
                            raise RuntimeError(f"Excel 绘图缺少 a 命名空间：{info.filename}")
                        data = data.replace(b"<avLst", b"<a:avLst").replace(
                            b"</avLst>",
                            b"</a:avLst>",
                        )
                    target_archive.writestr(info, data)
        os.replace(temp_path, output_path)
        temp_path = None
    except (BadZipFile, KeyError, ElementTree.ParseError, OSError, ValueError) as exc:
        raise RuntimeError(f"Excel 输出兼容性校验失败：{output_path}") from exc
    finally:
        if temp_path is not None and temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def _restore_openpyxl_style_tables(
    styles_xml: bytes,
    snapshot: XlsxSaveCompatibilitySnapshot | None,
) -> bytes:
    root = ElementTree.fromstring(styles_xml)
    cell_style_xfs = root.find(f"{_SPREADSHEET_XML_TAG}cellStyleXfs")
    cell_xfs = root.find(f"{_SPREADSHEET_XML_TAG}cellXfs")
    cell_styles = root.find(f"{_SPREADSHEET_XML_TAG}cellStyles")
    if cell_style_xfs is None or cell_xfs is None or cell_styles is None:
        raise RuntimeError("Excel 输出缺少必要样式结构。")

    style_count = len(cell_style_xfs)
    raw_style_ids = [xf.get("xfId") for xf in cell_xfs if xf.get("xfId") is not None]
    referenced_style_ids = {int(xf_id) for xf_id in raw_style_ids if xf_id and xf_id.isdigit()}
    if all(xf_id and xf_id.isdigit() and int(xf_id) < style_count for xf_id in raw_style_ids):
        return styles_xml

    restored_source_tables = False
    if snapshot is not None:
        source_cell_style_xfs = ElementTree.fromstring(snapshot.cell_style_xfs)
        source_cell_styles = ElementTree.fromstring(snapshot.cell_styles)
        output_style_names = {style.get("name") for style in cell_styles}
        source_style_names = {style.get("name") for style in source_cell_styles}
        if (
            all(xf_id and xf_id.isdigit() for xf_id in raw_style_ids)
            and all(xf_id < len(source_cell_style_xfs) for xf_id in referenced_style_ids)
            and not output_style_names - source_style_names
        ):
            for current, restored in (
                (cell_style_xfs, source_cell_style_xfs),
                (cell_styles, source_cell_styles),
            ):
                index = list(root).index(current)
                root.remove(current)
                root.insert(index, restored)
            restored_source_tables = True

    if not restored_source_tables:
        # 即使源样式父表无法恢复，也要保证文件结构可被旧版 Excel 打开。
        # cellXfs 已包含单元格的直接格式，失效的父样式引用降为 Normal 即可。
        for xf in cell_xfs:
            xf_id = xf.get("xfId")
            if xf_id is None or not xf_id.isdigit() or int(xf_id) >= style_count:
                xf.set("xfId", "0")
    ElementTree.register_namespace("", _SPREADSHEET_XML_NAMESPACE)
    return ElementTree.tostring(root, encoding="utf-8")


def ensure_xlsx_workbook(
    path: Path,
    temp_dir: Path,
    *,
    preserve_formatting: bool = False,
    warning_callback: Callable[[str], None] | None = None,
) -> Path:
    path = Path(path).expanduser().resolve()
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_EXCEL_SUFFIXES:
        raise ValueError(f"Excel 文件仅支持 .xlsx 或 .xls：{path}")

    file_kind = _detect_excel_file_kind(path)
    if suffix == ".xlsx" and file_kind == "xlsx":
        return path
    output_dir = _conversion_dir(path, temp_dir, preserve_formatting=preserve_formatting)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{path.stem}.xlsx"
    if output_path.exists():
        if _is_usable_xlsx(output_path):
            return output_path
        _remove_incomplete_conversion(output_path)

    if file_kind == "xlsx":
        shutil.copyfile(path, output_path)
    else:
        _convert_xls_to_xlsx(
            path,
            output_path,
            temp_dir=temp_dir,
            preserve_formatting=preserve_formatting,
            warning_callback=warning_callback,
        )
    if not _is_usable_xlsx(output_path):
        raise RuntimeError(f".xls 转换失败，未生成文件：{output_path}")
    return output_path


def _detect_excel_file_kind(path: Path) -> str:
    with path.open("rb") as file:
        header = file.read(8)
    if header.startswith(b"PK\x03\x04"):
        return "xlsx"
    if header.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "xls"
    return path.suffix.lower().lstrip(".")


def _conversion_dir(path: Path, temp_dir: Path, *, preserve_formatting: bool = False) -> Path:
    mode = "formatted" if preserve_formatting else "values"
    digest = hashlib.sha1(f"{path}\0{mode}".encode("utf-8")).hexdigest()[:12]
    return temp_dir / "xls_converted" / digest


def _convert_with_xlrd(source: Path, output_path: Path) -> None:
    """使用纯 Python xlrd + openpyxl 在内存中将 .xls 转为 .xlsx。

    特点：
    1. 纯只读文件流解析，绝不修改或触碰源文件（100% 解决 Windows 下批次文件校验哈希变化的问题）；
    2. 无需启动外部 Excel/WPS/COM 进程，速度快（0.01 秒级），跨平台稳定。
    """
    import xlrd
    import openpyxl

    rb = xlrd.open_workbook(str(source), formatting_info=False)
    wb = openpyxl.Workbook()
    try:
        # 移除默认新建的 Sheet
        wb.remove(wb.active)

        for sheet_name in rb.sheet_names():
            rs = rb.sheet_by_name(sheet_name)
            ws = wb.create_sheet(title=sheet_name)
            for row_idx in range(rs.nrows):
                row_vals: list[object] = []
                for col_idx in range(rs.ncols):
                    cell = rs.cell(row_idx, col_idx)
                    val = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt_tuple = xlrd.xldate_as_tuple(val, rb.datemode)
                            if dt_tuple[3:] == (0, 0, 0):
                                val = date(dt_tuple[0], dt_tuple[1], dt_tuple[2])
                            else:
                                val = datetime(*dt_tuple)
                        except Exception:
                            pass
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        val = bool(val)
                    elif cell.ctype == xlrd.XL_CELL_ERROR:
                        val = None
                    elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                        val = None
                    row_vals.append(val)
                ws.append(row_vals)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
    finally:
        wb.close()
        rb.release_resources()


def _convert_xls_to_xlsx(
    source: Path,
    output_path: Path,
    temp_dir: Path | None = None,
    *,
    preserve_formatting: bool = False,
    warning_callback: Callable[[str], None] | None = None,
) -> None:
    errors: list[str] = []

    if not preserve_formatting:
        # 仅提取数据的场景优先走纯 Python 快速路径；该路径不复制工作簿样式。
        if _try_xls_converter(
            "内置 xlrd 转换",
            _convert_with_xlrd,
            source,
            output_path,
            errors,
        ):
            return

    # 优先让 Excel/WPS/LibreOffice 在只读沙箱副本上保留格式、公式和结构。
    # 这些组件不可用时，格式不能成为阻断项，后面会自动退回 xlrd 数据模式。
    sandbox_dir = (temp_dir or output_path.parent) / "xls_sandbox"
    digest = hashlib.sha1(str(source).encode("utf-8")).hexdigest()[:8]
    sandbox_source = sandbox_dir / f"{digest}_{source.name}"
    try:
        sandbox_dir.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(source, sandbox_source)
    except Exception as exc:
        errors.append(f"创建只读转换副本失败：{exc}")
    else:
        try:
            if sys.platform.startswith("win") and _try_xls_converter(
                "Excel/WPS 转换",
                _convert_with_windows_com,
                sandbox_source,
                output_path,
                errors,
            ):
                return
            if _try_xls_converter(
                "LibreOffice 转换",
                _convert_with_libreoffice,
                sandbox_source,
                output_path,
                errors,
            ):
                return
        finally:
            if sandbox_source.exists():
                try:
                    sandbox_source.unlink()
                except OSError:
                    pass

    if preserve_formatting and _try_xls_converter(
        "内置 xlrd 兼容转换",
        _convert_with_xlrd,
        source,
        output_path,
        errors,
    ):
        _report_conversion_warning(
            warning_callback,
            f"{source.name} 无法使用本机表格组件完整保留格式，已自动切换兼容模式继续生成；"
            "数据会正常输出，但原表颜色、边框、换行和公式格式可能简化。",
        )
        return

    message = (
        "无法读取 .xls 文件内容。请确认文件未损坏、未加密，"
        "或先在 Excel/WPS 中另存为 .xlsx 后重试。"
    )
    raise RuntimeError(message + (" 详细信息：" + "；".join(errors) if errors else ""))


def _try_xls_converter(
    label: str,
    converter: Callable[[Path, Path], None],
    source: Path,
    output_path: Path,
    errors: list[str],
) -> bool:
    _remove_incomplete_conversion(output_path)
    try:
        converter(source, output_path)
    except Exception as exc:
        errors.append(f"{label}失败：{exc}")
        _remove_incomplete_conversion(output_path)
        return False
    if _is_usable_xlsx(output_path):
        return True
    errors.append(f"{label}失败：未生成有效的 .xlsx 文件")
    _remove_incomplete_conversion(output_path)
    return False


def _is_usable_xlsx(path: Path) -> bool:
    try:
        if not path.exists() or path.stat().st_size <= 0:
            return False
        with ZipFile(path) as archive:
            names = set(archive.namelist())
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                return False
            return archive.testzip() is None
    except Exception:
        # 这里只负责判断转换产物是否可用。旧系统上的压缩库、异常压缩方式或
        # 不完整文件都应触发下一个兼容转换方案，不能把校验异常直接暴露给用户。
        return False


def _remove_incomplete_conversion(path: Path) -> None:
    if not path.exists():
        return
    try:
        path.unlink()
    except OSError:
        pass


def _report_conversion_warning(
    callback: Callable[[str], None] | None,
    message: str,
) -> None:
    if callback is None:
        return
    try:
        callback(message)
    except Exception:
        # 警告展示失败不能反过来阻断已经成功的数据转换。
        pass


def _convert_with_windows_com(source: Path, output_path: Path) -> None:
    try:
        import pythoncom  # type: ignore[import-not-found]
        import win32com.client  # type: ignore[import-not-found]
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "缺少 Windows .xls 转换依赖 pywin32。请在 Windows 打包环境执行 "
            "`python -m pip install -r requirements.txt` 后重新打包。"
        ) from exc

    pythoncom.CoInitialize()
    app = None
    workbook = None
    try:
        last_error: Exception | None = None
        for prog_id in ("Excel.Application", "Ket.Application", "KET.Application", "ET.Application", "et.Application"):
            try:
                app = win32com.client.DispatchEx(prog_id)
                break
            except Exception as exc:
                last_error = exc
        if app is None:
            raise RuntimeError(f"未找到 Excel 或 WPS COM 组件：{last_error}")
        app.Visible = False
        app.DisplayAlerts = False
        workbook = app.Workbooks.Open(
            str(source),
            UpdateLinks=0,
            ReadOnly=True,
            AddToMru=False,
        )
        workbook.SaveAs(str(output_path), FileFormat=51)
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if app is not None:
            try:
                app.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def _convert_with_libreoffice(source: Path, output_path: Path) -> None:
    executable = _find_libreoffice()
    if executable is None:
        raise RuntimeError(_diagnose_missing_libreoffice())
    output_path.parent.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(
        [
            executable,
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(output_path.parent),
            str(source),
        ],
        check=False,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        env=_build_conversion_env(),
        timeout=300,
    )
    if result.returncode != 0:
        raise RuntimeError((result.stderr or result.stdout or "").strip() or f"退出码 {result.returncode}")
    converted = output_path.parent / f"{source.stem}.xlsx"
    if converted.exists() and converted != output_path:
        shutil.move(str(converted), str(output_path))


def _find_libreoffice() -> str | None:
    """Locate the LibreOffice / soffice executable on the host.

    macOS .app bundles launched from Finder inherit a minimal ``PATH`` that
    only contains ``/usr/bin:/bin:/usr/sbin:/sbin``. That means Homebrew
    installs at ``/opt/homebrew/bin`` (Apple Silicon) and
    ``/usr/local/bin`` (Intel) — the two most common install locations —
    are *not* visible to ``shutil.which`` even when ``soffice`` is installed
    on the machine. Without these fallbacks the user sees a confusing
    "未找到 libreoffice/soffice 命令" error even though the binary exists.
    """

    search_paths: list[str] = []
    path_env = os.environ.get("PATH", "")
    if path_env:
        search_paths.extend(path_env.split(os.pathsep))

    homebrew_candidates = (
        "/opt/homebrew/bin",
        "/usr/local/bin",
        "/opt/local/bin",
    )
    for candidate in homebrew_candidates:
        if candidate not in search_paths:
            search_paths.append(candidate)

    mac_app_paths = (
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "~/Applications/LibreOffice.app/Contents/MacOS/soffice",
    )
    for candidate in mac_app_paths:
        expanded = os.path.expanduser(candidate)
        if Path(expanded).exists():
            return expanded

    for command in ("soffice", "libreoffice"):
        executable = shutil.which(command, path=os.pathsep.join(search_paths))
        if executable:
            return executable

    return None


def _build_conversion_env() -> dict[str, str]:
    """Provide a stable PATH so subprocess can locate LibreOffice on macOS.

    Rebuilds ``PATH`` from the current environment plus the well-known
    Homebrew prefixes so the spawned ``soffice`` process can find its own
    resources (system fonts, locale data, etc.) — not just the binary.
    """

    env = os.environ.copy()
    candidates = ("/opt/homebrew/bin", "/usr/local/bin", "/opt/local/bin")
    current_path = env.get("PATH", "")
    parts = [p for p in current_path.split(os.pathsep) if p]
    for candidate in candidates:
        if Path(candidate).exists() and candidate not in parts:
            parts.append(candidate)
    env["PATH"] = os.pathsep.join(parts) if parts else current_path
    return env


def _diagnose_missing_libreoffice() -> str:
    """Return a helpful explanation when LibreOffice cannot be located."""

    path_env = os.environ.get("PATH", "")
    checked: list[str] = []
    for candidate in (
        "/opt/homebrew/bin/soffice",
        "/usr/local/bin/soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ):
        checked.append(f"{candidate}={'存在' if Path(candidate).exists() else '缺失'}")
    return (
        "未找到 libreoffice/soffice 命令。当前 PATH="
        f"{path_env or '<空>'}；已检查："
        + "，".join(checked)
        + "。请安装 LibreOffice（brew install --cask libreoffice），或将其可执行文件添加到 PATH。"
    )
