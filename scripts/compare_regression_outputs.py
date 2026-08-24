from __future__ import annotations

import argparse
import base64
import hashlib
import json
import math
import os
import sys
import tempfile
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Sequence
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook


SCHEMA_VERSION = 1
EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
DIFF_LIMIT_DEFAULT = 200
DISPLAY_VALUE_MAX = 500
LIST_IDENTITY_KEYS = ("coordinate", "title", "name", "key", "index", "range")


class RegressionComparisonError(RuntimeError):
    """Raised when a regression baseline cannot be captured or verified."""


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalized_scalar(value: Any) -> Any:
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        if math.isnan(value):
            return {"type": "float", "value": "nan"}
        if math.isinf(value):
            return {"type": "float", "value": "inf" if value > 0 else "-inf"}
        return value
    if isinstance(value, Decimal):
        return {"type": "decimal", "value": str(value)}
    if isinstance(value, (datetime, date, time)):
        return {"type": value.__class__.__name__, "value": value.isoformat()}
    if isinstance(value, bytes):
        return {
            "type": "bytes",
            "value": base64.b64encode(value).decode("ascii"),
        }
    if isinstance(value, (list, tuple)):
        return [_normalized_scalar(item) for item in value]
    if isinstance(value, dict):
        return {
            str(key): _normalized_scalar(item)
            for key, item in sorted(value.items(), key=lambda pair: str(pair[0]))
        }
    to_tree = getattr(value, "to_tree", None)
    if callable(to_tree):
        return {
            "type": f"{value.__class__.__module__}.{value.__class__.__qualname__}",
            "xml": _xml_snapshot(to_tree()),
        }
    attributes = getattr(value, "__dict__", None)
    if isinstance(attributes, dict):
        return {
            "type": f"{value.__class__.__module__}.{value.__class__.__qualname__}",
            "attributes": {
                str(key): _normalized_scalar(item)
                for key, item in sorted(attributes.items())
                if not str(key).startswith("_")
            },
        }
    return {
        "type": f"{value.__class__.__module__}.{value.__class__.__qualname__}",
        "value": str(value),
    }


def _xml_snapshot(element: Any) -> dict[str, Any]:
    tag = str(getattr(element, "tag", ""))
    if "}" in tag:
        tag = tag.rsplit("}", 1)[-1]
    attributes = {
        str(key).rsplit("}", 1)[-1]: str(value)
        for key, value in sorted(getattr(element, "attrib", {}).items())
    }
    text_value = getattr(element, "text", None)
    text = str(text_value) if text_value not in (None, "") else None
    return {
        "tag": tag,
        "attributes": attributes,
        "text": text,
        "children": [_xml_snapshot(child) for child in list(element)],
    }


def _serializable_snapshot(value: Any) -> Any:
    if value is None:
        return None
    to_tree = getattr(value, "to_tree", None)
    if callable(to_tree):
        return _xml_snapshot(to_tree())
    return _normalized_scalar(value)


def _style_snapshot(cell: Any) -> dict[str, Any] | None:
    if not cell.has_style:
        return None
    return {
        "font": _serializable_snapshot(cell.font),
        "fill": _serializable_snapshot(cell.fill),
        "border": _serializable_snapshot(cell.border),
        "alignment": _serializable_snapshot(cell.alignment),
        "protection": _serializable_snapshot(cell.protection),
        "number_format": cell.number_format,
        "quote_prefix": bool(cell.quotePrefix),
        "pivot_button": bool(cell.pivotButton),
    }


def _comment_snapshot(comment: Any) -> dict[str, Any] | None:
    if comment is None:
        return None
    return {
        "text": comment.text,
        "author": comment.author,
        "width": comment.width,
        "height": comment.height,
    }


def _hyperlink_snapshot(hyperlink: Any) -> dict[str, Any] | None:
    if hyperlink is None:
        return None
    return {
        "target": hyperlink.target,
        "location": hyperlink.location,
        "tooltip": hyperlink.tooltip,
        "display": hyperlink.display,
    }


def _cell_snapshot(cell: Any) -> dict[str, Any]:
    return {
        "coordinate": cell.coordinate,
        "value": _normalized_scalar(cell.value),
        "data_type": cell.data_type,
        "style": _style_snapshot(cell),
        "comment": _comment_snapshot(cell.comment),
        "hyperlink": _hyperlink_snapshot(cell.hyperlink),
    }


def _row_dimension_snapshot(index: int, dimension: Any) -> dict[str, Any]:
    return {
        "index": index,
        "height": dimension.height,
        "hidden": bool(dimension.hidden),
        "outline_level": dimension.outlineLevel,
        "collapsed": bool(dimension.collapsed),
        "thick_top": bool(dimension.thickTop),
        "thick_bottom": bool(dimension.thickBot),
        "style_id": dimension.style_id if dimension.has_style else None,
    }


def _column_dimension_snapshot(key: str, dimension: Any) -> dict[str, Any]:
    return {
        "key": key,
        "min": dimension.min,
        "max": dimension.max,
        "width": dimension.width,
        "best_fit": bool(dimension.bestFit),
        "hidden": bool(dimension.hidden),
        "outline_level": dimension.outlineLevel,
        "collapsed": bool(dimension.collapsed),
        "style_id": dimension.style_id if dimension.has_style else None,
    }


def _conditional_formatting_snapshot(worksheet: Any) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for conditional_formatting in worksheet.conditional_formatting:
        rules = worksheet.conditional_formatting[conditional_formatting]
        items.append(
            {
                "range": str(conditional_formatting.sqref),
                "rules": [_serializable_snapshot(rule) for rule in rules],
            }
        )
    return sorted(items, key=lambda item: item["range"])


def _table_snapshot(worksheet: Any) -> list[dict[str, Any]]:
    tables: list[dict[str, Any]] = []
    for name in sorted(worksheet.tables):
        table = worksheet.tables[name]
        tables.append(
            {
                "name": name,
                "ref": table.ref,
                "definition": _serializable_snapshot(table),
            }
        )
    return tables


def _anchor_snapshot(anchor: Any) -> Any:
    if isinstance(anchor, str):
        return anchor
    return _serializable_snapshot(anchor)


def _image_snapshot(image: Any) -> dict[str, Any]:
    try:
        payload = image._data()
        digest = hashlib.sha256(payload).hexdigest()
    except Exception as exc:
        digest = f"unavailable:{exc.__class__.__name__}"
    return {
        "format": getattr(image, "format", None),
        "width": getattr(image, "width", None),
        "height": getattr(image, "height", None),
        "anchor": _anchor_snapshot(getattr(image, "anchor", None)),
        "sha256": digest,
    }


def _header_footer_snapshot(item: Any) -> dict[str, Any]:
    return {
        section: {
            attribute: _normalized_scalar(getattr(getattr(item, section), attribute, None))
            for attribute in ("text", "font", "size", "color")
        }
        for section in ("left", "center", "right")
    }


def _worksheet_snapshot(worksheet: Any) -> dict[str, Any]:
    cells = [
        _cell_snapshot(cell)
        for _coordinate, cell in sorted(worksheet._cells.items())
        if cell.value is not None
        or cell.has_style
        or cell.comment is not None
        or cell.hyperlink is not None
    ]
    row_dimensions = [
        _row_dimension_snapshot(index, dimension)
        for index, dimension in sorted(worksheet.row_dimensions.items())
        if dimension.height is not None
        or dimension.hidden
        or dimension.outlineLevel
        or dimension.collapsed
        or dimension.has_style
        or dimension.thickTop
        or dimension.thickBot
    ]
    column_dimensions = [
        _column_dimension_snapshot(key, dimension)
        for key, dimension in sorted(worksheet.column_dimensions.items())
    ]
    data_validations = sorted(
        (_serializable_snapshot(item) for item in worksheet.data_validations.dataValidation),
        key=lambda item: json.dumps(item, ensure_ascii=False, sort_keys=True),
    )
    charts = [_serializable_snapshot(chart) for chart in worksheet._charts]
    images = [_image_snapshot(image) for image in worksheet._images]
    return {
        "title": worksheet.title,
        "state": worksheet.sheet_state,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "freeze_panes": str(worksheet.freeze_panes) if worksheet.freeze_panes else None,
        "merged_ranges": sorted(str(item) for item in worksheet.merged_cells.ranges),
        "cells": cells,
        "row_dimensions": row_dimensions,
        "column_dimensions": column_dimensions,
        "sheet_properties": _serializable_snapshot(worksheet.sheet_properties),
        "sheet_format": _serializable_snapshot(worksheet.sheet_format),
        "sheet_view": _serializable_snapshot(worksheet.sheet_view),
        "auto_filter": _serializable_snapshot(worksheet.auto_filter),
        "page_margins": _serializable_snapshot(worksheet.page_margins),
        "page_setup": _serializable_snapshot(worksheet.page_setup),
        "print_options": _serializable_snapshot(worksheet.print_options),
        "print_area": str(worksheet.print_area) if worksheet.print_area else None,
        "print_title_rows": worksheet.print_title_rows,
        "print_title_cols": worksheet.print_title_cols,
        "row_breaks": _serializable_snapshot(worksheet.row_breaks),
        "column_breaks": _serializable_snapshot(worksheet.col_breaks),
        "headers_and_footers": {
            "odd_header": _header_footer_snapshot(worksheet.oddHeader),
            "odd_footer": _header_footer_snapshot(worksheet.oddFooter),
            "even_header": _header_footer_snapshot(worksheet.evenHeader),
            "even_footer": _header_footer_snapshot(worksheet.evenFooter),
            "first_header": _header_footer_snapshot(worksheet.firstHeader),
            "first_footer": _header_footer_snapshot(worksheet.firstFooter),
        },
        "sheet_protection": _serializable_snapshot(worksheet.protection),
        "data_validations": data_validations,
        "conditional_formatting": _conditional_formatting_snapshot(worksheet),
        "tables": _table_snapshot(worksheet),
        "charts": charts,
        "images": images,
    }


def _defined_names_snapshot(workbook: Any) -> list[dict[str, Any]]:
    names: list[dict[str, Any]] = []
    for name, definition in sorted(workbook.defined_names.items()):
        names.append(
            {
                "name": name,
                "definition": _serializable_snapshot(definition),
            }
        )
    return names


def _workbook_properties_snapshot(workbook: Any) -> dict[str, Any]:
    properties = workbook.properties
    stable_fields = (
        "title",
        "subject",
        "creator",
        "keywords",
        "description",
        "lastModifiedBy",
        "category",
        "contentStatus",
        "identifier",
        "language",
        "version",
        "revision",
    )
    theme = workbook.loaded_theme
    return {
        "core": {
            field: _normalized_scalar(getattr(properties, field, None))
            for field in stable_fields
        },
        "security": _serializable_snapshot(workbook.security),
        "views": [_serializable_snapshot(view) for view in workbook.views],
        "custom": _normalized_scalar(list(workbook.custom_doc_props)),
        "theme_sha256": hashlib.sha256(theme).hexdigest() if theme else None,
    }


def _package_snapshot(path: Path) -> dict[str, Any]:
    try:
        with ZipFile(path) as archive:
            members = sorted(item.filename for item in archive.infolist())
            binary_parts = {}
            for member in members:
                lowered = member.lower()
                if lowered.endswith((".xml", ".rels")) or member == "[Content_Types].xml":
                    continue
                binary_parts[member] = hashlib.sha256(archive.read(member)).hexdigest()
            return {
                "members": members,
                "binary_parts": binary_parts,
            }
    except BadZipFile as exc:
        raise RegressionComparisonError(f"Excel 文件不是有效的 OOXML 压缩包：{path}") from exc


def workbook_semantic_snapshot(path: Path) -> dict[str, Any]:
    keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
    try:
        workbook = load_workbook(path, data_only=False, read_only=False, keep_vba=keep_vba)
    except Exception as exc:
        raise RegressionComparisonError(f"无法读取 Excel 输出 {path}：{exc}") from exc
    try:
        snapshot = {
            "sheet_order": list(workbook.sheetnames),
            "defined_names": _defined_names_snapshot(workbook),
            "workbook_properties": _workbook_properties_snapshot(workbook),
            "calculation": _serializable_snapshot(workbook.calculation),
            "worksheets": [_worksheet_snapshot(worksheet) for worksheet in workbook.worksheets],
            "package": _package_snapshot(path),
        }
    finally:
        workbook.close()

    if any(
        cell["data_type"] == "f"
        for worksheet in snapshot["worksheets"]
        for cell in worksheet["cells"]
    ):
        try:
            cached_workbook = load_workbook(
                path,
                data_only=True,
                read_only=False,
                keep_vba=keep_vba,
            )
        except Exception as exc:
            raise RegressionComparisonError(f"无法读取 Excel 公式缓存 {path}：{exc}") from exc
        try:
            for worksheet in snapshot["worksheets"]:
                cached_worksheet = cached_workbook[worksheet["title"]]
                for cell in worksheet["cells"]:
                    if cell["data_type"] == "f":
                        cell["cached_value"] = _normalized_scalar(
                            cached_worksheet[cell["coordinate"]].value
                        )
        finally:
            cached_workbook.close()
    return snapshot


def semantic_sha256(snapshot: dict[str, Any]) -> str:
    payload = json.dumps(
        snapshot,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _validated_root(root: Path, *, label: str) -> Path:
    if root.is_symlink():
        raise RegressionComparisonError(f"{label}不能是符号链接：{root}")
    if not root.is_dir():
        raise RegressionComparisonError(f"{label}不存在或不是目录：{root}")
    return root.resolve()


def _walk_tree(root: Path) -> tuple[list[str], list[tuple[str, Path]], dict[str, str]]:
    directories: list[str] = []
    files: list[tuple[str, Path]] = []
    symlinks: dict[str, str] = {}
    for current_root, dir_names, file_names in os.walk(root, followlinks=False):
        current = Path(current_root)
        retained_dirs: list[str] = []
        for name in sorted(dir_names):
            path = current / name
            relative = path.relative_to(root).as_posix()
            if path.is_symlink():
                symlinks[relative] = os.readlink(path)
            else:
                directories.append(relative)
                retained_dirs.append(name)
        dir_names[:] = retained_dirs
        for name in sorted(file_names):
            path = current / name
            relative = path.relative_to(root).as_posix()
            if path.is_symlink():
                symlinks[relative] = os.readlink(path)
            elif path.is_file():
                files.append((relative, path))
    return sorted(directories), sorted(files), dict(sorted(symlinks.items()))


def capture_tree(root: Path, *, excel_semantics: bool) -> dict[str, Any]:
    resolved = _validated_root(root, label="对比目录")
    directories, paths, symlinks = _walk_tree(resolved)
    files: dict[str, Any] = {}
    for relative, path in paths:
        if excel_semantics and path.suffix.lower() in EXCEL_SUFFIXES:
            snapshot = workbook_semantic_snapshot(path)
            files[relative] = {
                "kind": "excel",
                "semantic_sha256": semantic_sha256(snapshot),
                "semantic": snapshot,
            }
        else:
            files[relative] = {
                "kind": "file",
                "size": path.stat().st_size,
                "sha256": sha256_file(path),
            }
    return {
        "directories": directories,
        "symlinks": symlinks,
        "files": files,
    }


def capture_manifest(outputs: Path, sources: Path | None = None) -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "outputs": capture_tree(outputs, excel_semantics=True),
        "sources": capture_tree(sources, excel_semantics=False) if sources is not None else None,
    }


def _value_preview(value: Any) -> Any:
    if isinstance(value, str) and len(value) > DISPLAY_VALUE_MAX:
        return value[:DISPLAY_VALUE_MAX] + "…"
    return value


def _append_difference(
    differences: list[dict[str, Any]],
    *,
    path: str,
    before: Any,
    after: Any,
    limit: int,
) -> None:
    if len(differences) >= limit:
        return
    differences.append(
        {
            "path": path or "$",
            "before": _value_preview(before),
            "after": _value_preview(after),
        }
    )


def _list_identity_key(before: list[Any], after: list[Any]) -> str | None:
    combined = [*before, *after]
    if not combined or not all(isinstance(item, dict) for item in combined):
        return None
    for key in LIST_IDENTITY_KEYS:
        if not all(key in item for item in combined):
            continue
        before_values = [str(item[key]) for item in before]
        after_values = [str(item[key]) for item in after]
        if len(before_values) == len(set(before_values)) and len(after_values) == len(set(after_values)):
            return key
    return None


def _escaped_path_value(value: Any) -> str:
    return str(value).replace("~", "~0").replace("/", "~1")


def compare_values(
    before: Any,
    after: Any,
    *,
    path: str = "",
    differences: list[dict[str, Any]] | None = None,
    limit: int = DIFF_LIMIT_DEFAULT,
) -> list[dict[str, Any]]:
    if differences is None:
        differences = []
    if len(differences) >= limit:
        return differences
    if type(before) is not type(after):
        _append_difference(
            differences,
            path=path,
            before=before,
            after=after,
            limit=limit,
        )
        return differences
    if isinstance(before, dict):
        before_keys = set(before)
        after_keys = set(after)
        for key in sorted(before_keys - after_keys):
            _append_difference(
                differences,
                path=f"{path}/{key}",
                before=before[key],
                after={"missing": True},
                limit=limit,
            )
        for key in sorted(after_keys - before_keys):
            _append_difference(
                differences,
                path=f"{path}/{key}",
                before={"missing": True},
                after=after[key],
                limit=limit,
            )
        for key in sorted(before_keys & after_keys):
            compare_values(
                before[key],
                after[key],
                path=f"{path}/{key}",
                differences=differences,
                limit=limit,
            )
            if len(differences) >= limit:
                break
        return differences
    if isinstance(before, list):
        identity_key = _list_identity_key(before, after)
        if identity_key is not None:
            before_items = {str(item[identity_key]): item for item in before}
            after_items = {str(item[identity_key]): item for item in after}
            for value in sorted(before_items.keys() - after_items.keys()):
                _append_difference(
                    differences,
                    path=f"{path}/{identity_key}={_escaped_path_value(value)}",
                    before=before_items[value],
                    after={"missing": True},
                    limit=limit,
                )
            for value in sorted(after_items.keys() - before_items.keys()):
                _append_difference(
                    differences,
                    path=f"{path}/{identity_key}={_escaped_path_value(value)}",
                    before={"missing": True},
                    after=after_items[value],
                    limit=limit,
                )
            for value in sorted(before_items.keys() & after_items.keys()):
                compare_values(
                    before_items[value],
                    after_items[value],
                    path=f"{path}/{identity_key}={_escaped_path_value(value)}",
                    differences=differences,
                    limit=limit,
                )
                if len(differences) >= limit:
                    break
            return differences
        common = min(len(before), len(after))
        for index in range(common):
            compare_values(
                before[index],
                after[index],
                path=f"{path}/{index}",
                differences=differences,
                limit=limit,
            )
            if len(differences) >= limit:
                return differences
        if len(before) != len(after):
            _append_difference(
                differences,
                path=f"{path}/length",
                before=len(before),
                after=len(after),
                limit=limit,
            )
        return differences
    if before != after:
        _append_difference(
            differences,
            path=path,
            before=before,
            after=after,
            limit=limit,
        )
    return differences


def verify_manifest(
    baseline: dict[str, Any],
    outputs: Path,
    sources: Path | None = None,
    *,
    difference_limit: int = DIFF_LIMIT_DEFAULT,
) -> dict[str, Any]:
    if baseline.get("schema_version") != SCHEMA_VERSION:
        raise RegressionComparisonError(
            f"不支持的回归基线版本：{baseline.get('schema_version')!r}"
        )
    current = capture_manifest(outputs, sources)
    differences = compare_values(baseline, current, limit=difference_limit)
    return {
        "ok": not differences,
        "difference_count": len(differences),
        "difference_limit": difference_limit,
        "differences": differences,
        "output_file_count": len(current["outputs"]["files"]),
        "source_file_count": len(current["sources"]["files"]) if current["sources"] else 0,
    }


def _atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=".tmp",
        dir=path.parent,
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2, sort_keys=True)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _load_manifest(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise RegressionComparisonError(f"无法读取回归基线 {path}：{exc}") from exc
    if not isinstance(payload, dict):
        raise RegressionComparisonError("回归基线根节点必须是 JSON 对象。")
    return payload


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="捕获并验证 HR Toolkit 输出语义与源附件哈希。",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    capture = subparsers.add_parser("capture", help="保存修改前回归基线")
    capture.add_argument("--outputs", type=Path, required=True, help="修改前输出目录")
    capture.add_argument("--sources", type=Path, help="源附件目录；按字节与 SHA-256 记录")
    capture.add_argument("--manifest", type=Path, required=True, help="基线 JSON 路径")

    verify = subparsers.add_parser("verify", help="验证修改后输出和附件")
    verify.add_argument("--baseline", type=Path, required=True, help="capture 生成的基线 JSON")
    verify.add_argument("--outputs", type=Path, required=True, help="修改后输出目录")
    verify.add_argument("--sources", type=Path, help="修改后的源附件目录")
    verify.add_argument("--report", type=Path, help="可选的 JSON 对比报告")
    verify.add_argument(
        "--max-differences",
        type=int,
        default=DIFF_LIMIT_DEFAULT,
        help=f"最多记录的差异数量，默认 {DIFF_LIMIT_DEFAULT}",
    )
    return parser


def _print_verification(report: dict[str, Any]) -> None:
    if report["ok"]:
        print(
            "回归验证通过："
            f"输出文件 {report['output_file_count']} 个，"
            f"源附件 {report['source_file_count']} 个，未发现语义或哈希差异。"
        )
        return
    print(
        f"回归验证失败：记录到 {report['difference_count']} 处差异"
        f"（上限 {report['difference_limit']}）。",
        file=sys.stderr,
    )
    for item in report["differences"][:20]:
        print(
            f"- {item['path']}：{item['before']!r} -> {item['after']!r}",
            file=sys.stderr,
        )


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        if args.command == "capture":
            manifest = capture_manifest(args.outputs, args.sources)
            _atomic_write_json(args.manifest, manifest)
            output_count = len(manifest["outputs"]["files"])
            source_count = len(manifest["sources"]["files"]) if manifest["sources"] else 0
            print(
                f"已保存回归基线：输出文件 {output_count} 个，"
                f"源附件 {source_count} 个，路径 {args.manifest}"
            )
            return 0

        if args.max_differences <= 0:
            raise RegressionComparisonError("--max-differences 必须大于 0。")
        baseline = _load_manifest(args.baseline)
        report = verify_manifest(
            baseline,
            args.outputs,
            args.sources,
            difference_limit=args.max_differences,
        )
        if args.report is not None:
            _atomic_write_json(args.report, report)
        _print_verification(report)
        return 0 if report["ok"] else 1
    except RegressionComparisonError as exc:
        print(f"回归工具失败：{exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
