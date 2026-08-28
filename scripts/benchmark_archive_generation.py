#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import statistics
import sys
import tempfile
import time
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from hr_toolkit.common.resources import open_template_resource
from hr_toolkit.tools.archive_import import (
    DEFAULT_ARCHIVE_SUMMARY_TEMPLATE_RESOURCE,
    export_company_archive_tables,
)


def _build_summary_fixture(path: Path, row_count: int) -> None:
    with open_template_resource(DEFAULT_ARCHIVE_SUMMARY_TEMPLATE_RESOURCE) as source:
        with path.open("wb") as target:
            shutil.copyfileobj(source, target)
    workbook = load_workbook(path)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        worksheet.title = "性能公司"
        worksheet["A1"] = "性能公司人事档案编号表"
        for index in range(1, row_count + 1):
            row_index = index + 3
            suffix = f"{index:06d}"
            worksheet.cell(row_index, 1).value = "01"
            worksheet.cell(row_index, 2).value = f"性能员工{suffix}"
            worksheet.cell(row_index, 3).value = f"11010119800101{index:04d}"
            worksheet.cell(row_index, 6).value = "2020-01-02"
            worksheet.cell(row_index, 12).value = "√"
            worksheet.cell(row_index, 13).value = "√"
            worksheet.cell(row_index, 19).value = 2
            worksheet.cell(row_index, 20).value = "√"
            worksheet.cell(row_index, 22).value = 1
            worksheet.cell(row_index, 25).value = 1
        workbook.save(path)
    finally:
        workbook.close()


def _peak_rss_bytes() -> int | None:
    try:
        import resource
    except ImportError:
        return None
    peak = int(resource.getrusage(resource.RUSAGE_SELF).ru_maxrss)
    return peak if sys.platform == "darwin" else peak * 1024


def run_benchmark(*, row_count: int, repeat: int) -> dict[str, object]:
    if not 1 <= row_count <= 9999:
        raise ValueError("row_count 必须在 1 到 9999 之间。")
    if repeat <= 0:
        raise ValueError("repeat 必须大于 0。")

    with tempfile.TemporaryDirectory(prefix="hr_archive_benchmark_") as temp_root:
        root = Path(temp_root)
        summary_path = root / "archive_summary.xlsx"
        _build_summary_fixture(summary_path, row_count)
        source_sha256 = hashlib.sha256(summary_path.read_bytes()).hexdigest()
        durations: list[float] = []
        for run_index in range(1, repeat + 1):
            output_dir = root / f"output_{run_index}"
            started = time.perf_counter()
            result = export_company_archive_tables(summary_path, output_dir)
            durations.append(time.perf_counter() - started)
            if result.inserted_count != row_count or len(result.output_files) != 1:
                raise RuntimeError("性能基准输出数量异常。")
            if hashlib.sha256(summary_path.read_bytes()).hexdigest() != source_sha256:
                raise RuntimeError("性能基准运行修改了源档案汇总表。")

        peak_rss = _peak_rss_bytes()
        return {
            "rows": row_count,
            "repeat": repeat,
            "durations_seconds": [round(value, 4) for value in durations],
            "median_seconds": round(statistics.median(durations), 4),
            "minimum_seconds": round(min(durations), 4),
            "maximum_seconds": round(max(durations), 4),
            "peak_rss_bytes": peak_rss,
            "source_sha256": source_sha256,
        }


def main() -> None:
    parser = argparse.ArgumentParser(description="测量档案表生成的耗时、峰值内存和源文件不变性。")
    parser.add_argument("--rows", type=int, default=6000, help="模拟档案行数，默认 6000。")
    parser.add_argument("--repeat", type=int, default=1, help="重复次数，默认 1。")
    parser.add_argument("--output-json", type=Path, help="可选：保存 JSON 结果。")
    args = parser.parse_args()

    payload = run_benchmark(row_count=args.rows, repeat=args.repeat)
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    print(text)
    if args.output_json is not None:
        args.output_json.parent.mkdir(parents=True, exist_ok=True)
        args.output_json.write_text(f"{text}\n", encoding="utf-8")


if __name__ == "__main__":
    main()
