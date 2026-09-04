from __future__ import annotations

import tempfile
import unittest
import zipfile
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from hr_toolkit.tools.social_security import (
    DetailRecord,
    _write_detail_workbook,
    generate_social_security_reports,
)


class SocialSecurityTest(unittest.TestCase):
    def test_large_detail_output_does_not_rescan_sheet_width_per_record(self) -> None:
        records = [
            DetailRecord(
                id_card=f"3601111990{index:08d}"[:18],
                name=f"员工{index}",
                period="202601",
                billing_period="202601",
                period_split_input=False,
                account="测试账户",
                account_display="测试账户",
                company="测试公司",
                insured_place="测试地",
                project="测试项目",
                project_display="测试项目",
                cost_center="测试成本中心",
                start_period="202601",
                management_fee=0,
            )
            for index in range(100)
        ]
        original_getter = Worksheet.max_column.fget
        access_count = 0

        def counted_max_column(worksheet):
            nonlocal access_count
            access_count += 1
            return original_getter(worksheet)

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            with patch.object(Worksheet, "max_column", new=property(counted_max_column)):
                _write_detail_workbook(records, root / "明细.xlsx", root)

        self.assertLessEqual(access_count, 3)

    def test_generate_social_security_reports_from_mixed_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            roster = root / "参保人员花名册.xlsx"
            _write_roster(roster)
            _write_long_payment_file(input_dir / "北京春苗抚州账户2026年5月社保单位缴费明细.xlsx")
            _write_single_kind_file(input_dir / "2026-04——工伤保险（单位缴纳部分）职工明细.xlsx")
            progress = []

            result = generate_social_security_reports(
                input_dir,
                roster,
                output_dir,
                progress_callback=lambda current, total, message: progress.append(
                    (current, total, message)
                ),
            )
            payload = result.to_dict()

            self.assertEqual(payload["source_file_count"], 2)
            self.assertEqual(payload["source_record_count"], 3)
            self.assertEqual(payload["detail_record_count"], 2)
            self.assertEqual(payload["employee_count"], 2)
            self.assertEqual(payload["account_counts"], {"北京抚州": 1, "唐人四川": 1})
            self.assertEqual(payload["period_counts"], {"202605": 1, "202604": 1})
            self.assertTrue(result.detail_output_file and result.detail_output_file.exists())
            self.assertEqual(len(result.detail_output_files), 2)
            split_names = {path.name for path in result.detail_output_files}
            self.assertEqual(split_names, {"北京抚州-社保明细表.xlsx", "唐人四川-社保明细表.xlsx"})
            self.assertTrue(result.summary_output_file and result.summary_output_file.exists())
            self.assertEqual(progress[0][:2], (0, 5))
            self.assertEqual(progress[-1], (5, 5, "社保报表生成完成"))

            detail_wb = load_workbook(result.detail_output_file, data_only=False)
            detail_ws = detail_wb["社保明细表"]
            rows = {detail_ws.cell(row, 6).value: row for row in range(4, 6)}
            zhang_row = rows["360111199001010011"]
            li_row = rows["360111199002020022"]
            self.assertEqual(detail_ws.cell(zhang_row, 2).value, "北京春苗")
            self.assertEqual(detail_ws.cell(zhang_row, 3).value, "抚州")
            self.assertEqual(detail_ws.cell(zhang_row, 8).value, "202605")
            self.assertEqual(detail_ws.cell(zhang_row, 11).value, f"=ROUND(I{zhang_row}*J{zhang_row},2)")
            self.assertEqual(detail_ws.cell(zhang_row, 13).value, f"=ROUND(I{zhang_row}*L{zhang_row},2)")
            self.assertEqual(detail_ws.cell(zhang_row, 10).number_format, "0.00%")
            self.assertEqual(detail_ws.cell(zhang_row, 67).number_format, "0.00_ ")
            self.assertEqual(detail_ws.cell(zhang_row, 73).value, 20)
            self.assertEqual(detail_ws.cell(li_row, 2).value, "唐人数智")
            self.assertEqual(detail_ws.cell(li_row, 8).value, "202604")
            self.assertEqual(detail_ws.cell(li_row, 26).value, f"=ROUND(X{li_row}*Y{li_row},2)")
            template_wb = load_workbook(
                Path(__file__).resolve().parents[1] / "hr_toolkit" / "templates" / "social_security_detail_template.xlsx",
                data_only=False,
            )
            template_ws = template_wb["社保明细表模板"]
            for col_index in range(1, 77):
                self.assertEqual(detail_ws.cell(zhang_row, col_index)._style, template_ws.cell(4, col_index)._style)
            template_wb.close()
            detail_wb.close()

            split_detail = next(path for path in result.detail_output_files if path.name == "北京抚州-社保明细表.xlsx")
            split_wb = load_workbook(split_detail, data_only=False)
            split_ws = split_wb["社保明细表"]
            self.assertEqual(split_ws.cell(1, 1).value, "北京春苗2026年5月社保明细表")
            self.assertEqual(split_ws.max_row, 4)
            self.assertEqual(split_ws.cell(4, 5).value, "张三")
            split_wb.close()

            summary_wb = load_workbook(result.summary_output_file, data_only=False)
            self.assertIn("社保汇总表", summary_wb.sheetnames)
            self.assertIn("北京春苗", summary_wb.sheetnames)
            self.assertIn("唐人数智", summary_wb.sheetnames)
            self.assertIn("数据分析", summary_wb.sheetnames)
            self.assertIn("异常提醒", summary_wb.sheetnames)
            summary_wb.close()

    def test_generate_social_security_reports_from_zip(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            roster = root / "参保人员花名册.xlsx"
            source = root / "北京春苗抚州账户2026年5月社保单位缴费明细.xlsx"
            archive = root / "社保清单.zip"
            output_dir = root / "output"
            _write_roster(roster)
            _write_long_payment_file(source)
            with zipfile.ZipFile(archive, "w") as zip_file:
                zip_file.write(source, arcname=source.name)

            result = generate_social_security_reports([archive], roster, output_dir)

            self.assertEqual(result.source_record_count, 2)
            self.assertEqual(result.detail_record_count, 1)
            self.assertTrue(result.detail_output_file and result.detail_output_file.exists())
            self.assertTrue(result.summary_output_file and result.summary_output_file.exists())

    def test_zip_name_supplies_context_for_root_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            roster = root / "参保人员花名册.xlsx"
            source = root / "2026-04——工伤保险（单位缴纳部分）职工明细.xlsx"
            archive = root / "北京春苗抚州账户2026年5月社保单位缴费明细.zip"
            output_dir = root / "output"
            _write_roster(roster)
            _write_single_kind_file(source)
            with zipfile.ZipFile(archive, "w") as zip_file:
                zip_file.write(source, arcname=source.name)

            result = generate_social_security_reports([archive], roster, output_dir)

            self.assertEqual(result.period_counts, {"202604": 1})
            joined_warnings = "\n".join(result.warnings)
            self.assertIn("参保账户与花名册不一致", joined_warnings)
            self.assertIn("参保地与花名册不一致", joined_warnings)
            wb = load_workbook(result.detail_output_file, data_only=True)
            ws = wb["社保明细表"]
            self.assertEqual(ws.cell(4, 2).value, "北京春苗")
            self.assertEqual(ws.cell(4, 3).value, "抚州")
            self.assertEqual(ws.cell(4, 8).value, "202604")
            wb.close()

    def test_file_fee_month_overrides_container_month(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "唐人四川2026年5月社保单位缴费明细"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            input_dir.mkdir()
            _write_roster(roster)
            _write_single_kind_file(input_dir / "2026-04——工伤保险（单位缴纳部分）职工明细.xlsx")

            result = generate_social_security_reports(input_dir, roster, output_dir)

            self.assertEqual(result.period_counts, {"202604": 1})
            wb = load_workbook(result.detail_output_file, data_only=True)
            ws = wb["社保明细表"]
            self.assertEqual(ws.cell(4, 8).value, "202604")
            wb.close()

    def test_distinguishes_arrears_and_difference_in_same_month_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "唐人四川2026年5月社保单位缴费明细"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            input_dir.mkdir()
            _write_roster(
                roster,
                extra_rows=[
                    ["王五", "360111199003030033", "正常", date(2026, 2, 1), "唐人四川", "唐人数智科技股份有限公司", "四川项目部", "项目（成都）", "成本二", 30],
                ],
            )
            zhang = ("张三", "360111199001010011")
            li = ("李四", "360111199002020022")
            wang = ("王五", "360111199003030033")
            _write_single_kind_rows(
                input_dir / "2026-01——工伤保险（单位缴纳部分）职工明细.xlsx",
                [(*li, 4588, 0.001, 4.58), (*wang, 4588, 0.001, 4.58)],
            )
            _write_single_kind_rows(
                input_dir / "2026-02——工伤保险（单位缴纳部分）职工明细.xlsx",
                [(*li, 4588, 0.001, 4.58), (*wang, 4588, 0.001, 4.58)],
            )
            _write_single_kind_rows(
                input_dir / "2026-03——工伤保险（单位缴纳部分）职工明细.xlsx",
                [(*li, 4588, 0.001, 4.58), (*wang, 4588, 0.001, 4.58), (*zhang, 4588, 0.003, 13.76)],
            )
            _write_single_kind_rows(
                input_dir / "2026-04——工伤保险（单位缴纳部分）职工明细.xlsx",
                [(*li, 4588, 0.003, 13.76), (*wang, 4588, 0.003, 13.76), (*zhang, 4588, 0.003, 13.76)],
            )

            result = generate_social_security_reports(input_dir, roster, output_dir)

            self.assertEqual(len(result.source_files), 4)
            self.assertEqual(result.source_record_count, 10)
            self.assertEqual(result.detail_record_count, 3)
            self.assertEqual(result.period_counts, {"202604": 3})
            self.assertNotIn("待确认历史缴费", "\n".join(result.warnings))
            wb = load_workbook(result.detail_output_file, data_only=False)
            ws = wb["社保明细表"]
            rows = {ws.cell(row, 6).value: row for row in range(4, 7)}
            li_row = rows[li[1]]
            zhang_row = rows[zhang[1]]
            self.assertEqual(ws["A1"].value, "唐人数智2026年4月社保明细表")
            self.assertEqual(ws["BB2"].value, "工伤2026年1月-3月补差")
            self.assertEqual(ws["BJ2"].value, "个人社保\n补缴合计")
            self.assertEqual(ws["BK2"].value, "单位社保\n补缴合计")
            self.assertEqual(ws.cell(li_row, 8).value, "202604")
            self.assertEqual(ws.cell(li_row, 26).value, f"=ROUND(X{li_row}*Y{li_row},2)")
            self.assertEqual(ws.cell(li_row, 54).value, 4588)
            self.assertEqual(ws.cell(li_row, 55).value, 0.001)
            self.assertEqual(ws.cell(li_row, 55).number_format, "0.00%")
            self.assertEqual(ws.cell(li_row, 56).value, 13.74)
            self.assertIsNone(ws.cell(li_row, 62).value)
            self.assertEqual(ws.cell(li_row, 63).value, f"=AV{li_row}+BA{li_row}+BD{li_row}+BI{li_row}")
            self.assertEqual(
                ws.cell(li_row, 68).value,
                f"=M{li_row}+R{li_row}+W{li_row}+Z{li_row}+AK{li_row}+BK{li_row}+AC{li_row}+BN{li_row}",
            )
            self.assertEqual(ws.cell(li_row, 67).value, f"=K{li_row}+P{li_row}+U{li_row}+AI{li_row}+BJ{li_row}")
            self.assertEqual(ws.cell(li_row, 69).value, f"=BO{li_row}+BP{li_row}")
            self.assertEqual(ws.cell(li_row, 74).value, f"=ROUND((BQ{li_row}+BU{li_row})*6.72%,2)")
            self.assertEqual(ws.cell(li_row, 75).value, f"=BQ{li_row}+BU{li_row}+BV{li_row}")
            self.assertIsNone(ws.cell(li_row, 76).value)
            self.assertEqual(ws.cell(zhang_row, 8).value, "202604")
            self.assertEqual(ws.cell(zhang_row, 24).value, 4588)
            self.assertEqual(ws.cell(zhang_row, 25).value, 0.003)
            self.assertEqual(ws.cell(zhang_row, 26).value, 27.52)
            self.assertIsNone(ws.cell(zhang_row, 56).value)
            self.assertIsNone(ws.cell(zhang_row, 63).value)
            self.assertIsNone(ws.cell(zhang_row, 76).value)
            wb.close()

    def test_compatible_arrears_months_merge_into_front_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "北京春苗抚州账户2026年4月社保单位缴费明细.xlsx"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            _write_roster(roster)
            _write_long_payment_rows_with_nature(
                source,
                [
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 3, 1), 5000, 0.01, 50, "补缴"],
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 4, 1), 5000, 0.01, 50, "正常缴费"],
                ],
            )

            result = generate_social_security_reports(source, roster, output_dir)

            self.assertEqual(result.detail_record_count, 1)
            self.assertEqual(result.period_counts, {"202604": 1})
            wb = load_workbook(result.detail_output_file, data_only=False)
            ws = wb["社保明细表"]
            self.assertEqual(ws["H4"].value, "202604")
            self.assertEqual(ws["X4"].value, 5000)
            self.assertEqual(ws["Y4"].value, 0.01)
            self.assertEqual(ws["Z4"].value, 100)
            self.assertIsNone(ws["BD4"].value)
            self.assertIsNone(ws["BK4"].value)
            wb.close()

            summary_wb = load_workbook(result.summary_output_file, data_only=False)
            analysis = summary_wb["数据分析"]
            category_rows = {
                analysis.cell(row, 1).value: row
                for row in range(1, analysis.max_row + 1)
            }
            self.assertEqual(analysis.cell(category_rows["工伤"], 3).value, 5000)
            summary_wb.close()

    def test_same_base_is_not_multiplied_and_all_rows_keep_table_month(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "唐人长春2026年5月社保单位缴费明细.xlsx"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            _write_roster(roster)
            _write_long_payment_rows_with_nature(
                source,
                [
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 5, 1), 500, 0.01, 5, "补缴"],
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 5, 2), 500, 0.01, 5, "补缴"],
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 6, 1), 500, 0.01, 5, "补缴"],
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 6, 2), 500, 0.01, 5, "补缴"],
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 6, 3), 500, 0.01, 5, "补缴"],
                ],
            )

            result = generate_social_security_reports(source, roster, output_dir)

            self.assertEqual(result.detail_record_count, 1)
            self.assertEqual(result.period_counts, {"202605": 1})
            wb = load_workbook(result.detail_output_file, data_only=False)
            ws = wb["社保明细表"]
            self.assertEqual(ws["H4"].value, "202605")
            self.assertEqual(ws["X4"].value, 500)
            self.assertEqual(ws["Y4"].value, 0.01)
            self.assertEqual(ws["Z4"].value, 25)
            self.assertEqual(ws["BU4"].value, 20)
            wb.close()

    def test_different_bases_split_but_keep_the_same_table_month(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "2026-05——工伤保险（单位缴纳部分）职工明细.xlsx"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            _write_roster(roster)
            _write_long_payment_rows_with_nature(
                source,
                [
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 5, 1), 400, 0.01, 4, "补缴"],
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 6, 1), 500, 0.01, 5, "补缴"],
                ],
            )

            result = generate_social_security_reports(source, roster, output_dir)

            self.assertEqual(result.detail_record_count, 2)
            self.assertEqual(result.period_counts, {"202605": 1})
            wb = load_workbook(result.detail_output_file, data_only=False)
            ws = wb["社保明细表"]
            rows = {ws.cell(row, 24).value: row for row in range(4, 6)}
            self.assertEqual(set(rows), {400, 500})
            self.assertEqual(ws.cell(rows[400], 8).value, "202605")
            self.assertEqual(ws.cell(rows[500], 8).value, "202605")
            self.assertEqual(ws.cell(rows[400], 26).value, f"=ROUND(X{rows[400]}*Y{rows[400]},2)")
            self.assertEqual(ws.cell(rows[500], 26).value, f"=ROUND(X{rows[500]}*Y{rows[500]},2)")
            self.assertEqual(ws.cell(rows[400], 73).value, 20)
            self.assertIsNone(ws.cell(rows[500], 73).value)
            wb.close()

    def test_explicit_difference_marker_handles_single_history_month(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "北京春苗抚州账户2026年4月社保单位缴费明细.xlsx"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            _write_roster(roster)
            _write_long_payment_rows_with_nature(
                source,
                [
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 3, 1), 4588, 0.001, 4.58, "调整补收"],
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 4, 1), 4588, 0.003, 13.76, "正常缴费"],
                ],
            )

            result = generate_social_security_reports(source, roster, output_dir)

            wb = load_workbook(result.detail_output_file, data_only=False)
            ws = wb["社保明细表"]
            self.assertEqual(ws["A1"].value, "北京春苗2026年4月社保明细表")
            self.assertEqual(ws["H4"].value, "202604")
            self.assertEqual(ws["Z4"].value, "=ROUND(X4*Y4,2)")
            self.assertEqual(ws["BB4"].value, 4588)
            self.assertEqual(ws["BC4"].value, 0.001)
            self.assertEqual(ws["BC4"].number_format, "0.00%")
            self.assertEqual(ws["BD4"].value, 4.58)
            self.assertEqual(ws["BK4"].value, "=AV4+BA4+BD4+BI4")
            self.assertIsNone(ws["BX4"].value)
            wb.close()

    def test_explicit_arrears_marker_never_moves_amount_to_difference_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "北京春苗抚州账户2026年4月社保单位缴费明细.xlsx"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            _write_roster(roster)
            _write_long_payment_rows_with_nature(
                source,
                [
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 3, 1), 4588, 0.001, 4.58, "补缴"],
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 4, 1), 4588, 0.003, 13.76, "正常缴费"],
                ],
            )

            result = generate_social_security_reports(source, roster, output_dir)

            self.assertEqual(result.warnings, [])
            self.assertEqual(result.detail_record_count, 2)
            self.assertEqual(result.period_counts, {"202604": 1})
            wb = load_workbook(result.detail_output_file, data_only=False)
            ws = wb["社保明细表"]
            rows = {ws.cell(row, 25).value: row for row in range(4, 6)}
            arrears_row = rows[0.001]
            current_row = rows[0.003]
            self.assertEqual(ws.cell(arrears_row, 8).value, "202604")
            self.assertEqual(ws.cell(current_row, 8).value, "202604")
            self.assertEqual(ws.cell(arrears_row, 24).value, 4588)
            self.assertEqual(ws.cell(arrears_row, 25).value, 0.001)
            self.assertEqual(ws.cell(arrears_row, 26).value, 4.58)
            self.assertIsNone(ws.cell(arrears_row, 56).value)
            self.assertIsNone(ws.cell(arrears_row, 63).value)
            self.assertIsNone(ws.cell(arrears_row, 73).value)
            self.assertEqual(ws.cell(current_row, 26).value, f"=ROUND(X{current_row}*Y{current_row},2)")
            self.assertIsNone(ws.cell(current_row, 56).value)
            self.assertIsNone(ws.cell(current_row, 63).value)
            self.assertEqual(ws.cell(current_row, 73).value, 20)
            self.assertIsNone(ws.cell(current_row, 76).value)
            wb.close()

    def test_explicit_arrears_and_difference_remain_separate_for_same_person(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "北京春苗抚州账户2026年4月社保单位缴费明细.xlsx"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            _write_roster(roster)
            _write_long_payment_rows_with_nature(
                source,
                [
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 2, 1), 4588, 0.003, 13.76, "补缴"],
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 3, 1), 4588, 0.001, 4.58, "调整补收"],
                    ["张三", "360111199001010011", "工伤保险费", "工伤保险", date(2026, 4, 1), 4588, 0.003, 13.76, "正常缴费"],
                ],
            )

            result = generate_social_security_reports(source, roster, output_dir)

            self.assertEqual(result.warnings, [])
            self.assertEqual(result.detail_record_count, 1)
            self.assertEqual(result.period_counts, {"202604": 1})
            wb = load_workbook(result.detail_output_file, data_only=False)
            ws = wb["社保明细表"]
            self.assertEqual(ws["H4"].value, "202604")
            self.assertEqual(ws["X4"].value, 4588)
            self.assertEqual(ws["Y4"].value, 0.003)
            self.assertEqual(ws["Z4"].value, 27.52)
            self.assertEqual(ws["BD4"].value, 4.58)
            self.assertEqual(
                ws["BK4"].value,
                "=AV4+BA4+BD4+BI4",
            )
            self.assertIsNone(ws["BX4"].value)
            wb.close()

    def test_unsupported_difference_category_is_kept_visible_in_normal_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "北京春苗抚州账户2026年4月社保单位缴费明细.xlsx"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            _write_roster(roster)
            _write_long_payment_rows_with_nature(
                source,
                [
                    ["张三", "360111199001010011", "大病医疗保险费", "大病医疗保险", date(2026, 3, 1), 4588, 0.001, 5, "补差"],
                    ["张三", "360111199001010011", "大病医疗保险费", "大病医疗保险", date(2026, 4, 1), 4588, 0.005, 25, "正常缴费"],
                ],
            )

            result = generate_social_security_reports(source, roster, output_dir)

            self.assertIn("模板没有对应补差明细列", "\n".join(result.warnings))
            wb = load_workbook(result.detail_output_file, data_only=False)
            ws = wb["社保明细表"]
            self.assertEqual(ws["AK4"].value, 30)
            self.assertIsNone(ws["BI4"].value)
            self.assertIsNone(ws["BK4"].value)
            self.assertIsNone(ws["BX4"].value)
            wb.close()

    def test_does_not_guess_single_person_historical_rate_change_as_difference(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "唐人四川2026年4月社保单位缴费明细"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            input_dir.mkdir()
            _write_roster(roster)
            person = ("李四", "360111199002020022")
            for month in (1, 2):
                _write_single_kind_rows(
                    input_dir / f"2026-{month:02d}——工伤保险（单位缴纳部分）职工明细.xlsx",
                    [(*person, 4588, 0.001, 4.58)],
                )
            _write_single_kind_rows(
                input_dir / "2026-04——工伤保险（单位缴纳部分）职工明细.xlsx",
                [(*person, 4588, 0.003, 13.76)],
            )

            result = generate_social_security_reports(input_dir, roster, output_dir)

            self.assertIn("待确认历史缴费", "\n".join(result.warnings))
            wb = load_workbook(result.detail_output_file, data_only=False)
            ws = wb["社保明细表"]
            self.assertEqual(ws["Z4"].value, 22.92)
            self.assertIsNone(ws["BD4"].value)
            self.assertIsNone(ws["BK4"].value)
            self.assertIsNone(ws["BX4"].value)
            wb.close()

    def test_combined_wide_file_keeps_its_bill_month(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "唐人长春2026年5月社保单位缴费明细.xlsx"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            _write_roster(roster)
            _write_wide_payment_file(source)

            result = generate_social_security_reports(source, roster, output_dir)

            self.assertEqual(result.source_record_count, 10)
            self.assertEqual(result.detail_record_count, 1)
            self.assertEqual(result.period_counts, {"202605": 1})
            self.assertNotIn("待确认历史缴费", "\n".join(result.warnings))
            wb = load_workbook(result.detail_output_file, data_only=False)
            ws = wb["社保明细表"]
            self.assertEqual(ws["A1"].value, "唐人数智2026年5月社保明细表")
            self.assertEqual(ws["H4"].value, "202605")
            self.assertEqual(ws["M4"].value, 100)
            self.assertEqual(ws["K4"].value, 50)
            self.assertEqual(ws["R4"].value, 120)
            self.assertEqual(ws["P4"].value, 20)
            self.assertEqual(ws["W4"].value, 10)
            self.assertEqual(ws["U4"].value, 5)
            self.assertEqual(ws["Z4"].value, 8)
            self.assertEqual(ws["AK4"].value, 5)
            self.assertEqual(ws["AI4"].value, 2)
            self.assertIsNone(ws["BX4"].value)
            wb.close()

    def test_wide_history_without_nature_or_basis_is_not_guessed_as_arrears(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "唐人四川2026年5月社保单位缴费明细.xlsx"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            _write_roster(roster)
            _write_wide_amount_only_history_file(source)

            result = generate_social_security_reports(source, roster, output_dir)

            self.assertIn("待确认历史缴费", "\n".join(result.warnings))
            wb = load_workbook(result.detail_output_file, data_only=False)
            ws = wb["社保明细表"]
            self.assertEqual(ws["M4"].value, 200)
            self.assertIsNone(ws["BK4"].value)
            self.assertIsNone(ws["BX4"].value)
            wb.close()

    def test_warns_when_bill_account_differs_from_roster(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "北京春苗抚州账户2026年5月社保单位缴费明细"
            output_dir = root / "output"
            roster = root / "参保人员花名册.xlsx"
            input_dir.mkdir()
            _write_roster(roster)
            _write_single_kind_file(input_dir / "2026-05——工伤保险（单位缴纳部分）职工明细.xlsx")

            result = generate_social_security_reports(input_dir, roster, output_dir)

            joined_warnings = "\n".join(result.warnings)
            self.assertIn("参保账户与花名册不一致", joined_warnings)
            self.assertIn("参保地与花名册不一致", joined_warnings)
            wb = load_workbook(result.detail_output_file, data_only=True)
            ws = wb["社保明细表"]
            self.assertEqual(ws.cell(4, 2).value, "北京春苗")
            self.assertEqual(ws.cell(4, 3).value, "抚州")
            wb.close()


def _write_roster(path: Path, extra_rows: list[list[object]] | None = None) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "花名册"
    headers = [
        "*姓名.简体中文",
        "*身份证",
        "*参保状态",
        "*参保日期",
        "*参保方案.名称",
        "*参保单位.名称",
        "*责任部门.名称",
        "项目.项目名称",
        "成本中心.名称",
        "管理费",
    ]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(1, col_index).value = header
    rows = [
        ["张三", "360111199001010011", "正常", date(2026, 1, 1), "北京春苗抚州", "春苗人力资源（北京）有限公司", "抚州项目部", "项目（上饶市）", "成本一", 20],
        ["李四", "360111199002020022", "正常", date(2026, 2, 1), "唐人四川", "唐人数智科技股份有限公司", "四川项目部", "项目（成都）", "成本二", 30],
    ]
    rows.extend(extra_rows or [])
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    workbook.save(path)
    workbook.close()


def _write_long_payment_file(path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "缴费明细"
    headers = ["姓名", "身份证件号码", "参保费种", "征收品目", "费款所属日期起", "缴费基数", "费率", "本期应缴费额"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(1, col_index).value = header
    rows = [
        ["张三", "360111199001010011", "城镇企业职工基本养老保险", "个人缴纳部分", date(2026, 5, 1), 3000, 0.08, 240],
        ["张三", "360111199001010011", "城镇企业职工基本养老保险", "单位缴纳部分", date(2026, 5, 1), 3000, 0.16, 480],
    ]
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    workbook.save(path)
    workbook.close()


def _write_single_kind_file(path: Path, name: str = "李四", id_card: str = "360111199002020022") -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "职工明细"
    headers = ["姓名", "证件号码", "缴费基数", "费率", "应缴费额(元)"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(1, col_index).value = header
    row = [name, id_card, 3600, 0.01, 36]
    for col_index, value in enumerate(row, start=1):
        ws.cell(2, col_index).value = value
    workbook.save(path)
    workbook.close()


def _write_single_kind_rows(path: Path, rows: list[tuple[str, str, float, float, float]]) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "职工明细"
    headers = ["姓名", "证件号码", "缴费基数", "费率", "应缴费额(元)"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(1, col_index).value = header
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    workbook.save(path)
    workbook.close()


def _write_long_payment_rows_with_nature(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "缴费明细"
    headers = ["姓名", "身份证件号码", "参保费种", "征收品目", "费款所属日期起", "缴费基数", "费率", "本期应缴费额", "业务类型"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(1, col_index).value = header
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    workbook.save(path)
    workbook.close()


def _write_wide_payment_file(path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "缴费明细"
    ws.merge_cells("A1:F1")
    ws.merge_cells("G1:J1")
    ws.merge_cells("K1:L1")
    ws.merge_cells("M1:N1")
    ws["G1"] = "基本医疗保险费"
    ws["K1"] = "企业职工基本养老保险费"
    ws["M1"] = "失业保险费"
    ws["O1"] = "工伤保险费"
    headers = [
        "序号",
        "姓名",
        "证件类型",
        "证件号码",
        "费款所属期起",
        "费款所属期止",
        "职工基本医疗保险(单位缴纳)应缴费额(元)",
        "职工基本医疗保险(个人缴纳)应缴费额(元)",
        "职工大额医疗互助保险(单位缴纳)应缴费额(元)",
        "职工大额医疗互助保险(个人缴纳)应缴费额(元)",
        "职工基本养老保险(单位缴纳)应缴费额(元)",
        "职工基本养老保险(个人缴纳)应缴费额(元)",
        "失业保险(单位缴纳)应缴费额(元)",
        "失业保险(个人缴纳)应缴费额(元)",
        "工伤保险应缴费额(元)",
    ]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(2, col_index).value = header
    rows = [
        [1, "李四", "居民身份证", "360111199002020022", "2026-05", "2026-05", 70, 20, 5, 2, 100, 50, 10, 5, 8],
        [2, "李四", "居民身份证", "360111199002020022", "2026-06", "2026-06", 50, None, None, None, None, None, None, None, None],
    ]
    for row_index, row in enumerate(rows, start=3):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    workbook.save(path)
    workbook.close()


def _write_wide_amount_only_history_file(path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "缴费明细"
    headers = [
        "姓名",
        "证件号码",
        "费款所属期起",
        "费款所属期止",
        "职工基本养老保险(单位缴纳)应缴费额(元)",
    ]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(1, col_index).value = header
    rows = [
        ["李四", "360111199002020022", "2026-04", "2026-04", 100],
        ["李四", "360111199002020022", "2026-05", "2026-05", 100],
    ]
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    workbook.save(path)
    workbook.close()


if __name__ == "__main__":
    unittest.main()
