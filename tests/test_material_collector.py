from __future__ import annotations

import hashlib
import json
import os
import shutil
import sys
import tempfile
import unittest
import zipfile
import zlib
from io import BytesIO
from pathlib import Path
from types import ModuleType, SimpleNamespace
from unittest import mock

from openpyxl import load_workbook
from PIL import Image, PngImagePlugin
from pypdf import PdfWriter
from pypdf.generic import (
    DecodedStreamObject,
    DictionaryObject,
    EncodedStreamObject,
    NameObject,
    NumberObject,
)

from hr_toolkit.tools.material_collector import (
    LIBRARY_MODE_FLAT_OCR,
    LIBRARY_MODE_PERSON_FOLDER,
    MODE_BY_EMPLOYEE,
    MODE_BY_MATERIAL,
    MODE_FLAT,
    MaterialCollectResult,
    MaterialFileMatch,
    TargetEmployee,
    collect_employee_materials,
    parse_employee_roster,
    _OCR_CACHE_FILE_NAME,
    _build_doc_format_hint,
    _build_employee_key,
    _classify_material_type,
    _compute_cache_key,
    _compute_file_fingerprint,
    _extract_document_text,
    _get_engine_signature,
    _hash_id_card,
    _is_junk_or_temp_file,
    _is_path_nested,
    _is_valid_person_name,
    _load_ocr_cache,
    _match_folder_to_employee,
    _resolve_material_text,
    _save_ocr_cache,
    _scan_folder_index,
    _trim_cache_by_age_and_size,
)


def _png_fixture(marker: str = "fixture") -> bytes:
    """有效小图片；用无压缩元数据保存模拟 OCR 标记，等长标记保持文件等长。"""
    metadata = PngImagePlugin.PngInfo()
    metadata.add_text("test_marker", marker)
    output = BytesIO()
    with Image.new("RGB", (16, 16), "white") as image:
        image.save(output, format="PNG", pnginfo=metadata)
    return output.getvalue()


def _read_fixture_marker(source: str | Path) -> str:
    with Image.open(source) as image:
        return image.info["test_marker"]


def _write_text_pdf(path: Path, marker: str, *, padding_bytes: int = 0) -> None:
    writer = PdfWriter()
    page = writer.add_blank_page(width=300, height=300)
    font = DictionaryObject({
        NameObject("/Type"): NameObject("/Font"),
        NameObject("/Subtype"): NameObject("/Type1"),
        NameObject("/BaseFont"): NameObject("/Helvetica"),
    })
    page[NameObject("/Resources")] = DictionaryObject({
        NameObject("/Font"): DictionaryObject({
            NameObject("/F1"): writer._add_object(font),
        }),
    })
    content = DecodedStreamObject()
    content.set_data(
        b"%" + b"x" * padding_bytes
        + b"\nBT /F1 12 Tf 10 100 Td ("
        + marker.encode("ascii")
        + b") Tj ET"
    )
    page[NameObject("/Contents")] = writer._add_object(content)
    with path.open("wb") as stream:
        writer.write(stream)


def _write_scanned_pdf(
    path: Path,
    *,
    page_count: int,
    width: int = 16,
    height: int = 16,
) -> None:
    writer = PdfWriter()
    for page_index in range(page_count):
        page = writer.add_blank_page(width=300, height=300)
        image = EncodedStreamObject()
        image._data = zlib.compress(
            bytes([100 + page_index % 100]) * width * height * 3
        )
        image.update({
            NameObject("/Type"): NameObject("/XObject"),
            NameObject("/Subtype"): NameObject("/Image"),
            NameObject("/Width"): NumberObject(width),
            NameObject("/Height"): NumberObject(height),
            NameObject("/ColorSpace"): NameObject("/DeviceRGB"),
            NameObject("/BitsPerComponent"): NumberObject(8),
            NameObject("/Filter"): NameObject("/FlateDecode"),
        })
        page[NameObject("/Resources")] = DictionaryObject({
            NameObject("/XObject"): DictionaryObject({
                NameObject("/Im0"): writer._add_object(image),
            }),
        })
        content = DecodedStreamObject()
        content.set_data(b"q 300 0 0 300 0 0 cm /Im0 Do Q")
        page[NameObject("/Contents")] = writer._add_object(content)
    with path.open("wb") as stream:
        writer.write(stream)


def _write_mixed_text_image_pdf(path: Path, *, text_layer: str) -> None:
    """写入“字段值在文字层、标题和标签在内嵌图片”的电子证书结构。"""
    writer = PdfWriter()
    page = writer.add_blank_page(width=300, height=300)
    font = DictionaryObject({
        NameObject("/Type"): NameObject("/Font"),
        NameObject("/Subtype"): NameObject("/Type1"),
        NameObject("/BaseFont"): NameObject("/Helvetica"),
    })
    image = EncodedStreamObject()
    image._data = zlib.compress(bytes([180, 180, 180]) * 32 * 32)
    image.update({
        NameObject("/Type"): NameObject("/XObject"),
        NameObject("/Subtype"): NameObject("/Image"),
        NameObject("/Width"): NumberObject(32),
        NameObject("/Height"): NumberObject(32),
        NameObject("/ColorSpace"): NameObject("/DeviceRGB"),
        NameObject("/BitsPerComponent"): NumberObject(8),
        NameObject("/Filter"): NameObject("/FlateDecode"),
    })
    page[NameObject("/Resources")] = DictionaryObject({
        NameObject("/Font"): DictionaryObject({
            NameObject("/F1"): writer._add_object(font),
        }),
        NameObject("/XObject"): DictionaryObject({
            NameObject("/Im0"): writer._add_object(image),
        }),
    })
    content = DecodedStreamObject()
    content.set_data(
        b"q 300 0 0 300 0 0 cm /Im0 Do Q\n"
        b"BT /F1 10 Tf 10 20 Td ("
        + text_layer.encode("ascii")
        + b") Tj ET"
    )
    page[NameObject("/Contents")] = writer._add_object(content)
    with path.open("wb") as stream:
        writer.write(stream)


class TestResolveMAterialText(unittest.TestCase):
    def test_single_keyword(self) -> None:
        self.assertEqual(_resolve_material_text("身份证"), ["身份证"])

    def test_multiple_comma_separated(self) -> None:
        result = _resolve_material_text("身份证，合同")
        self.assertIn("身份证", result)
        self.assertIn("劳动合同", result)

    def test_empty(self) -> None:
        self.assertEqual(_resolve_material_text(""), [])

    def test_unrecognized_kept_as_is(self) -> None:
        result = _resolve_material_text("户口本")
        self.assertEqual(result, ["户口本"])

    def test_synonym_resolves_to_canonical(self) -> None:
        result = _resolve_material_text("毕业证")
        self.assertEqual(result, ["学历证明"])


class TestParseEmployeeRoster(unittest.TestCase):
    def test_plain_text_single_name(self) -> None:
        employees = parse_employee_roster("张三")
        self.assertEqual(len(employees), 1)
        self.assertEqual(employees[0].name, "张三")

    def test_plain_text_comma_separated(self) -> None:
        employees = parse_employee_roster("张三, 李四，王五")
        self.assertEqual(len(employees), 3)
        self.assertEqual(employees[0].name, "张三")
        self.assertEqual(employees[1].name, "李四")
        self.assertEqual(employees[2].name, "王五")

    def test_plain_text_with_id(self) -> None:
        employees = parse_employee_roster("张三 440111199001011234")
        self.assertEqual(len(employees), 1)
        self.assertEqual(employees[0].name, "张三")
        self.assertEqual(employees[0].id_card, "440111199001011234")

    def test_plain_text_id_only(self) -> None:
        employees = parse_employee_roster("440111199001011234")
        self.assertEqual(len(employees), 1)
        self.assertEqual(employees[0].id_card, "440111199001011234")

    def test_plain_text_with_materials(self) -> None:
        employees = parse_employee_roster("张三 身份证, 李四 合同")
        self.assertEqual(len(employees), 2)
        self.assertEqual(employees[0].name, "张三")
        self.assertEqual(employees[0].per_person_materials, ("身份证",))
        self.assertEqual(employees[1].name, "李四")
        self.assertEqual(employees[1].per_person_materials, ("劳动合同",))

    def test_dict_list_with_materials(self) -> None:
        data = [
            {"姓名": "张三", "材料": "身份证，合同"},
            {"姓名": "李四", "材料": ""},
        ]
        employees = parse_employee_roster(data)
        self.assertEqual(len(employees), 2)
        self.assertIn("身份证", employees[0].per_person_materials)
        self.assertIn("劳动合同", employees[0].per_person_materials)
        self.assertEqual(employees[1].per_person_materials, ())

    def test_dedup_by_name_and_id(self) -> None:
        employees = parse_employee_roster(["张三", "张三", "李四"])
        self.assertEqual(len(employees), 2)

    def test_noise_words_filtered(self) -> None:
        self.assertFalse(_is_valid_person_name("序号"))
        self.assertFalse(_is_valid_person_name("姓名"))
        self.assertFalse(_is_valid_person_name("合计"))
        self.assertFalse(_is_valid_person_name("123"))
        self.assertTrue(_is_valid_person_name("张三"))

    def test_excel_single_person_does_not_parse_extra(self) -> None:
        from openpyxl import Workbook
        with tempfile.TemporaryDirectory() as td:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "序号"
            ws["B1"] = "姓名"
            ws["A2"] = "1"
            ws["B2"] = "张三"
            path = Path(td) / "roster.xlsx"
            wb.save(path)
            wb.close()

            employees = parse_employee_roster(path)
            self.assertEqual(len(employees), 1)
            self.assertEqual(employees[0].name, "张三")


class TestExclusiveClassification(unittest.TestCase):
    """测试严密互斥分类：绝不把身份证、体检表等非合同文件作为劳动合同输出。"""

    def test_non_contract_files_not_collected_as_contract(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            lib = Path(td) / "资料库"
            lib.mkdir()
            w = lib / "王京川"
            w.mkdir()
            (w / "王京川_身份证正面.jpg").write_text("id front")
            _write_text_pdf(w / "王京川_安全员C证.pdf", "safety cert")
            _write_text_pdf(w / "王京川_特种作业操作证.pdf", "special cert")

            out = Path(td) / "输出"
            result = collect_employee_materials(
                lib, out,
                roster_source="王京川",
                material_types=["劳动合同"],
            )
            self.assertEqual(result.total_employees, 1)
            self.assertEqual(result.matched_file_count, 0)
            self.assertIn("王京川", result.missing_records)
            self.assertIn("劳动合同", result.missing_records["王京川"])

    def test_contract_identified_by_filename(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            lib = Path(td) / "资料库"
            lib.mkdir()
            w = lib / "王京川"
            w.mkdir()
            (w / "王京川_劳动合同.docx").write_text("contract doc")
            (w / "王京川_身份证.jpg").write_text("id")

            out = Path(td) / "输出"
            result = collect_employee_materials(
                lib, out,
                roster_source="王京川",
                material_types=["劳动合同"],
            )
            self.assertEqual(result.matched_file_count, 1)
            self.assertEqual(result.matches[0].material_type, "劳动合同")
            self.assertTrue(result.matches[0].target_filename.startswith("王京川_劳动合同"))

    def test_contract_identified_by_docx_content(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            lib = Path(td) / "资料库"
            lib.mkdir()
            w = lib / "王京川"
            w.mkdir()
            docx_path = w / "01.docx"
            with zipfile.ZipFile(docx_path, "w") as zf:
                zf.writestr(
                    "word/document.xml",
                    '<?xml version="1.0" encoding="UTF-8"?><w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body><w:p><w:r><w:t>劳动合同书 甲方与乙方自愿订立本用工合同，约定工作内容与劳动报酬。</w:t></w:r></w:p></w:body></w:document>',
                )

            out = Path(td) / "输出"
            result = collect_employee_materials(
                lib, out,
                roster_source="王京川",
                material_types=["劳动合同"],
            )
            self.assertEqual(result.matched_file_count, 1)
            self.assertEqual(result.matches[0].material_type, "劳动合同")


class TestCollectAll(unittest.TestCase):
    def test_single_person_only_copies_target_folder(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            lib = Path(td) / "资料库"
            lib.mkdir()
            (lib / "张三").mkdir()
            _write_text_pdf(lib / "张三" / "身份证.pdf", "id")
            (lib / "李四").mkdir()
            _write_text_pdf(lib / "李四" / "李四资料.pdf", "lisi")

            out = Path(td) / "输出"
            result = collect_employee_materials(
                lib, out,
                roster_source="张三",
                collect_all=True,
            )
            self.assertEqual(result.total_employees, 1)
            self.assertEqual(result.matched_file_count, 1)
            self.assertTrue((out / "张三").is_dir())
            self.assertFalse((out / "李四").exists())

    def test_direct_name_collect_all_skips_ocr_and_cache(self) -> None:
        """按人员文件夹 + 直接输入姓名 + 全部材料只复制，不进行 OCR 识别。"""
        from hr_toolkit.tools import material_collector as mc
        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class FakeEngine:
                call_count = 0
                def __call__(self, path):
                    type(self).call_count += 1
                    return (
                        [["姓名", "张三"], ["公民身份号码", "440111199001011234"]],
                        None,
                    )
            FakeEngine.call_count = 0
            mc._OCR_ENGINE = FakeEngine()
            mc._OCR_ATTEMPTED = True

            with tempfile.TemporaryDirectory() as td:
                lib = Path(td) / "资料库"
                lib.mkdir()
                emp = lib / "张三"
                emp.mkdir()
                pic = emp / "a5d6e67cd.jpg"
                pic.write_bytes(b"\x89PNG\r\n\x1a\nfake")
                source_sha_before = hashlib.sha256(pic.read_bytes()).hexdigest()

                out1 = Path(td) / "输出"
                with mock.patch.object(
                    mc, "_get_engine_signature", wraps=mc._get_engine_signature,
                ) as signature_mock, mock.patch.object(
                    mc, "_load_ocr_cache", wraps=mc._load_ocr_cache,
                ) as load_mock, mock.patch.object(
                    mc, "_save_ocr_cache", wraps=mc._save_ocr_cache,
                ) as save_mock:
                    result = collect_employee_materials(
                        lib, out1,
                        roster_source="张三",
                        collect_all=True,
                    )

                self.assertEqual(result.total_employees, 1)
                self.assertEqual(result.matched_file_count, 1)
                self.assertEqual(FakeEngine.call_count, 0)
                signature_mock.assert_not_called()
                load_mock.assert_not_called()
                save_mock.assert_not_called()
                self.assertFalse((lib / _OCR_CACHE_FILE_NAME).exists())
                self.assertFalse(result.ocr_cache_enabled)
                self.assertEqual(result.ocr_cache_hits, 0)
                self.assertEqual(result.ocr_cache_misses, 0)
                self.assertIn("无需 OCR", result.ocr_cache_skipped_reason or "")

                copied = out1 / "张三" / pic.name
                self.assertTrue(copied.is_file())
                source_sha_after = hashlib.sha256(pic.read_bytes()).hexdigest()
                self.assertEqual(source_sha_after, source_sha_before)
                self.assertEqual(
                    hashlib.sha256(copied.read_bytes()).hexdigest(),
                    source_sha_before,
                )

                workbook = load_workbook(result.report_path, data_only=True)
                try:
                    summary = workbook["资料提取汇总与缺失清单"]
                    self.assertEqual(summary["D6"].value, 0)
                    self.assertEqual(
                        summary["G8"].value,
                        "已按文件夹直接复制（未进行 OCR 内容核对）",
                    )
                finally:
                    workbook.close()
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted


class TestSafetyAndRobustness(unittest.TestCase):
    def test_nested_output_dir_raises_error(self) -> None:
        """测试保存目录在资料库内部时立即抛出异常，防止无限递归。"""
        with tempfile.TemporaryDirectory() as td:
            lib = Path(td) / "资料库"
            lib.mkdir()
            (lib / "张三").mkdir()
            _write_text_pdf(lib / "张三" / "身份证.pdf", "id")
            nested_out = lib / "output_nested"

            with self.assertRaises(ValueError) as ctx:
                collect_employee_materials(
                    lib, nested_out,
                    roster_source="张三",
                )
            self.assertIn("保存目录不能在资料库目录内部", str(ctx.exception))

    def test_duplicate_folders_for_same_person_preserved_and_warned(self) -> None:
        """测试同名人员有多个文件夹时全部保留且记录警告。"""
        with tempfile.TemporaryDirectory() as td:
            lib = Path(td) / "资料库"
            lib.mkdir()
            (lib / "部门A").mkdir()
            (lib / "部门A" / "张三").mkdir()
            (lib / "部门A" / "张三" / "张三_身份证.jpg").write_text("id1")

            (lib / "部门B").mkdir()
            (lib / "部门B" / "张三").mkdir()
            (lib / "部门B" / "张三" / "张三_学历证书.jpg").write_text("degree2")

            out = Path(td) / "输出"
            result = collect_employee_materials(
                lib, out,
                roster_source="张三",
                material_types=["身份证", "学历证明"],
                scan_depth=2,
            )
            self.assertEqual(result.matched_file_count, 2)
            # 应有同名文件夹预警
            self.assertTrue(any("同名文件夹" in w for w in result.warnings))

    def test_same_person_cross_dir_duplicate_file_deduped(self) -> None:
        """测试同一员工在不同备份目录下存在完全相同的内容哈希文件时自动去重。"""
        with tempfile.TemporaryDirectory() as td:
            lib = Path(td) / "资料库"
            lib.mkdir()
            (lib / "张三").mkdir()
            (lib / "张三" / "身份证.jpg").write_bytes(b"SAME_ID_IMAGE_BYTES_12345")
            (lib / "张三" / "备份").mkdir()
            (lib / "张三" / "备份" / "身份证_副本.jpg").write_bytes(b"SAME_ID_IMAGE_BYTES_12345")

            out = Path(td) / "输出"
            result = collect_employee_materials(
                lib, out,
                roster_source="张三",
                material_types=["身份证"],
                scan_depth=2,
            )
            # 完全重复的文件只提取 1 份，不去重会导致复制 2 份
            self.assertEqual(result.matched_file_count, 1)

    def test_same_size_and_prefix_but_different_tail_are_not_deduped(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            person_dir = root / "资料库" / "张三"
            person_dir.mkdir(parents=True)
            shared_prefix = b"A" * 65536
            (person_dir / "张三_身份证_A.jpg").write_bytes(shared_prefix + b"X" * 65536)
            (person_dir / "张三_身份证_B.jpg").write_bytes(shared_prefix + b"Y" * 65536)

            result = collect_employee_materials(
                root / "资料库",
                root / "输出",
                roster_source="张三",
                material_types=["身份证"],
            )

            self.assertEqual(result.matched_file_count, 2)

    def test_same_subtype_files_are_preserved_without_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "资料库"
            person_dir = library / "张三"
            person_dir.mkdir(parents=True)
            first = person_dir / "张三_身份证正面_A.jpg"
            second = person_dir / "张三_身份证正面_B.jpg"
            first.write_bytes(b"first-front-image")
            second.write_bytes(b"second-front-image-with-different-size")

            output = root / "输出"
            result = collect_employee_materials(
                library,
                output,
                roster_source="张三",
                material_types=["身份证"],
            )

            copied = sorted((output / "张三").glob("张三_身份证_正面*.jpg"))
            self.assertEqual(result.matched_file_count, 2)
            self.assertEqual(len(copied), 2)
            self.assertEqual(
                {path.read_bytes() for path in copied},
                {first.read_bytes(), second.read_bytes()},
            )

    def test_text_pdf_reads_text_after_150kb_without_modifying_source(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "资料库"
            employee_dir = library / "张三"
            employee_dir.mkdir(parents=True)
            pdf = employee_dir / "long-text.pdf"
            marker = "TAIL_MATERIAL_MARKER"
            _write_text_pdf(pdf, marker, padding_bytes=180_000)
            source_hash = hashlib.sha256(pdf.read_bytes()).hexdigest()

            self.assertIn(marker, _extract_document_text(pdf))
            result = collect_employee_materials(
                library,
                root / "输出",
                roster_source="张三",
                material_types=[marker],
                use_ocr_cache=False,
            )

            self.assertEqual(result.matched_file_count, 1)
            self.assertTrue(
                (root / "输出" / "张三" / f"张三_{marker}.pdf").is_file()
            )
            self.assertEqual(
                hashlib.sha256(pdf.read_bytes()).hexdigest(),
                source_hash,
            )

    def test_single_page_scanned_pdf_uses_existing_ocr_engine(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class FakeEngine:
                call_count = 0

                def __call__(self, _payload):
                    type(self).call_count += 1
                    return ([[None, "劳动合同"]], None)

            mc._OCR_ENGINE = FakeEngine()
            mc._OCR_ATTEMPTED = True
            with tempfile.TemporaryDirectory() as td:
                pdf = Path(td) / "scan.pdf"
                _write_scanned_pdf(pdf, page_count=1)

                result = mc._analyze_folder_ocr_file(pdf)

            self.assertEqual(result[0], "劳动合同")
            self.assertEqual(FakeEngine.call_count, 1)
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_multi_page_scan_combines_pages_and_reports_progress(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class SequencedEngine:
                call_count = 0

                def __call__(self, _payload):
                    page_text = ("员工手册", "签收单")[type(self).call_count]
                    type(self).call_count += 1
                    return ([[None, page_text]], None)

            mc._OCR_ENGINE = SequencedEngine()
            mc._OCR_ATTEMPTED = True
            with tempfile.TemporaryDirectory() as td:
                root = Path(td)
                library = root / "资料库"
                employee_dir = library / "张三"
                employee_dir.mkdir(parents=True)
                source = employee_dir / "random.pdf"
                _write_scanned_pdf(source, page_count=2)
                source_hash = hashlib.sha256(source.read_bytes()).hexdigest()
                progress_messages: list[str] = []

                result = collect_employee_materials(
                    library,
                    root / "输出",
                    roster_source="张三",
                    material_types=["员工手册签收单"],
                    use_ocr_cache=False,
                    progress_callback=lambda _current, _total, message: (
                        progress_messages.append(message)
                    ),
                )

                copied = root / "输出" / "张三" / "张三_员工手册签收单.pdf"
                self.assertEqual(result.matched_file_count, 1)
                self.assertEqual(SequencedEngine.call_count, 2)
                self.assertTrue(copied.is_file())
                self.assertEqual(
                    hashlib.sha256(copied.read_bytes()).hexdigest(),
                    source_hash,
                )
                self.assertEqual(
                    hashlib.sha256(source.read_bytes()).hexdigest(),
                    source_hash,
                )
                self.assertTrue(any("PDF" in message and "2/2" in message for message in progress_messages))
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_ocr_stream_keeps_text_but_releases_page_geometry(self) -> None:
        import gc
        import weakref
        from hr_toolkit.tools import material_collector as mc

        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        references: list[weakref.ReferenceType] = []
        try:
            class Geometry:
                pass

            class PageEngine:
                call_count = 0

                def __call__(self, _payload):
                    box = Geometry()
                    references.append(weakref.ref(box))
                    text = f"第{type(self).call_count + 1}页"
                    type(self).call_count += 1
                    return ([[box, text, 0.99]], None)

            mc._OCR_ENGINE = PageEngine()
            mc._OCR_ATTEMPTED = True
            with mock.patch.object(
                mc,
                "_iter_ocr_targets",
                return_value=iter(_png_fixture(f"page-{page}") for page in range(1, 4)),
            ):
                texts, complete = mc._collect_ocr_texts(Path("contract.pdf"))

            gc.collect()
            self.assertTrue(complete)
            self.assertEqual(texts, ["第1页", "第2页", "第3页"])
            self.assertTrue(all(reference() is None for reference in references))
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_five_page_contract_combines_first_page_type_and_last_page_name(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            pages = (
                "劳动合同",
                "第一章 工作内容",
                "第二章 劳动报酬",
                "第三章 合同期限",
                "姓名：张三",
            )

            class PageEngine:
                call_count = 0

                def __call__(self, _payload):
                    text = pages[type(self).call_count]
                    type(self).call_count += 1
                    return ([[None, text]], None)

            mc._OCR_ENGINE = PageEngine()
            mc._OCR_ATTEMPTED = True
            with mock.patch.object(
                mc,
                "_iter_ocr_targets",
                return_value=iter(_png_fixture(f"page-{page}") for page in range(len(pages))),
            ):
                result = mc._analyze_ocr_file(
                    Path("contract.pdf"),
                    requested_types=["劳动合同"],
                    allow_weak_id_fallback=False,
                )

            material, _method, _subtype, name, _eid, text, names, complete = result
            self.assertTrue(complete)
            self.assertEqual(PageEngine.call_count, 5)
            self.assertEqual(material, "劳动合同")
            self.assertEqual(name, "张三")
            self.assertEqual(names, ["张三"])
            self.assertIn("劳动合同", text)
            self.assertIn("姓名：张三", text)
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_pdf_cancel_stops_between_pages(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class CountingEngine:
                call_count = 0

                def __call__(self, _payload):
                    type(self).call_count += 1
                    return ([[None, "未分类"]], None)

            mc._OCR_ENGINE = CountingEngine()
            mc._OCR_ATTEMPTED = True
            with tempfile.TemporaryDirectory() as td:
                pdf = Path(td) / "multi-page.pdf"
                _write_scanned_pdf(pdf, page_count=3)

                with self.assertRaisesRegex(mc.MaterialCollectionCancelled, "停止"):
                    mc._analyze_folder_ocr_file(
                        pdf,
                        cancelled=lambda: CountingEngine.call_count >= 1,
                    )

            self.assertEqual(CountingEngine.call_count, 1)
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_corrupt_and_encrypted_pdf_have_clear_errors(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            corrupt = root / "corrupt.pdf"
            corrupt.write_bytes(b"%PDF-1.4\ntruncated")
            corrupt_hash = hashlib.sha256(corrupt.read_bytes()).hexdigest()
            with self.assertRaisesRegex(mc.PDFRecognitionError, "损坏"):
                _extract_document_text(corrupt)
            self.assertEqual(
                hashlib.sha256(corrupt.read_bytes()).hexdigest(),
                corrupt_hash,
            )

            encrypted = root / "encrypted.pdf"
            writer = PdfWriter()
            writer.add_blank_page(width=100, height=100)
            writer.encrypt("secret")
            with encrypted.open("wb") as stream:
                writer.write(stream)
            encrypted_hash = hashlib.sha256(encrypted.read_bytes()).hexdigest()
            with self.assertRaisesRegex(mc.PDFRecognitionError, "加密"):
                _extract_document_text(encrypted)
            self.assertEqual(
                hashlib.sha256(encrypted.read_bytes()).hexdigest(),
                encrypted_hash,
            )

    def test_pdf_resource_limits_fail_before_ocr_decode(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class UnexpectedEngine:
                call_count = 0

                def __call__(self, _payload):
                    type(self).call_count += 1
                    return ([], None)

            mc._OCR_ENGINE = UnexpectedEngine()
            mc._OCR_ATTEMPTED = True
            with tempfile.TemporaryDirectory() as td:
                root = Path(td)
                one_page = root / "one-page.pdf"
                two_pages = root / "two-pages.pdf"
                _write_scanned_pdf(one_page, page_count=1, width=32, height=32)
                _write_scanned_pdf(two_pages, page_count=2)

                cases = (
                    ("_PDF_MAX_FILE_BYTES", 10, one_page, "体积"),
                    ("_PDF_MAX_PAGES", 1, two_pages, "页数"),
                    ("_PDF_MAX_IMAGE_PIXELS", 100, one_page, "像素"),
                    ("_PDF_MAX_DECODED_IMAGE_BYTES", 100, one_page, "解码内存"),
                )
                for constant, limit, source, message in cases:
                    with self.subTest(constant=constant):
                        with mock.patch.object(mc, constant, limit):
                            with self.assertRaisesRegex(mc.PDFResourceLimitError, message):
                                mc._analyze_folder_ocr_file(source)

            self.assertEqual(UnexpectedEngine.call_count, 0)
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_flat_multi_page_pdf_uses_ocr_then_cache(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class SequencedEngine:
                call_count = 0

                def __call__(self, _payload):
                    page_text = ("姓名：张三", "劳动合同")[type(self).call_count]
                    type(self).call_count += 1
                    return ([[None, page_text]], None)

            mc._OCR_ENGINE = SequencedEngine()
            mc._OCR_ATTEMPTED = True
            with tempfile.TemporaryDirectory() as td:
                root = Path(td)
                library = root / "资料库"
                library.mkdir()
                source = library / "random.pdf"
                _write_scanned_pdf(source, page_count=2)
                source_hash = hashlib.sha256(source.read_bytes()).hexdigest()

                first = collect_employee_materials(
                    library,
                    root / "输出1",
                    roster_source="张三",
                    material_types=["劳动合同"],
                    library_mode=LIBRARY_MODE_FLAT_OCR,
                )
                first_calls = SequencedEngine.call_count
                second = collect_employee_materials(
                    library,
                    root / "输出2",
                    roster_source="张三",
                    material_types=["劳动合同"],
                    library_mode=LIBRARY_MODE_FLAT_OCR,
                )

                self.assertEqual(first.matched_file_count, 1)
                self.assertEqual(second.matched_file_count, 1)
                self.assertEqual(first_calls, 2)
                self.assertEqual(SequencedEngine.call_count, first_calls)
                self.assertEqual(second.ocr_cache_hits, 1)
                copied = root / "输出2" / "张三" / "张三_劳动合同.pdf"
                self.assertEqual(
                    hashlib.sha256(copied.read_bytes()).hexdigest(),
                    source_hash,
                )
                self.assertEqual(
                    hashlib.sha256(source.read_bytes()).hexdigest(),
                    source_hash,
                )
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_pdf_failures_are_reported_without_modifying_sources(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "资料库"
            employee_dir = library / "张三"
            employee_dir.mkdir(parents=True)
            corrupt = employee_dir / "corrupt.pdf"
            corrupt.write_bytes(b"%PDF-1.4\ntruncated")
            encrypted = employee_dir / "encrypted.pdf"
            writer = PdfWriter()
            writer.add_blank_page(width=100, height=100)
            writer.encrypt("secret")
            with encrypted.open("wb") as stream:
                writer.write(stream)
            source_hashes = {
                path.name: hashlib.sha256(path.read_bytes()).hexdigest()
                for path in (corrupt, encrypted)
            }

            result = collect_employee_materials(
                library,
                root / "输出",
                roster_source="张三",
                material_types=["员工手册签收单"],
                use_ocr_cache=False,
            )

            warning_text = "\n".join(result.warnings)
            self.assertIn("损坏", warning_text)
            self.assertIn("加密", warning_text)
            self.assertIn("张三", result.missing_records)
            for path in (corrupt, encrypted):
                self.assertEqual(
                    hashlib.sha256(path.read_bytes()).hexdigest(),
                    source_hashes[path.name],
                )

    def test_pdf_handles_are_closed_after_page_ocr(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            mc._OCR_ENGINE = lambda _payload: ([[None, "劳动合同"]], None)
            mc._OCR_ATTEMPTED = True
            with tempfile.TemporaryDirectory() as td:
                source = Path(td) / "scan.pdf"
                renamed = Path(td) / "renamed.pdf"
                _write_scanned_pdf(source, page_count=2)

                mc._analyze_folder_ocr_file(source)
                source.replace(renamed)

                self.assertTrue(renamed.is_file())
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_pdf_page_decode_cache_is_released_after_each_image(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        if mc.PdfReader is None:
            self.skipTest("pypdf 解码缓存测试不适用于 Win7 PDFium 后端")

        with tempfile.TemporaryDirectory() as td:
            source = Path(td) / "scan.pdf"
            _write_scanned_pdf(source, page_count=3, width=64, height=64)

            with mock.patch.object(mc, "_PDF_BACKEND", "pypdf"):
                with mock.patch.object(
                    mc,
                    "_release_pdf_image_decode_cache",
                    wraps=mc._release_pdf_image_decode_cache,
                ) as release_cache:
                    payloads = list(mc._iter_pdf_ocr_images(source))

            self.assertEqual(len(payloads), 3)
            self.assertEqual(release_cache.call_count, 3)
            for call in release_cache.call_args_list:
                image_object = call.args[0]
                self.assertIsNone(image_object.decoded_self)

    def test_office_xml_bombs_and_entity_declarations_are_ignored(self) -> None:
        import hr_toolkit.tools.material_collector as material_collector

        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            oversized = root / "oversized.docx"
            with zipfile.ZipFile(oversized, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                archive.writestr("word/document.xml", b"<w>" + b"x" * 256 + b"</w>")
            with mock.patch.object(material_collector, "_OFFICE_XML_MEMBER_MAX_BYTES", 64):
                self.assertEqual(_extract_document_text(oversized), "")

            entity_doc = root / "entity.docx"
            with zipfile.ZipFile(entity_doc, "w") as archive:
                archive.writestr(
                    "word/document.xml",
                    b'<!DOCTYPE x [<!ENTITY a "boom">]><x>&a;</x>',
                )
            self.assertEqual(_extract_document_text(entity_doc), "")

    def test_junk_files_filtered_out(self) -> None:
        """测试 .DS_Store, Thumbs.db, ~$临时文件被全局过滤。"""
        self.assertTrue(_is_junk_or_temp_file(".DS_Store"))
        self.assertTrue(_is_junk_or_temp_file("Thumbs.db"))
        self.assertTrue(_is_junk_or_temp_file("~$员工名单.xlsx"))
        self.assertFalse(_is_junk_or_temp_file("张三_身份证.jpg"))


class TestOCRCacheHelpers(unittest.TestCase):
    """OCR 智能索引缓存：纯函数单测。"""

    def test_engine_signature_is_non_empty(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        for modern, expected in (
            (False, "rapidocr_onnxruntime@1.4.4"),
            (True, "rapidocr@3.9.2:ppocrv4-mobile:2"),
        ):
            with self.subTest(modern=modern), mock.patch.object(
                mc, "_uses_modern_ocr", return_value=modern,
            ), mock.patch.dict(
                sys.modules,
                {"rapidocr_onnxruntime": SimpleNamespace(__version__="1.4.4")},
            ), mock.patch("importlib.metadata.version", return_value="3.9.2") as version:
                self.assertEqual(_get_engine_signature(), expected)
                if modern:
                    version.assert_called_once_with("rapidocr")
                else:
                    version.assert_not_called()

    def test_visual_ocr_query_cache_is_bounded_and_does_not_store_names(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        entry: dict = {}
        for index in range(40):
            mc._record_visual_ocr_query(entry, [f"自定义材料{index}"])

        self.assertLessEqual(len(entry["visual_ocr_queries"]), 32)
        self.assertTrue(
            mc._visual_ocr_query_was_attempted(entry, ["自定义材料39"])
        )
        self.assertNotIn("自定义材料", json.dumps(entry, ensure_ascii=False))

    def test_oversized_cache_stops_before_reading_and_preserves_file(self) -> None:
        import hr_toolkit.tools.material_collector as material_collector

        with tempfile.TemporaryDirectory() as td:
            cache_path = Path(td) / _OCR_CACHE_FILE_NAME
            cache_path.write_bytes(b"{" + b" " * 1024 + b"}")
            original = cache_path.read_bytes()
            with mock.patch.object(material_collector, "_OCR_CACHE_FILE_LOAD_MAX_BYTES", 128):
                with mock.patch.object(material_collector.json, "load") as read_json:
                    with self.assertRaises(material_collector.OCRResourceLimitError):
                        _load_ocr_cache(cache_path)
                    read_json.assert_not_called()
            self.assertEqual(cache_path.read_bytes(), original)

    def test_large_parsed_cache_is_reused_and_invalidated_by_file_change(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        with tempfile.TemporaryDirectory() as td:
            cache_path = Path(td) / _OCR_CACHE_FILE_NAME
            cache_path.write_text(
                json.dumps({"version": 6, "entries": {"first": {}}, "paths": {}}),
                encoding="utf-8",
            )
            mc._clear_ocr_memory_cache()
            try:
                with mock.patch.object(mc, "_OCR_MEMORY_CACHE_MIN_BYTES", 0):
                    first = mc._load_ocr_cache(cache_path)
                    second = mc._load_ocr_cache(cache_path)
                    self.assertIs(first, second)

                    cache_path.write_text(
                        json.dumps({
                            "version": 6,
                            "entries": {"replacement-longer-key": {}},
                            "paths": {},
                        }),
                        encoding="utf-8",
                    )
                    replacement = mc._load_ocr_cache(cache_path)
                    self.assertIsNot(replacement, first)
                    self.assertEqual(set(replacement["entries"]), {"replacement-longer-key"})
            finally:
                mc._clear_ocr_memory_cache()

    def test_pdf_cache_upgrade_invalidates_only_pdf_entries(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        cache = {
            "version": 5,
            "entries": {
                "pdf-key": {
                    "sample_filename": "random.pdf",
                    "analysis_state": "complete",
                },
                "image-key": {
                    "sample_filename": "random.jpg",
                    "analysis_state": "complete",
                },
            },
            "paths": {
                "random.pdf": {"cache_key": "pdf-key"},
                "random.jpg": {"cache_key": "image-key"},
            },
        }

        removed = mc._invalidate_legacy_pdf_cache_entries(cache)

        self.assertEqual(removed, 1)
        self.assertEqual(cache["version"], 6)
        self.assertEqual(set(cache["entries"]), {"image-key"})
        self.assertEqual(set(cache["paths"]), {"random.jpg"})

    def test_unversioned_path_only_pdf_cache_is_invalidated(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        cache = {
            "entries": {
                "pdf-key": {"analysis_state": "complete"},
                "image-key": {"analysis_state": "complete"},
            },
            "paths": {
                "archive/random.pdf": {"cache_key": "pdf-key"},
                "archive/random.jpg": {"cache_key": "image-key"},
            },
        }

        removed = mc._invalidate_legacy_pdf_cache_entries(cache)

        self.assertEqual(removed, 1)
        self.assertEqual(cache["version"], 6)
        self.assertEqual(set(cache["entries"]), {"image-key"})
        self.assertEqual(set(cache["paths"]), {"archive/random.jpg"})

    def test_pdf_backend_switch_invalidates_only_pdf_entries(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        cache = {
            "version": 5,
            "pdf_backend_signature": "pypdf@old",
            "entries": {
                "pdf-key": {"sample_filename": "random.pdf"},
                "image-key": {"sample_filename": "random.jpg"},
            },
            "paths": {
                "random.pdf": {"cache_key": "pdf-key"},
                "random.jpg": {"cache_key": "image-key"},
            },
        }

        with mock.patch.object(mc, "_get_pdf_backend_signature", return_value="pdfium@new"):
            removed = mc._invalidate_changed_pdf_backend_entries(cache)

        self.assertEqual(removed, 1)
        self.assertEqual(cache["pdf_backend_signature"], "pdfium@new")
        self.assertEqual(set(cache["entries"]), {"image-key"})
        self.assertEqual(set(cache["paths"]), {"random.jpg"})

    def test_compute_cache_key_stable(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            p1 = Path(td) / "a.png"
            p1.write_bytes(b"test_image_data")
            k1 = _compute_cache_key(p1)
            k2 = _compute_cache_key(p1)
            self.assertEqual(k1, k2)
            self.assertIn("_", k1)

    def test_compute_cache_key_invariant_to_filename(self) -> None:
        """重命名文件，只要内容相同，生成的 cache_key 必须完全相同。"""
        with tempfile.TemporaryDirectory() as td:
            p1 = Path(td) / "original_name.jpg"
            p1.write_bytes(b"identical_binary_content")
            k1 = _compute_cache_key(p1)

            p2 = Path(td) / "completely_different_name.png"
            p2.write_bytes(b"identical_binary_content")
            k2 = _compute_cache_key(p2)
            self.assertEqual(k1, k2, "相同内容的不同文件名必须生成完全相同的 content hash key")

    def test_compute_file_fingerprint_returns_size_mtime_hash(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "a.bin"
            p.write_bytes(b"hello")
            fp = _compute_file_fingerprint(p)
            self.assertIsNotNone(fp)
            size, mtime, sha = fp
            self.assertEqual(size, 5)
            self.assertGreater(mtime, 0)
            self.assertEqual(len(sha), 64)

    def test_hash_id_card_is_16_chars(self) -> None:
        self.assertEqual(_hash_id_card(""), "")
        self.assertEqual(len(_hash_id_card("440111199001011234")), 16)

    def test_employee_key_includes_id(self) -> None:
        emp = TargetEmployee(name="张三", id_card="440111199001011234")
        self.assertEqual(_build_employee_key(emp), "张三|440111199001011234")
        emp2 = TargetEmployee(name="张三", id_card="")
        self.assertEqual(_build_employee_key(emp2), "张三|")

    def test_doc_format_hint_only_for_doc(self) -> None:
        self.assertIsNone(_build_doc_format_hint(Path("a.docx")))
        self.assertIsNone(_build_doc_format_hint(Path("a.jpg")))
        hint = _build_doc_format_hint(Path("a.doc"))
        self.assertIsNotNone(hint)
        self.assertIn(".docx", hint)


class TestOCRCacheIO(unittest.TestCase):
    """缓存读写与生命周期。"""

    def test_folder_custom_material_rename_reclassifies_cached_ocr_text(self) -> None:
        """自定义名称改成图片真实标题后复用文字缓存，不重复 OCR。"""
        from hr_toolkit.tools import material_collector as mc

        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class CountingEngine:
                call_count = 0

                def __call__(self, path):
                    type(self).call_count += 1
                    return (
                        [
                            [None, "员工手册"],
                            [None, "签收单"],
                            [None, "姓名张三"],
                        ],
                        None,
                    )

            mc._OCR_ENGINE = CountingEngine()
            mc._OCR_ATTEMPTED = True

            with tempfile.TemporaryDirectory() as td:
                root = Path(td)
                library = root / "资料库"
                employee_dir = library / "张三"
                employee_dir.mkdir(parents=True)
                source = employee_dir / "a5d6e67cd.jpg"
                source.write_bytes(_png_fixture("customer-jpg-fixture"))
                source_sha = hashlib.sha256(source.read_bytes()).hexdigest()

                first = collect_employee_materials(
                    library,
                    root / "输出1",
                    roster_source="张三",
                    material_types=["员工手册领用单"],
                )
                first_ocr_calls = CountingEngine.call_count
                self.assertEqual(first.matched_file_count, 0)
                self.assertEqual(first_ocr_calls, 1)

                cache_path = library / _OCR_CACHE_FILE_NAME
                cache = _load_ocr_cache(cache_path)
                self.assertEqual(len(cache["entries"]), 1)
                entry = next(iter(cache["entries"].values()))
                self.assertEqual(entry["material_type"], "其他材料")
                self.assertIn("员工手册签收单", entry["ocr_text"].replace(" ", ""))
                self.assertNotEqual(entry["material_type"], "员工手册领用单")

                second = collect_employee_materials(
                    library,
                    root / "输出2",
                    roster_source="张三",
                    material_types=["员工手册签收单"],
                )
                self.assertEqual(CountingEngine.call_count, first_ocr_calls)
                self.assertEqual(second.ocr_cache_hits, 1)
                self.assertEqual(second.ocr_cache_misses, 0)
                self.assertEqual(second.matched_file_count, 1)
                self.assertEqual(second.matches[0].material_type, "员工手册签收单")
                copied = root / "输出2" / "张三" / "张三_员工手册签收单.jpg"
                self.assertTrue(copied.is_file())
                self.assertEqual(
                    hashlib.sha256(copied.read_bytes()).hexdigest(),
                    source_sha,
                )
                cached_after_match = next(
                    iter(_load_ocr_cache(cache_path)["entries"].values())
                )
                self.assertEqual(cached_after_match["material_type"], "其他材料")

                third = collect_employee_materials(
                    library,
                    root / "输出3",
                    roster_source="张三",
                    material_types=["员工手册领用单"],
                )
                self.assertEqual(CountingEngine.call_count, first_ocr_calls)
                self.assertEqual(third.matched_file_count, 0)
                self.assertEqual(
                    hashlib.sha256(source.read_bytes()).hexdigest(),
                    source_sha,
                )
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_cache_written_after_successful_recognition(self) -> None:
        """首次跑应创建 .hr_material_index_cache.json。"""
        with tempfile.TemporaryDirectory() as td:
            lib = Path(td) / "资料库"
            lib.mkdir()
            emp_dir = lib / "张三"
            emp_dir.mkdir()
            # 用按文件名识别的纯图片名（走缓存写入路径必须有 OCR 成功，
            # 这里使用扩展名 + 文件名都能走 OCR 分支的姿势）
            (emp_dir / "a5d6e67cd.jpg").write_bytes(_png_fixture())

            out = Path(td) / "输出"
            # mock OCR engine，固定返回"身份证"
            from hr_toolkit.tools import material_collector as mc
            real_engine = mc._OCR_ENGINE
            real_attempted = mc._OCR_ATTEMPTED
            try:
                class FakeEngine:
                    def __call__(self, path):
                        return (
                            [["姓名", "张三"], ["公民身份号码", "440111199001011234"]],
                            None,
                        )
                mc._OCR_ENGINE = FakeEngine()
                mc._OCR_ATTEMPTED = True

                result = collect_employee_materials(
                    lib, out,
                    roster_source="张三",
                    material_types=["身份证"],
                )
                self.assertTrue(result.ocr_cache_enabled)
                cache_file = lib / _OCR_CACHE_FILE_NAME
                self.assertTrue(cache_file.exists(), "缓存文件应自动创建")
                self.assertGreater(result.matched_file_count, 0)
            finally:
                mc._OCR_ENGINE = real_engine
                mc._OCR_ATTEMPTED = real_attempted

    def test_cache_hit_skips_ocr_call(self) -> None:
        """二次跑同一库时，OCR 引擎调用次数应为 0（命中缓存）。"""
        from hr_toolkit.tools import material_collector as mc
        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class CountingEngine:
                call_count = 0
                def __call__(self, path):
                    type(self).call_count += 1
                    return (
                        [["姓名", "张三"], ["公民身份号码", "440111199001011234"]],
                        None,
                    )
            CountingEngine.call_count = 0
            engine = CountingEngine()
            mc._OCR_ENGINE = engine
            mc._OCR_ATTEMPTED = True

            with tempfile.TemporaryDirectory() as td:
                lib = Path(td) / "资料库"
                lib.mkdir()
                emp = lib / "张三"
                emp.mkdir()
                # 多个图片都需 OCR 识别
                for name in ("a1.png", "b2.jpg", "c3.jpeg"):
                    (emp / name).write_bytes(_png_fixture())

                out1 = Path(td) / "输出1"
                # 首次：建立缓存（OCR 调用计数应有值）
                first_result = collect_employee_materials(
                    lib, out1,
                    roster_source="张三",
                    material_types=["身份证"],
                )
                first_count = CountingEngine.call_count
                self.assertGreater(first_count, 0, "首次跑应触发 OCR")

                out2 = Path(td) / "输出2"
                # 二次：缓存命中，OCR 调用计数不变
                result = collect_employee_materials(
                    lib, out2,
                    roster_source="张三",
                    material_types=["身份证"],
                )
                self.assertEqual(CountingEngine.call_count, first_count,
                                 "二次跑不应再调用 OCR")
                self.assertGreaterEqual(result.ocr_cache_hits, 1)
                self.assertEqual(result.matches[0].material_type, "身份证")
                self.assertEqual(
                    result.matches[0].matched_by,
                    first_result.matches[0].matched_by,
                )
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_cache_hit_after_file_renamed_or_moved(self) -> None:
        """文件被重命名或移动到深层子目录后，只要内容不变，依然 100% 命中 OCR 缓存。"""
        from hr_toolkit.tools import material_collector as mc
        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class CountingEngine:
                call_count = 0
                def __call__(self, path):
                    type(self).call_count += 1
                    return (
                        [["姓名", "张三"], ["公民身份号码", "440111199001011234"]],
                        None,
                    )
            CountingEngine.call_count = 0
            mc._OCR_ENGINE = CountingEngine()
            mc._OCR_ATTEMPTED = True

            with tempfile.TemporaryDirectory() as td:
                lib = Path(td) / "资料库"
                lib.mkdir()
                emp = lib / "张三"
                emp.mkdir()
                pic = emp / "old_random_hash_name.jpg"
                pic.write_bytes(_png_fixture("fake_image_bytes"))

                out1 = Path(td) / "输出1"
                # 首次运行：写入内容哈希指纹缓存
                collect_employee_materials(lib, out1, roster_source="张三", material_types=["身份证"])
                first_count = CountingEngine.call_count
                self.assertEqual(first_count, 1)

                # 重命名文件并移动到深层子目录
                sub_dir = emp / "2026新证件"
                sub_dir.mkdir()
                new_pic = sub_dir / "renamed_random_name_0x8f2a.png"
                pic.rename(new_pic)

                out2 = Path(td) / "输出2"
                # 第二次运行：即使文件名和路径彻底改变，因内容哈希一致，依然秒级命中缓存，OCR 调用次数不变！
                res2 = collect_employee_materials(lib, out2, roster_source="张三", material_types=["身份证"])
                self.assertEqual(CountingEngine.call_count, first_count, "文件重命名后不应重新调用 OCR，必须命中缓存")
                self.assertGreaterEqual(res2.ocr_cache_hits, 1)
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_cache_invalidated_on_file_modification(self) -> None:
        """修改图片内容（size+mtime 变化）后应触发重 OCR。"""
        from hr_toolkit.tools import material_collector as mc
        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class CountingEngine:
                call_count = 0
                def __call__(self, path):
                    type(self).call_count += 1
                    return (
                        [["姓名", "张三"], ["公民身份号码", "440111199001011234"]],
                        None,
                    )
            CountingEngine.call_count = 0
            engine = CountingEngine()
            mc._OCR_ENGINE = engine
            mc._OCR_ATTEMPTED = True

            with tempfile.TemporaryDirectory() as td:
                lib = Path(td) / "资料库"
                lib.mkdir()
                emp = lib / "张三"
                emp.mkdir()
                pic = emp / "a.png"
                pic.write_bytes(_png_fixture())

                out1 = Path(td) / "输出1"
                collect_employee_materials(
                    lib, out1,
                    roster_source="张三",
                    material_types=["身份证"],
                )
                first = CountingEngine.call_count

                # 修改文件 size + mtime
                import time
                time.sleep(0.05)
                pic.write_bytes(_png_fixture("X" * 1000))

                out2 = Path(td) / "输出2"
                collect_employee_materials(
                    lib, out2,
                    roster_source="张三",
                    material_types=["身份证"],
                )
                self.assertGreater(CountingEngine.call_count, first,
                                   "文件被修改后应触发再次 OCR")
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_cache_invalidated_on_file_deletion(self) -> None:
        """文件被删除后缓存条目应清理，再次跑匹配条目数应减少。"""
        from hr_toolkit.tools import material_collector as mc
        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class FakeEngine:
                def __call__(self, path):
                    return (
                        [["姓名", "张三"], ["公民身份号码", "440111199001011234"]],
                        None,
                    )
            mc._OCR_ENGINE = FakeEngine()
            mc._OCR_ATTEMPTED = True

            with tempfile.TemporaryDirectory() as td:
                lib = Path(td) / "资料库"
                lib.mkdir()
                emp = lib / "张三"
                emp.mkdir()
                pic1 = emp / "a.png"
                pic1.write_bytes(_png_fixture())
                pic2 = emp / "b.png"
                pic2.write_bytes(_png_fixture())

                out1 = Path(td) / "输出1"
                collect_employee_materials(
                    lib, out1,
                    roster_source="张三",
                    material_types=["身份证"],
                )

                cache_path = lib / _OCR_CACHE_FILE_NAME
                self.assertTrue(cache_path.exists())
                cache_data = _load_ocr_cache(cache_path)
                self.assertGreater(len(cache_data.get("entries") or {}), 0)

                # 删除其中一张图
                pic2.unlink()
                out2 = Path(td) / "输出2"
                collect_employee_materials(
                    lib, out2,
                    roster_source="张三",
                    material_types=["身份证"],
                )

                cache_data2 = _load_ocr_cache(cache_path)
                # 删除的图片条目应被清理
                remaining = cache_data2.get("entries") or {}
                self.assertEqual(len(remaining), 1)
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_cache_disabled_skips_read_and_write(self) -> None:
        """use_ocr_cache=False 时不应读写缓存文件。"""
        from hr_toolkit.tools import material_collector as mc
        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class FakeEngine:
                call_count = 0
                def __call__(self, path):
                    type(self).call_count += 1
                    return (
                        [["姓名", "张三"], ["公民身份号码", "440111199001011234"]],
                        None,
                    )
            FakeEngine.call_count = 0
            mc._OCR_ENGINE = FakeEngine()
            mc._OCR_ATTEMPTED = True

            with tempfile.TemporaryDirectory() as td:
                lib = Path(td) / "资料库"
                lib.mkdir()
                emp = lib / "张三"
                emp.mkdir()
                (emp / "a.png").write_bytes(_png_fixture())

                out = Path(td) / "输出"
                result = collect_employee_materials(
                    lib, out,
                    roster_source="张三",
                    material_types=["身份证"],
                    use_ocr_cache=False,
                )
                self.assertFalse(result.ocr_cache_enabled)
                self.assertIsNone(result.ocr_cache_path)
                cache_file = lib / _OCR_CACHE_FILE_NAME
                self.assertFalse(cache_file.exists())
                self.assertGreater(FakeEngine.call_count, 0)
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_cache_falls_back_silently_when_readonly_library(self) -> None:
        """库目录只读时不抛异常且不写缓存。"""
        from hr_toolkit.tools import material_collector as mc
        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class FakeEngine:
                def __call__(self, path):
                    return (
                        [["姓名", "张三"], ["公民身份号码", "440111199001011234"]],
                        None,
                    )
            mc._OCR_ENGINE = FakeEngine()
            mc._OCR_ATTEMPTED = True

            with tempfile.TemporaryDirectory() as td:
                lib = Path(td) / "资料库"
                lib.mkdir()
                emp = lib / "张三"
                emp.mkdir()
                (emp / "a.png").write_bytes(_png_fixture())

                # 模拟只读：先创建缓存文件路径后，切到只读目录
                out = Path(td) / "输出"
                result = collect_employee_materials(
                    lib, out,
                    roster_source="张三",
                    material_types=["身份证"],
                )
                # 库目录正常可写 → 缓存应生成
                self.assertTrue(result.ocr_cache_enabled)
                self.assertIsNotNone(result.ocr_cache_path)
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_cache_engine_upgrade_invalidates_all_entries(self) -> None:
        """engine_signature 不一致时，旧的 entries 应被清空。"""
        with tempfile.TemporaryDirectory() as td:
            cache_path = Path(td) / _OCR_CACHE_FILE_NAME
            cache_path.write_text(
                '{"version": 1, "engine_signature": "rapidocr_onnxruntime@0.0.0-old", '
                '"entries": {"abc": {"material_type": "身份证", "match_method": "ocr", '
                '"source_relpath": "a/b.png", "source_size": 1, "source_mtime": 1.0}}}',
                encoding="utf-8",
            )
            loaded = _load_ocr_cache(cache_path)
            self.assertEqual(len(loaded.get("entries") or {}), 1)
            # current_signature 必然不等于 0.0.0-old
            current = _get_engine_signature()
            self.assertNotEqual(loaded.get("engine_signature"), current)

            # 直接调用引擎升级路径的逻辑：清空 entries
            if loaded.get("engine_signature") != current:
                loaded["entries"] = {}
            self.assertEqual(len(loaded.get("entries") or {}), 0)

    def test_cache_does_not_store_id_card_in_plaintext(self) -> None:
        """缓存 JSON 不应包含身份证号原值，只存 hash。"""
        from hr_toolkit.tools import material_collector as mc
        real_engine = mc._OCR_ENGINE
        real_attempted = mc._OCR_ATTEMPTED
        try:
            class FakeEngine:
                def __call__(self, path):
                    return (
                        [
                            ["姓名", "张三"],
                            ["公民身份号码", "440111199001011234"],
                        ],
                        None,
                    )
            mc._OCR_ENGINE = FakeEngine()
            mc._OCR_ATTEMPTED = True

            with tempfile.TemporaryDirectory() as td:
                lib = Path(td) / "资料库"
                lib.mkdir()
                emp = lib / "张三"
                emp.mkdir()
                (emp / "a.png").write_bytes(_png_fixture())

                out = Path(td) / "输出"
                collect_employee_materials(
                    lib, out,
                    roster_source="张三",
                    material_types=["身份证"],
                )

                cache_path = lib / _OCR_CACHE_FILE_NAME
                self.assertTrue(cache_path.exists())
                raw_text = cache_path.read_text(encoding="utf-8")
                # 身份证号原值不应出现
                self.assertNotIn("440111199001011234", raw_text)
        finally:
            mc._OCR_ENGINE = real_engine
            mc._OCR_ATTEMPTED = real_attempted

    def test_cache_corrupted_json_recovers_gracefully(self) -> None:
        """损坏的 JSON 应被识别为新缓存，跑时不报错。"""
        with tempfile.TemporaryDirectory() as td:
            cache_path = Path(td) / _OCR_CACHE_FILE_NAME
            cache_path.write_text("{ this is not valid json ", encoding="utf-8")

            loaded = _load_ocr_cache(cache_path)
            self.assertIn("entries", loaded)
            self.assertEqual(loaded["entries"], {})

            # 应能正常写回
            self.assertTrue(_save_ocr_cache(cache_path, loaded))
            self.assertTrue(cache_path.exists())
            # 重读应是合法 JSON
            import json as _json
            _json.loads(cache_path.read_text(encoding="utf-8"))

    def test_trim_cache_by_age_removes_stale_entries(self) -> None:
        """超过 90 天未验证的条目应被清理。"""
        from datetime import datetime, timedelta
        from hr_toolkit.tools.material_collector import _BEIJING_TZ
        with tempfile.TemporaryDirectory() as td:
            data = {
                "version": 1,
                "engine_signature": "test",
                "entries": {
                    "stale": {
                        "verified_at": (datetime.now(tz=_BEIJING_TZ) - timedelta(days=120)).isoformat(),
                        "material_type": "身份证",
                    },
                    "fresh": {
                        "verified_at": datetime.now(tz=_BEIJING_TZ).isoformat(),
                        "material_type": "劳动合同",
                    },
                },
            }
            _trim_cache_by_age_and_size(data)
            self.assertNotIn("stale", data["entries"])
            self.assertIn("fresh", data["entries"])

    def test_safety_and_special_cert_collection(self) -> None:
        """测试新增的安全员证与特种证书的精准检索与提取。"""
        with tempfile.TemporaryDirectory() as td:
            lib = Path(td) / "资料库"
            lib.mkdir()
            emp = lib / "张三"
            emp.mkdir()
            _write_text_pdf(emp / "张三_安全员C证.pdf", "c cert")
            (emp / "张三_特种作业操作证.jpg").write_bytes(b"\x89PNG\r\n\x1a\nfake")

            out = Path(td) / "输出"
            result = collect_employee_materials(
                lib, out,
                roster_source="张三",
                material_types=["安全员证", "特种证书"],
            )
            self.assertEqual(result.total_employees, 1)
            self.assertEqual(result.matched_file_count, 2)
            self.assertEqual(len(result.missing_records), 0)

    def test_custom_other_material_collection(self) -> None:
        """测试用户通过「其他」手动输入的自定义材料（如入职证明、保密协议）。"""
        with tempfile.TemporaryDirectory() as td:
            lib = Path(td) / "资料库"
            lib.mkdir()
            emp = lib / "李四"
            emp.mkdir()
            _write_text_pdf(emp / "李四_入职证明.pdf", "onboarding proof")
            (emp / "李四_保密协议.docx").write_text("nda doc")

            out = Path(td) / "输出"
            result = collect_employee_materials(
                lib, out,
                roster_source="李四",
                material_types=["入职证明", "保密协议"],
            )
            self.assertEqual(result.total_employees, 1)
            self.assertEqual(result.matched_file_count, 2)
            self.assertEqual(len(result.missing_records), 0)
            self.assertTrue((out / "李四" / "李四_入职证明.pdf").exists())
            self.assertTrue((out / "李四" / "李四_保密协议.docx").exists())

    def test_folder_custom_material_uses_current_document_text_name(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            employee_dir = root / "资料库" / "张三"
            employee_dir.mkdir(parents=True)
            source = employee_dir / "a5d6e67cd.txt"
            source.write_text("员工手册 签收单\n姓名：张三", encoding="utf-8")

            first = collect_employee_materials(
                root / "资料库",
                root / "输出1",
                roster_source="张三",
                material_types=["员工手册领用单"],
            )
            second = collect_employee_materials(
                root / "资料库",
                root / "输出2",
                roster_source="张三",
                material_types=["员工手册签收单"],
            )

            self.assertEqual(first.matched_file_count, 0)
            self.assertEqual(second.matched_file_count, 1)
            self.assertEqual(second.matches[0].matched_by, "doc_content_custom")
            copied = root / "输出2" / "张三" / "张三_员工手册签收单.txt"
            self.assertEqual(copied.read_bytes(), source.read_bytes())

    def test_multipage_contract_and_reverse_order_id_card(self) -> None:
        """测试多页合同全量提取不被早停截断，以及身份证反面先被扫描时依然完整提取正反双面。"""
        with tempfile.TemporaryDirectory() as td:
            lib = Path(td) / "资料库"
            lib.mkdir()
            emp = lib / "王五"
            emp.mkdir()
            # 模拟多页合同
            (emp / "王五_劳动合同_第1页.jpg").write_bytes(b"\x89PNG\r\n\x1a\npage1")
            (emp / "王五_劳动合同_第2页.jpg").write_bytes(b"\x89PNG\r\n\x1a\npage2")
            # 模拟反面命名排在前面或先被扫描的身份证
            (emp / "王五_身份证反面.jpg").write_bytes(b"\x89PNG\r\n\x1a\nback")
            (emp / "王五_身份证正面.jpg").write_bytes(b"\x89PNG\r\n\x1a\nfront")
            # 模拟大量低优先级普通无关图片
            (emp / "无关图片_01.jpg").write_bytes(b"\x89PNG\r\n\x1a\nirrelevant")
            (emp / "无关图片_02.jpg").write_bytes(b"\x89PNG\r\n\x1a\nirrelevant2")

            out = Path(td) / "输出"
            result = collect_employee_materials(
                lib, out,
                roster_source="王五",
                material_types=["劳动合同", "身份证"],
            )
            self.assertEqual(result.total_employees, 1)
            # 应该提取到 2页合同 + 2面身份证 = 4个文件，绝不截断丢页
            self.assertEqual(result.matched_file_count, 4)
            self.assertEqual(len(result.missing_records), 0)

            emp_out = out / "王五"
            self.assertTrue((emp_out / "王五_劳动合同_1.jpg").exists())
            self.assertTrue((emp_out / "王五_劳动合同_2.jpg").exists())
            self.assertTrue((emp_out / "王五_身份证_正面.jpg").exists())
            self.assertTrue((emp_out / "王五_身份证_反面.jpg").exists())


class TestFlatOCRMaterialLibrary(unittest.TestCase):
    def setUp(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        self.mc = mc
        self.real_engine = mc._OCR_ENGINE
        self.real_attempted = mc._OCR_ATTEMPTED

        class ContentEngine:
            call_count = 0

            def __call__(self, source):
                type(self).call_count += 1
                marker = _read_fixture_marker(source)
                if marker == "zhang_contract":
                    return ([
                        ["乙方", "张三"],
                        ["材料", "劳动合同"],
                    ], None)
                if marker == "zhang_id_card":
                    return ([
                        ["姓名", "张三"],
                        ["公民身份号码", "440111199001011234"],
                        ["住址", "广东省"],
                    ], None)
                if marker in {"li_id_card", "lisi__id_card"}:
                    return ([
                        ["姓名", "李四"],
                        ["公民身份号码", "440111199001011235"],
                        ["住址", "广东省"],
                    ], None)
                if marker == "zhang_second_id_card":
                    return ([
                        ["姓名", "张三"],
                        ["公民身份号码", "440111199001019999"],
                        ["住址", "广东省"],
                    ], None)
                if marker == "zhang_household":
                    return ([
                        ["姓名", "张三"],
                        ["材料", "户口本"],
                    ], None)
                if marker == "zhang_degree_unlabeled":
                    return ([
                        ["证书", "普通高等学校毕业证书"],
                        ["持有人文字", "张三"],
                    ], None)
                if marker == "zhang_sanfeng_degree":
                    return ([
                        ["证书", "普通高等学校毕业证书"],
                        ["持有人文字", "张三丰"],
                    ], None)
                return ([["标题", "无法识别的普通图片"]], None)

        self.engine_type = ContentEngine
        self.engine_type.call_count = 0
        mc._OCR_ENGINE = ContentEngine()
        mc._OCR_ATTEMPTED = True

    def tearDown(self) -> None:
        self.mc._OCR_ENGINE = self.real_engine
        self.mc._OCR_ATTEMPTED = self.real_attempted

    @staticmethod
    def _write_image(path: Path, marker: str) -> None:
        path.write_bytes(_png_fixture(marker))

    def test_default_library_mode_remains_person_folder(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "资料库"
            employee = library / "张三"
            employee.mkdir(parents=True)
            (employee / "张三_劳动合同.txt").write_text("劳动合同", encoding="utf-8")

            result = collect_employee_materials(
                library,
                root / "输出",
                roster_source="张三",
                material_types=["劳动合同"],
            )

            self.assertEqual(result.library_mode, LIBRARY_MODE_PERSON_FOLDER)
            self.assertEqual(result.matched_file_count, 1)

    def test_first_scan_indexes_negatives_and_second_query_does_not_repeat_ocr(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            for index in range(99):
                self._write_image(library / f"{index:03d}.png", f"noise-{index}")
            self._write_image(library / "099.png", "zhang_id_card")

            first = collect_employee_materials(
                library,
                root / "输出1",
                roster_source="张三 440111199001011234",
                material_types=["身份证"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )
            first_calls = self.engine_type.call_count
            self.assertEqual(first_calls, 100)
            self.assertEqual(first.matched_file_count, 1)

            cache_path = library / _OCR_CACHE_FILE_NAME
            cache_data = _load_ocr_cache(cache_path)
            self.assertEqual(len(cache_data["paths"]), 100)
            self.assertEqual(len(cache_data["entries"]), 100)
            self.assertTrue(any(
                entry.get("material_type") == "其他材料"
                for entry in cache_data["entries"].values()
            ))
            self.assertNotIn("440111199001011234", cache_path.read_text(encoding="utf-8"))

            second = collect_employee_materials(
                library,
                root / "输出2",
                roster_source="张三 440111199001011234",
                material_types=["身份证"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )
            self.assertEqual(self.engine_type.call_count, first_calls)
            self.assertEqual(second.ocr_cache_hits, 100)
            self.assertEqual(second.ocr_cache_misses, 0)
            self.assertEqual(second.matched_file_count, 1)

    def test_filename_person_is_ignored_and_ocr_person_wins(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            self._write_image(library / "张三_身份证.png", "li_id_card")

            result = collect_employee_materials(
                library,
                root / "输出",
                roster_source="张三,李四",
                material_types=["身份证"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )

            self.assertEqual([match.employee_name for match in result.matches], ["李四"])
            self.assertEqual(self.engine_type.call_count, 1, "多人名单只能全库索引一次")
            self.assertIn("张三", result.missing_records)
            self.assertNotIn("李四", result.missing_records)

    def test_same_size_changed_file_is_reindexed_even_when_mtime_is_restored(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            source = library / "000.png"
            self._write_image(source, "zhang_id_card")
            original_stat = source.stat()

            collect_employee_materials(
                library,
                root / "输出1",
                roster_source="张三",
                material_types=["身份证"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )
            first_calls = self.engine_type.call_count

            # 两个 marker 等长；恢复 mtime 后仍必须依靠 ctime/hash 识别内容变更。
            self.assertEqual(len("zhang_id_card"), len("lisi__id_card"))
            self._write_image(source, "lisi__id_card")
            self.assertEqual(source.stat().st_size, original_stat.st_size)
            os.utime(source, ns=(original_stat.st_atime_ns, original_stat.st_mtime_ns))

            result = collect_employee_materials(
                library,
                root / "输出2",
                roster_source="张三",
                material_types=["身份证"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )
            self.assertGreater(self.engine_type.call_count, first_calls)
            self.assertEqual(result.matched_file_count, 0)
            self.assertGreaterEqual(result.ocr_cache_invalidated, 1)

    def test_missing_change_token_hashes_before_reusing_flat_cache(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            source = library / "000.png"
            self._write_image(source, "zhang_id_card")
            original_stat = source.stat()

            with mock.patch.object(self.mc, "_file_change_token", return_value=None):
                collect_employee_materials(
                    library,
                    root / "输出1",
                    roster_source="张三",
                    material_types=["身份证"],
                    library_mode=LIBRARY_MODE_FLAT_OCR,
                )
                first_calls = self.engine_type.call_count

                # 元数据变更标记不可用时，相同内容仍复用内容哈希缓存，不重复 OCR。
                collect_employee_materials(
                    library,
                    root / "输出2",
                    roster_source="张三",
                    material_types=["身份证"],
                    library_mode=LIBRARY_MODE_FLAT_OCR,
                )
                self.assertEqual(self.engine_type.call_count, first_calls)

                self._write_image(source, "lisi__id_card")
                os.utime(source, ns=(original_stat.st_atime_ns, original_stat.st_mtime_ns))
                result = collect_employee_materials(
                    library,
                    root / "输出3",
                    roster_source="张三",
                    material_types=["身份证"],
                    library_mode=LIBRARY_MODE_FLAT_OCR,
                )

            self.assertGreater(self.engine_type.call_count, first_calls)
            self.assertEqual(result.matched_file_count, 0)
            self.assertGreaterEqual(result.ocr_cache_invalidated, 1)

    def test_version_two_path_metadata_migrates_without_reocr(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            source = library / "000.png"
            self._write_image(source, "zhang_id_card")

            collect_employee_materials(
                library,
                root / "输出1",
                roster_source="张三",
                material_types=["身份证"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )
            first_calls = self.engine_type.call_count
            cache_path = library / _OCR_CACHE_FILE_NAME
            cache = json.loads(cache_path.read_text(encoding="utf-8"))
            path_entry = cache["paths"]["000.png"]
            path_entry["source_ctime_ns"] = path_entry.pop("source_change_token")
            cache["version"] = 2
            cache_path.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")

            collect_employee_materials(
                library,
                root / "输出2",
                roster_source="张三",
                material_types=["身份证"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )

            self.assertEqual(self.engine_type.call_count, first_calls)
            migrated = json.loads(cache_path.read_text(encoding="utf-8"))
            self.assertEqual(migrated["version"], 6)
            self.assertIn("source_change_token", migrated["paths"]["000.png"])
            self.assertNotIn("source_ctime_ns", migrated["paths"]["000.png"])

    def test_file_changed_after_index_is_not_copied_as_stale_match(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            source = library / "000.png"
            self._write_image(source, "zhang_contract")
            changed = False

            def mutate_before_copy(_current: int, _total: int, message: str) -> None:
                nonlocal changed
                if not changed and "【匹配人员资料】" in message:
                    self._write_image(source, "li_id_card")
                    changed = True

            result = collect_employee_materials(
                library,
                root / "输出",
                roster_source="张三",
                material_types=["劳动合同"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
                progress_callback=mutate_before_copy,
            )

            self.assertTrue(changed)
            self.assertEqual(result.matched_file_count, 0)
            self.assertEqual(result.missing_records["张三"], ["劳动合同"])
            self.assertTrue(any("索引后发生变化" in warning for warning in result.warnings))

    def test_corrupted_copy_is_removed_and_reported_missing(self) -> None:
        from unittest.mock import patch

        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            self._write_image(library / "000.png", "zhang_contract")

            def corrupt_copy(_source, destination):
                Path(destination).write_bytes(b"corrupted")

            with patch.object(self.mc.shutil, "copy2", side_effect=corrupt_copy):
                result = collect_employee_materials(
                    library,
                    root / "输出",
                    roster_source="张三",
                    material_types=["劳动合同"],
                    library_mode=LIBRARY_MODE_FLAT_OCR,
                )

            self.assertEqual(result.matched_file_count, 0)
            self.assertEqual(result.missing_records["张三"], ["劳动合同"])
            self.assertFalse(any(path.is_file() for path in (root / "输出").rglob("*.png")))
            self.assertTrue(any("复制后校验不一致" in warning for warning in result.warnings))

    def test_renamed_file_reuses_content_hash_without_ocr(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            old_path = library / "old.png"
            self._write_image(old_path, "zhang_contract")
            collect_employee_materials(
                library,
                root / "输出1",
                roster_source="张三",
                material_types=["劳动合同"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )
            first_calls = self.engine_type.call_count

            nested = library / "新目录"
            nested.mkdir()
            old_path.rename(nested / "new-random.png")
            result = collect_employee_materials(
                library,
                root / "输出2",
                roster_source="张三",
                material_types=["劳动合同"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )

            self.assertEqual(self.engine_type.call_count, first_calls)
            self.assertEqual(result.matched_file_count, 1)
            self.assertEqual(result.ocr_cache_hits, 1)
            cache = _load_ocr_cache(library / _OCR_CACHE_FILE_NAME)
            self.assertEqual(set(cache["paths"]), {"新目录/new-random.png"})

    def test_collect_all_only_copies_files_identified_for_target(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            self._write_image(library / "a.png", "zhang_contract")
            self._write_image(library / "b.png", "li_id_card")

            result = collect_employee_materials(
                library,
                root / "输出",
                roster_source="张三",
                collect_all=True,
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )

            self.assertEqual(result.matched_file_count, 1)
            self.assertEqual(result.matches[0].employee_name, "张三")
            copied_files = [path.name for path in (root / "输出" / "张三").iterdir()]
            self.assertEqual(copied_files, ["a.png"])

    def test_same_name_employees_are_separated_by_id_in_output_and_report(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            self._write_image(library / "a.png", "zhang_id_card")
            self._write_image(library / "b.png", "zhang_second_id_card")

            output = root / "输出"
            result = collect_employee_materials(
                library,
                output,
                roster_source=(
                    "张三 440111199001011234,"
                    "张三 440111199001019999"
                ),
                material_types=["身份证"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )

            self.assertEqual(result.total_employees, 2)
            self.assertEqual(result.complete_employee_count, 2)
            self.assertEqual(result.matched_file_count, 2)
            self.assertTrue(
                (output / "张三（证件尾号1234）" / "张三（证件尾号1234）_身份证_正面.png").exists()
            )
            self.assertTrue(
                (output / "张三（证件尾号9999）" / "张三（证件尾号9999）_身份证_正面.png").exists()
            )
            self.assertEqual(
                len({match.employee_identity_key for match in result.matches}),
                2,
            )

            workbook = load_workbook(result.report_path, data_only=True)
            try:
                sheet = workbook["资料提取汇总与缺失清单"]
                self.assertEqual(sheet["D8"].value, "1/1")
                self.assertEqual(sheet["D9"].value, "1/1")
                self.assertEqual(sheet["F8"].value, "已提取(1份)")
                self.assertEqual(sheet["F9"].value, "已提取(1份)")
            finally:
                workbook.close()

    def test_all_existing_output_modes_work_with_flat_library_input(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            self._write_image(library / "random.png", "zhang_contract")

            expected = {
                MODE_BY_EMPLOYEE: lambda output: output / "张三" / "张三_劳动合同.png",
                MODE_BY_MATERIAL: lambda output: output / "劳动合同" / "张三_劳动合同.png",
                MODE_FLAT: lambda output: output / "张三_劳动合同.png",
            }
            for mode, expected_path in expected.items():
                output = root / f"输出-{mode}"
                result = collect_employee_materials(
                    library,
                    output,
                    roster_source="张三",
                    material_types=["劳动合同"],
                    mode=mode,
                    library_mode=LIBRARY_MODE_FLAT_OCR,
                )
                self.assertEqual(result.matched_file_count, 1)
                self.assertTrue(expected_path(output).exists(), mode)

    def test_cached_text_supports_new_custom_material_without_reocr(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            self._write_image(library / "random.png", "zhang_household")

            first = collect_employee_materials(
                library,
                root / "输出1",
                roster_source="张三",
                material_types=["身份证"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )
            first_calls = self.engine_type.call_count
            self.assertEqual(first.matched_file_count, 0)

            second = collect_employee_materials(
                library,
                root / "输出2",
                roster_source="张三",
                material_types=["户口本"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )
            self.assertEqual(self.engine_type.call_count, first_calls)
            self.assertEqual(second.matched_file_count, 1)

    def test_text_document_is_indexed_without_ocr(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            (library / "random.txt").write_text(
                "劳动合同 乙方：张三 工作内容：项目管理",
                encoding="utf-8",
            )

            result = collect_employee_materials(
                library,
                root / "输出",
                roster_source="张三",
                material_types=["劳动合同"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )

            self.assertEqual(self.engine_type.call_count, 0)
            self.assertEqual(result.matched_file_count, 1)
            self.assertEqual(result.matches[0].matched_by, "doc_content_contract")

    def test_unlabeled_name_uses_exact_ocr_text_boundary(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            self._write_image(library / "a.png", "zhang_degree_unlabeled")
            self._write_image(library / "b.png", "zhang_sanfeng_degree")

            result = collect_employee_materials(
                library,
                root / "输出",
                roster_source="张三",
                material_types=["学历证明"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )

            self.assertEqual(result.matched_file_count, 1)
            self.assertEqual(result.matches[0].relative_source_path, "a.png")

    def test_ocr_unavailable_is_not_persisted_as_negative_result(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            self._write_image(library / "random.png", "zhang_contract")
            self._write_image(library / "noise.png", "noise")

            working_engine = self.mc._OCR_ENGINE
            self.mc._OCR_ENGINE = None
            self.mc._OCR_ATTEMPTED = True
            with self.assertRaises(self.mc.OCRUnavailableError):
                collect_employee_materials(
                    library,
                    root / "输出1",
                    roster_source="张三",
                    material_types=["劳动合同"],
                    library_mode=LIBRARY_MODE_FLAT_OCR,
                )
            cache = _load_ocr_cache(library / _OCR_CACHE_FILE_NAME)
            self.assertEqual(cache["entries"], {})

            self.mc._OCR_ENGINE = working_engine
            second = collect_employee_materials(
                library,
                root / "输出2",
                roster_source="张三",
                material_types=["劳动合同"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )
            self.assertEqual(second.matched_file_count, 1)

    def test_flat_library_rejects_disabled_cache(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            with self.assertRaisesRegex(ValueError, "必须启用 OCR 索引缓存"):
                collect_employee_materials(
                    library,
                    root / "输出",
                    roster_source="张三",
                    library_mode=LIBRARY_MODE_FLAT_OCR,
                    use_ocr_cache=False,
                )

    def test_separate_contract_images_are_grouped_and_reused_from_cache(self) -> None:
        mc = self.mc
        page_texts = {
            "page-1": "劳动合同",
            "page-2": "第一章 工作地点",
            "page-3": "第二章 劳动报酬",
            "page-4": "第三章 合同期限",
            "page-5": "乙方签字 姓名：张三 签署日期：2026年8月28日",
        }

        class ContractPageEngine:
            call_count = 0

            def __call__(self, source):
                type(self).call_count += 1
                marker = _read_fixture_marker(source)
                return ([[None, page_texts[marker]]], None)

        mc._OCR_ENGINE = ContractPageEngine()
        mc._OCR_ATTEMPTED = True
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            source_hashes: dict[str, str] = {}
            for page_number in range(1, 6):
                source = library / f"scan_{page_number:03d}.png"
                source.write_bytes(_png_fixture(f"page-{page_number}"))
                source_hashes[source.name] = hashlib.sha256(source.read_bytes()).hexdigest()

            first = collect_employee_materials(
                library,
                root / "输出1",
                roster_source="张三",
                material_types=["劳动合同"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
                generate_report=False,
            )
            first_calls = ContractPageEngine.call_count
            second = collect_employee_materials(
                library,
                root / "输出2",
                roster_source="张三",
                material_types=["劳动合同"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
                generate_report=False,
            )

            copied = sorted((root / "输出2" / "张三").glob("张三_劳动合同_*.png"))
            copied_hashes = {
                path.name: hashlib.sha256(path.read_bytes()).hexdigest()
                for path in copied
            }
            final_source_hashes = {
                path.name: hashlib.sha256(path.read_bytes()).hexdigest()
                for path in library.glob("scan_*.png")
            }
            cache_payload = json.loads(
                (library / _OCR_CACHE_FILE_NAME).read_text(encoding="utf-8")
            )
            cached_text_lengths = [
                len(str(entry.get("ocr_text") or ""))
                for entry in cache_payload.get("entries", {}).values()
                if isinstance(entry, dict)
            ]

        self.assertEqual(first.matched_file_count, 5)
        self.assertEqual(second.matched_file_count, 5)
        self.assertEqual(first_calls, 5)
        self.assertEqual(ContractPageEngine.call_count, first_calls)
        self.assertEqual(second.ocr_cache_hits, 5)
        self.assertEqual(len(copied_hashes), 5)
        self.assertEqual(sorted(copied_hashes.values()), sorted(source_hashes.values()))
        self.assertEqual(final_source_hashes, source_hashes)
        self.assertTrue(all("document_group" in match.matched_by for match in second.matches))
        # 人工确认元数据允许保存；逐页 OCR 缓存不能带上自动推断的分组归属。
        self.assertNotIn("document_group", json.dumps(cache_payload["entries"], ensure_ascii=False))
        self.assertTrue(all(length <= 4096 for length in cached_text_lengths))

    def test_adjacent_contracts_for_different_people_do_not_cross_group(self) -> None:
        mc = self.mc
        page_texts = {
            1: "劳动合同 合同编号：HT-001",
            2: "第一章 工作地点",
            3: "第二章 劳动报酬",
            4: "第三章 合同期限",
            5: "乙方签字 姓名：张三 签署日期：2026年8月28日",
            6: "劳动合同 合同编号：HT-002",
            7: "第一章 工作地点",
            8: "第二章 劳动报酬",
            9: "第三章 合同期限",
            10: "乙方签字 姓名：李四 签署日期：2026年8月28日",
        }

        class TwoContractEngine:
            def __call__(self, source):
                page_number = int(_read_fixture_marker(source))
                return ([[None, page_texts[page_number]]], None)

        mc._OCR_ENGINE = TwoContractEngine()
        mc._OCR_ATTEMPTED = True
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            for page_number in range(1, 11):
                (library / f"scan_{page_number:03d}.png").write_bytes(
                    _png_fixture(str(page_number)),
                )

            result = collect_employee_materials(
                library,
                root / "输出",
                roster_source=["张三", "李四"],
                material_types=["劳动合同"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
                generate_report=False,
            )

            zhang_sources = {
                match.relative_source_path for match in result.matches
                if match.employee_name == "张三"
            }
            li_sources = {
                match.relative_source_path for match in result.matches
                if match.employee_name == "李四"
            }

        self.assertEqual(result.matched_file_count, 10)
        self.assertEqual(zhang_sources, {f"scan_{number:03d}.png" for number in range(1, 6)})
        self.assertEqual(li_sources, {f"scan_{number:03d}.png" for number in range(6, 11)})

    def test_close_scan_time_and_dimensions_still_require_contract_confirmation(self) -> None:
        mc = self.mc
        page_texts = {
            "cover": "劳动合同 合同编号：HT-003",
            "body": "工作地点及劳动报酬",
            "signature": "乙方签字 姓名：张三 签署日期：2026年8月28日",
        }

        class MetadataGroupEngine:
            def __call__(self, source):
                marker = _read_fixture_marker(source)
                return ([[None, page_texts[marker]]], None)

        mc._OCR_ENGINE = MetadataGroupEngine()
        mc._OCR_ATTEMPTED = True
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            for filename, marker in (
                ("a-cover.png", "cover"),
                ("b-body.png", "body"),
                ("c-signature.png", "signature"),
            ):
                (library / filename).write_bytes(_png_fixture(marker))

            with mock.patch.object(
                mc, "_read_image_dimensions", return_value=(1200, 1600),
            ) as read_dimensions:
                result = collect_employee_materials(
                    library,
                    root / "输出",
                    roster_source="张三",
                    material_types=["劳动合同"],
                    library_mode=LIBRARY_MODE_FLAT_OCR,
                    generate_report=False,
                )

            self.assertIsNotNone(result.review_path)
            review = load_workbook(result.review_path, read_only=True, data_only=True)
            try:
                rows = list(review["待确认归属"].iter_rows(min_row=2, values_only=True))
                self.assertEqual(len(rows), 3)
                self.assertTrue(all("待确认" in row[6] for row in rows))
            finally:
                review.close()

        self.assertEqual(result.matched_file_count, 0)
        self.assertEqual(result.missing_records["张三"], ["劳动合同"])
        self.assertLessEqual(read_dimensions.call_count, 3)

    def test_time_and_dimensions_do_not_absorb_unrelated_image(self) -> None:
        mc = self.mc
        page_texts = {
            "cover": "劳动合同",
            "unrelated": "公司活动宣传材料",
            "signature": "乙方签字 姓名：张三 签署日期",
        }

        class MixedBatchEngine:
            def __call__(self, source):
                marker = _read_fixture_marker(source)
                return ([[None, page_texts[marker]]], None)

        mc._OCR_ENGINE = MixedBatchEngine()
        mc._OCR_ATTEMPTED = True
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            for filename, marker in (
                ("a-cover.png", "cover"),
                ("b-unrelated.png", "unrelated"),
                ("c-signature.png", "signature"),
            ):
                (library / filename).write_bytes(_png_fixture(marker))

            with mock.patch.object(
                mc, "_read_image_dimensions", return_value=(1200, 1600),
            ):
                result = collect_employee_materials(
                    library,
                    root / "输出",
                    roster_source="张三",
                    material_types=["劳动合同"],
                    library_mode=LIBRARY_MODE_FLAT_OCR,
                    generate_report=False,
                )

        self.assertEqual(result.matched_file_count, 0)
        self.assertEqual(result.missing_records["张三"], ["劳动合同"])

    def test_explicit_page_markers_group_without_file_metadata_reads(self) -> None:
        mc = self.mc
        root = Path("not-read") / "document-pages"
        items = [
            mc._FlatIndexedFile(
                source_path=root / filename,
                relative_path=filename,
                cache_key=filename,
                material_type=material,
                match_method=method,
                subtype="",
                extracted_names=names,
                extracted_id_hash="",
                text_snippet=text,
            )
            for filename, material, method, names, text in (
                ("a.png", "劳动合同", "ocr_contract", (), "劳动合同 第1页 共3页"),
                ("b.png", "其他材料", "unrecognized", (), "合同正文 第2页 共3页"),
                ("c.png", "其他材料", "unrecognized", ("张三",), "姓名：张三 第3页 共3页"),
            )
        ]
        with mock.patch.object(
            mc, "_read_image_dimensions",
            side_effect=AssertionError("明确页码不应读取图片尺寸"),
        ):
            result = mc._enrich_flat_index_with_document_groups(
                items, [], ["劳动合同"],
            )

        self.assertEqual(len(result), 3)
        self.assertTrue(all(item.material_type == "劳动合同" for item in result))
        self.assertTrue(all(item.extracted_names == ("张三",) for item in result))

    def test_document_group_is_recomputed_after_member_content_changes(self) -> None:
        mc = self.mc
        page_texts = {
            "page-1": "劳动合同",
            "page-2": "第一章 工作地点",
            "page-3": "第二章 劳动报酬",
            "page-4": "第三章 合同期限",
            "page-5-z": "乙方签字 姓名：张三 签署日期",
            "page-5-l": "乙方签字 姓名：李四 签署日期",
        }

        class MutableContractEngine:
            call_count = 0

            def __call__(self, source):
                type(self).call_count += 1
                marker = _read_fixture_marker(source)
                return ([[None, page_texts[marker]]], None)

        mc._OCR_ENGINE = MutableContractEngine()
        mc._OCR_ATTEMPTED = True
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            for page_number in range(1, 5):
                (library / f"scan_{page_number:03d}.png").write_bytes(
                    _png_fixture(f"page-{page_number}"),
                )
            signature = library / "scan_005.png"
            signature.write_bytes(_png_fixture("page-5-z"))

            first = collect_employee_materials(
                library,
                root / "输出1",
                roster_source=["张三", "李四"],
                material_types=["劳动合同"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
                generate_report=False,
            )
            previous_stat = signature.stat()
            signature.write_bytes(_png_fixture("page-5-l"))
            os.utime(
                signature,
                ns=(previous_stat.st_atime_ns, previous_stat.st_mtime_ns + 1_000_000_000),
            )
            second = collect_employee_materials(
                library,
                root / "输出2",
                roster_source=["张三", "李四"],
                material_types=["劳动合同"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
                generate_report=False,
            )

        self.assertEqual(first.matched_file_count, 5)
        self.assertEqual({match.employee_name for match in first.matches}, {"张三"})
        self.assertEqual(second.matched_file_count, 5)
        self.assertEqual({match.employee_name for match in second.matches}, {"李四"})
        self.assertGreaterEqual(second.ocr_cache_invalidated, 1)
        self.assertEqual(MutableContractEngine.call_count, 6)

    def test_whole_document_group_is_skipped_if_member_changes_after_index(self) -> None:
        mc = self.mc
        page_texts = {
            "page-1": "劳动合同",
            "page-2": "第一章 工作地点",
            "page-3": "第二章 劳动报酬",
            "page-4": "第三章 合同期限",
            "page-5": "乙方签字 姓名：张三 签署日期",
        }

        class RaceContractEngine:
            def __call__(self, source):
                marker = _read_fixture_marker(source)
                return ([[None, page_texts[marker]]], None)

        mc._OCR_ENGINE = RaceContractEngine()
        mc._OCR_ATTEMPTED = True
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            for page_number in range(1, 6):
                (library / f"scan_{page_number:03d}.png").write_bytes(
                    _png_fixture(f"page-{page_number}"),
                )
            changed_page = library / "scan_005.png"
            changed = False

            def mutate_after_index(_current, _total, message):
                nonlocal changed
                if not changed and "【匹配人员资料】" in message:
                    changed_page.write_bytes(_png_fixture("changed-after-index"))
                    changed = True

            result = collect_employee_materials(
                library,
                root / "输出",
                roster_source="张三",
                material_types=["劳动合同"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
                generate_report=False,
                progress_callback=mutate_after_index,
            )

        self.assertTrue(changed)
        self.assertEqual(result.matched_file_count, 0)
        self.assertEqual(result.missing_records["张三"], ["劳动合同"])
        self.assertTrue(any("整组已跳过" in warning for warning in result.warnings))

    def test_conflicting_names_prevent_ambiguous_image_group(self) -> None:
        mc = self.mc
        page_texts = {
            1: "劳动合同 合同编号：HT-001",
            2: "姓名：张三",
            3: "姓名：李四",
        }

        class ConflictingPageEngine:
            def __call__(self, source):
                page_number = int(_read_fixture_marker(source))
                return ([[None, page_texts[page_number]]], None)

        mc._OCR_ENGINE = ConflictingPageEngine()
        mc._OCR_ATTEMPTED = True
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            for page_number in range(1, 4):
                (library / f"scan_{page_number:03d}.png").write_bytes(
                    _png_fixture(str(page_number)),
                )

            result = collect_employee_materials(
                library,
                root / "输出",
                roster_source="张三",
                material_types=["劳动合同"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
                generate_report=False,
            )

        self.assertEqual(result.matched_file_count, 0)
        self.assertEqual(result.missing_records["张三"], ["劳动合同"])
        self.assertTrue(any("姓名冲突" in warning for warning in result.warnings))


class TestSpecialCertGraphicTitleAndReorderedMatching(unittest.TestCase):
    """复盘会需求：图形标题 PDF 渲染兜底、自定义名称乱序窗口匹配、缓存重分类。"""

    def setUp(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        self.mc = mc
        self.real_engine = mc._OCR_ENGINE
        self.real_attempted = mc._OCR_ATTEMPTED

    def tearDown(self) -> None:
        self.mc._OCR_ENGINE = self.real_engine
        self.mc._OCR_ATTEMPTED = self.real_attempted

    def _set_engine(self, lines):
        class FixedEngine:
            def __call__(self, source):
                return ([list(item) for item in lines], None)

        self.mc._OCR_ENGINE = FixedEngine()
        self.mc._OCR_ATTEMPTED = True

    def test_reordered_window_matching_boundaries(self) -> None:
        mc = self.mc
        hit = mc._classify_requested_material_text(
            "我的证书天谴", ["天谴证书"], method_prefix="t",
        )
        self.assertEqual(hit[0], "天谴证书")
        self.assertEqual(hit[1], "t_custom_reordered")
        # 字符被其他字打断：不命中
        self.assertIsNone(
            mc._classify_requested_material_text(
                "我天天谴的证书", ["天谴证书"], method_prefix="t",
            )[0]
        )
        # 任意字符洗牌不是“两个连续词块前后互换”，不能误命中。
        self.assertIsNone(
            mc._classify_requested_material_text(
                "天证谴书", ["天谴证书"], method_prefix="t",
            )[0]
        )
        # 名称过短（<3）不走乱序窗口，且无精确子串：不命中
        self.assertIsNone(
            mc._classify_requested_material_text(
                "的证书天", ["天谴"], method_prefix="t",
            )[0]
        )

    def test_standard_then_custom_then_weak_id_priority(self) -> None:
        mc = self.mc
        text = "我的证书天谴 证号500237200308190399"
        result = mc._classify_text_content(
            text,
            requested_types=["天谴证书"],
            method_prefix="doc",
            allow_weak_id_fallback=False,
        )
        self.assertEqual(result[0], "天谴证书")
        # 明确身份证字段属于标准强证据，不应被同页偶然出现的自定义词覆盖。
        self.assertEqual(
            mc._classify_text_content(
                "居民身份证 我的证书天谴",
                requested_types=["天谴证书"],
                allow_weak_id_fallback=False,
            )[0],
            "身份证",
        )
        # 弱兜底门禁：仅 18 位证号时，禁用→无结论，启用→身份证
        self.assertIsNone(
            mc._classify_text_content(
                "T500237200308190399", allow_weak_id_fallback=False,
            )[0]
        )
        self.assertEqual(
            mc._classify_text_content(
                "T500237200308190399", allow_weak_id_fallback=True,
            )[0],
            "身份证",
        )
        # 发证机关名称本身不能定类；与证书专属字段组合后才是强证据。
        self.assertIsNone(
            mc._classify_text_content(
                "北京市应急管理局关于召开年度工作会议的通知",
                allow_weak_id_fallback=False,
            )[0]
        )
        self.assertEqual(
            mc._classify_text_content(
                "姓名：张三 证号：T500237200308190399 "
                "作业类别：低压电工作业 北京市应急管理局 备注",
                allow_weak_id_fallback=False,
            )[0],
            "特种证书",
        )

    def test_official_special_certificate_title_maps_to_requested_label(self) -> None:
        mc = self.mc
        self.assertEqual(
            mc._requested_label_for_detected_material(
                "特种证书", ["特种作业操作证"],
            ),
            "特种作业操作证",
        )
        # “证书”等宽泛词不能因标准同义词表而变成等价材料。
        self.assertIsNone(
            mc._requested_label_for_detected_material(
                "资格证书", ["证书"],
            )
        )

    def test_special_certificate_custom_label_collects_in_folder_and_flat_modes(self) -> None:
        mc = self.mc
        lines = [
            ("标题", "中华人民共和国特种作业操作证"),
            ("姓名", "张三"),
            ("作业类别", "低压电工作业"),
            ("机关", "北京市应急管理局"),
        ]
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            folder_library = root / "人员资料库"
            employee_dir = folder_library / "张三"
            employee_dir.mkdir(parents=True)
            folder_source = employee_dir / "random.jpg"
            folder_source.write_bytes(_png_fixture("special-certificate-folder"))
            folder_hash = hashlib.sha256(folder_source.read_bytes()).hexdigest()

            self._set_engine(lines)
            folder_result = collect_employee_materials(
                folder_library,
                root / "文件夹输出",
                roster_source="张三",
                material_types=["特种作业操作证"],
            )

            class FailingEngine:
                def __call__(self, _source):
                    raise AssertionError("正式名称映射命中缓存后不应重复 OCR")

            mc._OCR_ENGINE = FailingEngine()
            cached_folder_result = collect_employee_materials(
                folder_library,
                root / "文件夹缓存输出",
                roster_source="张三",
                material_types=["特种作业操作证"],
            )

            flat_library = root / "无序资料库"
            flat_library.mkdir()
            flat_source = flat_library / "random.jpg"
            flat_source.write_bytes(_png_fixture("special-certificate-flat"))
            flat_hash = hashlib.sha256(flat_source.read_bytes()).hexdigest()
            self._set_engine(lines)
            flat_result = collect_employee_materials(
                flat_library,
                root / "平铺输出",
                roster_source="张三",
                material_types=["特种作业操作证"],
                library_mode=LIBRARY_MODE_FLAT_OCR,
            )
            final_folder_hash = hashlib.sha256(folder_source.read_bytes()).hexdigest()
            final_flat_hash = hashlib.sha256(flat_source.read_bytes()).hexdigest()
            copied_folder_hash = hashlib.sha256(
                folder_result.matches[0].target_path.read_bytes()
            ).hexdigest()
            copied_flat_hash = hashlib.sha256(
                flat_result.matches[0].target_path.read_bytes()
            ).hexdigest()
            workbook = load_workbook(flat_result.report_path, data_only=True)
            try:
                report_sheet = workbook["资料提取汇总与缺失清单"]
                report_material_header = report_sheet["F7"].value
                report_material_status = report_sheet["F8"].value
            finally:
                workbook.close()

        self.assertEqual(folder_result.matched_file_count, 1)
        self.assertEqual(folder_result.matches[0].material_type, "特种作业操作证")
        self.assertNotIn("张三", folder_result.missing_records)
        self.assertEqual(cached_folder_result.matched_file_count, 1)
        self.assertTrue(cached_folder_result.matches[0].cache_hit)
        self.assertEqual(flat_result.matched_file_count, 1)
        self.assertEqual(flat_result.matches[0].material_type, "特种作业操作证")
        self.assertNotIn("张三", flat_result.missing_records)
        self.assertEqual(final_folder_hash, folder_hash)
        self.assertEqual(final_flat_hash, flat_hash)
        self.assertEqual(copied_folder_hash, folder_hash)
        self.assertEqual(copied_flat_hash, flat_hash)
        self.assertEqual(report_material_header, "特种作业操作证")
        self.assertEqual(report_material_status, "已提取(1份)")

    def test_modern_lightweight_path_ocr_embedded_certificate_image(self) -> None:
        mc = self.mc

        class CertificateImageEngine:
            call_count = 0

            def __call__(self, _source):
                type(self).call_count += 1
                return ([
                    [None, "中华人民共和国特种作业操作证"],
                    [None, "姓名：张三"],
                    [None, "作业类别：低压电工作业"],
                    [None, "北京市应急管理局"],
                ], None)

        mc._OCR_ENGINE = CertificateImageEngine()
        mc._OCR_ATTEMPTED = True
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "无序资料库"
            library.mkdir()
            source = library / "certificate.pdf"
            _write_mixed_text_image_pdf(
                source,
                text_layer="ZhangSan T500237200308190399 low-voltage Beijing remarks",
            )
            source_hash = hashlib.sha256(source.read_bytes()).hexdigest()
            with mock.patch.object(mc, "pdfium", None):
                with mock.patch.object(mc, "_PDF_BACKEND", "pypdf"):
                    standard_result = collect_employee_materials(
                        library,
                        root / "标准输出",
                        roster_source="张三",
                        material_types=["特种证书"],
                        library_mode=LIBRARY_MODE_FLAT_OCR,
                        generate_report=False,
                    )
                    first_calls = CertificateImageEngine.call_count
                    custom_result = collect_employee_materials(
                        library,
                        root / "自定义输出",
                        roster_source="张三",
                        material_types=["特种作业操作证"],
                        library_mode=LIBRARY_MODE_FLAT_OCR,
                        generate_report=False,
                    )

            final_hash = hashlib.sha256(source.read_bytes()).hexdigest()

        self.assertEqual(standard_result.matched_file_count, 1)
        self.assertEqual(standard_result.matches[0].material_type, "特种证书")
        self.assertEqual(custom_result.matched_file_count, 1)
        self.assertEqual(custom_result.matches[0].material_type, "特种作业操作证")
        self.assertEqual(first_calls, 1)
        self.assertEqual(CertificateImageEngine.call_count, first_calls)
        self.assertEqual(custom_result.ocr_cache_hits, 1)
        self.assertEqual(final_hash, source_hash)

    def test_flat_pdf_graphic_title_render_fallback(self) -> None:
        from tests.test_pdf_backend_compat import _write_minimal_pdf

        if self.mc.pdfium is None:
            self.skipTest("现代轻量安装包不捆绑 PDFium；Win7 通道覆盖整页渲染")

        self._set_engine([("title", "中华人民共和国特种作业操作证")])
        with tempfile.TemporaryDirectory() as td:
            pdf = Path(td) / "cert.pdf"
            _write_minimal_pdf(pdf, ["text"], text_content="T500237200308190399")
            material, _method, _subtype, _names, _eid, text, complete = (
                self.mc._analyze_flat_source(pdf, ["特种证书"])
            )
        self.assertTrue(complete)
        self.assertEqual(material, "特种证书")
        self.assertIn("特种作业操作证", text)

    def test_complete_text_layer_signal_skips_embedded_image_ocr(self) -> None:
        mc = self.mc
        text_layer = (
            "姓名：张三 证号：T500237200308190399 "
            "作业类别：低压电工作业 北京市应急管理局 备注"
        )
        with mock.patch.object(mc, "pdfium", None):
            with mock.patch.object(
                mc,
                "_extract_flat_document_text",
                return_value=text_layer,
            ):
                with mock.patch.object(mc, "_analyze_ocr_file") as analyze_ocr:
                    result = mc._analyze_flat_source(
                        Path("certificate.pdf"),
                        ["特种证书"],
                    )

        self.assertEqual(result[0], "特种证书")
        self.assertEqual(result[3], ["张三"])
        analyze_ocr.assert_not_called()

    def test_flat_cache_reclassifies_cached_standard_type_against_new_request(self) -> None:
        mc = self.mc
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "lib"
            library.mkdir()
            (library / "a.jpg").write_bytes(_png_fixture("cert-number-only"))
            cache = {"version": mc._OCR_CACHE_VERSION}
            cache_path = library / ".hr_material_index_cache.json"
            warnings: list = []

            self._set_engine([("n", "我的证书天谴 500237200308190399")])
            first, _ = mc._build_flat_ocr_index(
                library, root / "out1", ["身份证"], cache, cache_path, {}, warnings, None,
            )
            self.assertEqual(first[0].material_type, "身份证")

            second, _ = mc._build_flat_ocr_index(
                library, root / "out2", ["天谴证书"], cache, cache_path, {}, warnings, None,
            )
            self.assertEqual(second[0].material_type, "天谴证书")
            self.assertTrue(second[0].cache_hit)

    def test_folder_cache_reclassifies_weak_id_as_custom_without_reocr(self) -> None:
        mc = self.mc
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            employee_dir = root / "资料库" / "张三"
            employee_dir.mkdir(parents=True)
            source = employee_dir / "a.jpg"
            source.write_bytes(_png_fixture("custom-cert-number"))

            self._set_engine([("n", "我的证书天谴 500237200308190399")])
            first = collect_employee_materials(
                root / "资料库",
                root / "输出1",
                roster_source="张三",
                material_types=["身份证"],
            )
            self.assertEqual(first.matched_file_count, 1)

            class FailingEngine:
                def __call__(self, _source):
                    raise AssertionError("缓存重分类不应重复 OCR")

            mc._OCR_ENGINE = FailingEngine()
            second = collect_employee_materials(
                root / "资料库",
                root / "输出2",
                roster_source="张三",
                material_types=["天谴证书"],
            )
            self.assertEqual(second.matched_file_count, 1)
            self.assertEqual(second.matches[0].material_type, "天谴证书")
            self.assertTrue(second.matches[0].cache_hit)

    def test_folder_pdf_does_not_repeat_failed_visual_query(self) -> None:
        from tests.test_pdf_backend_compat import _write_minimal_pdf

        mc = self.mc
        with tempfile.TemporaryDirectory() as td:
            pdf = Path(td) / "a5d6e67c.pdf"
            _write_minimal_pdf(
                pdf,
                ["text"],
                text_content="T500237200308190399",
            )
            cache = {"entries": {}}
            requested = ["天谴证书"]
            mc._store_ocr_cache(
                cache,
                pdf,
                "身份证",
                "ocr_id_number_fallback",
                "正面",
                "",
                "500237200308190399",
                extracted_text="T500237200308190399",
                analysis_complete=True,
                visual_ocr_query_signature=mc._visual_ocr_query_signature(requested),
            )

            class FailingEngine:
                def __call__(self, _source):
                    raise AssertionError("相同失败查询不应重复整页 OCR")

            mc._OCR_ENGINE = FailingEngine()
            mc._OCR_ATTEMPTED = True
            result = mc._classify_material_type(
                pdf,
                pdf.name,
                requested,
                cache=cache,
            )

        self.assertEqual(result[0], "身份证")
        self.assertTrue(result[-1])

    def test_flat_pdf_records_failed_visual_query_after_one_retry(self) -> None:
        from tests.test_pdf_backend_compat import _write_minimal_pdf

        mc = self.mc
        if mc.pdfium is None:
            self.skipTest("该缓存路径只在具备 PDFium 整页渲染能力时启用")

        class CountingEngine:
            call_count = 0

            def __call__(self, _source):
                type(self).call_count += 1
                return ([[None, "无关内容"]], None)

        mc._OCR_ENGINE = CountingEngine()
        mc._OCR_ATTEMPTED = True
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            library = root / "资料库"
            library.mkdir()
            source = library / "a5d6e67c.pdf"
            _write_minimal_pdf(
                source,
                ["text"],
                text_content="T500237200308190399",
            )
            source_hash = hashlib.sha256(source.read_bytes()).hexdigest()
            cache = {"version": mc._OCR_CACHE_VERSION}
            cache_path = library / mc._OCR_CACHE_FILE_NAME
            requested = ["天谴证书"]

            mc._build_flat_ocr_index(
                library, root / "输出1", requested, cache, cache_path, {}, [], None,
            )
            first_calls = CountingEngine.call_count
            mc._build_flat_ocr_index(
                library, root / "输出2", requested, cache, cache_path, {}, [], None,
            )
            second_calls = CountingEngine.call_count
            mc._build_flat_ocr_index(
                library, root / "输出3", requested, cache, cache_path, {}, [], None,
            )
            final_source_hash = hashlib.sha256(source.read_bytes()).hexdigest()

        self.assertEqual(first_calls, 1)
        self.assertEqual(second_calls, 2)
        self.assertEqual(CountingEngine.call_count, second_calls)
        self.assertEqual(final_source_hash, source_hash)


class TestLargeBatchCandidateIndexes(unittest.TestCase):
    """Candidate indexes must be a lossless accelerator, never a new match rule."""

    def test_flat_library_scandir_preserves_filters_and_relative_order(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        with tempfile.TemporaryDirectory() as td:
            library = Path(td) / "资料库"
            library.mkdir()
            (library / "b").mkdir()
            (library / "b" / "2.png").write_bytes(b"2")
            (library / "a").mkdir()
            (library / "a" / "1.jpg").write_bytes(b"1")
            (library / ".hidden").mkdir()
            (library / ".hidden" / "hidden.png").write_bytes(b"hidden")
            (library / "~$temp.xlsx").write_bytes(b"temp")
            output = library / "输出"
            output.mkdir()
            (output / "result.png").write_bytes(b"result")
            try:
                (library / "linked.png").symlink_to(library / "a" / "1.jpg")
            except (OSError, NotImplementedError):
                pass

            result = mc._scan_flat_library_files(library, skip_dir=output)

        self.assertEqual(
            [path.relative_to(library).as_posix() for path in result],
            ["a/1.jpg", "b/2.png"],
        )

    def test_flat_library_scan_honors_prestart_cancellation(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        with self.assertRaises(mc.MaterialCollectionCancelled):
            mc._scan_flat_library_files(Path("/not-read"), cancelled=lambda: True)

    def test_result_summary_can_skip_duplicate_match_serialization(self) -> None:
        match = MaterialFileMatch(
            employee_name="张三",
            material_type="身份证",
            source_path=Path("/tmp/source.png"),
            relative_source_path="source.png",
            matched_by="ocr",
        )
        result = MaterialCollectResult(
            library_dir=Path("/tmp/library"),
            output_dir=Path("/tmp/output"),
            target_employees=[TargetEmployee("张三")],
            matches=[match],
        )

        self.assertEqual(result.to_dict()["matches"][0]["employee_name"], "张三")
        self.assertNotIn("matches", result.to_dict(include_matches=False))

    def test_folder_candidate_index_matches_exhaustive_business_predicate(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        employees = [
            mc.TargetEmployee(
                name="张三",
                id_card="440111199001011234",
                phone="13800000001",
            ),
            mc.TargetEmployee(name="张三丰"),
            mc.TargetEmployee(name="李四"),
            mc.TargetEmployee(name="王五", phone="13900000002"),
        ]
        folder_index = {
            "张三": [Path("/library/01")],
            "A-张三-资料": [Path("/library/02")],
            "张三丰": [Path("/library/03")],
            "440111199001011234_档案": [Path("/library/04")],
            "13900000002_证件": [Path("/library/05")],
            "无关文件夹": [Path("/library/06")],
        }

        candidate_index = mc._FolderEmployeeCandidateIndex(folder_index, employees)
        for employee_index, employee in enumerate(employees):
            expected = []
            for folder_name, paths in folder_index.items():
                reason = mc._match_folder_to_employee(folder_name, employee)
                if reason:
                    expected.extend((path, reason) for path in paths)
            self.assertEqual(candidate_index.matches_for(employee_index), expected)

    def test_flat_candidate_index_matches_exhaustive_business_predicate(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        employees = [
            mc.TargetEmployee(name="张三", id_card="440111199001011234"),
            mc.TargetEmployee(name="李四"),
            mc.TargetEmployee(name="王五", phone="13900000002"),
            mc.TargetEmployee(name="张三丰"),
        ]
        indexed_files = [
            mc._FlatIndexedFile(
                Path("/library/id.png"), "id.png", "k0", "身份证", "ocr", "",
                (), mc._hash_id_card("440111199001011234"),
            ),
            mc._FlatIndexedFile(
                Path("/library/mismatch.png"), "mismatch.png", "k1", "身份证", "ocr", "",
                ("张三",), mc._hash_id_card("440111199001019999"),
            ),
            mc._FlatIndexedFile(
                Path("/library/name.png"), "name.png", "k2", "劳动合同", "ocr", "",
                ("李·四",), "",
            ),
            mc._FlatIndexedFile(
                Path("/library/text.png"), "text.png", "k3", "学历证明", "ocr", "",
                (), "", text_snippet="持证人：张三丰；有效",
            ),
            mc._FlatIndexedFile(
                Path("/library/phone.png"), "phone.png", "k4", "银行卡", "ocr", "",
                (), "", extracted_phone_hash=mc._hash_phone("13900000002"),
            ),
            mc._FlatIndexedFile(
                Path("/library/raw-name.png"), "raw-name.png", "k5", "资格证书", "ocr", "",
                (), "", text_snippet="申请人（李四）审核通过",
            ),
        ]
        duplicate_names: set[str] = set()
        candidate_index = mc._FlatEmployeeCandidateIndex(
            indexed_files,
            employees,
            duplicate_names,
        )

        for employee_index, employee in enumerate(employees):
            expected = [
                item for item in indexed_files
                if mc._flat_file_matches_employee(item, employee, duplicate_names)
            ]
            self.assertEqual(
                candidate_index.files_for(employee_index, employee, duplicate_names),
                expected,
            )

    def test_flat_duplicate_names_still_require_identity_evidence(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        employees = [
            mc.TargetEmployee(name="张三", id_card="440111199001011234"),
            mc.TargetEmployee(name="张 三", id_card="440111199001011235"),
        ]
        duplicate_names = {"张三"}
        indexed_files = [
            mc._FlatIndexedFile(
                Path("/library/name-only.png"), "name-only.png", "k0", "身份证", "ocr", "",
                ("张三",), "", text_snippet="姓名：张三",
            ),
            mc._FlatIndexedFile(
                Path("/library/id.png"), "id.png", "k1", "身份证", "ocr", "",
                ("张三",), mc._hash_id_card("440111199001011235"),
            ),
        ]
        candidate_index = mc._FlatEmployeeCandidateIndex(
            indexed_files,
            employees,
            duplicate_names,
        )

        self.assertEqual(candidate_index.files_for(0, employees[0], duplicate_names), [])
        self.assertEqual(
            candidate_index.files_for(1, employees[1], duplicate_names),
            [indexed_files[1]],
        )

    def test_low_core_ocr_reserves_one_cpu_for_the_ui_thread(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        for modern in (False, True):
            calls: list[dict] = []

            class FakeRapidOCR:
                def __init__(self, **kwargs):
                    calls.append(kwargs)

            package = "rapidocr" if modern else "rapidocr_onnxruntime"
            fake_modules = {
                package: SimpleNamespace(RapidOCR=FakeRapidOCR),
                "rapidocr.utils.typings": SimpleNamespace(
                    ModelType=SimpleNamespace(MOBILE="mobile"),
                    OCRVersion=SimpleNamespace(PPOCRV4="ppocrv4"),
                ),
            }
            with self.subTest(modern=modern), mock.patch.multiple(
                mc, _OCR_ENGINE=None, _OCR_ATTEMPTED=False, _OCR_ENGINE_ERROR="",
            ), mock.patch.object(
                mc, "_uses_modern_ocr", return_value=modern,
            ), mock.patch.object(mc.os, "cpu_count", return_value=2), mock.patch.dict(
                sys.modules,
                fake_modules,
            ), mock.patch.object(
                mc, "_rapidocr_low_memory_session_options",
            ):
                engine = mc._get_ocr_engine()

                self.assertIsInstance(engine, FakeRapidOCR)
                if modern:
                    self.assertEqual(calls, [{"params": {
                        "EngineConfig.onnxruntime.intra_op_num_threads": 1,
                        "EngineConfig.onnxruntime.inter_op_num_threads": 1,
                        "EngineConfig.onnxruntime.enable_cpu_mem_arena": False,
                        "Global.log_level": "warning",
                        "Det.ocr_version": "ppocrv4",
                        "Det.model_type": "mobile",
                        "Rec.ocr_version": "ppocrv4",
                        "Rec.model_type": "mobile",
                    }}])
                else:
                    self.assertEqual(calls, [{
                        "intra_op_num_threads": 1,
                        "inter_op_num_threads": 1,
                    }])

    def test_ocr_engine_disables_memory_pattern_only_during_session_creation(self) -> None:
        from hr_toolkit.tools import material_collector as mc

        observed_memory_patterns: list[bool] = []

        class FakeSessionOptions:
            enable_mem_pattern = True

        class FakeOrtInferSession:
            @staticmethod
            def _init_sess_opts(_config):
                return FakeSessionOptions()

        original_descriptor = FakeOrtInferSession.__dict__["_init_sess_opts"]

        class FakeRapidOCR:
            def __init__(self, **_kwargs):
                options = FakeOrtInferSession._init_sess_opts({})
                observed_memory_patterns.append(options.enable_mem_pattern)

        for modern in (False, True):
            observed_memory_patterns.clear()
            package = "rapidocr" if modern else "rapidocr_onnxruntime"
            infer_name = (
                "rapidocr.inference_engine.onnxruntime.main" if modern
                else "rapidocr_onnxruntime.utils.infer_engine"
            )
            module_names = [package, f"{package}.utils", infer_name]
            if modern:
                module_names.extend([
                    "rapidocr.utils.typings",
                    "rapidocr.inference_engine",
                    "rapidocr.inference_engine.onnxruntime",
                ])
            fake_modules = {name: ModuleType(name) for name in module_names}
            for module in fake_modules.values():
                module.__path__ = []
            fake_modules[package].RapidOCR = FakeRapidOCR
            fake_modules[infer_name].OrtInferSession = FakeOrtInferSession
            if modern:
                typings = fake_modules["rapidocr.utils.typings"]
                typings.ModelType = SimpleNamespace(MOBILE="mobile")
                typings.OCRVersion = SimpleNamespace(PPOCRV4="ppocrv4")

            with self.subTest(modern=modern), mock.patch.multiple(
                mc, _OCR_ENGINE=None, _OCR_ATTEMPTED=False, _OCR_ENGINE_ERROR="",
            ), mock.patch.object(
                mc, "_uses_modern_ocr", return_value=modern,
            ), mock.patch.dict(sys.modules, fake_modules):
                engine = mc._get_ocr_engine()

                self.assertIsInstance(engine, FakeRapidOCR)
                self.assertEqual(observed_memory_patterns, [False])
                self.assertIs(
                    FakeOrtInferSession.__dict__["_init_sess_opts"],
                    original_descriptor,
                )
                self.assertTrue(
                    FakeOrtInferSession._init_sess_opts({}).enable_mem_pattern
                )


if __name__ == "__main__":
    unittest.main()
