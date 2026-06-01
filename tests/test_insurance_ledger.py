from __future__ import annotations

import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from hr_toolkit.tools.insurance_ledger import generate_insurance_ledger


class InsuranceLedgerTest(unittest.TestCase):
    def test_generate_insurance_ledger(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            roster = root / "最新人员花名册.xlsx"
            input_dir.mkdir()
            _write_pzdx_policy(input_dir / "PZDX202536010000000011.xlsx")
            _write_peac_policy(input_dir / "PEAC202536010000000431.xlsx")
            _write_roster(roster)

            result = generate_insurance_ledger(input_dir, roster, output_dir)

            self.assertEqual(result.policy_count, 2)
            self.assertEqual(result.insured_person_count, 5)
            self.assertEqual(result.roster_person_count, 5)
            self.assertEqual(result.add_warning_count, 1)
            self.assertEqual(result.reduce_warning_count, 1)
            self.assertTrue(result.output_file and result.output_file.exists())
            self.assertTrue(result.roster_warning_file and result.roster_warning_file.exists())

            wb = load_workbook(result.output_file, data_only=True)
            ledger = wb["保险台账"]
            self.assertEqual([ledger.cell(2, col).value for col in range(1, 11)], ["序号", "姓名", "身份证号码", "项目/部门", "保单号1", "保额", "保单号2", "保额", "保额合计", "预警"])
            rows = {ledger.cell(row, 2).value: [ledger.cell(row, col).value for col in range(1, 11)] for row in range(3, 8)}
            self.assertEqual(rows["金三"][3:9], ["运维一部", "PZDX202536010000000011", 60, "PEAC202536010000000431", 60, 120])
            self.assertEqual(rows["金七"][3:10], ["运维五部", "PZDX202536010000000011", 60, None, None, 60, "需减保"])

            warning = wb["人员增减预警"]
            warning_rows = {
                warning.cell(row, 3).value: [warning.cell(row, col).value for col in range(1, 7)]
                for row in range(2, warning.max_row + 1)
            }
            self.assertEqual(warning_rows["金八"][1], "需加保")
            self.assertEqual(warning_rows["金七"][1], "需减保")
            wb.close()

            roster_wb = load_workbook(result.roster_warning_file, data_only=True)
            roster_ws = roster_wb["花名册"]
            warning_col = next(col for col in range(1, roster_ws.max_column + 1) if roster_ws.cell(1, col).value == "保险预警")
            self.assertEqual(roster_ws.cell(7, warning_col).value, "需加保")
            roster_wb.close()

    def test_generate_from_zip(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "PZDX202536010000000011.xlsx"
            archive = root / "保单.zip"
            roster = root / "最新人员花名册.xlsx"
            output_dir = root / "output"
            _write_pzdx_policy(source)
            _write_roster(roster)
            with zipfile.ZipFile(archive, "w") as zip_file:
                zip_file.write(source, arcname=source.name)

            result = generate_insurance_ledger([archive], roster, output_dir)

            self.assertEqual(len(result.source_files), 1)
            self.assertEqual(result.policy_count, 1)
            self.assertEqual(result.insured_person_count, 5)
            self.assertTrue(result.output_file and result.output_file.exists())

    def test_reads_demand6_analysis_roster_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            roster = root / "人力资源分析表.xlsx"
            input_dir.mkdir()
            _write_pzdx_policy(input_dir / "PZDX202536010000000011.xlsx")
            _write_demand6_analysis_roster(roster)

            result = generate_insurance_ledger(input_dir, roster, output_dir)

            self.assertEqual(result.roster_person_count, 5)
            self.assertEqual(result.add_warning_count, 1)
            self.assertEqual(result.reduce_warning_count, 1)
            self.assertTrue(result.roster_warning_file and result.roster_warning_file.exists())

            wb = load_workbook(result.output_file, data_only=True)
            ledger = wb["保险台账"]
            rows = {ledger.cell(row, 2).value: [ledger.cell(row, col).value for col in range(1, 9)] for row in range(3, 8)}
            self.assertEqual(rows["金三"][3], "南昌")
            self.assertEqual(rows["金七"][3], "抚州")
            self.assertEqual(rows["金七"][7], "需减保")
            warning = wb["人员增减预警"]
            warning_names = [warning.cell(row, 3).value for row in range(2, warning.max_row + 1)]
            self.assertIn("金八", warning_names)
            self.assertNotIn("姓名", warning_names)
            wb.close()

            roster_wb = load_workbook(result.roster_warning_file, data_only=True)
            roster_ws = roster_wb["花名册"]
            warning_col = next(col for col in range(1, 30) if roster_ws.cell(3, col).value == "保险预警")
            self.assertEqual(roster_ws.cell(9, warning_col).value, "需加保")
            roster_wb.close()


def _write_pzdx_policy(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(5, 1).value = "保单号码:PZDX202536010000000011"
    headers = ["雇员编号", "雇员姓名", "性别", "身份证号码", "雇员工种", "限额币别", "每人伤残死亡限额", "每人医疗费用限额"]
    for col, header in enumerate(headers, start=1):
        ws.cell(6, col).value = header
    rows = [
        [1, "金三", "男性", "330424123456241614", None, "CNY", "600,000.00", "60,000.00"],
        [2, "金四", "男性", "413026123456095128", None, "CNY", "600,000.00", "60,000.00"],
        [3, "金五", "男性", "320911123456171215", None, "CNY", "600,000.00", "60,000.00"],
        [4, "金六", "男性", "411521123456155377", None, "CNY", "600,000.00", "60,000.00"],
        [5, "金七", "男性", "362523123456066014", None, "CNY", "600,000.00", "60,000.00"],
    ]
    for row_index, row in enumerate(rows, start=7):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    wb.save(path)
    wb.close()


def _write_peac_policy(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(5, 7).value = "保险单号： PEAC202536010000000431"
    headers = ["序号", "姓名", "证件号", "年龄", "性别", "职业岗位", "分组组别", "姓   名", "证件号"]
    for col, header in enumerate(headers, start=1):
        ws.cell(7, col).value = header
    rows = [
        [1, "金三", "330424123456241614", 51, "男性", "维护工程师", 1, None, None],
        [2, "金四", "413026123456095128", 28, "男性", "维护工程师", 1, None, None],
        [3, "金五", "320911123456171215", 44, "男性", "维护工程师", 1, None, None],
        [4, "金六", "411521123456155377", 29, "男性", "维护工程师", 1, None, None],
    ]
    for row_index, row in enumerate(rows, start=9):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    wb.save(path)
    wb.close()


def _write_roster(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "花名册"
    headers = ["姓名", "身份证号码", "项目/部门", "员工状态"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    rows = [
        ["金三", "330424123456241614", "运维一部", "在职"],
        ["金四", "413026123456095128", "运维二部", "在职"],
        ["金五", "320911123456171215", "运维三部", "在职"],
        ["金六", "411521123456155377", "运维四部", "在职"],
        ["金七", "362523123456066014", "运维五部", "离职"],
        ["金八", "360111123456251234", "运维六部", "在职"],
    ]
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    wb.save(path)
    wb.close()


def _write_demand6_analysis_roster(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "总表"
    ws["A1"] = "总表"
    ws = wb.create_sheet("花名册")
    ws["A1"] = "2026年4月人员花名册"
    headers = ["序号", "部门", "部门/项目", "姓名", "身份证号码", "出生日期", "年龄", "性别", "岗位", "人员分类"]
    for col, header in enumerate(headers, start=1):
        ws.cell(3, col).value = header
    rows = [
        [1, "运营商事业部", "南昌", "金三", "330424123456241614"],
        [2, "运营商事业部", "南昌", "金四", "413026123456095128"],
        [3, "运营商事业部", "九江", "金五", "320911123456171215"],
        [4, "运营商事业部", "上饶", "金六", "411521123456155377"],
        [5, "运营商事业部", "抚州", "金七", "362523123456066014"],
        [6, "运营商事业部", "赣州", "金八", "360111123456251234"],
    ]
    leave_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    for row_index, row in enumerate(rows, start=4):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
            if row[3] == "金七":
                ws.cell(row_index, col_index).fill = leave_fill
    footer_row = 11
    footer = ["对应异动汇总表中的字段", "手动填写", "地市", "姓名", "身份证号码"]
    for col_index, value in enumerate(footer, start=1):
        ws.cell(footer_row, col_index).value = value
    wb.save(path)
    wb.close()


if __name__ == "__main__":
    unittest.main()
