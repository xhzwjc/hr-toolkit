from __future__ import annotations

import queue
import tempfile
import threading
import unittest
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from unittest.mock import Mock, patch

from openpyxl import Workbook

from hr_toolkit.gui import HRToolkitApp, make_result_output_dir
from hr_toolkit.project_store import CATEGORY_RESULTS, CATEGORY_UPLOADS, ProjectStore
from hr_toolkit.tools.folder_rename import (
    FILE_TYPE_FOLDER,
    FILE_TYPE_PDF,
    rename_files_by_excel,
)


class _FakeResult:
    def __init__(self, output_dir: Path) -> None:
        self.output_dir = output_dir
        self.warnings: list[str] = []

    def to_dict(self) -> dict[str, object]:
        return {
            "output_dir": str(self.output_dir),
            "output_file": str(self.output_dir / "结果.xlsx"),
            "record_count": 1,
            "warnings": [],
        }


class _Value:
    def __init__(self, value: str) -> None:
        self.value = value

    def get(self) -> str:
        return self.value


class HistoryGuiIntegrationTests(unittest.TestCase):
    def test_result_directory_failure_is_explained_before_processing_starts(self) -> None:
        app = HRToolkitApp.__new__(HRToolkitApp)
        app.root = object()
        with (
            patch("hr_toolkit.gui.make_result_output_dir", side_effect=PermissionError("read only")),
            patch("hr_toolkit.gui.messagebox.showerror") as showerror,
            patch("hr_toolkit.gui.runlog.log_exception"),
        ):
            result = app._prepare_result_output_dir(Path("/not-writable"))
        self.assertIsNone(result)
        self.assertEqual(showerror.call_args.args[0], "无法创建保存目录")
        self.assertIn("写入权限", showerror.call_args.args[1])

    def test_result_directories_are_atomically_unique(self) -> None:
        with tempfile.TemporaryDirectory() as temp_root:
            parent = Path(temp_root) / "results"
            with patch("hr_toolkit.gui._default_result_dir_name", return_value="结果_固定时间"):
                with ThreadPoolExecutor(max_workers=2) as executor:
                    paths = list(executor.map(lambda _index: make_result_output_dir(parent), range(2)))
            self.assertEqual(len(set(paths)), 2)
            self.assertTrue(all(path.is_dir() for path in paths))

    def test_worker_snapshots_inputs_and_writes_formal_result_directly_in_project(self) -> None:
        with tempfile.TemporaryDirectory() as temp_root:
            root = Path(temp_root)
            source = root / "工资.xlsx"
            source.write_bytes(b"input")
            project_store = ProjectStore.create(root / "工作项目", "工资处理")

            app = HRToolkitApp.__new__(HRToolkitApp)
            app._tool_run_token = 1
            app.current_tool = "salary_split"
            app.project_store = project_store
            app.current_project_path = project_store.root
            app._run_cancel_events = {1: threading.Event()}
            app._history_task_by_token = {}
            app._project_batch_by_token = {}
            app.status_queue = queue.Queue()

            def fake_tool(input_path: Path, output_dir: Path) -> _FakeResult:
                self.assertNotEqual(input_path, source)
                self.assertEqual(input_path.read_bytes(), b"input")
                self.assertTrue(output_dir.is_relative_to(project_store.root))
                (output_dir / "结果.xlsx").write_bytes(input_path.read_bytes())
                return _FakeResult(output_dir)

            original_import_sources = project_store.import_sources

            def import_then_change_source(*args, **kwargs):
                records = original_import_sources(*args, **kwargs)
                source.write_bytes(b"changed-after-snapshot")
                return records

            with (
                patch("hr_toolkit.gui.runlog.log_line"),
                patch("hr_toolkit.gui.runlog.log_exception"),
                patch.object(project_store, "import_sources", side_effect=import_then_change_source),
            ):
                app._start_tool_worker(fake_tool, source, root / "ignored-output")
                status, token, payload = app.status_queue.get(timeout=5)

            self.assertEqual((status, token), ("success", 1), str(payload))
            self.assertIsInstance(payload, _FakeResult)
            batch_id = app._project_batch_by_token[1]
            detail = project_store.get_batch(batch_id)
            assert detail is not None
            self.assertEqual(detail.summary.status, "success")
            uploaded = detail.files_for(CATEGORY_UPLOADS)
            results = detail.files_for(CATEGORY_RESULTS)
            self.assertEqual([item.display_name for item in uploaded], ["工资.xlsx"])
            self.assertEqual([item.display_name for item in results], ["结果.xlsx"])
            self.assertEqual(uploaded[0].path(project_store.workspace).read_bytes(), b"input")
            self.assertEqual(results[0].path(project_store.workspace).read_bytes(), b"input")
            self.assertEqual(payload.output_dir, detail.directories[CATEGORY_RESULTS])
            project_store.close()

    def test_worker_does_not_resolve_away_a_selected_symbolic_link(self) -> None:
        with tempfile.TemporaryDirectory() as temp_root:
            root = Path(temp_root)
            target = root / "真实工资.xlsx"
            target.write_bytes(b"private")
            selected_link = root / "工资链接.xlsx"
            try:
                selected_link.symlink_to(target)
            except (OSError, NotImplementedError):
                self.skipTest("当前文件系统不支持符号链接")
            project_store = ProjectStore.create(root / "工作项目", "工资处理")

            app = HRToolkitApp.__new__(HRToolkitApp)
            app._tool_run_token = 3
            app.current_tool = "salary_split"
            app.project_store = project_store
            app.current_project_path = project_store.root
            app._run_cancel_events = {3: threading.Event()}
            app._history_task_by_token = {}
            app._project_batch_by_token = {}
            app.status_queue = queue.Queue()

            with (
                patch("hr_toolkit.gui.runlog.log_line"),
                patch("hr_toolkit.gui.runlog.log_exception"),
            ):
                app._start_tool_worker(lambda input_path, output_dir: None, selected_link, root / "ignored")
                status, token, payload = app.status_queue.get(timeout=5)

            self.assertEqual((status, token), ("error", 3))
            self.assertRegex(str(payload), "链接")
            self.assertEqual(project_store.list_batches(), ())
            project_store.close()

    def test_workspace_import_is_cancelled_before_close_releases_project_lock(self) -> None:
        app = HRToolkitApp.__new__(HRToolkitApp)
        store = Mock()
        app.root = Mock()
        app.project_store = store
        app.current_project_path = Path("/tmp/project")
        app._tool_running = False
        app._history_task_by_token = {}
        app._project_batch_by_token = {}
        app._run_cancel_events = {}
        app.history_store = None
        app._workspace_project_generation = 1
        app._workspace_write_token = 0
        app._workspace_write_tasks = {}
        app._workspace_close_requested = False
        app._workspace_queue = queue.Queue()
        app._update_workspace_action_states = Mock()
        app._refresh_workspace_tree = Mock()
        app._update_sidebar_project_summary = Mock()
        app._save_workspace_preferences = Mock()
        started = threading.Event()
        cancelled = threading.Event()

        def background_write(is_cancelled) -> None:
            started.set()
            self.assertTrue(cancelled.wait(timeout=5))
            self.assertTrue(is_cancelled())

        app._workspace_run_write("导入文件", background_write)
        self.assertTrue(started.wait(timeout=5))
        self.assertTrue(app._project_change_is_blocked())
        with patch("hr_toolkit.gui.messagebox.askyesno", return_value=True):
            app._request_close()
        cancelled.set()
        self.assertFalse(app.root.destroy.called)

        status = app._workspace_queue.get(timeout=5)
        app._workspace_queue.put(status)
        app._poll_workspace_queue()
        self.assertFalse(app._workspace_write_tasks)
        store.close.assert_called_once_with()
        app.root.destroy.assert_called_once_with()

    def test_project_finalization_failure_keeps_switch_guard_active(self) -> None:
        app = HRToolkitApp.__new__(HRToolkitApp)
        app.root = Mock()
        app.status_queue = queue.Queue()
        app.status_queue.put(
            (
                "project_finalize_error",
                9,
                (RuntimeError("处理失败"), OSError("清单无法写入")),
            )
        )
        app._tool_run_token = 9
        app._history_task_by_token = {}
        app._project_batch_by_token = {9: "batch-id"}
        app._run_cancel_events = {9: threading.Event()}
        app.current_view = "tool"
        app.current_project_path = Path("/tmp/project")
        app._finish_tool_run = Mock()
        app._record_last_run = Mock()
        app._write_log = Mock()
        app._show_error_after_log = Mock()
        app._refresh_workspace_tree = Mock()
        app._update_workspace_action_states = Mock()
        app._update_sidebar_project_summary = Mock()

        with patch("hr_toolkit.gui.runlog.log_exception"):
            app._poll_status_queue()

        self.assertEqual(app._project_batch_by_token, {9: "batch-id"})
        self.assertIn(9, app._run_cancel_events)
        app._finish_tool_run.assert_called_once_with()
        app._show_error_after_log.assert_called_once()

    def test_folder_rename_changes_only_the_project_result_copy(self) -> None:
        with tempfile.TemporaryDirectory() as temp_root:
            root = Path(temp_root)
            source = root / "人员资料"
            (source / "张三").mkdir(parents=True)
            (source / "张三" / "说明.txt").write_text("record", encoding="utf-8")
            project_store = ProjectStore.create(root / "工作项目", "资料整理")

            app = HRToolkitApp.__new__(HRToolkitApp)
            app._tool_run_token = 2
            app.current_tool = "folder_rename"
            app.rename_mode = _Value("追加文字")
            app.project_store = project_store
            app.current_project_path = project_store.root
            app._run_cancel_events = {2: threading.Event()}
            app._history_task_by_token = {}
            app._project_batch_by_token = {}
            app.status_queue = queue.Queue()

            def fake_rename(root_dir: Path, *, mode: str) -> _FakeResult:
                self.assertEqual(mode, "append")
                (root_dir / "张三").rename(root_dir / "张三-合同")
                return _FakeResult(root_dir)

            with (
                patch("hr_toolkit.gui.runlog.log_line"),
                patch("hr_toolkit.gui.runlog.log_exception"),
            ):
                app._start_tool_worker(fake_rename, root_dir=source, mode="append")
                status, token, payload = app.status_queue.get(timeout=5)

            self.assertEqual((status, token), ("success", 2), str(payload))
            self.assertTrue((source / "张三" / "说明.txt").is_file())
            self.assertFalse((source / "张三-合同").exists())
            detail = project_store.get_batch(app._project_batch_by_token[2])
            assert detail is not None
            result_root = detail.directories[CATEGORY_RESULTS] / source.name
            self.assertTrue((result_root / "张三-合同" / "说明.txt").is_file())
            self.assertFalse((result_root / "张三").exists())
            project_store.close()

    def test_excel_folder_rename_uses_archived_roster_and_result_copy(self) -> None:
        with tempfile.TemporaryDirectory() as temp_root:
            root = Path(temp_root)
            source = root / "人员资料"
            source.mkdir()
            original_file = source / "1.pdf"
            original_file.write_bytes(b"record")
            roster = root / "人员名单.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["姓名"])
            worksheet.append(["张三"])
            workbook.save(roster)
            workbook.close()
            roster_bytes = roster.read_bytes()
            confirmed_preview = rename_files_by_excel(source, roster, file_type=FILE_TYPE_PDF, dry_run=True)
            expected_operations = [
                (operation.source.name, operation.target.name)
                for operation in confirmed_preview.operations
            ]
            project_store = ProjectStore.create(root / "工作项目", "名单改名")

            app = HRToolkitApp.__new__(HRToolkitApp)
            app._tool_run_token = 3
            app.current_tool = "folder_rename"
            app.rename_mode = _Value("按 Excel 人名顺序批量重命名")
            app.project_store = project_store
            app.current_project_path = project_store.root
            app._run_cancel_events = {3: threading.Event()}
            app._history_task_by_token = {}
            app._project_batch_by_token = {}
            app.status_queue = queue.Queue()

            def checked_rename(
                root_dir: Path,
                excel_path: Path,
                *,
                file_type: str,
                expected_operations: list[tuple[str, str]],
                expected_warnings: list[str],
            ):
                self.assertNotEqual(root_dir, source)
                self.assertNotEqual(excel_path, roster)
                self.assertTrue(root_dir.is_relative_to(project_store.root))
                self.assertTrue(excel_path.is_relative_to(project_store.root))
                self.assertEqual(excel_path.read_bytes(), roster_bytes)
                return rename_files_by_excel(
                    root_dir,
                    excel_path,
                    file_type=file_type,
                    expected_operations=expected_operations,
                    expected_warnings=expected_warnings,
                )

            with (
                patch("hr_toolkit.gui.runlog.log_line"),
                patch("hr_toolkit.gui.runlog.log_exception"),
            ):
                app._start_tool_worker(
                    checked_rename,
                    root_dir=source,
                    excel_path=roster,
                    file_type=FILE_TYPE_PDF,
                    expected_operations=expected_operations,
                    expected_warnings=list(confirmed_preview.warnings),
                )
                status, token, payload = app.status_queue.get(timeout=5)

            self.assertEqual((status, token), ("success", 3), str(payload))
            self.assertEqual(original_file.read_bytes(), b"record")
            self.assertEqual(roster.read_bytes(), roster_bytes)
            detail = project_store.get_batch(app._project_batch_by_token[3])
            assert detail is not None
            result_root = detail.directories[CATEGORY_RESULTS] / source.name
            self.assertEqual((result_root / "张三.pdf").read_bytes(), b"record")
            self.assertFalse((result_root / "1.pdf").exists())
            archived_rosters = list(detail.directories[CATEGORY_UPLOADS].rglob(roster.name))
            self.assertTrue(archived_rosters)
            self.assertTrue(any(path.read_bytes() == roster_bytes for path in archived_rosters))
            project_store.close()

    def test_excel_folder_rename_handles_only_empty_person_folders(self) -> None:
        with tempfile.TemporaryDirectory() as temp_root:
            root = Path(temp_root)
            source = root / "testname"
            source.mkdir()
            for name in ("54", "4343", "2331221"):
                (source / name).mkdir()
            finder_metadata = source / ".DS_Store"
            finder_metadata.write_bytes(b"finder metadata")
            source_before = sorted(
                path.relative_to(source).as_posix() for path in source.rglob("*")
            )
            metadata_before = finder_metadata.read_bytes()

            roster = root / "人员名单.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["姓名"])
            for name in ("王京川", "张三", "王五"):
                worksheet.append([name])
            workbook.save(roster)
            workbook.close()
            roster_before = roster.read_bytes()
            confirmed_preview = rename_files_by_excel(
                source,
                roster,
                file_type=FILE_TYPE_FOLDER,
                dry_run=True,
            )
            expected_operations = [
                (operation.source.name, operation.target.name)
                for operation in confirmed_preview.operations
            ]
            project_store = ProjectStore.create(root / "工作项目", "名单改名")

            app = HRToolkitApp.__new__(HRToolkitApp)
            app._tool_run_token = 4
            app.current_tool = "folder_rename"
            app.rename_mode = _Value("按 Excel 人名顺序批量重命名")
            app.project_store = project_store
            app.current_project_path = project_store.root
            app._run_cancel_events = {4: threading.Event()}
            app._history_task_by_token = {}
            app._project_batch_by_token = {}
            app.status_queue = queue.Queue()

            with (
                patch("hr_toolkit.gui.runlog.log_line"),
                patch("hr_toolkit.gui.runlog.log_exception"),
            ):
                app._start_tool_worker(
                    rename_files_by_excel,
                    root_dir=source,
                    excel_path=roster,
                    file_type=FILE_TYPE_FOLDER,
                    expected_operations=expected_operations,
                    expected_warnings=list(confirmed_preview.warnings),
                )
                status, token, payload = app.status_queue.get(timeout=5)

            self.assertEqual((status, token), ("success", 4), str(payload))
            self.assertEqual(
                sorted(path.relative_to(source).as_posix() for path in source.rglob("*")),
                source_before,
            )
            self.assertEqual(finder_metadata.read_bytes(), metadata_before)
            self.assertEqual(roster.read_bytes(), roster_before)

            detail = project_store.get_batch(app._project_batch_by_token[4])
            assert detail is not None
            result_root = detail.directories[CATEGORY_RESULTS] / source.name
            self.assertEqual(
                {path.name for path in result_root.iterdir()},
                {"王京川", "张三", "王五"},
            )
            self.assertTrue(all(path.is_dir() for path in result_root.iterdir()))
            upload_root = detail.directories[CATEGORY_UPLOADS] / source.name
            self.assertEqual(
                sorted(path.name for path in upload_root.iterdir()),
                ["2331221", "4343", "54"],
            )
            project_store.close()

    def test_excel_folder_rename_gui_previews_then_dispatches_excel_mode(self) -> None:
        with tempfile.TemporaryDirectory() as temp_root:
            root = Path(temp_root)
            source = root / "人员资料"
            source.mkdir()
            (source / "1.pdf").write_bytes(b"record")
            roster = root / "人员名单.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["姓名"])
            worksheet.append(["张三"])
            workbook.save(roster)
            workbook.close()

            app = HRToolkitApp.__new__(HRToolkitApp)
            app.input_path = _Value(str(source))
            app.summary_path = _Value(str(roster))
            app.rename_mode = _Value("按 Excel 人名顺序批量重命名")
            app.rename_file_type = _Value("PDF")
            app.rename_text = _Value("")
            app.rename_target_name = _Value("")
            app.rename_replacement_name = _Value("")
            app._clear_log = Mock()
            app._write_log = Mock()
            app._begin_tool_run = Mock()
            app._start_tool_worker = Mock()

            with (
                patch("hr_toolkit.gui.app.messagebox.askyesno", return_value=True),
                patch("hr_toolkit.gui.app.messagebox.showwarning") as showwarning,
                patch("hr_toolkit.gui.app.messagebox.showerror") as showerror,
            ):
                app._run_folder_rename()

            showwarning.assert_not_called()
            showerror.assert_not_called()
            app._begin_tool_run.assert_called_once_with()
            app._start_tool_worker.assert_called_once_with(
                rename_files_by_excel,
                root_dir=source,
                excel_path=roster,
                file_type=FILE_TYPE_PDF,
                expected_operations=[("1.pdf", "张三.pdf")],
                expected_warnings=[],
            )
            self.assertTrue((source / "1.pdf").is_file())

    def test_excel_folder_rename_mode_alone_shows_roster_row(self) -> None:
        app = HRToolkitApp.__new__(HRToolkitApp)
        app.current_tool = "folder_rename"
        app.rename_mode = _Value("按 Excel 人名顺序批量重命名")

        app._update_summary_controls(apply_layout=False)
        self.assertTrue(app._summary_row_visible)

        app.rename_mode = _Value("追加文字")
        app._update_summary_controls(apply_layout=False)
        self.assertFalse(app._summary_row_visible)

    def test_folder_rename_context_keeps_before_and_after_snapshots(self) -> None:
        with tempfile.TemporaryDirectory() as temp_root:
            root = Path(temp_root) / "人员资料"
            root.mkdir()
            (root / "说明.txt").write_text("record", encoding="utf-8")
            app = HRToolkitApp.__new__(HRToolkitApp)

            def rename_tool(root_dir: Path, *, mode: str) -> None:
                return None

            sources, parameters, output_dir = app._history_context_from_call(
                rename_tool,
                (),
                {"root_dir": root, "mode": "append"},
            )

            self.assertEqual(output_dir, root)
            self.assertEqual(parameters["root_dir"], root.name)
            self.assertEqual(len(sources), 1)
            self.assertIsNone(sources[0].suffixes)
            self.assertTrue(sources[0].preserve_directories)

    def test_stored_result_removes_absolute_paths_but_keeps_counts(self) -> None:
        payload = HRToolkitApp._history_result_for_storage(
            {
                "output_file": "/Users/hr/候选人/结果.xlsx",
                "source_files": ["/Users/hr/候选人/原表.xlsx"],
                "record_count": 8,
            }
        )
        self.assertEqual(payload["output_file"], "结果.xlsx")
        self.assertEqual(payload["source_files"], ["原表.xlsx"])
        self.assertEqual(payload["record_count"], 8)


if __name__ == "__main__":
    unittest.main()
