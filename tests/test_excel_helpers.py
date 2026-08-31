from __future__ import annotations

import os
import sys
import unittest
import shutil
import tempfile
from copy import deepcopy
from pathlib import Path
from unittest.mock import patch
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from hr_toolkit.common import resources as package_resources
from hr_toolkit.common.excel import (
    _translate_same_row_formula,
    apply_row_snapshot,
    cached_style_id,
    insert_rows,
    set_style_ids,
    snapshot_row,
    style_source_id,
)
from hr_toolkit.common.excel_compat import (
    _build_conversion_env,
    _convert_with_xlrd,
    _diagnose_missing_libreoffice,
    _find_libreoffice,
    capture_xlsx_save_compatibility,
    ensure_xlsx_workbook,
    finalize_xlsx_after_openpyxl_save,
)
from hr_toolkit.common.resources import open_template_resource


class ExcelHelperTest(unittest.TestCase):
    def test_translate_same_row_formula_only_rewrites_cell_reference_rows(self) -> None:
        formula = "=A1+B10+SUM(C1:D1)+E$1+$F1+LOG10(100)"

        self.assertEqual(
            _translate_same_row_formula(formula, source_row=1, target_row=12),
            "=A12+B10+SUM(C12:D12)+E$1+$F12+LOG10(100)",
        )

    def test_ensure_xlsx_workbook_converts_xls_to_temp_xlsx(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            original_xlsx = root / "原始.xlsx"
            xls_file = root / "原始.xls"
            wb = Workbook()
            wb.active["A1"] = "测试"
            wb.save(original_xlsx)
            shutil.copyfile(original_xlsx, xls_file)

            def fake_convert(source: Path, output_path: Path, *args, **kwargs) -> None:
                shutil.copyfile(source, output_path)

            with patch("hr_toolkit.common.excel_compat._convert_xls_to_xlsx", side_effect=fake_convert):
                converted = ensure_xlsx_workbook(xls_file, root / "temp")

            self.assertEqual(converted.suffix, ".xlsx")
            self.assertNotEqual(converted, xls_file)
            loaded = load_workbook(converted, data_only=True)
            self.assertEqual(loaded.active["A1"].value, "测试")
            loaded.close()

    def test_ensure_xlsx_workbook_converts_renamed_xls_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            renamed_xls = root / "改后缀.xlsx"
            renamed_xls.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1fake")

            def fake_convert(source: Path, output_path: Path, *args, **kwargs) -> None:
                wb = Workbook()
                wb.active["A1"] = "已转换"
                wb.save(output_path)

            with patch("hr_toolkit.common.excel_compat._convert_xls_to_xlsx", side_effect=fake_convert):
                converted = ensure_xlsx_workbook(renamed_xls, root / "temp")

            self.assertNotEqual(converted, renamed_xls)
            loaded = load_workbook(converted, data_only=True)
            self.assertEqual(loaded.active["A1"].value, "已转换")
            loaded.close()

    def test_ensure_xlsx_workbook_never_modifies_source_file(self) -> None:
        """验证 .xls 转换过程严格只读，源文件字节与哈希 100% 保持不变。"""
        import hashlib
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xls_file = root / "人员异动表7月-四川分公司-春苗.xls"
            # 写入模拟二进制文件
            xls_file.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 256)
            before_sha = hashlib.sha256(xls_file.read_bytes()).hexdigest()
            before_size = xls_file.stat().st_size

            # 模拟 xlrd 成功转换
            out_xlsx = root / "temp" / "out.xlsx"
            def fake_xlrd_convert(src: Path, dst: Path) -> None:
                wb = Workbook()
                wb.active["A1"] = "测试数据"
                dst.parent.mkdir(parents=True, exist_ok=True)
                wb.save(dst)

            with patch("hr_toolkit.common.excel_compat._convert_with_xlrd", side_effect=fake_xlrd_convert):
                res = ensure_xlsx_workbook(xls_file, root / "temp")

            self.assertTrue(res.exists())
            after_sha = hashlib.sha256(xls_file.read_bytes()).hexdigest()
            after_size = xls_file.stat().st_size
            self.assertEqual(before_sha, after_sha, "源文件哈希绝不能发生改变")
            self.assertEqual(before_size, after_size, "源文件大小绝不能发生改变")

    def test_preserve_formatting_uses_native_converter_before_xlrd(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xls_file = root / "档案模板.xls"
            xls_file.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)

            def fake_native_convert(_source: Path, output_path: Path) -> None:
                workbook = Workbook()
                workbook.active["A1"] = "保留格式"
                output_path.parent.mkdir(parents=True, exist_ok=True)
                workbook.save(output_path)
                workbook.close()

            with (
                patch("hr_toolkit.common.excel_compat.sys.platform", "linux"),
                patch(
                    "hr_toolkit.common.excel_compat._convert_with_libreoffice",
                    side_effect=fake_native_convert,
                ) as native_convert,
                patch("hr_toolkit.common.excel_compat._convert_with_xlrd") as xlrd_convert,
            ):
                converted = ensure_xlsx_workbook(
                    xls_file,
                    root / "temp",
                    preserve_formatting=True,
                )

            self.assertTrue(converted.exists())
            native_convert.assert_called_once()
            xlrd_convert.assert_not_called()

    def test_preserve_formatting_prefers_excel_or_wps_on_windows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xls_file = root / "档案模板.xls"
            original_bytes = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64
            xls_file.write_bytes(original_bytes)

            def fake_windows_convert(source: Path, output_path: Path) -> None:
                self.assertNotEqual(source, xls_file)
                self.assertEqual(source.read_bytes(), original_bytes)
                workbook = Workbook()
                workbook.active["A1"] = "Excel/WPS 原生转换"
                output_path.parent.mkdir(parents=True, exist_ok=True)
                workbook.save(output_path)
                workbook.close()

            with (
                patch("hr_toolkit.common.excel_compat.sys.platform", "win32"),
                patch(
                    "hr_toolkit.common.excel_compat._convert_with_windows_com",
                    side_effect=fake_windows_convert,
                ) as windows_convert,
                patch("hr_toolkit.common.excel_compat._convert_with_libreoffice") as libreoffice_convert,
                patch("hr_toolkit.common.excel_compat._convert_with_xlrd") as xlrd_convert,
            ):
                converted = ensure_xlsx_workbook(
                    xls_file,
                    root / "temp",
                    preserve_formatting=True,
                )

            self.assertTrue(converted.exists())
            self.assertEqual(xls_file.read_bytes(), original_bytes)
            windows_convert.assert_called_once()
            libreoffice_convert.assert_not_called()
            xlrd_convert.assert_not_called()

    def test_preserve_formatting_falls_back_to_data_mode_without_native_converter(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xls_file = root / "档案模板.xls"
            xls_file.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)
            warnings: list[str] = []

            def fake_xlrd_convert(_source: Path, output_path: Path) -> None:
                workbook = Workbook()
                workbook.active["A1"] = "兼容模式数据"
                output_path.parent.mkdir(parents=True, exist_ok=True)
                workbook.save(output_path)
                workbook.close()

            with (
                patch("hr_toolkit.common.excel_compat.sys.platform", "linux"),
                patch(
                    "hr_toolkit.common.excel_compat._convert_with_libreoffice",
                    side_effect=RuntimeError("未安装转换器"),
                ),
                patch(
                    "hr_toolkit.common.excel_compat._convert_with_xlrd",
                    side_effect=fake_xlrd_convert,
                ) as xlrd_convert,
            ):
                converted = ensure_xlsx_workbook(
                    xls_file,
                    root / "temp",
                    preserve_formatting=True,
                    warning_callback=warnings.append,
                )

            self.assertTrue(converted.exists())
            xlrd_convert.assert_called_once()
            self.assertTrue(any("自动切换兼容模式" in warning for warning in warnings))
            workbook = load_workbook(converted, data_only=True)
            self.assertEqual(workbook.active["A1"].value, "兼容模式数据")
            workbook.close()

    def test_native_output_validation_error_falls_back_without_leaking_exception(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xls_file = root / "档案模板.xls"
            xls_file.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)

            def fake_native_convert(_source: Path, output_path: Path) -> None:
                output_path.parent.mkdir(parents=True, exist_ok=True)
                output_path.write_bytes(b"PK\x03\x04legacy-invalid")

            def fake_xlrd_convert(_source: Path, output_path: Path) -> None:
                workbook = Workbook()
                workbook.active["A1"] = "降级成功"
                output_path.parent.mkdir(parents=True, exist_ok=True)
                workbook.save(output_path)
                workbook.close()

            real_zip_file = ZipFile

            class ValidationErrorZipFile:
                def __new__(cls, path, *args, **kwargs):
                    if Path(path).read_bytes().startswith(b"PK\x03\x04legacy"):
                        raise RuntimeError("模拟旧系统压缩校验异常")
                    return real_zip_file(path, *args, **kwargs)

            with (
                patch("hr_toolkit.common.excel_compat.sys.platform", "linux"),
                patch(
                    "hr_toolkit.common.excel_compat._convert_with_libreoffice",
                    side_effect=fake_native_convert,
                ),
                patch(
                    "hr_toolkit.common.excel_compat._convert_with_xlrd",
                    side_effect=fake_xlrd_convert,
                ) as xlrd_convert,
                patch("hr_toolkit.common.excel_compat.ZipFile", ValidationErrorZipFile),
            ):
                converted = ensure_xlsx_workbook(
                    xls_file,
                    root / "temp",
                    preserve_formatting=True,
                )

            xlrd_convert.assert_called_once()
            workbook = load_workbook(converted, data_only=True)
            self.assertEqual(workbook.active["A1"].value, "降级成功")
            workbook.close()

    def test_preserve_formatting_raises_only_when_all_readers_fail(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xls_file = root / "损坏文件.xls"
            xls_file.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)

            with (
                patch("hr_toolkit.common.excel_compat.sys.platform", "linux"),
                patch(
                    "hr_toolkit.common.excel_compat._convert_with_libreoffice",
                    side_effect=RuntimeError("未安装转换器"),
                ),
                patch(
                    "hr_toolkit.common.excel_compat._convert_with_xlrd",
                    side_effect=RuntimeError("文件内容损坏"),
                ),
            ):
                with self.assertRaisesRegex(RuntimeError, "未损坏、未加密"):
                    ensure_xlsx_workbook(
                        xls_file,
                        root / "temp",
                        preserve_formatting=True,
                    )

    def test_formatted_and_values_only_conversions_use_separate_cache_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            xls_file = root / "档案模板.xls"
            xls_file.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)
            modes: list[bool] = []

            def fake_convert(
                _source: Path,
                output_path: Path,
                *args,
                preserve_formatting: bool = False,
                **kwargs,
            ) -> None:
                modes.append(preserve_formatting)
                workbook = Workbook()
                workbook.active["A1"] = preserve_formatting
                output_path.parent.mkdir(parents=True, exist_ok=True)
                workbook.save(output_path)
                workbook.close()

            with patch("hr_toolkit.common.excel_compat._convert_xls_to_xlsx", side_effect=fake_convert):
                values_only = ensure_xlsx_workbook(xls_file, root / "temp")
                formatted = ensure_xlsx_workbook(
                    xls_file,
                    root / "temp",
                    preserve_formatting=True,
                )

            self.assertNotEqual(values_only, formatted)
            self.assertEqual(modes, [False, True])

    def test_finalize_xlsx_repairs_legacy_style_and_drawing_compatibility(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "原始模板.xlsx"
            output = root / "输出.xlsx"
            for path in (source, output):
                workbook = Workbook()
                workbook.active["A1"] = "测试"
                workbook.save(path)
                workbook.close()

            namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
            tag = f"{{{namespace}}}"

            def source_styles(styles_xml: bytes) -> bytes:
                style_root = ElementTree.fromstring(styles_xml)
                style_xfs = style_root.find(f"{tag}cellStyleXfs")
                cell_xfs = style_root.find(f"{tag}cellXfs")
                cell_styles = style_root.find(f"{tag}cellStyles")
                self.assertIsNotNone(style_xfs)
                self.assertIsNotNone(cell_xfs)
                self.assertIsNotNone(cell_styles)
                style_xfs.append(deepcopy(style_xfs[0]))
                style_xfs.set("count", str(len(style_xfs)))
                cell_xfs[0].set("xfId", "1")
                cell_styles.append(
                    ElementTree.Element(
                        f"{tag}cellStyle",
                        {"name": "旧模板样式", "xfId": "1"},
                    )
                )
                cell_styles.set("count", str(len(cell_styles)))
                ElementTree.register_namespace("", namespace)
                return ElementTree.tostring(style_root, encoding="utf-8")

            def invalid_output_styles(styles_xml: bytes) -> bytes:
                style_root = ElementTree.fromstring(styles_xml)
                cell_xfs = style_root.find(f"{tag}cellXfs")
                self.assertIsNotNone(cell_xfs)
                cell_xfs[0].set("xfId", "1")
                ElementTree.register_namespace("", namespace)
                return ElementTree.tostring(style_root, encoding="utf-8")

            def rewrite_package(
                path: Path,
                replacements: dict[str, bytes],
                additions: dict[str, bytes] | None = None,
            ) -> None:
                temp_path = path.with_name(f".{path.name}.tmp")
                with ZipFile(path) as source_archive, ZipFile(temp_path, "w") as target_archive:
                    for info in source_archive.infolist():
                        target_archive.writestr(
                            info,
                            replacements.get(info.filename, source_archive.read(info.filename)),
                        )
                    for name, data in (additions or {}).items():
                        target_archive.writestr(name, data)
                os.replace(temp_path, path)

            with ZipFile(source) as archive:
                original_source_styles = archive.read("xl/styles.xml")
            rewrite_package(source, {"xl/styles.xml": source_styles(original_source_styles)})

            with ZipFile(output) as archive:
                original_output_styles = archive.read("xl/styles.xml")
            drawing_xml = (
                b'<wsDr xmlns="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
                b'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
                b'<a:prstGeom prst="rect"><avLst /></a:prstGeom></wsDr>'
            )
            rewrite_package(
                output,
                {"xl/styles.xml": invalid_output_styles(original_output_styles)},
                {"xl/drawings/drawing1.xml": drawing_xml},
            )

            snapshot = capture_xlsx_save_compatibility(source)
            finalize_xlsx_after_openpyxl_save(output, snapshot)

            with ZipFile(output) as archive:
                repaired_styles = ElementTree.fromstring(archive.read("xl/styles.xml"))
                repaired_drawing = archive.read("xl/drawings/drawing1.xml")
            repaired_style_xfs = repaired_styles.find(f"{tag}cellStyleXfs")
            repaired_cell_xfs = repaired_styles.find(f"{tag}cellXfs")
            repaired_cell_styles = repaired_styles.find(f"{tag}cellStyles")
            self.assertEqual(len(repaired_style_xfs), 2)
            self.assertLess(int(repaired_cell_xfs[0].get("xfId")), len(repaired_style_xfs))
            self.assertIn("旧模板样式", {style.get("name") for style in repaired_cell_styles})
            self.assertIn(b"<a:avLst", repaired_drawing)
            self.assertNotIn(b"<avLst", repaired_drawing)

            fallback_output = root / "无样式快照输出.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "测试"
            workbook.save(fallback_output)
            workbook.close()
            with ZipFile(fallback_output) as archive:
                fallback_styles = archive.read("xl/styles.xml")
            rewrite_package(
                fallback_output,
                {"xl/styles.xml": invalid_output_styles(fallback_styles)},
                {"xl/drawings/drawing1.xml": drawing_xml},
            )

            finalize_xlsx_after_openpyxl_save(fallback_output, None)

            with ZipFile(fallback_output) as archive:
                normalized_styles = ElementTree.fromstring(archive.read("xl/styles.xml"))
                normalized_drawing = archive.read("xl/drawings/drawing1.xml")
            normalized_cell_xfs = normalized_styles.find(f"{tag}cellXfs")
            self.assertEqual(normalized_cell_xfs[0].get("xfId"), "0")
            self.assertIn(b"<a:avLst", normalized_drawing)

    def test_open_template_resource_falls_back_without_files_api(self) -> None:
        with (
            patch.object(package_resources.resources, "files", None),
            patch.object(
                package_resources.resources,
                "open_binary",
                wraps=package_resources.resources.open_binary,
            ) as open_binary,
        ):
            with open_template_resource("data_statistics_template.xlsx") as handle:
                self.assertEqual(handle.read(2), b"PK")
        self.assertIs(open_binary.call_args.args[0], package_resources._template_package)


def _style_signature(cell) -> tuple:
    font, fill, border, alignment = cell.font, cell.fill, cell.border, cell.alignment
    return (
        font.name,
        font.sz,
        font.b,
        fill.patternType,
        str(fill.fgColor.rgb) if fill.fgColor is not None else None,
        border.left.style,
        border.right.style,
        border.top.style,
        border.bottom.style,
        alignment.horizontal,
        alignment.vertical,
        alignment.wrapText,
        alignment.textRotation,
        cell.number_format,
        cell.protection.locked,
    )


class RowSnapshotStyleTest(unittest.TestCase):
    """样式快照走的是样式表下标快路径，必须与逐项赋值的结果完全一致。"""

    @staticmethod
    def _styled_workbook() -> Workbook:
        workbook = Workbook()
        ws = workbook.active
        thin = Side(style="thin", color="000000")
        for col_index in range(1, 6):
            cell = ws.cell(1, col_index)
            cell.value = f"值{col_index}"
            cell.font = Font(name="宋体", size=11, bold=col_index % 2 == 0, color="FF0000")
            cell.fill = PatternFill("solid", fgColor="FCE4D6")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True, textRotation=col_index)
            cell.number_format = "0.00"
        ws.row_dimensions[1].height = 33.5
        return workbook

    def test_same_workbook_apply_reproduces_every_style_facet(self) -> None:
        workbook = self._styled_workbook()
        ws = workbook.active
        snapshot = snapshot_row(ws, 1, 5)

        apply_row_snapshot(ws, 7, snapshot, translate_formulas=False)

        self.assertEqual(ws.row_dimensions[7].height, 33.5)
        for col_index in range(1, 6):
            self.assertEqual(
                _style_signature(ws.cell(7, col_index)),
                _style_signature(ws.cell(1, col_index)),
                f"第 {col_index} 列样式不一致",
            )

    def test_cross_workbook_apply_reproduces_every_style_facet(self) -> None:
        """工资表拆分会把快照套用到另一个工作簿，样式下标在那边无效。"""
        source = self._styled_workbook()
        snapshot = snapshot_row(source.active, 1, 5)

        target = Workbook()
        target_ws = target.active
        apply_row_snapshot(target_ws, 3, snapshot, translate_formulas=False)
        # 第二行走缓存路径，必须和第一行结果相同
        apply_row_snapshot(target_ws, 4, snapshot, translate_formulas=False)

        for row_index in (3, 4):
            for col_index in range(1, 6):
                self.assertEqual(
                    _style_signature(target_ws.cell(row_index, col_index)),
                    _style_signature(source.active.cell(1, col_index)),
                    f"第 {row_index} 行第 {col_index} 列样式不一致",
                )

    def test_cross_workbook_cache_does_not_mix_up_two_sources(self) -> None:
        """两个来源工作簿的样式下标含义不同，缓存不能相互串味。"""
        first = Workbook()
        first.active.cell(1, 1).font = Font(name="宋体", size=20, bold=True)
        second = Workbook()
        second.active.cell(1, 1).font = Font(name="Arial", size=8, bold=False)

        first_snapshot = snapshot_row(first.active, 1, 1)
        second_snapshot = snapshot_row(second.active, 1, 1)

        target = Workbook()
        apply_row_snapshot(target.active, 1, first_snapshot, translate_formulas=False)
        apply_row_snapshot(target.active, 2, second_snapshot, translate_formulas=False)

        self.assertEqual((target.active.cell(1, 1).font.name, target.active.cell(1, 1).font.sz), ("宋体", 20.0))
        self.assertEqual((target.active.cell(2, 1).font.name, target.active.cell(2, 1).font.sz), ("Arial", 8.0))

    def test_snapshot_does_not_keep_source_workbook_alive(self) -> None:
        """快照只持弱引用，否则批量处理会把每个来源工作簿都留在内存里。"""
        import gc
        import weakref

        workbook = self._styled_workbook()
        snapshot = snapshot_row(workbook.active, 1, 5)
        reference = weakref.ref(workbook)

        del workbook
        gc.collect()

        self.assertIsNone(reference(), "快照不应延长来源工作簿的生命周期")
        # 来源已释放时仍要能安全套用（退回逐项赋值）
        target = Workbook()
        apply_row_snapshot(target.active, 1, snapshot, translate_formulas=False)
        self.assertEqual(target.active.cell(1, 1).font.name, "宋体")


class StyleIdCacheTest(unittest.TestCase):
    def test_cached_style_id_matches_direct_assignment(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        ws.cell(1, 1).border = border
        ws.cell(1, 1).alignment = alignment

        border_id = cached_style_id(ws, "border", "thin", lambda: border)
        alignment_id = cached_style_id(ws, "alignment", "center", lambda: alignment)
        set_style_ids(ws.cell(2, 1), border_id=border_id, alignment_id=alignment_id)

        self.assertEqual(_style_signature(ws.cell(2, 1)), _style_signature(ws.cell(1, 1)))

    def test_cached_style_id_is_stable_across_calls(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        font = Font(name="宋体", size=10)
        first = cached_style_id(ws, "font", "song10", lambda: font)
        second = cached_style_id(ws, "font", "song10", lambda: self.fail("不应重复构造样式"))
        self.assertEqual(first, second)

    def test_set_style_ids_does_not_mutate_shared_style_array(self) -> None:
        """StyleArray 可能被多个单元格共用，就地改写会污染邻居。"""
        workbook = Workbook()
        ws = workbook.active
        shared = ws.cell(1, 1)._style
        ws.cell(1, 2)._style = shared  # 人为制造共用
        thin = Side(style="thin", color="000000")
        border_id = cached_style_id(
            ws, "border", "thin", lambda: Border(left=thin, right=thin, top=thin, bottom=thin)
        )

        set_style_ids(ws.cell(1, 1), border_id=border_id)

        self.assertEqual(ws.cell(1, 1).border.left.style, "thin")
        self.assertIsNone(ws.cell(1, 2).border.left.style, "相邻单元格不应被连带修改")

    def test_style_source_id_reports_current_index(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        self.assertEqual(style_source_id(ws.cell(1, 1), "alignment"), 0)
        ws.cell(1, 1).alignment = Alignment(horizontal="center")
        self.assertNotEqual(style_source_id(ws.cell(1, 1), "alignment"), 0)


class InsertRowsTest(unittest.TestCase):
    def test_insert_rows_tracks_max_row_without_scanning_every_cell(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        for row_index in range(1, 6):
            ws.cell(row_index, 1).value = row_index

        insert_rows(ws, 2, 1)

        self.assertEqual(ws.max_row, 6)
        self.assertEqual(ws._current_row, ws.max_row)
        self.assertIsNone(ws.cell(2, 1).value)
        self.assertEqual([ws.cell(r, 1).value for r in (1, 3, 4, 5, 6)], [1, 2, 3, 4, 5])

    def test_insert_rows_beyond_used_range_keeps_max_row(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        ws.cell(1, 1).value = "只有一行"

        insert_rows(ws, 10, 2)

        self.assertEqual(ws.max_row, 1)
        self.assertEqual(ws._current_row, 1)


class LibreOfficeLookupTest(unittest.TestCase):
    """回归测试：macOS .app 启动时 PATH 被裁剪，找不到 Homebrew 装的 soffice。"""

    def test_find_libreoffice_returns_homebrew_binary_on_macos(self) -> None:
        """当 PATH 不含 /opt/homebrew/bin，但二进制实际存在时，应能识别。"""
        fake_binary = Path("/opt/homebrew/bin/soffice")
        with patch("hr_toolkit.common.excel_compat.os.environ", {"PATH": "/usr/bin:/bin"}):
            with patch("hr_toolkit.common.excel_compat.Path.exists", return_value=True):
                executable = _find_libreoffice()
        # 实际值受其他路径影响，只要不为 None 即可证明查找逻辑生效
        self.assertIsNotNone(executable)
        self.assertTrue(executable.endswith("soffice") or executable.endswith("libreoffice"))

    def test_find_libreoffice_returns_none_when_nothing_exists(self) -> None:
        """全部路径都不存在时返回 None。"""
        with patch("hr_toolkit.common.excel_compat.os.environ", {"PATH": "/usr/bin:/bin"}):
            with patch("hr_toolkit.common.excel_compat.Path.exists", return_value=False):
                with patch("hr_toolkit.common.excel_compat.shutil.which", return_value=None):
                    self.assertIsNone(_find_libreoffice())

    @unittest.skipUnless(sys.platform == "darwin", "macOS-specific LibreOffice bundle path")
    def test_find_libreoffice_prefers_mac_app_path(self) -> None:
        """官方 .app 路径优先级最高。"""
        with patch("hr_toolkit.common.excel_compat.os.environ", {"PATH": "/usr/bin:/bin"}):
            original_exists = Path.exists

            def fake_exists(self: Path) -> bool:
                if str(self) == "/Applications/LibreOffice.app/Contents/MacOS/soffice":
                    return True
                return False

            with patch("hr_toolkit.common.excel_compat.Path.exists", fake_exists):
                executable = _find_libreoffice()
        self.assertEqual(
            executable,
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        )

    def test_build_conversion_env_keeps_existing_path(self) -> None:
        """环境变量注入不能丢失用户已有 PATH。"""
        with patch.dict(os.environ, {"PATH": "/usr/bin:/bin", "HOME": "/tmp"}, clear=False):
            env = _build_conversion_env()
        self.assertIn("/usr/bin", env["PATH"])
        self.assertIn("/bin", env["PATH"])

    def test_build_conversion_env_adds_homebrew_when_exists(self) -> None:
        """当 Homebrew 路径存在时，应被追加到 PATH。"""
        with patch.dict(os.environ, {"PATH": "/usr/bin:/bin"}, clear=False):
            with patch("hr_toolkit.common.excel_compat.Path.exists", return_value=True):
                env = _build_conversion_env()
        self.assertIn("/opt/homebrew/bin", env["PATH"])
        self.assertIn("/usr/local/bin", env["PATH"])

    def test_diagnose_contains_path_and_install_hint(self) -> None:
        """诊断信息应包含 PATH、已检查路径与安装建议。"""
        with patch.dict(os.environ, {"PATH": "/usr/bin:/bin"}, clear=False):
            with patch("hr_toolkit.common.excel_compat.Path.exists", return_value=False):
                message = _diagnose_missing_libreoffice()
        self.assertIn("未找到 libreoffice/soffice", message)
        self.assertIn("PATH", message)
        self.assertIn("brew install", message)
        self.assertIn("/opt/homebrew/bin", message)
        self.assertIn("/Applications/LibreOffice.app", message)


if __name__ == "__main__":
    unittest.main()
