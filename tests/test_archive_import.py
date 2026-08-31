from __future__ import annotations

import hashlib
import tempfile
import unittest
import zipfile
import shutil
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from hr_toolkit.tools.archive_import import (
    _find_footer_start_row,
    export_company_archive_tables,
    import_archive_transfers,
)


class ArchiveImportTest(unittest.TestCase):
    def test_footer_scan_reads_worksheet_dimensions_once(self) -> None:
        class CountingSheet:
            title = "性能公司"

            def __init__(self) -> None:
                self.max_row_reads = 0
                self.max_column_reads = 0

            @property
            def max_row(self) -> int:
                self.max_row_reads += 1
                return 6003

            @property
            def max_column(self) -> int:
                self.max_column_reads += 1
                return 37

            @staticmethod
            def cell(row_index: int, col_index: int):
                value = "对应行" if row_index == 6003 and col_index == 1 else None
                return type("Cell", (), {"value": value})()

        ws = CountingSheet()

        self.assertEqual(_find_footer_start_row(ws, 4), 6003)
        self.assertEqual(ws.max_row_reads, 1)
        self.assertEqual(ws.max_column_reads, 1)

    def test_import_archive_transfers_by_company_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            target = root / "档案汇总.xlsx"
            _write_transfer_file(input_dir / "茂名项目部人事档案移交表.xlsx")
            _write_summary_file(target)
            _write_summary_file(input_dir / "档案汇总.xlsx")

            result = import_archive_transfers(input_dir, target, output_dir)

            self.assertEqual(result.source_record_count, 3)
            self.assertEqual(result.inserted_count, 2)
            self.assertEqual(result.updated_count, 1)
            self.assertTrue(any("不是档案移交表" in warning for warning in result.warnings))
            self.assertTrue(result.output_file and result.output_file.exists())

            wb = load_workbook(result.output_file, data_only=False)
            ws1 = wb["公司1"]
            self.assertEqual(ws1.cell(5, 2).value, "张三")
            self.assertEqual(ws1.cell(5, 1).value, "11")
            self.assertEqual(ws1.cell(5, 4).value, '=MIDB(C5,7,4)&"-"&MIDB(C5,11,2)&"-"&MIDB(C5,13,2)')
            self.assertEqual(ws1.cell(5, 9).value, '=A5&"-"&TEXT(G5,"00000000")&"-"&TEXT(J5,"00")&"-"&H5')
            self.assertEqual(ws1.cell(5, 12).value, "√")
            self.assertEqual(ws1.cell(5, 19).value, 4)
            self.assertIn("驾照复印件", ws1.cell(5, 30).value)
            self.assertIn("解除合同协议书", ws1.cell(5, 30).value)

            ws2 = wb["公司2"]
            self.assertEqual(ws2.cell(4, 2).value, "已存在")
            self.assertEqual(ws2.cell(4, 12).value, "√")
            self.assertEqual(ws2.cell(4, 19).value, 4)

            ws3 = wb["公司3"]
            self.assertEqual(ws3.cell(4, 2).value, "张五")
            self.assertIsNone(ws3.cell(4, 12).value)
            self.assertIsNone(ws3.cell(4, 19).value)

    def test_existing_materials_merge_using_template_field_semantics(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "人事档案移交表.xlsx"
            target = root / "档案汇总.xlsx"
            output_dir = root / "output"
            _write_transfer_file(source)
            _write_summary_file(target)

            source_wb = load_workbook(source)
            source_ws = source_wb["移交表"]
            source_ws.cell(2, 20).value = "离职证明\n（前司）"
            source_ws.cell(3, 20).value = "√"
            # 公司2：三项数量用于验证“旧值为空则写入”；存在性字段仍只表示已有。
            source_ws.cell(4, 9).value = "✅"
            source_ws.cell(4, 10).value = "✓"
            source_ws.cell(4, 12).value = 2
            source_ws.cell(4, 13).value = 2
            source_ws.cell(4, 14).value = "✔"
            source_ws.cell(4, 15).value = "✅"
            source_ws.cell(4, 16).value = "✔"
            source_ws.cell(4, 20).value = "✅"
            # 公司3：数量字段的空值、数字/勾选冲突、双勾选三种边界。
            source_ws.cell(5, 12).value = "✅"
            source_ws.cell(5, 13).value = "✅"
            source_wb.save(source)
            source_wb.close()

            target_wb = load_workbook(target)
            ws1 = target_wb["公司1"]
            ws1.cell(4, 1).value = "11"
            ws1.cell(4, 2).value = "张三"
            ws1.cell(4, 3).value = "4600271987030XXXXX"
            ws1.cell(4, 4).value = '=MIDB(C4,7,4)&"-"&MIDB(C4,11,2)&"-"&MIDB(C4,13,2)'
            ws1.cell(4, 6).value = "2020-01-01"
            ws1.cell(4, 6).comment = Comment("人工批注不得变化", "HR")
            ws1.cell(4, 12).value = "√"
            ws1.cell(4, 13).value = 1
            ws1.cell(4, 14).value = None
            ws1.cell(4, 19).value = 3
            ws1.cell(4, 20).value = 1
            ws1.cell(4, 21).value = 1
            ws1.cell(4, 22).value = 2
            ws1.cell(4, 25).value = 1
            ws1.cell(4, 27).value = None
            ws1.cell(4, 30).value = "驾照复印件：√；备注：补充说明"
            ws1.cell(4, 19).fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

            ws2 = target_wb["公司2"]
            ws2.cell(4, 1).value = "11"
            ws2.cell(4, 6).value = "2020-01-01"
            ws2.cell(4, 19).value = None
            ws2.cell(4, 20).value = "✅"
            ws2.cell(4, 21).value = None
            ws2.cell(4, 22).value = None
            ws2.cell(4, 25).value = None

            ws3 = target_wb.copy_worksheet(ws1)
            ws3.title = "公司3"
            ws3.cell(4, 2).value = "张五"
            ws3.cell(4, 3).value = "4409211994103XXXXX"
            ws3.cell(4, 19).value = 5
            ws3.cell(4, 22).value = "√"
            ws3.cell(4, 25).value = 5
            target_wb.save(target)
            target_wb.close()

            source_sha256 = _sha256(source)
            target_sha256 = _sha256(target)
            before_wb = load_workbook(target, data_only=False)

            result = import_archive_transfers(source, target, output_dir)

            self.assertEqual(_sha256(source), source_sha256)
            self.assertEqual(_sha256(target), target_sha256)
            self.assertEqual(result.source_record_count, 3)
            self.assertEqual(result.inserted_count, 0)
            self.assertEqual(result.updated_count, 3)
            self.assertEqual(result.skipped_count, 0)
            self.assertTrue(
                any("保密协议数量冲突" in warning and "已保留原值" in warning for warning in result.warnings)
            )

            after_wb = load_workbook(result.output_file, data_only=False)
            try:
                self.assertEqual(after_wb.sheetnames, before_wb.sheetnames)
                for sheet_name in before_wb.sheetnames:
                    before_ws = before_wb[sheet_name]
                    after_ws = after_wb[sheet_name]
                    self.assertEqual(before_ws.max_row, after_ws.max_row)
                    self.assertEqual(before_ws.max_column, after_ws.max_column)
                    self.assertEqual(
                        sorted(str(item) for item in before_ws.merged_cells.ranges),
                        sorted(str(item) for item in after_ws.merged_cells.ranges),
                    )

                out1 = after_wb["公司1"]
                self.assertEqual(out1.cell(4, 19).value, 7)
                self.assertEqual(out1.cell(4, 25).value, 3)
                self.assertEqual(out1.cell(4, 22).value, 4)
                self.assertEqual(out1.cell(4, 12).value, "√")
                self.assertEqual(out1.cell(4, 13).value, 1)
                self.assertEqual(out1.cell(4, 14).value, "√")
                self.assertEqual(out1.cell(4, 20).value, 1)
                self.assertEqual(out1.cell(4, 21).value, 1)
                self.assertEqual(out1.cell(4, 27).value, "√")
                self.assertEqual(
                    out1.cell(4, 30).value,
                    "驾照复印件：√；备注：补充说明；解除合同协议书：√",
                )
                self.assertEqual(
                    _visible_comment_text(out1.cell(4, 6).comment.text),
                    "人工批注不得变化\n\n"
                    "第一次入职：2020/1/1\n"
                    "第二次入职：2026/4/9",
                )

                out2 = after_wb["公司2"]
                self.assertEqual(out2.cell(4, 19).value, 4)
                self.assertEqual(out2.cell(4, 25).value, 2)
                self.assertEqual(out2.cell(4, 22).value, 2)
                self.assertEqual(out2.cell(4, 20).value, "✅")
                self.assertEqual(out2.cell(4, 21).value, "✅")
                self.assertEqual(
                    _visible_comment_text(out2.cell(4, 6).comment.text),
                    "第一次入职：2020/1/1\n"
                    "第二次入职：2026/4/14",
                )

                out3 = after_wb["公司3"]
                self.assertEqual(out3.cell(4, 19).value, 5)
                self.assertEqual(out3.cell(4, 25).value, 5)
                self.assertEqual(out3.cell(4, 22).value, "√")
                self.assertEqual(
                    _visible_comment_text(out3.cell(4, 6).comment.text),
                    "人工批注不得变化\n\n"
                    "第一次入职：2020/1/1\n"
                    "第二次入职：2026/3/13",
                )

                allowed_value_changes = {
                    ("公司1", "N4"),
                    ("公司1", "S4"),
                    ("公司1", "V4"),
                    ("公司1", "Y4"),
                    ("公司1", "AA4"),
                    ("公司1", "AD4"),
                    ("公司2", "L4"),
                    ("公司2", "M4"),
                    ("公司2", "N4"),
                    ("公司2", "S4"),
                    ("公司2", "U4"),
                    ("公司2", "V4"),
                    ("公司2", "Y4"),
                    ("公司2", "AA4"),
                }
                _assert_only_expected_cell_values_changed(
                    self,
                    before_wb,
                    after_wb,
                    allowed_value_changes,
                    allowed_comment_changes={
                        ("公司1", "F4"),
                        ("公司2", "F4"),
                        ("公司3", "F4"),
                    },
                )
            finally:
                after_wb.close()
                before_wb.close()

    def test_reemployment_dates_are_kept_in_sorted_comments_without_changing_values(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "重复入职离职移交表.xlsx"
            target = root / "档案汇总.xlsx"
            output_dir = root / "output"
            _write_history_transfer_file(
                source,
                entry_date=date(2026, 1, 1),
                exit_date=date(2026, 8, 31),
            )
            shutil.copyfile(
                Path(__file__).resolve().parents[1]
                / "hr_toolkit"
                / "templates"
                / "archive_summary_template.xlsx",
                target,
            )

            target_wb = load_workbook(target)
            target_ws = target_wb["模板"]
            target_ws.title = "公司1"
            headers = _header_columns(target_ws, 3)
            target_ws.cell(4, headers["姓名"]).value = "张三"
            target_ws.cell(4, headers["身份证"]).value = "4600271987030XXXXX"
            target_ws.cell(4, headers["入职时间"]).value = date(2024, 5, 2)
            target_ws.cell(4, headers["入职时间"]).comment = Comment("人工入职说明不得变化", "甲方HR")
            target_ws.cell(4, headers["离职时间"]).value = date(2024, 12, 31)
            target_ws.cell(4, headers["离职时间"]).comment = Comment("人工离职说明不得变化", "甲方HR")
            target_wb.save(target)
            target_wb.close()

            source_sha256 = _sha256(source)
            target_sha256 = _sha256(target)
            before_wb = load_workbook(target, data_only=False)

            result = import_archive_transfers(source, target, output_dir)

            self.assertEqual(_sha256(source), source_sha256)
            self.assertEqual(_sha256(target), target_sha256)
            self.assertEqual(result.updated_count, 1)
            self.assertEqual(result.skipped_count, 0)
            after_wb = load_workbook(result.output_file, data_only=False)
            try:
                after_ws = after_wb["公司1"]
                self.assertEqual(_date_only(after_ws.cell(4, headers["入职时间"]).value), date(2024, 5, 2))
                self.assertEqual(_date_only(after_ws.cell(4, headers["离职时间"]).value), date(2024, 12, 31))

                entry_comment = after_ws.cell(4, headers["入职时间"]).comment
                self.assertIsNotNone(entry_comment)
                self.assertEqual(entry_comment.author, "甲方HR")
                self.assertEqual(
                    _visible_comment_text(entry_comment.text),
                    "人工入职说明不得变化\n\n"
                    "第一次入职：2024/5/2\n"
                    "第二次入职：2026/1/1",
                )

                exit_comment = after_ws.cell(4, headers["离职时间"]).comment
                self.assertIsNotNone(exit_comment)
                self.assertEqual(exit_comment.author, "甲方HR")
                self.assertEqual(
                    _visible_comment_text(exit_comment.text),
                    "人工离职说明不得变化\n\n"
                    "第一次离职：2024/12/31\n"
                    "第二次离职：2026/8/31",
                )
                _assert_only_expected_cell_values_changed(
                    self,
                    before_wb,
                    after_wb,
                    set(),
                    allowed_comment_changes={("公司1", "F4"), ("公司1", "AK4")},
                )
            finally:
                after_wb.close()
                before_wb.close()

            # 再次导入更早日期：批注按日期排序，已有日期去重，主单元格仍保持原值。
            third_source = root / "第三次入职离职移交表.xlsx"
            _write_history_transfer_file(
                third_source,
                entry_date=date(2023, 3, 4),
                exit_date=date(2025, 6, 30),
            )
            third_result = import_archive_transfers(
                third_source,
                result.output_file,
                root / "third-output",
            )
            third_wb = load_workbook(third_result.output_file, data_only=False)
            try:
                third_ws = third_wb["公司1"]
                self.assertEqual(_date_only(third_ws.cell(4, headers["入职时间"]).value), date(2024, 5, 2))
                self.assertEqual(_date_only(third_ws.cell(4, headers["离职时间"]).value), date(2024, 12, 31))
                self.assertEqual(
                    _visible_comment_text(third_ws.cell(4, headers["入职时间"]).comment.text),
                    "人工入职说明不得变化\n\n"
                    "第一次入职：2023/3/4\n"
                    "第二次入职：2024/5/2\n"
                    "第三次入职：2026/1/1",
                )
                self.assertEqual(
                    _visible_comment_text(third_ws.cell(4, headers["离职时间"]).comment.text),
                    "人工离职说明不得变化\n\n"
                    "第一次离职：2024/12/31\n"
                    "第二次离职：2025/6/30\n"
                    "第三次离职：2026/8/31",
                )
                third_entry_comment_text = third_ws.cell(4, headers["入职时间"]).comment.text
                third_exit_comment_text = third_ws.cell(4, headers["离职时间"]).comment.text
            finally:
                third_wb.close()

            duplicate_source = root / "重复日期移交表.xlsx"
            _write_history_transfer_file(
                duplicate_source,
                entry_date=date(2026, 1, 1),
                exit_date=date(2026, 8, 31),
            )
            duplicate_result = import_archive_transfers(
                duplicate_source,
                third_result.output_file,
                root / "duplicate-output",
            )
            self.assertEqual(duplicate_result.updated_count, 0)
            self.assertEqual(duplicate_result.skipped_count, 1)
            duplicate_wb = load_workbook(duplicate_result.output_file, data_only=False)
            try:
                duplicate_ws = duplicate_wb["公司1"]
                self.assertEqual(
                    duplicate_ws.cell(4, headers["入职时间"]).comment.text,
                    third_entry_comment_text,
                )
                self.assertEqual(
                    duplicate_ws.cell(4, headers["离职时间"]).comment.text,
                    third_exit_comment_text,
                )
            finally:
                duplicate_wb.close()

    def test_single_date_fills_empty_cell_without_creating_history_comment(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "首次入职离职移交表.xlsx"
            target = root / "档案汇总.xlsx"
            _write_history_transfer_file(
                source,
                entry_date=date(2026, 2, 3),
                exit_date=date(2026, 9, 4),
            )
            shutil.copyfile(
                Path(__file__).resolve().parents[1]
                / "hr_toolkit"
                / "templates"
                / "archive_summary_template.xlsx",
                target,
            )
            target_wb = load_workbook(target)
            target_ws = target_wb["模板"]
            target_ws.title = "公司1"
            headers = _header_columns(target_ws, 3)
            target_ws.cell(4, headers["姓名"]).value = "张三"
            target_ws.cell(4, headers["身份证"]).value = "4600271987030XXXXX"
            target_ws.cell(4, headers["入职时间"]).comment = Comment("人工说明", "甲方HR")
            target_wb.save(target)
            target_wb.close()

            result = import_archive_transfers(source, target, root / "output")

            output_wb = load_workbook(result.output_file, data_only=False)
            try:
                output_ws = output_wb["公司1"]
                self.assertEqual(_date_only(output_ws.cell(4, headers["入职时间"]).value), date(2026, 2, 3))
                self.assertEqual(_date_only(output_ws.cell(4, headers["离职时间"]).value), date(2026, 9, 4))
                self.assertEqual(output_ws.cell(4, headers["入职时间"]).comment.text, "人工说明")
                self.assertIsNone(output_ws.cell(4, headers["离职时间"]).comment)
            finally:
                output_wb.close()

    def test_exit_date_and_exit_time_headers_are_bidirectionally_compatible(self) -> None:
        cases = (
            ("离职日期", "离职时间"),
            ("离职时间", "离职日期"),
        )
        for source_header, target_header in cases:
            with self.subTest(source_header=source_header, target_header=target_header):
                with tempfile.TemporaryDirectory() as tmp:
                    root = Path(tmp)
                    source = root / "人事档案移交表.xlsx"
                    target = root / "档案汇总.xlsx"
                    _write_history_transfer_file(
                        source,
                        entry_date=date(2026, 1, 1),
                        exit_date=date(2026, 7, 31),
                        exit_date_header=source_header,
                    )
                    _write_summary_file(target, exit_date_header=target_header)
                    source_sha256 = _sha256(source)
                    target_sha256 = _sha256(target)

                    result = import_archive_transfers(source, target, root / "output")

                    self.assertEqual(_sha256(source), source_sha256)
                    self.assertEqual(_sha256(target), target_sha256)
                    self.assertFalse(any("已自动追加" in warning for warning in result.warnings))
                    output_wb = load_workbook(result.output_file, data_only=False)
                    try:
                        output_ws = output_wb["公司1"]
                        headers = _header_columns(output_ws, 3)
                        self.assertIn(target_header, headers)
                        self.assertNotIn(
                            "离职时间" if target_header == "离职日期" else "离职日期",
                            headers,
                        )
                        row_index = next(
                            row
                            for row in range(4, output_ws.max_row + 1)
                            if output_ws.cell(row, headers["身份证"]).value == "4600271987030XXXXX"
                        )
                        self.assertEqual(
                            _date_only(output_ws.cell(row_index, headers[target_header]).value),
                            date(2026, 7, 31),
                        )
                        other_text = str(output_ws.cell(row_index, headers["其他"]).value or "")
                        self.assertNotIn("离职时间：", other_text)
                        self.assertNotIn("离职日期：", other_text)
                    finally:
                        output_wb.close()

    def test_exit_date_alias_keeps_existing_value_and_writes_history_comment(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "人事档案移交表.xlsx"
            target = root / "档案汇总.xlsx"
            _write_history_transfer_file(
                source,
                entry_date=date(2026, 1, 1),
                exit_date=date(2026, 7, 31),
                exit_date_header="离职时间",
            )
            _write_summary_file(target, exit_date_header="离职日期")
            target_wb = load_workbook(target)
            target_ws = target_wb["公司1"]
            headers = _header_columns(target_ws, 3)
            target_ws.cell(4, headers["姓名"]).value = "张三"
            target_ws.cell(4, headers["身份证"]).value = "4600271987030XXXXX"
            target_ws.cell(4, headers["离职日期"]).value = date(2025, 12, 31)
            target_wb.save(target)
            target_wb.close()

            result = import_archive_transfers(source, target, root / "output")

            output_wb = load_workbook(result.output_file, data_only=False)
            try:
                output_ws = output_wb["公司1"]
                headers = _header_columns(output_ws, 3)
                exit_cell = output_ws.cell(4, headers["离职日期"])
                self.assertEqual(_date_only(exit_cell.value), date(2025, 12, 31))
                self.assertEqual(
                    _visible_comment_text(exit_cell.comment.text),
                    "第一次离职：2025/12/31\n第二次离职：2026/7/31",
                )
            finally:
                output_wb.close()

    def test_import_appends_exit_date_column_when_legacy_target_has_no_supported_header(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "人事档案移交表.xlsx"
            target = root / "旧版档案汇总.xlsx"
            _write_history_transfer_file(
                source,
                entry_date=date(2026, 1, 1),
                exit_date=date(2026, 7, 31),
                exit_date_header="离职日期",
            )
            _write_summary_file(target, exit_date_header=None)
            target_wb = load_workbook(target)
            target_ws = target_wb["公司1"]
            target_headers = _header_columns(target_ws, 3)
            target_ws.cell(4, target_headers["姓名"]).value = "张三"
            target_ws.cell(4, target_headers["身份证"]).value = "4600271987030XXXXX"
            target_ws.cell(4, target_headers["其他"]).value = "解除合同协议书：√"
            target_wb.save(target)
            target_wb.close()
            source_sha256 = _sha256(source)
            target_sha256 = _sha256(target)

            result = import_archive_transfers(source, target, root / "output")

            self.assertEqual(_sha256(source), source_sha256)
            self.assertEqual(_sha256(target), target_sha256)
            self.assertTrue(any("已自动追加“离职时间”列" in warning for warning in result.warnings))
            output_wb = load_workbook(result.output_file, data_only=False)
            try:
                for sheet_name in ("公司1", "公司2"):
                    headers = _header_columns(output_wb[sheet_name], 3)
                    self.assertEqual(headers["离职时间"], 37)
                    self.assertNotIn("离职日期", headers)

                output_ws = output_wb["公司1"]
                headers = _header_columns(output_ws, 3)
                row_index = next(
                    row
                    for row in range(4, output_ws.max_row + 1)
                    if output_ws.cell(row, headers["身份证"]).value == "4600271987030XXXXX"
                )
                exit_cell = output_ws.cell(row_index, headers["离职时间"])
                self.assertEqual(_date_only(exit_cell.value), date(2026, 7, 31))
                self.assertEqual(exit_cell.number_format, r"yyyy\-mm\-dd")
                other_text = str(output_ws.cell(row_index, headers["其他"]).value or "")
                self.assertEqual(other_text, "解除合同协议书：√")
            finally:
                output_wb.close()

    def test_export_appends_exit_date_column_to_legacy_company_archive(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "人事档案移交表.xlsx"
            _write_history_transfer_file(
                source,
                entry_date=date(2026, 1, 1),
                exit_date=date(2026, 7, 31),
                exit_date_header="离职日期",
            )
            imported = import_archive_transfers(source, None, root / "summary-output")
            existing = root / "公司1旧版档案表.xlsx"
            _write_existing_company_archive(existing, exit_date_header=None)
            summary_sha256 = _sha256(imported.output_file)
            existing_sha256 = _sha256(existing)

            result = export_company_archive_tables(
                imported.output_file,
                root / "export-output",
                existing_archive_path=existing,
            )

            self.assertEqual(_sha256(imported.output_file), summary_sha256)
            self.assertEqual(_sha256(existing), existing_sha256)
            self.assertTrue(any("已自动追加“离职时间”列" in warning for warning in result.warnings))
            output_wb = load_workbook(result.output_files[0], data_only=False)
            try:
                output_ws = output_wb["公司1"]
                headers = _header_columns(output_ws, 3)
                self.assertEqual(headers["离职时间"], 37)
                row_index = next(
                    row
                    for row in range(4, output_ws.max_row + 1)
                    if output_ws.cell(row, headers["身份证"]).value == "4600271987030XXXXX"
                )
                self.assertEqual(
                    _date_only(output_ws.cell(row_index, headers["离职时间"]).value),
                    date(2026, 7, 31),
                )
                other_text = str(output_ws.cell(row_index, headers["其他"]).value or "")
                self.assertNotIn("离职", other_text)
            finally:
                output_wb.close()

    def test_dry_run_does_not_write_output(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_file = root / "茂名项目部人事档案移交表.xlsx"
            target = root / "档案汇总.xlsx"
            output_dir = root / "output"
            _write_transfer_file(input_file)
            _write_summary_file(target)

            result = import_archive_transfers(input_file, target, output_dir, dry_run=True)

            self.assertEqual(result.source_record_count, 3)
            self.assertEqual(result.inserted_count, 3)
            self.assertIsNone(result.output_file)
            self.assertFalse(output_dir.exists())

    def test_duplicate_new_id_card_in_same_batch_is_skipped(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_file = root / "茂名项目部人事档案移交表.xlsx"
            target = root / "档案汇总.xlsx"
            output_dir = root / "output"
            _write_transfer_file(input_file, duplicate_first=True)
            _write_summary_file(target)

            result = import_archive_transfers(input_file, target, output_dir)

            self.assertEqual(result.source_record_count, 4)
            self.assertEqual(result.inserted_count, 2)
            self.assertEqual(result.updated_count, 1)
            self.assertEqual(result.skipped_count, 1)
            self.assertTrue(any("在本次导入中重复" in warning for warning in result.warnings))

            wb = load_workbook(result.output_file, data_only=False)
            ws = wb["公司1"]
            names = [ws.cell(row, 2).value for row in range(4, 8)]
            self.assertEqual(names.count("张三"), 1)
            wb.close()

    def test_import_from_zip_with_default_template(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "茂名项目部人事档案移交表.xlsx"
            archive = root / "移交表.zip"
            output_dir = root / "output"
            _write_transfer_file(source)
            with zipfile.ZipFile(archive, "w") as zip_file:
                zip_file.write(source, arcname=source.name)

            result = import_archive_transfers([archive], None, output_dir)

            self.assertTrue(result.using_default_template)
            self.assertEqual(result.source_record_count, 3)
            self.assertEqual(result.inserted_count, 3)
            self.assertTrue(result.output_file and result.output_file.exists())
            wb = load_workbook(result.output_file, data_only=False)
            self.assertNotIn("模板", wb.sheetnames)
            self.assertIn("公司1", wb.sheetnames)
            self.assertIn("公司2", wb.sheetnames)
            self.assertIn("公司3", wb.sheetnames)
            wb.close()

    def test_export_company_archive_tables_from_summary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_file = root / "茂名项目部人事档案移交表.xlsx"
            target = root / "档案汇总.xlsx"
            import_output = root / "import"
            export_output = root / "export"
            _write_transfer_file(input_file)
            _write_summary_file(target)
            imported = import_archive_transfers(input_file, target, import_output)

            result = export_company_archive_tables(imported.output_file, export_output)

            self.assertEqual(result.company_counts["公司1"], 2)
            self.assertEqual(result.company_counts["公司2"], 1)
            self.assertEqual(result.company_counts["公司3"], 1)
            self.assertEqual(len(result.output_files), 3)
            self.assertTrue((export_output / "公司1-档案表.xlsx").exists())
            wb = load_workbook(export_output / "公司3-档案表.xlsx", data_only=False)
            self.assertEqual(wb.sheetnames, ["公司3"])
            self.assertEqual(wb["公司3"].cell(4, 2).value, "张五")
            wb.close()

    def test_export_reports_progress_without_changing_result(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            summary = root / "档案汇总.xlsx"
            output_dir = root / "output"
            _write_summary_file(summary)
            progress: list[tuple[int, int, str]] = []

            result = export_company_archive_tables(
                summary,
                output_dir,
                progress_callback=lambda current, total, message: progress.append((current, total, message)),
            )

            self.assertEqual(len(result.output_files), 2)
            self.assertTrue(progress)
            self.assertEqual(progress[-1][0], progress[-1][1])
            self.assertTrue(any("保存" in message for _current, _total, message in progress))

    def test_export_cancellation_stops_before_writing_result(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            summary = root / "档案汇总.xlsx"
            output_dir = root / "output"
            _write_summary_file(summary)
            workbook = load_workbook(summary)
            workbook.remove(workbook["公司2"])
            workbook.save(summary)
            workbook.close()
            source_sha256 = _sha256(summary)
            cancel_requested = False

            def progress(current: int, total: int, message: str) -> None:
                nonlocal cancel_requested
                if total > 0 and current >= total and message.startswith("正在生成"):
                    cancel_requested = True

            with self.assertRaisesRegex(RuntimeError, "本次处理已停止"):
                export_company_archive_tables(
                    summary,
                    output_dir,
                    progress_callback=progress,
                    cancelled=lambda: cancel_requested,
                )

            self.assertEqual(_sha256(summary), source_sha256)
            self.assertFalse(output_dir.exists() and any(output_dir.glob("*.xlsx")))

    def test_export_appends_existing_archive_and_creates_missing_company_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_file = root / "茂名项目部人事档案移交表.xlsx"
            target = root / "档案汇总.xlsx"
            import_output = root / "import"
            existing_dir = root / "existing"
            export_output = root / "export"
            existing_dir.mkdir()
            _write_transfer_file(input_file)
            imported = import_archive_transfers(input_file, None, import_output)
            existing_company1 = existing_dir / "公司1档案表.xlsx"
            _write_existing_company_archive(existing_company1)

            result = export_company_archive_tables(imported.output_file, export_output, existing_archive_path=existing_dir)

            self.assertEqual(result.created_count, 2)
            self.assertEqual(result.inserted_count, 3)
            self.assertEqual(result.updated_count, 0)
            self.assertEqual(result.skipped_count, 0)
            self.assertTrue((export_output / "公司1-档案表.xlsx").exists())
            self.assertTrue((export_output / "公司2-档案表.xlsx").exists())
            self.assertTrue((export_output / "公司3-档案表.xlsx").exists())

            wb1 = load_workbook(export_output / "公司1-档案表.xlsx", data_only=False)
            ws1 = wb1["公司1"]
            self.assertEqual(ws1["A1"].value, "公司1人员档案编号表")
            self.assertEqual(ws1.cell(4, 2).value, "旧员工")
            self.assertEqual(ws1.cell(4, 3).value, "111111199901019999")
            self.assertEqual(ws1.cell(5, 2).value, "张三")
            self.assertEqual(ws1.cell(5, 1).value, "11")
            self.assertEqual(ws1.cell(5, 4).value, '=MIDB(C5,7,4)&"-"&MIDB(C5,11,2)&"-"&MIDB(C5,13,2)')
            self.assertEqual(ws1.cell(5, 9).value, '=A5&"-"&TEXT(G5,"00000000")&"-"&TEXT(J5,"00")&"-"&H5')
            self.assertEqual(ws1.cell(5, 2).alignment.horizontal, "center")
            self.assertEqual(ws1.cell(5, 2).border.left.style, "thin")
            self.assertFalse(ws1.cell(5, 2).font.bold)
            wb1.close()

            wb3 = load_workbook(export_output / "公司3-档案表.xlsx", data_only=False)
            ws3 = wb3["公司3"]
            self.assertEqual(ws3["A1"].value, "公司3人员档案编号表")
            self.assertEqual(ws3.cell(4, 2).value, "张五")
            self.assertEqual(ws3.cell(4, 1).value, "11")
            self.assertEqual(ws3.cell(4, 2).alignment.horizontal, "center")
            self.assertEqual(ws3.cell(4, 2).border.left.style, "thin")
            self.assertFalse(ws3.cell(4, 2).font.bold)
            wb3.close()

    def test_export_adds_quantity_materials_to_existing_employee(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "人事档案移交表.xlsx"
            summary_output = root / "summary"
            existing_dir = root / "existing"
            export_output = root / "export"
            existing_dir.mkdir()
            _write_transfer_file(source)
            source_wb = load_workbook(source)
            source_ws = source_wb["移交表"]
            source_ws.cell(2, 20).value = "离职时间"
            source_ws.cell(3, 20).value = date(2026, 7, 1)
            source_wb.save(source)
            source_wb.close()
            imported = import_archive_transfers(source, None, summary_output)

            existing = existing_dir / "公司1档案表.xlsx"
            _write_existing_company_archive(existing)
            existing_wb = load_workbook(existing)
            existing_ws = existing_wb["公司1"]
            headers = _header_columns(existing_ws, 3)
            existing_ws.cell(4, headers["姓名"]).value = "张三"
            existing_ws.cell(4, headers["身份证"]).value = "4600271987030XXXXX"
            existing_ws.cell(4, headers["劳动合同"]).value = 3
            existing_ws.cell(4, headers["保密协议"]).value = 1
            existing_ws.cell(4, headers["入职员工须知"]).value = 2
            existing_ws.cell(4, headers["照片"]).value = 1
            existing_ws.cell(4, headers["入职时间"]).value = date(2020, 1, 1)
            existing_entry_comment = Comment("旧档案入职说明", "甲方HR")
            existing_entry_comment.width = 260
            existing_entry_comment.height = 110
            existing_ws.cell(4, headers["入职时间"]).comment = existing_entry_comment
            existing_ws.cell(4, headers["离职时间"]).value = date(2021, 2, 3)
            existing_exit_comment = Comment("旧档案离职说明", "甲方HR")
            existing_exit_comment.width = 280
            existing_exit_comment.height = 120
            existing_ws.cell(4, headers["离职时间"]).comment = existing_exit_comment
            existing_wb.save(existing)
            existing_wb.close()
            saved_existing_wb = load_workbook(existing)
            saved_existing_ws = saved_existing_wb["公司1"]
            expected_entry_size = (
                saved_existing_ws.cell(4, headers["入职时间"]).comment.width,
                saved_existing_ws.cell(4, headers["入职时间"]).comment.height,
            )
            expected_exit_size = (
                saved_existing_ws.cell(4, headers["离职时间"]).comment.width,
                saved_existing_ws.cell(4, headers["离职时间"]).comment.height,
            )
            saved_existing_wb.close()
            summary_sha256 = _sha256(imported.output_file)
            existing_sha256 = _sha256(existing)

            result = export_company_archive_tables(
                imported.output_file,
                export_output,
                existing_archive_path=existing_dir,
            )

            self.assertEqual(_sha256(imported.output_file), summary_sha256)
            self.assertEqual(_sha256(existing), existing_sha256)
            self.assertEqual(result.updated_count, 1)
            output_wb = load_workbook(export_output / "公司1-档案表.xlsx", data_only=False)
            try:
                output_ws = output_wb["公司1"]
                output_headers = _header_columns(output_ws, 3)
                self.assertEqual(output_ws.cell(4, output_headers["劳动合同"]).value, 7)
                self.assertEqual(output_ws.cell(4, output_headers["保密协议"]).value, 3)
                self.assertEqual(output_ws.cell(4, output_headers["入职员工须知"]).value, 4)
                self.assertEqual(output_ws.cell(4, output_headers["照片"]).value, 1)
                self.assertEqual(
                    _date_only(output_ws.cell(4, output_headers["入职时间"]).value),
                    date(2020, 1, 1),
                )
                self.assertEqual(
                    _date_only(output_ws.cell(4, output_headers["离职时间"]).value),
                    date(2021, 2, 3),
                )
                entry_comment = output_ws.cell(4, output_headers["入职时间"]).comment
                self.assertEqual(entry_comment.author, "甲方HR")
                self.assertEqual((entry_comment.width, entry_comment.height), expected_entry_size)
                self.assertEqual(
                    _visible_comment_text(entry_comment.text),
                    "旧档案入职说明\n\n"
                    "第一次入职：2020/1/1\n"
                    "第二次入职：2026/4/9",
                )
                exit_comment = output_ws.cell(4, output_headers["离职时间"]).comment
                self.assertEqual(exit_comment.author, "甲方HR")
                self.assertEqual((exit_comment.width, exit_comment.height), expected_exit_size)
                self.assertEqual(
                    _visible_comment_text(exit_comment.text),
                    "旧档案离职说明\n\n"
                    "第一次离职：2021/2/3\n"
                    "第二次离职：2026/7/1",
                )
            finally:
                output_wb.close()


def _write_transfer_file(path: Path, *, duplicate_first: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "移交表"
    ws["A1"] = "茂名项目部人事档案移交表"
    headers = [
        "公司",
        "姓名",
        "身份证",
        "出生日期",
        "年齡",
        "入职时间",
        "入职公式",
        "出生年月公式",
        "电子照片",
        "入职登记表",
        "劳动合同",
        "保密协议",
        "入职须知",
        "员工三级安全教育",
        "身份证复印件",
        "银行卡复印件",
        "驾照复印件",
        "解除合同协议书",
        "备注",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(2, col).value = header
    rows = [
        ["公司1", "张三", "4600271987030XXXXX", None, None, "2026-04-09", None, None, "√", "√", 4, 2, 2, "√", "√", "√", "√", "√", "补充说明"],
        ["公司2", "已存在", "440921198009XXXXXX", None, None, "2026-04-14", None, None, None, "√", 4, None, None, None, None, None, None, None, None],
        ["公司3", "张五", "4409211994103XXXXX", None, None, "2026-03-13", None, None, None, None, None, None, None, None, None, None, None, None, None],
    ]
    if duplicate_first:
        rows.append(["公司1", "张三重复", "4600271987030XXXXX", None, None, "2026-04-09", None, None, None, None, None, None, None, None, None, None, None, None, None])
    for row_index, row in enumerate(rows, start=3):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    wb.save(path)


def _write_history_transfer_file(
    path: Path,
    *,
    entry_date: date,
    exit_date: date,
    exit_date_header: str = "离职时间",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "移交表"
    headers = ["公司", "姓名", "身份证", "入职时间", exit_date_header]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(2, col_index).value = header
    values = ["公司1", "张三", "4600271987030XXXXX", entry_date, exit_date]
    for col_index, value in enumerate(values, start=1):
        ws.cell(3, col_index).value = value
    wb.save(path)
    wb.close()


def _write_existing_company_archive(
    path: Path,
    *,
    exit_date_header: str | None = "离职时间",
) -> None:
    template = Path(__file__).resolve().parents[1] / "hr_toolkit" / "templates" / "archive_company_template.xlsx"
    shutil.copyfile(template, path)
    wb = load_workbook(path)
    ws = wb["公司1"]
    headers = _header_columns(ws, 3)
    exit_date_column = headers["离职时间"]
    if exit_date_header is None:
        ws.delete_cols(exit_date_column, 1)
    elif exit_date_header != "离职时间":
        ws.cell(3, exit_date_column).value = exit_date_header
    ws.cell(4, 1).value = "11"
    ws.cell(4, 2).value = "旧员工"
    ws.cell(4, 3).value = "111111199901019999"
    wb.save(path)
    wb.close()


def _write_summary_file(
    path: Path,
    *,
    exit_date_header: str | None = "离职时间",
) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "公司1"
    ws2 = wb.create_sheet("公司2")
    for ws in (ws1, ws2):
        ws["A1"] = "人事档案编号表"
        headers = [
            "编号",
            "姓名",
            "身份证",
            "出生日期",
            "年齡",
            "入职时间",
            "入职公式",
            "序号",
            "档案号",
            "出生年月公式",
            "档案柜号",
            "员工入职表",
            "身份证复印件",
            "银行卡复印件",
            "体检报告单",
            "学历证书",
            "学位证书",
            "相关资格证书",
            "劳动合同",
            "照片",
            "离职证明",
            "入职员工须知",
            "员工手册签收单",
            "安全生产责任书",
            "保密协议",
            "竞业协议",
            "三级安全教育登记（登记卡+试卷）",
            "员工健康情况调查表",
            "员工进场记录",
            "其他",
            "员工异动审批表",
            "入职考试试卷",
            "员工转正审批表",
            "转正考试试卷",
            "增购社保申请单",
            "离职申请单",
        ]
        if exit_date_header is not None:
            headers.append(exit_date_header)
        for col, header in enumerate(headers, start=1):
            ws.cell(3, col).value = header
        ws.cell(4, 2).value = "模板行"
        ws.cell(4, 3).value = "000000000000000000"
        ws.cell(5, 1).value = "对应行2的序号，如是抚州项目则标02"
    ws2.cell(4, 2).value = "已存在"
    ws2.cell(4, 3).value = "440921198009XXXXXX"
    ws2.cell(4, 12).value = None
    wb.save(path)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _header_columns(ws, header_row: int) -> dict[str, int]:
    return {
        str(ws.cell(header_row, col_index).value).replace("\n", "").replace(" ", ""): col_index
        for col_index in range(1, ws.max_column + 1)
        if ws.cell(header_row, col_index).value not in (None, "")
    }


def _visible_comment_text(value: str) -> str:
    return value.replace("\u2060", "").replace("\u2063", "")


def _date_only(value):
    return value.date() if hasattr(value, "date") else value


def _assert_only_expected_cell_values_changed(
    test_case: unittest.TestCase,
    before_wb,
    after_wb,
    allowed_value_changes: set[tuple[str, str]],
    *,
    allowed_comment_changes: set[tuple[str, str]] | None = None,
) -> None:
    allowed_comment_changes = allowed_comment_changes or set()
    actual_value_changes: set[tuple[str, str]] = set()
    actual_comment_changes: set[tuple[str, str]] = set()
    for sheet_name in before_wb.sheetnames:
        before_ws = before_wb[sheet_name]
        after_ws = after_wb[sheet_name]
        for row_index in range(1, before_ws.max_row + 1):
            for col_index in range(1, before_ws.max_column + 1):
                before_cell = before_ws.cell(row_index, col_index)
                after_cell = after_ws.cell(row_index, col_index)
                key = (sheet_name, before_cell.coordinate)
                if before_cell.value != after_cell.value:
                    actual_value_changes.add(key)
                if _comment_signature(before_cell.comment) != _comment_signature(after_cell.comment):
                    actual_comment_changes.add(key)
                if key not in allowed_value_changes:
                    test_case.assertEqual(before_cell.value, after_cell.value, key)
                    test_case.assertEqual(before_cell.data_type, after_cell.data_type, key)
                test_case.assertEqual(before_cell.style_id, after_cell.style_id, key)
                test_case.assertEqual(before_cell.number_format, after_cell.number_format, key)
                if key not in allowed_comment_changes:
                    test_case.assertEqual(
                        _comment_signature(before_cell.comment),
                        _comment_signature(after_cell.comment),
                        key,
                    )
                test_case.assertEqual(before_cell.hyperlink, after_cell.hyperlink, key)
    test_case.assertEqual(actual_value_changes, allowed_value_changes)
    test_case.assertEqual(actual_comment_changes, allowed_comment_changes)


def _comment_signature(comment):
    if comment is None:
        return None
    return (comment.text, comment.author, comment.width, comment.height)


if __name__ == "__main__":
    unittest.main()
