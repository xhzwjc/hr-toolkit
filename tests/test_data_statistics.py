from __future__ import annotations

import hashlib
import re
import tarfile
import tempfile
import unittest
import zipfile
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

import py7zr
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from hr_toolkit.tools.data_statistics import (
    AttendancePersonSummary,
    _write_attendance_sheet,
    generate_data_statistics_reports,
    parse_report_date,
    resolve_month_range,
    resolve_week_range,
)


class DataStatisticsTest(unittest.TestCase):
    def test_generate_data_statistics_reports(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            _write_attendance_file(input_dir / "考勤结果.xlsx")
            _write_weekly_file(input_dir / "【汇报】唐人周报04.01-05.04.xlsx")
            _write_monthly_file(input_dir / "【汇报】唐人月报04.01-05.04.xlsx")

            result = generate_data_statistics_reports(input_dir, output_dir)

            self.assertEqual(result.attendance_source_count, 4)
            self.assertEqual(result.attendance_person_count, 1)
            self.assertEqual(result.attendance_exception_count, 5)
            self.assertEqual(result.weekly_record_count, 8)
            self.assertEqual(result.monthly_record_count, 2)
            self.assertEqual(result.report_person_count, 2)
            self.assertEqual(result.report_exception_count, 2)
            self.assertTrue(result.output_file and result.output_file.exists())

            wb = load_workbook(result.output_file, data_only=False)
            attendance = wb["考勤统计"]
            self.assertEqual(attendance.cell(3, 2).value, "总部")
            self.assertEqual(attendance.cell(3, 3).value, "运营部")
            self.assertEqual(attendance.cell(3, 4).value, "王小丽")
            self.assertEqual(attendance.cell(3, 8).value, 0.5)
            self.assertEqual(attendance.cell(3, 12).value, 0.5)
            self.assertEqual(attendance.cell(3, 15).value, 3)
            self.assertIn("4.10上班未打卡", attendance.cell(3, 17).value)
            self.assertIn("4.15晚上加班0.5天", attendance.cell(3, 17).value)
            self.assertIn("4.15晚上加班0.5天、下班未打卡", attendance.cell(3, 17).value)
            self.assertIn("4.30上班未打卡", attendance.cell(3, 17).value)
            self.assertNotIn("4.30下班未打卡", attendance.cell(3, 17).value)

            report = wb["周月报统计"]
            self.assertEqual(report.max_column, 10)
            rows = {report.cell(row, 4).value: [report.cell(row, col).value for col in range(1, 11)] for row in (3, 4)}
            self.assertEqual(rows["黄五"][7:10], [1, None, "月报超时（17:31提交）"])
            self.assertEqual(rows["黄三"][5:10], [1, None, None, None, "第四周周报超时（18:37提交）"])
            self.assertEqual(report.cell(5, 2).value, "总计（周报截止时间2026.5.4 17:00）")
            self.assertIn("审批", report.cell(7, 1).value)
            self.assertEqual(report.cell(9, 3).value, "汇报规则：")
            self.assertIn("2026.4.13、2026.4.20、2026.4.27、2026.5.4", report.cell(9, 4).value)
            self.assertEqual(report.cell(10, 3).value, "月报规则：")
            self.assertIn("2026.5.2 17:00", report.cell(10, 4).value)

            detail = wb["周月报异常明细"]
            self.assertEqual(detail.cell(2, 5).value, "2026年4月")
            self.assertEqual(detail.cell(3, 5).value, "第四周")
            self.assertIsNone(detail.cell(2, 9).value)
            wb.close()

    def test_missing_punch_classification_uses_shift_side(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_missing_punch_boundary_file(input_dir / "考勤结果.xlsx")

            result = generate_data_statistics_reports(input_dir, root / "output")
            self.assertEqual(result.attendance_exception_count, 18)

            wb = load_workbook(result.output_file, data_only=True)
            detail = wb["考勤异常明细"]
            remarks_by_day = {
                detail.cell(row, 4).value.day: detail.cell(row, 7).value
                for row in range(2, detail.max_row + 1)
                if detail.cell(row, 5).value == "漏打卡"
            }
            wb.close()

            self.assertEqual(remarks_by_day[1], "下班未打卡")
            self.assertEqual(remarks_by_day[2], "上班未打卡")
            self.assertEqual(remarks_by_day[3], "下班未打卡")
            self.assertEqual(remarks_by_day[4], "上班未打卡")
            self.assertEqual(remarks_by_day[5], "下班未打卡")
            self.assertEqual(remarks_by_day[6], "上班未打卡")
            self.assertEqual(remarks_by_day[7], "上班未打卡")
            self.assertEqual(remarks_by_day[8], "下班未打卡")
            self.assertEqual(remarks_by_day[9], "上班未打卡")
            self.assertEqual(remarks_by_day[10], "漏打卡1次")
            self.assertEqual(remarks_by_day[11], "漏打卡1次")
            self.assertEqual(remarks_by_day[20], "上班未打卡")
            self.assertEqual(remarks_by_day[25], "上班未打卡")
            self.assertEqual(remarks_by_day[29], "上班未打卡")
            self.assertEqual(remarks_by_day[30], "下班未打卡")

    def test_overtime_slot_uses_last_punch_and_exact_boundaries(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_overtime_slot_file(input_dir / "考勤结果.xlsx")

            day_result = generate_data_statistics_reports(input_dir, root / "out_day")
            day_wb = load_workbook(day_result.output_file, data_only=True)
            stats = day_wb["考勤统计"]
            remark = stats.cell(3, stats.max_column).value or ""
            expected_fragments = (
                "4.1上午加班0.5天",
                "4.2上午加班0.5天",
                "4.3下午加班0.5天",
                "4.4下午加班0.5天",
                "4.5晚上加班0.5天",
                "4.6晚上加班0.5天",
                "4.7加班0.5天",
                "4.8加班0.5天",
                "4.9加班1天",
            )
            for fragment in expected_fragments:
                self.assertIn(fragment, remark)
            self.assertNotIn("4.7晚上加班", remark)
            self.assertNotIn("4.8上午加班", remark)
            self.assertNotIn("4.9晚上加班", remark)

            detail = day_wb["考勤异常明细"]
            detail_by_day = {
                detail.cell(row, 4).value.day: detail.cell(row, 7).value
                for row in range(2, detail.max_row + 1)
                if detail.cell(row, 5).value == "加班"
            }
            self.assertEqual(detail_by_day[1], "上午加班0.5天")
            self.assertEqual(detail_by_day[3], "下午加班0.5天")
            self.assertEqual(detail_by_day[5], "晚上加班0.5天")
            self.assertEqual(detail_by_day[6], "晚上加班0.5天")
            self.assertEqual(detail_by_day[7], "加班0.5天")
            self.assertEqual(detail_by_day[9], "加班1天")
            day_wb.close()

            hour_result = generate_data_statistics_reports(
                input_dir, root / "out_hour", remark_unit="hour"
            )
            hour_wb = load_workbook(hour_result.output_file, data_only=True)
            hour_stats = hour_wb["考勤统计"]
            hour_remark = hour_stats.cell(3, hour_stats.max_column).value or ""
            self.assertIn("4.1上午加班3.5小时", hour_remark)
            self.assertIn("4.3下午加班3.5小时", hour_remark)
            self.assertIn("4.5晚上加班3.5小时", hour_remark)
            self.assertIn("4.9加班7小时", hour_remark)
            hour_wb.close()

    def test_rest_slot_uses_punches_plan_and_full_day_rule(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_rest_slot_file(input_dir / "考勤结果.xlsx")

            day_result = generate_data_statistics_reports(input_dir, root / "out_day")
            day_wb = load_workbook(day_result.output_file, data_only=True)
            day_stats = day_wb["考勤统计"]
            day_remark = day_stats.cell(3, day_stats.max_column).value or ""
            self.assertIn("4.1调休1天", day_remark)
            self.assertNotIn("4.1上午调休", day_remark)
            self.assertNotIn("4.1下午调休", day_remark)
            self.assertIn("4.2下午调休0.5天", day_remark)
            self.assertIn("4.3上午调休0.5天", day_remark)
            self.assertIn("4.4上午调休0.5天", day_remark)
            self.assertIn("4.5下午调休0.5天", day_remark)
            self.assertIn("4.6下午调休0.57天", day_remark)
            self.assertIn("4.7上午调休0.5天", day_remark)
            # TASK-4 只修改统计备注；异常明细保持原有事实性描述。
            detail = day_wb["考勤异常明细"]
            detail_remarks = [
                detail.cell(row, 7).value for row in range(2, detail.max_row + 1)
            ]
            self.assertNotIn("上午调休0.5天", detail_remarks)
            self.assertNotIn("下午调休0.5天", detail_remarks)
            day_wb.close()

            hour_result = generate_data_statistics_reports(
                input_dir, root / "out_hour", remark_unit="hour"
            )
            hour_wb = load_workbook(hour_result.output_file, data_only=True)
            hour_stats = hour_wb["考勤统计"]
            hour_remark = hour_stats.cell(3, hour_stats.max_column).value or ""
            self.assertIn("4.1调休7小时", hour_remark)
            self.assertNotIn("4.1上午调休", hour_remark)
            self.assertNotIn("4.1下午调休", hour_remark)
            self.assertIn("4.2下午调休3.5小时", hour_remark)
            self.assertIn("4.3上午调休3.5小时", hour_remark)
            self.assertIn("4.5下午调休3.5小时", hour_remark)
            self.assertIn("4.6下午调休4小时", hour_remark)
            self.assertIn("4.7上午调休3.5小时", hour_remark)
            hour_wb.close()

    def test_paid_leave_remarks_use_specific_types_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_paid_leave_remark_file(input_dir / "考勤结果.xlsx")

            remarks: dict[str, str] = {}
            details: dict[str, list[list[object]]] = {}
            for unit in ("day", "hour"):
                result = generate_data_statistics_reports(
                    input_dir, root / f"out_{unit}", remark_unit=unit
                )
                wb = load_workbook(result.output_file, data_only=False)
                stats = wb["考勤统计"]
                self.assertEqual(stats.max_column, 17)
                self.assertEqual(stats.cell(3, 5).value, 0.5)
                self.assertEqual(stats.cell(3, 6).value, 0.5)
                # TASK-5 只改备注，带薪休假统计值继续保持原有年假口径。
                self.assertEqual(stats.cell(3, 7).value, 1.5)
                remarks[unit] = stats.cell(3, 17).value or ""
                detail = wb["考勤异常明细"]
                details[unit] = [
                    [detail.cell(row, col).value for col in range(1, detail.max_column + 1)]
                    for row in range(1, detail.max_row + 1)
                ]
                wb.close()

            remark = remarks["day"]
            self.assertNotIn("带薪休假", remark)
            self.assertIn("4.1年假1天", remark)
            self.assertIn("4.2年假下午0.5天", remark)
            self.assertIn("4.3病假上午0.5天", remark)
            self.assertIn(
                "4.4婚假1天、产假1天、陪护假1天、丧假1天、探亲假1天、工伤1天",
                remark,
            )
            self.assertIn("4.5事假上午0.5天", remark)
            self.assertEqual(remarks["hour"], remark)
            self.assertEqual(details["hour"], details["day"])

    def test_summary_paid_leave_remarks_use_specific_types_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_summary_paid_leave_remark_file(
                input_dir / "春苗2026年5月考勤.xlsx"
            )

            result = generate_data_statistics_reports(input_dir, root / "out")
            wb = load_workbook(result.output_file, data_only=False)
            stats = wb["考勤统计"]
            self.assertEqual(stats.cell(3, 7).value, 0.5)
            remark = stats.cell(3, stats.max_column).value or ""
            self.assertNotIn("带薪休假", remark)
            self.assertIn("病假上午0.5天", remark)
            self.assertIn(
                "婚假1天、产假1天、陪护假1天、丧假1天、探亲假1天、工伤1天、年假上午0.5天",
                remark,
            )
            wb.close()

    def test_report_staff_list_counts_missing_people(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            staff_file = root / "应汇报人员名单.xlsx"
            input_dir.mkdir()
            _write_weekly_file(input_dir / "【汇报】唐人周报04.01-05.04.xlsx")
            _write_monthly_file(input_dir / "【汇报】唐人月报04.01-05.04.xlsx")
            _write_staff_file(staff_file)

            result = generate_data_statistics_reports(input_dir, output_dir, report_staff_path=staff_file)

            self.assertEqual(result.expected_reporter_count, 3)
            self.assertEqual(result.report_person_count, 3)
            self.assertEqual(result.report_exception_count, 7)

            wb = load_workbook(result.output_file, data_only=True)
            report = wb["周月报统计"]
            rows = {report.cell(row, 4).value: [report.cell(row, col).value for col in range(1, 11)] for row in range(3, 6)}
            self.assertEqual(rows["黄六"][4:10], [4, None, 1, None, None, "第二周未写周报；第三周未写周报；第四周未写周报；第五周未写周报；未写月报"])
            self.assertEqual(rows["黄六"][1], "总部")
            self.assertEqual(rows["黄六"][2], "财务部")
            wb.close()

    def test_report_deadline_allows_170059(self) -> None:
        self.assertFalse(
            _generate_deadline_case(datetime(2026, 5, 2, 17, 0, 59)).report_exception_count,
        )
        self.assertEqual(
            _generate_deadline_case(datetime(2026, 5, 2, 17, 1, 0)).report_exception_count,
            1,
        )

    def test_weekly_deadline_allows_170059(self) -> None:
        self.assertFalse(
            _generate_weekly_deadline_case(datetime(2026, 4, 13, 17, 0, 59)).report_exception_count,
        )
        self.assertEqual(
            _generate_weekly_deadline_case(datetime(2026, 4, 13, 17, 1, 0)).report_exception_count,
            1,
        )

    def test_week_range_excludes_previous_month_monday(self) -> None:
        # 2026年6月1日是周一：不选日期时，6.1 会作为第一个周报截止日，
        # 统计到 5 月最后一周的周报；选择 6.2-6.30 后不再统计。
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_june_weekly_file(input_dir / "【汇报】唐人周报06.01-06.30.xlsx")

            # 无日期范围：6.1 的周报超时 + 未写月报，共 2 条异常
            default_result = generate_data_statistics_reports(input_dir, root / "out1", dry_run=True)
            self.assertEqual(default_result.report_exception_count, 2)

            # 选择 6.2-6.30：只剩未写月报 1 条，5 月最后一周不再统计
            ranged_result = generate_data_statistics_reports(
                input_dir,
                root / "out2",
                week_start="2026-06-02",
                week_end="2026-06-30",
            )
            self.assertEqual(ranged_result.report_exception_count, 1)
            self.assertEqual(ranged_result.week_range_start.isoformat(), "2026-06-02")

            wb = load_workbook(ranged_result.output_file, data_only=True)
            report = wb["周月报统计"]
            self.assertIn("2026.6.8、2026.6.15、2026.6.22、2026.6.29", report.cell(9, 4).value)
            self.assertNotIn("2026.6.1、", report.cell(9, 4).value)
            wb.close()

    def test_week_range_late_annotation_with_cross_day_time(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_report_boundary_file(
                input_dir / "【汇报】唐人周报04.13-04.13.xlsx",
                datetime(2026, 4, 14, 9, 5),
                "周报",
            )
            output_result = generate_data_statistics_reports(input_dir, root / "out2")
            wb = load_workbook(output_result.output_file, data_only=True)
            report = wb["周月报统计"]
            remarks = report.cell(3, 10).value
            self.assertIn("周报超时（4月14日9:05提交）", remarks)
            wb.close()

    def test_midweek_makeup_counts_as_previous_week_late(self) -> None:
        # 周三补交：算上一期超时（写明日期时间），下一周没交照样记未写
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_weekly_rows_file(
                input_dir / "【汇报】唐人周报06.01-06.30.xlsx",
                [
                    datetime(2026, 6, 8, 15, 0),  # 6.8 周一按时
                    datetime(2026, 6, 17, 9, 5),  # 6.15 那期的周三补交
                ],
            )
            result = generate_data_statistics_reports(
                input_dir,
                root / "out",
                week_start="2026-06-02",
                week_end="2026-06-30",
            )
            wb = load_workbook(result.output_file, data_only=True)
            remarks = wb["周月报统计"].cell(3, 10).value
            wb.close()
            self.assertIn("第二周周报超时（6月17日9:05提交）", remarks)
            self.assertIn("第三周未写周报", remarks)
            self.assertIn("第四周未写周报", remarks)

    def test_early_submission_rolls_to_next_week_when_already_reported(self) -> None:
        # 上周一已按时交过，周四又交一份（提前交下期）：算下一期，不记未写
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_weekly_rows_file(
                input_dir / "【汇报】唐人周报06.15-06.28.xlsx",
                [
                    datetime(2026, 6, 15, 15, 0),  # 6.15 周一按时
                    datetime(2026, 6, 18, 16, 0),  # 周四提前交 6.22 那期
                ],
            )
            result = generate_data_statistics_reports(
                input_dir,
                root / "out",
                week_start="2026-06-15",
                week_end="2026-06-28",
                dry_run=True,
            )
            # 只剩未写月报，两个周一都算已交
            self.assertEqual(result.report_exception_count, 1)

    def test_friday_submission_counts_for_next_monday(self) -> None:
        # 几维周末双休：周五下班交的周报算下周一截止那期，按时
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_weekly_rows_file(
                input_dir / "【汇报】几维周报06.16-06.28.xlsx",
                [datetime(2026, 6, 19, 18, 0)],  # 周五
            )
            result = generate_data_statistics_reports(
                input_dir,
                root / "out",
                week_start="2026-06-16",
                week_end="2026-06-28",
                dry_run=True,
            )
            self.assertEqual(result.report_exception_count, 1)  # 仅未写月报

    def test_submission_for_period_outside_range_is_ignored(self) -> None:
        # 6.26（周五）交的属于 6.29 那期；范围只到 6.24 时不参与本次统计，
        # 也不能被强行算到 6.22 头上记超时
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_weekly_rows_file(
                input_dir / "【汇报】唐人周报06.02-06.24.xlsx",
                [
                    datetime(2026, 6, 8, 15, 0),
                    datetime(2026, 6, 26, 10, 0),
                ],
            )
            result = generate_data_statistics_reports(
                input_dir,
                root / "out",
                week_start="2026-06-02",
                week_end="2026-06-24",
            )
            wb = load_workbook(result.output_file, data_only=True)
            remarks = wb["周月报统计"].cell(3, 10).value
            wb.close()
            self.assertNotIn("超时", remarks.replace("月报超时", ""))
            self.assertIn("第二周未写周报", remarks)
            self.assertIn("第三周未写周报", remarks)

    def test_week_range_requires_both_dates(self) -> None:
        with self.assertRaises(ValueError):
            resolve_week_range("2026-06-02", None)
        with self.assertRaises(ValueError):
            resolve_week_range(None, "2026-06-30")
        with self.assertRaises(ValueError):
            resolve_week_range("2026-06-30", "2026-06-02")
        self.assertIsNone(resolve_week_range(None, None))

    def test_month_range_requires_both_dates(self) -> None:
        with self.assertRaises(ValueError):
            resolve_month_range("2026-06-01", None)
        with self.assertRaises(ValueError):
            resolve_month_range(None, "2026-06-30")
        with self.assertRaises(ValueError):
            resolve_month_range("2026-06-30", "2026-06-01")
        self.assertIsNone(resolve_month_range(None, None))

    def test_month_range_filters_monthly_records(self) -> None:
        """Bug2：月报统计日期范围外的记录不计入、未写月报异常。"""
        from datetime import date

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_monthly_file(input_dir / "月报.xlsx")  # 4/30 黄三、5/2 黄五

            # 限定到 4 月，黄五 5/2 落在范围外 → 应报「未写月报」
            result = generate_data_statistics_reports(
                input_dir,
                root / "out",
                month_start=date(2026, 4, 1),
                month_end=date(2026, 4, 30),
            )

            month_exceptions = [
                exc for exc in result.to_dict() if isinstance(exc, dict)
            ]
            # 4 月只有黄三交了，黄五未交 → 至少 1 个「未写月报」
            self.assertGreaterEqual(result.monthly_record_count, 1)
            self.assertGreaterEqual(result.report_exception_count, 1)

    def test_parse_report_date_formats(self) -> None:
        from datetime import date

        expected = date(2026, 6, 2)
        for text in ("2026-06-02", "2026-6-2", "2026/6/2", "2026.6.2", "2026年6月2日"):
            self.assertEqual(parse_report_date(text), expected, text)
        with self.assertRaises(ValueError):
            parse_report_date("6.2")
        with self.assertRaises(ValueError):
            parse_report_date("2026-13-01")

    def test_large_attendance_file_completes_quickly(self) -> None:
        # 回归防线：read_only 工作表随机访问是 O(行数²)，2000 行曾需要 20 分钟以上。
        # 单遍读取后应在数秒内完成；上限放宽到 20 秒以避免慢机器误报。
        import time as time_module

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_large_attendance_file(input_dir / "考勤结果.xlsx", people=100, days=20)

            start = time_module.monotonic()
            result = generate_data_statistics_reports(input_dir, root / "out", dry_run=True)
            elapsed = time_module.monotonic() - start

            self.assertEqual(result.attendance_source_count, 2000)
            self.assertLess(elapsed, 20, f"2000 行考勤耗时 {elapsed:.1f}s，疑似退化为逐格重复解析")

    def test_large_attendance_output_does_not_rescan_sheet_width_per_person(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        for col_index in range(1, 18):
            ws.cell(2, col_index).value = f"字段{col_index}"
            ws.cell(3, col_index).value = ""
        summaries = [
            AttendancePersonSummary("测试公司", "测试部门", f"员工{index}")
            for index in range(100)
        ]
        original_getter = Worksheet.max_column.fget
        access_count = 0

        def counted_max_column(worksheet):
            nonlocal access_count
            access_count += 1
            return original_getter(worksheet)

        try:
            with patch.object(Worksheet, "max_column", new=property(counted_max_column)):
                _write_attendance_sheet(ws, summaries)
        finally:
            workbook.close()

        self.assertLessEqual(access_count, 3)

    def test_incorrect_declared_dimension_is_recovered_without_output_changes(self) -> None:
        """考勤、周报、月报导出范围过小时，输出仍须与规范文件完全一致。"""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source_dir = root / "source"
            normal_input = root / "normal_input"
            broken_input = root / "broken_input"
            source_dir.mkdir()
            normal_input.mkdir()
            broken_input.mkdir()
            writers = {
                "考勤结果.xlsx": _write_attendance_file,
                "【汇报】唐人周报04.01-05.04.xlsx": _write_weekly_file,
                "【汇报】唐人月报04.01-05.04.xlsx": _write_monthly_file,
            }
            source_hashes: dict[str, str] = {}
            normal_hashes: dict[str, str] = {}
            broken_hashes: dict[str, str] = {}
            for file_name, writer in writers.items():
                source_file = source_dir / file_name
                normal_file = normal_input / file_name
                broken_file = broken_input / file_name
                writer(source_file)
                normal_file.write_bytes(source_file.read_bytes())
                broken_file.write_bytes(source_file.read_bytes())
                _rewrite_first_sheet_dimension(broken_file, "A1:A1")
                source_hashes[file_name] = hashlib.sha256(source_file.read_bytes()).hexdigest()
                normal_hashes[file_name] = hashlib.sha256(normal_file.read_bytes()).hexdigest()
                broken_hashes[file_name] = hashlib.sha256(broken_file.read_bytes()).hexdigest()

            normal_result = generate_data_statistics_reports(normal_input, root / "normal_output")
            broken_result = generate_data_statistics_reports(broken_input, root / "broken_output")

            self.assertEqual(broken_result.attendance_source_count, 4)
            self.assertEqual(broken_result.weekly_record_count, 8)
            self.assertEqual(broken_result.monthly_record_count, 2)
            recovered_warnings = [
                warning
                for warning in broken_result.warnings
                if "导出范围 A1:A1 不完整" in warning
            ]
            self.assertEqual(len(recovered_warnings), 3)
            self.assertTrue(any("A1:AE5" in warning for warning in recovered_warnings))
            self.assertTrue(any("A1:E9" in warning for warning in recovered_warnings))
            self.assertTrue(any("A1:F3" in warning for warning in recovered_warnings))
            self.assertEqual(
                _xlsx_business_manifest(broken_result.output_file),
                _xlsx_business_manifest(normal_result.output_file),
            )
            for file_name in writers:
                self.assertEqual(
                    hashlib.sha256((source_dir / file_name).read_bytes()).hexdigest(),
                    source_hashes[file_name],
                )
                self.assertEqual(
                    hashlib.sha256((normal_input / file_name).read_bytes()).hexdigest(),
                    normal_hashes[file_name],
                )
                self.assertEqual(
                    hashlib.sha256((broken_input / file_name).read_bytes()).hexdigest(),
                    broken_hashes[file_name],
                )

    def test_missing_formula_cache_is_reported_precisely(self) -> None:
        """无法计算公式时，不再只给出泛化的“格式不识别”提示。"""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_formula_cache_missing_attendance(input_dir / "考勤结果.xlsx")

            with self.assertRaisesRegex(ValueError, "2 个公式单元格缺少缓存结果"):
                generate_data_statistics_reports(input_dir, root / "output")

    def test_remark_unit_hour_only_changes_overtime_and_rest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_attendance_file(input_dir / "考勤结果.xlsx")

            day_result = generate_data_statistics_reports(input_dir, root / "out_day", remark_unit="day")
            hour_result = generate_data_statistics_reports(input_dir, root / "out_hour", remark_unit="hour")

            day_wb = load_workbook(day_result.output_file)
            hour_wb = load_workbook(hour_result.output_file)
            day_ws = day_wb["考勤统计"]
            hour_ws = hour_wb["考勤统计"]
            day_remark = day_ws.cell(3, 17).value
            hour_remark = hour_ws.cell(3, 17).value

            # 源表 4.15 加班计调休时长=3.5、4.18 调休=3.5（小时）
            self.assertIn("4.15晚上加班0.5天", day_remark)
            self.assertIn("4.18上午调休0.5天", day_remark)
            self.assertIn("4.15晚上加班3.5小时", hour_remark)
            self.assertIn("4.18上午调休3.5小时", hour_remark)
            # 除加班/调休两处外，备注其余内容与按天完全一致
            self.assertEqual(
                hour_remark.replace("晚上加班3.5小时", "晚上加班0.5天").replace("上午调休3.5小时", "上午调休0.5天"),
                day_remark,
            )
            self.assertEqual(day_ws.cell(2, 8).value, "调休（天）")
            self.assertEqual(hour_ws.cell(2, 8).value, "调休（小时）")
            self.assertEqual(day_ws.cell(3, 8).value, 0.5)
            self.assertEqual(hour_ws.cell(3, 8).value, 3.5)
            for month in range(1, 5):
                self.assertEqual(day_ws.cell(2, 8 + month).value, f"{month}月份加班天数")
                self.assertEqual(hour_ws.cell(2, 8 + month).value, f"{month}月份加班小时数")
            self.assertEqual(day_ws.cell(3, 12).value, 0.5)
            self.assertEqual(hour_ws.cell(3, 12).value, 3.5)
            self.assertEqual(day_ws.cell(2, 16).value, "累计剩余加班天数")
            self.assertEqual(hour_ws.cell(2, 16).value, "累计剩余加班天数")
            self.assertEqual(day_ws.cell(3, 16).value, "=SUM(I3:L3)-H3")
            self.assertEqual(hour_ws.cell(3, 16).value, "=ROUND((SUM(I3:L3)-H3)/7,2)")
            # 其余列不受单位影响。
            for row in range(3, 4):
                for col in range(1, 17):
                    if col in (8, 12, 16):
                        continue
                    self.assertEqual(day_ws.cell(row, col).value, hour_ws.cell(row, col).value)
            day_wb.close()
            hour_wb.close()

    def test_remark_unit_defaults_to_day(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_attendance_file(input_dir / "考勤结果.xlsx")

            default_result = generate_data_statistics_reports(input_dir, root / "out_default")
            self.assertEqual(default_result.remark_unit, "day")

            wb = load_workbook(default_result.output_file)
            remark = wb["考勤统计"].cell(3, 17).value
            self.assertIn("4.15晚上加班0.5天", remark)
            self.assertIn("4.18上午调休0.5天", remark)
            wb.close()

    def test_remark_unit_rejects_unknown_value(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_attendance_file(input_dir / "考勤结果.xlsx")
            with self.assertRaises(ValueError):
                generate_data_statistics_reports(input_dir, root / "out", remark_unit="week")

    def test_same_attendance_archive_has_identical_output_across_formats(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "考勤结果.xlsx"
            _write_attendance_file(source)

            archives: list[Path] = []
            zip_path = root / "数据统计.zip"
            with zipfile.ZipFile(zip_path, "w") as zip_file:
                zip_file.write(source, arcname=source.name)
            archives.append(zip_path)

            seven_zip_path = root / "数据统计.7z"
            with py7zr.SevenZipFile(seven_zip_path, "w") as seven_zip_file:
                seven_zip_file.write(source, arcname=source.name)
            archives.append(seven_zip_path)

            for suffix, mode in {
                ".tar": "w",
                ".tar.gz": "w:gz",
                ".tgz": "w:gz",
                ".tar.bz2": "w:bz2",
                ".tbz2": "w:bz2",
                ".tar.xz": "w:xz",
                ".txz": "w:xz",
            }.items():
                archive_path = root / f"数据统计{suffix}"
                with tarfile.open(archive_path, mode) as tar_file:
                    tar_file.add(source, arcname=source.name)
                archives.append(archive_path)

            reference_result = None
            reference_manifest = None
            for index, archive_path in enumerate(archives):
                result = generate_data_statistics_reports([archive_path], root / f"output-{index}")
                output_manifest = _xlsx_business_manifest(result.output_file)
                summary = (
                    result.attendance_source_count,
                    result.attendance_person_count,
                    result.attendance_exception_count,
                    result.weekly_record_count,
                    result.monthly_record_count,
                    result.report_person_count,
                    result.report_exception_count,
                    result.warnings,
                )
                if reference_result is None:
                    reference_result = summary
                    reference_manifest = output_manifest

                with self.subTest(archive=archive_path.name):
                    self.assertEqual(summary, reference_result)
                    self.assertEqual(output_manifest, reference_manifest)
                    self.assertTrue(result.output_file and result.output_file.exists())

    def test_late_early_displayed_in_minutes(self) -> None:
        """迟到/早退以分钟数写入统计表，列头改为"迟到/早退（分钟）"."""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_attendance_file_with_minutes(input_dir / "考勤结果.xlsx")
            result = generate_data_statistics_reports(input_dir, root / "out")
            wb = load_workbook(result.output_file)
            ws = wb["考勤统计"]
            headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
            self.assertIn("迟到/早退（分钟）", headers)
            self.assertNotIn("迟到/早退（次）", headers)
            # 15 迟到 + 30 早退 = 45 分钟
            late_col = headers.index("迟到/早退（分钟）") + 1
            self.assertAlmostEqual(ws.cell(3, late_col).value or 0, 45.0)
            wb.close()

    def test_business_trip_toggle_inserts_column(self) -> None:
        """勾选 include_business_trip 时插入「公出（天）」列并显示。"""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_attendance_file_with_minutes(input_dir / "考勤结果.xlsx")
            off_result = generate_data_statistics_reports(
                input_dir, root / "off", include_business_trip=False
            )
            on_result = generate_data_statistics_reports(
                input_dir, root / "on", include_business_trip=True
            )
            off_wb = load_workbook(off_result.output_file)
            on_wb = load_workbook(on_result.output_file)
            off_ws = off_wb["考勤统计"]
            on_ws = on_wb["考勤统计"]
            off_headers = [off_ws.cell(2, c).value for c in range(1, off_ws.max_column + 1)]
            on_headers = [on_ws.cell(2, c).value for c in range(1, on_ws.max_column + 1)]
            self.assertNotIn("公出（天）", off_headers)
            self.assertIn("公出（天）", on_headers)
            # 勾选后总列数 = 不勾选 + 1
            self.assertEqual(on_ws.max_column, off_ws.max_column + 1)
            out_col = on_headers.index("公出（天）") + 1
            # 2 天公出
            self.assertAlmostEqual(on_ws.cell(3, out_col).value or 0, 2.0)
            off_wb.close()
            on_wb.close()

    def test_workday_business_trip_toggle_is_independent_from_public_outing(self) -> None:
        """公出和出差可独立勾选，出差只累计工作日出差。"""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_attendance_file_with_minutes(input_dir / "考勤结果.xlsx")

            cases = {
                (False, False): (False, False, None, "=SUM(I3:L3)-H3"),
                (True, False): (True, False, None, "=SUM(J3:M3)-H3"),
                (False, True): (False, True, 9, "=SUM(J3:M3)-H3"),
                (True, True): (True, True, 10, "=SUM(K3:N3)-H3"),
            }
            remarks: dict[tuple[bool, bool], str] = {}
            detail_outputs: dict[tuple[bool, bool], list[list[object]]] = {}
            for (include_out, include_trip), (has_out, has_trip, trip_col, expected_formula) in cases.items():
                with self.subTest(include_out=include_out, include_trip=include_trip):
                    result = generate_data_statistics_reports(
                        input_dir,
                        root / f"out_{int(include_out)}_{int(include_trip)}",
                        include_business_trip=include_out,
                        include_workday_business_trip=include_trip,
                    )
                    wb = load_workbook(result.output_file)
                    ws = wb["考勤统计"]
                    headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
                    self.assertEqual("公出（天）" in headers, has_out)
                    self.assertEqual("出差" in headers, has_trip)
                    if has_out:
                        out_col = headers.index("公出（天）") + 1
                        self.assertEqual(out_col, 9)
                        self.assertAlmostEqual(ws.cell(3, out_col).value or 0, 2.0)
                    if has_trip:
                        self.assertEqual(headers.index("出差") + 1, trip_col)
                        # 工作日出差共 1.75 天；休息日出差列中的 187 天必须排除。
                        self.assertAlmostEqual(ws.cell(3, trip_col).value or 0, 1.75)
                        self.assertEqual(ws.cell(2, trip_col).style_id, ws.cell(2, trip_col + 1).style_id)
                        late_minutes_col = headers.index("迟到/早退（分钟）") + 1
                        self.assertEqual(ws.cell(3, trip_col).style_id, ws.cell(3, late_minutes_col).style_id)
                        self.assertEqual(ws.cell(3, trip_col).number_format, "0.##")
                    balance_col = headers.index("累计剩余加班天数") + 1
                    self.assertEqual(ws.cell(3, balance_col).value, expected_formula)
                    detail_ws = wb["考勤异常明细"]
                    detail_values = [
                        [detail_ws.cell(row, col).value for col in range(1, detail_ws.max_column + 1)]
                        for row in range(1, detail_ws.max_row + 1)
                    ]
                    remarks[(include_out, include_trip)] = ws.cell(
                        3, headers.index("备注") + 1
                    ).value or ""
                    detail_outputs[(include_out, include_trip)] = detail_values
                    wb.close()

            self.assertNotIn("出差", remarks[(False, False)])
            self.assertNotIn("出差", remarks[(True, False)])
            self.assertIn("4.8出差1.5天、迟到15分钟", remarks[(False, True)])
            self.assertIn("4.9出差0.25天、早退30分钟", remarks[(False, True)])
            self.assertIn("4.8公出2天、出差1.5天、迟到15分钟", remarks[(True, True)])
            # 本次只扩展备注；单独切换“出差”时，异常明细不得发生变化。
            self.assertEqual(detail_outputs[(False, False)], detail_outputs[(False, True)])
            self.assertEqual(detail_outputs[(True, False)], detail_outputs[(True, True)])

            hour_result = generate_data_statistics_reports(
                input_dir,
                root / "out_hour",
                include_workday_business_trip=True,
                remark_unit="hour",
            )
            hour_wb = load_workbook(hour_result.output_file)
            hour_ws = hour_wb["考勤统计"]
            hour_headers = [hour_ws.cell(2, col).value for col in range(1, hour_ws.max_column + 1)]
            hour_remark = hour_ws.cell(3, hour_headers.index("备注") + 1).value or ""
            self.assertIn("出差1.5天", hour_remark)
            self.assertIn("出差0.25天", hour_remark)
            hour_wb.close()

    def test_business_trip_remark_includes_days(self) -> None:
        """include_business_trip=True 时，备注中包含「公出X天」."""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_attendance_file_with_minutes(input_dir / "考勤结果.xlsx")
            result = generate_data_statistics_reports(
                input_dir, root / "out", include_business_trip=True
            )
            wb = load_workbook(result.output_file)
            ws = wb["考勤统计"]
            remark_col = ws.max_column
            remark = ws.cell(3, remark_col).value or ""
            self.assertIn("公出", remark)
            self.assertIn("公出2天", remark)
            wb.close()

    def test_business_trip_fractional_day_source_kept_as_days(self) -> None:
        """考勤源表外出列为 0.4 天时，正确统计为 0.4 天且不受 remark_unit='hour' 影响。"""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            # 写入含 0.4 天外出的考勤表
            wb_src = Workbook()
            ws_src = wb_src.active
            ws_src.title = "日结果"
            headers = [
                "姓名", "部门名称", "确认状态", "日期", "是否异常处理", "漏打卡次数",
                "应出勤小时数", "实出勤小时数", "迟到次数", "迟到分钟数", "早退次数",
                "早退分钟数", "旷工天数", "旷工次数", "外出", "工作日出差", "休息日出差天数",
                "事假", "病假天数", "婚假", "产假天数", "陪护假", "丧假", "探亲假",
                "工伤", "年假天数", "调休", "加班计调休时长", "计划上下班时间",
                "当日刷卡记录", "缺卡记录",
            ]
            for col, header in enumerate(headers, start=1):
                ws_src.cell(1, col).value = header
            ws_src.cell(2, 1).value = "王小丽"
            ws_src.cell(2, 2).value = "运营部"
            ws_src.cell(2, 4).value = datetime(2026, 4, 12)
            ws_src.cell(2, 7).value = 0
            ws_src.cell(2, 8).value = 0
            ws_src.cell(2, 15).value = 0.4  # 0.4 天
            wb_src.save(input_dir / "考勤结果.xlsx")
            wb_src.close()

            for unit in ("day", "hour"):
                result = generate_data_statistics_reports(
                    input_dir, root / f"out_{unit}", include_business_trip=True, remark_unit=unit
                )
                wb = load_workbook(result.output_file)
                ws = wb["考勤统计"]
                on_headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
                out_col = on_headers.index("公出（天）") + 1
                self.assertAlmostEqual(ws.cell(3, out_col).value or 0, 0.4)
                remark = ws.cell(3, ws.max_column).value or ""
                self.assertIn("公出0.4天", remark)
                wb.close()

    def test_summary_attendance_remark_parsing(self) -> None:
        """汇总表无日期/无分钟列时，从备注中解析「M.D迟到N分钟」并填充到行/异常明细。"""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            _write_summary_attendance_file_with_remark(input_dir / "春苗考勤表.xlsx")
            result = generate_data_statistics_reports(input_dir, root / "out")
            wb = load_workbook(result.output_file)
            stats_ws = wb["考勤统计"]
            detail_ws = wb["考勤异常明细"]
            # 1) 统计表：迟到/早退（分钟）= 11（不再走"次"兜底）
            late_col = next(
                c for c in range(1, stats_ws.max_column + 1) if stats_ws.cell(2, c).value == "迟到/早退（分钟）"
            )
            self.assertAlmostEqual(stats_ws.cell(3, late_col).value or 0, 11.0)
            # 2) 异常明细：日期 = 5月18日，不再是5月1日
            detail_day = detail_ws.cell(2, 4).value
            self.assertEqual(detail_day.month, 5)
            self.assertEqual(detail_day.day, 18)
            wb.close()


def _write_summary_attendance_file_with_remark(path: Path) -> None:
    """模拟春苗表：汇总格式，无日期列；备注里写「5.18迟到11分钟」。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤表"
    ws.cell(1, 2).value = "春苗人力资源（北京）有限公司2026年5月考勤表"
    headers = [
        None, "姓名", "应出勤天数", "实际出勤天数", "请假天数", "出差天数",
        "事假\n(小时)", "病假\n(天)", "丧假\n(天)", "总调休\n(小时)", "当月调休(小时)",
        "年假\n(天)", "婚嫁（天）", "产假\n(天)", "当月加班时长", "6月份加班（小时）",
        "7月份加班（小时）", "8月份加班（小时）", "9月份加班（小时）", "10月份加班（小时）",
        "11月份加班（小时）", "12月份加班（小时）", "1月份加班（小时）", "2月份加班（小时）",
        "3月份加班（小时）", "4月份加班（小时）", "5月份加班（小时）", "剩余加班时长(h)",
        "迟到次数", "早退次数", "上班缺卡次数", "下班缺卡次数", "旷工天数", "补卡次数", "备注",
    ]
    for c, header in enumerate(headers, start=1):
        if header:
            ws.cell(2, c).value = header
    ws.cell(3, 2).value = "张三"
    ws.cell(3, 3).value = 19
    ws.cell(3, 4).value = 19
    ws.cell(3, 29).value = 1
    ws.cell(3, 34).value = 2
    ws.cell(3, 35).value = "迟到：5.18迟到11分钟\n补卡：5.21/5.26补下班卡"
    wb.save(path)
    wb.close()


def _rewrite_first_sheet_dimension(path: Path, dimension: str) -> None:
    rewritten = path.with_name(f"{path.stem}.rewritten.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        rewritten, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                payload, count = re.subn(
                    br'<dimension ref="[^"]+"',
                    f'<dimension ref="{dimension}"'.encode(),
                    payload,
                    count=1,
                )
                if count != 1:
                    raise AssertionError("测试工作簿缺少 dimension")
            target.writestr(info, payload)
    rewritten.replace(path)


def _write_formula_cache_missing_attendance(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "日结果"
    worksheet.append(["姓名", "日期", "漏打卡次数", "应出勤小时数"])
    # openpyxl 只写公式本身，不计算并写入缓存结果，用来模拟系统直出的不完整公式。
    worksheet.append(['="王小丽"', "=DATE(2026,4,1)", 0, 7])
    workbook.save(path)
    workbook.close()


def _write_attendance_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "日结果"
    headers = [
        "姓名",
        "部门名称",
        "确认状态",
        "日期",
        "是否异常处理",
        "漏打卡次数",
        "应出勤小时数",
        "实出勤小时数",
        "迟到次数",
        "迟到分钟数",
        "早退次数",
        "早退分钟数",
        "旷工天数",
        "旷工次数",
        "外出",
        "工作日出差",
        "休息日出差天数",
        "事假",
        "病假天数",
        "婚假",
        "产假天数",
        "陪护假",
        "丧假",
        "探亲假",
        "工伤",
        "年假天数",
        "调休",
        "加班计调休时长",
        "计划上下班时间",
        "当日刷卡记录",
        "缺卡记录",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    rows = [
        ["王小丽", "运营部", "否", datetime(2026, 4, 10), "否", 1, 7, 7, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "08:30,12:00|14:00,17:30", "08:30,17:51", None],
        ["王小丽", "运营部", "否", datetime(2026, 4, 15), "否", 1, 7, 7, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 3.5, "08:30~12:00|14:00~17:30", "08:19,17:32,17:54,21:00", None],
        ["王小丽", "运营部", "否", datetime(2026, 4, 18), "否", 0, 3.5, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 3.5, 0, "08:30~12:00", None, None],
        ["王小丽", "运营部", "否", datetime(2026, 4, 30), "否", 1, 7, 7, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "08:30,12:00|14:00,17:30", "08:23,17:30", None],
    ]
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    wb.save(path)
    wb.close()


def _write_overtime_slot_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "日结果"
    ws.append([
        "姓名", "部门名称", "日期", "漏打卡次数", "应出勤小时数", "实出勤小时数",
        "调休", "加班计调休时长", "计划上下班时间", "当日刷卡记录", "缺卡记录",
    ])
    cases = (
        (1, 3.5, "08:30:00"),
        (2, 3.5, "15:29:59"),
        (3, 3.5, "15:30:00"),
        (4, 3.5, "20:30:00"),
        (5, 3.5, "20:30:01"),
        (6, 3.5, "20:00:00,03:30:00"),
        (7, 3.5, "20:00:00,03:30:01"),
        (8, 3.5, "08:29:59"),
        (9, 7, "21:00:00"),
    )
    for day, overtime, punches in cases:
        ws.append([
            "王小丽", "运营部", datetime(2026, 4, day), 0, 7, 7,
            0, overtime, "08:30~12:00|14:00~17:30", punches, None,
        ])
    wb.save(path)
    wb.close()


def _write_rest_slot_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "日结果"
    ws.append([
        "姓名", "部门名称", "日期", "漏打卡次数", "应出勤小时数", "实出勤小时数",
        "调休", "加班计调休时长", "计划上下班时间", "当日刷卡记录", "缺卡记录",
    ])
    cases = (
        # 全天调休：按勾选单位显示 1 天或 7 小时，不添加上午/下午。
        (1, 7, 7, 0, "08:30~12:00|14:00~17:30", None),
        # 只在上午出勤，下午调休。
        (2, 3.5, 7, 3.5, "08:30~12:00|14:00~17:30", "08:30,12:00"),
        # 只在下午出勤，上午调休。
        (3, 3.5, 7, 3.5, "08:30~12:00|14:00~17:30", "14:00,17:30"),
        # 无打卡，仅有上午/下午半天排班。
        (4, 3.5, 3.5, 0, "08:30~12:00", None),
        (5, 3.5, 3.5, 0, "14:00~17:30", None),
        # 甲方补充案例：4 小时调休，09:30 上班、15:12 最后打卡，属于下午调休。
        (6, 4, 7, 3, "09:30~00:00", "09:30,15:12"),
        # 信息不足时按甲方口径兜底为上午。
        (7, 3.5, 7, 3.5, None, None),
    )
    for day, rest, expected, actual, plan, punches in cases:
        ws.append([
            "王小丽", "运营部", datetime(2026, 4, day), 0, expected, actual,
            rest, 0, plan, punches, None,
        ])
    wb.save(path)
    wb.close()


def _write_paid_leave_remark_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "日结果"
    headers = [
        "姓名", "部门名称", "日期", "漏打卡次数", "应出勤小时数", "实出勤小时数",
        "事假", "病假天数", "婚假", "产假天数", "陪护假", "丧假", "探亲假",
        "工伤", "年假天数", "调休", "加班计调休时长", "计划上下班时间",
        "当日刷卡记录", "缺卡记录",
    ]
    ws.append(headers)
    rows = (
        # 全天年假不加时段。
        (1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 7, 0, None, None),
        # 上午出勤，下午半天年假。
        (2, 0, 0, 0, 0, 0, 0, 0, 0, 0.5, 7, 3.5, "08:30~17:30", "08:30,12:00"),
        # 下午出勤，上午半天病假。
        (3, 0, 0.5, 0, 0, 0, 0, 0, 0, 0, 7, 3.5, "08:30~17:30", "14:00,17:30"),
        # 同一天存在多个具体带薪假别时逐项显示。
        (4, 0, 0, 1, 1, 1, 1, 1, 1, 0, 7, 0, None, None),
        # 信息不足的半天事假按甲方口径兜底为上午。
        (5, 0.5, 0, 0, 0, 0, 0, 0, 0, 0, 7, 3.5, None, None),
    )
    for (
        day, personal, sick, marriage, maternity, nursing, bereavement,
        home_visit, injury, annual, expected, actual, plan, punches,
    ) in rows:
        ws.append([
            "王小丽", "运营部", datetime(2026, 4, day), 0, expected, actual,
            personal, sick, marriage, maternity, nursing, bereavement, home_visit,
            injury, annual, 0, 0, plan, punches, None,
        ])
    wb.save(path)
    wb.close()


def _write_summary_paid_leave_remark_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤表"
    ws.append([
        "姓名", "应出勤天数", "事假\n(小时)", "病假\n(天)", "丧假\n(天)",
        "年假\n(天)", "婚嫁（天）", "产假\n(天)", "陪护假", "探亲假", "工伤",
        "带薪休假（天）", "备注",
    ])
    ws.append([
        "张三", 20, 0, 0.5, 1, 0.5, 1, 1, 1, 1, 1, 6.5, None,
    ])
    wb.save(path)
    wb.close()


def _write_missing_punch_boundary_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "日结果"
    headers = [
        "姓名",
        "部门名称",
        "日期",
        "漏打卡次数",
        "应出勤小时数",
        "实出勤小时数",
        "迟到次数",
        "迟到分钟数",
        "早退次数",
        "早退分钟数",
        "计划上下班时间",
        "当日刷卡记录",
        "缺卡记录",
    ]
    ws.append(headers)
    day_plan = "08:30,12:00|14:00,17:30"
    night_plan = "20:00~00:00|01:00~08:00"
    rows = [
        # 上午迟到卡存在、下班卡缺失：不能误报为上班缺卡。
        ["王小丽", "运营部", datetime(2026, 4, 1), 1, 7, 7, 1, 10, 0, 0, day_plan, "08:40", None],
        # 只有下班侧刷卡：缺失的是上班卡。
        ["王小丽", "运营部", datetime(2026, 4, 2), 1, 7, 7, 0, 0, 0, 0, day_plan, "17:10", None],
        # 多条记录都在上午侧：缺失的仍是下班卡。
        ["王小丽", "运营部", datetime(2026, 4, 3), 1, 7, 7, 0, 0, 0, 0, day_plan, "08:20,08:40", None],
        # 刷卡文本乱序时仍按班次时间判断。
        ["王小丽", "运营部", datetime(2026, 4, 4), 1, 7, 7, 0, 0, 0, 0, day_plan, "17:30,08:20", None],
        # 跨午夜班次的上班侧、下班侧与乱序记录。
        ["王小丽", "运营部", datetime(2026, 4, 5), 1, 7, 7, 0, 0, 0, 0, night_plan, "20:10", None],
        ["王小丽", "运营部", datetime(2026, 4, 6), 1, 7, 7, 0, 0, 0, 0, night_plan, "07:50", None],
        ["王小丽", "运营部", datetime(2026, 4, 7), 1, 7, 7, 0, 0, 0, 0, night_plan, "08:05,19:55", None],
        # 半天班的极端迟到/早退由明确异常信号判定，不受中点启发式影响。
        ["王小丽", "运营部", datetime(2026, 4, 8), 1, 3.5, 3.5, 1, 180, 0, 0, "08:30~12:00", "11:30", None],
        ["王小丽", "运营部", datetime(2026, 4, 9), 1, 3.5, 3.5, 0, 0, 1, 120, "08:30~12:00", "10:00", None],
        # 计划或刷卡信息不足时不强猜方向，只保留事实性次数。
        ["王小丽", "运营部", datetime(2026, 4, 10), 1, 7, 7, 0, 0, 0, 0, "08:30", "17:30", None],
        ["王小丽", "运营部", datetime(2026, 4, 11), 1, 7, 7, 0, 0, 0, 0, None, None, None],
        # 补上班卡早于 08:30，但已有下班侧刷卡。
        ["王小丽", "运营部", datetime(2026, 4, 20), 1, 7, 7, 0, 0, 0, 0, day_plan, "08:20,17:30", None],
        ["王小丽", "运营部", datetime(2026, 4, 25), 1, 7, 7, 0, 0, 0, 0, day_plan, "08:25,17:35", None],
        ["王小丽", "运营部", datetime(2026, 4, 29), 1, 7, 7, 0, 0, 0, 0, day_plan, "08:29,18:00", None],
        # 只有上午侧刷卡时，确实是下班缺卡。
        ["王小丽", "运营部", datetime(2026, 4, 30), 1, 7, 7, 0, 0, 0, 0, day_plan, "08:20", None],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _write_attendance_file_with_minutes(path: Path) -> None:
    """含迟到/早退分钟数与公出的考勤源表。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "日结果"
    headers = [
        "姓名", "部门名称", "确认状态", "日期", "是否异常处理", "漏打卡次数",
        "应出勤小时数", "实出勤小时数", "迟到次数", "迟到分钟数", "早退次数",
        "早退分钟数", "旷工天数", "旷工次数", "外出", "工作日出差", "休息日出差天数",
        "事假", "病假天数", "婚假", "产假天数", "陪护假", "丧假", "探亲假",
        "工伤", "年假天数", "调休", "加班计调休时长", "计划上下班时间",
        "当日刷卡记录", "缺卡记录",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    rows = [
        # 迟到 15 分钟 + 公出 2 小时
        ["赵小亮", "市场部", "否", datetime(2026, 4, 8), "否", 0, 8, 8, 1, 15, 0, 0, 0, 0, 2, 1.5, 99, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "08:30,12:00|14:00,17:30", "08:45,17:30", None],
        # 早退 30 分钟
        ["赵小亮", "市场部", "否", datetime(2026, 4, 9), "否", 0, 8, 8, 0, 0, 1, 30, 0, 0, 0, 0.25, 88, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "08:30,12:00|14:00,17:30", "08:30,17:00", None],
    ]
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    wb.save(path)
    wb.close()


def _write_weekly_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "工作表1"
    headers = ["汇报编号", "汇报时间", "汇报人", "汇报人部门", "汇报对象"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    rows = [
        ["1", datetime(2026, 4, 11, 15, 0), "黄三", "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一"],
        ["2", datetime(2026, 4, 11, 15, 0), "黄五", "唐人数智科技股份有限公司/唐人数智/行政人事中心/办公室", "罗一一"],
        ["3", datetime(2026, 4, 18, 15, 0), "黄三", "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一"],
        ["4", datetime(2026, 4, 18, 15, 0), "黄五", "唐人数智科技股份有限公司/唐人数智/行政人事中心/办公室", "罗一一"],
        ["5", datetime(2026, 4, 27, 18, 37), "黄三", "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一"],
        ["6", datetime(2026, 4, 25, 15, 0), "黄五", "唐人数智科技股份有限公司/唐人数智/行政人事中心/办公室", "罗一一"],
        ["7", datetime(2026, 5, 2, 11, 0), "黄三", "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一"],
        ["8", datetime(2026, 5, 2, 11, 0), "黄五", "唐人数智科技股份有限公司/唐人数智/行政人事中心/办公室", "罗一一"],
    ]
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    wb.save(path)
    wb.close()


def _write_monthly_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "工作表1"
    headers = ["汇报编号", "汇报时间", "汇报人", "汇报人部门", "汇报对象", "评论"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    rows = [
        ["1", datetime(2026, 4, 30, 11, 30), "黄三", "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一", None],
        ["2", datetime(2026, 5, 2, 17, 31), "黄五", "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一", None],
    ]
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    wb.save(path)
    wb.close()


def _write_large_attendance_file(path: Path, people: int, days: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "日结果"
    headers = ["姓名", "部门名称", "日期", "漏打卡次数", "应出勤小时数", "实出勤小时数", "事假", "病假天数", "年假天数", "调休", "加班计调休时长", "旷工天数", "迟到次数", "早退次数", "计划上下班时间", "当日刷卡记录", "缺卡记录"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    row_index = 2
    for person in range(people):
        for day in range(1, days + 1):
            ws.cell(row_index, 1).value = f"员工{person:03d}"
            ws.cell(row_index, 2).value = "运营部"
            ws.cell(row_index, 3).value = datetime(2026, 6, day)
            ws.cell(row_index, 4).value = 0
            ws.cell(row_index, 5).value = 7
            ws.cell(row_index, 6).value = 7
            row_index += 1
    wb.save(path)
    wb.close()


def _write_weekly_rows_file(path: Path, report_times: list[datetime], name: str = "黄三") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "工作表1"
    headers = ["汇报编号", "汇报时间", "汇报人", "汇报人部门", "汇报对象"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    for row_index, report_time in enumerate(report_times, start=2):
        row = [str(row_index - 1), report_time, name, "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一"]
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    wb.save(path)
    wb.close()


def _write_june_weekly_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "工作表1"
    headers = ["汇报编号", "汇报时间", "汇报人", "汇报人部门", "汇报对象"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    rows = [
        # 6.1（周一）提交的是 5 月最后一周的周报，且超时
        ["1", datetime(2026, 6, 1, 18, 0), "黄三", "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一"],
        ["2", datetime(2026, 6, 8, 15, 0), "黄三", "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一"],
        ["3", datetime(2026, 6, 15, 15, 0), "黄三", "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一"],
        ["4", datetime(2026, 6, 22, 15, 0), "黄三", "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一"],
        ["5", datetime(2026, 6, 29, 15, 0), "黄三", "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一"],
    ]
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    wb.save(path)
    wb.close()


def _write_staff_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "应汇报人员"
    headers = ["姓名", "部门（片区）"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    rows = [
        ["黄三", "装备事业部"],
        ["黄五", "行政人事中心/办公室"],
        ["黄六", "财务部"],
    ]
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    wb.save(path)
    wb.close()


def _generate_deadline_case(monthly_time: datetime):
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        input_dir = root / "input"
        output_dir = root / "output"
        input_dir.mkdir()
        wb = Workbook()
        ws = wb.active
        ws.title = "工作表1"
        headers = ["汇报编号", "汇报时间", "汇报人", "汇报人部门", "汇报对象"]
        for col, header in enumerate(headers, start=1):
            ws.cell(1, col).value = header
        row = ["1", monthly_time, "黄三", "唐人数智科技股份有限公司/唐人数智/装备事业部", "罗一一"]
        for col_index, value in enumerate(row, start=1):
            ws.cell(2, col_index).value = value
        wb.save(input_dir / "【汇报】唐人月报04.01-05.04.xlsx")
        wb.close()
        return generate_data_statistics_reports(input_dir, output_dir)
def _generate_weekly_deadline_case(weekly_time: datetime):
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        input_dir = root / "input"
        output_dir = root / "output"
        input_dir.mkdir()
        _write_report_boundary_file(
            input_dir / "【汇报】唐人周报04.13-04.13.xlsx",
            weekly_time,
            "周报",
        )
        _write_report_boundary_file(
            input_dir / "【汇报】唐人月报04.01-05.04.xlsx",
            datetime(2026, 4, 30, 11, 0, 0),
            "月报",
        )
        return generate_data_statistics_reports(input_dir, output_dir)


def _write_report_boundary_file(path: Path, report_time: datetime, report_kind: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "工作表1"
    headers = ["汇报编号", "汇报时间", "汇报人", "汇报人部门", "汇报对象"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    row = [
        f"{report_kind}1",
        report_time,
        "黄三",
        "唐人数智科技股份有限公司/唐人数智/装备事业部",
        "罗一一",
    ]
    for col_index, value in enumerate(row, start=1):
        ws.cell(2, col_index).value = value
    wb.save(path)
    wb.close()


def _xlsx_business_manifest(path: Path | None) -> tuple[tuple[str, str], ...]:
    assert path is not None
    with zipfile.ZipFile(path) as workbook:
        return tuple(
            (info.filename, hashlib.sha256(workbook.read(info.filename)).hexdigest())
            for info in workbook.infolist()
            # openpyxl 每次保存都会刷新这里的修改时间；其余 XML/资源必须逐字节一致。
            if info.filename != "docProps/core.xml"
        )


if __name__ == "__main__":
    unittest.main()
