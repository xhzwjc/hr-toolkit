from __future__ import annotations

import importlib.util
import json
import os
import re
import shutil
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "compare_regression_outputs.py"
SPEC = importlib.util.spec_from_file_location("compare_regression_outputs", SCRIPT_PATH)
assert SPEC is not None and SPEC.loader is not None
regression = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = regression
SPEC.loader.exec_module(regression)


def _write_semantic_workbook(path: Path, *, created_year: int = 2024) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "汇总"
    worksheet.append(["姓名", "金额"])
    worksheet.append(["张三", 100])
    worksheet["D1"] = "=SUM(B2:B2)"
    worksheet["D1"].number_format = "0.00"
    worksheet["E1"] = "重点"
    worksheet["E1"].font = Font(name="微软雅黑", bold=True, color="FF0000")
    worksheet["E1"].fill = PatternFill("solid", fgColor="FFFF00")
    worksheet["A2"].comment = Comment("原始批注", "HR")
    worksheet["A2"].hyperlink = "https://example.test/employee/1"
    worksheet.freeze_panes = "C2"
    worksheet.row_dimensions[1].height = 24
    worksheet.column_dimensions["A"].width = 18
    worksheet.merge_cells("H1:I1")
    worksheet["H1"] = "合并标题"

    validation = DataValidation(type="whole", operator="between", formula1="0", formula2="1000")
    validation.add(worksheet["B2"])
    worksheet.add_data_validation(validation)
    worksheet.conditional_formatting.add(
        "B2",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="00FF00")),
    )

    table = Table(displayName="SalaryTable", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    workbook.defined_names.add(DefinedName("业务区域", attr_text="'汇总'!$A$1:$B$2"))
    workbook.properties.created = datetime(created_year, 1, 1)
    workbook.properties.modified = datetime(created_year, 1, 2)
    workbook.save(path)
    workbook.close()


def _write_formula_cache(source: Path, target: Path, value: int) -> None:
    pattern = re.compile(
        rb'(<c r="D1"[^>]*>.*?<f>.*?</f>)<v(?:\s*/>|>.*?</v>)(</c>)',
        flags=re.DOTALL,
    )
    with ZipFile(source) as source_archive, ZipFile(target, "w") as target_archive:
        replaced = False
        for item in source_archive.infolist():
            payload = source_archive.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                payload, count = pattern.subn(
                    rb"\g<1><v>" + str(value).encode("ascii") + rb"</v>\g<2>",
                    payload,
                    count=1,
                )
                replaced = count == 1
            target_archive.writestr(item, payload)
    if not replaced:
        raise AssertionError("formula cache fixture was not updated")


class OutputRegressionSnapshotTests(unittest.TestCase):
    def test_semantic_snapshot_ignores_volatile_core_timestamps(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            first = root / "first.xlsx"
            second = root / "second.xlsx"
            _write_semantic_workbook(first, created_year=2024)
            _write_semantic_workbook(second, created_year=2026)

            first_snapshot = regression.workbook_semantic_snapshot(first)
            second_snapshot = regression.workbook_semantic_snapshot(second)

            self.assertEqual(first_snapshot, second_snapshot)
            self.assertEqual(
                regression.semantic_sha256(first_snapshot),
                regression.semantic_sha256(second_snapshot),
            )

    def test_snapshot_detects_business_relevant_workbook_changes(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            baseline_path = root / "baseline.xlsx"
            _write_semantic_workbook(baseline_path)
            baseline = regression.workbook_semantic_snapshot(baseline_path)

            def value_change(workbook) -> None:
                workbook["汇总"]["A2"] = "李四"

            def formula_change(workbook) -> None:
                workbook["汇总"]["D1"] = "=SUM(B2:B3)"

            def style_change(workbook) -> None:
                workbook["汇总"]["E1"].font = Font(name="宋体", bold=False)

            def comment_change(workbook) -> None:
                workbook["汇总"]["A2"].comment = Comment("修改后批注", "HR")

            def hyperlink_change(workbook) -> None:
                workbook["汇总"]["A2"].hyperlink = "https://example.test/employee/2"

            def dimension_change(workbook) -> None:
                workbook["汇总"].column_dimensions["A"].width = 22

            def merge_change(workbook) -> None:
                worksheet = workbook["汇总"]
                worksheet.unmerge_cells("H1:I1")
                worksheet.merge_cells("H1:J1")

            def freeze_change(workbook) -> None:
                workbook["汇总"].freeze_panes = "D2"

            def defined_name_change(workbook) -> None:
                workbook.defined_names["业务区域"].attr_text = "'汇总'!$A$1:$A$2"

            def print_area_change(workbook) -> None:
                workbook["汇总"].print_area = "A1:A2"

            def header_change(workbook) -> None:
                workbook["汇总"].oddHeader.center.text = "内部资料"

            def workbook_property_change(workbook) -> None:
                workbook.properties.title = "薪酬结果"

            mutators = {
                "value": value_change,
                "formula": formula_change,
                "style": style_change,
                "comment": comment_change,
                "hyperlink": hyperlink_change,
                "dimension": dimension_change,
                "merge": merge_change,
                "freeze": freeze_change,
                "defined_name": defined_name_change,
                "print_area": print_area_change,
                "header": header_change,
                "workbook_property": workbook_property_change,
            }
            for name, mutate in mutators.items():
                with self.subTest(name=name):
                    changed_path = root / f"changed-{name}.xlsx"
                    shutil.copy2(baseline_path, changed_path)
                    workbook = load_workbook(changed_path)
                    mutate(workbook)
                    workbook.save(changed_path)
                    workbook.close()
                    self.assertNotEqual(
                        regression.workbook_semantic_snapshot(changed_path),
                        baseline,
                    )

    def test_snapshot_detects_formula_cached_result_changes(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.xlsx"
            first = root / "cached-100.xlsx"
            second = root / "cached-200.xlsx"
            _write_semantic_workbook(source)
            _write_formula_cache(source, first, 100)
            _write_formula_cache(source, second, 200)

            first_snapshot = regression.workbook_semantic_snapshot(first)
            second_snapshot = regression.workbook_semantic_snapshot(second)
            first_formula = next(
                cell
                for cell in first_snapshot["worksheets"][0]["cells"]
                if cell["coordinate"] == "D1"
            )
            second_formula = next(
                cell
                for cell in second_snapshot["worksheets"][0]["cells"]
                if cell["coordinate"] == "D1"
            )

            self.assertEqual(first_formula["value"], second_formula["value"])
            self.assertEqual(first_formula["cached_value"], 100)
            self.assertEqual(second_formula["cached_value"], 200)
            self.assertNotEqual(first_snapshot, second_snapshot)

    def test_manifest_reports_actual_cell_and_source_hash_changes(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            outputs = root / "outputs"
            sources = root / "sources"
            outputs.mkdir()
            sources.mkdir()
            workbook_path = outputs / "业务结果.xlsx"
            _write_semantic_workbook(workbook_path)
            source = sources / "客户附件.txt"
            source.write_text("原始附件", encoding="utf-8")

            baseline = regression.capture_manifest(outputs, sources)
            unchanged = regression.verify_manifest(baseline, outputs, sources)
            self.assertTrue(unchanged["ok"])

            workbook = load_workbook(workbook_path)
            workbook["汇总"]["A2"] = "李四"
            workbook.save(workbook_path)
            workbook.close()
            source.write_text("附件被修改", encoding="utf-8")

            changed = regression.verify_manifest(baseline, outputs, sources)
            self.assertFalse(changed["ok"])
            paths = [item["path"] for item in changed["differences"]]
            self.assertTrue(any("coordinate=A2/value" in path for path in paths), paths)
            self.assertTrue(any("sources/files/客户附件.txt/sha256" in path for path in paths), paths)

    def test_non_excel_files_and_empty_directories_are_part_of_contract(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            outputs = root / "outputs"
            outputs.mkdir()
            (outputs / "空目录").mkdir()
            report = outputs / "结果.txt"
            report.write_text("before", encoding="utf-8")
            baseline = regression.capture_manifest(outputs)

            report.write_text("after", encoding="utf-8")
            (outputs / "空目录").rmdir()
            changed = regression.verify_manifest(baseline, outputs)

            self.assertFalse(changed["ok"])
            paths = [item["path"] for item in changed["differences"]]
            self.assertTrue(any("结果.txt/sha256" in path for path in paths), paths)
            self.assertTrue(any("directories" in path for path in paths), paths)

    def test_symlinks_are_recorded_without_following_targets(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            outputs = root / "outputs"
            outputs.mkdir()
            outside = root / "outside.txt"
            outside.write_text("outside", encoding="utf-8")
            link = outputs / "link.txt"
            try:
                link.symlink_to(outside)
            except OSError:
                self.skipTest("当前文件系统不支持符号链接")

            captured = regression.capture_tree(outputs, excel_semantics=True)

            self.assertEqual(captured["files"], {})
            self.assertEqual(captured["symlinks"], {"link.txt": os.readlink(link)})


class OutputRegressionCliTests(unittest.TestCase):
    def test_capture_verify_and_json_report_exit_codes(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            outputs = root / "outputs"
            sources = root / "sources"
            outputs.mkdir()
            sources.mkdir()
            _write_semantic_workbook(outputs / "结果.xlsx")
            (sources / "附件.bin").write_bytes(b"source")
            baseline = root / "baseline.json"
            report = root / "report.json"

            self.assertEqual(
                regression.main(
                    [
                        "capture",
                        "--outputs",
                        str(outputs),
                        "--sources",
                        str(sources),
                        "--manifest",
                        str(baseline),
                    ]
                ),
                0,
            )
            self.assertEqual(
                regression.main(
                    [
                        "verify",
                        "--baseline",
                        str(baseline),
                        "--outputs",
                        str(outputs),
                        "--sources",
                        str(sources),
                        "--report",
                        str(report),
                    ]
                ),
                0,
            )
            self.assertTrue(json.loads(report.read_text(encoding="utf-8"))["ok"])

            (sources / "附件.bin").write_bytes(b"changed")
            self.assertEqual(
                regression.main(
                    [
                        "verify",
                        "--baseline",
                        str(baseline),
                        "--outputs",
                        str(outputs),
                        "--sources",
                        str(sources),
                        "--report",
                        str(report),
                    ]
                ),
                1,
            )
            self.assertFalse(json.loads(report.read_text(encoding="utf-8"))["ok"])

    def test_verify_rejects_missing_source_snapshot(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            outputs = root / "outputs"
            sources = root / "sources"
            outputs.mkdir()
            sources.mkdir()
            (outputs / "result.txt").write_text("ok", encoding="utf-8")
            (sources / "source.txt").write_text("source", encoding="utf-8")
            baseline = regression.capture_manifest(outputs, sources)

            report = regression.verify_manifest(baseline, outputs)

            self.assertFalse(report["ok"])
            self.assertTrue(any(item["path"] == "/sources" for item in report["differences"]))


if __name__ == "__main__":
    unittest.main()
