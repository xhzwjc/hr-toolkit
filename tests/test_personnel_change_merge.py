from __future__ import annotations

import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from hr_toolkit.tools.personnel_change_merge import merge_personnel_changes, update_roster_from_change_summaries


class PersonnelChangeMergeTest(unittest.TestCase):
    def test_merge_multiple_project_change_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()

            _write_change_file(
                input_dir / "项目A异动表.xlsx",
                {
                    "增员": [
                        ["员工1", "44162219901007667X", "生产人员", "2026-04-01"],
                        ["员工2", "441622198404210312", "管理人员", "2026-04-02"],
                    ],
                    "减员": [["员工3", "44162219800516649X", "离职", "2026-04-30"]],
                },
            )
            _write_change_file(
                input_dir / "项目B异动表.xlsx",
                {
                    "增员": [["员工4", "44132419860927333X", "生产人员", "2026-04-03"]],
                    "奖罚扣补": [["员工5", 0, 100, 0, 50]],
                },
            )

            result = merge_personnel_changes(input_dir, output_dir)
            payload = result.to_dict()

            self.assertEqual(payload["source_file_count"], 2)
            self.assertEqual(payload["record_count"], 4)
            self.assertEqual(payload["sheet_counts"]["增员"], 3)
            self.assertEqual(payload["sheet_counts"]["减员"], 1)
            self.assertTrue(result.output_file and result.output_file.exists())

            wb = load_workbook(result.output_file, data_only=True)
            add_ws = wb["增员"]
            self.assertEqual([add_ws.cell(3, col).value for col in (1, 4, 5, 10)], [1, "员工1", "44162219901007667X", "生产人员"])
            self.assertEqual([add_ws.cell(5, col).value for col in (1, 4, 5, 10)], [3, "员工4", "44132419860927333X", "生产人员"])
            self.assertEqual(add_ws.cell(5, 19).value, "2026-04-03")
            self.assertIsNone(add_ws.cell(6, 4).value)

            leave_ws = wb["减员"]
            self.assertEqual([leave_ws.cell(3, col).value for col in (1, 4, 5, 9)], [1, "员工3", "44162219800516649X", "离职"])
            self.assertEqual(leave_ws.cell(3, 7).value, "2026-04-30")

    def test_dry_run_does_not_write_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            _write_change_file(input_dir / "项目A异动表.xlsx", {"增员": [["员工1", "44162219901007667X", "生产人员", "2026-04-01"]]})

            result = merge_personnel_changes(input_dir, output_dir, dry_run=True)

            self.assertEqual(result.record_count, 1)
            self.assertIsNone(result.output_file)
            self.assertFalse(output_dir.exists())

    def test_missing_template_sheet_reports_clear_error(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            template_path = root / "缺少调动模板.xlsx"
            input_dir.mkdir()
            _write_change_file(input_dir / "项目A异动表.xlsx", {"增员": [["员工1", "44162219901007667X", "生产人员", "2026-04-01"]]})
            _write_change_file(template_path, {}, omit_sheets={"调动"})

            with self.assertRaisesRegex(ValueError, "异动汇总表缺少工作表：调动"):
                merge_personnel_changes(input_dir, output_dir, template_path=template_path)

    def test_existing_summary_is_appended_without_clearing_original_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "2026年4月项目A异动表.xlsx"
            output_dir = root / "output"
            existing_summary = root / "已有异动汇总表.xlsx"
            _write_change_file(source, {"增员": [["新增员工", "44162219901007667X", "生产人员", "2026-04-01"]]})
            _write_real_summary_template(existing_summary)
            wb = load_workbook(existing_summary)
            add_ws = wb["增员"]
            add_ws.cell(3, 1).value = 1
            add_ws.cell(3, 4).value = "原有员工"
            add_ws.cell(3, 5).value = "36040219791125XXXX"
            wb.save(existing_summary)
            wb.close()

            result = merge_personnel_changes(source, output_dir, template_path=existing_summary)

            self.assertTrue(result.append_mode)
            self.assertEqual(result.inserted_count, 1)
            wb = load_workbook(result.output_file, data_only=True)
            add_ws = wb["增员"]
            self.assertEqual(add_ws.cell(3, 4).value, "原有员工")
            self.assertEqual(add_ws.cell(4, 4).value, "新增员工")
            self.assertEqual(add_ws.cell(4, 5).value, "44162219901007667X")
            self.assertEqual(add_ws.cell(4, 10).value, "生产人员")
            wb.close()

    def test_rows_are_routed_to_monthly_summaries_by_row_dates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "2026年7月项目A异动表.xlsx"
            output_dir = root / "output"
            _write_change_file(
                source,
                {
                    "增员": [
                        ["四月员工", "44162219901007667X", "生产人员", "2026-04-01"],
                        ["五月员工", "441622198404210312", "管理人员", "2026-05-01"],
                    ],
                    "减员": [["五月离职", "44162219800516649X", "离职", "2026-05-20"]],
                },
            )

            result = merge_personnel_changes(source, output_dir)

            self.assertIsNone(result.output_file)
            self.assertEqual([path.name for path in result.output_files], ["2026年4月异动汇总表.xlsx", "2026年5月异动汇总表.xlsx"])
            self.assertEqual(result.period_counts["2026年4月"]["增员"], 1)
            self.assertEqual(result.period_counts["2026年5月"]["增员"], 1)
            self.assertEqual(result.period_counts["2026年5月"]["减员"], 1)
            april = load_workbook(output_dir / "2026年4月异动汇总表.xlsx", data_only=True)
            may = load_workbook(output_dir / "2026年5月异动汇总表.xlsx", data_only=True)
            self.assertEqual(april["增员"].cell(3, 4).value, "四月员工")
            self.assertEqual(may["增员"].cell(3, 4).value, "五月员工")
            self.assertEqual(may["减员"].cell(3, 4).value, "五月离职")
            april.close()
            may.close()

    def test_summary_folder_is_matched_by_month_and_missing_month_is_created(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "项目A异动表.xlsx"
            output_dir = root / "output"
            summary_dir = root / "summaries"
            summary_dir.mkdir()
            april_summary = summary_dir / "2026年4月异动汇总表.xlsx"
            _write_real_summary_template(april_summary)
            wb = load_workbook(april_summary)
            ws = wb["增员"]
            ws.cell(3, 1).value = 1
            ws.cell(3, 4).value = "原四月员工"
            ws.cell(3, 5).value = "36040219791125XXXX"
            wb.save(april_summary)
            wb.close()
            _write_change_file(
                source,
                {
                    "增员": [
                        ["新增四月", "44162219901007667X", "生产人员", "2026-04-01"],
                        ["新增五月", "441622198404210312", "管理人员", "2026-05-01"],
                    ]
                },
            )

            result = merge_personnel_changes(source, output_dir, template_path=summary_dir)

            self.assertTrue(result.append_mode)
            self.assertEqual([path.name for path in result.output_files], ["2026年4月异动汇总表.xlsx", "2026年5月异动汇总表.xlsx"])
            april = load_workbook(output_dir / "2026年4月异动汇总表.xlsx", data_only=True)
            may = load_workbook(output_dir / "2026年5月异动汇总表.xlsx", data_only=True)
            self.assertEqual(april["增员"].cell(3, 4).value, "原四月员工")
            self.assertEqual(april["增员"].cell(4, 4).value, "新增四月")
            self.assertEqual(may["增员"].cell(3, 4).value, "新增五月")
            april.close()
            may.close()

    def test_default_summary_uses_packaged_clean_template(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "项目A异动表.xlsx"
            output_dir = root / "output"
            _write_change_file(source, {"增员": [["内置模板员工", "44162219901007667X", "生产人员", "2026-04-01"]]})

            result = merge_personnel_changes(source, output_dir)

            wb = load_workbook(result.output_file, data_only=True)
            self.assertIn("奖罚扣补", wb.sheetnames)
            self.assertEqual(wb["增员"].cell(1, 1).value, "2026年4月增员表")
            self.assertEqual(wb["增员"].cell(3, 4).value, "内置模板员工")
            self.assertEqual(wb["增员"].cell(3, 4).alignment.horizontal, "center")
            self.assertEqual(wb["增员"].cell(3, 4).alignment.vertical, "center")
            self.assertEqual(wb["增员"].cell(3, 4).border.left.style, "thin")
            self.assertIsNone(wb["增员"].cell(7, 1).value)
            wb.close()

    def test_zip_input_is_extracted_and_merged(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "项目A异动表.xlsx"
            zip_path = root / "项目A.zip"
            output_dir = root / "output"
            _write_change_file(source, {"增员": [["压缩包员工", "44162219901007667X", "生产人员", "2026-04-01"]]})
            with zipfile.ZipFile(zip_path, "w") as archive:
                archive.write(source, "项目A异动表.xlsx")

            result = merge_personnel_changes(zip_path, output_dir)

            self.assertEqual(len(result.source_files), 1)
            self.assertEqual(result.record_count, 1)
            wb = load_workbook(result.output_files[0], data_only=True)
            self.assertEqual(wb["增员"].cell(3, 4).value, "压缩包员工")
            wb.close()

    def test_real_project_sheet_names_map_to_month_summary_and_roster(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            source = input_dir / "问题6-2026年4月南昌分公司异动表.xlsx"
            summary_template = root / "问题6-2026年4月异动汇总表.xlsx"
            analysis_template = input_dir / "问题6-2026年4月人力资源分析.xlsx"
            _write_real_change_file(source)
            _write_real_summary_template(summary_template)
            _write_analysis_template(analysis_template)

            result = merge_personnel_changes(input_dir, output_dir, template_path=summary_template)

            self.assertEqual(result.period, "2026年4月")
            self.assertEqual(result.sheet_counts["增员"], 1)
            self.assertEqual(result.sheet_counts["减员"], 1)
            self.assertEqual(result.output_file.name, "2026年4月异动汇总表.xlsx")
            self.assertEqual(result.roster_output_file.name, "2026年4月人力资源分析表_更新后.xlsx")
            self.assertEqual(result.roster_added_count, 1)
            self.assertEqual(result.roster_marked_count, 1)

            summary = load_workbook(result.output_file, data_only=False)
            add_ws = summary["增员"]
            self.assertEqual(add_ws["A1"].value, "2026年4月增员表")
            self.assertEqual(add_ws.cell(3, 2).value, "春苗南昌")
            self.assertEqual(add_ws.cell(3, 3).value, "南昌")
            self.assertEqual(add_ws.cell(3, 9).value, "维护员")
            self.assertEqual(add_ws.cell(3, 11).value, "基站")
            self.assertEqual(add_ws.cell(3, 12).value, "1597067xxxx")
            leave_ws = summary["减员"]
            self.assertEqual(leave_ws.cell(3, 3).value, "南昌")
            self.assertEqual(leave_ws.cell(3, 6).number_format, "yyyy/m/d")
            self.assertEqual(leave_ws.cell(3, 7).number_format, "yyyy/m/d")
            self.assertEqual(leave_ws.cell(3, 8).number_format, "yyyy/m/d")

            analysis = load_workbook(result.roster_output_file, data_only=False)
            roster = analysis["花名册"]
            self.assertEqual(roster.cell(6, 3).value, "南昌")
            self.assertEqual(roster.cell(6, 4).value, "钱一")
            self.assertEqual(roster.cell(6, 13).value, "春苗南昌")
            self.assertEqual(roster.cell(6, 17).value, "2026-04-01")
            self.assertEqual(roster.cell(6, 4).alignment.horizontal, "center")
            self.assertEqual(roster.cell(6, 4).alignment.vertical, "center")
            self.assertEqual(roster.cell(6, 4).border.left.style, "thin")
            self.assertEqual(roster.cell(4, 1).fill.fgColor.rgb, "00FFF2CC")

    def test_multiple_projects_months_and_change_types_update_roster(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            _write_change_file(
                input_dir / "项目A异动表.xlsx",
                {
                    "增员": [["三月新增", "360111199603170001", "生产人员", "2026-03-15"]],
                    "减员": [["孙一", "360121197710127xxx", "离职", "2026-05-20"]],
                    "转正": [["四月转正", "2026-02-01", "2026-04-01"]],
                    "调动": [["五月调动", "基站", "线路", "2026-05-03"]],
                },
            )
            _write_change_file(
                input_dir / "项目B异动表.xlsx",
                {
                    "增员": [["五月新增", "360111199603170002", "管理人员", "2026-05-06"]],
                    "转正": [["六月转正", "2026-03-01", "2026-06-01"]],
                    "调动": [["四月调动", "线路", "基站", "2026-04-08"]],
                },
            )
            _write_analysis_template(input_dir / "人力资源分析表.xlsx")

            result = merge_personnel_changes(input_dir, output_dir)

            self.assertEqual(
                [path.name for path in result.output_files],
                ["2026年3月异动汇总表.xlsx", "2026年4月异动汇总表.xlsx", "2026年5月异动汇总表.xlsx", "2026年6月异动汇总表.xlsx"],
            )
            self.assertEqual(result.sheet_counts, {"增员": 2, "减员": 1, "转正": 2, "调动": 2})
            self.assertEqual(result.roster_added_count, 2)
            self.assertEqual(result.roster_marked_count, 1)
            may = load_workbook(output_dir / "2026年5月异动汇总表.xlsx", data_only=True)
            self.assertEqual(may["减员"].cell(3, 4).value, "孙一")
            self.assertEqual(may["调动"].cell(3, 5).value, "五月调动")
            may.close()

            roster_book = load_workbook(result.roster_output_file, data_only=False)
            roster = roster_book["花名册"]
            names = [roster.cell(row, 4).value for row in range(4, 8)]
            self.assertIn("三月新增", names)
            self.assertIn("五月新增", names)
            self.assertEqual(roster.cell(4, 1).fill.fgColor.rgb, "00FFF2CC")
            self.assertEqual(roster.cell(6, 4).alignment.horizontal, "center")
            self.assertEqual(roster.cell(6, 4).alignment.vertical, "center")
            self.assertEqual(roster.cell(6, 4).border.left.style, "thin")
            roster_book.close()

    def test_update_roster_from_existing_summary_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            summary_dir = root / "summaries"
            roster_dir = root / "roster"
            input_dir.mkdir()
            _write_change_file(
                input_dir / "项目A异动表.xlsx",
                {
                    "增员": [["汇总新增", "360111199603170003", "生产人员", "2026-04-15"]],
                    "减员": [["孙一", "360121197710127xxx", "离职", "2026-04-20"]],
                    "转正": [["汇总转正", "2026-02-01", "2026-04-01"]],
                    "调动": [["汇总调动", "基站", "线路", "2026-04-08"]],
                },
            )
            roster_template = root / "人力资源花名册.xlsx"
            _write_analysis_template(roster_template)
            merge_result = merge_personnel_changes(input_dir, summary_dir)

            roster_result = update_roster_from_change_summaries(merge_result.output_files, roster_template, roster_dir)

            self.assertEqual(roster_result.record_count, 4)
            self.assertEqual(roster_result.roster_added_count, 1)
            self.assertEqual(roster_result.roster_marked_count, 1)
            self.assertTrue(roster_result.output_file and roster_result.output_file.exists())
            roster_book = load_workbook(roster_result.output_file, data_only=False)
            roster = roster_book["花名册"]
            names = [roster.cell(row, 4).value for row in range(4, 7)]
            self.assertIn("汇总新增", names)
            self.assertEqual(roster.cell(4, 1).fill.fgColor.rgb, "00FFF2CC")
            roster_book.close()


def _write_change_file(path: Path, rows_by_sheet: dict[str, list[list]], omit_sheets: set[str] | None = None) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    omit_sheets = omit_sheets or set()
    sheet_headers = {
        "增员": ["序号", "姓名", "身份证号码", "人员分类", "入职日期"],
        "减员": ["序号", "姓名", "身份证号码", "备注", "离职日期"],
        "转正": ["序号", "姓名", "入职日期", "转正日期"],
        "调动": ["序号", "姓名", "原部门", "现部门", "调整日期"],
    }
    for sheet_name, headers in sheet_headers.items():
        if sheet_name in omit_sheets:
            continue
        ws = workbook.create_sheet(sheet_name)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws["A1"] = f"2026年4月{sheet_name}表"
        for col_index, header in enumerate(headers, start=1):
            ws.cell(2, col_index).value = header
        for row_index in range(3, 8):
            ws.cell(row_index, 1).value = row_index - 2

        rows = rows_by_sheet.get(sheet_name, [])
        for offset, values in enumerate(rows):
            row_index = 3 + offset
            ws.cell(row_index, 1).value = offset + 1
            for col_index, value in enumerate(values, start=2):
                ws.cell(row_index, col_index).value = value
    workbook.save(path)


def _write_real_change_file(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    add = workbook.create_sheet("增补表")
    add["A1"] = "南昌分公司增员表"
    add_headers = ["序号", "公司", "部门（片区）", "姓名", "职务", "人员分类", "专业", "工作联系电话", "身份证号码", "性别", "学历", "毕业学校", "婚否", "家庭住址", "入职日期", "用工状态", "试用期工资", "备注"]
    for col, header in enumerate(add_headers, start=1):
        add.cell(2, col).value = header
    for col, value in enumerate([1, "春苗南昌", "南昌", "钱一", "维护员", "生产人员", "基站", "1597067xxxx", "360111199603170xxx", "男", "本科", "学校", None, None, "2026-04-01"], start=1):
        add.cell(3, col).value = value

    leave = workbook.create_sheet("离职")
    leave_headers = ["序号", "公司", "部门（片区）", "姓名", "身份证号码", "入职日期", "离职日期", "薪资结算日期", "备注"]
    for col, header in enumerate(leave_headers, start=1):
        leave.cell(1, col).value = header
    for col, value in enumerate([1, "春苗北京", "南昌", "孙一", "360121197710127xxx", "2026-02-01", "2026-04-30", "2026-04-30"], start=1):
        leave.cell(2, col).value = value
    for sheet_name in ("转正", "调整", "奖、罚、扣、补"):
        ws = workbook.create_sheet(sheet_name)
        headers = ["序号", "公司", "部门（片区）", "姓名"]
        for col, header in enumerate(headers, start=1):
            ws.cell(2, col).value = header
    workbook.save(path)


def _write_real_summary_template(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    headers_by_sheet = {
        "增员": ["序号", "入职公司", "地市", "姓名", "身份证号码", "出生日期", "年龄", "性别", "岗位", "人员分类", "所属专业", "联系方式", "学历", "毕业学校", "专业", "毕业时间", "婚否", "家庭住址", "入职日期", "用工状态", "试用期工资", "备注"],
        "减员": ["序号", "公司", "地市", "姓名", "身份证号码", "入职日期", "离职日期", "薪资结算日期", "备注"],
        "转正": ["序号", "公司", "部门（片区）", "姓名", "职务", "入职日期", "转正日期", "试用期工资（元）", "转正后岗级", "转正后工资（元）", "备注"],
        "调动": ["序号", "公司", "地市", "部门（片区）", "姓名", "原部门", "原职位", "原岗级", "原金额（元）", "现部门", "现职位", "异动类型", "调整后金额（元）", "增减额度（元）", "调整日期", "备注"],
        "奖罚扣补": ["序号", "地市", "部门（片区）", "姓名", "罚（元）", "奖（元）", "扣（元）", "补（元）", "备注"],
    }
    for sheet_name, headers in headers_by_sheet.items():
        ws = workbook.create_sheet(sheet_name)
        ws["A1"] = f"2026年4月{sheet_name}表"
        for col, header in enumerate(headers, start=1):
            ws.cell(2, col).value = header
        for col in range(1, len(headers) + 1):
            ws.cell(3, col).value = None
    workbook.save(path)


def _write_analysis_template(path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "花名册"
    headers = ["序号", "部门", "部门/项目", "姓名", "身份证号码", "出生日期", "年龄", "性别", "岗位", "人员分类", "所属专业", "联系方式", "入职公司", "学历", "毕业院校", "专业", "入职时间"]
    for col, header in enumerate(headers, start=1):
        ws.cell(3, col).value = header
    rows = [
        [1, "运营商事业部", "南昌", "孙一", "360121197710127xxx", None, None, "男", "维护员", "生产人员", "基站", "123", "北京春苗", "本科", "x", "x", "2026-02-01"],
        [2, "运营商事业部", "南昌", "孙二", "360122198204207xxx", None, None, "女", "维护员", "生产人员", "基站", "124", "北京春苗", "本科", "x", "x", "2026-01-01"],
    ]
    for row_index, row in enumerate(rows, start=4):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index).value = value
    ws.cell(6, 1).value = "对应异动汇总表中的字段"
    workbook.save(path)


if __name__ == "__main__":
    unittest.main()
