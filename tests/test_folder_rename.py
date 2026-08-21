from __future__ import annotations

import hashlib
import io
import json
import shutil
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path

from openpyxl import Workbook

from hr_toolkit.cli import main as cli_main
from hr_toolkit.tools.folder_rename import (
    FILE_TYPE_ALL,
    FILE_TYPE_FOLDER,
    FILE_TYPE_IMAGE,
    MODE_APPEND,
    MODE_EXCEL_BATCH,
    MODE_REMOVE,
    MODE_REPLACE,
    rename_files_by_excel,
    rename_person_folders,
)


class FolderRenameTest(unittest.TestCase):
    @staticmethod
    def _write_name_workbook(path: Path, names: list[str], *, header_row: int = 1) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(header_row, 1, "姓名")
        for row, name in enumerate(names, start=header_row + 1):
            worksheet.cell(row, 1, name)
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _sha256(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    def test_append_suffix_to_all_folders(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "张三").mkdir()
            (root / "李四").mkdir()
            (root / "说明.txt").write_text("ignore", encoding="utf-8")

            # “-劳动合同”带前缀追加;右侧注释验证“劳动合同”不带前缀时直传
            preview = rename_person_folders(root, mode=MODE_APPEND, text="-劳动合同", dry_run=True)
            self.assertEqual(preview.operation_count, 2)
            self.assertTrue((root / "张三").exists())

            result = rename_person_folders(root, mode=MODE_APPEND, text="-劳动合同")

            self.assertEqual(result.operation_count, 2)
            self.assertTrue((root / "张三-劳动合同").exists())
            self.assertTrue((root / "李四-劳动合同").exists())
            self.assertTrue((root / "说明.txt").exists())

    def test_append_text_passed_through_unchanged(self) -> None:
        """bug4 修复:用户输入什么就追加什么，不自动加分隔符"""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "张三").mkdir()

            result = rename_person_folders(root, mode=MODE_APPEND, text="劳动合同")

            self.assertEqual(result.operation_count, 1)
            self.assertTrue((root / "张三劳动合同").exists())
            self.assertFalse((root / "张三-劳动合同").exists())

    def test_append_suffix_to_one_person(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "张三").mkdir()
            (root / "李四").mkdir()

            result = rename_person_folders(root, mode=MODE_APPEND, text="-身份证", target_name="张三")

            self.assertEqual(result.operation_count, 1)
            self.assertTrue((root / "张三-身份证").exists())
            self.assertTrue((root / "李四").exists())

    def test_remove_suffix_variants_from_all_folders(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "张三-劳动合同").mkdir()
            (root / "李四_劳动合同").mkdir()
            (root / "赵露思劳动合同").mkdir()
            (root / "王五-身份证").mkdir()

            result = rename_person_folders(root, mode=MODE_REMOVE, text="_劳动合同")

            self.assertEqual(result.operation_count, 3)
            self.assertTrue((root / "张三").exists())
            self.assertTrue((root / "李四").exists())
            self.assertTrue((root / "赵露思").exists())
            self.assertTrue((root / "王五-身份证").exists())

    def test_remove_suffix_from_one_person(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "张三_身份证").mkdir()
            (root / "李四_身份证").mkdir()

            result = rename_person_folders(root, mode=MODE_REMOVE, text="身份证", target_name="张三")

            self.assertEqual(result.operation_count, 1)
            self.assertTrue((root / "张三").exists())
            self.assertTrue((root / "李四_身份证").exists())

    def test_replace_one_folder_name(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "张三").mkdir()
            (root / "李四").mkdir()

            result = rename_person_folders(
                root,
                mode=MODE_REPLACE,
                target_name="张三",
                replacement_name="章五",
            )

            self.assertEqual(result.operation_count, 1)
            self.assertFalse((root / "张三").exists())
            self.assertTrue((root / "章五").exists())
            self.assertTrue((root / "李四").exists())

    def test_skip_existing_target(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "张三").mkdir()
            (root / "张三-劳动合同").mkdir()

            result = rename_person_folders(root, mode=MODE_APPEND, text="-劳动合同")

            self.assertEqual(result.operation_count, 0)
            self.assertTrue(any("已存在" in warning or "已包含后缀" in warning for warning in result.warnings))

    def test_excel_batch_preview_and_execute_change_only_names(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            root.mkdir()
            sources = {
                "10.png": b"image-ten",
                "2.jpg": b"image-two",
                "1.pdf": b"pdf-one",
            }
            for name, content in sources.items():
                (root / name).write_bytes(content)
            (root / ".DS_Store").write_bytes(b"hidden")
            (root / "~$临时.docx").write_bytes(b"temporary")
            roster = base / "人员名单.xlsx"
            self._write_name_workbook(roster, ["张三", "李四", "王五"])

            source_hashes = {name: self._sha256(root / name) for name in sources}
            roster_hash = self._sha256(roster)
            preview = rename_files_by_excel(root, roster, file_type=FILE_TYPE_ALL, dry_run=True)

            self.assertEqual(preview.mode, MODE_EXCEL_BATCH)
            self.assertEqual(
                [operation.source.name for operation in preview.operations],
                ["1.pdf", "2.jpg", "10.png"],
            )
            self.assertEqual(
                [operation.target.name for operation in preview.operations],
                ["张三.pdf", "李四.jpg", "王五.png"],
            )
            self.assertEqual({name: self._sha256(root / name) for name in sources}, source_hashes)
            self.assertEqual(self._sha256(roster), roster_hash)

            result = rename_files_by_excel(
                root,
                roster,
                file_type=FILE_TYPE_ALL,
                expected_operations=[
                    (operation.source.name, operation.target.name)
                    for operation in preview.operations
                ],
                expected_warnings=list(preview.warnings),
            )

            self.assertEqual(result.operation_count, 3)
            for operation in preview.operations:
                self.assertFalse(operation.source.exists())
                self.assertEqual(
                    self._sha256(operation.target),
                    source_hashes[operation.source.name],
                )
                self.assertEqual(operation.source.suffix, operation.target.suffix)
            self.assertTrue((root / ".DS_Store").is_file())
            self.assertTrue((root / "~$临时.docx").is_file())
            self.assertEqual(self._sha256(roster), roster_hash)

    def test_excel_batch_fewer_names_keeps_extra_item_and_warns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            root.mkdir()
            (root / "1.pdf").write_bytes(b"one")
            (root / "2.pdf").write_bytes(b"two")
            roster = base / "名单.xlsx"
            self._write_name_workbook(roster, ["张三"])

            result = rename_files_by_excel(root, roster)

            self.assertTrue((root / "张三.pdf").is_file())
            self.assertEqual((root / "张三.pdf").read_bytes(), b"one")
            self.assertEqual((root / "2.pdf").read_bytes(), b"two")
            self.assertTrue(any("少 1 个" in warning and "2.pdf" in warning for warning in result.warnings))

    def test_excel_batch_more_names_reports_every_unmatched_name(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            root.mkdir()
            (root / "1.pdf").write_bytes(b"one")
            roster = base / "名单.xlsx"
            self._write_name_workbook(roster, ["张三", "李四", "王五"])

            preview = rename_files_by_excel(root, roster, dry_run=True)

            self.assertEqual(preview.operation_count, 1)
            self.assertTrue(
                any("多 2 人" in warning and "李四" in warning and "王五" in warning for warning in preview.warnings)
            )
            self.assertTrue((root / "1.pdf").is_file())

    def test_excel_batch_file_type_filter_keeps_nonmatching_items(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            root.mkdir()
            (root / "1.jpg").write_bytes(b"jpg")
            (root / "2.pdf").write_bytes(b"pdf")
            (root / "3.png").write_bytes(b"png")
            (root / "4").mkdir()
            roster = base / "名单.xlsx"
            self._write_name_workbook(roster, ["张三", "李四"])

            result = rename_files_by_excel(root, roster, file_type=FILE_TYPE_IMAGE)

            self.assertEqual(result.operation_count, 2)
            self.assertEqual((root / "张三.jpg").read_bytes(), b"jpg")
            self.assertEqual((root / "李四.png").read_bytes(), b"png")
            self.assertEqual((root / "2.pdf").read_bytes(), b"pdf")
            self.assertTrue((root / "4").is_dir())

    def test_excel_batch_folder_filter_preserves_folder_contents(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            (root / "档案10").mkdir(parents=True)
            (root / "档案2").mkdir()
            (root / "档案2" / "说明.txt").write_text("record", encoding="utf-8")
            (root / "忽略.pdf").write_bytes(b"pdf")
            roster = base / "名单.xlsx"
            self._write_name_workbook(roster, ["张三", "李四"])

            result = rename_files_by_excel(root, roster, file_type=FILE_TYPE_FOLDER)

            self.assertEqual(result.operation_count, 2)
            self.assertEqual((root / "张三" / "说明.txt").read_text(encoding="utf-8"), "record")
            self.assertTrue((root / "李四").is_dir())
            self.assertTrue((root / "忽略.pdf").is_file())

    def test_excel_batch_excludes_roster_copy_inside_root(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            root.mkdir()
            (root / "1.pdf").write_bytes(b"one")
            (root / "2.pdf").write_bytes(b"two")
            external_roster = base / "留存" / "名单.xlsx"
            external_roster.parent.mkdir()
            self._write_name_workbook(external_roster, ["张三", "李四"])
            roster_copy = root / external_roster.name
            shutil.copy2(external_roster, roster_copy)
            roster_hash = self._sha256(roster_copy)

            preview = rename_files_by_excel(root, external_roster, file_type=FILE_TYPE_ALL, dry_run=True)

            self.assertEqual([item.source.name for item in preview.operations], ["1.pdf", "2.pdf"])
            self.assertEqual(self._sha256(roster_copy), roster_hash)

    def test_excel_batch_existing_target_is_never_overwritten(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            root.mkdir()
            (root / "1.pdf").write_bytes(b"source")
            (root / "张三.pdf").mkdir()
            (root / "张三.pdf" / "原内容.txt").write_text("keep", encoding="utf-8")
            roster = base / "名单.xlsx"
            self._write_name_workbook(roster, ["张三"])

            result = rename_files_by_excel(root, roster)

            self.assertEqual(result.operation_count, 0)
            self.assertEqual((root / "1.pdf").read_bytes(), b"source")
            self.assertEqual((root / "张三.pdf" / "原内容.txt").read_text(encoding="utf-8"), "keep")
            self.assertTrue(any("已存在" in warning for warning in result.warnings))

    def test_excel_batch_duplicate_and_invalid_names_do_not_shift_pairing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            root.mkdir()
            for index in range(1, 4):
                (root / f"{index}.pdf").write_bytes(str(index).encode())
            roster = base / "名单.xlsx"
            self._write_name_workbook(roster, ["坏/名", "李四", "李四"])

            result = rename_files_by_excel(root, roster)

            self.assertEqual(result.operation_count, 1)
            self.assertEqual((root / "1.pdf").read_bytes(), b"1")
            self.assertEqual((root / "李四.pdf").read_bytes(), b"2")
            self.assertEqual((root / "3.pdf").read_bytes(), b"3")
            self.assertTrue(any("不能用于改名" in warning for warning in result.warnings))
            self.assertTrue(any("目标名称重复" in warning for warning in result.warnings))

    def test_excel_batch_no_matching_items_returns_explicit_warning(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            root.mkdir()
            (root / "说明.txt").write_text("keep", encoding="utf-8")
            roster = base / "名单.xlsx"
            self._write_name_workbook(roster, ["张三", "李四"])

            preview = rename_files_by_excel(root, roster, file_type=FILE_TYPE_IMAGE, dry_run=True)

            self.assertEqual(preview.operation_count, 0)
            self.assertTrue(any("多 2 人" in warning for warning in preview.warnings))
            self.assertTrue((root / "说明.txt").is_file())

    def test_excel_batch_finds_name_header_below_title_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            root.mkdir()
            (root / "1.pdf").write_bytes(b"one")
            roster = base / "名单.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.cell(1, 1, "项目人员姓名名单")
            worksheet.cell(3, 1, "姓名")
            worksheet.cell(4, 1, "张三")
            workbook.save(roster)
            workbook.close()

            preview = rename_files_by_excel(root, roster, dry_run=True)

            self.assertEqual(preview.operation_count, 1)
            self.assertEqual(preview.operations[0].target.name, "张三.pdf")

    def test_excel_batch_rejects_windows_reserved_and_case_duplicate_names(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            root.mkdir()
            for index in range(1, 4):
                (root / f"{index}.pdf").write_bytes(str(index).encode())
            roster = base / "名单.xlsx"
            self._write_name_workbook(roster, ["CON", "Alice", "ALICE"])

            preview = rename_files_by_excel(root, roster, dry_run=True)

            self.assertEqual(preview.operation_count, 1)
            self.assertEqual(preview.operations[0].source.name, "2.pdf")
            self.assertEqual(preview.operations[0].target.name, "Alice.pdf")
            self.assertTrue(any("Windows 保留名称" in warning for warning in preview.warnings))
            self.assertTrue(any("目标名称重复" in warning for warning in preview.warnings))

    def test_excel_batch_aborts_when_confirmed_preview_changes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            root.mkdir()
            (root / "1.pdf").write_bytes(b"one")
            roster = base / "名单.xlsx"
            self._write_name_workbook(roster, ["张三"])
            preview = rename_files_by_excel(root, roster, dry_run=True)
            expected_operations = [
                (operation.source.name, operation.target.name)
                for operation in preview.operations
            ]
            (root / "0.pdf").write_bytes(b"new")

            with self.assertRaisesRegex(RuntimeError, "预览确认后发生了变化"):
                rename_files_by_excel(
                    root,
                    roster,
                    expected_operations=expected_operations,
                    expected_warnings=list(preview.warnings),
                )

            self.assertEqual((root / "0.pdf").read_bytes(), b"new")
            self.assertEqual((root / "1.pdf").read_bytes(), b"one")
            self.assertFalse((root / "张三.pdf").exists())

    def test_cli_excel_mode_previews_without_changing_source(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / "待改名"
            root.mkdir()
            source = root / "1.pdf"
            source.write_bytes(b"source")
            roster = base / "名单.xlsx"
            self._write_name_workbook(roster, ["张三"])
            output = io.StringIO()

            with redirect_stdout(output):
                exit_code = cli_main(
                    [
                        "folder-rename",
                        "--root",
                        str(root),
                        "--mode",
                        "excel",
                        "--excel",
                        str(roster),
                        "--file-type",
                        "pdf",
                        "--json",
                    ]
                )

            payload = json.loads(output.getvalue())
            self.assertEqual(exit_code, 0)
            self.assertTrue(payload["dry_run"])
            self.assertEqual(payload["operations"][0]["source_name"], "1.pdf")
            self.assertEqual(payload["operations"][0]["target_name"], "张三.pdf")
            self.assertEqual(source.read_bytes(), b"source")


if __name__ == "__main__":
    unittest.main()
