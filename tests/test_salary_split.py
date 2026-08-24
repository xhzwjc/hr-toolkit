from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from hr_toolkit.tools.salary_split import split_salary_by_company


class SalarySplitTest(unittest.TestCase):
    def test_split_sample_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            sample = root / "脱敏工资表.xlsx"
            out_dir = root / "output"
            _write_salary_split_sample(sample)
            result = split_salary_by_company(sample, out_dir)
            payload = result.to_dict()

            self.assertEqual(payload["company_count"], 3)
            self.assertEqual(payload["employee_count"], 23)
            self.assertFalse((out_dir / "_salary_split_manifest.json").exists())

            companies = {item["company"]: item for item in payload["outputs"]}
            self.assertEqual(companies["春苗北京"]["employee_count"], 18)
            self.assertEqual(companies["唐人"]["employee_count"], 4)
            self.assertEqual(companies["岩亨"]["employee_count"], 1)

            wb = load_workbook(companies["春苗北京"]["file_path"], data_only=False)
            detail = wb["明细表"]
            self.assertEqual(detail["A18"].value, "河源无线代维合计")
            self.assertEqual(detail["A25"].value, "河源传输代维合计")
            self.assertEqual(detail["A26"].value, "广东分公司（河源项目部）总计")
            self.assertEqual(detail["B6"].value, "员工1")
            self.assertEqual(detail["B24"].value, "员工23")
            self.assertEqual(detail["AU6"].value, "春苗北京")
            self.assertEqual(detail["P18"].value, "=SUM(P6:P17)")
            self.assertEqual(detail["P25"].value, "=SUM(P19:P24)")
            self.assertEqual(detail["P26"].value, "=P18+P25")

            summary = wb["汇总表"]
            self.assertEqual(summary["A6"].value, "广东河源市2026年4月移动基站代维项目")
            self.assertEqual(summary["A7"].value, "广东河源市2026年4月移动线路代维项目")
            self.assertEqual(summary["C6"].value, "='明细表'!P18")
            self.assertEqual(summary["C7"].value, "='明细表'!P25")
            wb.close()

            empty_section_wb = load_workbook(companies["岩亨"]["file_path"], data_only=False)
            empty_detail = empty_section_wb["明细表"]
            detail_labels = [empty_detail.cell(row, 1).value for row in range(1, empty_detail.max_row + 1)]
            self.assertIn("河源无线代维合计", detail_labels)
            self.assertNotIn("河源传输代维合计", detail_labels)
            empty_summary = empty_section_wb["汇总表"]
            summary_labels = [empty_summary.cell(row, 1).value for row in range(1, empty_summary.max_row + 1)]
            self.assertIn("广东河源市2026年4月移动基站代维项目", summary_labels)
            self.assertNotIn("广东河源市2026年4月移动线路代维项目", summary_labels)
            self.assertEqual(empty_summary["A7"].value, "合计")
            self.assertEqual(empty_summary["C7"].value, "=SUM(C6:C6)")
            empty_section_wb.close()

    def test_manifest_is_optional(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            sample = root / "脱敏工资表.xlsx"
            out_dir = root / "output"
            _write_salary_split_sample(sample)
            split_salary_by_company(sample, out_dir, write_manifest=True)
            manifest_path = out_dir / "_salary_split_manifest.json"
            self.assertTrue(manifest_path.exists())

            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(manifest["tool_name"], "需求4-工资表按入职公司拆分")


    def test_split_case1_area_subtotals_only(self) -> None:
        """Case 1: Only Area Subtotals (*小计 -> *总计), without intermediate category totals."""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            sample = root / "area_subtotals_sample.xlsx"
            out_dir = root / "output"
            _write_case1_area_subtotals_sample(sample)
            result = split_salary_by_company(sample, out_dir)
            payload = result.to_dict()

            self.assertEqual(payload["company_count"], 2)
            self.assertEqual(payload["employee_count"], 12)

            companies = {item["company"]: item for item in payload["outputs"]}
            self.assertEqual(companies["公司A"]["employee_count"], 8)
            self.assertEqual(companies["公司B"]["employee_count"], 4)

            wb = load_workbook(companies["公司A"]["file_path"], data_only=False)
            detail = wb["明细表"]
            self.assertEqual(detail["A8"].value, "A区代维小计")
            self.assertEqual(detail["A11"].value, "B区代维小计")
            self.assertEqual(detail["A14"].value, "C区代维小计")
            self.assertEqual(detail["A17"].value, "D区代维小计")
            self.assertEqual(detail["A18"].value, "分公司总计")
            self.assertEqual(detail["P8"].value, "=SUM(P6:P7)")
            self.assertEqual(detail["P11"].value, "=SUM(P9:P10)")
            self.assertEqual(detail["P14"].value, "=SUM(P12:P13)")
            self.assertEqual(detail["P17"].value, "=SUM(P15:P16)")

            summary = wb["汇总表"]
            self.assertEqual(summary["A6"].value, "2026年A区代维小计")
            self.assertEqual(summary["A7"].value, "2026年B区代维小计")
            self.assertEqual(summary["A8"].value, "2026年C区代维小计")
            self.assertEqual(summary["A9"].value, "2026年D区代维小计")
            self.assertEqual(summary["A10"].value, "分公司总计")
            self.assertEqual(summary["C10"].value, "=SUM(C6:C9)")
            wb.close()

    def test_split_case3_hierarchical_subtotals_and_groups(self) -> None:
        """Case 3: Hierarchical Area Subtotals + Category Totals (*小计 -> *合计 -> *总计)."""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            sample = root / "hierarchical_sample.xlsx"
            out_dir = root / "output"
            _write_case3_hierarchical_sample(sample)
            result = split_salary_by_company(sample, out_dir)
            payload = result.to_dict()

            self.assertEqual(payload["company_count"], 2)
            self.assertEqual(payload["employee_count"], 8)

            companies = {item["company"]: item for item in payload["outputs"]}
            wb = load_workbook(companies["公司A"]["file_path"], data_only=False)
            detail = wb["明细表"]
            self.assertEqual(detail["A7"].value, "A区基站小计")
            self.assertEqual(detail["A9"].value, "B区基站小计")
            self.assertEqual(detail["A10"].value, "无线专业合计")
            self.assertEqual(detail["A12"].value, "A区线路小计")
            self.assertEqual(detail["A14"].value, "B区线路小计")
            self.assertEqual(detail["A15"].value, "传输专业合计")
            self.assertEqual(detail["A16"].value, "项目部总计")

            summary = wb["汇总表"]
            self.assertEqual(summary["A6"].value, "A区基站小计")
            self.assertEqual(summary["A7"].value, "B区基站小计")
            self.assertEqual(summary["A8"].value, "无线专业合计")
            self.assertEqual(summary["A9"].value, "A区线路小计")
            self.assertEqual(summary["A10"].value, "B区线路小计")
            self.assertEqual(summary["A11"].value, "传输专业合计")
            self.assertEqual(summary["A12"].value, "项目部总计")
            wb.close()

    def test_split_case4_flat_grand_total_only(self) -> None:
        """Case 4: Flat data without intermediate subtotals (Data -> *总计)."""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            sample = root / "flat_sample.xlsx"
            out_dir = root / "output"
            _write_case4_flat_sample(sample)
            result = split_salary_by_company(sample, out_dir)
            payload = result.to_dict()

            self.assertEqual(payload["company_count"], 2)
            self.assertEqual(payload["employee_count"], 6)

            companies = {item["company"]: item for item in payload["outputs"]}
            wb = load_workbook(companies["公司A"]["file_path"], data_only=False)
            detail = wb["明细表"]
            self.assertEqual(detail["A10"].value, "全公司总计")
            self.assertEqual(detail["P10"].value, "=SUM(P6:P9)")
            wb.close()

    def test_split_many_areas_variable_distribution(self) -> None:
        """Variable area counts: 8 areas across 2 departments."""
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            sample = root / "eight_areas_sample.xlsx"
            out_dir = root / "output"
            _write_many_areas_sample(sample, num_areas=8)
            result = split_salary_by_company(sample, out_dir)
            payload = result.to_dict()

            self.assertEqual(payload["company_count"], 2)
            self.assertEqual(payload["employee_count"], 16)

            for item in payload["outputs"]:
                wb = load_workbook(item["file_path"], data_only=False)
                self.assertIn("明细表", wb.sheetnames)
                self.assertIn("汇总表", wb.sheetnames)
                self.assertGreater(wb["明细表"].max_row, 6)
                wb.close()

    def test_split_real_version2_template_when_available(self) -> None:
        """Verify real 薪资表模板-版本2.xlsx with 31 sheets and 4 hiring companies."""
        v2_path = Path("/Users/salem/Downloads/问题描述附件/薪资表模板-版本2.xlsx")
        if not v2_path.exists():
            self.skipTest("薪资表模板-版本2.xlsx not present in Downloads")

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            out_dir = root / "output_v2"
            result = split_salary_by_company(v2_path, out_dir)
            payload = result.to_dict()

            self.assertEqual(payload["company_count"], 4)
            self.assertEqual(payload["employee_count"], 34)

            companies = {item["company"]: item for item in payload["outputs"]}
            self.assertIn("春苗北京", companies)
            self.assertIn("岩亨", companies)
            self.assertIn("唐人", companies)
            self.assertIn("北京春苗", companies)

            for comp_name, comp_info in companies.items():
                wb = load_workbook(comp_info["file_path"], data_only=False)
                self.assertEqual(len(wb.sheetnames), 31)
                self.assertIn("明细表", wb.sheetnames)
                self.assertIn("汇总表", wb.sheetnames)
                self.assertIn("车辆补贴2025.8", wb.sheetnames)
                self.assertIn("补贴明细3月", wb.sheetnames)

                detail = wb["明细表"]
                summary = wb["汇总表"]
                self.assertGreater(detail.max_row, 6)
                self.assertGreater(summary.max_row, 6)
                wb.close()

    def test_split_problem4_template_when_available(self) -> None:
        """Verify problem 4 template with multiple amount columns (车租 at col 12, 应发 at col 16)."""
        external_path = Path("/Users/salem/Documents/2026年8月人事月度工作/薪酬管理/工资表拆分/20260817_111124_工资表拆分_2026-08-16/上传资料/问题4-薪资表模板(1).xlsx")
        repository_fixture = Path(__file__).resolve().parents[1] / "附件" / "问题4-薪资表模板(1).xlsx"
        p4_path = external_path if external_path.exists() else repository_fixture
        if not p4_path.exists():
            self.skipTest("问题4-薪资表模板(1).xlsx not present")

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            out_dir = root / "output_p4"
            result = split_salary_by_company(p4_path, out_dir)
            payload = result.to_dict()

            companies = {item["company"]: item for item in payload["outputs"]}
            self.assertIn("春苗北京", companies)

            wb = load_workbook(companies["春苗北京"]["file_path"], data_only=False)
            detail = wb["明细表"]
            # Row 18 is wireless subtotal: check car rental (col 12) and total salary (col 16)
            self.assertEqual(detail["A18"].value, "河源无线代维合计")
            self.assertEqual(detail["L18"].value, "=SUM(L6:L17)")
            self.assertEqual(detail["P18"].value, "=SUM(P6:P17)")
            self.assertEqual(detail["U18"].value, "=SUM(U6:U17)")
            self.assertEqual(detail["W18"].value, "=SUM(W6:W17)")

            # Row 26 is grand total
            self.assertEqual(detail["A26"].value, "广东分公司（河源项目部）总计")
            self.assertEqual(detail["L26"].value, "=L18+L25")
            self.assertEqual(detail["P26"].value, "=P18+P25")
            self.assertEqual(detail["U26"].value, "=U18+U25")

            summary = wb["汇总表"]
            self.assertEqual(summary["C6"].value, "='明细表'!P18")
            self.assertEqual(summary["D6"].value, "='明细表'!U18")
            self.assertIsNone(summary["V8"].value)

            # Check merged ranges in detail sheet
            merged_coords = {m.coord for m in detail.merged_cells.ranges}
            self.assertIn("A18:C18", merged_coords)
            self.assertIn("A25:C25", merged_coords)
            self.assertIn("A26:C26", merged_coords)
            self.assertIn("AR26:AY26", merged_coords)

            wb.close()



def _write_salary_split_sample(path: Path) -> None:
    """Create a sanitized two-section salary sheet matching the production layout."""
    workbook = Workbook()
    detail = workbook.active
    detail.title = "明细表"
    summary = workbook.create_sheet("汇总表")

    detail["A1"] = "广东分公司（河源项目部）工资表"
    headers = [f"金额{index}" for index in range(1, 46)] + ["项目", "入职公司"]
    headers[0] = "序号"
    headers[1] = "姓名"
    headers[3] = "身份证号码"
    headers[15] = "应发小计"
    for column, header in enumerate(headers, start=1):
        detail.cell(5, column).value = header

    section_one = [
        *(f"员工{index}" for index in range(1, 13)),
        "唐人员工13",
        "唐人员工14",
        "唐人员工15",
        "岩亨员工16",
    ]
    section_two = ["唐人员工17", *(f"员工{index}" for index in range(18, 24))]
    current_row = 6
    for section_name, names in (("河源无线代维合计", section_one), ("河源传输代维合计", section_two)):
        for name in names:
            detail.cell(current_row, 1).value = current_row - 5
            detail.cell(current_row, 2).value = name
            detail.cell(current_row, 4).value = f"44010019900101{current_row:04d}"
            detail.cell(current_row, 16).value = 100
            detail.cell(current_row, 46).value = section_name.replace("合计", "项目")
            if name.startswith("员工"):
                detail.cell(current_row, 47).value = "春苗北京"
            elif name.startswith("唐人"):
                detail.cell(current_row, 47).value = "唐人"
            else:
                detail.cell(current_row, 47).value = "岩亨"
            current_row += 1
        detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        detail.cell(current_row, 1).value = section_name
        detail.cell(current_row, 16).value = "=SUM(P6:P6)"
        current_row += 1

    detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    detail.cell(current_row, 1).value = "广东分公司（河源项目部）总计"
    detail.cell(current_row, 16).value = "=SUM(P6:P6)"

    summary["A1"] = "工资汇总表"
    summary["A6"] = "广东河源市2026年4月移动基站代维项目"
    summary["A7"] = "广东河源市2026年4月移动线路代维项目"
    summary["A8"] = "合计"
    summary.merge_cells(start_row=9, start_column=1, end_row=9, end_column=21)
    summary["A9"] = "制表："
    workbook.save(path)
    workbook.close()


def _write_case1_area_subtotals_sample(path: Path) -> None:
    """Create Case 1: Only Area Subtotals (A区小计, B区小计, C区小计, D区小计 -> 总计)."""
    workbook = Workbook()
    detail = workbook.active
    detail.title = "明细表"
    summary = workbook.create_sheet("汇总表")

    headers = [f"金额{i}" for i in range(1, 40)] + ["入职公司"]
    headers[0] = "序号"
    headers[1] = "姓名"
    headers[3] = "身份证号码"
    headers[15] = "应发小计"
    for col, h in enumerate(headers, 1):
        detail.cell(5, col).value = h

    current_row = 6
    areas = ["A区代维小计", "B区代维小计", "C区代维小计", "D区代维小计"]
    for area_idx, area_name in enumerate(areas):
        for emp_idx in range(3):
            detail.cell(current_row, 1).value = current_row - 5
            detail.cell(current_row, 2).value = f"员工{current_row}"
            detail.cell(current_row, 4).value = f"11010119900101{current_row:04d}"
            detail.cell(current_row, 16).value = 500
            detail.cell(current_row, len(headers)).value = "公司A" if (area_idx + emp_idx) % 3 != 0 else "公司B"
            current_row += 1
        detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        detail.cell(current_row, 1).value = area_name
        detail.cell(current_row, 16).value = f"=SUM(P{current_row-3}:P{current_row-1})"
        current_row += 1

    detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    detail.cell(current_row, 1).value = "分公司总计"
    detail.cell(current_row, 16).value = "=SUM(P6:P20)"

    summary["A1"] = "工资汇总表"
    for i, name in enumerate(areas):
        r = 6 + i
        summary.cell(r, 1).value = f"2026年{name}"
        summary.cell(r, 2).value = 3
        summary.cell(r, 3).value = f"=明细表!P{9 + i*4}"
    sum_total_r = 6 + len(areas)
    summary.cell(sum_total_r, 1).value = "分公司总计"
    summary.cell(sum_total_r, 3).value = f"=SUM(C6:C{sum_total_r-1})"
    summary.cell(sum_total_r + 1, 1).value = "总经理："

    workbook.save(path)
    workbook.close()


def _write_case3_hierarchical_sample(path: Path) -> None:
    """Create Case 3: Hierarchical Area Subtotals + Category Totals."""
    workbook = Workbook()
    detail = workbook.active
    detail.title = "明细表"
    summary = workbook.create_sheet("汇总表")

    headers = [f"金额{i}" for i in range(1, 40)] + ["入职公司"]
    headers[0] = "序号"
    headers[1] = "姓名"
    headers[3] = "身份证号码"
    headers[15] = "应发小计"
    for col, h in enumerate(headers, 1):
        detail.cell(5, col).value = h

    groups = [
        ("无线专业合计", ["A区基站小计", "B区基站小计"]),
        ("传输专业合计", ["A区线路小计", "B区线路小计"]),
    ]

    current_row = 6
    summary_r = 6
    summary_leaf_rows = []

    for group_name, sub_names in groups:
        group_sub_rows = []
        for sub_name in sub_names:
            first_data = current_row
            for e in range(2):
                detail.cell(current_row, 1).value = current_row - 5
                detail.cell(current_row, 2).value = f"员工{current_row}"
                detail.cell(current_row, 4).value = f"11010119900101{current_row:04d}"
                detail.cell(current_row, 16).value = 500
                detail.cell(current_row, len(headers)).value = "公司A" if current_row % 2 == 0 else "公司B"
                current_row += 1
            last_data = current_row - 1
            detail.cell(current_row, 1).value = sub_name
            detail.cell(current_row, 16).value = f"=SUM(P{first_data}:P{last_data})"
            group_sub_rows.append(current_row)

            summary.cell(summary_r, 1).value = sub_name
            summary.cell(summary_r, 3).value = f"=明细表!P{current_row}"
            summary_leaf_rows.append(summary_r)
            summary_r += 1

            current_row += 1

        detail.cell(current_row, 1).value = group_name
        detail.cell(current_row, 16).value = "=" + "+".join(f"P{r}" for r in group_sub_rows)

        summary.cell(summary_r, 1).value = group_name
        summary.cell(summary_r, 3).value = f"=SUM(C{summary_r-2}:C{summary_r-1})"
        summary_r += 1

        current_row += 1

    detail.cell(current_row, 1).value = "项目部总计"
    detail.cell(current_row, 16).value = f"=P{current_row-1}+P{current_row-5}"

    summary.cell(summary_r, 1).value = "项目部总计"
    summary.cell(summary_r, 3).value = f"=C{summary_r-1}+C{summary_r-4}"
    summary.cell(summary_r + 1, 1).value = "总经理："

    workbook.save(path)
    workbook.close()


def _write_case4_flat_sample(path: Path) -> None:
    """Create Case 4: Flat data without intermediate subtotals."""
    workbook = Workbook()
    detail = workbook.active
    detail.title = "明细表"
    summary = workbook.create_sheet("汇总表")

    headers = [f"金额{i}" for i in range(1, 40)] + ["入职公司"]
    headers[0] = "序号"
    headers[1] = "姓名"
    headers[3] = "身份证号码"
    headers[15] = "应发小计"
    for col, h in enumerate(headers, 1):
        detail.cell(5, col).value = h

    current_row = 6
    for i in range(6):
        detail.cell(current_row, 1).value = i + 1
        detail.cell(current_row, 2).value = f"员工{i+1}"
        detail.cell(current_row, 4).value = f"11010119900101{i+1:04d}"
        detail.cell(current_row, 16).value = 1000
        detail.cell(current_row, len(headers)).value = "公司A" if i < 4 else "公司B"
        current_row += 1

    detail.cell(current_row, 1).value = "全公司总计"
    detail.cell(current_row, 16).value = "=SUM(P6:P11)"

    summary["A1"] = "汇总表"
    summary["A6"] = "全公司总计"
    summary["C6"] = "=明细表!P12"
    summary["A7"] = "总经理："

    workbook.save(path)
    workbook.close()


def _write_many_areas_sample(path: Path, num_areas: int = 8) -> None:
    """Create a sample with many areas (e.g. 8 areas)."""
    workbook = Workbook()
    detail = workbook.active
    detail.title = "明细表"
    summary = workbook.create_sheet("汇总表")

    headers = [f"金额{i}" for i in range(1, 40)] + ["入职公司"]
    headers[0] = "序号"
    headers[1] = "姓名"
    headers[3] = "身份证号码"
    headers[15] = "应发小计"
    for col, h in enumerate(headers, 1):
        detail.cell(5, col).value = h

    current_row = 6
    for i in range(num_areas):
        detail.cell(current_row, 1).value = i * 2 + 1
        detail.cell(current_row, 2).value = f"员工{i*2+1}"
        detail.cell(current_row, 4).value = f"11010119900101{i*2+1:04d}"
        detail.cell(current_row, 16).value = 600
        detail.cell(current_row, len(headers)).value = "公司A"
        current_row += 1

        detail.cell(current_row, 1).value = i * 2 + 2
        detail.cell(current_row, 2).value = f"员工{i*2+2}"
        detail.cell(current_row, 4).value = f"11010119900101{i*2+2:04d}"
        detail.cell(current_row, 16).value = 600
        detail.cell(current_row, len(headers)).value = "公司B"
        current_row += 1

        detail.cell(current_row, 1).value = f"区域{i+1}小计"
        detail.cell(current_row, 16).value = f"=SUM(P{current_row-2}:P{current_row-1})"
        summary.cell(6 + i, 1).value = f"区域{i+1}小计"
        summary.cell(6 + i, 3).value = f"=明细表!P{current_row}"
        current_row += 1

    detail.cell(current_row, 1).value = "分公司总计"
    detail.cell(current_row, 16).value = f"=SUM(P6:P{current_row-1})"

    tot_r = 6 + num_areas
    summary.cell(tot_r, 1).value = "分公司总计"
    summary.cell(tot_r, 3).value = f"=SUM(C6:C{tot_r-1})"
    summary.cell(tot_r + 1, 1).value = "总经理："

    workbook.save(path)
    workbook.close()


if __name__ == "__main__":
    unittest.main()
