from __future__ import annotations

import tempfile
import unittest
import shutil
from datetime import date
from io import BytesIO
from pathlib import Path
from unittest.mock import patch
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

from hr_toolkit.tools.salary_merge import AMOUNT_NUMBER_FORMAT, SUMMARY_TITLE, merge_monthly_salary
from hr_toolkit.tools.salary_split import split_salary_by_company


class SalaryMergeTest(unittest.TestCase):
    def test_accept_monthly_gross_header_and_first_id_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / '工资表_202608.xlsx'
            _write_current_salary_file(source, amount=1035)
            result = merge_monthly_salary(source, root / 'output')
            self.assertEqual((result.employee_count, result.record_count), (1, 1))
            self.assertEqual(result.warnings, [])
            wb = load_workbook(result.output_file, data_only=True)
            try:
                self.assertEqual(wb['汇总']['C5'].value, 'TEST-001')
                self.assertEqual(wb['汇总']['K5'].value, 1035)
            finally:
                wb.close()

    def test_monthly_formula_without_cache_uses_actual_ranges(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / '工资表_202608.xlsx'
            _write_current_salary_file(source)
            result = merge_monthly_salary(source, root / 'output')
            self.assertEqual(result.warnings, [])
            wb = load_workbook(result.output_file, data_only=True)
            try:
                # M is income, P is a deduction; O is itself an uncached formula.
                self.assertEqual(wb['汇总']['K5'].value, 1035)
            finally:
                wb.close()

    def test_monthly_formula_prefers_saved_excel_cache(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / '工资表_202608.xlsx'
            _write_current_salary_file(source, amount='=VLOOKUP(B6,其他表!A:C,3,0)')
            _set_formula_cache(source, 'Q6', 1234.56)
            result = merge_monthly_salary(source, root / 'output')
            self.assertEqual(result.warnings, [])
            wb = load_workbook(result.output_file, data_only=True)
            try:
                self.assertEqual(wb['汇总']['K5'].value, 1234.56)
            finally:
                wb.close()

    def test_old_and_new_templates_merge_without_overwriting_existing_months(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            old = root / '旧格式_202607.xlsx'
            new = root / '新格式_202608.xlsx'
            _write_salary_file(old, [('测试人员', 'TEST-001', 500)])
            _write_current_salary_file(new)
            initial = merge_monthly_salary([old, new], root / 'initial')
            self.assertEqual((initial.employee_count, initial.record_count), (1, 2))
            _write_current_salary_file(new, amount=9999)
            appended = merge_monthly_salary(new, root / 'appended', existing_summary_path=initial.output_file)
            self.assertEqual(appended.skipped_record_count, 1)
            wb = load_workbook(appended.output_file, data_only=True)
            try:
                self.assertEqual(wb['汇总']['J5'].value, 500)
                self.assertEqual(wb['汇总']['K5'].value, 1035)
            finally:
                wb.close()

    def test_split_then_merge_monthly_gross_without_excel_recalculation(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / '工资表_202608.xlsx'
            _write_current_salary_file(source)
            split = split_salary_by_company(source, root / 'split')
            result = merge_monthly_salary([item.file_path for item in split.outputs], root / 'merged')
            self.assertEqual((result.employee_count, result.record_count), (1, 1))
            self.assertEqual(result.warnings, [])
            wb = load_workbook(result.output_file, data_only=True)
            try:
                self.assertEqual(wb['汇总']['K5'].value, 1035)
            finally:
                wb.close()

    def test_legacy_header_takes_precedence_and_preserves_legacy_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / '工资表_202608.xlsx'
            _write_salary_file(source, [('测试人员', 'TEST-001', 500)])
            wb = load_workbook(source)
            ws = wb['明细表']
            ws['Q4'] = '本月应发工资'
            ws['Q5'] = 999
            ws['E5'] = 800
            ws['M5'] = 50
            ws['P5'] = '=ROUND(SUM(E5:L5)-SUM(M5:O5),2)'
            wb.save(source)
            wb.close()
            result = merge_monthly_salary(source, root / 'output')
            wb = load_workbook(result.output_file, data_only=True)
            try:
                self.assertEqual(wb['汇总']['K5'].value, 750)
            finally:
                wb.close()

    def test_uncached_monthly_formula_rounding_and_shifted_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            for income, expected in ((2.675, 2.68), (-2.675, -2.68)):
                with self.subTest(income=income):
                    source = root / '工资表_202608.xlsx'
                    _write_current_salary_file(source)
                    wb = load_workbook(source)
                    ws = wb['明细表']
                    ws['F6'] = income
                    ws['G6'] = 0
                    ws['Q6'] = '= round( sum($F$6:$F$6) - sum($G6:$G6), 2)'
                    wb.save(source)
                    wb.close()
                    result = merge_monthly_salary(source, root / 'output')
                    wb = load_workbook(result.output_file, data_only=True)
                    try:
                        self.assertEqual(wb['汇总']['K5'].value, expected)
                    finally:
                        wb.close()

    def test_unsupported_uncached_formula_reports_reason_not_zero(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            for formula in (
                '=VLOOKUP(B6,其他表!A:C,3,0)',
                '=ROUND(SUM(E6:M7)-SUM(N6:P6),2)',
                '=ROUND(SUM(E6:Q6)-SUM(N6:P6),2)',
                '=ROUND(' + ' ' * 512 + 'SUM(E6:M6)-SUM(N6:P6),2)',
            ):
                with self.subTest(formula=formula):
                    source = root / '工资表_202608.xlsx'
                    _write_current_salary_file(source, amount=formula)
                    with self.assertRaisesRegex(ValueError, 'Q6.*重新计算.*保存'):
                        merge_monthly_salary(source, root / 'output')
                    self.assertFalse((root / 'output' / '个人薪资汇总表.xlsx').exists())

    def test_uncached_dependency_errors_are_not_treated_as_zero(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            for formula in ('=1/0', '=O6+1', '=SUM(N6:P6)', '=E7+1', '=' + '+'.join(['1'] * 100)):
                with self.subTest(formula=formula):
                    source = root / '工资表_202608.xlsx'
                    _write_current_salary_file(source)
                    wb = load_workbook(source)
                    wb['明细表']['O6'] = formula
                    wb.save(source)
                    wb.close()
                    with self.assertRaisesRegex(ValueError, 'Q6.*重新计算.*保存'):
                        merge_monthly_salary(source, root / 'output')

    def test_cached_and_uncached_same_row_arithmetic_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            for formula, cache, expected in (
                ('=-(-N6*2)+(+10)/2', None, 1035),
                ('=SUM(N6:P6)', 25, 1035),
            ):
                with self.subTest(formula=formula):
                    source = root / '工资表_202608.xlsx'
                    _write_current_salary_file(source)
                    wb = load_workbook(source)
                    wb['明细表']['O6'] = formula
                    wb.save(source)
                    wb.close()
                    if cache is not None:
                        _set_formula_cache(source, 'O6', cache)
                    result = merge_monthly_salary(source, root / 'output')
                    wb = load_workbook(result.output_file, data_only=True)
                    try:
                        self.assertEqual(wb['汇总']['K5'].value, expected)
                    finally:
                        wb.close()

    def test_merge_monthly_salary_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()

            _write_salary_file(
                input_dir / "河源项目部工资表_202604.xlsx",
                [
                    ("员工1", "44162219901007667X", 500),
                    ("员工2", "441622198404210312", 600),
                ],
            )
            _write_salary_file(
                input_dir / "河源项目部工资表_202605.xlsx",
                [
                    ("员工1", "44162219901007667X", 700),
                    ("员工3", "44162219800516649X", 800),
                ],
            )

            result = merge_monthly_salary(input_dir, output_dir, year=2026)
            payload = result.to_dict()

            self.assertEqual(payload["source_file_count"], 2)
            self.assertEqual(payload["employee_count"], 3)
            self.assertEqual(payload["record_count"], 4)
            self.assertEqual(payload["applied_record_count"], 4)
            self.assertEqual(payload["skipped_record_count"], 0)
            self.assertEqual(payload["months"][0], "202601")
            self.assertEqual(payload["months"][-1], "202612")
            self.assertTrue(result.output_file and result.output_file.exists())

            wb = load_workbook(result.output_file, data_only=True)
            ws = wb["汇总"]
            self.assertEqual(ws["A1"].value, SUMMARY_TITLE)
            self.assertEqual(ws["A1"].fill.fill_type, None)
            self.assertEqual(ws["D3"].value, 202601)
            self.assertEqual(ws["D3"].fill.fill_type, None)
            self.assertEqual(ws["D5"].number_format, AMOUNT_NUMBER_FORMAT)
            rows = {
                ws.cell(row, 3).value: [ws.cell(row, col).value for col in range(1, 16)]
                for row in range(5, 8)
            }
            self.assertEqual(rows["44162219901007667X"][3:8], [0, 0, 0, 500, 700])
            self.assertEqual(ws["D6"].value, 0)
            self.assertEqual(ws["D6"].number_format, AMOUNT_NUMBER_FORMAT)
            self.assertEqual(rows["441622198404210312"][3:8], [0, 0, 0, 600, 0])
            self.assertEqual(rows["44162219800516649X"][3:8], [0, 0, 0, 0, 800])

    def test_append_to_existing_summary_and_skip_existing_month(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            first_input_dir = root / "first_input"
            first_output_dir = root / "first_output"
            second_input_dir = root / "second_input"
            second_output_dir = root / "second_output"
            first_input_dir.mkdir()
            second_input_dir.mkdir()

            _write_salary_file(
                first_input_dir / "河源项目部工资表_202601.xlsx",
                [("员工1", "44162219901007667X", 100)],
            )
            _write_salary_file(
                first_input_dir / "河源项目部工资表_202602.xlsx",
                [
                    ("员工1", "44162219901007667X", 200),
                    ("员工2", "441622198404210312", 300),
                ],
            )
            first_result = merge_monthly_salary(first_input_dir, first_output_dir, year=2026)

            _write_salary_file(
                second_input_dir / "河源项目部工资表_202601.xlsx",
                [("员工1", "44162219901007667X", 999)],
            )
            _write_salary_file(
                second_input_dir / "河源项目部工资表_202603.xlsx",
                [
                    ("员工1", "44162219901007667X", 400),
                    ("员工3", "44162219800516649X", 500),
                ],
            )

            result = merge_monthly_salary(
                second_input_dir,
                second_output_dir,
                existing_summary_path=first_result.output_file,
            )
            payload = result.to_dict()

            self.assertEqual(payload["source_file_count"], 2)
            self.assertEqual(payload["employee_count"], 3)
            self.assertEqual(payload["record_count"], 3)
            self.assertEqual(payload["applied_record_count"], 2)
            self.assertEqual(payload["skipped_record_count"], 1)
            self.assertTrue(any("已存在金额，未覆盖" in warning for warning in payload["warnings"]))

            wb = load_workbook(result.output_file, data_only=True)
            ws = wb["汇总"]
            rows = {
                ws.cell(row, 3).value: [ws.cell(row, col).value for col in range(1, 16)]
                for row in range(5, 8)
            }
            self.assertEqual(rows["44162219901007667X"][3:7], [100, 200, 400, 0])
            self.assertEqual(rows["441622198404210312"][3:7], [0, 300, 0, 0])
            self.assertEqual(rows["44162219800516649X"][3:7], [0, 0, 500, 0])

    def test_detect_month_from_cell_and_accept_id_card_alias(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()

            _write_salary_file(
                input_dir / "河源项目部工资表.xlsx",
                [("员工1", "44162219901007667X", 1200)],
                id_header="身份证号",
                month_date=date(2026, 6, 1),
            )

            result = merge_monthly_salary(input_dir, output_dir)

            self.assertEqual(result.months, [f"2026{month:02d}" for month in range(1, 13)])
            self.assertEqual(result.employee_count, 1)
            self.assertEqual(result.record_count, 1)
            wb = load_workbook(result.output_file, data_only=True)
            ws = wb["汇总"]
            self.assertEqual(ws["I3"].value, 202606)
            self.assertEqual(ws["I5"].value, 1200)

    def test_merge_accepts_xls_files_after_conversion(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            xlsx_file = input_dir / "河源项目部工资表_202607.xlsx"
            xls_file = input_dir / "河源项目部工资表_202607.xls"
            _write_salary_file(xlsx_file, [("员工1", "44162219901007667X", 900)])
            shutil.move(xlsx_file, xls_file)

            def fake_convert(source: Path, output_path: Path) -> None:
                shutil.copyfile(source, output_path)

            with patch("hr_toolkit.common.excel_compat._convert_xls_to_xlsx", side_effect=fake_convert):
                result = merge_monthly_salary(input_dir, output_dir)

            self.assertEqual(result.source_files[0].endswith(".xlsx"), True)
            wb = load_workbook(result.output_file, data_only=True)
            ws = wb["汇总"]
            self.assertEqual(ws["J3"].value, 202607)
            self.assertEqual(ws["J5"].value, 900)
            wb.close()


def _set_formula_cache(path: Path, coordinate: str, value: float) -> None:
    """Emulate Excel's saved formula result in a synthetic fixture only."""
    buffer = BytesIO()
    namespace = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
    with ZipFile(path) as source, ZipFile(buffer, 'w') as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == 'xl/worksheets/sheet1.xml':
                sheet = ElementTree.fromstring(payload)
                cell = sheet.find(f'.//{namespace}c[@r="{coordinate}"]')
                assert cell is not None
                cached = cell.find(f'{namespace}v')
                assert cached is not None
                cached.text = str(value)
                payload = ElementTree.tostring(sheet, encoding='utf-8')
            target.writestr(item, payload)
    path.write_bytes(buffer.getvalue())


def _write_current_salary_file(
    path: Path, *, amount: str | float = '=ROUND(SUM(E6:M6)-SUM(N6:P6),2)'
) -> None:
    """Synthetic two-row header, duplicate ID column and uncached payroll formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = '明细表'
    ws['A1'] = '测试项目2026年8月工资表'
    headers = {
        'A4': '序号',
        'B4': '姓名',
        'D4': '身份证号',
        'Q4': '本月应发工资',
        'R4': '入职公司',
        'S4': '身份证号',
    }
    for coordinate, value in headers.items():
        ws[coordinate] = value
        col = ws[coordinate].column
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
    values = {
        'A6': 1,
        'B6': '测试人员',
        'D6': 'TEST-001',
        'E6': 1000,
        'M6': 100,
        'N6': 10,
        'O6': '=20+5',
        'P6': 30,
        'Q6': amount,
        'R6': '测试公司',
        'S6': 'OTHER-ID',
        'A7': '总计',
    }
    for coordinate, value in values.items():
        ws[coordinate] = value
    summary = wb.create_sheet('汇总表')
    summary['A1'] = '2026年8月工资汇总表'
    summary['A4'] = '项目名称'
    summary['C4'] = '本月应发工资'
    summary['A6'] = '测试项目'
    summary['C6'] = "='明细表'!Q7"
    summary['A7'] = '合计'
    wb.save(path)
    wb.close()


def _write_salary_file(
    path: Path,
    employees: list[tuple[str, str, int]],
    *,
    id_header: str = "身份证号码",
    month_date: date | None = None,
) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "明细表"
    if month_date is not None:
        ws["A3"] = month_date
    ws["A4"] = "序号"
    ws["B4"] = "姓名"
    ws["D4"] = id_header
    ws["P4"] = "应发小计"
    for index, (name, id_card, amount) in enumerate(employees, start=1):
        row = index + 4
        ws.cell(row, 1).value = index
        ws.cell(row, 2).value = name
        ws.cell(row, 4).value = id_card
        ws.cell(row, 16).value = amount
    workbook.save(path)


if __name__ == "__main__":
    unittest.main()
